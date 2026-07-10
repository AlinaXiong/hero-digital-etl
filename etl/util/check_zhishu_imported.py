# -*- coding: utf-8 -*-
"""多线程检查 Excel 合同是否已导入智书,并导出未导入清单。

默认只读取项目内 ``resources/source/zhishu_import_check``。
该目录里若只有一个 ``.xlsx`` 文件则自动使用;若有多个文件,请用参数或环境变量指定具体文件。
输出只包含 3 列,用于后续混合增补清单:

    合同编号 / 关联业财订单（必填） / 智书合同类型

接口按截图调用:

    POST https://open.qfei.cn/open-apis/contract/v1/contracts/search?user_id_type=user_id
    {"contract_number": "..."}

认证信息从环境变量读取,不要写死在代码里。常用配置:

    ZHISHU_AUTHORIZATION=Bearer xxx
    # 或 ZHISHU_BEARER_TOKEN=xxx
    # 或 ZHISHU_APP_ID=xxx + ZHISHU_APP_SECRET=xxx 自动获取 tenant_access_token
    # 如网关用 Cookie,可配置 ZHISHU_COOKIE=...

运行:

    python run.py check_zhishu_imported
    python -m etl.util.check_zhishu_imported --input resources/source/zhishu_import_check/已有标黄.xlsx
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from threading import Lock

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if __package__ is None or __package__ == '':
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from etl.util import common as c


def _env_value(name, default=''):
    value = os.getenv(name)
    if value is None or str(value).strip() == '':
        return default
    return str(value).strip()


TASK_NAME = 'zhishu_import_check'
SCRIPT_VERSION = '20260710-final-retry-v5'
OUTPUT_DIR = c.OUT_DIR / TASK_NAME
SOURCE_DIR = c.SRC_DIR / TASK_NAME
PROJECT_INPUT_FILE = SOURCE_DIR / '已有标黄.xlsx'
DEFAULT_INPUT_FILE = Path(_env_value('ZHISHU_CHECK_INPUT_FILE', SOURCE_DIR))
DEFAULT_OUTPUT_FILE = OUTPUT_DIR / f'未导入智书清单_{c.today_suffix()}.xlsx'
DEFAULT_SEARCH_URL = (
    'https://open.qfei.cn/open-apis/contract/v1/contracts/search'
    '?user_id_type=user_id'
)
DEFAULT_TOKEN_URL = 'https://open.qfei.cn/open-apis/auth/v3/tenant_access_token/internal'

OUTPUT_COLUMNS = ('合同编号', '关联业财订单（必填）', '智书合同类型', '泛微项目编码')
CONTRACT_COLUMN_CANDIDATES = (
    '合同编号', '合同号', '合同编码', 'contract_number（合同编码）', 'contract_number',
)
ORDER_COLUMN_CANDIDATES = (
    '关联业财订单（必填）',
    '关联业财订单',
    '订单编号',
    'custom_1024_90a78c8120994f95b2dbfedd297c7d81（相关单据-订单信息）',
    'custom_1_7f977c0d30064dd199434f706470c669（订单编号）',
)
TYPE_COLUMN_CANDIDATES = (
    '智书合同类型',
    '合同类型',
    'contractCategory(智书框架合同类型)',
    'contractCategory（智书框架合同类型）',
)
PROJECT_CODE_COLUMN_CANDIDATES = (
    '泛微项目编码',
    '项目编码',
    '泛微项目编号',
    '项目编号',
    'project_code',
    'projectCode',
)


CONTRACT_TYPE_ITEMS = (
    ('ZBJJ', '主播专项-主播经纪'),
    ('PTJJ', '主播专项-平台经纪'),
    ('ZBZM', '主播专项-主播招募'),
    ('ZBGK', '主播专项-主播挂靠'),
    ('ZBZR', '主播专项-主播转让'),
    ('MLZB', '主播专项-马来直播'),
    ('INZB', '主播专项-印尼直播'),
    ('ZBQT', '主播专项-其他'),
    ('NDA', '其他-保密协议'),
    ('FSYHL', '其他-反商业贿赂协议'),
    ('BC', '其他-补充/变更协议'),
    ('MOU', '其他-战略合作协议'),
    ('QTLX', '其他-其他类型'),
    ('NBJZ', '内部合同-内部结转'),
    ('LC', '内部合同-理财产品'),
    ('DK', '内部合同-银行贷款'),
    ('GNDF', '单次支出-国内赛事及活动支出'),
    ('XZDF', '单次支出-行政运营及人力支出'),
    ('JYDF', '单次支出-简易支出'),
    ('TA', '单次支出-其他支出'),
    ('HLWDF', '单次支出-互联网产品支出'),
    ('DCDF', '单次支出-地产运营支出'),
    ('DSDF', '单次支出-电商运营支出'),
    ('ENDF', '单次支出-海外赛事业及活动支出'),
    ('GRDF', '单次支出-个人合作支出'),
    ('JJDF', '单次支出-奖金补贴支出'),
    ('ZCDF', '单次支出-资产采购租赁及经营支出'),
    ('GGDF', '单次支出-广告赞助支出'),
    ('ZHDF', '单次支出-整合营销支出'),
    ('IPDF', '单次支出-衍生品支出'),
    ('SXDF', '单次支出-视效支出'),
    ('SSDS', '单次收入-赛事及活动收入'),
    ('DCDS', '单次收入-地产运营收入'),
    ('DSDS', '单次收入-电商运营收入'),
    ('ZCDS', '单次收入-资产采购租赁及经营收入'),
    ('GGDS', '单次收入-广告赞助收入'),
    ('ZHDS', '单次收入-整合营销收入'),
    ('IPDS', '单次收入-衍生品收入'),
    ('SXDS', '单次收入-视效收入'),
    ('XZDS', '单次收入-行政运营及人力收入'),
    ('QTDS', '单次收入-其他收入'),
    ('HLWDS', '单次收入-互联网产品收入'),
    ('TYDJ', '单次收支-通用单次收支'),
    ('SSKF', '框架支出-赛事及活动支出框架'),
    ('DCKF', '框架支出-地产运营支出框架'),
    ('DSKF', '框架支出-电商运营支出框架'),
    ('ZCKF', '框架支出-资产采购租赁及经营支出框架'),
    ('GGKF', '框架支出-广告赞助支出框架'),
    ('ZHKF', '框架支出-整合营销支出框架'),
    ('IPKF', '框架支出-衍生品支出框架'),
    ('SXKF', '框架支出-视效支出框架'),
    ('XZKF', '框架支出-行政运营及人力支出框架'),
    ('QTKF', '框架支出-其他支出框架'),
    ('HLWKF', '框架支出-互联网产品支出框架'),
    ('SSKS', '框架收入-赛事及活动收入框架'),
    ('DCKS', '框架收入-地产运营收入框架'),
    ('DSKS', '框架收入-电商运营收入框架'),
    ('ZCKS', '框架收入-资产采购租赁及经营收入框架'),
    ('GGKS', '框架收入-广告赞助收入'),
    ('ZHKS', '框架收入-整合营销收入框架'),
    ('IPKS', '框架收入-衍生品收入'),
    ('SXKS', '框架收入-视效收入框架'),
    ('XZKS', '框架收入-行政运营及人力收入框架'),
    ('QTKS', '框架收入-其他收入框架'),
    ('HLWKS', '框架收入-互联网产品收入框架'),
    ('MCNKS', '框架收入-MCN平台收入框架'),
    ('TYKJ', '框架收支-通用框架收支'),
    ('SSOF', '订单支出-赛事及活动支出订单'),
    ('DCOF', '订单支出-地产运营支出订单'),
    ('DSOF', '订单支出-电商运营支出订单'),
    ('ZCOF', '订单支出-资产采购租赁及经营支出订单'),
    ('GGOF', '订单支出-广告赞助支出订单'),
    ('ZHOF', '订单支出-整合营销支出订单'),
    ('IPOF', '订单支出-衍生品支出订单'),
    ('SXOF', '订单支出-视效支出订单'),
    ('XZOF', '订单支出-行政运营及人力支出订单'),
    ('QTOF', '订单支出-其他支出订单'),
    ('HLWOF', '订单支出-互联网产品支出订单'),
    ('SSOS', '订单收入-赛事及活动收入订单'),
    ('DCOS', '订单收入-地产运营收入订单'),
    ('DSOS', '订单收入-电商运营订单收入'),
    ('ZCOS', '订单收入-资产采购租赁及经营收入订单'),
    ('GGOS', '订单收入-广告赞助收入订单'),
    ('ZHOS', '订单收入-整合营销收入订单'),
    ('IPOS', '订单收入-衍生品收入订单'),
    ('SXOS', '订单收入-视效订单收入'),
    ('XZOS', '订单收入-行政运营及人力收入订单'),
    ('QTOS', '订单收入-其他收入订单'),
    ('HLWOS', '订单收入-互联网产品收入订单'),
    ('MCNOS', '订单收入-MCN平台收入订单'),
    ('TZ', '资本专项-投资协议'),
    ('RZ', '资本专项-融资协议'),
    ('JK', '资本专项-借款协议'),
    ('QTCP', '资本专项-其他'),
)

CODE_TO_TYPE = dict(CONTRACT_TYPE_ITEMS)
TYPE_TO_CODE = {label: code for code, label in CONTRACT_TYPE_ITEMS}
VALID_TYPE_LABELS = set(TYPE_TO_CODE)
TYPE_FALLBACK_BY_GROUP = {
    '主播专项': '主播专项-其他',
    '其他': '其他-其他类型',
    '单次支出': '单次支出-其他支出',
    '单次收入': '单次收入-其他收入',
    '单次收支': '单次收支-通用单次收支',
    '框架支出': '框架支出-其他支出框架',
    '框架收入': '框架收入-其他收入框架',
    '框架收支': '框架收支-通用框架收支',
    '订单支出': '订单支出-其他支出订单',
    '订单收入': '订单收入-其他收入订单',
    '资本专项': '资本专项-其他',
}
UNKNOWN_NON_BLANK_TYPE_FALLBACK = '其他-其他类型'
NOT_IMPORTED_CODES = {'110107'}
_TOKEN_CACHE = None
_TOKEN_LOCK = Lock()


@dataclass
class ContractInput:
    contract_number: str
    related_orders: str = ''
    contract_type: str = ''
    project_code: str = ''
    source_rows: str = ''
    type_note: str = ''


@dataclass
class QueryResult:
    contract_number: str
    imported: bool
    elapsed_ms: int
    response_summary: str = ''
    status_names: str = ''


def _text(value):
    if value is None:
        return ''
    text = str(value).strip()
    return '' if text in ('', 'nan', 'None', 'NaT') else text


def _contract_key(value):
    return re.sub(r'\s+', '', _text(value)).upper()


def _header_key(value):
    text = _text(value).lower()
    return re.sub(r'[\s_()（）\[\]【】:：;；,，/\\-]+', '', text)


def _first_existing_index(headers, candidates):
    header_map = {_header_key(header): index for index, header in enumerate(headers)}
    for candidate in candidates:
        index = header_map.get(_header_key(candidate))
        if index is not None:
            return index
    return None


def _split_multi_values(value):
    text = _text(value)
    if not text:
        return []
    return [part.strip() for part in re.split(r'[\r\n,，;；]+', text) if part.strip()]


def _join_unique(values):
    result = []
    seen = set()
    for value in values:
        for part in _split_multi_values(value):
            if part and part not in seen:
                seen.add(part)
            result.append(part)
    return '; '.join(result)


def _resolve_input_file(value):
    path = Path(value)
    if path.is_dir():
        named_file = path / PROJECT_INPUT_FILE.name
        if named_file.exists():
            return named_file
        files = sorted(item for item in path.glob('*.xlsx') if not item.name.startswith('~$'))
        if len(files) == 1:
            return files[0]
        if not files:
            raise FileNotFoundError(
                f'输入目录没有 Excel 文件: {path}\n'
                f'请把待检查文件放到 {SOURCE_DIR}, 推荐文件名: {PROJECT_INPUT_FILE.name}'
            )
        raise RuntimeError(
            '输入目录存在多个 Excel,请只保留一个,或用 --input / ZHISHU_CHECK_INPUT_FILE 指定具体文件: '
            + ', '.join(item.name for item in files)
        )
    return path


def normalize_contract_type(value):
    """把合同类型规整为智书枚举标签;不合法时按大类兜到“其他”类型。"""
    text = _text(value)
    if not text:
        return '', ''
    if text in CODE_TO_TYPE:
        return CODE_TO_TYPE[text], f'合同类型编码 {text} 已转为名称'
    if text in VALID_TYPE_LABELS:
        return text, ''

    group = text.split('-', 1)[0].strip()
    fallback = TYPE_FALLBACK_BY_GROUP.get(group)
    if fallback:
        return fallback, f'原类型“{text}”不在智书枚举中,按大类“{group}”兜底'
    return UNKNOWN_NON_BLANK_TYPE_FALLBACK, f'原类型“{text}”不在智书枚举中,未识别大类,兜底为其他类型'


def _detect_header_row(worksheet, max_scan_rows=20):
    for row_number, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, max_scan_rows), values_only=True),
            start=1):
        headers = [_text(value) for value in row]
        if _first_existing_index(headers, CONTRACT_COLUMN_CANDIDATES) is not None:
            return row_number, headers
    return None, []


def read_contract_inputs(input_file, sheet_name=None, all_sheets=False):
    """从普通列表或智书导入模板中抽取合同编号、业财订单、合同类型、泛微项目编码。"""
    input_file = _resolve_input_file(input_file)
    if not input_file.exists():
        raise FileNotFoundError(f'输入 Excel 不存在: {input_file}')

    read_progress_rows = int(_env_value('ZHISHU_READ_PROGRESS_ROWS', '20000'))
    print(f'[智书导入检查] 读取输入文件: {input_file}', flush=True)
    print('[智书导入检查] 正在打开 Excel...', flush=True)
    wb = load_workbook(input_file, read_only=True, data_only=True)
    if sheet_name:
        worksheets = [wb[sheet_name]]
    elif all_sheets:
        worksheets = wb.worksheets
    else:
        worksheets = [wb.worksheets[0]]
    print(f'[智书导入检查] Excel 已打开,待扫描 sheet {len(worksheets)} 个', flush=True)

    by_contract = {}
    skipped_sheets = []
    for sheet_index, ws in enumerate(worksheets, start=1):
        print(
            f'[智书导入检查] 扫描 sheet {sheet_index}/{len(worksheets)}: '
            f'{ws.title} ({ws.max_row} 行 x {ws.max_column} 列)',
            flush=True,
        )
        header_row, headers = _detect_header_row(ws)
        if not header_row:
            skipped_sheets.append(ws.title)
            print(f'[智书导入检查] sheet {ws.title} 未找到合同编号表头,跳过', flush=True)
            continue

        contract_idx = _first_existing_index(headers, CONTRACT_COLUMN_CANDIDATES)
        order_idx = _first_existing_index(headers, ORDER_COLUMN_CANDIDATES)
        type_idx = _first_existing_index(headers, TYPE_COLUMN_CANDIDATES)
        project_code_idx = _first_existing_index(headers, PROJECT_CODE_COLUMN_CANDIDATES)
        if contract_idx is None:
            skipped_sheets.append(ws.title)
            print(f'[智书导入检查] sheet {ws.title} 未找到合同编号列,跳过', flush=True)
            continue

        sheet_contract_count = 0
        scanned_count = 0
        for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            scanned_count += 1
            raw_contract = row[contract_idx] if contract_idx < len(row) else ''
            contract_number = _text(raw_contract)
            key = _contract_key(contract_number)
            if read_progress_rows > 0 and scanned_count % read_progress_rows == 0:
                print(
                    f'[智书导入检查] sheet {ws.title} 已扫描 {scanned_count} 行; '
                    f'本 sheet 命中 {sheet_contract_count}; 累计合同 {len(by_contract)}',
                    flush=True,
                )
            if not key:
                continue
            sheet_contract_count += 1
            order_value = _text(row[order_idx]) if order_idx is not None and order_idx < len(row) else ''
            type_value = _text(row[type_idx]) if type_idx is not None and type_idx < len(row) else ''
            project_code_value = (
                _text(row[project_code_idx])
                if project_code_idx is not None and project_code_idx < len(row)
                else ''
            )

            item = by_contract.setdefault(key, {
                'contract_number': contract_number,
                'orders': [],
                'types': [],
                'project_codes': [],
                'source_rows': [],
            })
            item['source_rows'].append(f'{ws.title}!{row_number}')
            if order_value:
                item['orders'].append(order_value)
            if type_value:
                item['types'].append(type_value)
            if project_code_value:
                item['project_codes'].append(project_code_value)
        print(
            f'[智书导入检查] sheet {ws.title} 完成: 扫描 {scanned_count} 行; '
            f'命中 {sheet_contract_count}; 累计合同 {len(by_contract)}',
            flush=True,
        )

    result = []
    for item in by_contract.values():
        normalized_type, type_note = normalize_contract_type(_join_unique(item['types']))
        result.append(ContractInput(
            contract_number=item['contract_number'],
            related_orders=_join_unique(item['orders']),
            contract_type=normalized_type,
            project_code=_join_unique(item['project_codes']),
            source_rows='; '.join(item['source_rows'][:5]) + ('; ...' if len(item['source_rows']) > 5 else ''),
            type_note=type_note,
        ))

    result.sort(key=lambda item: item.contract_number)
    print(f'[智书导入检查] 读取合同 {len(result)} 个; 跳过无合同编号表头 sheet {len(skipped_sheets)} 个', flush=True)
    if skipped_sheets:
        print('[智书导入检查] 跳过 sheet: ' + ', '.join(skipped_sheets[:10]), flush=True)
    return result


def _extract_token(payload):
    if not isinstance(payload, dict):
        return ''
    candidates = (
        payload,
        payload.get('data') or {},
        payload.get('tenant_access_token') or {},
    )
    for item in candidates:
        if not isinstance(item, dict):
            continue
        for key in (
                'tenant_access_token',
                'tenantAccessToken',
                'access_token',
                'accessToken',
                'token',
        ):
            token = _text(item.get(key))
            if token:
                return token
    return ''


def _fetch_tenant_access_token():
    app_id = os.getenv('ZHISHU_APP_ID', '').strip()
    app_secret = os.getenv('ZHISHU_APP_SECRET', '').strip()
    if not app_id or not app_secret:
        raise RuntimeError(
            '缺少智书接口认证配置: 请在 .env 配置 ZHISHU_APP_ID / ZHISHU_APP_SECRET, '
            '或直接配置 ZHISHU_AUTHORIZATION / ZHISHU_BEARER_TOKEN'
        )

    token_url = _env_value('ZHISHU_TOKEN_URL', DEFAULT_TOKEN_URL)
    body = json.dumps({
        'appId': app_id,
        'appSecret': app_secret,
    }, ensure_ascii=False).encode('utf-8')
    req = urllib.request.Request(
        token_url,
        data=body,
        headers={
            'Content-Type': 'application/json; charset=utf-8',
            'Accept': 'application/json',
        },
        method='POST',
    )
    with urllib.request.urlopen(req, timeout=float(_env_value('ZHISHU_TOKEN_TIMEOUT_SECONDS', '30'))) as resp:
        raw = resp.read().decode('utf-8')
    payload = json.loads(raw) if raw else {}
    _api_business_ok(payload)
    token = _extract_token(payload)
    if not token:
        raise RuntimeError(f'智书 token 接口未返回 tenant_access_token: {_summarize_response(payload)}')
    return token


def _tenant_access_token():
    global _TOKEN_CACHE
    if _TOKEN_CACHE:
        return _TOKEN_CACHE
    with _TOKEN_LOCK:
        if not _TOKEN_CACHE:
            _TOKEN_CACHE = _fetch_tenant_access_token()
            print('[智书导入检查] 已获取 tenant_access_token')
    return _TOKEN_CACHE


def _request_headers():
    headers = {
        'Content-Type': 'application/json; charset=utf-8',
        'Accept': 'application/json',
    }
    authorization = os.getenv('ZHISHU_AUTHORIZATION', '').strip()
    bearer = os.getenv('ZHISHU_BEARER_TOKEN', '').strip() or os.getenv('ZHISHU_ACCESS_TOKEN', '').strip()
    cookie = os.getenv('ZHISHU_COOKIE', '').strip()
    user_agent = os.getenv('ZHISHU_USER_AGENT', '').strip()
    if authorization:
        headers['Authorization'] = authorization
    elif bearer:
        headers['Authorization'] = f'Bearer {bearer}'
    else:
        headers['Authorization'] = f'Bearer {_tenant_access_token()}'
    if cookie:
        headers['Cookie'] = cookie
    if user_agent:
        headers['User-Agent'] = user_agent
    return headers


def _summarize_response(payload):
    if isinstance(payload, dict):
        data = payload.get('data')
        if isinstance(data, dict) and isinstance(data.get('items'), list):
            return (
                f'code={payload.get("code")}; msg={payload.get("msg")}; '
                f'data.items={len(data.get("items") or [])}; has_more={data.get("has_more")}'
            )
        if 'code' in payload or 'msg' in payload or 'success' in payload:
            return f'code={payload.get("code")}; success={payload.get("success")}; msg={payload.get("msg")}'
        keys = ','.join(str(key) for key in list(payload.keys())[:8])
        return '{' + keys + '}'
    if isinstance(payload, list):
        return f'list[{len(payload)}]'
    return _text(payload)[:80]


def _api_business_ok(payload):
    if not isinstance(payload, dict) or 'code' not in payload:
        return True
    code = payload.get('code')
    if str(code) in ('0', '200'):
        return True
    if str(code) in NOT_IMPORTED_CODES:
        return True
    message = payload.get('msg') or payload.get('message') or payload.get('error')
    raise RuntimeError(f'智书查询接口返回失败: code={code}, msg={message}')


def _short_error(value, limit=220):
    text = re.sub(r'\s+', ' ', _text(value))
    return text if len(text) <= limit else text[:limit] + '...'


def _format_request_error(exc):
    if isinstance(exc, urllib.error.HTTPError):
        body = ''
        try:
            body = exc.read().decode('utf-8', errors='replace')
        except Exception:
            body = ''
        return _short_error(f'HTTP {exc.code} {exc.reason}: {body}')
    if isinstance(exc, urllib.error.URLError):
        return _short_error(f'URL Error: {exc.reason}')
    return _short_error(str(exc))


def _failure_summary(failed, max_items=3):
    counts = {}
    for reason in failed.values():
        key = _short_error(reason, limit=120)
        counts[key] = counts.get(key, 0) + 1
    return '; '.join(
        f'{count}x {reason}'
        for reason, count in sorted(counts.items(), key=lambda item: item[1], reverse=True)[:max_items]
    )


def _known_total_value(obj):
    if not isinstance(obj, dict):
        return None
    for key in ('total', 'count', 'total_count', 'totalCount'):
        if key not in obj:
            continue
        try:
            return int(obj[key] or 0)
        except (TypeError, ValueError):
            return None
    return None


def _payload_has_contract(payload):
    """智书搜索接口口径:
    - code=0 且 data.items 有数据: 已导入
    - code=110107/未查询到该合同: 未导入
    """
    if not isinstance(payload, dict):
        return False
    if str(payload.get('code')) in NOT_IMPORTED_CODES:
        return False
    data = payload.get('data')
    if not isinstance(data, dict):
        return False
    items = data.get('items')
    return isinstance(items, list) and len(items) > 0


def _contract_status_names(payload):
    if not isinstance(payload, dict):
        return ''
    data = payload.get('data')
    if not isinstance(data, dict):
        return ''
    items = data.get('items')
    if not isinstance(items, list):
        return ''
    names = []
    seen = set()
    for item in items:
        if not isinstance(item, dict):
            continue
        status = _text(item.get('contract_status_name'))
        if status and status not in seen:
            seen.add(status)
            names.append(status)
    return '; '.join(names)


def query_one_contract(contract_number, search_url, timeout, retries):
    payload = {'contract_number': contract_number}
    body = json.dumps(payload, ensure_ascii=False).encode('utf-8')
    headers = _request_headers()
    last_error = None
    started = time.perf_counter()

    for attempt in range(1, retries + 2):
        req = urllib.request.Request(search_url, data=body, headers=headers, method='POST')
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                raw = resp.read().decode('utf-8')
            data = json.loads(raw) if raw else {}
            _api_business_ok(data)
            elapsed_ms = int((time.perf_counter() - started) * 1000)
            return QueryResult(
                contract_number=contract_number,
                imported=_payload_has_contract(data),
                elapsed_ms=elapsed_ms,
                response_summary=_summarize_response(data),
                status_names=_contract_status_names(data),
            )
        except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError, ConnectionError,
                json.JSONDecodeError, RuntimeError) as exc:
            last_error = _format_request_error(exc)
            if attempt <= retries:
                time.sleep(min(1.5 * attempt, 6))

    raise RuntimeError(f'{contract_number} 查询失败: {last_error}')


def probe_contracts(contract_numbers, search_url, timeout, retries):
    numbers = [number for number in (_text(value) for value in contract_numbers) if number]
    if not numbers:
        return
    print(f'[智书导入检查] 探测合同: {", ".join(numbers)}', flush=True)
    for number in numbers:
        result = query_one_contract(number, search_url, timeout, retries)
        print(
            f'[智书导入检查][PROBE] {number}: '
            f'imported={result.imported}; {result.response_summary}',
            flush=True,
        )


def check_imported(
        contract_inputs,
        search_url,
        workers,
        timeout,
        retries,
        assume_not_imported=False,
        progress_interval=10,
        progress_every_done=1,
        debug_sample=5,
        final_retries=2,
        final_retry_delay=15,
        final_retry_workers=2,
):
    if assume_not_imported:
        return {
            item.contract_number: QueryResult(item.contract_number, imported=False, elapsed_ms=0, response_summary='assumed')
            for item in contract_inputs
        }

    results = {}
    failed = {}
    total = len(contract_inputs)
    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_map = {
            executor.submit(query_one_contract, item.contract_number, search_url, timeout, retries): item.contract_number
            for item in contract_inputs
        }
        pending = set(future_map)
        started = time.perf_counter()
        last_report_at = started
        last_report_done = 0

        while pending:
            done, pending = wait(
                pending,
                timeout=max(float(progress_interval), 1.0),
                return_when=FIRST_COMPLETED,
            )
            for future in done:
                contract_number = future_map[future]
                try:
                    results[contract_number] = future.result()
                    if debug_sample > 0 and len(results) <= debug_sample:
                        result = results[contract_number]
                        print(
                            f'[智书导入检查][DEBUG] {contract_number}: '
                            f'imported={result.imported}; {result.response_summary}',
                            flush=True,
                        )
                except Exception as exc:
                    failed[contract_number] = str(exc)

            done_count = len(results) + len(failed)
            imported_count = sum(1 for result in results.values() if result.imported)
            not_imported_count = sum(1 for result in results.values() if not result.imported)
            elapsed = time.perf_counter() - started
            now = time.perf_counter()
            should_report = (
                done_count == total
                or done_count - last_report_done >= max(int(progress_every_done), 1)
                or now - last_report_at >= max(float(progress_interval), 1.0)
                or not done
            )
            if should_report:
                top_failures = _failure_summary(failed, max_items=2) if failed else ''
                print(
                    f'[智书导入检查] 查询进度 {done_count}/{total}; '
                    f'已导入 {imported_count}; 未导入 {not_imported_count}; '
                    f'失败 {len(failed)}; 剩余 {len(pending)}; 耗时 {elapsed:.0f}s'
                    + (f'; 失败原因: {top_failures}' if top_failures else ''),
                    flush=True,
                )
                last_report_done = done_count
                last_report_at = now

    for final_round in range(1, max(int(final_retries), 0) + 1):
        if not failed:
            break

        retry_numbers = list(failed)
        retry_workers = max(1, min(int(final_retry_workers), len(retry_numbers)))
        if final_retry_delay:
            delay = max(float(final_retry_delay), 0)
            print(
                f'[智书导入检查] 收尾重试第 {final_round}/{final_retries} 轮: '
                f'{len(retry_numbers)} 条失败合同,等待 {delay:g}s 后以 {retry_workers} 并发重试',
                flush=True,
            )
            time.sleep(delay)
        else:
            print(
                f'[智书导入检查] 收尾重试第 {final_round}/{final_retries} 轮: '
                f'{len(retry_numbers)} 条失败合同,以 {retry_workers} 并发重试',
                flush=True,
            )

        retry_failed = {}
        with ThreadPoolExecutor(max_workers=retry_workers) as executor:
            future_map = {
                executor.submit(query_one_contract, contract_number, search_url, timeout, retries): contract_number
                for contract_number in retry_numbers
            }
            pending = set(future_map)
            started = time.perf_counter()
            last_report_at = started
            last_report_done = 0

            while pending:
                done, pending = wait(
                    pending,
                    timeout=max(float(progress_interval), 1.0),
                    return_when=FIRST_COMPLETED,
                )
                for future in done:
                    contract_number = future_map[future]
                    try:
                        results[contract_number] = future.result()
                        failed.pop(contract_number, None)
                    except Exception as exc:
                        retry_failed[contract_number] = str(exc)

                done_count = len(retry_numbers) - len(pending)
                elapsed = time.perf_counter() - started
                now = time.perf_counter()
                should_report = (
                    done_count == len(retry_numbers)
                    or done_count - last_report_done >= max(int(progress_every_done), 1)
                    or now - last_report_at >= max(float(progress_interval), 1.0)
                    or not done
                )
                if should_report:
                    top_failures = _failure_summary(retry_failed, max_items=2) if retry_failed else ''
                    print(
                        f'[智书导入检查] 收尾重试进度 {done_count}/{len(retry_numbers)}; '
                        f'本轮仍失败 {len(retry_failed)}; 剩余 {len(pending)}; 耗时 {elapsed:.0f}s'
                        + (f'; 失败原因: {top_failures}' if top_failures else ''),
                        flush=True,
                    )
                    last_report_done = done_count
                    last_report_at = now

        failed.update(retry_failed)
        recovered = len(retry_numbers) - len(retry_failed)
        print(
            f'[智书导入检查] 收尾重试第 {final_round}/{final_retries} 轮完成: '
            f'恢复 {recovered}; 仍失败 {len(failed)}',
            flush=True,
        )

    if failed:
        sample = '; '.join(f'{key}: {value}' for key, value in list(failed.items())[:5])
        summary = _failure_summary(failed, max_items=8)
        raise RuntimeError(f'智书查询失败 {len(failed)} 条; 原因汇总: {summary}; 示例: {sample}')
    return results


def _write_output_rows(worksheet, rows):
    worksheet.append(list(OUTPUT_COLUMNS))
    header_fill = PatternFill('solid', fgColor='E7EAED')
    header_font = Font(name='Microsoft YaHei', bold=True, color='374151')
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row in rows:
        worksheet.append([
            row.contract_number,
            row.related_orders,
            row.contract_type,
            row.project_code,
        ])

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            cell.border = border

    widths = (28, 34, 34, 24)
    for col_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(col_index)].width = width
    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = worksheet.dimensions


def _write_audit_sheet(workbook, contract_inputs, results):
    worksheet = workbook.create_sheet('查询核对')
    headers = ('合同编号', '是否已导入智书', 'contract_status_name', '接口耗时ms', '接口返回摘要', '来源行')
    worksheet.append(headers)
    header_fill = PatternFill('solid', fgColor='E7EAED')
    header_font = Font(name='Microsoft YaHei', bold=True, color='374151')
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for item in contract_inputs:
        result = results[item.contract_number]
        worksheet.append([
            item.contract_number,
            'Y' if result.imported else 'N',
            result.status_names,
            result.elapsed_ms,
            result.response_summary,
            item.source_rows,
        ])

    widths = (28, 14, 24, 12, 60, 40)
    for col_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(col_index)].width = width
    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = worksheet.dimensions
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            cell.border = border


def _write_status_summary_sheet(workbook, results):
    worksheet = workbook.create_sheet('状态统计')
    worksheet.append(['contract_status_name', '合同数'])
    header_fill = PatternFill('solid', fgColor='E7EAED')
    header_font = Font(name='Microsoft YaHei', bold=True, color='374151')
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    counts = {}
    for result in results.values():
        if not result.imported:
            status_values = ['未导入']
        else:
            status_values = _split_multi_values(result.status_names) or ['(空)']
        for status in status_values:
            counts[status] = counts.get(status, 0) + 1

    for status, count in sorted(counts.items(), key=lambda item: (-item[1], item[0])):
        worksheet.append([status, count])

    worksheet.column_dimensions['A'].width = 28
    worksheet.column_dimensions['B'].width = 12
    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = worksheet.dimensions
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center')
            cell.border = border


def write_not_imported_excel(rows, output_file, contract_inputs=None, results=None):
    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = '未导入智书'
    _write_output_rows(ws, rows)
    if contract_inputs is not None and results is not None:
        _write_audit_sheet(wb, contract_inputs, results)
        _write_status_summary_sheet(wb, results)
    wb.save(output_file)
    return output_file


def run_with_options(args):
    contract_inputs = read_contract_inputs(args.input, sheet_name=args.sheet, all_sheets=args.all_sheets)
    if not contract_inputs:
        raise RuntimeError('输入 Excel 未读取到任何合同编号')

    search_url = args.url or _env_value('ZHISHU_SEARCH_URL', DEFAULT_SEARCH_URL)
    workers = args.workers or int(_env_value('ZHISHU_CHECK_WORKERS', '12'))
    timeout = args.timeout or float(_env_value('ZHISHU_CHECK_TIMEOUT_SECONDS', '30'))
    retries = args.retries if args.retries is not None else int(_env_value('ZHISHU_CHECK_RETRIES', '2'))
    progress_interval = float(_env_value('ZHISHU_PROGRESS_INTERVAL_SECONDS', '10'))
    progress_every_done = int(_env_value('ZHISHU_PROGRESS_EVERY_DONE', '1'))
    debug_sample = int(_env_value('ZHISHU_DEBUG_SAMPLE', '5'))
    arg_final_retries = getattr(args, 'final_retries', None)
    arg_final_retry_delay = getattr(args, 'final_retry_delay', None)
    arg_final_retry_workers = getattr(args, 'final_retry_workers', None)
    final_retries = (
        arg_final_retries
        if arg_final_retries is not None
        else int(_env_value('ZHISHU_FINAL_RETRIES', '2'))
    )
    final_retry_delay = (
        arg_final_retry_delay
        if arg_final_retry_delay is not None
        else float(_env_value('ZHISHU_FINAL_RETRY_DELAY_SECONDS', '15'))
    )
    final_retry_workers = (
        arg_final_retry_workers
        if arg_final_retry_workers is not None
        else int(_env_value('ZHISHU_FINAL_RETRY_WORKERS', '2'))
    )

    started = datetime.now()
    print(f'[智书导入检查] 脚本版本: {SCRIPT_VERSION}', flush=True)
    print(f'[智书导入检查] 接口: {search_url}')
    print(f'[智书导入检查] 并发: {workers}, 超时: {timeout}s, 重试: {retries}')
    print(
        f'[智书导入检查] 收尾重试: {final_retries} 轮, '
        f'等待 {final_retry_delay:g}s, 并发 {final_retry_workers}'
    )
    print(f'[智书导入检查] 进度打印: 每完成 {progress_every_done} 条或每 {progress_interval}s')
    probe_values = _split_multi_values(_env_value('ZHISHU_PROBE_CONTRACTS', 'HH-J-2026030012'))
    probe_contracts(probe_values, search_url, timeout, retries)

    results = check_imported(
        contract_inputs,
        search_url=search_url,
        workers=workers,
        timeout=timeout,
        retries=retries,
        assume_not_imported=args.assume_not_imported,
        progress_interval=progress_interval,
        progress_every_done=progress_every_done,
        debug_sample=debug_sample,
        final_retries=final_retries,
        final_retry_delay=final_retry_delay,
        final_retry_workers=final_retry_workers,
    )

    not_imported = [
        item for item in contract_inputs
        if not results[item.contract_number].imported
    ]
    output_file = Path(args.output or _env_value('ZHISHU_CHECK_OUTPUT_FILE', DEFAULT_OUTPUT_FILE))
    write_not_imported_excel(not_imported, output_file, contract_inputs=contract_inputs, results=results)

    fallback_count = sum(1 for item in not_imported if item.type_note)
    status_counts = {}
    for result in results.values():
        status_values = ['未导入'] if not result.imported else (_split_multi_values(result.status_names) or ['(空)'])
        for status in status_values:
            status_counts[status] = status_counts.get(status, 0) + 1
    elapsed = (datetime.now() - started).total_seconds()
    print(f'[智书导入检查] 输入合同: {len(contract_inputs)}')
    print(f'[智书导入检查] 已导入智书: {len(contract_inputs) - len(not_imported)}')
    print(f'[智书导入检查] 未导入智书: {len(not_imported)}')
    print('[智书导入检查] contract_status_name 统计: ' + '; '.join(
        f'{status}={count}'
        for status, count in sorted(status_counts.items(), key=lambda item: (-item[1], item[0]))[:20]
    ))
    print(f'[智书导入检查] 未导入清单类型兜底: {fallback_count}')
    print(f'[智书导入检查] 输出: {output_file} ({elapsed:.1f}s)')
    return output_file


def build_arg_parser():
    parser = argparse.ArgumentParser(description='多线程检查 Excel 合同是否已导入智书,并导出未导入清单。')
    parser.add_argument('--input', default=DEFAULT_INPUT_FILE, help='输入 Excel 文件路径')
    parser.add_argument('--output', default=None, help='输出 Excel 文件路径')
    parser.add_argument('--sheet', default='', help='只读取指定 sheet; 默认只读取第一个 sheet')
    parser.add_argument('--all-sheets', action='store_true', help='扫描全部 sheet; 默认只读取第一个 sheet')
    parser.add_argument('--url', default='', help='智书合同查询接口 URL')
    parser.add_argument('--workers', type=int, default=0, help='并发线程数,默认读取 ZHISHU_CHECK_WORKERS 或 12')
    parser.add_argument('--timeout', type=float, default=0, help='单次请求超时时间秒,默认 30')
    parser.add_argument('--retries', type=int, default=None, help='失败重试次数,默认 2')
    parser.add_argument('--final-retries', type=int, default=None, help='整批结束后对失败合同的收尾重试轮数,默认 2')
    parser.add_argument(
        '--final-retry-delay',
        type=float,
        default=None,
        help='每轮收尾重试前等待秒数,默认 15',
    )
    parser.add_argument(
        '--final-retry-workers',
        type=int,
        default=None,
        help='收尾重试并发数,默认 2',
    )
    parser.add_argument(
        '--assume-not-imported',
        action='store_true',
        help='不请求接口,仅用于本地验证 Excel 读取和输出格式',
    )
    return parser


def run(argv=None):
    if argv is None:
        argv = sys.argv[2:] if len(sys.argv) > 1 and sys.argv[1] == 'check_zhishu_imported' else []
    args = build_arg_parser().parse_args(argv)
    return run_with_options(args)


def main():
    args = build_arg_parser().parse_args()
    run_with_options(args)


if __name__ == '__main__':
    main()
