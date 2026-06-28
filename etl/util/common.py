# -*- coding: utf-8 -*-
"""公共能力:路径配置、数据库只读连接、各类映射、归一化、Excel 读写、行过滤/去重。

所有清洗任务(ap_opening_payment 应付期初,以及将来的应收/预付/预收)都从这里取用,
避免在每个任务里重复写数据库查询和映射逻辑。任务文件本身只关心"过滤 + 字段映射"。

数据库账密从环境变量读取;若项目根有 .env 则自动加载(.env 不提交版本库)。
需要的变量:FW_*(泛微 vspn_xtyy 取工号)、ZT_*(中台库取供应商编码和核算主体编码)。
"""
import os
import re
import json
import atexit
import unicodedata
import html
import http.client
import time
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from sqlalchemy import create_engine
from sqlalchemy.engine import URL

# ============================ 路径 ============================
ROOT      = Path(__file__).resolve().parents[2]
SRC_DIR   = ROOT / 'resources' / 'source'       # 源表(泛微导出)
RULES_DIR = ROOT / 'resources' / 'rules'        # 映射规则
TPL_DIR   = ROOT / 'resources' / 'templates'    # 导入模版
OUT_DIR   = ROOT / 'output'                # 产出
RULE_XLSX = RULES_DIR / '业财项目_数据映射规则.xlsx'
SUPPLIER_VENDOR_MAPPING_JSON = SRC_DIR / 'supplier_vendor_aliases.json'
CUSTOMER_ALIAS_MAPPING_JSON = SRC_DIR / 'customer_aliases.json'
SUPPLIER_VENDOR_NAME_MATCH_JSON = SRC_DIR / 'supplier_vendor_name_matches.json'
PROJECT_ORDER_MAPPING_ENV = 'PROJECT_ORDER_MAPPING_XLSX'
PROJECT_ORDER_MAPPING_XLSX_NAME = '业财项目_项目订单清洗汇总终版.xlsx'
PROJECT_ORDER_MAPPING_SHEETS = {
    '全量项目': '全量项目_清洗后',
    '全量订单': '全量订单主表_清洗后',
    '订单明细': '全量订单明细行表_清洗后',
}
CLEANED_PROJECT_SOURCE_SHEETS = ('赛事专项项目_清洗前', 'MCN专项项目_清洗前')
PROJECT_ORDER_CLEANABLE_COLUMN = '是否可洗流程'
_PROJECT_ORDER_MAPPING_CACHE = None
_CLEANED_PROJECT_MAPPING_CACHE = None
_CLEANABLE_ORDER_INFO_CACHE = None

ATTACHMENT_COOKIE_ENV = 'WEAVER_CONTRACT_ATTACHMENT_COOKIE'
ATTACHMENT_BASE_URL_ENV = 'WEAVER_CONTRACT_ATTACHMENT_BASE_URL'
ATTACHMENT_LOGIN_USERID_ENV = 'WEAVER_CONTRACT_ATTACHMENT_LOGIN_USERID'
ATTACHMENT_AUTHORIZEMODE_ID_ENV = 'WEAVER_CONTRACT_ATTACHMENT_AUTHORIZEMODE_ID'
ATTACHMENT_AUTHORIZEFIELD_ID_ENV = 'WEAVER_CONTRACT_ATTACHMENT_AUTHORIZEFIELD_ID'
ATTACHMENT_DOWNLOAD_ROOT_ENV = 'WEAVER_CONTRACT_ATTACHMENT_DOWNLOAD_ROOT'
ATTACHMENT_DOWNLOAD_ENABLED_ENV = 'WEAVER_CONTRACT_ATTACHMENT_DOWNLOAD_ENABLED'
ATTACHMENT_DOWNLOAD_WORKERS_ENV = 'WEAVER_CONTRACT_ATTACHMENT_DOWNLOAD_WORKERS'
ATTACHMENT_DOWNLOAD_RETRIES_ENV = 'WEAVER_CONTRACT_ATTACHMENT_DOWNLOAD_RETRIES'
DEFAULT_ATTACHMENT_BASE_URL = 'http://oaportal.heroesports.com'
DEFAULT_ATTACHMENT_LOGIN_USERID = '3837'
DEFAULT_ATTACHMENT_AUTHORIZEMODE_ID = '5'
DEFAULT_ATTACHMENT_AUTHORIZEFIELD_ID = '6461'


# ============================ 配置 / 数据库 ============================
def _load_env():
    env = ROOT / '.env'
    if env.exists():
        for line in env.read_text(encoding='utf-8').splitlines():
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                k, v = line.split('=', 1)
                os.environ.setdefault(k.strip(), v.strip())


def _sql_echo_enabled():
    return os.getenv('SQL_ECHO', '').strip() == '1' or os.getenv('DEBUG_SQL', '').strip() == '1'


def _sqlalchemy_echo_enabled():
    return os.getenv('SQLALCHEMY_ECHO', '').strip() == '1'


_ENGINE_CACHE = {}
_SSH_TUNNELS = {}


def _env_bool(name, default=False):
    value = os.getenv(name, '').strip().lower()
    if not value:
        return default
    return value in ('1', 'true', 't', 'y', 'yes', '是')


def _effective_db_prefix(prefix):
    """兼容老代码:设置 HAND_AS_ZT=1 后,所有 ZT 查询改走汉得生产环境。"""
    if prefix == 'ZT' and _env_bool('HAND_AS_ZT', False):
        return 'HAND'
    return prefix


def _close_ssh_tunnels():
    for tunnel in list(_SSH_TUNNELS.values()):
        try:
            tunnel.stop()
        except Exception:
            pass


atexit.register(_close_ssh_tunnels)


def _ssh_tunnel_config(prefix):
    enabled = _env_bool(f'{prefix}_SSH_ENABLED', False) or bool(os.getenv(f'{prefix}_SSH_HOST', '').strip())
    if not enabled:
        return None
    required = {
        'ssh_host': f'{prefix}_SSH_HOST',
        'ssh_port': f'{prefix}_SSH_PORT',
        'ssh_user': f'{prefix}_SSH_USER',
        'ssh_pass': f'{prefix}_SSH_PASS',
    }
    missing = [env_name for env_name in required.values() if not os.getenv(env_name, '').strip()]
    if missing:
        raise RuntimeError(f'缺少 SSH 跳板环境变量: {", ".join(missing)}')
    return {
        'ssh_host': os.environ[required['ssh_host']].strip(),
        'ssh_port': int(os.environ[required['ssh_port']]),
        'ssh_user': os.environ[required['ssh_user']].strip(),
        'ssh_pass': os.environ[required['ssh_pass']],
    }


def _db_config(prefix):
    try:
        return {
            'host': os.environ[f'{prefix}_HOST'].strip(),
            'port': int(os.environ[f'{prefix}_PORT']),
            'username': os.environ[f'{prefix}_USER'].strip(),
            'password': os.environ[f'{prefix}_PASS'],
        }
    except KeyError as e:
        raise RuntimeError(f'缺少数据库环境变量 {e};请在 .env 或环境变量配置 {prefix}_HOST/PORT/USER/PASS') from e


def _db_endpoint(prefix, config):
    tunnel_config = _ssh_tunnel_config(prefix)
    if not tunnel_config:
        return config['host'], config['port']
    try:
        from sshtunnel import SSHTunnelForwarder
    except ImportError as exc:
        raise RuntimeError('使用 SSH 跳板数据库连接需要安装 sshtunnel: pip install sshtunnel') from exc

    tunnel_key = (
        prefix,
        tunnel_config['ssh_host'],
        tunnel_config['ssh_port'],
        tunnel_config['ssh_user'],
        config['host'],
        config['port'],
    )
    tunnel = _SSH_TUNNELS.get(tunnel_key)
    if tunnel is None or not tunnel.is_active:
        tunnel = SSHTunnelForwarder(
            (tunnel_config['ssh_host'], tunnel_config['ssh_port']),
            ssh_username=tunnel_config['ssh_user'],
            ssh_password=tunnel_config['ssh_pass'],
            remote_bind_address=(config['host'], config['port']),
            local_bind_address=('127.0.0.1', 0),
        )
        tunnel.start()
        _SSH_TUNNELS[tunnel_key] = tunnel
        print(
            f'[数据库连接] {prefix} SSH 隧道: '
            f'127.0.0.1:{tunnel.local_bind_port} -> {config["host"]}:{config["port"]}'
        )
    return '127.0.0.1', tunnel.local_bind_port


def _db_engine(prefix, database):
    """按前缀(FW/ZT/HAND)建 SQLAlchemy engine。只跑 SELECT,不写生产库。"""
    prefix = _effective_db_prefix(prefix)
    config = _db_config(prefix)
    host, port = _db_endpoint(prefix, config)

    cache_key = (prefix, database, host, port, _sqlalchemy_echo_enabled())
    if cache_key not in _ENGINE_CACHE:
        url = URL.create(
            'mysql+pymysql',
            username=config['username'],
            password=config['password'],
            host=host,
            port=port,
            database=database,
            query={'charset': 'utf8mb4'},
        )
        _ENGINE_CACHE[cache_key] = create_engine(
            url,
            echo=_sqlalchemy_echo_enabled(),
            pool_pre_ping=True,
            connect_args={'connect_timeout': 20},
        )
    return _ENGINE_CACHE[cache_key]


def render_sql(prefix, database, sql, params=None):
    """渲染成可直接复制到 MySQL 执行的 SQL。仅用于调试输出。"""
    if params is None:
        params = ()
    elif isinstance(params, list):
        params = tuple(params)
    raw_conn = _db_engine(prefix, database).raw_connection()
    try:
        with raw_conn.cursor() as cursor:
            rendered = cursor.mogrify(sql, params)
            return rendered.decode('utf-8') if isinstance(rendered, bytes) else rendered
    finally:
        raw_conn.close()


def query_db(prefix, database, sql, params=None):
    """执行 SELECT 并返回 DataFrame。SQL 可使用 PyMySQL 参数风格(%s / %(name)s)。

    调试 SQL:运行前设置 SQL_ECHO=1 或 DEBUG_SQL=1。
    """
    if params is None:
        params = ()
    elif isinstance(params, list):
        params = tuple(params)
    if _sql_echo_enabled():
        print(f'\n-- SQL [{prefix}.{database}] --')
        print(render_sql(prefix, database, sql, params).strip())
    with _db_engine(prefix, database).connect() as conn:
        result = conn.exec_driver_sql(sql, params)
        return pd.DataFrame(result.fetchall(), columns=result.keys())


_load_env()
OUT_DIR.mkdir(parents=True, exist_ok=True)


def read_fw_field_dictionary(table_name, language_id=7):
    """查询泛微建模表字段含义。
    主表和明细字段都按 workflow_bill.tablename 查;明细字段看返回列 detail_table。"""
    return query_db(
        'FW',
        'vspn_xtyy',
        '''
        SELECT
            f.id AS field_id,
            f.fieldname AS field_name,
            l.labelname AS label_name,
            f.fielddbtype AS field_db_type,
            f.fieldhtmltype AS field_html_type,
            f.type AS field_type,
            f.detailtable AS detail_table,
            f.dsporder AS display_order
        FROM workflow_bill b
        JOIN workflow_billfield f
            ON f.billid = b.id
        LEFT JOIN htmllabelinfo l
            ON l.indexid = f.fieldlabel
           AND l.languageid = %s
        WHERE b.tablename = %s
        ORDER BY f.viewtype, f.detailtable, f.dsporder, f.id
        ''',
        [language_id, table_name],
    )


def validate_fw_fields(table_name, expected_fields, language_id=7):
    """用泛微字段字典校验真实 SQL 字段名/含义。

    expected_fields: {detail_table: {field_name: expected_label}};主表 detail_table 用空字符串。
    """
    field_df = read_fw_field_dictionary(table_name, language_id=language_id).assign(
        detail_table=lambda df: df['detail_table'].fillna('').astype(str),
        label_key=lambda df: df['label_name'].map(normalize_name),
    )
    actual_labels = {
        (row['detail_table'], row['field_name']): row['label_name']
        for _, row in field_df.iterrows()
    }
    problems = []
    for detail_table, fields in expected_fields.items():
        for field_name, expected_label in fields.items():
            actual_label = actual_labels.get((detail_table or '', field_name))
            if actual_label is None:
                problems.append(f'{detail_table or table_name}.{field_name}: 字段不存在')
            elif normalize_name(actual_label) != normalize_name(expected_label):
                problems.append(
                    f'{detail_table or table_name}.{field_name}: 期望含义={expected_label}, 实际含义={actual_label}')
    if problems:
        raise RuntimeError('泛微字段字典校验失败:\n' + '\n'.join(problems))
    return True


def today_suffix():
    """返回产出文件使用的运行日期后缀。"""
    return date.today().strftime('%Y%m%d')


# ============================ 归一化 / 格式化 ============================
# NFKD 折不掉的特殊拉丁字母(无组合记号),显式映射到基础字母。覆盖土耳其语等外文供应商/人名。
_SPECIAL_LETTERS = str.maketrans({
    'ı': 'i', 'İ': 'I', 'ø': 'o', 'Ø': 'O', 'ł': 'l', 'Ł': 'L',
    'đ': 'd', 'Đ': 'D', 'ð': 'd', 'Ð': 'D', 'ħ': 'h', 'ŧ': 't',
    'ß': 'ss', 'æ': 'ae', 'Æ': 'AE', 'œ': 'oe', 'Œ': 'OE',
})
_SQL_FOLD_LETTERS = (
    ('ı', 'i'), ('İ', 'i'), ('ø', 'o'), ('Ø', 'o'), ('ł', 'l'), ('Ł', 'l'),
    ('ß', 'ss'), ('æ', 'ae'), ('Æ', 'ae'), ('œ', 'oe'), ('Œ', 'oe'),
    ('á', 'a'), ('à', 'a'), ('â', 'a'), ('ä', 'a'), ('ã', 'a'), ('å', 'a'),
    ('ç', 'c'), ('é', 'e'), ('è', 'e'), ('ê', 'e'), ('ë', 'e'),
    ('í', 'i'), ('ì', 'i'), ('î', 'i'), ('ï', 'i'),
    ('ñ', 'n'), ('ó', 'o'), ('ò', 'o'), ('ô', 'o'), ('ö', 'o'), ('õ', 'o'),
    ('ú', 'u'), ('ù', 'u'), ('û', 'u'), ('ü', 'u'),
    ('ğ', 'g'), ('ş', 's'),
)


def _fold_accents(text):
    """去掉变音符号:ş->s、ç->c、ö->o、ı->i…,把外文名折成基础拉丁字母(并把全角等兼容字符规整)。"""
    text = text.translate(_SPECIAL_LETTERS)
    decomposed = unicodedata.normalize('NFKD', text)
    return ''.join(ch for ch in decomposed if not unicodedata.combining(ch))


def normalize_name(value):
    """名称归一化(用于按名称匹配):折叠变音符号(土耳其语 ş/ı 等)、去掉所有空格/标点/符号
    (含 & 、,。()（）等)、忽略大小写;只保留字母数字与中文。
    消除主体/供应商/人名在两套系统间的全半角、特殊字符、标点、大小写差异。"""
    text = _fold_accents(str(value).strip())
    # 去掉空白、标点(category P*)、符号(category S*,含 &);保留字母/数字/中文
    text = ''.join(ch for ch in text
                   if not (ch.isspace() or unicodedata.category(ch)[0] in ('P', 'S')))
    return text.casefold()


def sql_normalized_name(column):
    """生成 MySQL 名称归一化表达式,用于 WHERE IN 缩小维表查询范围。
    输出逻辑尽量贴近 normalize_name:小写、折叠常见拉丁变音字符、移除空白和标点。"""
    expr = f'LOWER(COALESCE({column}, ""))'
    for source, target in _SQL_FOLD_LETTERS:
        expr = f"REPLACE({expr}, '{source}', '{target}')"
    return f"REGEXP_REPLACE({expr}, '[[:space:][:punct:]]', '')"


def remove_slashes(value):
    """去掉所有 '/'。费用科目里 '/' 既是层级分隔符又可能是层级名内部字符,两侧统一去掉才能对齐。"""
    return str(value).strip().replace('/', '')


def format_code(value):
    """浮点编码(1000.0)->整数串'1000';空值->''。"""
    text = str(value).strip()
    if text in ('', 'nan', 'None'):
        return ''
    try:
        return str(int(float(text)))
    except ValueError:
        return text


_ISO_DATE_RE = re.compile(r'^(\d{4}-\d{2}-\d{2})')


def format_date(value):
    """日期 -> 'yyyy-mm-dd';空值/非法值(如 0000-00-00)->''。

    多数源值已是 'YYYY-MM-DD' 字符串(char(10)),直接截取避免逐格 pd.to_datetime(慢)。
    """
    if pd.isna(value):
        return ''
    if isinstance(value, str):
        match = _ISO_DATE_RE.match(value.strip())
        if match:
            iso = match.group(1)
            return '' if iso == '0000-00-00' else iso
    parsed = pd.to_datetime(value, errors='coerce')
    return '' if pd.isna(parsed) else parsed.strftime('%Y-%m-%d')


def round_amount(value):
    """金额保留 2 位小数;空值->''。"""
    return '' if pd.isna(value) else round(float(value), 2)


def _ordered_unique(values):
    result = []
    seen = set()
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def clean_codes(values):
    """把一组 ID/编码规整成去重后的字符串列表,用于 IN 查询参数。"""
    result = []
    seen = set()
    for value in values:
        code = format_code(value)
        if code and code not in seen:
            seen.add(code)
            result.append(code)
    return result


def _clean_codes(values):
    return clean_codes(values)


def in_placeholders(values):
    """按参数个数生成 PyMySQL IN 占位符: %s,%s,%s。"""
    return ','.join(['%s'] * len(values))


def _in_placeholders(values):
    return in_placeholders(values)


def clean_text_values(values):
    """把一组文本值去空、去重后保留原始顺序。"""
    result = []
    seen = set()
    for value in values:
        if pd.isna(value):
            continue
        text = str(value).strip()
        if not text or text in ('nan', 'None'):
            continue
        if text not in seen:
            seen.add(text)
            result.append(text)
    return result


def normalized_name_values(values):
    """把一组名称规整成 normalize_name 后的去重列表,用于名称匹配查询。"""
    result = []
    seen = set()
    for value in clean_text_values(values):
        key = normalize_name(value)
        if key and key not in seen:
            seen.add(key)
            result.append(key)
    return result


def parse_browser_ids(value):
    """解析泛微 browser 字段里的 ID 列表,保持原始顺序并去重。"""
    if pd.isna(value):
        return []
    text = str(value).strip()
    if not text or text in ('nan', 'None'):
        return []
    return _ordered_unique(
        format_code(part)
        for part in re.split(r'[,，;；、|\s]+', text)
    )


def clean_fw_select_name(value, language_id=7):
    """清洗泛微 workflow_selectitem.SELECTNAME。

    多语言选项常见格式类似 ``~`~`7 纸质`~`8 Paper`~`~``;优先取中文(7),
    普通纯文本选项则原样返回。
    """
    text = _cell_text(value)
    if not text:
        return ''
    marker = '~`~`'
    legacy_marker = '`~`'

    def repair_legacy_mojibake(label):
        """部分泛微节点名的中文被按 Big5 解码,需按 Big5->GBK 还原。"""
        try:
            repaired = label.encode('big5').decode('gbk')
        except UnicodeError:
            return label
        # 只接受明显还原成常见简体审批节点/动作的结果,避免误伤正常文本。
        if any(token in repaired for token in (
            '归档', '审批', '审核', '法务', '财务', '用印', '上传', '确认', '提交', '查看',
        )):
            return repaired
        return label

    if marker not in text and legacy_marker not in text:
        return text

    def first_legacy_label(label):
        parts = [part.strip('`~ ') for part in label.split(legacy_marker)]
        for part in parts:
            if part.startswith(str(language_id)):
                return part[len(str(language_id)):].strip()
        for part in parts:
            if part and not re.match(r'^\d+\s+', part):
                return part
        return re.sub(r'`~`\d+\s*', '', label).strip('`~ ')

    if marker in text:
        for part in text.split(marker):
            part = part.strip('`~ ')
            if part.startswith(str(language_id)):
                return repair_legacy_mojibake(first_legacy_label(part[len(str(language_id)):].strip()))
        cleaned = re.sub(r'~`~`\d+\s*', '', text)
        cleaned = cleaned.replace('`~`~', '').replace(marker, '').strip('`~ ')
        return repair_legacy_mojibake(first_legacy_label(cleaned))

    # 旧节点名常见格式: 中文`~`8 English`~`9 繁体。language 7 没有显式编号,
    # 因此优先取第一个无编号片段。
    return repair_legacy_mojibake(first_legacy_label(text))


def build_fw_select_option_map(table_name, field_name, language_id=7):
    """泛微建模下拉字段编码 -> 选项中文名。

    table_name 为 workflow_bill.tablename;field_name 为实际数据库字段名。
    """
    field_df = read_fw_field_dictionary(table_name, language_id=language_id)
    matched = field_df[field_df['field_name'] == field_name]
    if matched.empty:
        return {}
    field_ids = clean_codes(matched['field_id'])
    if not field_ids:
        return {}

    option_df = query_db(
        'FW',
        'vspn_xtyy',
        'SELECT fieldid, selectvalue, selectname '
        'FROM workflow_selectitem '
        f'WHERE fieldid IN ({in_placeholders(field_ids)}) '
        'ORDER BY fieldid, listorder, selectvalue',
        field_ids,
    )
    return {
        format_code(row['selectvalue']): clean_fw_select_name(row['selectname'], language_id=language_id)
        for _, row in option_df.iterrows()
        if clean_fw_select_name(row['selectname'], language_id=language_id)
    }


def build_fw_select_option_maps(table_name, field_names, language_id=7):
    """批量读取同一个泛微建模表的多个下拉字段选项。"""
    field_names = list(field_names)
    if not field_names:
        return {}
    field_df = read_fw_field_dictionary(table_name, language_id=language_id)
    field_df = field_df[field_df['field_name'].isin(field_names)]
    if field_df.empty:
        return {field_name: {} for field_name in field_names}

    field_id_to_name = {
        format_code(row['field_id']): row['field_name']
        for _, row in field_df.iterrows()
        if format_code(row['field_id'])
    }
    field_ids = list(field_id_to_name)
    option_df = query_db(
        'FW',
        'vspn_xtyy',
        'SELECT fieldid, selectvalue, selectname '
        'FROM workflow_selectitem '
        f'WHERE fieldid IN ({in_placeholders(field_ids)}) '
        'ORDER BY fieldid, listorder, selectvalue',
        field_ids,
    )
    result = {field_name: {} for field_name in field_names}
    for _, row in option_df.iterrows():
        field_name = field_id_to_name.get(format_code(row['fieldid']))
        option_name = clean_fw_select_name(row['selectname'], language_id=language_id)
        if field_name and option_name:
            result[field_name][format_code(row['selectvalue'])] = option_name
    return result


# ============================ 映射字典 ============================
# 币种 -> ISO 码
CURRENCY_TO_ISO = {
    '人民币': 'CNY', '美元': 'USD', '马来西亚令吉': 'MYR', '泰铢': 'THB', '印尼盾': 'IDR',
    '韩元': 'KRW', '港币': 'HKD', '新加坡元': 'SGD', '沙特里亚尔': 'SAR', '菲律宾比索': 'PHP',
    '欧元': 'EUR', '日元': 'JPY', '英镑': 'GBP', '瑞士法郎': 'CHF',
    '伊拉克第纳尔': 'IQD', '科威特第纳尔': 'KWD', '埃及镑': 'EGP',
}


def to_iso_currency(currency_name):
    text = '' if pd.isna(currency_name) else str(currency_name).strip()
    return CURRENCY_TO_ISO.get(text, text)


def build_employee_code_map():
    """经办人姓名 -> 工号。来源:泛微 vspn_xtyy。
    hrmresource.JOBTITLE 关联 hrmjobtitles.id 后取 hrmjobtitles.JOBTITLENAME;键用 normalize_name 归一化。
    取到占位值 'Default' 也算匹配、不留空(同名有真实工号时优先真实工号)-- 20260614Leo确认。"""
    employee_df = query_db(
        'FW',
        'vspn_xtyy',
        'SELECT r.LASTNAME employee_name, j.JOBTITLENAME employee_code '
        'FROM hrmresource r LEFT JOIN hrmjobtitles j ON r.JOBTITLE=j.id',
    )
    employee_df['key'] = employee_df['employee_name'].map(normalize_name)
    employee_code_map = {}
    for key, employee_codes in employee_df.groupby('key')['employee_code']:
        values = [str(code).strip() for code in employee_codes.dropna().unique()
                  if str(code).strip() not in ('', 'nan')]
        real_codes = [code for code in values if code != 'Default']
        # 取到 Default 也算匹配、不留空;同名有真实工号时优先真实工号 -- Leo确认
        chosen = real_codes[0] if real_codes else (values[0] if values else None)
        if key and chosen:
            employee_code_map[key] = chosen
    return employee_code_map


def build_fw_employee_info_map_for_ids(user_ids):
    """泛微用户ID -> 员工信息。

    uf_dgfktz.jbr 存 hrmresource.id;姓名取 hrmresource.LASTNAME,
    工号沿用现有口径取 hrmjobtitles.JOBTITLENAME。
    """
    user_ids = clean_codes(user_ids)
    if not user_ids:
        return {}
    employee_df = query_db(
        'FW',
        'vspn_xtyy',
        'SELECT r.id, r.LASTNAME employee_name, j.JOBTITLENAME employee_code, '
        'r.LOGINID loginid, r.WORKCODE workcode, r.EMAIL email, r.MOBILE mobile, r.TELEPHONE telephone '
        'FROM hrmresource r LEFT JOIN hrmjobtitles j ON r.JOBTITLE = j.id '
        f'WHERE r.id IN ({in_placeholders(user_ids)})',
        user_ids,
    )
    employee_map = {}
    for _, row in employee_df.iterrows():
        employee_id = format_code(row['id'])
        employee_name = _cell_text(row['employee_name'])
        employee_code = _cell_text(row['employee_code'])
        if employee_id:
            employee_map[employee_id] = {
                'name': employee_name,
                'code': employee_code,
                'loginid': _cell_text(row.get('loginid')),
                'workcode': _cell_text(row.get('workcode')),
                'email': _cell_text(row.get('email')),
                'mobile': _cell_text(row.get('mobile')),
                'telephone': _cell_text(row.get('telephone')),
            }
    return employee_map


def build_applicant_status_map(fw_user_ids, status_by_number, status_by_name_unique):
    """泛微用户ID -> '在职'/'离职'/''(未匹配)。

    泛微 hrmjobtitles.JOBTITLENAME 常为占位'Default', 故两级匹配飞书:
      1) 泛微 JOBTITLENAME 工号直接命中飞书工号;
      2) 泛微姓名(LASTNAME)在飞书唯一命中(重名则留空)。
    status_by_number: 工号 -> enum_name; status_by_name_unique: 唯一姓名 -> enum_name。
    """
    ids = clean_codes(fw_user_ids)
    if not ids or (not status_by_number and not status_by_name_unique):
        return {}
    info = query_db(
        'FW', 'vspn_xtyy',
        'SELECT r.id, r.LASTNAME AS nm, j.JOBTITLENAME AS gonghao '
        'FROM hrmresource r LEFT JOIN hrmjobtitles j ON r.JOBTITLE = j.id '
        f'WHERE r.id IN ({in_placeholders(ids)})',
        ids,
    )

    def _label(status):
        return '在职' if status == 'hired' else '离职'

    result = {}
    for _, row in info.iterrows():
        employee_id = format_code(row['id'])
        if not employee_id:
            continue
        gonghao = _cell_text(row['gonghao'])
        if gonghao in status_by_number:
            result[employee_id] = _label(status_by_number[gonghao])
            continue
        name = _cell_text(row['nm'])
        if name in status_by_name_unique:
            result[employee_id] = _label(status_by_name_unique[name])
            continue
        result[employee_id] = ''
    return result


def build_fw_company_name_map_for_ids(company_ids):
    """泛微公司主体ID -> 公司主体名称。

    uf_dgfktz.gszt 存 uf_gstt.id;名称取 uf_gstt.gsmc。
    """
    company_ids = clean_codes(company_ids)
    if not company_ids:
        return {}
    company_df = query_db(
        'FW',
        'vspn_xtyy',
        'SELECT id, gsmc company_name '
        f'FROM uf_gstt WHERE id IN ({in_placeholders(company_ids)})',
        company_ids,
    )
    return {
        format_code(row['id']): _cell_text(row['company_name'])
        for _, row in company_df.iterrows()
        if _cell_text(row['company_name'])
    }


def build_fw_currency_name_map_for_ids(currency_ids):
    """泛微币种ID -> 币种名称。

    uf_dgfktz.fkbz 存 fnacurrency.id;名称取 fnacurrency.CURRENCYNAME。
    """
    currency_ids = clean_codes(currency_ids)
    if not currency_ids:
        return {}
    currency_df = query_db(
        'FW',
        'vspn_xtyy',
        'SELECT id, CURRENCYNAME currency_name '
        f'FROM fnacurrency WHERE id IN ({in_placeholders(currency_ids)})',
        currency_ids,
    )
    return {
        format_code(row['id']): _cell_text(row['currency_name'])
        for _, row in currency_df.iterrows()
        if _cell_text(row['currency_name'])
    }


def build_fw_department_name_map_for_ids(department_ids):
    """泛微部门ID -> 部门名称。

    uf_xtyykp.sqrbm 等字段存 hrmdepartment.id;名称取 hrmdepartment.DEPARTMENTNAME。
    """
    department_ids = clean_codes(department_ids)
    if not department_ids:
        return {}
    department_df = query_db(
        'FW',
        'vspn_xtyy',
        'SELECT id, DEPARTMENTNAME department_name '
        f'FROM hrmdepartment WHERE id IN ({in_placeholders(department_ids)})',
        department_ids,
    )
    return {
        format_code(row['id']): _cell_text(row['department_name'])
        for _, row in department_df.iterrows()
        if _cell_text(row['department_name'])
    }


def build_fw_customer_name_map_for_ids(customer_values):
    """泛微客户ID -> 客户名称。

    browser.khk 的配置是: select id,khmc,khmc from uf_khgys。
    """
    customer_ids = clean_codes(
        customer_id
        for value in customer_values
        for customer_id in parse_browser_ids(value)
    )
    if not customer_ids:
        return {}
    customer_df = query_db(
        'FW',
        'vspn_xtyy',
        'SELECT id, khmc customer_name '
        'FROM uf_khgys '
        f'WHERE id IN ({in_placeholders(customer_ids)})',
        customer_ids,
    )
    return {
        format_code(row['id']): _cell_text(row['customer_name'])
        for _, row in customer_df.iterrows()
        if _cell_text(row['customer_name'])
    }


def build_fw_contract_code_map_for_ids(contract_values):
    """泛微合同ID -> 合同编号。

    开票表 uf_xtyykp.kpht 字段 (type=161 浏览框) 绑定 browser.xtyy_httz,
    其配置为 select id,htbh,htbh from uf_htsp,故合同 ID 一定落在 uf_htsp(协同运营-合同台账)。
    uf_htsp 查不到的 ID 视为源系统已删除的孤儿合同,留空走异常清单。
    """
    contract_ids = clean_codes(
        contract_id
        for value in contract_values
        for contract_id in parse_browser_ids(value)
    )
    if not contract_ids:
        return {}

    result = {}
    contract_df = query_db(
        'FW',
        'vspn_xtyy',
        'SELECT id, htbh contract_code '
        'FROM uf_htsp '
        f'WHERE id IN ({in_placeholders(contract_ids)})',
        contract_ids,
    )
    for _, row in contract_df.iterrows():
        contract_id = format_code(row['id'])
        contract_code = _cell_text(row['contract_code'])
        if contract_id and contract_code:
            result[contract_id] = contract_code
    return result


def build_fw_budget_subject_path_map_for_ids(subject_ids):
    """泛微预算科目ID -> 预算科目完整路径。

    uf_dgfktz_dt1.yskm 存 fnabudgetfeetype.id;
    fnabudgetfeetype.ALLSUPSUBJECTIDS 存祖先ID链,再按 ID 查名称并拼出路径。
    """
    subject_ids = clean_codes(subject_ids)
    if not subject_ids:
        return {}
    subject_df = query_db(
        'FW',
        'vspn_xtyy',
        'SELECT id, NAME subject_name, ALLSUPSUBJECTIDS ancestor_ids '
        f'FROM fnabudgetfeetype WHERE id IN ({in_placeholders(subject_ids)})',
        subject_ids,
    )
    ancestor_ids = []
    for value in subject_df['ancestor_ids']:
        for part in str(value or '').split(','):
            code = format_code(part)
            if code:
                ancestor_ids.append(code)
    ancestor_ids = clean_codes(ancestor_ids)
    if not ancestor_ids:
        return {
            format_code(row['id']): _cell_text(row['subject_name'])
            for _, row in subject_df.iterrows()
        }

    ancestor_df = query_db(
        'FW',
        'vspn_xtyy',
        'SELECT id, NAME subject_name '
        f'FROM fnabudgetfeetype WHERE id IN ({in_placeholders(ancestor_ids)})',
        ancestor_ids,
    )
    name_by_id = {
        format_code(row['id']): _cell_text(row['subject_name'])
        for _, row in ancestor_df.iterrows()
        if _cell_text(row['subject_name'])
    }
    subject_map = {}
    for _, row in subject_df.iterrows():
        subject_id = format_code(row['id'])
        ids = [format_code(part) for part in str(row['ancestor_ids'] or '').split(',') if format_code(part)]
        subject_map[subject_id] = '/'.join(name_by_id.get(code, '') for code in ids if name_by_id.get(code))
    return subject_map


def build_fw_cost_center_info_map_for_ids(cost_center_values):
    """泛微成本中心(browser.cbzx01)ID -> 成本中心名称/编号。

    各表单成本中心字段(uf_dgfktz.rzdw / uf_yfkxx.cbzx / uf_xtyykp.cbzx /
    uf_plfy.cbzx / uf_xtyynbsz.zrcbzx,zccbzx / uf_lgptfk.cbzx)都指向成本中心
    建模表 uf_cbzx,取 uf_cbzx.mc(成本中心名称)、uf_cbzx.bh(成本中心编号)。
    """
    cost_center_ids = clean_codes(
        cost_center_id
        for value in cost_center_values
        for cost_center_id in parse_browser_ids(value)
    )
    if not cost_center_ids:
        return {}
    cost_center_df = query_db(
        'FW',
        'vspn_xtyy',
        f'SELECT id, bh, mc FROM uf_cbzx WHERE id IN ({in_placeholders(cost_center_ids)})',
        cost_center_ids,
    )
    return {
        format_code(row['id']): {
            'code': _cell_text(row['bh']),
            'name': _cell_text(row['mc']),
        }
        for _, row in cost_center_df.iterrows()
        if _cell_text(row['mc']) or _cell_text(row['bh'])
    }


def build_fw_cost_center_map_for_ids(cost_center_values):
    """泛微成本中心(browser.cbzx01)ID -> 成本中心名称。"""
    return {
        cost_center_id: info.get('name', '')
        for cost_center_id, info in build_fw_cost_center_info_map_for_ids(cost_center_values).items()
        if info.get('name', '')
    }


def build_fw_cost_center_code_map_for_ids(cost_center_values):
    """泛微成本中心(browser.cbzx01)ID -> 成本中心编号。"""
    return {
        cost_center_id: info.get('code', '')
        for cost_center_id, info in build_fw_cost_center_info_map_for_ids(cost_center_values).items()
        if info.get('code', '')
    }


def build_vendor_map():
    """供应商名称 -> 中台供应商编码 vender_code。来源:中台 hfins_base.hfbs_system_vender。
    按 description / taxpayer_name 建键(均 normalize_name 归一化)。"""
    vendor_df = query_db(
        'ZT',
        'hfins_base',
        'SELECT vender_code vendor_code, description vendor_name, taxpayer_name taxpayer_name '
        'FROM hfbs_system_vender',
    )
    vendor_map = {}
    for _, row in vendor_df.iterrows():
        for name in (row['vendor_name'], row['taxpayer_name']):
            key = normalize_name(name)
            if key and key not in ('nan', 'None') and key not in vendor_map:
                vendor_map[key] = str(row['vendor_code']).strip()
    return vendor_map


def _load_json_rows(mapping_file):
    if not mapping_file.exists():
        return []

    with mapping_file.open('r', encoding='utf-8-sig') as f:
        data = json.load(f)
    return list(data.values()) if isinstance(data, dict) else data


def _load_id_alias_mapping(mapping_file, source_keys, log_label, log_prefix=''):
    """读取“源泛微ID -> targetId”归并规则。"""
    rows = _load_json_rows(mapping_file)

    mapping = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        source_id = ''
        for key in source_keys:
            source_id = format_code(row.get(key))
            if source_id:
                break
        target_id = format_code(row.get('targetId') or row.get('target_id'))
        if source_id and target_id:
            mapping[source_id] = target_id
    prefix = f'{log_prefix} ' if log_prefix else ''
    if mapping_file.exists():
        print(f'{prefix}{log_label}: {mapping_file} ({len(mapping)} 条)')
    return mapping


def load_same_supplier_mapping(mapping_file=SUPPLIER_VENDOR_MAPPING_JSON, log_prefix=''):
    """读取“视为同一个供应商”规则:供应商泛微Id -> targetId。"""
    return _load_id_alias_mapping(
        mapping_file,
        ['供应商泛微Id', 'Id', 'id', 'supplierId'],
        '供应商ID归并规则',
        log_prefix,
    )


def load_same_customer_mapping(mapping_file=CUSTOMER_ALIAS_MAPPING_JSON, log_prefix=''):
    """读取“视为同一个客户”规则:客户泛微Id -> targetId。"""
    return _load_id_alias_mapping(
        mapping_file,
        ['客户泛微Id', 'Id', 'id', 'customerId'],
        '客户ID归并规则',
        log_prefix,
    )


def resolve_alias_id(source_id, alias_map):
    """把 JSON 中声明为同一主体的源ID逐级归并到 targetId。"""
    current_id = format_code(source_id)
    seen = set()
    while current_id and current_id in alias_map and current_id not in seen:
        seen.add(current_id)
        current_id = alias_map[current_id]
    return current_id


def resolve_same_supplier_id(supplier_id, same_supplier_map):
    """把 JSON 中声明为同一供应商的源ID归并到 targetId。"""
    return resolve_alias_id(supplier_id, same_supplier_map)


def resolve_same_customer_id(customer_id, same_customer_map):
    """把 JSON 中声明为同一客户的源ID归并到 targetId。"""
    return resolve_alias_id(customer_id, same_customer_map)


def build_fw_supplier_status_map(supplier_values):
    """泛微供应商ID -> 供应商状态。

    browser.gysk 的配置是: select id,khmc,khmc from uf_khgys where zt='0'。
    所以多供应商时优先选 zt=0 的供应商,避开已失效供应商。
    """
    supplier_id_values = clean_codes(
        supplier_id
        for value in supplier_values
        for supplier_id in parse_browser_ids(value)
    )
    if not supplier_id_values:
        return {}

    supplier_df = query_db(
        'FW',
        'vspn_xtyy',
        'SELECT id, khmc supplier_name, zt status_code, rzzt certification_status '
        'FROM uf_khgys '
        f'WHERE id IN ({in_placeholders(supplier_id_values)})',
        supplier_id_values,
    )
    result = {}
    for _, row in supplier_df.iterrows():
        supplier_id = format_code(row['id'])
        if supplier_id:
            result[supplier_id] = {
                'name': '' if pd.isna(row['supplier_name']) else str(row['supplier_name']).strip(),
                'status_code': format_code(row['status_code']),
                'certification_status': format_code(row['certification_status']),
            }
    return result


def choose_fw_supplier_id(supplier_id_values, supplier_status_map):
    """多供应商ID时,优先选择泛微供应商库里未失效的供应商。"""
    ids = _ordered_unique(format_code(value) for value in supplier_id_values)
    if not ids:
        return ''
    if len(ids) == 1:
        return ids[0]

    def score(supplier_id):
        status = supplier_status_map.get(supplier_id, {})
        status_code = status.get('status_code', '')
        certification_status = status.get('certification_status', '')
        return (
            0 if status_code == '0' else 1,
            0 if certification_status == '0' else 1,
            ids.index(supplier_id),
        )

    return min(ids, key=score)


def build_hand_vendor_info_by_ids(target_ids):
    """Hand 供应商ID(vender_id) -> 供应商信息。"""
    target_ids = clean_codes(target_ids)
    if not target_ids:
        return {}

    vendor_df = query_db(
        'ZT',
        'hfins_base',
        'SELECT vender_id, vender_code, description vendor_name, taxpayer_name '
        'FROM hfbs_system_vender '
        f'WHERE vender_id IN ({in_placeholders(target_ids)})',
        target_ids,
    )
    result = {}
    for _, row in vendor_df.iterrows():
        vendor_id = format_code(row['vender_id'])
        vendor_code = '' if pd.isna(row['vender_code']) else str(row['vender_code']).strip()
        vendor_name = '' if pd.isna(row['vendor_name']) else str(row['vendor_name']).strip()
        taxpayer_name = '' if pd.isna(row['taxpayer_name']) else str(row['taxpayer_name']).strip()
        if vendor_id:
            result[vendor_id] = {
                'code': '' if vendor_code in ('nan', 'None') else vendor_code,
                'name': '' if vendor_name in ('nan', 'None') else vendor_name,
                'taxpayer_name': '' if taxpayer_name in ('nan', 'None') else taxpayer_name,
                'match_method': 'supplier_id',
            }
    return result


def build_hand_vendor_info_by_names(names):
    """供应商名称 -> 唯一命中的 Hand 供应商信息。

    按 description / taxpayer_name 归一化精确匹配;如果同名命中多个不同编码则不自动采用。
    """
    keys = normalized_name_values(names)
    if not keys:
        return {}
    description_key = sql_normalized_name('description')
    taxpayer_key = sql_normalized_name('taxpayer_name')
    placeholders = in_placeholders(keys)
    vendor_df = query_db(
        'ZT',
        'hfins_base',
        'SELECT vender_id, vender_code, description vendor_name, taxpayer_name '
        'FROM hfbs_system_vender '
        f'WHERE {description_key} IN ({placeholders}) '
        f'   OR {taxpayer_key} IN ({placeholders})',
        keys + keys,
    )

    grouped = {}
    for _, row in vendor_df.iterrows():
        vendor_id = format_code(row['vender_id'])
        vendor_code = _cell_text(row['vender_code'])
        vendor_name = _cell_text(row['vendor_name'])
        taxpayer_name = _cell_text(row['taxpayer_name'])
        if not vendor_code:
            continue
        info = {
            'id': vendor_id,
            'code': vendor_code,
            'name': vendor_name,
            'taxpayer_name': taxpayer_name,
            'match_method': 'disabled_supplier_name',
        }
        for name in (vendor_name, taxpayer_name):
            key = normalize_name(name)
            if key in keys:
                grouped.setdefault(key, []).append(info)

    result = {}
    for key, candidates in grouped.items():
        by_code = {}
        for candidate in candidates:
            by_code.setdefault(candidate['code'], candidate)
        if len(by_code) == 1:
            item = next(iter(by_code.values())).copy()
            item['candidate_count'] = len(candidates)
            result[key] = item
    return result


def load_supplier_vendor_name_match_map(mapping_file=SUPPLIER_VENDOR_NAME_MATCH_JSON, log_prefix=''):
    """读取自动发现的禁用供应商名称匹配缓存:泛微供应商ID -> Hand 供应商编码。"""
    rows = _load_json_rows(mapping_file)
    result = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        supplier_id = format_code(
            row.get('供应商泛微Id')
            or row.get('supplierId')
            or row.get('id')
        )
        vendor_code = _cell_text(
            row.get('handVendorCode')
            or row.get('vender_code')
            or row.get('vendor_code')
        )
        if supplier_id and vendor_code:
            result[supplier_id] = {
                'code': vendor_code,
                'id': format_code(row.get('handVendorId') or row.get('vender_id') or row.get('targetId')),
                'name': _cell_text(row.get('handVendorName') or row.get('供应商名称')),
                'source_name': _cell_text(row.get('供应商名称')),
                'match_method': _cell_text(row.get('匹配方式')) or 'disabled_supplier_name_cache',
            }
    prefix = f'{log_prefix} ' if log_prefix else ''
    if mapping_file.exists():
        print(f'{prefix}禁用供应商名称匹配缓存: {mapping_file} ({len(result)} 条)')
    return result


def append_supplier_vendor_name_matches(rows, mapping_file=SUPPLIER_VENDOR_NAME_MATCH_JSON):
    """把本次按名称匹配到的禁用供应商写入新的缓存 JSON,供后续复用。"""
    new_rows = []
    for row in rows:
        supplier_id = format_code(row.get('供应商泛微Id') or row.get('supplierId') or row.get('id'))
        vendor_code = _cell_text(row.get('handVendorCode') or row.get('vender_code') or row.get('vendor_code'))
        if supplier_id and vendor_code:
            new_rows.append(row)
    if not new_rows:
        return 0

    existing_rows = _load_json_rows(mapping_file)
    by_supplier = {}
    for row in existing_rows:
        if not isinstance(row, dict):
            continue
        supplier_id = format_code(row.get('供应商泛微Id') or row.get('supplierId') or row.get('id'))
        if supplier_id:
            by_supplier[supplier_id] = row
    changed = 0
    for row in new_rows:
        supplier_id = format_code(row.get('供应商泛微Id') or row.get('supplierId') or row.get('id'))
        old = by_supplier.get(supplier_id)
        if old != row:
            by_supplier[supplier_id] = row
            changed += 1
    if not changed:
        return 0

    mapping_file.parent.mkdir(parents=True, exist_ok=True)
    output_rows = sorted(by_supplier.values(), key=lambda item: format_code(item.get('供应商泛微Id') or item.get('supplierId') or item.get('id')))
    mapping_file.write_text(json.dumps(output_rows, ensure_ascii=False, indent=2), encoding='utf-8')
    return changed


def _series_like(values, index):
    if values is None:
        return pd.Series([''] * len(index), index=index)
    if isinstance(values, pd.Series):
        return values.reindex(index)
    return pd.Series(values, index=index)


def build_supplier_vendor_missing_report(
        supplier_values, supplier_texts, document_numbers, selected_by_index, source_to_target, vendor_by_source_id):
    """生成 Hand 按供应商ID查不到的诊断清单。"""
    supplier_series = supplier_values if isinstance(supplier_values, pd.Series) else pd.Series(supplier_values)
    text_series = _series_like(supplier_texts, supplier_series.index)
    doc_series = _series_like(document_numbers, supplier_series.index)

    grouped = {}
    for index in supplier_series.index:
        source_id = selected_by_index.get(index, '')
        if not source_id or vendor_by_source_id.get(source_id):
            continue
        item = grouped.setdefault(source_id, {
            '泛微供应商ID': source_id,
            '泛微供应商文本': [],
            '单据编号': [],
        })
        supplier_text = _cell_text(text_series.get(index, ''))
        document_number = _cell_text(doc_series.get(index, ''))
        if supplier_text and supplier_text not in item['泛微供应商文本']:
            item['泛微供应商文本'].append(supplier_text)
        if document_number and document_number not in item['单据编号']:
            item['单据编号'].append(document_number)

    rows = []
    for item in grouped.values():
        rows.append({
            '泛微供应商ID': item['泛微供应商ID'],
            '泛微供应商文本': ' | '.join(item['泛微供应商文本'][:3]),
            '单据数': len(item['单据编号']),
            '示例流程编号': ' | '.join(item['单据编号'][:5]),
        })
    report_df = pd.DataFrame(rows, columns=[
        '泛微供应商ID', '泛微供应商文本', '单据数', '示例流程编号',
    ])
    if not report_df.empty:
        report_df = report_df.sort_values(
            '泛微供应商ID',
            key=lambda series: pd.to_numeric(series, errors='coerce').fillna(float('inf')),
        ).reset_index(drop=True)
    return report_df


def build_supplier_vendor_info_map_for_rows(
        supplier_values, supplier_texts=None, document_numbers=None, same_supplier_map=None,
        missing_report_file=None, log_prefix=''):
    """逐行确定 Hand 供应商。

    supplier_values 为泛微 gys 字段序列。多供应商ID时先查泛微供应商库 uf_khgys,
    优先选择 zt=0 的有效供应商;选中的供应商ID再套同供应商归并规则,
    最后用 Hand hfbs_system_vender.vender_id 查编码。

    返回值按传入 Series 的 index 对齐: {index: {'code': ..., 'name': ...}}。
    若传入 missing_report_file,每次额外输出 Hand 按ID查不到的供应商诊断 Excel。
    """
    supplier_series = supplier_values if isinstance(supplier_values, pd.Series) else pd.Series(supplier_values)
    mapping = same_supplier_map if same_supplier_map is not None else load_same_supplier_mapping(log_prefix=log_prefix)
    supplier_status_map = build_fw_supplier_status_map(supplier_series)

    source_to_target = {}
    candidate_ids_by_index = {}
    for index, value in supplier_series.items():
        ids = parse_browser_ids(value)
        selected_id = choose_fw_supplier_id(ids, supplier_status_map)
        candidate_ids = _ordered_unique([selected_id] + ids)
        candidate_ids_by_index[index] = candidate_ids
        for source_id in candidate_ids:
            if source_id and source_id not in source_to_target:
                source_to_target[source_id] = resolve_same_supplier_id(source_id, mapping)

    vendor_by_target_id = build_hand_vendor_info_by_ids(source_to_target.values())
    vendor_by_source_id = {
        source_id: vendor_by_target_id.get(target_id, {})
        for source_id, target_id in source_to_target.items()
    }

    vendor_by_row = {}
    changed_multi_count = 0
    hand_fallback_count = 0
    selected_by_index = {}
    for index, value in supplier_series.items():
        ids = parse_browser_ids(value)
        candidate_ids = candidate_ids_by_index.get(index, [])
        preferred_id = candidate_ids[0] if candidate_ids else ''
        selected_id = next((supplier_id for supplier_id in candidate_ids
                            if vendor_by_source_id.get(supplier_id)), preferred_id)
        selected_by_index[index] = selected_id
        if len(ids) > 1 and selected_id and selected_id != ids[0]:
            changed_multi_count += 1
        if selected_id and preferred_id and selected_id != preferred_id:
            hand_fallback_count += 1
        vendor_by_row[index] = vendor_by_source_id.get(selected_id, {})

    prefix = f'{log_prefix} ' if log_prefix else ''
    print(f'{prefix}多供应商按有效状态改选: {changed_multi_count} 行')
    print(f'{prefix}供应商首选ID未命中后续ID命中: {hand_fallback_count} 行')
    if missing_report_file is not None:
        report_file = Path(missing_report_file)
        report_df = build_supplier_vendor_missing_report(
            supplier_series, supplier_texts, document_numbers,
            selected_by_index, source_to_target, vendor_by_source_id,
        )
        if report_df.empty:
            if report_file.exists():
                try:
                    report_file.unlink()
                    print(f'{prefix}Hand按ID查不到的供应商清单为空，已删除旧文件: {report_file}')
                except PermissionError:
                    print(f'{prefix}Hand按ID查不到的供应商清单为空，但旧文件被占用未删除: {report_file}')
            else:
                print(f'{prefix}Hand按ID查不到的供应商清单为空，不生成文件')
        else:
            report_file.parent.mkdir(parents=True, exist_ok=True)
            report_df.to_excel(report_file, index=False)
            print(f'{prefix}Hand按ID查不到的供应商清单已写出: {report_file} ({len(report_df)} 条)')
    return vendor_by_row


def normalize_bank_account(value):
    """银行账号归一化:去空白,保留字母数字及原始账号字符用于比较。"""
    text = _cell_text(value)
    if text.endswith('.0') and text[:-2].isdigit():
        text = text[:-2]
    return re.sub(r'\s+', '', text)


def build_fw_supplier_bank_account_map_for_ids(bank_account_values):
    """泛微供应商银行账号浏览框 ID -> 银行账号文本。

    对公付款/供应商预付里的银行账号字段经常存 uf_khgys_dt1.id,不能直接写入导入模板。
    """
    bank_account_ids = clean_codes(
        bank_account_id
        for value in bank_account_values
        for bank_account_id in parse_browser_ids(value)
    )
    if not bank_account_ids:
        return {}
    bank_df = query_db(
        'FW',
        'vspn_xtyy',
        'SELECT id, yhzh bank_account '
        'FROM uf_khgys_dt1 '
        f'WHERE id IN ({in_placeholders(bank_account_ids)})',
        bank_account_ids,
    )
    return {
        format_code(row['id']): _cell_text(row['bank_account'])
        for _, row in bank_df.iterrows()
        if _cell_text(row['bank_account'])
    }


def build_hand_vendor_bank_account_info_for_codes(vendor_codes):
    """Hand 供应商编码 -> 银行账号信息。

    来源: hfins_base.hfbs_vender_account。优先默认账户(primary_flag=1),再优先 CNY。
    返回结构:
      {vender_code: {'default': '...', 'accounts': [...], 'normalized': {'...': account_dict}}}
    """
    codes = _ordered_unique(_cell_text(value) for value in vendor_codes)
    codes = [code for code in codes if code]
    if not codes:
        return {}

    account_df = query_db(
        'ZT',
        'hfins_base',
        'SELECT vender_code, bank_account_number, bank_account_name, primary_flag, enabled_flag, '
        '       pk_currtype, bank_code, bank_location_name, account_id, last_update_date '
        'FROM hfbs_vender_account '
        f'WHERE vender_code IN ({in_placeholders(codes)}) '
        "  AND bank_account_number IS NOT NULL AND bank_account_number <> '' "
        "  AND (enabled_flag IS NULL OR enabled_flag = 1 OR enabled_flag = '1') "
        'ORDER BY vender_code, primary_flag DESC, '
        "         CASE WHEN pk_currtype = 'CNY' THEN 0 ELSE 1 END, "
        '         last_update_date DESC, account_id DESC',
        codes,
    )

    result = {}
    for _, row in account_df.iterrows():
        vendor_code = _cell_text(row['vender_code'])
        account_number = _cell_text(row['bank_account_number'])
        normalized = normalize_bank_account(account_number)
        if not vendor_code or not account_number or not normalized:
            continue
        bucket = result.setdefault(vendor_code, {
            'default': '',
            'accounts': [],
            'normalized': {},
        })
        account = {
            'bank_account_number': account_number,
            'bank_account_name': _cell_text(row.get('bank_account_name', '')),
            'primary_flag': format_code(row.get('primary_flag', '')),
            'enabled_flag': format_code(row.get('enabled_flag', '')),
            'currency': _cell_text(row.get('pk_currtype', '')),
            'bank_code': _cell_text(row.get('bank_code', '')),
            'bank_location_name': _cell_text(row.get('bank_location_name', '')),
        }
        bucket['accounts'].append(account)
        bucket['normalized'].setdefault(normalized, account)
        if not bucket['default']:
            bucket['default'] = account_number
    return result


def resolve_hand_vendor_bank_accounts(vendor_codes, source_accounts=None):
    """按收款方供应商编码解析导入银行账号。

    - 源银行账号存在且在该供应商 Hand 银行账户中:使用 Hand 中的规范账号。
    - 源银行账号为空或不属于该供应商:使用该供应商默认银行账号。
    """
    vendor_series = vendor_codes if isinstance(vendor_codes, pd.Series) else pd.Series(vendor_codes)
    source_series = _series_like(source_accounts, vendor_series.index)
    bank_info_by_code = build_hand_vendor_bank_account_info_for_codes(vendor_series)

    resolved = {}
    for index, vendor_code in vendor_series.items():
        vendor_code = _cell_text(vendor_code)
        source_norm = normalize_bank_account(source_series.get(index, ''))
        bank_info = bank_info_by_code.get(vendor_code, {})
        normalized_accounts = bank_info.get('normalized', {})
        if source_norm and source_norm in normalized_accounts:
            resolved[index] = normalized_accounts[source_norm]['bank_account_number']
        else:
            resolved[index] = bank_info.get('default', '')
    return pd.Series(resolved, index=vendor_series.index)


def collect_hand_vendor_bank_account_issues(
        output_df, source_accounts=None, vendor_code_col='收款方编码',
        bank_col='银行账号', doc_col='来源单据编号'):
    """校验输出银行账号是否存在于收款方 Hand 供应商银行卡中。"""
    if vendor_code_col not in output_df.columns or bank_col not in output_df.columns:
        return pd.DataFrame()

    vendor_series = output_df[vendor_code_col]
    source_series = _series_like(source_accounts, output_df.index)
    bank_info_by_code = build_hand_vendor_bank_account_info_for_codes(vendor_series)

    rows = []
    for index, vendor_code in vendor_series.items():
        vendor_code = _cell_text(vendor_code)
        if not vendor_code:
            continue
        source_account = _cell_text(source_series.get(index, ''))
        source_norm = normalize_bank_account(source_account)
        output_account = _cell_text(output_df.at[index, bank_col])
        output_norm = normalize_bank_account(output_account)
        bank_info = bank_info_by_code.get(vendor_code, {})
        normalized_accounts = bank_info.get('normalized', {})
        default_account = bank_info.get('default', '')
        reason = ''
        if not output_norm:
            reason = '收款方在Hand供应商银行卡中未找到可用银行账号'
        elif output_norm not in normalized_accounts:
            reason = '输出银行账号未在该收款方Hand供应商银行卡中找到'
        elif source_norm and source_norm not in normalized_accounts:
            reason = '源银行账号未在该收款方Hand供应商银行卡中找到,已使用默认银行账号'
        if not reason:
            continue
        rows.append({
            doc_col: _cell_text(output_df.at[index, doc_col]) if doc_col in output_df.columns else '',
            vendor_code_col: vendor_code,
            '源银行账号': source_account,
            '输出银行账号': output_account,
            'Hand默认银行账号': default_account,
            'Hand可用银行账号数': len(bank_info.get('accounts', [])),
            '校验结果': reason,
        })

    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows).drop_duplicates().reset_index(drop=True)


def build_customer_map():
    """客户名称 -> 中台客户编码 customer_code。来源:中台 hfins_base.hfbs_system_customer。
    按 description / taxpayer_name 建键(均 normalize_name 归一化)。"""
    customer_df = query_db(
        'ZT',
        'hfins_base',
        'SELECT customer_code, description customer_name, taxpayer_name '
        'FROM hfbs_system_customer',
    )
    customer_map = {}
    for _, row in customer_df.iterrows():
        for name in (row['customer_name'], row['taxpayer_name']):
            key = normalize_name(name)
            if key and key not in ('nan', 'none') and key not in customer_map:
                customer_map[key] = str(row['customer_code']).strip()
    return customer_map


def build_hand_customer_info_by_ids(target_ids):
    """Hand 客户ID(customer_id) -> 客户信息。"""
    target_ids = clean_codes(target_ids)
    if not target_ids:
        return {}

    customer_df = query_db(
        'ZT',
        'hfins_base',
        'SELECT customer_id, customer_code, description customer_name, taxpayer_name '
        'FROM hfbs_system_customer '
        f'WHERE customer_id IN ({in_placeholders(target_ids)})',
        target_ids,
    )
    result = {}
    for _, row in customer_df.iterrows():
        customer_id = format_code(row['customer_id'])
        customer_code = _cell_text(row['customer_code'])
        customer_name = _cell_text(row['customer_name'])
        taxpayer_name = _cell_text(row['taxpayer_name'])
        if customer_id:
            result[customer_id] = {
                'code': customer_code,
                'name': customer_name,
                'taxpayer_name': taxpayer_name,
                'match_method': 'customer_id',
            }
    return result


def build_customer_map_for_names(names):
    """客户名称 -> 中台客户编码。

    只按传入名称缩小查询范围;按 description / taxpayer_name 建键。
    """
    keys = normalized_name_values(names)
    if not keys:
        return {}
    description_key = sql_normalized_name('description')
    taxpayer_key = sql_normalized_name('taxpayer_name')
    placeholders = in_placeholders(keys)
    customer_df = query_db(
        'ZT',
        'hfins_base',
        'SELECT customer_code, description customer_name, taxpayer_name '
        'FROM hfbs_system_customer '
        f'WHERE {description_key} IN ({placeholders}) '
        f'   OR {taxpayer_key} IN ({placeholders})',
        keys + keys,
    )
    customer_map = {}
    for _, row in customer_df.iterrows():
        customer_code = _cell_text(row['customer_code'])
        if not customer_code:
            continue
        for name in (row['customer_name'], row['taxpayer_name']):
            key = normalize_name(name)
            if key and key not in ('nan', 'none') and key not in customer_map:
                customer_map[key] = customer_code
    return customer_map


def build_lov_meaning_map(lov_code):
    """HZero 值集 meaning -> value。用于把业务侧展示值转成汉得编码。"""
    lov_df = query_db(
        'ZT',
        'hzero_platform',
        'SELECT v.value, v.meaning '
        'FROM hpfm_lov l JOIN hpfm_lov_value v ON v.lov_id = l.lov_id '
        'WHERE l.lov_code = %s AND v.enabled_flag = 1 '
        'ORDER BY v.order_seq, v.value',
        [lov_code],
    )
    return {
        str(row['meaning']).strip(): str(row['value']).strip()
        for _, row in lov_df.iterrows()
        if str(row['meaning']).strip() and str(row['meaning']).strip() != 'nan'
    }


def build_tax_type_description_map(preferred_descriptions=None):
    """税率 -> 汉得税率类型描述。preferred_descriptions 可指定每个税率优先使用的 description。
    键统一为小数税率,例如 6% 为 0.06。"""
    tax_df = query_db(
        'ZT',
        'hfins_base',
        'SELECT tax_type_code, description, tax_type_rate, sale_tax_flag, input_tax_flag, enabled_flag '
        'FROM hfbs_tax_type '
        'WHERE enabled_flag = 1',
    )

    tax_df['rate_key'] = pd.to_numeric(tax_df['tax_type_rate'], errors='coerce').round(4)
    description_to_rate = {
        str(row['description']).strip(): float(row['rate_key'])
        for _, row in tax_df.iterrows()
        if pd.notna(row['rate_key']) and str(row['description']).strip()
    }

    tax_map = {}
    for rate, descriptions in (preferred_descriptions or {}).items():
        for description in descriptions:
            if description in description_to_rate:
                tax_map[round(float(rate), 4)] = description
                break

    # 未显式指定的税率,优先取销项税、再取非进项税,最后保底取该税率第一条。
    sort_df = tax_df.assign(
        sale_priority=(tax_df['sale_tax_flag'] == 1).astype(int),
        non_input_priority=(tax_df['input_tax_flag'] == 0).astype(int),
    ).sort_values(['rate_key', 'sale_priority', 'non_input_priority', 'tax_type_code'],
                  ascending=[True, False, False, True])
    for _, row in sort_df.iterrows():
        if pd.isna(row['rate_key']):
            continue
        tax_map.setdefault(round(float(row['rate_key']), 4), str(row['description']).strip())
    return tax_map


def build_accounting_entity_map():
    """公司主体名称 -> 核算主体编号。来源:中台 hfins_base_account.hfac_accounting_entity。
    按 acc_entity_name 建键。"""
    entity_df = query_db(
        'ZT',
        'hfins_base_account',
        'SELECT acc_entity_code, acc_entity_name '
        'FROM hfac_accounting_entity',
    )
    entity_map = {}
    for _, row in entity_df.iterrows():
        code = str(row['acc_entity_code']).strip()
        if not code:
            continue
        key = normalize_name(row.get('acc_entity_name', ''))
        if key and key != 'nan' and key not in entity_map:
            entity_map[key] = code
    return entity_map


def build_accounting_entity_map_for_names(names):
    """公司主体名称 -> Hand 核算主体编号。

    只按传入名称缩小查询范围,避免每个任务都全表读取 hfac_accounting_entity。
    """
    keys = normalized_name_values(names)
    if not keys:
        return {}
    entity_key = sql_normalized_name('acc_entity_name')
    entity_df = query_db(
        'ZT',
        'hfins_base_account',
        'SELECT acc_entity_code, acc_entity_name '
        'FROM hfac_accounting_entity '
        f'WHERE {entity_key} IN ({in_placeholders(keys)})',
        keys,
    )
    entity_map = {}
    for _, row in entity_df.iterrows():
        code = _cell_text(row['acc_entity_code'])
        key = normalize_name(row['acc_entity_name'])
        if key and code and key not in entity_map:
            entity_map[key] = code
    return entity_map


def build_fw_company_map():
    """泛微公司主体ID -> 公司主体名称。来源:泛微 vspn_xtyy.uf_gstt。"""
    company_df = query_db(
        'FW',
        'vspn_xtyy',
        'SELECT id company_id, gsmc company_name '
        'FROM uf_gstt',
    )
    company_map = {}
    for _, row in company_df.iterrows():
        company_id = format_code(row['company_id'])
        company_name = str(row['company_name']).strip()
        if company_id and company_name and company_name != 'nan':
            company_map[company_id] = company_name
    return company_map


def _cell_text(value):
    """Excel 单元格文本归一化: 空值/nan/None -> ''。"""
    if pd.isna(value):
        return ''
    text = str(value).strip()
    return '' if text in ('', 'nan', 'None', 'NaT') else text


def attachment_download_enabled(cookie):
    flag = os.getenv(ATTACHMENT_DOWNLOAD_ENABLED_ENV, '').strip().lower()
    if flag in ('0', 'false', 'n', 'no', '否'):
        return False
    return bool(_cell_text(cookie))


def attachment_download_workers(default=8):
    raw = os.getenv(ATTACHMENT_DOWNLOAD_WORKERS_ENV, '').strip()
    if not raw:
        return default
    try:
        workers = int(raw)
    except ValueError:
        return default
    return max(1, min(workers, 32))


def attachment_download_retries(default=10):
    raw = os.getenv(ATTACHMENT_DOWNLOAD_RETRIES_ENV, '').strip()
    if not raw:
        return default
    try:
        retries = int(raw)
    except ValueError:
        return default
    return max(1, min(retries, 20))


def build_attachment_referer(base_url, imagefileid, docid):
    query = urllib.parse.urlencode({
        'pdfimagefileid': imagefileid,
        'authorizemodeId': os.getenv(ATTACHMENT_AUTHORIZEMODE_ID_ENV, DEFAULT_ATTACHMENT_AUTHORIZEMODE_ID),
        'authorizefieldid': os.getenv(ATTACHMENT_AUTHORIZEFIELD_ID_ENV, DEFAULT_ATTACHMENT_AUTHORIZEFIELD_ID),
        'docisLock': 'false',
        'formmode_authorize': 'formmode_authorize',
        'authorizeformmodebillId': docid,
        'f_weaver_belongto_usertype': '0',
        'f_weaver_belongto_userid': os.getenv(ATTACHMENT_LOGIN_USERID_ENV, DEFAULT_ATTACHMENT_LOGIN_USERID),
        'canDownload': 'true',
        'canPrint': 'true',
    })
    return f'{base_url.rstrip("/")}/docs/pdfview3.x/web/pdfViewer.jsp?&{query}'


def _download_attachment_file(meta, cookie, log_prefix='附件下载'):
    target_path = Path(meta['target_path'])
    if target_path.exists() and target_path.stat().st_size > 0:
        return 'skipped_exists', ''
    target_path.parent.mkdir(parents=True, exist_ok=True)
    base_url = os.getenv(ATTACHMENT_BASE_URL_ENV, DEFAULT_ATTACHMENT_BASE_URL)
    url = f'{base_url.rstrip("/")}/weaver/weaver.file.FileDownload?fileid={meta["imagefileid"]}'
    headers = {
        'Accept': '*/*',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Connection': 'keep-alive',
        'User-Agent': (
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/146.0.0.0 Safari/537.36'
        ),
        'Cookie': cookie,
        'Referer': build_attachment_referer(base_url, meta['imagefileid'], meta['docid']),
    }
    temp_path = target_path.with_suffix(target_path.suffix + '.part')
    max_attempts = attachment_download_retries()
    last_error = ''
    for attempt in range(1, max_attempts + 1):
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=120) as resp:
                data = resp.read()
                final_url = resp.geturl()
                content_type = resp.headers.get('Content-Type', '')
                if 'login' in final_url.lower():
                    raise RuntimeError(f'跳转到登录页: {final_url}')
                if 'text/html' in content_type.lower() and not _cell_text(meta.get('attachment_name')).lower().endswith(('.html', '.htm')):
                    snippet = data[:200].decode('utf-8', errors='ignore')
                    raise RuntimeError(f'返回 HTML,疑似无权限或会话失效: {snippet}')
                with open(temp_path, 'wb') as file:
                    file.write(data)
            os.replace(temp_path, target_path)
            return ('downloaded', '') if attempt == 1 else ('downloaded', f'重试成功: 第{attempt}次')
        except urllib.error.HTTPError as exc:
            last_error = str(exc)
            retryable = exc.code in (429, 500, 502, 503, 504)
        except (http.client.IncompleteRead, urllib.error.URLError, TimeoutError, ConnectionError, OSError) as exc:
            last_error = str(exc)
            retryable = True
        except RuntimeError as exc:
            if temp_path.exists():
                temp_path.unlink(missing_ok=True)
            return 'failed', str(exc)

        if temp_path.exists():
            temp_path.unlink(missing_ok=True)
        if not retryable or attempt >= max_attempts:
            break
        print(
            f'[{log_prefix}] 下载失败将重试({attempt}/{max_attempts}): '
            f'{_cell_text(meta.get("contract_number（合同编码）"))} '
            f'{_cell_text(meta.get("attachment_name"))}; {last_error}',
            flush=True,
        )
        time.sleep(min(2 * attempt, 10))
    return 'failed', last_error


def download_attachment_manifest(manifest_df, cookie, log_prefix='附件下载'):
    """多线程下载附件,并把 status/error 写回 manifest_df 副本。"""
    if manifest_df.empty:
        return manifest_df.copy()
    result_df = manifest_df.copy()
    workers = attachment_download_workers()
    print(f'[{log_prefix}] 使用 {workers} 个下载线程')
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {
            executor.submit(_download_attachment_file, meta, cookie, log_prefix): index
            for index, meta in result_df.iterrows()
        }
        total = len(futures)
        for done_count, future in enumerate(as_completed(futures), 1):
            index = futures[future]
            try:
                status, error = future.result()
            except Exception as exc:  # 防御单个任务异常,不让整体下载中断。
                status, error = 'failed', str(exc)
            result_df.at[index, 'status'] = status
            result_df.at[index, 'error'] = error
            if error and (status == 'failed' or '重试成功' in error):
                meta = result_df.loc[index]
                print(
                    f'[{log_prefix}] {status}: '
                    f'{_cell_text(meta.get("contract_number（合同编码）"))} '
                    f'{_cell_text(meta.get("attachment_name"))}; {error}',
                    flush=True,
                )
            if done_count % 50 == 0 or done_count == total:
                print(f'[{log_prefix}] 下载进度: {done_count}/{total}')
    return result_df


def download_attachment_manifest_16_workers(manifest_df, cookie, log_prefix='附件下载'):
    old_workers = os.environ.get(ATTACHMENT_DOWNLOAD_WORKERS_ENV)
    os.environ[ATTACHMENT_DOWNLOAD_WORKERS_ENV] = '16'
    try:
        return download_attachment_manifest(manifest_df, cookie, log_prefix=log_prefix)
    finally:
        if old_workers is None:
            os.environ.pop(ATTACHMENT_DOWNLOAD_WORKERS_ENV, None)
        else:
            os.environ[ATTACHMENT_DOWNLOAD_WORKERS_ENV] = old_workers


def split_fanwei_project_codes(value):
    """拆分清洗表里的「原泛微项目编码」;同一格可能维护多个泛微项目编号。"""
    text = _cell_text(value)
    if not text:
        return []
    return [
        item.strip()
        for item in re.split(r'[;；,，\n\r]+', text)
        if item.strip()
    ]


def _find_project_order_mapping_file():
    configured = os.getenv(PROJECT_ORDER_MAPPING_ENV, '').strip()
    if configured:
        path = Path(configured)
        if path.exists():
            return path
        print(f'[项目订单映射] {PROJECT_ORDER_MAPPING_ENV} 指向的文件不存在:', configured)

    search_dirs = [
        SRC_DIR / 'other_cleaned_data',
        SRC_DIR / 'project_order',
        SRC_DIR,
        Path.home() / 'Downloads',
    ]
    for search_dir in search_dirs:
        path = search_dir / PROJECT_ORDER_MAPPING_XLSX_NAME
        if path.exists():
            return path
    return None


def _dedupe_headers(headers):
    seen = {}
    deduped_headers = []
    for header in headers:
        header = header or '未命名'
        if header in seen:
            seen[header] += 1
            header = f'{header}.{seen[header]}'
        else:
            seen[header] = 0
        deduped_headers.append(header)
    return deduped_headers


def _read_cleaned_order_sheet(path, sheet_name, required_columns):
    raw_df = pd.read_excel(path, sheet_name=sheet_name, dtype=str, keep_default_na=False, header=None)
    required = set(required_columns)

    header_index = None
    headers = []
    for index, row in raw_df.iterrows():
        normalized = [_cell_text(value).replace('\n', '') for value in row.tolist()]
        if required.issubset(set(normalized)):
            header_index = index
            headers = normalized
            break
    if header_index is None:
        raise ValueError(f'{sheet_name} 未找到表头列: {sorted(required)}')

    df = raw_df.iloc[header_index + 1:].copy()
    df.columns = _dedupe_headers(headers)
    return df


def _find_cleaned_column(columns, *patterns):
    for column in columns:
        text = _cell_text(column).replace('\n', '')
        lowered = text.lower()
        if any(pattern in text or pattern in lowered for pattern in patterns):
            return column
    return None


def _project_name_key(value):
    text = html.unescape(_cell_text(value))
    text = re.sub(r'<[^>]+>', ' ', text)
    return normalize_name(text)


def _read_cleaned_project_source_rows(mapping_file):
    rows = []
    for sheet_name in CLEANED_PROJECT_SOURCE_SHEETS:
        try:
            df = pd.read_excel(mapping_file, sheet_name=sheet_name, dtype=str, keep_default_na=False)
        except ValueError:
            continue
        id_col = _find_cleaned_column(df.columns, 'id')
        code_col = _find_cleaned_column(df.columns, 'prj_dim_value', '项目编号')
        name_col = _find_cleaned_column(df.columns, 'project_name', '项目名称')
        if id_col is None or code_col is None or name_col is None:
            continue
        for _, row in df.iterrows():
            project_id = format_code(row.get(id_col))
            project_code = _cell_text(row.get(code_col))
            project_name = _cell_text(row.get(name_col))
            if not (project_id and project_code):
                continue
            rows.append({
                '泛微项目ID': project_id,
                '项目编号': project_code,
                '项目名称': project_name,
                '项目名称key': _project_name_key(project_name),
                '映射来源': sheet_name,
            })
    return rows


def load_cleaned_project_mapping():
    """读取项目清洗前表,支持按 合同所属项目编号ID+项目名称 定位清洗后项目编号。"""
    global _CLEANED_PROJECT_MAPPING_CACHE
    if _CLEANED_PROJECT_MAPPING_CACHE is not None:
        return _CLEANED_PROJECT_MAPPING_CACHE

    mapping_file = _find_project_order_mapping_file()
    if mapping_file is None:
        _CLEANED_PROJECT_MAPPING_CACHE = ({}, {}, {}, {}, None)
        return _CLEANED_PROJECT_MAPPING_CACHE

    rows = _read_cleaned_project_source_rows(mapping_file)
    by_id_name = {}
    by_id = {}
    by_code = {}
    for row in rows:
        project_id = row['泛微项目ID']
        project_name_key = row['项目名称key']
        project_code = row['项目编号']
        if project_name_key:
            by_id_name.setdefault((project_id, project_name_key), row)
        by_id.setdefault(project_id, []).append(row)
        by_code.setdefault(project_code, row)

    by_unique_id = {}
    for project_id, items in by_id.items():
        project_codes = {_cell_text(item['项目编号']) for item in items if _cell_text(item['项目编号'])}
        if len(project_codes) == 1:
            by_unique_id[project_id] = items[0]

    print(
        '[项目清洗映射] 使用:',
        mapping_file,
        '| 清洗前项目记录数:', len(rows),
        '| ID+名称映射数:', len(by_id_name),
        '| 唯一ID映射数:', len(by_unique_id),
    )
    _CLEANED_PROJECT_MAPPING_CACHE = (by_id_name, by_id, by_unique_id, by_code, mapping_file)
    return _CLEANED_PROJECT_MAPPING_CACHE


def cleaned_project_mapping(project_id_value, project_name='', project_code=''):
    by_id_name, by_id, by_unique_id, by_code, _ = load_cleaned_project_mapping()
    project_name_key = _project_name_key(project_name)
    project_ids = parse_browser_ids(project_id_value) or [format_code(project_id_value)]
    for project_id in project_ids:
        project_id = format_code(project_id)
        if project_id and project_name_key:
            mapped = by_id_name.get((project_id, project_name_key))
            if mapped:
                return mapped
    if project_name_key:
        for project_id in project_ids:
            for mapped in by_id.get(format_code(project_id), []):
                mapped_name_key = mapped.get('项目名称key', '')
                if mapped_name_key and (project_name_key in mapped_name_key or mapped_name_key in project_name_key):
                    return mapped
    for project_code_item in split_fanwei_project_codes(project_code):
        mapped = by_code.get(project_code_item)
        if mapped:
            return mapped
    for project_id in project_ids:
        mapped = by_unique_id.get(format_code(project_id))
        if mapped:
            return mapped
    return {}


def _filter_cleanable_order_rows(order_df):
    """0621 订单主表口径:仅处理“是否可洗流程”包含 Y 的订单行。"""
    if PROJECT_ORDER_CLEANABLE_COLUMN not in order_df.columns:
        print(f'[项目订单映射] {PROJECT_ORDER_MAPPING_SHEETS["全量订单"]} 未找到 {PROJECT_ORDER_CLEANABLE_COLUMN} 列,暂不按可洗流程过滤。')
        return order_df
    flag = order_df[PROJECT_ORDER_CLEANABLE_COLUMN].map(_cell_text).str.upper()
    return order_df[flag.str.contains('Y', na=False)].copy()


def split_person_names(value):
    """拆分清洗表中的人员姓名列表,支持中文顿号/逗号/分号/换行。"""
    text = _cell_text(value)
    if not text:
        return []
    return [
        item.strip()
        for item in re.split(r'[;；,，、/\n\r]+', text)
        if item.strip()
    ]


def _merge_cleanable_order_info(target, incoming):
    for field in ('订单编号', '订单标题', '映射来源'):
        if not target.get(field) and incoming.get(field):
            target[field] = incoming[field]
    for field in ('项目经理A', '项目经理B'):
        if not target.get(field) and incoming.get(field):
            target[field] = incoming[field]
    manager_names = target.setdefault('项目经理候选', [])
    for name in incoming.get('项目经理候选', []):
        if name and name not in manager_names:
            manager_names.append(name)
    project_codes = target.setdefault('项目编号候选', [])
    for code in incoming.get('项目编号候选', []):
        if code and code not in project_codes:
            project_codes.append(code)
    return target


def _copy_cleanable_order_info(info):
    copied = dict(info)
    copied['项目编号候选'] = list(info.get('项目编号候选', []))
    copied['项目经理候选'] = list(info.get('项目经理候选', []))
    return copied


def load_cleanable_order_info():
    """读取清洗终版中标记 Y 的订单主表,返回 (订单号映射, 项目号映射, 文件路径)。"""
    global _CLEANABLE_ORDER_INFO_CACHE
    if _CLEANABLE_ORDER_INFO_CACHE is not None:
        return _CLEANABLE_ORDER_INFO_CACHE

    mapping_file = _find_project_order_mapping_file()
    if mapping_file is None:
        print('[项目订单映射] 未找到项目&订单清洗后 Excel,无法按可洗订单范围过滤合同。')
        _CLEANABLE_ORDER_INFO_CACHE = ({}, {}, None)
        return _CLEANABLE_ORDER_INFO_CACHE

    try:
        order_df = _read_cleaned_order_sheet(
            mapping_file,
            PROJECT_ORDER_MAPPING_SHEETS['全量订单'],
            ['原泛微项目编码', '订单编号', '订单标题', PROJECT_ORDER_CLEANABLE_COLUMN],
        )
    except ValueError as error:
        print(f'[项目订单映射] 读取可洗订单失败: {error}')
        _CLEANABLE_ORDER_INFO_CACHE = ({}, {}, mapping_file)
        return _CLEANABLE_ORDER_INFO_CACHE

    before_count = len(order_df)
    order_df = _filter_cleanable_order_rows(order_df)
    by_order = {}
    by_project = {}
    for _, row in order_df.iterrows():
        order_code = _cell_text(row.get('订单编号'))
        if not order_code:
            continue
        project_codes = []
        for column in ('原泛微项目编码', '项目编号', '来源项目编码'):
            for project_code in split_fanwei_project_codes(row.get(column, '')):
                if project_code and project_code not in project_codes:
                    project_codes.append(project_code)
        manager_a = _cell_text(row.get('项目经理A'))
        manager_b = _cell_text(row.get('项目经理B'))
        manager_names = []
        for value in (manager_a, manager_b):
            for name in split_person_names(value):
                if name not in manager_names:
                    manager_names.append(name)
        info = {
            '订单编号': order_code,
            '订单标题': _cell_text(row.get('订单标题')),
            '项目编号候选': project_codes,
            '项目经理A': manager_a,
            '项目经理B': manager_b,
            '项目经理候选': manager_names,
            '映射来源': PROJECT_ORDER_MAPPING_SHEETS['全量订单'],
        }
        if order_code in by_order:
            _merge_cleanable_order_info(by_order[order_code], info)
        else:
            by_order[order_code] = _copy_cleanable_order_info(info)
        for project_code in project_codes:
            project_items = by_project.setdefault(project_code, [])
            existing = next((item for item in project_items if item.get('订单编号') == order_code), None)
            if existing:
                _merge_cleanable_order_info(existing, info)
            else:
                project_items.append(_copy_cleanable_order_info(info))

    print(
        '[项目订单映射] 可洗订单主表:',
        f'{len(order_df)}/{before_count} 行,',
        f'订单 {len(by_order)} 个,',
        f'项目 {len(by_project)} 个',
    )
    _CLEANABLE_ORDER_INFO_CACHE = (by_order, by_project, mapping_file)
    return _CLEANABLE_ORDER_INFO_CACHE


def cleanable_order_info_for_order(order_code):
    by_order, _, _ = load_cleanable_order_info()
    return by_order.get(_cell_text(order_code), {})


def cleanable_order_infos_for_project(project_code):
    _, by_project, _ = load_cleanable_order_info()
    return by_project.get(_cell_text(project_code), [])


def _empty_order_presence_df():
    return pd.DataFrame(columns=['泛微项目编号', '订单编号', '订单标题', '映射来源'])


def _empty_order_candidates_df():
    return pd.DataFrame(columns=['泛微项目编号', '订单编号', '订单标题', '项目编号', '项目名称', '映射来源'])


def _mcn_order_project_code(value):
    text = _cell_text(value)
    match = re.match(r'^(.+)-\d{3,}$', text)
    return match.group(1) if match else ''


def _build_full_project_lookup(mapping_file):
    """读取 0619 新版全量项目表,建立 新项目编码 -> 原泛微项目编码列表 的映射。"""
    project_df = _read_cleaned_order_sheet(
        mapping_file,
        PROJECT_ORDER_MAPPING_SHEETS['全量项目'],
        ['原泛微项目编码', '项目编码', '项目名称'],
    )

    rows = []
    project_lookup = {}
    seen_lookup = set()
    for _, row in project_df.iterrows():
        fanwei_projects = split_fanwei_project_codes(row.get('原泛微项目编码', ''))
        project_code = _cell_text(row.get('项目编码', ''))
        project_name = _cell_text(row.get('项目名称', ''))
        if not fanwei_projects:
            continue

        for fanwei_project in fanwei_projects:
            rows.append({
                '泛微项目编号': fanwei_project,
                '订单编号': '',
                '订单标题': '',
                '映射来源': PROJECT_ORDER_MAPPING_SHEETS['全量项目'],
            })

            if not project_code:
                continue
            lookup_key = (project_code, fanwei_project)
            if lookup_key in seen_lookup:
                continue
            seen_lookup.add(lookup_key)
            project_lookup.setdefault(project_code, []).append({
                '泛微项目编号': fanwei_project,
                '项目编号': project_code,
                '项目名称': project_name,
            })

    presence_df = pd.DataFrame(rows, columns=['泛微项目编号', '订单编号', '订单标题', '映射来源'])
    return project_lookup, presence_df.drop_duplicates()


def _order_project_items(order_row, project_lookup):
    """取订单行关联的所有原泛微项目编号。

    0619 全量订单表自身带「原泛微项目编码」;如果为空,用「项目编号」去全量项目表反查。
    「原泛微项目编码」可在同一格维护多个编码,需要拆分后逐个保留。
    如果一个新项目编码对应多个原泛微项目编码,这些原编码都保留为订单候选。
    """
    items = []
    seen_projects = set()
    order_project_code = _cell_text(order_row.get('项目编号', ''))
    order_project_name = _cell_text(order_row.get('项目名称', ''))
    direct_fanwei_projects = split_fanwei_project_codes(order_row.get('原泛微项目编码', ''))

    def add_item(fanwei_project, project_code, project_name, source):
        fanwei_project = _cell_text(fanwei_project)
        if not fanwei_project or fanwei_project in seen_projects:
            return
        seen_projects.add(fanwei_project)
        items.append({
            '泛微项目编号': fanwei_project,
            '项目编号': _cell_text(project_code) or order_project_code,
            '项目名称': _cell_text(project_name) or order_project_name,
            '映射来源': source,
        })

    for direct_fanwei_project in direct_fanwei_projects:
        add_item(
            direct_fanwei_project,
            order_project_code,
            order_project_name,
            PROJECT_ORDER_MAPPING_SHEETS['全量订单'],
        )
    if not direct_fanwei_projects:
        for project_item in project_lookup.get(order_project_code, []):
            add_item(
                project_item.get('泛微项目编号', ''),
                project_item.get('项目编号', ''),
                order_project_name or project_item.get('项目名称', ''),
                f"{PROJECT_ORDER_MAPPING_SHEETS['全量项目']}+{PROJECT_ORDER_MAPPING_SHEETS['全量订单']}",
            )

    return items


def _build_project_order_presence(mapping_file):
    """读取新版全量项目+全量订单表,汇总所有出现过的泛微项目编号。"""
    project_lookup, project_presence_df = _build_full_project_lookup(mapping_file)
    order_df = _read_cleaned_order_sheet(
        mapping_file,
        PROJECT_ORDER_MAPPING_SHEETS['全量订单'],
        ['原泛微项目编码', '订单编号', '订单标题', '项目编号'],
    )
    order_df = _filter_cleanable_order_rows(order_df)

    rows = project_presence_df.to_dict('records')
    for _, order_row in order_df.iterrows():
        for project_item in _order_project_items(order_row, project_lookup):
            rows.append({
                '泛微项目编号': project_item['泛微项目编号'],
                '订单编号': _cell_text(order_row.get('订单编号', '')),
                '订单标题': _cell_text(order_row.get('订单标题', '')),
                '映射来源': project_item['映射来源'],
            })

    detail_df = _build_order_detail_candidates(mapping_file)
    for _, detail_row in detail_df.iterrows():
        rows.append({
            '泛微项目编号': detail_row['泛微项目编号'],
            '订单编号': detail_row['订单编号'],
            '订单标题': detail_row['订单标题'],
            '映射来源': detail_row['映射来源'],
        })

    if not rows:
        return _empty_order_presence_df()
    return pd.DataFrame(rows, columns=['泛微项目编号', '订单编号', '订单标题', '映射来源']).drop_duplicates()


def _build_order_detail_candidates(mapping_file):
    try:
        detail_df = _read_cleaned_order_sheet(
            mapping_file,
            PROJECT_ORDER_MAPPING_SHEETS['订单明细'],
            ['原泛微项目编码', '原泛微MCN订单编码', '订单编号'],
        )
    except ValueError:
        return _empty_order_candidates_df()

    rows = []
    for _, detail_row in detail_df.iterrows():
        order_code = _cell_text(detail_row.get('订单编号', ''))
        if not order_code:
            continue
        mcn_order_code = _cell_text(detail_row.get('原泛微MCN订单编码', ''))
        project_keys = split_fanwei_project_codes(detail_row.get('原泛微项目编码', ''))
        for key in (mcn_order_code, _mcn_order_project_code(mcn_order_code)):
            if key and key not in project_keys:
                project_keys.append(key)
        order_title = _cell_text(detail_row.get('合作内容', '')) or _cell_text(detail_row.get('订单内容', ''))
        project_name = _cell_text(detail_row.get('开票内容', ''))
        for project_key in project_keys:
            rows.append({
                '泛微项目编号': project_key,
                '订单编号': order_code,
                '订单标题': order_title,
                '项目编号': _mcn_order_project_code(mcn_order_code) or project_key,
                '项目名称': project_name,
                '映射来源': PROJECT_ORDER_MAPPING_SHEETS['订单明细'],
            })

    if not rows:
        return _empty_order_candidates_df()
    return pd.DataFrame(
        rows,
        columns=['泛微项目编号', '订单编号', '订单标题', '项目编号', '项目名称', '映射来源'],
    ).drop_duplicates()


def _build_project_order_candidates(mapping_file):
    project_lookup, _ = _build_full_project_lookup(mapping_file)
    order_df = _read_cleaned_order_sheet(
        mapping_file,
        PROJECT_ORDER_MAPPING_SHEETS['全量订单'],
        ['原泛微项目编码', '订单编号', '订单标题', '项目编号', '项目名称'],
    )
    before_count = len(order_df)
    order_df = _filter_cleanable_order_rows(order_df)

    rows = []
    for _, order_row in order_df.iterrows():
        order_code = _cell_text(order_row.get('订单编号', ''))
        if not order_code:
            continue
        for project_item in _order_project_items(order_row, project_lookup):
            rows.append({
                '泛微项目编号': project_item['泛微项目编号'],
                '订单编号': order_code,
                '订单标题': _cell_text(order_row.get('订单标题', '')),
                '项目编号': project_item['项目编号'],
                '项目名称': project_item['项目名称'],
                '映射来源': project_item['映射来源'],
            })

    if not rows:
        result_df = _empty_order_candidates_df()
    else:
        result_df = pd.DataFrame(
            rows,
            columns=['泛微项目编号', '订单编号', '订单标题', '项目编号', '项目名称', '映射来源'],
        )
    detail_df = _build_order_detail_candidates(mapping_file)
    result_df = pd.concat([result_df, detail_df], ignore_index=True).drop_duplicates()
    print(f'[项目订单映射] 可洗订单过滤: {len(order_df)}/{before_count} 行')
    return result_df[
        ['泛微项目编号', '订单编号', '订单标题', '项目编号', '项目名称', '映射来源']
    ]


def load_project_order_mapping():
    """读取项目&订单清洗后的 Excel,返回一对一映射和一对多候选。"""
    global _PROJECT_ORDER_MAPPING_CACHE
    if _PROJECT_ORDER_MAPPING_CACHE is not None:
        return _PROJECT_ORDER_MAPPING_CACHE

    mapping_file = _find_project_order_mapping_file()
    if mapping_file is None:
        print('[项目订单映射] 未找到项目&订单清洗后 Excel,订单字段保持为空。')
        _PROJECT_ORDER_MAPPING_CACHE = ({}, {}, None)
        return _PROJECT_ORDER_MAPPING_CACHE

    mapping_df = _build_project_order_candidates(mapping_file)

    safe_map = {}
    ambiguous_map = {}
    for project_code, group in mapping_df.groupby('泛微项目编号', sort=False):
        orders = group['订单编号'].drop_duplicates()
        if len(orders) == 1:
            row = group.iloc[0]
            safe_map[project_code] = {
                '订单编号': row['订单编号'],
                '订单标题': row['订单标题'],
                '项目编号': row['项目编号'],
                '项目名称': row['项目名称'],
                '映射来源': row['映射来源'],
            }
        else:
            ambiguous_map[project_code] = group.to_dict('records')

    print(
        '[项目订单映射] 使用:',
        mapping_file,
        '| 候选记录数:', len(mapping_df),
        '| 一对一项目数:', len(safe_map),
        '| 多候选项目数:', len(ambiguous_map),
    )
    _PROJECT_ORDER_MAPPING_CACHE = (safe_map, ambiguous_map, mapping_file)
    return _PROJECT_ORDER_MAPPING_CACHE


def project_order_mapping_value(project_code, field):
    safe_map, _, _ = load_project_order_mapping()
    return safe_map.get(_cell_text(project_code), {}).get(field, '')


def collect_order_mapping_issues(
        source_df, doc_col='流程编号', project_col='项目编号',
        project_id_col='项目编号ID', project_name_col='项目名称'):
    """输出项目->订单映射的未匹配和多候选清单。"""
    safe_map, ambiguous_map, mapping_file = load_project_order_mapping()
    if doc_col not in source_df.columns or project_col not in source_df.columns:
        return {}
    issue_columns = {
        '来源单据编号': source_df[doc_col].map(_cell_text),
        '泛微项目编号': source_df[project_col].map(_cell_text),
    }
    if project_id_col in source_df.columns:
        issue_columns['泛微项目ID'] = source_df[project_id_col].map(_cell_text)
    if project_name_col in source_df.columns:
        issue_columns['泛微项目名称'] = source_df[project_name_col].map(_cell_text)
    rows = pd.DataFrame(issue_columns).drop_duplicates()
    rows = rows[rows['泛微项目编号'] != '']

    sheets = {}
    if mapping_file is None:
        sheets['订单映射_文件缺失'] = pd.DataFrame([{
            '说明': (
                f'未找到 {PROJECT_ORDER_MAPPING_XLSX_NAME},'
                f'可通过环境变量 {PROJECT_ORDER_MAPPING_ENV} 指定清洗后 Excel。'
            ),
        }])
        return sheets

    ambiguous_rows = []
    for _, row in rows.iterrows():
        candidates = ambiguous_map.get(row['泛微项目编号'])
        if candidates:
            ambiguous_rows.append({
                '来源单据编号': row['来源单据编号'],
                '泛微项目编号': row['泛微项目编号'],
                '候选订单编号': '; '.join(_cell_text(item.get('订单编号')) for item in candidates),
                '候选订单标题': '; '.join(_cell_text(item.get('订单标题')) for item in candidates),
                '候选项目编号': '; '.join(_cell_text(item.get('项目编号')) for item in candidates),
                '候选映射来源': '; '.join(_cell_text(item.get('映射来源')) for item in candidates),
            })
    if ambiguous_rows:
        sheets['订单映射_多候选'] = pd.DataFrame(ambiguous_rows)

    order_presence_df = _build_project_order_presence(mapping_file)
    order_presence = {
        project_code: group.to_dict('records')
        for project_code, group in order_presence_df.groupby('泛微项目编号', sort=False)
    }

    mapped_projects = set(safe_map) | set(ambiguous_map)
    unmatched = rows[~rows['泛微项目编号'].isin(mapped_projects)].copy()
    if len(unmatched) > 0:
        unmatched_reasons = []
        unmatched_sources = []
        unmatched_order_values = []
        for _, row in unmatched.iterrows():
            appearances = order_presence.get(row['泛微项目编号'], [])
            if appearances:
                unmatched_reasons.append(f'{PROJECT_ORDER_MAPPING_XLSX_NAME} 项目/订单映射表中有该泛微项目编号,但没有可用订单编号')
                unmatched_sources.append('; '.join(_cell_text(item.get('映射来源')) for item in appearances))
                unmatched_order_values.append('; '.join(_cell_text(item.get('订单编号')) or '(空)' for item in appearances))
            else:
                unmatched_reasons.append(f'{PROJECT_ORDER_MAPPING_XLSX_NAME} 全量项目/全量订单均未出现该泛微项目编号')
                unmatched_sources.append('')
                unmatched_order_values.append('')
        unmatched.insert(2, '未匹配原因', unmatched_reasons)
        unmatched['订单映射表出现位置'] = unmatched_sources
        unmatched['订单映射表订单编号字段值'] = unmatched_order_values
        sheets['订单映射_未匹配'] = unmatched
    return sheets


def _joined_row_key(row, columns):
    return remove_slashes(''.join(_cell_text(row[column]) for column in columns if column < len(row)))


def _first_subject_pair(row, pairs):
    """按优先级读取 (费用项目编码, 费用项目描述) 列对。"""
    for code_column, name_column in pairs:
        code = _cell_text(row[code_column]) if code_column < len(row) else ''
        name = _cell_text(row[name_column]) if name_column < len(row) else ''
        if code and code != '\\':
            return code, name
    return '', ''


def _read_rule_sheet(sheet_name):
    return pd.read_excel(RULE_XLSX, sheet_name=sheet_name, header=None, engine='calamine')


def build_subject_map():
    """费用科目(预算科目)-> (费用项目编码, 费用项目描述)。
    来源:规则表「赛事/MCN 新旧预算项科目-调整后」。键=各级科目名拼接去'/';值=调整后三级预算编码+三级费用项。"""
    subject_map = {}
    event_base_df = _read_rule_sheet('赛事泛微新旧科目映射底表')
    event_base_map = {}
    for _, row in event_base_df.iloc[1:].iterrows():
        key = remove_slashes(_cell_text(row[5]))
        code, name = _first_subject_pair(row, ((6, 7),))
        if key and code:
            event_base_map[key] = (code, name)

    event_budget_df = _read_rule_sheet('赛事新旧预算项科目-调整后')
    for _, row in event_budget_df.iloc[2:].iterrows():
        keys = {_joined_row_key(row, (2, 5, 8, 11, 14)), remove_slashes(_cell_text(row[15]))}
        for key in (k for k in keys if k):
            subject_code, subject_name = _first_subject_pair(row, ((21, 22), (19, 20), (17, 18)))
            if not subject_code:
                subject_code, subject_name = event_base_map.get(key, ('', ''))
            if subject_code:
                subject_map[key] = (subject_code, subject_name)

    mcn_base_df = _read_rule_sheet('MCN泛微新旧科目映射底表')
    mcn_code_to_name = {}
    mcn_code_to_subject = {}
    mcn_name_to_subject = {}
    for _, row in mcn_base_df.iloc[2:].iterrows():
        old_code = _cell_text(row[1])
        old_name = _cell_text(row[2])
        subject_code, subject_name = _first_subject_pair(row, ((5, 6),))
        if old_code and old_name:
            mcn_code_to_name[old_code] = old_name
        if old_code and subject_code:
            mcn_code_to_subject[old_code] = (subject_code, subject_name)
        if old_name and subject_code:
            mcn_name_to_subject[old_name] = (subject_code, subject_name)

    mcn_budget_df = _read_rule_sheet('MCN新旧预算项科目-调整后')
    for _, row in mcn_budget_df.iloc[2:].iterrows():
        codes = [_cell_text(row[column]) for column in (0, 2, 4) if column < len(row)]
        names = [_cell_text(row[column]) or mcn_code_to_name.get(code, '') for column, code in zip((1, 3, 5), codes)]
        keys = {
            _joined_row_key(row, (1, 3, 5)),
            remove_slashes(''.join(name for name in names if name)),
            remove_slashes(''.join(code for code in codes if code)),
        }
        subject_code, subject_name = _first_subject_pair(row, ((11, 12), (9, 10), (7, 8)))
        if not subject_code:
            deepest_code = next((code for code in reversed(codes) if code), '')
            deepest_name = next((name for name in reversed(names) if name), '')
            subject_code, subject_name = (
                mcn_code_to_subject.get(deepest_code)
                or mcn_name_to_subject.get(deepest_name)
                or ('', '')
            )
        if subject_code:
            for key in (k for k in keys if k):
                subject_map.setdefault(key, (subject_code, subject_name))
    return subject_map


# ============================ 去重 / 统计 ============================
# 行过滤口径(流程来源/日期/状态等)各任务差异较大,放在各任务文件内的 filter_main,不在此公共层。
def dedup_rows(output_df, key_cols):
    """按 key_cols 完全相同则合并为一条。返回(去重后, 参与合并的全部行[供核对])。"""
    group_key = output_df[key_cols].apply(lambda row: '|'.join(map(str, row.tolist())), axis=1)
    duplicate_mask = group_key.duplicated(keep=False)
    # 注意:key 要对齐到 mask 子集;否则 0 重复时空表 .assign 整列会把空表撑回全部行
    merged_rows = output_df[duplicate_mask].copy()
    merged_rows['_key'] = group_key[duplicate_mask]
    merged_rows = merged_rows.sort_values('_key').drop(columns='_key')
    deduped_rows = output_df[~group_key.duplicated(keep='first')].reset_index(drop=True)
    print(f'分组去重: 参与合并 {int(duplicate_mask.sum())} 行 -> 最终输出 {len(deduped_rows)} 行')
    return deduped_rows, merged_rows


def required_columns(rule_sheet, table_name):
    """必输字段以【规则表的「是否必填」列=Y】。
    从规则表 rule_sheet 内、表名列=table_name 的行里,取「是否必填」=Y 的字段名。
    规则列:0=模块 1=表名 2=字段名 3=是否必填。"""
    df = pd.read_excel(RULE_XLSX, sheet_name=rule_sheet, header=None, engine='calamine')
    columns = []
    current_table = ''
    for _, row in df.iloc[2:].iterrows():
        table = str(row[1]).strip()
        if table and table != 'nan':
            current_table = table
        if current_table == table_name and str(row[3]).strip() == 'Y':
            field = str(row[2]).strip()
            if field and field != 'nan' and field not in columns:
                columns.append(field)
    return columns


def required_column_remarks(rule_sheet, table_name):
    """读取规则表目标表下各字段的备注。用于问题清单汇总页补充说明。"""
    df = pd.read_excel(RULE_XLSX, sheet_name=rule_sheet, header=None, engine='calamine')
    remarks = {}
    current_table = ''
    for _, row in df.iloc[2:].iterrows():
        table = str(row[1]).strip()
        if table and table != 'nan':
            current_table = table
        if current_table != table_name:
            continue
        field = str(row[2]).strip()
        if not field or field == 'nan' or field in remarks:
            continue
        remarks[field] = _cell_text(row[5] if len(row) > 5 else '')
    return remarks


def report_fill(output_df, columns):
    """打印各列非空填充率。只统计 output_df 里存在的列。"""
    for column in columns:
        if column not in output_df.columns:
            continue
        filled_count = (output_df[column].astype(str).str.strip() != '').sum()
        print(f'  {column} 填充率: {filled_count}/{len(output_df)} = {filled_count/len(output_df)*100:.2f}%')


def collect_field_issues(output_df, source_df, required_cols, source_field_map, doc_col='来源单据编号'):
    """驱动于必输字段:遍历所有必输字段,凡【部分缺失】(0<缺失<全部)的,各生成一张
    「缺失_<字段>」明细 sheet。每张只两列:来源单据编号 + 泛微原表-<源字段>(没匹配上的原始值)。
    不写死具体字段;以后新增必输字段自动纳入。
    全空字段(如规则说明无需填写/暂未映射的字段)只在「必输字段未达100%」汇总里体现,不导出整表明细。
        output_df         输出宽表(含 doc_col;用于判断缺失)
        source_df         与 output_df 同索引的主子合并表(取泛微原始字段值)
        required_cols     必输字段(通常来自 required_columns(模版))
        source_field_map  {输出必输字段: source_df 里对应的泛微源字段名}
        doc_col           单据编号列(默认输出里的「来源单据编号」)"""
    sheets = {}
    total = len(output_df)
    for column in required_cols:
        if column not in output_df.columns:
            continue
        blank_mask = output_df[column].astype(str).str.strip() == ''
        missing_count = int(blank_mask.sum())
        if not (0 < missing_count < total):
            continue
        data = {doc_col: output_df.loc[blank_mask, doc_col].astype(str)}
        source_field = source_field_map.get(column)
        if source_field and source_field in source_df.columns:
            data[f'泛微原表-{source_field}'] = source_df.loc[blank_mask, source_field].astype(str)
        sheets[f'缺失_{column}'] = pd.DataFrame(data).drop_duplicates().reset_index(drop=True)
    return sheets


def fill_summary(output_df, columns, rule_sheet=None, table_name=None):
    """返回必输字段中【填充率未达100%】的汇总(供问题清单)。全部满 100% 时返回空表。
    若规则备注含「无需填写」且该必输字段整列为空,汇总页备注写「无需填写」。"""
    total = len(output_df)
    rule_remarks = required_column_remarks(rule_sheet, table_name) if rule_sheet and table_name else {}
    rows = []
    for column in columns:
        column_exists = column in output_df.columns
        filled = int((output_df[column].astype(str).str.strip() != '').sum()) if column_exists else 0
        if filled < total:
            if filled == 0 and '无需填写' in rule_remarks.get(column, ''):
                remark = '无需填写'
            elif not column_exists:
                remark = '输出表缺少该列'
            else:
                remark = ''
            rows.append({'必输字段': column, '填充数': filled, '缺失数': total - filled,
                         '总数': total, '填充率': f'{filled / total * 100:.2f}%', '备注': remark})
    return pd.DataFrame(rows, columns=['必输字段', '填充数', '缺失数', '总数', '填充率', '备注'])


# ============================ 未匹配清单辅助列 ============================
# 未匹配清单里可能出现的单据号列名(应收用「来源单据号」,应付/预付用「来源单据编号」)。
BUDGET_ISSUE_DOC_COLUMNS = ('来源单据编号', '来源单据号')


def build_budget_issue_map(source_df, doc_col='流程编号',
                           cost_center_col='成本中心', subject_col='预算科目'):
    """{来源单据号 -> (成本中心, 预算项)},供未匹配清单按单据号补充辅助列。
    成本中心 = 泛微「成本中心」字段(已解析为名称,来源 uf_cbzx.mc);预算项 = 预算科目完整路径。
    同一单据有多个值时,按出现顺序去重后用「; 」拼接。
    source_df 缺成本中心/预算科目列时,对应项留空(如应收无预算科目则预算项为空)。"""
    has_cost = cost_center_col in source_df.columns
    has_subject = subject_col in source_df.columns
    if doc_col not in source_df.columns or not (has_cost or has_subject):
        return {}
    blanks = [None] * len(source_df)
    cost_by_doc = {}
    budget_by_doc = {}
    for doc, cost, subject in zip(
            source_df[doc_col],
            source_df[cost_center_col] if has_cost else blanks,
            source_df[subject_col] if has_subject else blanks):
        doc_text = _cell_text(doc)
        if not doc_text:
            continue
        cost_text = _cell_text(cost)
        if cost_text:
            cost_list = cost_by_doc.setdefault(doc_text, [])
            if cost_text not in cost_list:
                cost_list.append(cost_text)
        path = _cell_text(subject)
        if path:
            budget_list = budget_by_doc.setdefault(doc_text, [])
            if path not in budget_list:
                budget_list.append(path)
    return {
        doc: ('; '.join(cost_by_doc.get(doc, [])), '; '.join(budget_by_doc.get(doc, [])))
        for doc in set(cost_by_doc) | set(budget_by_doc)
    }


def attach_budget_issue_columns(sheets, budget_map, doc_columns=BUDGET_ISSUE_DOC_COLUMNS):
    """给未匹配清单各 sheet 末尾补「成本中心」「预算项」两列(按单据号关联 budget_map)。
    识别不到单据号列的 sheet(如必输字段汇总页)原样跳过。原地修改并返回 sheets。"""
    for sheet_df in sheets.values():
        if sheet_df is None or len(sheet_df) == 0:
            continue
        doc_col = next((col for col in doc_columns if col in sheet_df.columns), None)
        if doc_col is None:
            continue
        docs = sheet_df[doc_col].map(_cell_text)
        sheet_df['成本中心'] = docs.map(lambda doc: budget_map.get(doc, ('', ''))[0])
        sheet_df['预算项'] = docs.map(lambda doc: budget_map.get(doc, ('', ''))[1])
    return sheets


# ============================ Excel 输出 ============================
def _clean_cell_value(value):
    """Excel(openpyxl)不允许的控制字符会报 IllegalCharacterError, 写入前剔除。"""
    if pd.isna(value):
        return ''
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub('', value)
    return value


def _fill_sheet(worksheet, output_df):
    """把 output_df 按列顺序写进 worksheet(保留表头样式,清空旧数据行)。"""
    for col_idx, column_name in enumerate(output_df.columns, start=1):
        worksheet.cell(row=1, column=col_idx).value = column_name
    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row)
    for row in output_df.itertuples(index=False, name=None):
        worksheet.append([_clean_cell_value(v) for v in row])


def write_to_template(output_df, template_path, output_path, sheet_name):
    """写进导入模版单个 sheet(保留表头与 lov 下拉页),从第 2 行覆盖写入。"""
    wb = load_workbook(template_path)
    _fill_sheet(wb[sheet_name], output_df)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def write_template_sheets(template_path, output_path, sheet_to_df, extra_sheets=None):
    """一次把多个 sheet 写进同一个导入模版(保留各表头与 lov 页),单次加载/保存。

    sheet_to_df: {sheet名: DataFrame};写入模版已有 sheet(列顺序需与表头一致)。
    extra_sheets: {sheet名: DataFrame};新建 sheet(如核对页),避免重复 load/save 大文件。
    """
    wb = load_workbook(template_path)
    for sheet_name, output_df in sheet_to_df.items():
        _fill_sheet(wb[sheet_name], output_df)
    for sheet_name, output_df in (extra_sheets or {}).items():
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        worksheet = wb.create_sheet(sheet_name)
        worksheet.append([_clean_cell_value(c) for c in output_df.columns])
        for row in output_df.itertuples(index=False, name=None):
            worksheet.append([_clean_cell_value(v) for v in row])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def write_exceptions(output_path, sheets):
    """导出未匹配/待核对清单。sheets: {sheet名: DataFrame}。空表的 sheet 不生成。"""
    non_empty = {name: df for name, df in sheets.items() if len(df) > 0}
    if not non_empty:
        print('  (无任何未匹配/待核对项,跳过清单文件)')
        return None
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path) as writer:
        for sheet_name, sheet_df in non_empty.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
    return output_path
