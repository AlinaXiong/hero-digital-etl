# -*- coding: utf-8 -*-
"""反商业贿赂协议签署情况 —— 把已签署的反商业贿赂协议供应商补登到台账。

规则:合同审批台账(MCN uf_htk + 赛事 uf_htsp)里,
    合同编号含 'H-P' 且 合同名称含 '贿赂'(兼容「反商业贿赂」「反贿赂」)的合同,
    视为反商业贿赂协议;把对应供应商补登到第一个 sheet 末尾。

同时会按第一个 sheet 已有的合同编号批量查询并刷新「合同状态」、
「供应商ID」「供应商汉得编码」「供应商汉得名称」列,
用于补齐底稿里前序合同缺失/过期的信息。

不改动模板 resources/templates/contract/反商业贿赂协议签署情况.xlsx,
而是以它为底稿,另存为带日期后缀的新文件
    output/anti_bribery_signers_db/反商业贿赂协议签署情况_<YYYYMMDD>.xlsx。

去重键 = (供应商名称, 合同编号):若该组合已记录则跳过,可重复跑(幂等)。
新增行尽量补全:供应商名称 / 合同编号 / 合同状态 / 签约时间 / 供应商ID /
    供应商汉得编码 / 供应商汉得名称。

跑法:在项目根执行  python run.py anti_bribery_signers_db
"""
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

if __package__ is None or __package__ == '':
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from etl.util import common as c
from etl.contract import contract_general_db as cg


# ============================ 文件 / 过滤口径 ============================
TASK_NAME = 'anti_bribery_signers_db'
# 底稿(只读):已有的签署情况台账,提供已记录的去重基准与表头结构。
TEMPLATE_FILE_CANDIDATES = (
    cg.TEMPLATE_DIR / '【可查】反商业贿赂协议签署情况.xlsx',
    cg.TEMPLATE_DIR / '反商业贿赂协议签署情况.xlsx',
)
TEMPLATE_FILE = next((path for path in TEMPLATE_FILE_CANDIDATES if path.exists()), TEMPLATE_FILE_CANDIDATES[0])
# 产出(带日期后缀):底稿内容 + 本次新增,模板本身不改动。
OUTPUT_DIR = c.OUT_DIR / TASK_NAME
OUTPUT_FILE = OUTPUT_DIR / f'反商业贿赂协议签署情况_{c.today_suffix()}.xlsx'

# 合同编号「带有 H-P」、合同名称「带有 贿赂」即视为反商业贿赂协议。
CONTRACT_NUMBER_LIKE = '%H-P%'   # UPPER(htbh) LIKE,兼容 H-P / H-PZ 等
CONTRACT_NAME_LIKE = '%贿赂%'     # 兼容「反商业贿赂协议」「反贿赂协议」

# 第一个 sheet 的列名(只写其中存在的列)。
HEADER_SUPPLIER_NAME = '供应商名称'
HEADER_CONTRACT_NO = '合同编号'
HEADER_CONTRACT_STATUS = '合同状态'
HEADER_SIGN_TIME = '签约时间'
HEADER_REMARK = '备注'
HEADER_SUPPLIER_ID = '供应商ID'
HEADER_HAND_CODE = '供应商汉得编码'
HEADER_HAND_NAME = '供应商汉得名称'

REMARK_CUSTOMER_FALLBACK = '无供应商，取客户'

# 取数:MCN 合同库 uf_htk(合同名称字段 htbt)。
SQL_HTK = """
SELECT
    h.htbh  AS `合同编号`,
    h.htbt  AS `合同标题`,
    h.htzt  AS `合同签署状态ID`,
    h.htgys AS `合同供应商ID`,
    h.htkh  AS `合同客户ID`,
    h.htqdrq AS `合同签订日期`
FROM uf_htk h
WHERE UPPER(h.htbh) LIKE %s AND h.htbt LIKE %s
ORDER BY h.htbh, h.id
"""

# 取数:赛事 合同审批台账 uf_htsp(合同名称字段 htmc,供应商字段 gys,客户字段 kh,签订日期 qdsj)。
SQL_HTSP = """
SELECT
    h.htbh AS `合同编号`,
    h.htmc AS `合同标题`,
    h.htzt AS `合同签署状态ID`,
    h.gys  AS `合同供应商ID`,
    h.kh   AS `合同客户ID`,
    h.qdsj AS `合同签订日期`
FROM uf_htsp h
WHERE UPPER(h.htbh) LIKE %s AND h.htmc LIKE %s
ORDER BY h.htbh, h.id
"""

# 按文档已有合同编号刷新状态。编号有时从表格复制会带空白/换行,用同一规则归一后匹配。
CONTRACT_NO_KEY_EXPR = (
    "UPPER(REPLACE(REPLACE(REPLACE(htbh, ' ', ''), CHAR(13), ''), CHAR(10), ''))"
)

SQL_HTK_STATUS_BY_NO = f"""
SELECT
    {CONTRACT_NO_KEY_EXPR} AS `合同编号键`,
    htzt AS `合同签署状态ID`
FROM uf_htk
WHERE {CONTRACT_NO_KEY_EXPR} IN %(contract_no_keys)s
ORDER BY htbh, id
"""

SQL_HTSP_STATUS_BY_NO = f"""
SELECT
    {CONTRACT_NO_KEY_EXPR} AS `合同编号键`,
    htzt AS `合同签署状态ID`
FROM uf_htsp
WHERE {CONTRACT_NO_KEY_EXPR} IN %(contract_no_keys)s
ORDER BY htbh, id
"""

SQL_HTK_PARTY_BY_NO = f"""
SELECT
    {CONTRACT_NO_KEY_EXPR} AS `合同编号键`,
    htgys AS `合同供应商ID`,
    htkh  AS `合同客户ID`
FROM uf_htk
WHERE {CONTRACT_NO_KEY_EXPR} IN %(contract_no_keys)s
ORDER BY htbh, id
"""

SQL_HTSP_PARTY_BY_NO = f"""
SELECT
    {CONTRACT_NO_KEY_EXPR} AS `合同编号键`,
    gys AS `合同供应商ID`,
    kh  AS `合同客户ID`
FROM uf_htsp
WHERE {CONTRACT_NO_KEY_EXPR} IN %(contract_no_keys)s
ORDER BY htbh, id
"""

SQL_FW_SUPPLIER_BY_NAME = """
SELECT
    id AS `供应商ID`,
    khmc AS `供应商名称`,
    zt AS `供应商状态ID`,
    rzzt AS `认证状态ID`
FROM uf_khgys
WHERE khmc IN %(supplier_names)s
ORDER BY khmc, id
"""


# ============================ 小工具 ============================
def _text(value):
    if pd.isna(value):
        return ''
    text = str(value).strip()
    return '' if text in ('', 'nan', 'None', 'NaT') else text


def _norm_no(value):
    """合同编号去重键:去空白、转大写。"""
    return re.sub(r'\s+', '', _text(value)).upper()


def _dedupe_key(supplier_name, contract_no):
    """去重键 = (归一化供应商名称, 归一化合同编号)。

    供应商名称用 common.normalize_name,容忍历史数据里的空格/换行/标点差异。
    """
    return (c.normalize_name(supplier_name), _norm_no(contract_no))


def _join(parts):
    """多供应商时按 '; ' 拼接,去重且保留顺序、忽略空值。"""
    result, seen = [], set()
    for part in parts:
        text = _text(part)
        if text and text not in seen:
            seen.add(text)
            result.append(text)
    return '; '.join(result)


def _first_non_blank_local(*values):
    for value in values:
        text = _text(value)
        if text:
            return text
    return ''


def _supplier_name_variants(name):
    """供应商名称查询兜底:只生成低风险的格式/漏字变体。"""
    text = _text(name)
    if not text:
        return []
    variants = [text]
    paren_variants = []
    for value in variants:
        paren_variants.append(value.replace('（', '(').replace('）', ')'))
        paren_variants.append(value.replace('(', '（').replace(')', '）'))
    variants.extend(paren_variants)
    variants.extend(value.replace('责任有限公司', '有限责任公司') for value in list(variants))

    result = []
    seen = set()
    for value in variants:
        value = _text(value)
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def _chunks(values, size=500):
    for index in range(0, len(values), size):
        yield values[index:index + size]


# ============================ 取数 / 解析 ============================
def _fetch_contract_status_map(contract_numbers):
    """按合同编号查状态名称。若同编号两表都存在,MCN(uf_htk)优先。"""
    contract_no_keys = []
    seen = set()
    for contract_no in contract_numbers:
        key = _norm_no(contract_no)
        if key and key not in seen:
            seen.add(key)
            contract_no_keys.append(key)
    if not contract_no_keys:
        return {}

    htk_status = c.build_fw_select_option_maps(cg.FW_TABLE, ['htzt']).get('htzt', {})
    htsp_status = c.build_fw_select_option_maps(cg.FW_TABLE_HTSP, ['htzt']).get('htzt', {})
    status_by_contract_no = {}

    for batch in _chunks(contract_no_keys):
        params = {'contract_no_keys': tuple(batch)}
        htk_df = c.query_db('FW', 'vspn_xtyy', SQL_HTK_STATUS_BY_NO, params)
        for row in htk_df.to_dict('records'):
            key = _text(row['合同编号键'])
            status = htk_status.get(c.format_code(row['合同签署状态ID']), '')
            if key and status and key not in status_by_contract_no:
                status_by_contract_no[key] = status

        htsp_df = c.query_db('FW', 'vspn_xtyy', SQL_HTSP_STATUS_BY_NO, params)
        for row in htsp_df.to_dict('records'):
            key = _text(row['合同编号键'])
            status = htsp_status.get(c.format_code(row['合同签署状态ID']), '')
            if key and status and key not in status_by_contract_no:
                status_by_contract_no[key] = status

    print(f'[反商业贿赂] 文档已有合同状态命中: {len(status_by_contract_no)} / {len(contract_no_keys)}')
    return status_by_contract_no


def _fetch_contract_party_map(contract_numbers):
    """按合同编号查对方主体信息。若同编号两表都存在,MCN(uf_htk)优先。"""
    contract_no_keys = []
    seen = set()
    for contract_no in contract_numbers:
        key = _norm_no(contract_no)
        if key and key not in seen:
            seen.add(key)
            contract_no_keys.append(key)
    if not contract_no_keys:
        return {}

    rows = []
    for batch in _chunks(contract_no_keys):
        params = {'contract_no_keys': tuple(batch)}
        htk_df = c.query_db('FW', 'vspn_xtyy', SQL_HTK_PARTY_BY_NO, params)
        htk_df['数据来源'] = '泛微(MCN)'
        htsp_df = c.query_db('FW', 'vspn_xtyy', SQL_HTSP_PARTY_BY_NO, params)
        htsp_df['数据来源'] = '泛微(赛事)'
        rows.extend(pd.concat([htk_df, htsp_df], ignore_index=True).to_dict('records'))

    if not rows:
        print(f'[反商业贿赂] 文档已有合同对方主体命中: 0 / {len(contract_no_keys)}')
        return {}

    df = pd.DataFrame(rows)
    supplier_info_map = cg.build_supplier_info_map_for_values(df['合同供应商ID'])
    customer_info_map = cg.build_customer_info_map_for_values(df['合同客户ID'])

    party_by_contract_no = {}
    customer_fallback = 0
    for source in rows:
        key = _text(source['合同编号键'])
        if not key or key in party_by_contract_no:
            continue

        supplier_ids = c.parse_browser_ids(source['合同供应商ID'])
        supplier_infos = [supplier_info_map.get(supplier_id, {}) for supplier_id in supplier_ids]
        party_name = _join(info.get('source_name', '') for info in supplier_infos)

        if party_name:
            party_ids, infos = supplier_ids, supplier_infos
        else:
            party_ids = c.parse_browser_ids(source['合同客户ID'])
            infos = [customer_info_map.get(customer_id, {}) for customer_id in party_ids]
            party_name = _join(info.get('source_name', '') for info in infos)
            if party_name:
                customer_fallback += 1

        if not (party_ids or infos):
            continue
        party_by_contract_no[key] = {
            HEADER_SUPPLIER_ID: _join(party_ids),
            HEADER_HAND_CODE: _join(info.get('code', '') for info in infos),
            HEADER_HAND_NAME: _join(info.get('name', '') for info in infos),
        }

    print(
        f'[反商业贿赂] 文档已有合同对方主体命中: '
        f'{len(party_by_contract_no)} / {len(contract_no_keys)}; 无供应商改取客户 {customer_fallback} 条'
    )
    return party_by_contract_no


def _fetch_supplier_party_by_name_map(supplier_names):
    """按表格里的供应商名称兜底反查供应商ID和汉得信息。"""
    supplier_names = c.clean_text_values(supplier_names)
    supplier_name_keys = c.normalized_name_values(supplier_names)
    if not supplier_names:
        return {}

    query_names = []
    variant_to_source_key = {}
    seen_query_names = set()
    for supplier_name in supplier_names:
        source_key = c.normalize_name(supplier_name)
        for variant in _supplier_name_variants(supplier_name):
            variant_key = c.normalize_name(variant)
            if variant_key:
                variant_to_source_key.setdefault(variant_key, source_key)
            if variant not in seen_query_names:
                seen_query_names.add(variant)
                query_names.append(variant)

    rows = []
    for batch in _chunks(query_names):
        rows.extend(c.query_db(
            'FW',
            'vspn_xtyy',
            SQL_FW_SUPPLIER_BY_NAME,
            {'supplier_names': tuple(batch)},
        ).to_dict('records'))
    if not rows:
        print(f'[反商业贿赂] 供应商名称兜底命中: 0 / {len(supplier_name_keys)}')
        return {}

    ids_by_name_key = {}
    status_map = {}
    for row in rows:
        supplier_id = c.format_code(row['供应商ID'])
        name_key = variant_to_source_key.get(c.normalize_name(row['供应商名称']))
        if not supplier_id or not name_key:
            continue
        ids_by_name_key.setdefault(name_key, []).append(supplier_id)
        status_map[supplier_id] = {
            'name': _text(row['供应商名称']),
            'status_code': c.format_code(row['供应商状态ID']),
            'certification_status': c.format_code(row['认证状态ID']),
        }

    selected_ids = {
        name_key: c.choose_fw_supplier_id(supplier_ids, status_map)
        for name_key, supplier_ids in ids_by_name_key.items()
    }
    supplier_info_map = cg.build_supplier_info_map_for_values(selected_ids.values())
    hand_info_by_name = c.build_hand_vendor_info_by_names(supplier_names)

    result = {}
    for name_key, supplier_id in selected_ids.items():
        supplier_info = supplier_info_map.get(supplier_id, {})
        hand_info = hand_info_by_name.get(name_key, {})
        hand_code = _first_non_blank_local(supplier_info.get('code', ''), hand_info.get('code', ''))
        hand_name = _first_non_blank_local(
            supplier_info.get('name', ''),
            hand_info.get('name', ''),
            status_map.get(supplier_id, {}).get('name', ''),
        )
        result[name_key] = {
            HEADER_SUPPLIER_ID: supplier_id,
            HEADER_HAND_CODE: hand_code,
            HEADER_HAND_NAME: hand_name,
        }

    hand_only = 0
    for name in supplier_names:
        name_key = c.normalize_name(name)
        if not name_key or name_key in result:
            continue
        hand_info = hand_info_by_name.get(name_key, {})
        if not hand_info.get('code'):
            continue
        result[name_key] = {
            HEADER_SUPPLIER_ID: '',
            HEADER_HAND_CODE: hand_info.get('code', ''),
            HEADER_HAND_NAME: hand_info.get('name', ''),
        }
        hand_only += 1

    print(
        f'[反商业贿赂] 供应商名称兜底命中: {len(result)} / {len(supplier_name_keys)}'
        f'{f"; 仅命中汉得 {hand_only} 条" if hand_only else ""}'
    )
    return result


def _fetch_anti_bribery_contracts():
    """查两套合同台账,返回带「合同状态」「供应商各字段」的候选行列表。"""
    htk_df = c.query_db('FW', 'vspn_xtyy', SQL_HTK, [CONTRACT_NUMBER_LIKE, CONTRACT_NAME_LIKE])
    htk_df['数据来源'] = '泛微(MCN)'
    htsp_df = c.query_db('FW', 'vspn_xtyy', SQL_HTSP, [CONTRACT_NUMBER_LIKE, CONTRACT_NAME_LIKE])
    htsp_df['数据来源'] = '泛微(赛事)'
    print(f'[反商业贿赂] 命中合同:MCN(uf_htk) {len(htk_df)} 条; 赛事(uf_htsp) {len(htsp_df)} 条')

    # 合同状态名称:两表各自的 htzt 选项码表。
    htk_status = c.build_fw_select_option_maps(cg.FW_TABLE, ['htzt']).get('htzt', {})
    htsp_status = c.build_fw_select_option_maps(cg.FW_TABLE_HTSP, ['htzt']).get('htzt', {})

    df = pd.concat([htk_df, htsp_df], ignore_index=True)
    if df.empty:
        return []

    # 对方主体:复用一般流程的汉得匹配,拿到 泛微名称 / 汉得编码 / 汉得名称。
    # 反商业贿赂协议可能签给供应商,也可能签给客户;无供应商时回退取客户。
    supplier_info_map = cg.build_supplier_info_map_for_values(df['合同供应商ID'])
    customer_info_map = cg.build_customer_info_map_for_values(df['合同客户ID'])

    contracts = []
    customer_fallback = 0
    seen_contract = set()  # 同编号在两表都命中时,按 MCN 优先取一条
    for source in df.to_dict('records'):
        contract_no = _text(source['合同编号'])
        if contract_no and contract_no in seen_contract:
            continue
        seen_contract.add(contract_no)

        status_map = htk_status if source['数据来源'] == '泛微(MCN)' else htsp_status
        supplier_ids = c.parse_browser_ids(source['合同供应商ID'])
        supplier_infos = [supplier_info_map.get(supplier_id, {}) for supplier_id in supplier_ids]
        party_name = _join(info.get('source_name', '') for info in supplier_infos)

        if party_name:
            party_ids, infos, remark = supplier_ids, supplier_infos, ''
        else:
            # 无供应商 -> 取客户(合同客户ID)。
            party_ids = c.parse_browser_ids(source['合同客户ID'])
            infos = [customer_info_map.get(customer_id, {}) for customer_id in party_ids]
            party_name = _join(info.get('source_name', '') for info in infos)
            remark = REMARK_CUSTOMER_FALLBACK if party_name else ''
            if party_name:
                customer_fallback += 1

        contracts.append({
            HEADER_SUPPLIER_NAME: party_name,
            HEADER_CONTRACT_NO: contract_no,
            HEADER_CONTRACT_STATUS: status_map.get(c.format_code(source['合同签署状态ID']), ''),
            HEADER_SIGN_TIME: c.format_date(source['合同签订日期']),
            HEADER_REMARK: remark,
            HEADER_SUPPLIER_ID: _join(party_ids),
            HEADER_HAND_CODE: _join(info.get('code', '') for info in infos),
            HEADER_HAND_NAME: _join(info.get('name', '') for info in infos),
        })
    print(f'[反商业贿赂] 无供应商改取客户: {customer_fallback} 条')
    return contracts


# ============================ 写入第一个 sheet ============================
def _header_columns(worksheet):
    """第一个 sheet 的表头文本 -> 列号(1-based)。"""
    columns = {}
    for cell in worksheet[1]:
        text = _text(cell.value)
        if text and text not in columns:
            columns[text] = cell.column
    return columns


def _existing_keys_and_last_row(worksheet, columns):
    """已记录去重键集合,以及最后一行有内容的行号(供续写定位)。"""
    name_col = columns.get(HEADER_SUPPLIER_NAME)
    no_col = columns.get(HEADER_CONTRACT_NO)
    keys = set()
    last_row = 1  # 至少保留表头行
    for row in range(2, worksheet.max_row + 1):
        supplier_name = _text(worksheet.cell(row, name_col).value) if name_col else ''
        contract_no = _text(worksheet.cell(row, no_col).value) if no_col else ''
        if not (supplier_name or contract_no):
            continue
        keys.add(_dedupe_key(supplier_name, contract_no))
        last_row = row
    return keys, last_row


def _existing_contract_numbers(worksheet, columns):
    """读取第一个 sheet 已有合同编号,用于刷新状态。"""
    no_col = columns.get(HEADER_CONTRACT_NO)
    if not no_col:
        return []
    contract_numbers = []
    for row in range(2, worksheet.max_row + 1):
        contract_no = _text(worksheet.cell(row, no_col).value)
        if contract_no:
            contract_numbers.append(contract_no)
    return contract_numbers


def _refresh_existing_contract_statuses(worksheet, columns):
    """刷新底稿已有行的合同状态,不新增/删除行。"""
    no_col = columns.get(HEADER_CONTRACT_NO)
    status_col = columns.get(HEADER_CONTRACT_STATUS)
    if not no_col or not status_col:
        print(f'[反商业贿赂] 未找到「{HEADER_CONTRACT_NO}」或「{HEADER_CONTRACT_STATUS}」列,跳过已有状态刷新')
        return 0, 0, 0

    status_by_contract_no = _fetch_contract_status_map(_existing_contract_numbers(worksheet, columns))
    updated = unchanged = missing = 0
    for row in range(2, worksheet.max_row + 1):
        contract_no = _text(worksheet.cell(row, no_col).value)
        if not contract_no:
            continue
        new_status = status_by_contract_no.get(_norm_no(contract_no), '')
        if not new_status:
            missing += 1
            continue
        old_status = _text(worksheet.cell(row, status_col).value)
        if old_status == new_status:
            unchanged += 1
            continue
        worksheet.cell(row, status_col, new_status)
        updated += 1

    print(f'[反商业贿赂] 已有合同状态刷新: 更新 {updated} 条; 不变 {unchanged} 条; 未查到 {missing} 条')
    return updated, unchanged, missing


def _refresh_existing_contract_parties(worksheet, columns):
    """刷新底稿已有行的供应商ID/汉得编码/汉得名称,不新增/删除行。"""
    no_col = columns.get(HEADER_CONTRACT_NO)
    name_col = columns.get(HEADER_SUPPLIER_NAME)
    target_headers = (HEADER_SUPPLIER_ID, HEADER_HAND_CODE, HEADER_HAND_NAME)
    target_columns = {header: columns.get(header) for header in target_headers}
    if not no_col or not name_col or any(not col for col in target_columns.values()):
        print('[反商业贿赂] 未找到合同编号/供应商名称或供应商补充列,跳过已有供应商字段刷新')
        return 0, 0, 0

    party_by_contract_no = _fetch_contract_party_map(_existing_contract_numbers(worksheet, columns))
    supplier_by_name = _fetch_supplier_party_by_name_map(
        _text(worksheet.cell(row, name_col).value)
        for row in range(2, worksheet.max_row + 1)
    )
    updated_cells = unchanged_cells = missing_rows = fallback_rows = 0
    for row in range(2, worksheet.max_row + 1):
        contract_no = _text(worksheet.cell(row, no_col).value)
        supplier_name = _text(worksheet.cell(row, name_col).value)
        party = party_by_contract_no.get(_norm_no(contract_no), {}) if contract_no else {}
        fallback_party = supplier_by_name.get(c.normalize_name(supplier_name), {}) if supplier_name else {}
        if fallback_party and (
            not party
            or any(not party.get(header) for header in target_headers)
        ):
            party = {
                header: _first_non_blank_local(party.get(header, ''), fallback_party.get(header, ''))
                for header in target_headers
            }
            fallback_rows += 1
        if not party:
            missing_rows += 1
            continue
        for header, col in target_columns.items():
            new_value = party.get(header, '')
            if not new_value:
                continue
            old_value = _text(worksheet.cell(row, col).value)
            if old_value == new_value:
                unchanged_cells += 1
                continue
            worksheet.cell(row, col, new_value)
            updated_cells += 1

    print(
        f'[反商业贿赂] 已有供应商字段刷新: 更新单元格 {updated_cells} 个; '
        f'不变 {unchanged_cells} 个; 名称兜底 {fallback_rows} 行; 未查到对方主体 {missing_rows} 行'
    )
    return updated_cells, unchanged_cells, missing_rows


def _timestamped_path(path):
    return path.with_name(f'{path.stem}_{datetime.now().strftime("%H%M%S")}{path.suffix}')


def append_anti_bribery_signers(template_file=TEMPLATE_FILE, output_file=OUTPUT_FILE):
    """以模板为底稿,把命中的反商业贿赂协议供应商补登到第一个 sheet 末尾,
    按 (供应商名称, 合同编号) 去重,并刷新已有行合同状态/供应商字段,
    另存为带日期后缀的新文件;模板本身不改动。"""
    contracts = _fetch_anti_bribery_contracts()
    print(f'[反商业贿赂] 去重后候选合同(MCN 优先): {len(contracts)} 条')

    workbook = load_workbook(template_file)
    worksheet = workbook.worksheets[0]
    columns = _header_columns(worksheet)
    if HEADER_SUPPLIER_NAME not in columns and HEADER_CONTRACT_NO not in columns:
        raise RuntimeError(
            f'第一个 sheet「{worksheet.title}」未找到「{HEADER_SUPPLIER_NAME}」或「{HEADER_CONTRACT_NO}」列,无法写入')

    _refresh_existing_contract_statuses(worksheet, columns)
    _refresh_existing_contract_parties(worksheet, columns)

    existing_keys, last_row = _existing_keys_and_last_row(worksheet, columns)
    next_row = last_row + 1
    appended = 0
    skipped = 0
    for contract in contracts:
        key = _dedupe_key(contract[HEADER_SUPPLIER_NAME], contract[HEADER_CONTRACT_NO])
        if key in existing_keys:
            skipped += 1
            continue
        existing_keys.add(key)
        for header, col in columns.items():
            if header in contract:
                worksheet.cell(next_row, col, contract[header])
        next_row += 1
        appended += 1

    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(output_file)
        saved_to = output_file
    except PermissionError:
        saved_to = _timestamped_path(output_file)
        print(f'产出文件被占用,改写到: {saved_to}')
        workbook.save(saved_to)

    print(f'[反商业贿赂] 新增 {appended} 条; 跳过(已记录) {skipped} 条; 模板未改动')
    print('已写出:', saved_to)
    return appended


def run():
    append_anti_bribery_signers()


if __name__ == '__main__':
    run()
