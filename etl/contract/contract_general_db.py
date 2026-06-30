# -*- coding: utf-8 -*-
"""合同迁移 —— 智书一般流程(DB 直连版)。

按「智书合同字段-一般流程.xlsx」生成一般流程导入模板:
    1. 字段模板
    2. 关联合同 / 相关单据-订单信息 / 采购申请 / 订单信息明细
    3. 对方信息 / 我方主体列表
    4. 付款计划 / 收款计划
    5. 合同附件(仅写附件名称) / 其他附件(模板保留); 文件下载请运行 contract_general_attachments_db

跑法:在项目根执行  python run.py contract_general_db
"""
import html
import os
import re
import sys
import time
from functools import lru_cache
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

if __package__ is None or __package__ == '':
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from etl.util import common as c
from etl.lark import feishu
from etl.process.ap_prepayment_opening_db import build_fw_project_code_map_for_ids


# ============================ 文件 / 模板 ============================
TASK_NAME = 'contract_general_db'
TEMPLATE_DIR = c.TPL_DIR / 'contract'
OUTPUT_DIR = c.OUT_DIR / TASK_NAME
DATE_SUFFIX = c.today_suffix()

TEMPLATE_FILE = TEMPLATE_DIR / '智书合同字段-一般流程.xlsx'
ANTI_BRIBERY_TEMPLATE_FILE = TEMPLATE_DIR / '签署反商业贿赂协议6.25终版.xlsx'
RULE_CSV = c.RULES_DIR / '业财项目_数据映射规则 - 合同数据映射规则-for法务.csv'
# 专项品类映射: 泛微采购品类树(uf_xt_cgpldy)叶子 -> 原一级/二级/三级路径 -> 终版二级分类。
SPECIAL_CATEGORY_CSV = c.RULES_DIR / '预算项&项目类型底稿 - 供应商分类【新旧mapping】.csv'
SPECIAL_CATEGORY_TABLE = 'uf_xt_cgpldy'
ORDER_INIT_FILE = c.SRC_DIR / 'other_cleaned_data' / '订单申请初始化导入-基础信息+财务信息.xlsx'
ORDER_INIT_SHEET = '基础信息+财务信息'
ORDER_INIT_MAPPING_SOURCE = f'{ORDER_INIT_FILE.name}:{ORDER_INIT_SHEET}(OA编号)'
PROJECT_FILTER_ENV = 'PROJECT_FILTER_XLSX'
PROJECT_FILTER_DEFAULT_FILE = c.RULES_DIR / '数据清洗涉及泛微项目编码_0629_分类.xlsx'
PROJECT_FILTER_FILE = Path(os.getenv(PROJECT_FILTER_ENV, '').strip() or PROJECT_FILTER_DEFAULT_FILE)
EVENT_PROJECT_SHEET = '赛事'
MCN_PROJECT_SHEET = 'MCN'
FORCED_GENERAL_CONTRACT_CATEGORY = '框架支出-其他支出框架'
FORCED_GENERAL_CONTRACT_NUMBERS = (
    'HH-O2-202501008',
    'HH-O2-202501010',
    'HH-O2-202501012',
    'HH-O2-202501015',
    'HH-O2-202503001',
    'HH-O2-202503004',
    'HH-O1-202505003',
    'HH-O1-202506002',
    'HH-O2-202601001',
    'HH-O2-202601004',
    'JZ-O1-202601001',
    'HH-O2-202603002',
    'HH-O2-202603003',
    'SHHJ-O1-202603001',
    'SHHJ-O1-202403001-S1',
)
OUTPUT_FILE = OUTPUT_DIR / f'智书合同字段_一般流程_{DATE_SUFFIX}.xlsx'
EXCEPTION_FILE = OUTPUT_DIR / f'未匹配清单_一般流程_{DATE_SUFFIX}.xlsx'

SHEET_MAIN = '字段模板'
SHEET_OPTIONS = '选项'
SHEET_RELATION = '关联合同'
SHEET_RELATED_ORDER = '相关单据-订单信息'
SHEET_PURCHASE_REQUEST = '采购申请'
SHEET_ORDER_DETAIL = '订单信息明细'
SHEET_COUNTERPARTY = '对方信息'
SHEET_OUR_PARTY = '我方主体列表'
SHEET_PAYMENT_PLAN = '付款计划'
SHEET_COLLECTION_PLAN = '收款计划'
SHEET_CONTRACT_ATTACHMENT = '合同附件'
SHEET_OTHER_ATTACHMENT = '其他附件'
ANTI_BRIBERY_SOURCE_SHEETS = (
    '归档的反贿赂供应商',
    '未归档的反贿赂签署协议',
)
ATTACHMENT_FOLDER_MAIN = '主文件'
ATTACHMENT_FOLDER_ARCHIVE_SCAN = '归档扫描件'

OUTPUT_SHEETS = (
    SHEET_MAIN,
    SHEET_RELATION,
    SHEET_RELATED_ORDER,
    SHEET_PURCHASE_REQUEST,
    SHEET_ORDER_DETAIL,
    SHEET_COUNTERPARTY,
    SHEET_OUR_PARTY,
    SHEET_PAYMENT_PLAN,
    SHEET_COLLECTION_PLAN,
    SHEET_CONTRACT_ATTACHMENT,
    SHEET_OTHER_ATTACHMENT,
)

FW_TABLE = 'uf_htk'
MIGRATION_STATUS_CODES = (1, 2)
ANCHOR_CONTRACT_TYPE_CODE = 3

DEFAULT_PROPERTY_TYPE = '固定总价'
DEFAULT_VALIDITY_TYPE = '固定期限'
DEFAULT_ACCEPTANCE_REQUIRED = '否'
DEFAULT_PRINT_MODE = '黑白单面打印'
DEFAULT_INVOICE_TYPE = '增值税专用发票'
# 税率按百分数口径(13 = 13%); 赛事源直接取 srsl/zcsl, 仅在源无值时用此兜底默认(6%)。
DEFAULT_TAX_RATE = '6'
DEFAULT_TAX_ITEM = '生产生活服务'
DEFAULT_BANK_FEE_BEARER = '各自承担'
DEFAULT_FIRST_SEAL_PARTY = '我方'
DEFAULT_SIGN_FORM = '纸质签约'
DEFAULT_SEAL_NUMBER = 2
DEFAULT_PREPAID = '否'
DEFAULT_PAYMENT_NATURE = '一般付款'
DEFAULT_CONTRACT_CREATOR_NAME = '黄劭文'

ATTACHMENT_COOKIE_ENV = c.ATTACHMENT_COOKIE_ENV
ATTACHMENT_BASE_URL_ENV = c.ATTACHMENT_BASE_URL_ENV
ATTACHMENT_LOGIN_USERID_ENV = c.ATTACHMENT_LOGIN_USERID_ENV
ATTACHMENT_AUTHORIZEMODE_ID_ENV = c.ATTACHMENT_AUTHORIZEMODE_ID_ENV
ATTACHMENT_AUTHORIZEFIELD_ID_ENV = c.ATTACHMENT_AUTHORIZEFIELD_ID_ENV
ATTACHMENT_DOWNLOAD_ROOT_ENV = c.ATTACHMENT_DOWNLOAD_ROOT_ENV
ATTACHMENT_DOWNLOAD_ENABLED_ENV = c.ATTACHMENT_DOWNLOAD_ENABLED_ENV
ATTACHMENT_DOWNLOAD_WORKERS_ENV = c.ATTACHMENT_DOWNLOAD_WORKERS_ENV
ATTACHMENT_DOWNLOAD_RETRIES_ENV = c.ATTACHMENT_DOWNLOAD_RETRIES_ENV
download_attachment_manifest = c.download_attachment_manifest
download_attachment_manifest_16_workers = c.download_attachment_manifest_16_workers
ATTACHMENT_TYPE_DRAFT = '合同初稿'
ATTACHMENT_TYPE_REVISED = '合同修订稿'
ATTACHMENT_TYPE_SIGNED = '合同签署稿'
ATTACHMENT_TYPE_EFFECTIVE = '合同生效稿'

# 审批流程表单(formtable_main_*)里的三个稿件字段 -> 固定归类。
# OA「合同相关」显示名: 合同修订稿/合同签署版/合同生效版。
# 注意: 字段代码因流程/表单而异; 下面是默认(主播协议审批流程 等)的字段名。
FORM_DOC_FIELD_TYPES = (
    ('htsxb', ATTACHMENT_TYPE_EFFECTIVE),  # 合同生效版(= uf_htk.htqdg)
    ('htqsb', ATTACHMENT_TYPE_SIGNED),     # 合同签署版
    ('htxdg', ATTACHMENT_TYPE_REVISED),    # 合同修订稿
)
# 个别流程的表单字段名与默认不同, 按「表单数据表名」覆盖默认字段(界面标签相同)。
# 合同补充/终止流程(流程类型 48)= formtable_main_26: 字段名为 htcg/htfwxdg/htqdg。
FORM_DOC_FIELD_TYPES_BY_TABLE = {
    'formtable_main_26': (
        ('htqdg', ATTACHMENT_TYPE_EFFECTIVE),   # 合同生效版
        ('htfwxdg', ATTACHMENT_TYPE_SIGNED),    # 合同签署版
        ('htcg', ATTACHMENT_TYPE_REVISED),      # 合同修订稿
    ),
}
# 未匹配到上述稿件字段的表单, 退回单一稿件字段 htqdg, 归生效稿。
FORM_DOC_FALLBACK_FIELD = 'htqdg'

# 固定附件覆盖: 个别合同业务确认只下载指定附件(按 docid), 绕过常规稿件归类与目标稿件过滤。
# SHSY-B-202408018-S1(困困经纪约补充协议): 附件只挂在修订稿字段会被过滤, 业务确认仅下载补充协议 docx 一份。
FIXED_CONTRACT_ATTACHMENT_DOCIDS = {
    'SHSY-B-202408018-S1': {1784677},  # 困困别睡辣补充协议.docx (imagefileid 1958491)
}

# 赛事源 uf_htsp 的稿件字段(逗号分隔 docid) -> 固定归类。
SAISHI_DOC_FIELD_TYPES = (
    ('赛事生效稿DOCID', ATTACHMENT_TYPE_EFFECTIVE),
    ('赛事签署稿DOCID', ATTACHMENT_TYPE_SIGNED),
    ('赛事初稿DOCID', ATTACHMENT_TYPE_DRAFT),
)
INVALID_PATH_CHARS = re.compile(r'[<>:"/\\|?*]+')
DOC_ID_SPLITTER = re.compile(r'[,\uff0c]+')
SIGNED_APPROVAL_NODE_KEYWORDS = ('用印', '申请人性质', '上传电子版', '法务确认')


# ============================ 泛微源 SQL ============================
SOURCE_SQL = """
SELECT
    h.id AS `ID`,
    h.htbh AS `合同编号`,
    h.htbt AS `合同标题`,
    h.htlx AS `合同类型ID`,
    h.htejlx AS `合同二级类型ID`,
    h.htzt AS `合同签署状态ID`,
    h.htlc AS `合同流程ID`,
    h.htcjrq AS `合同创建日期`,
    h.htqdrq AS `合同签订日期`,
    h.htyxqqssj AS `合同有效期起始时间`,
    h.htyxqjzsj AS `合同有效期截止时间`,
    h.htzhry AS `合同执行人员ID`,
    h.htyyfw AS `合同用印范围ID`,
    h.htkh AS `合同客户ID`,
    h.htgys AS `合同供应商ID`,
    h.htje AS `合同金额`,
    h.htyjsr AS `合同预计收入`,
    h.htyjzc AS `合同预计支出`,
    h.htzy AS `合同摘要`,
    h.htqdg AS `合同附件DOCID`,
    h.cbzx AS `成本中心ID`,
    h.glkjxy AS `关联框架协议ID`,
    h.bglx AS `变更类型ID`,
    h.bczzbhsc AS `补充/终止编号生成`,
    h.htszxmbh AS `合同所属项目编号ID`,
    h.htszxm AS `合同所属项目`,
    NULL AS `收入税率`,
    NULL AS `支出税率`,
    NULL AS `押金`,
    NULL AS `保证金`,
    NULL AS `采购申请单ID`,
    NULL AS `专项分类编码`,
    h.modedatacreater AS `合同创建人ID`,
    h.modedatacreater AS `申请人ID`,
    rb.workflowid AS `流程类型ID`,
    wb.workflowname AS `流程名称`,
    nb.nodename AS `合同审批状态`
FROM uf_htk h
LEFT JOIN workflow_requestbase rb ON rb.requestid = h.htlc
LEFT JOIN workflow_base wb ON wb.id = rb.workflowid
LEFT JOIN workflow_nownode nn ON nn.requestid = h.htlc
LEFT JOIN workflow_nodebase nb ON nb.id = COALESCE(rb.currentnodeid, nn.nownodeid, rb.lastnodeid)
WHERE (h.htlx <> %(anchor_contract_type_code)s OR h.htlx IS NULL)
  AND h.htzt IN %(migration_status_codes)s
ORDER BY h.htbh, h.id
"""

FORCED_GENERAL_CONTRACT_SOURCE_SQL = SOURCE_SQL.replace(
    """WHERE (h.htlx <> %(anchor_contract_type_code)s OR h.htlx IS NULL)
  AND h.htzt IN %(migration_status_codes)s""",
    """WHERE h.htbh IN %(forced_contract_numbers)s""",
)

STATS_SQL = """
SELECT
    COUNT(*) AS all_count,
    SUM(CASE WHEN htlx = %(anchor_contract_type_code)s THEN 1 ELSE 0 END) AS anchor_type_count,
    SUM(CASE
        WHEN (htlx <> %(anchor_contract_type_code)s OR htlx IS NULL)
         AND htzt IN %(migration_status_codes)s
        THEN 1 ELSE 0 END) AS kept_count,
    SUM(CASE
        WHEN (htlx <> %(anchor_contract_type_code)s OR htlx IS NULL)
         AND (htzt NOT IN %(migration_status_codes)s OR htzt IS NULL)
        THEN 1 ELSE 0 END) AS excluded_status_count
FROM uf_htk
"""

STATUS_BREAKDOWN_SQL = """
SELECT htzt AS `合同签署状态ID`, COUNT(*) AS `合同数`
FROM uf_htk
WHERE htlx <> %(anchor_contract_type_code)s OR htlx IS NULL
GROUP BY htzt
ORDER BY htzt
"""

# ---- 赛事源: 合同审批台账 uf_htsp(独立码表, htzt=0 审批中, 1 归档) ----
FW_TABLE_HTSP = 'uf_htsp'
HTSP_MIGRATION_STATUS_CODES = (0, 1)  # 0=审批中, 1=归档(业务口径覆盖归档/已归档/审批完成)

# 字段对齐到 uf_htk 源的统一列名, 复用下游 resolve/builder。
SOURCE_SQL_HTSP = """
SELECT
    h.id AS `ID`,
    h.htbh AS `合同编号`,
    h.htmc AS `合同标题`,
    h.htlx AS `合同类型ID`,
    h.htejfl AS `合同二级类型ID`,
    h.htzt AS `合同签署状态ID`,
    h.lcqqid AS `合同流程ID`,
    h.sqrq AS `合同创建日期`,
    h.qdsj AS `合同签订日期`,
    h.htksrq AS `合同有效期起始时间`,
    h.htjsrq AS `合同有效期截止时间`,
    h.htzhr AS `合同执行人员ID`,
    h.gszt AS `合同用印范围ID`,
    h.kh AS `合同客户ID`,
    h.gys AS `合同供应商ID`,
    NULL AS `合同金额`,
    h.srje AS `合同预计收入`,
    h.zcje AS `合同预计支出`,
    h.htzy AS `合同摘要`,
    h.htsxg AS `合同附件DOCID`,
    h.htcg AS `赛事初稿DOCID`,
    h.htqsg AS `赛事签署稿DOCID`,
    h.htsxg AS `赛事生效稿DOCID`,
    h.cbzx AS `成本中心ID`,
    COALESCE(h.glht, h.kjht) AS `关联框架协议ID`,
    NULL AS `变更类型ID`,
    NULL AS `补充/终止编号生成`,
    h.xmbh AS `合同所属项目编号ID`,
    h.xmmc AS `合同所属项目`,
    h.srsl AS `收入税率`,
    h.zcsl AS `支出税率`,
    h.yj AS `押金`,
    h.bzj AS `保证金`,
    h.cgsqddx AS `采购申请单ID`,
    h.zxflcg AS `专项分类编码`,
    h.modedatacreater AS `合同创建人ID`,
    h.sqr AS `申请人ID`,
    rb.workflowid AS `流程类型ID`,
    wb.workflowname AS `流程名称`,
    nb.nodename AS `合同审批状态`
FROM uf_htsp h
LEFT JOIN workflow_requestbase rb ON rb.requestid = h.lcqqid
LEFT JOIN workflow_base wb ON wb.id = rb.workflowid
LEFT JOIN workflow_nownode nn ON nn.requestid = h.lcqqid
LEFT JOIN workflow_nodebase nb ON nb.id = COALESCE(rb.currentnodeid, nn.nownodeid, rb.lastnodeid)
WHERE h.htzt IN %(htsp_status_codes)s
ORDER BY h.htbh, h.id
"""

EXPECTED_FW_FIELDS = {
    '': {
        'htbh': '合同编号',
        'htbt': '合同标题',
        'htlx': '合同类型',
        'htejlx': '合同二级类型',
        'htzt': '合同签署状态',
        'htlc': '合同流程',
        'htqdrq': '合同签订日期',
        'htyxqqssj': '合同有效期起始时间',
        'htyxqjzsj': '合同有效期截止时间',
        'htzhry': '合同执行人员',
        'htyyfw': '合同用印范围',
        'htkh': '合同客户',
        'htgys': '合同供应商',
        'htje': '合同金额',
        'htyjsr': '合同预计收入',
        'htyjzc': '合同预计支出',
        'htzy': '合同摘要',
        'htqdg': '合同签定稿',
        'glkjxy': '关联框架协议',
        'bglx': '变更类型',
        'bczzbhsc': '补充/终止编号生成',
        'htszxmbh': '合同所属项目编号',
    },
}


# ============================ 小工具 ============================
def _query_fw(sql):
    return c.query_db('FW', 'vspn_xtyy', sql, {
        'anchor_contract_type_code': ANCHOR_CONTRACT_TYPE_CODE,
        'migration_status_codes': MIGRATION_STATUS_CODES,
    })


def _text(value):
    if pd.isna(value):
        return ''
    text = str(value).strip()
    return '' if text in ('', 'nan', 'None', 'NaT') else text


def _first_non_blank(*values):
    for value in values:
        text = _text(value)
        if text:
            return text
    return ''


def _split_multi_values(value):
    text = _text(value)
    if not text:
        return []
    return [
        item.strip()
        for item in re.split(r'[;；,，\n\r]+', text)
        if item.strip()
    ]


def _split_joined_field(value):
    text = _text(value)
    if not text:
        return []
    return [
        item.strip()
        for item in re.split(r'[;；\n\r]+', text)
        if item.strip()
    ]


@lru_cache(maxsize=1)
def load_contract_project_filter_codes():
    """读取合同一般流程项目白名单:赛事/MCN sheet 的 OA编号。"""
    if not PROJECT_FILTER_FILE.exists():
        raise FileNotFoundError(f'项目白名单不存在: {PROJECT_FILTER_FILE}')

    result = {}
    for sheet_name in (EVENT_PROJECT_SHEET, MCN_PROJECT_SHEET):
        df = pd.read_excel(PROJECT_FILTER_FILE, sheet_name=sheet_name, dtype=object)
        if df.empty:
            result[sheet_name] = set()
            continue
        if 'OA编号' in df.columns:
            code_column = 'OA编号'
        elif '原泛微项目编码' in df.columns:
            code_column = '原泛微项目编码'
        else:
            code_column = df.columns[0]
        codes = []
        for value in df[code_column]:
            codes.extend(c.split_fanwei_project_codes(value))
        result[sheet_name] = set(codes)

    print(
        '[合同迁移-一般流程-项目白名单] 使用:',
        PROJECT_FILTER_FILE,
        f'| 赛事 {len(result[EVENT_PROJECT_SHEET])} 个',
        f'| MCN {len(result[MCN_PROJECT_SHEET])} 个',
    )
    return result


def _project_filter_sheet_for_source(source_name):
    source_name = _text(source_name)
    if source_name == '泛微(赛事)':
        return EVENT_PROJECT_SHEET
    if source_name == '泛微(MCN)':
        return MCN_PROJECT_SHEET
    return ''


def _project_code_filter_reason(row):
    sheet_name = _project_filter_sheet_for_source(row.get('数据来源'))
    # 规则1: 框架合同(KF/KS/KJ)不经过白名单, 始终保留。
    if _is_framework_contract(row.get('合同编号')):
        return sheet_name, ''
    project_codes = c.split_fanwei_project_codes(row.get('泛微项目编号'))
    if not sheet_name:
        return '', '未知数据来源'
    if not project_codes:
        return sheet_name, '泛微项目编码为空'
    allowed_codes = load_contract_project_filter_codes().get(sheet_name, set())
    if any(code in allowed_codes for code in project_codes):
        return sheet_name, ''
    return sheet_name, f'泛微项目编码不在{sheet_name} sheet'


def _filter_by_contract_project_whitelist(source_df):
    if source_df.empty:
        return source_df, pd.DataFrame()

    check = source_df.apply(_project_code_filter_reason, axis=1)
    sheet_names = check.map(lambda item: item[0])
    reasons = check.map(lambda item: item[1])
    force_mask = _force_include_mask(source_df)
    keep_mask = reasons.map(_text).eq('') | force_mask
    filtered = source_df.loc[keep_mask].copy()
    excluded = source_df.loc[~keep_mask].copy()
    if not excluded.empty:
        excluded['项目白名单Sheet'] = sheet_names.loc[~keep_mask].to_numpy()
        excluded['项目白名单校验'] = reasons.loc[~keep_mask].to_numpy()
    print(f'[合同迁移-一般流程] 泛微项目编码白名单过滤: {len(filtered)}/{len(source_df)} 行')
    return filtered, excluded


# 规则2: 合同审批状态硬过滤 —— 只保留这些审批节点(精确匹配清洗后的 nodename); 空/其他状态一律剔除。
KEPT_APPROVAL_STATUSES = (
    '上传修订版',
    '申请人确认签约性质',
    '法务确认',
    '用印',
    '上传电子档',
    '归档',
)


def _filter_by_approval_status(source_df):
    """规则2: 仅保留合同审批状态∈KEPT_APPROVAL_STATUSES 的合同(含框架合同), 空/其他状态剔除。"""
    if source_df.empty:
        return source_df, pd.DataFrame()
    status = source_df.get('合同审批状态', pd.Series('', index=source_df.index)).map(_text)
    keep_mask = status.isin(KEPT_APPROVAL_STATUSES) | _force_include_mask(source_df)
    filtered = source_df.loc[keep_mask].copy()
    excluded = source_df.loc[~keep_mask].copy()
    if not excluded.empty:
        excluded['审批状态过滤'] = status.loc[~keep_mask].map(
            lambda value: '审批状态为空' if not value else f'审批状态不在保留清单:{value}'
        ).to_numpy()
    print(f'[合同迁移-一般流程] 合同审批状态过滤: {len(filtered)}/{len(source_df)} 行')
    return filtered, excluded


@lru_cache(maxsize=1)
def load_saishi_order_init_mapping():
    """赛事合同:按初始化导入表 OA编号(即旧项目/OA编号) -> 订单编号。"""
    if not ORDER_INIT_FILE.exists():
        print(f'[合同迁移-一般流程] 赛事订单初始化表不存在,赛事订单保持原映射: {ORDER_INIT_FILE}')
        return {}
    try:
        df = pd.read_excel(ORDER_INIT_FILE, sheet_name=ORDER_INIT_SHEET, dtype=object)
    except Exception as error:
        print(f'[合同迁移-一般流程] 赛事订单初始化表读取失败,赛事订单保持原映射: {error}')
        return {}

    required = {'OA编号', '订单编号'}
    missing = required - set(df.columns)
    if missing:
        print(f'[合同迁移-一般流程] 赛事订单初始化表缺少列 {sorted(missing)},赛事订单保持原映射。')
        return {}

    by_oa = {}
    for _, row in df.iterrows():
        order_code = _text(row.get('订单编号'))
        oa_codes = c.split_fanwei_project_codes(row.get('OA编号'))
        if not (oa_codes and order_code):
            continue
        item = {
            '订单编号': order_code,
            '订单标题': _text(row.get('订单标题')),
            '申请人员工编码': _text(row.get('申请人员工编码')),
            '项目编号': _text(row.get('项目编号')),
            '项目名称': _text(row.get('项目名称')),
            '映射来源': ORDER_INIT_MAPPING_SOURCE,
        }
        for oa_code in oa_codes:
            items = by_oa.setdefault(oa_code, [])
            if order_code not in {_text(existing.get('订单编号')) for existing in items}:
                items.append(item)

    print(
        '[合同迁移-一般流程] 赛事订单初始化映射:',
        f'{len(df)} 行,',
        f'OA编号 {len(by_oa)} 个,',
        f'订单 {sum(len(items) for items in by_oa.values())} 个',
    )
    return by_oa


def _is_saishi_source(row):
    return _text(row.get('数据来源')) == '泛微(赛事)'


def _saishi_order_init_items_for_source(row):
    if not _is_saishi_source(row):
        return []
    mapping = load_saishi_order_init_mapping()
    for key in c.split_fanwei_project_codes(row.get('泛微项目编号')):
        items = mapping.get(key)
        if items:
            return items
    return []


def _order_init_items_for_source(row):
    if _is_saishi_source(row):
        return _saishi_order_init_items_for_source(row)

    items = []
    seen_orders = set()
    for project_code in (row.get('项目编号'), row.get('泛微项目编号'), row.get('清洗后项目编号')):
        for info in c.cleanable_order_infos_for_project(project_code):
            order_code = _text(info.get('订单编号'))
            if not order_code or order_code in seen_orders:
                continue
            seen_orders.add(order_code)
            items.append({
                '订单编号': order_code,
                '订单标题': _text(info.get('订单标题')),
                '申请人员工编码': '',
                '项目编号': _first_non_blank(*(info.get('项目编号候选') or [])),
                '项目名称': '',
                '映射来源': _text(info.get('映射来源')) or ORDER_INIT_MAPPING_SOURCE,
            })
    return items


def _join_order_item_field(items, field):
    return ';'.join(c.clean_text_values(item.get(field, '') for item in items))


def _number(value, default=0.0):
    if pd.isna(value) or _text(value) == '':
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _round_amount(value):
    return round(_number(value), 2)


def _format_tax_rate(value):
    """税率按源库百分数口径输出, 去掉多余的尾随零: 13.00->'13'、1.0000->'1'、13.5->'13.5'。

    空值返回 '', 由调用方决定是否回退默认税率。
    """
    text = _text(value)
    if not text:
        return ''
    try:
        number = float(text)
    except (TypeError, ValueError):
        return text
    return f'{number:.4f}'.rstrip('0').rstrip('.') or '0'


@lru_cache(maxsize=8192)
def _normalize_cached(text):
    return re.sub(r'\s+', '', text)


def _normalize_field_name(value):
    # 热路径: 模板表头/字段名在 5 万行 × 数十字段上反复归一化, 缓存避免重复正则。
    return _normalize_cached(_text(value))


def _contract_number_key(value):
    return re.sub(r'\s+', '', _text(value)).upper()


def _forced_contract_number_variants(contract_number):
    text = _text(contract_number)
    if not text:
        return []
    variants = [text]
    if 'HH-O2-' in text:
        variants.append(text.replace('HH-O2-', 'HH-02-'))
    if 'HH-02-' in text:
        variants.append(text.replace('HH-02-', 'HH-O2-'))
    return list(dict.fromkeys(variants))


def _forced_general_contract_number_keys():
    keys = set()
    for contract_number in FORCED_GENERAL_CONTRACT_NUMBERS:
        keys.update(_contract_number_key(value) for value in _forced_contract_number_variants(contract_number))
    keys.discard('')
    return keys


def _force_include_mask(source_df):
    if source_df.empty:
        return pd.Series(False, index=source_df.index)
    forced_by_flag = source_df.get('强制追加导出', pd.Series('', index=source_df.index)).map(_text).eq('Y')
    forced_keys = _forced_general_contract_number_keys()
    forced_by_number = source_df.get('合同编号', pd.Series('', index=source_df.index)).map(
        lambda value: _contract_number_key(value) in forced_keys)
    return forced_by_flag | forced_by_number


def _is_force_included_contract(row):
    return (
        _text(row.get('强制追加导出')) == 'Y'
        or _contract_number_key(row.get('合同编号')) in _forced_general_contract_number_keys()
    )


def _first_browser_id(value):
    ids = c.parse_browser_ids(value)
    return ids[0] if ids else ''


def _lookup_first_browser_value(mapping, value):
    for item_id in c.parse_browser_ids(value):
        mapped = mapping.get(item_id, '')
        if mapped:
            return mapped
    return ''


def _sanitize_path_part(value, fallback):
    text = _text(value) or fallback
    text = html.unescape(text)
    text = INVALID_PATH_CHARS.sub('_', text)
    text = text.replace('\r', ' ').replace('\n', ' ').strip(' .')
    return (text or fallback)[:180]


def _to_int_id(value):
    text = _text(value)
    if not text:
        return None
    try:
        return int(float(text)) if re.fullmatch(r'\d+(?:\.0+)?', text) else None
    except (TypeError, ValueError):
        return None


def _split_docids(raw_value):
    text = _text(raw_value)
    if not text:
        return []
    result = []
    for part in DOC_ID_SPLITTER.split(text):
        part = part.strip()
        docid = _to_int_id(part)
        if docid is not None:
            result.append(docid)
    return result


def _timed(label, func):
    """跑一个步骤并打印耗时(秒)。用于定位性能瓶颈。"""
    start = time.perf_counter()
    print(f'[计时] ▶ {label} ...', flush=True)
    result = func()
    print(f'[计时] ✓ {label}: {time.perf_counter() - start:.1f}s', flush=True)
    return result


def _chunked(items, size=5000):
    batch = []
    for item in items:
        batch.append(item)
        if len(batch) >= size:
            yield batch
            batch = []
    if batch:
        yield batch


def _build_target_filename(filename, imagefileid):
    original = _sanitize_path_part(filename, str(imagefileid))
    if '.' not in Path(original).name:
        return f'{original}_{imagefileid}'
    return original


def _unique_attachment_path_preserve_name(target_dir, target_name, imagefileid, used_paths):
    target_path = target_dir / target_name
    if target_path not in used_paths:
        used_paths.add(target_path)
        return target_path

    marker = _sanitize_path_part(imagefileid, 'duplicate')
    for counter in range(1, 1000):
        folder_name = f'duplicate_{marker}' if counter == 1 else f'duplicate_{marker}_{counter}'
        target_path = target_dir / folder_name / target_name
        if target_path not in used_paths:
            used_paths.add(target_path)
            return target_path

    raise RuntimeError(f'Unable to build unique attachment path under {target_dir}')


def _attachment_download_root():
    configured = os.getenv(ATTACHMENT_DOWNLOAD_ROOT_ENV, '').strip()
    if configured:
        return Path(configured)
    return OUTPUT_DIR / f'一般流程合同附件_{DATE_SUFFIX}'


def _load_attachment_maps(docids):
    """一次 JOIN 取齐附件三表, 返回 (docimage_map, imagefile_map, docdetail_map)。

    原先按 docid/imagefileid 分三趟批量查(docimagefile / imagefile / docdetail),
    对远程库是几百次往返。这里改成每个 docid 批次一条
    docimagefile ⋈ imagefile ⋈ docdetail 的 JOIN, 配合更大的批量, 往返从数百降到十几次。
    """
    docimage_map, imagefile_map, docdetail_map = {}, {}, {}
    docids = [docid for docid in docids if docid is not None]
    if not docids:
        return docimage_map, imagefile_map, docdetail_map
    for batch in _chunked(sorted(set(docids))):
        df = c.query_db(
            'FW',
            'vspn_xtyy',
            'SELECT d.DOCID, d.IMAGEFILEID, d.IMAGEFILENAME, d.DOCFILETYPE, d.VERSIONID, '
            'i.IMAGEFILENAME AS IMG_IMAGEFILENAME, i.IMAGEFILETYPE, i.FILESIZE, '
            'dd.DOCSUBJECT, dd.DOCEXTENDNAME, dd.DOCCREATEDATE, dd.DOCCREATETIME '
            'FROM docimagefile d '
            'LEFT JOIN imagefile i ON i.IMAGEFILEID = d.IMAGEFILEID '
            'LEFT JOIN docdetail dd ON dd.ID = d.DOCID '
            f'WHERE d.DOCID IN ({c.in_placeholders(batch)}) '
            'ORDER BY d.DOCID, d.IMAGEFILEID',
            batch,
        )
        for _, row in df.iterrows():
            docid = _to_int_id(row['DOCID'])
            imagefileid = _to_int_id(row['IMAGEFILEID'])
            if docid is None or imagefileid is None:
                continue
            docimage_map.setdefault(docid, []).append({
                'docid': docid,
                'imagefileid': imagefileid,
                'filename': _text(row.get('IMAGEFILENAME')),
                'docfiletype': _text(row.get('DOCFILETYPE')),
                'versionid': _text(row.get('VERSIONID')),
            })
            if imagefileid not in imagefile_map:
                imagefile_map[imagefileid] = {
                    'filename': _text(row.get('IMG_IMAGEFILENAME')),
                    'imagefiletype': _text(row.get('IMAGEFILETYPE')),
                    'filesize': _text(row.get('FILESIZE')),
                }
            if docid not in docdetail_map:
                docdetail_map[docid] = {
                    'subject': html.unescape(_text(row.get('DOCSUBJECT'))),
                    'extend_name': _text(row.get('DOCEXTENDNAME')),
                    'created_date': _text(row.get('DOCCREATEDATE')),
                    'created_time': _text(row.get('DOCCREATETIME')),
                }
    return docimage_map, imagefile_map, docdetail_map


def _load_workflow_shared_docids(request_ids):
    mapping = {}
    request_ids = c.clean_codes(request_id for request_id in request_ids if _text(request_id))
    if not request_ids:
        return mapping
    for batch in _chunked(sorted(set(request_ids))):
        df = c.query_db(
            'FW',
            'vspn_xtyy',
            'SELECT REQUESTID, DOCID, NODEID, id AS share_id '
            'FROM workflow_docshareinfo '
            f'WHERE REQUESTID IN ({c.in_placeholders(batch)}) '
            'ORDER BY REQUESTID, id',
            batch,
        )
        for _, row in df.iterrows():
            request_id = c.format_code(row['REQUESTID'])
            docid = _to_int_id(row['DOCID'])
            if not (request_id and docid is not None):
                continue
            mapping.setdefault(request_id, []).append({
                'docid': docid,
                'nodeid': _text(row.get('NODEID')),
                'share_id': _to_int_id(row.get('share_id')) or 0,
            })
    for request_id, items in list(mapping.items()):
        if not items:
            continue
        latest = max(items, key=lambda item: item.get('share_id') or 0)
        latest_nodeid = latest.get('nodeid')
        if latest_nodeid:
            mapping[request_id] = [item for item in items if item.get('nodeid') == latest_nodeid]
    return mapping


def _classify_attachment_type(doc_row, image_info, doc_info, explicit_type=''):
    # 主取数(审批流程表单三字段)已带固定归类, 直接采用。
    if explicit_type:
        return explicit_type
    # 仅 docshareinfo 兜底走到这里: 按文件名关键字判断, 默认归初稿/未分类。
    filename = _first_non_blank(doc_row.get('filename'), image_info.get('filename'), doc_info.get('subject'))
    normalized = _normalize_field_name(filename).lower()
    if '生效' in normalized:
        return ATTACHMENT_TYPE_EFFECTIVE
    if '签署' in normalized or '签定' in normalized or '签订' in normalized:
        return ATTACHMENT_TYPE_SIGNED
    return ATTACHMENT_TYPE_DRAFT


def _preferred_attachment_type(source):
    approval_status = _text(source.get('合同审批状态'))
    if any(keyword in approval_status for keyword in SIGNED_APPROVAL_NODE_KEYWORDS):
        return ATTACHMENT_TYPE_SIGNED
    status_text = _text(source.get('合同签署状态'))
    status_id = c.format_code(source.get('合同签署状态ID'))
    if '归档' in status_text or status_id == '2':
        return ATTACHMENT_TYPE_EFFECTIVE
    return ATTACHMENT_TYPE_SIGNED


def _attachment_name_startswith_contract(attachment_name, contract_number):
    normalized_name = _normalize_field_name(attachment_name).upper()
    normalized_contract = _normalize_field_name(contract_number).upper()
    return bool(normalized_contract and normalized_name.startswith(normalized_contract))


def _attachment_skip_reason(source, attachment_type):
    preferred_type = _preferred_attachment_type(source)
    if attachment_type != preferred_type:
        return f'跳过非目标稿件:{attachment_type or "未分类"},目标:{preferred_type}'
    return ''


def _is_archived_contract(source):
    status_text = _text(source.get('合同签署状态'))
    status_id = c.format_code(source.get('合同签署状态ID'))
    return '归档' in status_text or status_id == '2'


def _assign_attachment_targets(candidate_rows, retention_mode='legacy'):
    """按合同维度分配目标附件类型。

    legacy 规则:
      - 文件名以合同编号开头 -> 合同附件
      - 若目标稿件只有一个文件,即便不以合同编号开头,也作为合同附件
      - 若目标稿件多个文件且都不以合同编号开头,第一个作为合同附件,其余作为其他附件
      - 其余不以合同编号开头的文件作为其他附件

    main_archive 规则:
      - 已归档合同取生效版;主文件落「主文件」,归档扫描件落「归档扫描件」。
      - 非归档/用印等合同取签署版;只落主文件和其他附件,不落归档扫描件。
      - 文件名以合同编号开头时只保留第一条作为主文件;未以合同编号开头的文件为其他附件。
      - 没有合同编号开头文件时,单文件或多文件第一条作为主文件,多文件其余为其他附件。
    """
    grouped = {}
    for row in candidate_rows:
        grouped.setdefault(_text(row.get('contract_number（合同编码）')), []).append(row)

    if retention_mode == 'main_archive':
        return _assign_attachment_targets_main_archive(grouped)

    assigned = []
    for contract_number, items in grouped.items():
        starts = [
            _attachment_name_startswith_contract(item.get('attachment_name'), contract_number)
            for item in items
        ]
        any_starts = any(starts)
        for index, item in enumerate(items):
            if starts[index]:
                item['attachment_sheet'] = SHEET_CONTRACT_ATTACHMENT
                item['attachment_rule'] = '目标稿件且文件名以合同编号开头'
            elif len(items) == 1:
                item['attachment_sheet'] = SHEET_CONTRACT_ATTACHMENT
                item['attachment_rule'] = '目标稿件仅1个文件,虽未以合同编号开头仍作为合同附件'
            elif not any_starts and index == 0:
                item['attachment_sheet'] = SHEET_CONTRACT_ATTACHMENT
                item['attachment_rule'] = '目标稿件多个文件均未以合同编号开头,取第一个作为合同附件'
            else:
                item['attachment_sheet'] = SHEET_OTHER_ATTACHMENT
                item['attachment_rule'] = '目标稿件但文件名未以合同编号开头'
            assigned.append(item)
    return assigned


def _assign_attachment_targets_main_archive(grouped):
    assigned = []
    for contract_number, items in grouped.items():
        if not items:
            continue
        starts = [
            _attachment_name_startswith_contract(item.get('attachment_name'), contract_number)
            for item in items
        ]
        start_indexes = [index for index, startswith_contract in enumerate(starts) if startswith_contract]
        main_index = start_indexes[0] if start_indexes else 0
        archived = _is_archived_contract(items[main_index].get('_source', {}))

        for index, item in enumerate(items):
            if index == main_index:
                main_item = item.copy()
                main_item['attachment_sheet'] = ATTACHMENT_FOLDER_MAIN
                if starts[index]:
                    main_item['attachment_rule'] = '目标稿件且文件名以合同编号开头,保留1条作为主文件'
                elif len(items) == 1:
                    main_item['attachment_rule'] = '目标稿件仅1个文件,虽未以合同编号开头仍作为主文件'
                else:
                    main_item['attachment_rule'] = '目标稿件多个文件均未以合同编号开头,取第一个作为主文件'
                assigned.append(main_item)

                if archived:
                    archive_item = item.copy()
                    archive_item['attachment_sheet'] = ATTACHMENT_FOLDER_ARCHIVE_SCAN
                    archive_item['attachment_rule'] = '已归档合同归档扫描件与主文件相同,保留1条'
                    assigned.append(archive_item)
                continue

            if starts[index]:
                # 同一合同编号开头的目标稿件只保留第一条作为主文件/归档扫描件。
                continue
            other_item = item.copy()
            other_item['attachment_sheet'] = SHEET_OTHER_ATTACHMENT
            other_item['attachment_rule'] = '目标稿件但文件名未以合同编号开头,作为其他附件'
            assigned.append(other_item)
    return assigned


def _resolve_request_form_tables(request_ids):
    """requestid -> 审批流程表单数据表名(formtable_main_*)。"""
    mapping = {}
    request_ids = c.clean_codes(request_id for request_id in request_ids if _text(request_id))
    if not request_ids:
        return mapping
    for batch in _chunked(sorted(set(request_ids))):
        df = c.query_db(
            'FW',
            'vspn_xtyy',
            'SELECT rb.requestid AS rid, bill.tablename AS tablename '
            'FROM workflow_requestbase rb '
            'JOIN workflow_base wb ON wb.id = rb.workflowid '
            'JOIN workflow_bill bill ON bill.id = wb.formid '
            f'WHERE rb.requestid IN ({c.in_placeholders(batch)})',
            batch,
        )
        for _, row in df.iterrows():
            request_id = c.format_code(row['rid'])
            tablename = _text(row.get('tablename'))
            if request_id and tablename:
                mapping[request_id] = tablename
    return mapping


def _existing_columns(tablename, candidates):
    df = c.query_db(
        'FW',
        'vspn_xtyy',
        'SELECT COLUMN_NAME FROM information_schema.COLUMNS '
        f'WHERE TABLE_NAME = %s AND COLUMN_NAME IN ({c.in_placeholders(candidates)})',
        [tablename, *candidates],
    )
    return {str(name).lower() for name in df['COLUMN_NAME'].tolist()}


def _load_form_doc_fields(request_form_tables):
    """requestid -> [(attachment_type, docid), ...],来自审批流程表单的稿件字段。"""
    result = {}
    requests_by_table = {}
    for request_id, tablename in request_form_tables.items():
        requests_by_table.setdefault(tablename, []).append(request_id)

    for tablename, request_ids in requests_by_table.items():
        field_types = FORM_DOC_FIELD_TYPES_BY_TABLE.get(tablename, FORM_DOC_FIELD_TYPES)
        typed_fields = [field for field, _ in field_types]
        type_by_field = dict(field_types)
        columns = _existing_columns(tablename, [*typed_fields, FORM_DOC_FALLBACK_FIELD])
        present_typed = [field for field in typed_fields if field in columns]
        use_fallback = not present_typed and FORM_DOC_FALLBACK_FIELD in columns
        select_fields = present_typed or ([FORM_DOC_FALLBACK_FIELD] if use_fallback else [])
        if not select_fields:
            continue
        select_clause = ', '.join(['requestid AS rid', *select_fields])
        for batch in _chunked(sorted(set(request_ids))):
            df = c.query_db(
                'FW',
                'vspn_xtyy',
                f'SELECT {select_clause} FROM {tablename} '
                f'WHERE requestid IN ({c.in_placeholders(batch)})',
                batch,
            )
            for _, row in df.iterrows():
                request_id = c.format_code(row['rid'])
                if not request_id:
                    continue
                entries = []
                for field in select_fields:
                    attachment_type = type_by_field.get(field, ATTACHMENT_TYPE_EFFECTIVE)
                    for docid in _split_docids(row.get(field)):
                        entries.append((attachment_type, docid))
                if entries:
                    result[request_id] = entries
    return result


def _collect_attachment_docrefs(source_df):
    raw_effective_docids = {}
    docrefs_by_contract = {}
    all_docids = []

    for source in source_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        raw_effective_docids[contract_number] = _text(source.get('合同附件DOCID', ''))
        docrefs_by_contract.setdefault(contract_number, [])

    contracts_with_form_docs = set()

    # 赛事(uf_htsp): 附件 docid 直接挂在 htcg/htqsg/htsxg 字段, 各自固定归类。
    is_saishi = source_df.get('数据来源', pd.Series('', index=source_df.index)).map(_text) == '泛微(赛事)'
    for source in source_df[is_saishi].to_dict('records'):
        contract_number = _text(source['合同编号'])
        for field, attachment_type in SAISHI_DOC_FIELD_TYPES:
            for docid in _split_docids(source.get(field)):
                docrefs_by_contract[contract_number].append({
                    'docid': docid,
                    'source': 'htsp_field',
                    'attachment_type': attachment_type,
                    'share_id': 0,
                })
                all_docids.append(docid)
                contracts_with_form_docs.add(contract_number)

    # MCN(uf_htk): 附件来自审批流程表单三字段(修订稿/签署版/生效版), 各自固定归类。
    mcn_df = source_df[~is_saishi]
    request_form_tables = _resolve_request_form_tables(mcn_df['合同流程ID'].map(c.format_code))
    form_doc_fields = _load_form_doc_fields(request_form_tables)
    for source in mcn_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        request_id = c.format_code(source.get('合同流程ID'))
        for attachment_type, docid in form_doc_fields.get(request_id, []):
            docrefs_by_contract[contract_number].append({
                'docid': docid,
                'source': 'form_field',
                'attachment_type': attachment_type,
                'share_id': 0,
            })
            all_docids.append(docid)
            contracts_with_form_docs.add(contract_number)

    # 兜底: 仅当合同三字段全空时, 才回退 workflow_docshareinfo(按文件名启发式归类)。
    fallback_request_ids = [
        c.format_code(source.get('合同流程ID'))
        for source in mcn_df.to_dict('records')
        if _text(source['合同编号']) not in contracts_with_form_docs
    ]
    workflow_docids = _load_workflow_shared_docids(fallback_request_ids)
    for source in mcn_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        if contract_number in contracts_with_form_docs:
            continue
        request_id = c.format_code(source.get('合同流程ID'))
        for item in workflow_docids.get(request_id, []):
            docid = item['docid']
            docrefs_by_contract[contract_number].append({
                'docid': docid,
                'source': 'workflow_docshareinfo',
                'attachment_type': '',
                'nodeid': item.get('nodeid', ''),
                'share_id': item.get('share_id', 0),
            })
            all_docids.append(docid)

    return docrefs_by_contract, raw_effective_docids, all_docids


def build_contract_attachment_manifest(source_df, retention_mode='legacy'):
    docrefs_by_contract, raw_effective_docids, all_docids = _timed(
        '  ├─附件docref收集(表单/赛事字段)', lambda: _collect_attachment_docrefs(source_df))
    print(f'[计时]   ├─附件 docid 总数: {len(all_docids)}', flush=True)
    docimage_map, imagefile_map, docdetail_map = _timed(
        '  ├─附件三表JOIN取数(_load_attachment_maps)', lambda: _load_attachment_maps(all_docids))

    candidate_rows = []
    missing_rows = []
    seen = set()
    used_paths = set()
    download_root = _attachment_download_root()

    _t0 = time.perf_counter()
    for source in source_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        contract_dir = _sanitize_path_part(contract_number, f'contract_{_text(source.get("ID"))}')
        fixed_docids = FIXED_CONTRACT_ATTACHMENT_DOCIDS.get(contract_number)
        contract_docrefs = docrefs_by_contract.get(contract_number, [])
        if fixed_docids:
            contract_docrefs = [dr for dr in contract_docrefs if dr['docid'] in fixed_docids]
        for docref in contract_docrefs:
            docid = docref['docid']
            doc_rows = docimage_map.get(docid, [])
            if not doc_rows:
                missing_rows.append({
                    'contract_number（合同编码）': contract_number,
                    '合同ID': _text(source.get('ID')),
                    '合同名称': _text(source.get('合同标题')),
                    'raw_docids': raw_effective_docids.get(contract_number, ''),
                    'docid': docid,
                    'source': docref.get('source', ''),
                    'status': 'missing_docimagefile',
                    'error': 'docimagefile 无记录',
                })
                continue

            for doc_row in doc_rows:
                imagefileid = doc_row['imagefileid']
                dedupe_key = (contract_number, docid, imagefileid)
                if dedupe_key in seen:
                    continue
                seen.add(dedupe_key)

                image_info = imagefile_map.get(imagefileid, {})
                doc_info = docdetail_map.get(docid, {})
                attachment_name = html.unescape(_first_non_blank(
                    doc_row.get('filename'),
                    image_info.get('filename'),
                    doc_info.get('subject'),
                    str(imagefileid),
                ))
                attachment_type = _classify_attachment_type(
                    doc_row, image_info, doc_info, docref.get('attachment_type', ''))
                skip_reason = '' if fixed_docids else _attachment_skip_reason(source, attachment_type)
                if skip_reason:
                    continue
                candidate_rows.append({
                    'contract_number（合同编码）': contract_number,
                    '合同ID': _text(source.get('ID')),
                    '合同名称': _text(source.get('合同标题')),
                    'attachment_type': attachment_type,
                    'raw_docids': raw_effective_docids.get(contract_number, ''),
                    'docid': docid,
                    'imagefileid': imagefileid,
                    'attachment_name': attachment_name,
                    'attachment_sheet': '',
                    'attachment_rule': '',
                    'imagefiletype': image_info.get('imagefiletype', ''),
                    'filesize': image_info.get('filesize', ''),
                    'target_path': '',
                    'source': docref.get('source', ''),
                    'nodeid': docref.get('nodeid', ''),
                    'share_id': docref.get('share_id', ''),
                    'doc_created_at': ' '.join(
                        item for item in [doc_info.get('created_date', ''), doc_info.get('created_time', '')] if item
                    ),
                    'status': 'pending',
                    'error': '',
                    '_source': source,
                })

    rows = _assign_attachment_targets(candidate_rows, retention_mode=retention_mode)
    for row in rows:
        row.pop('_source', None)
        contract_number = _text(row.get('contract_number（合同编码）'))
        contract_dir = _sanitize_path_part(contract_number, f'contract_{_text(row.get("合同ID"))}')
        folder_value = row.get('attachment_type')
        if retention_mode == 'main_archive':
            folder_value = row.get('attachment_sheet') or folder_value
        target_dir = download_root / contract_dir / _sanitize_path_part(folder_value, folder_value)
        target_name = _build_target_filename(row.get('attachment_name'), row.get('imagefileid'))
        target_path = _unique_attachment_path_preserve_name(
            target_dir, target_name, row.get('imagefileid'), used_paths)
        row['target_path'] = str(target_path)

    print(f'[计时]   └─附件清单逐行构建({len(rows)}行): {time.perf_counter() - _t0:.1f}s', flush=True)
    return pd.DataFrame(rows), pd.DataFrame(missing_rows)


def _sheet_headers(sheet_name):
    wb = load_workbook(TEMPLATE_FILE, read_only=True, data_only=True)
    ws = wb[sheet_name]
    return [_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]


def _template_headers():
    return {sheet_name: _sheet_headers(sheet_name) for sheet_name in OUTPUT_SHEETS}


def _timestamped_path(path):
    return path.with_name(f'{path.stem}_{datetime.now().strftime("%H%M%S")}{path.suffix}')


def _write_template_sheets_with_fallback(template_file, output_file, sheet_to_df, extra_sheets=None):
    try:
        return c.write_template_sheets(template_file, output_file, sheet_to_df, extra_sheets)
    except PermissionError:
        fallback = _timestamped_path(output_file)
        print(f'输出文件被占用,改写到: {fallback}')
        return c.write_template_sheets(template_file, fallback, sheet_to_df, extra_sheets)


def _write_exceptions_with_fallback(output_file, sheets):
    try:
        return c.write_exceptions(output_file, sheets)
    except PermissionError:
        fallback = _timestamped_path(output_file)
        print(f'未匹配清单被占用,改写到: {fallback}')
        return c.write_exceptions(fallback, sheets)


def _new_row(headers):
    return {header: '' for header in headers}


def _set(row, field_name, value):
    normalized = _normalize_field_name(field_name)
    for header in row:
        if _normalize_field_name(header) == normalized:
            row[header] = value
            return
    raise KeyError(f'模板缺少字段: {field_name}')


def _read_general_required_rules(headers_by_sheet):
    raw = pd.read_csv(RULE_CSV, encoding='utf-8-sig').iloc[1:, :15].copy()
    raw.columns = [
        'module', 'flow', 'table_name', 'field_name', 'required', 'enum', 'note',
        'event_table', 'event_field', 'event_enum', 'event_note',
        'mcn_table', 'mcn_field', 'mcn_enum', 'mcn_note',
    ]
    for column in raw.columns:
        raw[column] = raw[column].where(raw[column].notna(), '')
    raw['table_name'] = raw['table_name'].replace('', pd.NA).ffill().fillna('')
    raw = raw[raw['flow'] == '一般流程']

    table_to_sheet = {
        '字段模板（主表）': SHEET_MAIN,
        '关联合同': SHEET_RELATION,
        '相关单据-订单信息': SHEET_RELATED_ORDER,
        '采购申请': SHEET_PURCHASE_REQUEST,
        '订单信息明细': SHEET_ORDER_DETAIL,
        '对方信息': SHEET_COUNTERPARTY,
        '我方主体列表': SHEET_OUR_PARTY,
        '付款计划': SHEET_PAYMENT_PLAN,
        '收款计划': SHEET_COLLECTION_PLAN,
        '合同附件': SHEET_CONTRACT_ATTACHMENT,
        '其他附件': SHEET_OTHER_ATTACHMENT,
    }
    required = {sheet_name: [] for sheet_name in headers_by_sheet}
    remarks = {sheet_name: {} for sheet_name in headers_by_sheet}
    header_lookup = {
        sheet_name: {_normalize_field_name(header): header for header in headers}
        for sheet_name, headers in headers_by_sheet.items()
    }

    for _, rule in raw.iterrows():
        sheet_name = table_to_sheet.get(_text(rule['table_name']))
        if not sheet_name:
            continue
        normalized_field = _normalize_field_name(rule['field_name'])
        actual_field = header_lookup[sheet_name].get(normalized_field)
        if not actual_field:
            continue
        note = _first_non_blank(rule['note'], rule['event_note'], rule['mcn_note'])
        if note and actual_field not in remarks[sheet_name]:
            remarks[sheet_name][actual_field] = note
        if _text(rule['required']).upper() == 'Y' and actual_field not in required[sheet_name]:
            required[sheet_name].append(actual_field)
    return required, remarks


def _fill_summary(output_df, required_cols, remarks=None):
    total = len(output_df)
    rows = []
    for column in required_cols:
        column_exists = column in output_df.columns
        filled = int((output_df[column].astype(str).str.strip() != '').sum()) if column_exists else 0
        if filled < total:
            rows.append({
                '必输字段': column,
                '填充数': filled,
                '缺失数': total - filled,
                '总数': total,
                '填充率': '0.00%' if total == 0 else f'{filled / total * 100:.2f}%',
                '备注': (remarks or {}).get(column, ''),
            })
    return pd.DataFrame(rows, columns=['必输字段', '填充数', '缺失数', '总数', '填充率', '备注'])


def _collect_missing_details(output_df, source_df, required_cols, source_field_map, doc_col):
    sheets = {}
    total = len(output_df)
    if total == 0 or doc_col not in output_df.columns:
        return sheets
    source_by_doc = pd.DataFrame()
    source_doc_col = '合同编号'
    if source_doc_col in source_df.columns:
        source_by_doc = (
            source_df.assign(_doc_key=source_df[source_doc_col].map(_text))
            .drop_duplicates('_doc_key')
            .set_index('_doc_key')
        )
        # 下面 source_by_doc.at[...] 是逐行调用,摘掉 attrs 避免每次 deepcopy 大映射(含赛事收支计划)。
        source_by_doc.attrs = {}
    for column in required_cols:
        if column not in output_df.columns:
            continue
        blank_mask = output_df[column].astype(str).str.strip() == ''
        missing_count = int(blank_mask.sum())
        if not (0 < missing_count < total):
            continue
        missing_docs = output_df.loc[blank_mask, doc_col].map(_text)
        data = {doc_col: missing_docs}
        source_field = source_field_map.get(column)
        if source_field and source_field in source_df.columns and not source_by_doc.empty:
            data[f'泛微原表-{source_field}'] = missing_docs.map(
                lambda value: _text(source_by_doc.at[value, source_field]) if value in source_by_doc.index else ''
            )
        sheets[f'缺失_{column[:22]}'] = pd.DataFrame(data).drop_duplicates().reset_index(drop=True)
    return sheets


# ============================ 映射构建 ============================
def build_fw_company_info_map_for_values(company_values):
    company_ids = c.clean_codes(
        company_id
        for value in company_values
        for company_id in c.parse_browser_ids(value)
    )
    if not company_ids:
        return {}
    company_name_map = c.build_fw_company_name_map_for_ids(company_ids)
    entity_map = c.build_accounting_entity_map_for_names(company_name_map.values())
    return {
        company_id: {
            'name': company_name,
            'code': entity_map.get(c.normalize_name(company_name), ''),
        }
        for company_id, company_name in company_name_map.items()
    }


def build_customer_info_map_for_values(customer_values):
    customer_name_map = c.build_fw_customer_name_map_for_ids(customer_values)
    same_customer_map = c.load_same_customer_mapping(log_prefix='[合同迁移-一般流程]')
    source_to_target = {
        customer_id: c.resolve_same_customer_id(customer_id, same_customer_map)
        for customer_id in customer_name_map
    }
    target_name_map = c.build_fw_customer_name_map_for_ids(source_to_target.values())
    customer_by_target_id = c.build_hand_customer_info_by_ids(source_to_target.values())
    customer_code_map = c.build_customer_map_for_names(
        list(customer_name_map.values()) + list(target_name_map.values()))

    result = {}
    for customer_id, target_id in source_to_target.items():
        source_name = customer_name_map.get(customer_id, '')
        target_name = target_name_map.get(target_id, '')
        customer_info = customer_by_target_id.get(target_id, {})
        code = _first_non_blank(
            customer_info.get('code', ''),
            customer_code_map.get(c.normalize_name(target_name), ''),
            customer_code_map.get(c.normalize_name(source_name), ''),
        )
        name = _first_non_blank(customer_info.get('name', ''), target_name, source_name)
        result[customer_id] = {
            'name': name,
            'code': code,
            'source_name': source_name,
            'target_id': target_id,
            'match_method': '客户归并ID' if target_id != customer_id else customer_info.get('match_method', '客户名称'),
        }
    return result


def build_supplier_info_map_for_values(supplier_values):
    supplier_status_map = c.build_fw_supplier_status_map(supplier_values)
    same_supplier_map = c.load_same_supplier_mapping(log_prefix='[合同迁移-一般流程]')
    name_match_cache = c.load_supplier_vendor_name_match_map(log_prefix='[合同迁移-一般流程]')
    source_to_target = {
        supplier_id: c.resolve_same_supplier_id(supplier_id, same_supplier_map)
        for supplier_id in supplier_status_map
    }
    target_status_map = c.build_fw_supplier_status_map(source_to_target.values())
    vendor_by_target_id = c.build_hand_vendor_info_by_ids(source_to_target.values())

    vendor_by_source_id = {}
    names_to_match = []
    disabled_unmatched_ids = []
    for supplier_id, target_id in source_to_target.items():
        supplier_info = supplier_status_map.get(supplier_id, {})
        vendor_info = vendor_by_target_id.get(target_id, {})
        if vendor_info.get('code'):
            vendor_by_source_id[supplier_id] = vendor_info
            continue
        if supplier_info.get('status_code') != '0':
            cached_info = name_match_cache.get(supplier_id, {})
            if cached_info.get('code'):
                vendor_by_source_id[supplier_id] = cached_info
                continue
            supplier_name = supplier_info.get('name', '')
            if supplier_name:
                names_to_match.append(supplier_name)
                disabled_unmatched_ids.append(supplier_id)

    vendor_by_name = c.build_hand_vendor_info_by_names(names_to_match)
    discovered_rows = []
    for supplier_id in disabled_unmatched_ids:
        supplier_info = supplier_status_map.get(supplier_id, {})
        supplier_name = supplier_info.get('name', '')
        vendor_info = vendor_by_name.get(c.normalize_name(supplier_name), {})
        if not vendor_info.get('code'):
            continue
        vendor_by_source_id[supplier_id] = vendor_info
        discovered_rows.append({
            '供应商泛微Id': supplier_id,
            '供应商名称': supplier_name,
            '泛微状态': supplier_info.get('status_code', ''),
            '匹配方式': '禁用供应商按名称匹配汉得供应商',
            'handVendorId': vendor_info.get('id', ''),
            'handVendorCode': vendor_info.get('code', ''),
            'handVendorName': vendor_info.get('name', ''),
            'handTaxpayerName': vendor_info.get('taxpayer_name', ''),
        })
    changed_count = c.append_supplier_vendor_name_matches(discovered_rows)
    if changed_count:
        print(f'[合同迁移-一般流程] 禁用供应商按名称匹配汉得并写入缓存: {changed_count} 条')

    result = {}
    for supplier_id, target_id in source_to_target.items():
        vendor_info = vendor_by_source_id.get(supplier_id) or vendor_by_target_id.get(target_id, {})
        supplier_info = supplier_status_map.get(supplier_id, {})
        target_supplier_info = supplier_status_map.get(target_id) or target_status_map.get(target_id, {})
        target_name = target_supplier_info.get('name', '')
        match_method = vendor_info.get('match_method', '')
        if target_id != supplier_id and vendor_info.get('code'):
            match_method = '供应商归并ID'
        result[supplier_id] = {
            'name': _first_non_blank(vendor_info.get('name', ''), target_name, supplier_info.get('name', '')),
            'code': vendor_info.get('code', ''),
            'source_name': supplier_info.get('name', ''),
            'target_id': target_id,
            'target_name': target_name,
            'status_code': supplier_info.get('status_code', ''),
            'match_method': match_method or ('供应商名称缓存' if name_match_cache.get(supplier_id) else '供应商ID'),
        }
    return result


def build_contract_info_map_for_ids(contract_values, option_table=FW_TABLE):
    contract_ids = c.clean_codes(
        contract_id
        for value in contract_values
        for contract_id in c.parse_browser_ids(value)
    )
    if not contract_ids:
        return {}
    if option_table == FW_TABLE_HTSP:
        source_table = FW_TABLE_HTSP
        secondary_option_field = 'htejfl'
        select_sql = (
            'SELECT id, htbh, htmc AS htbt, htlx, htejfl AS htejlx, xmbh AS htszxmbh, xmmc AS htszxm '
            'FROM uf_htsp '
            f'WHERE id IN ({c.in_placeholders(contract_ids)})'
        )
    else:
        source_table = FW_TABLE
        secondary_option_field = 'htejlx'
        select_sql = (
            'SELECT id, htbh, htbt, htlx, htejlx, htszxmbh, htszxm '
            'FROM uf_htk '
            f'WHERE id IN ({c.in_placeholders(contract_ids)})'
        )
    contract_df = c.query_db(
        'FW',
        'vspn_xtyy',
        select_sql,
        contract_ids,
    )
    option_maps = c.build_fw_select_option_maps(source_table, ['htlx', secondary_option_field])
    result = {}
    for _, row in contract_df.iterrows():
        contract_id = c.format_code(row['id'])
        if not contract_id:
            continue
        type_id = c.format_code(row.get('htlx'))
        secondary_id = c.format_code(row.get('htejlx'))
        result[contract_id] = {
            'id': contract_id,
            'number': _text(row.get('htbh')),
            'title': _text(row.get('htbt')),
            'type_id': type_id,
            'type': option_maps['htlx'].get(type_id, ''),
            'secondary_type_id': secondary_id,
            'secondary_type': option_maps[secondary_option_field].get(secondary_id, ''),
            'project_id': _text(row.get('htszxmbh')),
            'project_name': _text(row.get('htszxm')),
            'source_table': source_table,
        }
    return result


def build_cost_center_platform_map(cost_center_values):
    """成本中心ID -> 一级成本中心(平台)名称。

    链路: 合同.cbzx -> uf_cbzx.id, uf_cbzx.sjcbzx -> uf_yjcbzx.id, uf_yjcbzx.mc。
    """
    cost_center_ids = c.clean_codes(
        cost_center_id
        for value in cost_center_values
        for cost_center_id in c.parse_browser_ids(value)
    )
    if not cost_center_ids:
        return {}
    result = {}
    for batch in _chunked(sorted(set(cost_center_ids))):
        df = c.query_db(
            'FW',
            'vspn_xtyy',
            'SELECT cbzx.id AS cbzx_id, yj.mc AS platform '
            'FROM uf_cbzx cbzx LEFT JOIN uf_yjcbzx yj ON yj.id = cbzx.sjcbzx '
            f'WHERE cbzx.id IN ({c.in_placeholders(batch)})',
            batch,
        )
        for _, row in df.iterrows():
            cost_center_id = c.format_code(row['cbzx_id'])
            platform = _text(row.get('platform'))
            if cost_center_id and platform:
                result[cost_center_id] = platform
    return result


def build_feishu_employee_id_map(code_values):
    """员工工号(V编号) -> 飞书 user_id。

    来源: 汉得 hfins_base.hfbs_employee.feishu_employee_id, 按 employee_code(员工编号)关联。
    泛微侧的工号取自 hrmjobtitles.JOBTITLENAME(即 `合同执行人员工号`), 与汉得 employee_code 同口径。
    """
    codes = c.clean_codes(_text(code) for code in code_values if _text(code))
    if not codes:
        return {}
    result = {}
    for batch in _chunked(sorted(set(codes))):
        df = c.query_db(
            'ZT',
            'hfins_base',
            'SELECT employee_code, feishu_employee_id FROM hfbs_employee '
            f'WHERE employee_code IN ({c.in_placeholders(batch)})',
            batch,
        )
        for _, row in df.iterrows():
            code = _text(row.get('employee_code'))
            feishu_id = _text(row.get('feishu_employee_id'))
            if code and feishu_id:
                result[code] = feishu_id
    return result


def build_employee_info_map_by_codes(code_values, status_by_number, status_by_name):
    """员工工号(V编号) -> 姓名/飞书 user_id/在离职状态。用于赛事订单申请人工号兜底。"""
    codes = c.clean_codes(_text(code) for code in code_values if _text(code))
    if not codes:
        return {}
    result = {}
    for batch in _chunked(sorted(set(codes))):
        df = c.query_db(
            'ZT',
            'hfins_base',
            'SELECT employee_code, name, feishu_employee_id FROM hfbs_employee '
            f'WHERE employee_code IN ({c.in_placeholders(batch)})',
            batch,
        )
        for _, row in df.iterrows():
            code = _text(row.get('employee_code'))
            if not code:
                continue
            name = _text(row.get('name'))
            result[code] = {
                'name': name,
                'code': code,
                'user_id': _text(row.get('feishu_employee_id')),
                'status': _employee_status_label(name, code, status_by_number, status_by_name),
            }
    return result


def build_feishu_employee_id_map_by_name(name_values):
    """员工姓名 -> 飞书 user_id。

    仅姓名唯一且 feishu_employee_id 唯一时返回,避免重名员工误配。
    """
    names = c.clean_text_values(_text(name) for name in name_values if _text(name))
    if not names:
        return {}
    rows = []
    for batch in _chunked(sorted(set(names))):
        df = c.query_db(
            'ZT',
            'hfins_base',
            'SELECT name, feishu_employee_id FROM hfbs_employee '
            f'WHERE name IN ({c.in_placeholders(batch)})',
            batch,
        )
        rows.extend(df.to_dict('records'))
    grouped = {}
    for row in rows:
        name = _text(row.get('name'))
        feishu_id = _text(row.get('feishu_employee_id'))
        if name and feishu_id:
            grouped.setdefault(name, set()).add(feishu_id)
    return {
        name: next(iter(ids))
        for name, ids in grouped.items()
        if len(ids) == 1
    }


def build_feishu_employee_id_map_by_contact(contact_values):
    """泛微登录名/手机号/邮箱 -> 飞书 user_id。仅唯一命中时返回。"""
    contacts = c.clean_text_values(_text(value) for value in contact_values if _text(value) and _text(value) != 'Default')
    if not contacts:
        return {}
    rows = []
    for batch in _chunked(sorted(set(contacts))):
        placeholders = c.in_placeholders(batch)
        df = c.query_db(
            'ZT',
            'hfins_base',
            'SELECT email, phone, login_identity, taxpayers_phone, feishu_employee_id FROM hfbs_employee '
            f'WHERE email IN ({placeholders}) '
            f'   OR phone IN ({placeholders}) '
            f'   OR login_identity IN ({placeholders}) '
            f'   OR taxpayers_phone IN ({placeholders})',
            [*batch, *batch, *batch, *batch],
        )
        rows.extend(df.to_dict('records'))
    grouped = {}
    for row in rows:
        feishu_id = _text(row.get('feishu_employee_id'))
        if not feishu_id:
            continue
        for field in ('email', 'phone', 'login_identity', 'taxpayers_phone'):
            contact = _text(row.get(field))
            if contact:
                grouped.setdefault(contact, set()).add(feishu_id)
    return {
        contact: next(iter(ids))
        for contact, ids in grouped.items()
        if len(ids) == 1
    }


def _employee_status_label(name, code, status_by_number, status_by_name):
    code = _text(code)
    name = _text(name)
    if code and code in status_by_number:
        return '在职' if status_by_number[code] == 'hired' else '离职'
    if name and name in status_by_name:
        return '在职' if status_by_name[name] == 'hired' else '离职'
    return ''


def _build_employee_info_by_name(names, status_by_number, status_by_name):
    unique_names = []
    for name in names:
        for item in c.split_person_names(name):
            if item and item not in unique_names:
                unique_names.append(item)
    if DEFAULT_CONTRACT_CREATOR_NAME not in unique_names:
        unique_names.append(DEFAULT_CONTRACT_CREATOR_NAME)
    if not unique_names:
        return {}
    employee_code_map = c.build_employee_code_map()
    name_to_code = {
        name: employee_code_map.get(c.normalize_name(name), '')
        for name in unique_names
    }
    feishu_id_map = build_feishu_employee_id_map(name_to_code.values())
    return {
        name: {
            'name': name,
            'code': code,
            'user_id': feishu_id_map.get(_text(code), ''),
            'status': _employee_status_label(name, code, status_by_number, status_by_name),
        }
        for name, code in name_to_code.items()
    }


def _cleanable_order_infos_for_source(row):
    order_code = _text(row.get('订单编号'))
    if order_code:
        if _is_saishi_source(row) and _text(row.get('订单映射来源')) == ORDER_INIT_MAPPING_SOURCE:
            order_names = _split_joined_field(row.get('订单名称'))
            project_candidates = c.clean_text_values([
                row.get('泛微项目编号'),
                row.get('项目编号'),
                row.get('清洗后项目编号'),
            ])
            return [
                {
                    '订单编号': code,
                    '订单标题': order_names[index] if index < len(order_names) else '',
                    '项目编号候选': project_candidates,
                    '项目经理候选': [],
                    '映射来源': ORDER_INIT_MAPPING_SOURCE,
                }
                for index, code in enumerate(_split_multi_values(order_code))
            ]
        infos = []
        for code in _split_multi_values(order_code):
            info = c.cleanable_order_info_for_order(code)
            if info:
                infos.append(info)
        if infos:
            return infos
    if _is_saishi_source(row):
        return []
    infos = []
    seen_orders = set()
    for project_code in (row.get('项目编号'), row.get('泛微项目编号'), row.get('清洗后项目编号')):
        for info in c.cleanable_order_infos_for_project(project_code):
            order = _text(info.get('订单编号'))
            if order and order not in seen_orders:
                seen_orders.add(order)
                infos.append(info)
    return infos


def _order_manager_names_for_source(row):
    names = []
    for info in _cleanable_order_infos_for_source(row):
        for name in info.get('项目经理候选', []):
            if name and name not in names:
                names.append(name)
    return names


def _first_split_value(value):
    values = _split_multi_values(value)
    return values[0] if values else ''


def _candidate_employee(name='', code='', user_id='', status='', source=''):
    return {
        'name': _text(name),
        'code': _text(code),
        'user_id': _text(user_id),
        'status': _text(status),
        'source': source,
    }


def _employee_is_left(info):
    return _text(info.get('status')) == '离职'


def _employee_is_present(info):
    return bool(_text(info.get('name')) or _text(info.get('code')) or _text(info.get('user_id')))


def _apply_saishi_contract_creator_rules(df, status_by_number, status_by_name):
    """赛事合同创建人:申请人 -> 合同执行人 -> 订单初始化表申请人员工编码。"""
    saishi_mask = df['数据来源'].map(_text).eq('泛微(赛事)')
    if not bool(saishi_mask.any()):
        return df

    order_applicant_codes = []
    for value in df.loc[saishi_mask, '订单申请人员工编码']:
        order_applicant_codes.extend(_split_multi_values(value))
    order_applicant_info = build_employee_info_map_by_codes(order_applicant_codes, status_by_number, status_by_name)

    for index, row in df.loc[saishi_mask].iterrows():
        applicant = _candidate_employee(
            row.get('申请人'),
            row.get('申请人工号'),
            row.get('申请人user_id'),
            row.get('申请人状态'),
            '合同申请人',
        )
        executor = _candidate_employee(
            row.get('合同执行人员'),
            row.get('合同执行人员工号'),
            row.get('合同执行人飞书ID'),
            row.get('合同执行人状态'),
            '合同执行人',
        )
        order_code = _first_split_value(row.get('订单申请人员工编码'))
        order_applicant = order_applicant_info.get(order_code, {
            'name': '',
            'code': order_code,
            'user_id': '',
            'status': _employee_status_label('', order_code, status_by_number, status_by_name),
            'source': '订单申请人员工编码',
        })
        order_applicant = _candidate_employee(
            order_applicant.get('name'),
            order_applicant.get('code'),
            order_applicant.get('user_id'),
            order_applicant.get('status'),
            '订单申请人员工编码',
        )

        selected = applicant
        reason = '赛事合同:取合同申请人'
        if (not _employee_is_present(selected)) or _employee_is_left(selected):
            selected = executor
            reason = '赛事合同:合同申请人离职/缺失,取合同执行人'
            if (not _employee_is_present(selected)) or _employee_is_left(selected):
                selected = order_applicant
                reason = '赛事合同:合同申请人与执行人离职/缺失,取订单申请人员工编码'

        df.at[index, '订单申请人员工编码'] = order_code
        df.at[index, '合同创建人调整方式'] = reason
        df.at[index, '合同创建人'] = selected.get('name', '')
        df.at[index, '合同创建人工号'] = selected.get('code', '')
        df.at[index, '合同创建人user_id'] = selected.get('user_id', '')
        df.at[index, '合同创建人状态'] = selected.get('status', '')
    return df


def _apply_contract_creator_rules(df, status_by_number, status_by_name):
    df['原合同创建人'] = df['合同创建人']
    df['原合同创建人工号'] = df['合同创建人工号']
    df['原合同创建人user_id'] = df['合同创建人user_id']
    df['原合同创建人状态'] = df['合同创建人状态']
    df['订单项目经理候选'] = ''
    df['合同创建人调整方式'] = '创建人在职/无需调整'
    df = _apply_saishi_contract_creator_rules(df, status_by_number, status_by_name)
    left_mask = (
        df['合同创建人状态'].map(_text).eq('离职')
        & ~df['数据来源'].map(_text).eq('泛微(赛事)')
    )
    if not bool(left_mask.any()):
        return df

    manager_names = []
    for _, row in df.loc[left_mask].iterrows():
        manager_names.extend(_order_manager_names_for_source(row))
    employee_info_by_name = _build_employee_info_by_name(manager_names, status_by_number, status_by_name)
    default_info = employee_info_by_name.get(DEFAULT_CONTRACT_CREATOR_NAME, {
        'name': DEFAULT_CONTRACT_CREATOR_NAME,
        'code': '',
        'user_id': '',
        'status': _employee_status_label(DEFAULT_CONTRACT_CREATOR_NAME, '', status_by_number, status_by_name),
    })

    for index, row in df.loc[left_mask].iterrows():
        manager_names = _order_manager_names_for_source(row)
        df.at[index, '订单项目经理候选'] = '、'.join(manager_names)
        selected = None
        for manager_name in manager_names:
            info = employee_info_by_name.get(manager_name)
            if info and info.get('status') == '在职' and info.get('user_id'):
                selected = info
                break
        if selected:
            df.at[index, '合同创建人调整方式'] = '原创建人离职,更新为订单项目经理'
        else:
            selected = default_info
            df.at[index, '合同创建人调整方式'] = '原创建人离职,项目经理离职/缺失,更新为黄劭文'
        df.at[index, '合同创建人'] = selected.get('name', '')
        df.at[index, '合同创建人工号'] = selected.get('code', '')
        df.at[index, '合同创建人user_id'] = selected.get('user_id', '')
        df.at[index, '合同创建人状态'] = selected.get('status', '')
    return df


def build_purchase_request_code_map(values):
    """采购申请对象ID(uf_htsp.cgsqddx -> uf_cgspxx.id) -> 采购申请单编号(uf_cgspxx.dh)。"""
    request_ids = c.clean_codes(
        request_id
        for value in values
        for request_id in c.parse_browser_ids(value)
    )
    if not request_ids:
        return {}
    result = {}
    for batch in _chunked(sorted(set(request_ids))):
        df = c.query_db(
            'FW',
            'vspn_xtyy',
            f'SELECT id, dh FROM uf_cgspxx WHERE id IN ({c.in_placeholders(batch)})',
            batch,
        )
        for _, row in df.iterrows():
            request_id = c.format_code(row['id'])
            code = _text(row.get('dh'))
            if request_id and code:
                result[request_id] = code
    return result


# ============================ 分类 / 金额 ============================
def _contract_prefix(contract_number):
    text = _text(contract_number).upper()
    match = re.search(r'H-[A-Z]+', text)
    return match.group(0) if match else ''


# 框架合同: 合同类型缩写 KF/KS/KJ。无独立缩写码表, 用合同编号前缀(H-KF/H-KS/H-KJ)判定。
FRAMEWORK_CONTRACT_PREFIXES = ('H-KF', 'H-KS', 'H-KJ')


def _is_framework_contract(contract_number):
    return _contract_prefix(contract_number) in FRAMEWORK_CONTRACT_PREFIXES


def _mcn_pay_letter(contract_number):
    """MCN 代字式编号(主体缩写-S/F/O…)的收支字母。S=收入 F=支出 O=其他。"""
    text = _text(contract_number).upper()
    for segment in text.split('-'):
        match = re.fullmatch(r'([SFO])\d*', segment.strip())
        if match:
            return match.group(1)
    return ''


def _contains_any(text, keywords):
    text = _text(text)
    folded = text.casefold()
    return any(keyword and keyword.casefold() in folded for keyword in keywords)


# 一级成本中心(平台) -> 业务域。MD「合同数据迁移-二级分类映射规则」以成本中心为准。
COST_CENTER_PLATFORM_AREAS = {
    '国内赛事业务平台': 'event',
    '赛事商务产品平台': 'event',
    '海外业务平台': 'overseas',
    '全球战略平台': 'overseas',
    '全球生态事业部': 'overseas',
    '资产技术平台': 'asset',
    '商业化事业部': 'commercial',
    'IP衍生事业部': 'derivative',
    '视效事业部': 'visual',
    '虚拟体育事业部': 'visual',
    '运营人力平台': 'admin',
    '采购中心': 'admin',
}


def _business_area(row):
    # 优先用成本中心平台(MD 规则口径), 命中即返回。
    platform = _text(row.get('成本中心平台'))
    if platform:
        for name, area in COST_CENTER_PLATFORM_AREAS.items():
            if name in platform:
                return area
    # 兜底: 成本中心未命中时, 退回项目/标题关键字推断。
    corpus = ' '.join([
        _text(row.get('合同所属项目')),
        _text(row.get('项目名称')),
        _text(row.get('合同标题')),
        _text(row.get('合同二级类型')),
    ])
    if _contains_any(corpus, ['海外', '全球']):
        return 'overseas'
    if _contains_any(corpus, ['资管', '资产技术', '资产采购', '资产租赁', '设备采购', '设备租赁']):
        return 'asset'
    if _contains_any(corpus, ['商业化', '广告', '赞助']):
        return 'commercial'
    if _contains_any(corpus, ['衍生', 'IP']):
        return 'derivative'
    if _contains_any(corpus, ['视效', '视频制作', '视觉']):
        return 'visual'
    if _contains_any(corpus, ['运营人力', '行政', '人力', 'IT', '采购中心']):
        return 'admin'
    if _contains_any(corpus, ['赛事', '活动', '赛']):
        return 'event'
    return ''


def _area_secondary(area, pay_side, contract_shape):
    names = {
        ('event', '支出', '单次'): '国内赛事及活动支出',
        ('overseas', '支出', '单次'): '海外赛事业及活动支出',
        ('asset', '支出', '单次'): '资产采购租赁及经营支出',
        ('commercial', '支出', '单次'): '广告赞助支出',
        ('derivative', '支出', '单次'): '衍生品支出',
        ('visual', '支出', '单次'): '视效支出',
        ('admin', '支出', '单次'): '行政运营及人力支出',
        ('event', '支出', '框架'): '赛事及活动支出框架',
        ('overseas', '支出', '框架'): '赛事及活动支出框架',
        ('admin', '支出', '框架'): '赛事及活动支出框架',
        ('asset', '支出', '框架'): '资产采购租赁及经营支出框架',
        ('commercial', '支出', '框架'): '广告赞助支出框架',
        ('derivative', '支出', '框架'): '衍生品支出框架',
        ('visual', '支出', '框架'): '视效支出框架',
        ('event', '支出', '订单'): '赛事及活动支出订单',
        ('overseas', '支出', '订单'): '赛事及活动支出订单',
        ('asset', '支出', '订单'): '资产采购租赁及经营支出订单',
        ('commercial', '支出', '订单'): '广告赞助支出订单',
        ('derivative', '支出', '订单'): '衍生品支出订单',
        ('visual', '支出', '订单'): '视效支出订单',
        ('admin', '支出', '订单'): '行政运营及人力支出订单',
        ('event', '收入', '单次'): '赛事及活动收入',
        ('overseas', '收入', '单次'): '赛事及活动收入',
        ('asset', '收入', '单次'): '资产采购租赁及经营收入',
        ('commercial', '收入', '单次'): '广告赞助收入',
        ('derivative', '收入', '单次'): '衍生品收入',
        ('visual', '收入', '单次'): '视效收入',
        ('event', '收入', '框架'): '赛事及活动收入框架',
        ('overseas', '收入', '框架'): '赛事及活动收入框架',
        ('asset', '收入', '框架'): '资产采购租赁及经营收入',
        ('admin', '收入', '框架'): '资产采购租赁及经营收入',
        ('commercial', '收入', '框架'): '广告赞助收入',
        ('derivative', '收入', '框架'): '衍生品收入',
        ('visual', '收入', '框架'): '视效收入',
        ('event', '收入', '订单'): '赛事及活动收入订单',
        ('overseas', '收入', '订单'): '赛事及活动收入订单',
        ('asset', '收入', '订单'): '资产采购租赁及经营收入',
        ('admin', '收入', '订单'): '资产采购租赁及经营收入',
        ('commercial', '收入', '订单'): '广告赞助收入订单',
        ('derivative', '收入', '订单'): '衍生品收入订单',
        ('visual', '收入', '订单'): '视效收入订单',
    }
    return names.get((area, pay_side, contract_shape), '')


def resolve_contract_category(row):
    if _is_force_included_contract(row):
        return FORCED_GENERAL_CONTRACT_CATEGORY

    prefix = _contract_prefix(row.get('合同编号'))
    title = _text(row.get('合同标题'))
    supplier_name = _text(row.get('合同供应商名称'))
    area = _business_area(row)

    if prefix in ('H-DF', 'H-F'):
        if '云账户' in supplier_name or _contains_any(title, ['个人合作']):
            return '单次支出-个人合作支出'
        if _contains_any(title, ['奖金', '补贴', 'prize', 'subsidy']):
            return '单次支出-奖金补贴支出'
        return '单次支出-' + (_area_secondary(area, '支出', '单次') or '其他支出')
    if prefix == 'H-KF':
        return '框架支出-' + (_area_secondary(area, '支出', '框架') or '其他支出框架')
    if prefix == 'H-OF':
        return '订单支出-' + (_area_secondary(area, '支出', '订单') or '其他支出订单')
    if prefix in ('H-DS', 'H-S'):
        return '单次收入-' + (_area_secondary(area, '收入', '单次') or '其他收入')
    if prefix == 'H-KS':
        return '框架收入-' + (_area_secondary(area, '收入', '框架') or '其他收入')
    if prefix == 'H-OS':
        return '订单收入-' + (_area_secondary(area, '收入', '订单') or '其他收入')
    if prefix == 'H-DJ':
        return '单次收支-通用单次收支'
    if prefix == 'H-KJ':
        return '框架收支-通用框架收支'
    if prefix == 'H-N':
        return '内部合同-内部结转'
    if prefix == 'H-P' and '保密' in title:
        return '其他-保密协议'
    if prefix == 'H-P' and '反商业贿赂' in title:
        return '其他-反商业贿赂协议'
    if prefix in ('H-P', 'H-PZ'):
        return '其他-战略合作协议'
    if prefix == 'H-Q':
        return '其他-其他类型'

    # MCN 代字式编号(无 H- 前缀): 用代字 S/F 字母定收支, 成本中心平台细分二级。
    letter = _mcn_pay_letter(row.get('合同编号'))
    if letter == 'F':
        return '单次支出-' + (_area_secondary(area, '支出', '单次') or '其他支出')
    if letter == 'S':
        return '单次收入-' + (_area_secondary(area, '收入', '单次') or '其他收入')

    contract_type = _text(row.get('合同类型')) + _text(row.get('合同二级类型'))
    if '收入' in contract_type and '支出' not in contract_type:
        return '单次收入-其他收入'
    if '支出' in contract_type and '收入' not in contract_type:
        return '单次支出-其他支出'
    return '其他-其他类型'


def resolve_contract_category_basis(row):
    if _is_force_included_contract(row):
        return f'强制追加导出合同; 智书框架合同类型固定为{FORCED_GENERAL_CONTRACT_CATEGORY}'

    prefix = _contract_prefix(row.get('合同编号')) or '空'
    area = _business_area(row) or '未命中成本中心/项目关键词'
    return (
        f'合同编号前缀={prefix}; 成本中心平台={_text(row.get("成本中心平台")) or "空"}; '
        f'合同类型={_text(row.get("合同类型")) or "空"}; '
        f'合同二级类型={_text(row.get("合同二级类型")) or "空"}; '
        f'分类线索={area}; 规则来源=合同数据迁移-二级分类映射规则'
    )


def resolve_pay_type(row):
    prefix = _contract_prefix(row.get('合同编号'))
    if prefix in ('H-DS', 'H-KS', 'H-OS', 'H-S'):
        return '收入类'
    if prefix in ('H-DF', 'H-KF', 'H-OF', 'H-F'):
        return '支出类'
    if prefix in ('H-DJ', 'H-KJ', 'H-N'):
        return '既收又支'
    if prefix in ('H-P', 'H-PZ', 'H-Q'):
        return '无金额'

    letter = _mcn_pay_letter(row.get('合同编号'))
    if letter == 'S':
        return '收入类'
    if letter == 'F':
        return '支出类'

    in_amount = abs(_number(row.get('合同预计收入')))
    out_amount = abs(_number(row.get('合同预计支出')))
    if in_amount and out_amount:
        return '既收又支'
    if in_amount:
        return '收入类'
    if out_amount:
        return '支出类'
    return '无金额'


def resolve_amounts(row):
    pay_type = _text(row.get('收支类型'))
    contract_amount = abs(_number(row.get('合同金额')))
    in_amount = abs(_number(row.get('合同预计收入')))
    out_amount = abs(_number(row.get('合同预计支出')))
    if pay_type == '收入类' and not in_amount:
        in_amount = contract_amount
    if pay_type == '支出类' and not out_amount:
        out_amount = contract_amount
    if pay_type == '既收又支' and not (in_amount or out_amount) and contract_amount:
        in_amount = contract_amount
    if pay_type == '收入类':
        amount = in_amount
    elif pay_type == '支出类':
        amount = out_amount
    elif pay_type == '既收又支':
        amount = contract_amount or in_amount + out_amount
    else:
        amount = 0
    return round(amount, 2), round(in_amount, 2), round(out_amount, 2)


def resolve_signed_amounts(row):
    pay_type = _text(row.get('收支类型'))
    contract_amount = _number(row.get('合同金额'))
    in_amount = _number(row.get('合同预计收入'))
    out_amount = _number(row.get('合同预计支出'))
    if pay_type == '收入类' and (not in_amount or contract_amount < 0):
        in_amount = contract_amount
    if pay_type == '支出类' and (not out_amount or contract_amount < 0):
        out_amount = contract_amount
    if pay_type == '既收又支' and not (in_amount or out_amount) and contract_amount:
        in_amount = contract_amount
    if pay_type == '收入类':
        amount = in_amount
    elif pay_type == '支出类':
        amount = out_amount
    elif pay_type == '既收又支':
        amount = contract_amount or in_amount + out_amount
    else:
        amount = 0
    return round(amount, 2), round(in_amount, 2), round(out_amount, 2)


def _main_contract_code_candidates(contract_number):
    code = _text(contract_number)
    candidates = []
    while code:
        next_code = re.sub(r'-(?:S\d+|N)$', '', code, flags=re.IGNORECASE)
        if next_code == code:
            break
        if next_code and next_code not in candidates:
            candidates.append(next_code)
        code = next_code
    return list(reversed(candidates))


def _is_supplement_contract_number(contract_number):
    return bool(re.search(r'-(?:S\d+|N)$', _text(contract_number), flags=re.IGNORECASE))


def _apply_supplement_amount_rollup(df):
    # pandas 3.0 在每次列/行 boxing 时会 deepcopy df.attrs;本函数有逐行 iterrows + df.at,
    # 带着 read_source 注入的大映射(attrs)会让本步从几秒劣化到数千秒。
    # 这里临时摘掉 attrs、末尾恢复(本函数只按列计算,不读取 attrs)。
    saved_attrs = df.attrs
    df.attrs = {}
    contract_numbers = {_text(value) for value in df['合同编号']}
    df['主合同编号'] = ''
    df['是否补充协议'] = ''
    df['金额汇总目标合同编号'] = df['合同编号'].map(_text)
    for index, row in df.iterrows():
        contract_number = _text(row.get('合同编号'))
        for main_code in _main_contract_code_candidates(contract_number):
            if main_code in contract_numbers:
                df.at[index, '主合同编号'] = main_code
                df.at[index, '是否补充协议'] = 'Y'
                df.at[index, '金额汇总目标合同编号'] = main_code
                break

    payment_sum = {}
    collection_sum = {}
    child_count = {}
    child_codes = {}
    for _, row in df.iterrows():
        target = _text(row.get('金额汇总目标合同编号')) or _text(row.get('合同编号'))
        contract_number = _text(row.get('合同编号'))
        payment_sum[target] = payment_sum.get(target, 0.0) + _number(row.get('支出总额_签名'))
        collection_sum[target] = collection_sum.get(target, 0.0) + _number(row.get('收入总额_签名'))
        if target != contract_number:
            child_count[target] = child_count.get(target, 0) + 1
            child_codes.setdefault(target, []).append(contract_number)

    df['付款计划汇总金额'] = df['合同编号'].map(lambda value: round(payment_sum.get(_text(value), 0.0), 2))
    df['收款计划汇总金额'] = df['合同编号'].map(lambda value: round(collection_sum.get(_text(value), 0.0), 2))
    df['补充协议数量'] = df['合同编号'].map(lambda value: child_count.get(_text(value), 0))
    df['补充协议编号'] = df['合同编号'].map(lambda value: '、'.join(child_codes.get(_text(value), [])))
    df.attrs = saved_attrs
    return df


# ============================ 专项品类映射 ============================
_SPECIAL_CATEGORY_CACHE = None


def _strip_category_suffix(value):
    """去掉终版二级分类的英文后缀, 如 '市场调研及数据分析-RES' -> '市场调研及数据分析'。"""
    return re.sub(r'-[A-Za-z]+$', '', _text(value))


def build_special_category_path_map():
    """泛微采购品类树(uf_xt_cgpldy): 叶子id -> [一级, 二级, 三级] 名称链(沿 sjpl 上溯)。"""
    tree = c.query_db('FW', 'vspn_xtyy', f'SELECT id, plmc, sjpl FROM {SPECIAL_CATEGORY_TABLE}')
    name = {}
    parent = {}
    for _, row in tree.iterrows():
        node_id = c.format_code(row['id'])
        if not node_id:
            continue
        name[node_id] = _text(row['plmc'])
        parent[node_id] = c.format_code(row['sjpl']) or None
    path_map = {}
    for node_id in name:
        parts = []
        cur = node_id
        seen = set()
        while cur and cur in name and cur not in seen:
            seen.add(cur)
            parts.append(name[cur])
            cur = parent.get(cur)
        path_map[node_id] = list(reversed(parts))
    return path_map


def load_special_category_final_mapping():
    """供应商分类【新旧mapping】CSV: 原一级/二级/三级 -> 终版二级分类(去英文后缀)。

    返回 (full, ac):
      full —— '一级/二级/三级' 精确键(无冲突);
      ac   —— '一级/三级' 回退键(剔除冲突项), 用于吸收海外分支二级命名漂移。
    """
    raw = pd.read_csv(SPECIAL_CATEGORY_CSV, header=None, skiprows=2, encoding='utf-8-sig')
    full = {}
    ac = {}
    ac_conflict = set()
    for _, row in raw.iterrows():
        first, second, third = _text(row[0]), _text(row[1]), _text(row[2])
        final_label = _strip_category_suffix(row[6])
        if not first or not final_label:
            continue
        full[f'{first}/{second}/{third}'] = final_label
        ac_key = f'{first}/{third}'
        if ac_key in ac and ac[ac_key] != final_label:
            ac_conflict.add(ac_key)
        ac[ac_key] = final_label
    for key in ac_conflict:
        ac.pop(key, None)
    return full, ac


def _special_category_maps():
    global _SPECIAL_CATEGORY_CACHE
    if _SPECIAL_CATEGORY_CACHE is None:
        path_map = build_special_category_path_map()
        full, ac = load_special_category_final_mapping()
        _SPECIAL_CATEGORY_CACHE = (path_map, full, ac)
    return _SPECIAL_CATEGORY_CACHE


_EMPLOYMENT_STATUS_CACHE = None


def _employment_status_map():
    """飞书在职/离职映射; 缺凭据或接口异常时降级为空映射(申请人状态留空), 不中断迁移。"""
    global _EMPLOYMENT_STATUS_CACHE
    if _EMPLOYMENT_STATUS_CACHE is None:
        try:
            _EMPLOYMENT_STATUS_CACHE = feishu.get_employee_status_maps()
            by_number, by_name = _EMPLOYMENT_STATUS_CACHE
            print(f'[合同迁移-一般流程] 飞书员工状态: 工号 {len(by_number)} / 唯一姓名 {len(by_name)}')
        except Exception as error:
            print(f'[合同迁移-一般流程] 飞书员工状态获取失败, 申请人状态留空: {error}')
            _EMPLOYMENT_STATUS_CACHE = ({}, {})
    return _EMPLOYMENT_STATUS_CACHE


def resolve_special_category(code):
    """专项分类编码(zxflcg, 形如 '45_60', 末段为品类树叶子id) -> 终版二级分类。"""
    code = _text(code)
    if '_' not in code:
        return ''
    path_map, full, ac = _special_category_maps()
    leaf = c.format_code(code.rsplit('_', 1)[-1])
    parts = path_map.get(leaf)
    if not parts:
        return ''
    if '/'.join(parts) in full:
        return full['/'.join(parts)]
    if len(parts) >= 2:
        return ac.get(f'{parts[0]}/{parts[-1]}', '')
    return ''


# ============================ 源值解析 ============================
def resolve_source_values(source_df, option_table=FW_TABLE):
    df = source_df.copy()
    option_maps = c.build_fw_select_option_maps(
        option_table,
        ['htlx', 'htejlx', 'htzt', 'bglx'],
    )
    employee_map = _timed('  ├─员工映射', lambda: c.build_fw_employee_info_map_for_ids(
        pd.concat([df['合同执行人员ID'], df['合同创建人ID'], df['申请人ID']], ignore_index=True)))
    company_info_map = _timed('  ├─我方主体映射', lambda: build_fw_company_info_map_for_values(df['合同用印范围ID']))
    customer_info_map = _timed('  ├─客户映射', lambda: build_customer_info_map_for_values(df['合同客户ID']))
    supplier_info_map = _timed('  ├─供应商映射(汉得匹配)', lambda: build_supplier_info_map_for_values(df['合同供应商ID']))
    project_map = _timed('  ├─项目映射', lambda: build_fw_project_code_map_for_ids(df['合同所属项目编号ID']))
    relation_info_map = _timed(
        '  ├─关联框架映射',
        lambda: build_contract_info_map_for_ids(df['关联框架协议ID'], option_table=option_table),
    )
    cost_center_platform_map = _timed('  └─成本中心平台映射',
                                      lambda: build_cost_center_platform_map(df.get('成本中心ID', pd.Series(dtype=object))))

    df['成本中心平台'] = df.get('成本中心ID', pd.Series('', index=df.index)).map(
        lambda value: _lookup_first_browser_value(cost_center_platform_map, value))
    df['合同类型'] = df['合同类型ID'].map(lambda value: option_maps.get('htlx', {}).get(c.format_code(value), ''))
    df['合同二级类型'] = df['合同二级类型ID'].map(lambda value: option_maps.get('htejlx', {}).get(c.format_code(value), ''))
    df['合同签署状态'] = df['合同签署状态ID'].map(lambda value: option_maps.get('htzt', {}).get(c.format_code(value), ''))
    df['变更类型'] = df['变更类型ID'].map(lambda value: option_maps.get('bglx', {}).get(c.format_code(value), ''))
    df['流程名称'] = df['流程名称'].map(c.clean_fw_select_name)
    df['合同审批状态'] = df.get('合同审批状态', pd.Series('', index=df.index)).map(c.clean_fw_select_name)
    df['合同执行人员'] = df['合同执行人员ID'].map(
        lambda value: employee_map.get(c.format_code(value), {}).get('name', ''))
    df['合同执行人员工号'] = df['合同执行人员ID'].map(
        lambda value: employee_map.get(c.format_code(value), {}).get('code', ''))
    df['合同创建人'] = df['合同创建人ID'].map(
        lambda value: employee_map.get(c.format_code(value), {}).get('name', ''))
    df['合同创建人工号'] = df['合同创建人ID'].map(
        lambda value: employee_map.get(c.format_code(value), {}).get('code', ''))
    df['合同创建人联系方式'] = df['合同创建人ID'].map(
        lambda value: ';'.join(c.clean_text_values(
            employee_map.get(c.format_code(value), {}).get(field, '')
            for field in ('workcode', 'loginid', 'mobile', 'telephone', 'email')
        )))
    df['申请人'] = df['申请人ID'].map(
        lambda value: employee_map.get(c.format_code(value), {}).get('name', ''))
    df['申请人工号'] = df['申请人ID'].map(
        lambda value: employee_map.get(c.format_code(value), {}).get('code', ''))
    df['申请人联系方式'] = df['申请人ID'].map(
        lambda value: ';'.join(c.clean_text_values(
            employee_map.get(c.format_code(value), {}).get(field, '')
            for field in ('workcode', 'loginid', 'mobile', 'telephone', 'email')
        )))
    status_by_number, status_by_name = _employment_status_map()
    applicant_status = c.build_applicant_status_map(df['申请人ID'], status_by_number, status_by_name)
    creator_status = c.build_applicant_status_map(df['合同创建人ID'], status_by_number, status_by_name)
    executor_status = c.build_applicant_status_map(df['合同执行人员ID'], status_by_number, status_by_name)
    df['申请人状态'] = df['申请人ID'].map(lambda value: applicant_status.get(c.format_code(value), ''))
    df['合同创建人状态'] = df['合同创建人ID'].map(lambda value: creator_status.get(c.format_code(value), ''))
    df['合同执行人状态'] = df['合同执行人员ID'].map(lambda value: executor_status.get(c.format_code(value), ''))
    feishu_id_map = _timed('  ├─合同执行人/创建人飞书ID映射(汉得)',
                           lambda: build_feishu_employee_id_map(
                               pd.concat([df['合同执行人员工号'], df['合同创建人工号'], df['申请人工号']],
                                         ignore_index=True)))
    feishu_id_by_name = _timed('  ├─申请人飞书ID姓名兜底(汉得)',
                               lambda: build_feishu_employee_id_map_by_name(df['申请人']))
    contact_values = []
    for value in pd.concat([df['合同创建人联系方式'], df['申请人联系方式']], ignore_index=True):
        contact_values.extend(c.clean_text_values(_text(value).split(';')))
    feishu_id_by_contact = _timed('  ├─申请人/创建人飞书ID联系方式兜底(汉得)',
                                  lambda: build_feishu_employee_id_map_by_contact(contact_values))
    df['合同执行人飞书ID'] = df['合同执行人员工号'].map(lambda code: feishu_id_map.get(_text(code), ''))
    df['合同创建人user_id'] = df.apply(
        lambda row: feishu_id_map.get(_text(row.get('合同创建人工号')), '')
        or next((
            feishu_id_by_contact.get(contact)
            for contact in c.clean_text_values(_text(row.get('合同创建人联系方式')).split(';'))
            if feishu_id_by_contact.get(contact)
        ), ''),
        axis=1,
    )
    df['申请人user_id'] = df.apply(
        lambda row: feishu_id_map.get(_text(row.get('申请人工号')), '')
        or next((
            feishu_id_by_contact.get(contact)
            for contact in c.clean_text_values(_text(row.get('申请人联系方式')).split(';'))
            if feishu_id_by_contact.get(contact)
        ), '')
        or feishu_id_by_name.get(_text(row.get('申请人')), ''),
        axis=1,
    )
    purchase_request_map = _timed('  ├─采购申请单编号映射',
                                  lambda: build_purchase_request_code_map(df.get('采购申请单ID', pd.Series(dtype=object))))
    df['采购申请单编号'] = df.get('采购申请单ID', pd.Series('', index=df.index)).map(
        lambda value: ';'.join(
            purchase_request_map[request_id]
            for request_id in c.parse_browser_ids(value)
            if purchase_request_map.get(request_id)
        ))
    df['泛微项目编号'] = df['合同所属项目编号ID'].map(
        lambda value: _lookup_first_browser_value(project_map, value) or _text(value))
    df['泛微项目名称'] = df['合同所属项目']
    _t0 = time.perf_counter()
    cleaned_project_info = df.apply(
        lambda row: c.cleaned_project_mapping(
            row.get('合同所属项目编号ID'),
            row.get('合同所属项目'),
            row.get('泛微项目编号'),
        ),
        axis=1,
    )
    df['项目编号'] = cleaned_project_info.map(
        lambda info: _text(info.get('项目编号')) if info else '')
    df['项目名称'] = cleaned_project_info.map(
        lambda info: _text(info.get('项目名称')) if info else '')
    df['项目映射来源'] = cleaned_project_info.map(
        lambda info: _text(info.get('映射来源')) if info else '')
    df['项目编号'] = df.apply(lambda row: _first_non_blank(row['项目编号'], row['泛微项目编号']), axis=1)
    df['项目名称'] = df.apply(lambda row: _first_non_blank(row['项目名称'], row['合同所属项目']), axis=1)
    df['清洗后项目编号'] = df['项目编号']
    df['清洗后项目名称'] = df['项目名称']
    order_init_items = df.apply(_order_init_items_for_source, axis=1)
    has_order_init_items = order_init_items.map(bool)
    df['订单编号'] = ''
    df['订单名称'] = ''
    df['订单映射来源'] = ''
    df['订单申请人员工编码'] = ''
    if bool(has_order_init_items.any()):
        df.loc[has_order_init_items, '订单编号'] = order_init_items[has_order_init_items].map(
            lambda items: _join_order_item_field(items, '订单编号'))
        df.loc[has_order_init_items, '订单名称'] = order_init_items[has_order_init_items].map(
            lambda items: _join_order_item_field(items, '订单标题'))
        df.loc[has_order_init_items, '订单申请人员工编码'] = order_init_items[has_order_init_items].map(
            lambda items: _join_order_item_field(items, '申请人员工编码'))
        df.loc[has_order_init_items, '订单映射来源'] = order_init_items[has_order_init_items].map(
            lambda items: _join_order_item_field(items, '映射来源'))
    saishi_mask = df['数据来源'].map(_text).eq('泛微(赛事)')
    saishi_has_order = saishi_mask & has_order_init_items
    if bool(saishi_mask.any()):
        print(f'[合同迁移-一般流程] 赛事订单初始化表命中 {int(saishi_has_order.sum())}/{int(saishi_mask.sum())} 行')
    df = _apply_contract_creator_rules(df, status_by_number, status_by_name)
    print(f'[计时] ✓   项目/订单映射(逐行apply): {time.perf_counter() - _t0:.1f}s', flush=True)

    def supplier_names(value):
        return '; '.join(
            supplier_info_map.get(supplier_id, {}).get('source_name', '')
            for supplier_id in c.parse_browser_ids(value)
            if supplier_info_map.get(supplier_id, {}).get('source_name', '')
        )

    _t0 = time.perf_counter()
    df['合同供应商名称'] = df['合同供应商ID'].map(supplier_names)
    df['合同分类'] = df.apply(resolve_contract_category, axis=1)
    df['合同分类依据'] = df.apply(resolve_contract_category_basis, axis=1)
    df['收支类型'] = df.apply(resolve_pay_type, axis=1)
    df['专项品类'] = df.get('专项分类编码', pd.Series('', index=df.index)).map(resolve_special_category)
    amount_tuples = df.apply(resolve_amounts, axis=1)
    df['合同总额_解析'] = amount_tuples.map(lambda item: item[0])
    df['收入总额_解析'] = amount_tuples.map(lambda item: item[1])
    df['支出总额_解析'] = amount_tuples.map(lambda item: item[2])
    signed_amount_tuples = df.apply(resolve_signed_amounts, axis=1)
    df['合同总额_签名'] = signed_amount_tuples.map(lambda item: item[0])
    df['收入总额_签名'] = signed_amount_tuples.map(lambda item: item[1])
    df['支出总额_签名'] = signed_amount_tuples.map(lambda item: item[2])
    print(f'[计时] ✓   分类/金额(逐行apply): {time.perf_counter() - _t0:.1f}s', flush=True)

    def relation_items(value):
        items = []
        for relation_id in c.parse_browser_ids(value):
            info = relation_info_map.get(relation_id, {})
            if info.get('number'):
                item = dict(info)
                item['id'] = relation_id
                items.append(item)
        return items

    df['关联合同信息'] = df['关联框架协议ID'].map(relation_items)
    df['关联合同编号'] = df['关联合同信息'].map(lambda items: ';'.join(item['number'] for item in items))
    df['关联合同名称'] = df['关联合同信息'].map(lambda items: ';'.join(item.get('title', '') for item in items))

    df.attrs['company_info_map'] = company_info_map
    df.attrs['customer_info_map'] = customer_info_map
    df.attrs['supplier_info_map'] = supplier_info_map
    df.attrs['relation_info_map'] = relation_info_map
    return df


def _merge_attrs(target_df, source_dfs):
    """把多个已解析 df 的 attrs(各 id->info 映射)合并到 target_df.attrs。"""
    for key in ('company_info_map', 'customer_info_map', 'supplier_info_map', 'relation_info_map'):
        merged = {}
        for df in source_dfs:
            merged.update(df.attrs.get(key, {}))
        target_df.attrs[key] = merged
    return target_df


@lru_cache(maxsize=1)
def load_anti_bribery_contract_number_keys():
    """读取反商业贿赂协议任务的来源模板合同编号,供一般流程排除。"""
    if not ANTI_BRIBERY_TEMPLATE_FILE.exists():
        print(f'[合同迁移-一般流程] 反贿赂模板不存在,跳过反贿赂合同排除: {ANTI_BRIBERY_TEMPLATE_FILE}')
        return frozenset()

    keys = set()
    source_counts = []
    for sheet_name in ANTI_BRIBERY_SOURCE_SHEETS:
        try:
            df = pd.read_excel(ANTI_BRIBERY_TEMPLATE_FILE, sheet_name=sheet_name, dtype=object)
        except Exception as error:
            print(f'[合同迁移-一般流程] 反贿赂模板sheet读取失败({sheet_name}),跳过该sheet: {error}')
            continue
        if '合同编号' not in df.columns:
            print(f'[合同迁移-一般流程] 反贿赂模板sheet缺少「合同编号」列,跳过: {sheet_name}')
            continue
        sheet_keys = {_contract_number_key(value) for value in df['合同编号']}
        sheet_keys.discard('')
        keys.update(sheet_keys)
        source_counts.append(f'{sheet_name} {len(sheet_keys)} 个')

    print(
        '[合同迁移-一般流程] 反贿赂合同编号排除池:',
        '; '.join(source_counts) if source_counts else '无有效来源',
        f'合计去重 {len(keys)} 个',
    )
    return frozenset(keys)


def _exclude_anti_bribery_contracts(*source_dfs):
    anti_keys = load_anti_bribery_contract_number_keys()
    filtered_dfs = []
    excluded_dfs = []
    for source_df in source_dfs:
        if source_df.empty or not anti_keys:
            filtered_dfs.append(source_df)
            continue
        exclude_mask = source_df['合同编号'].map(lambda value: _contract_number_key(value) in anti_keys)
        excluded = source_df.loc[exclude_mask].copy()
        filtered = source_df.loc[~exclude_mask].copy()
        filtered_dfs.append(filtered)
        if not excluded.empty:
            excluded_dfs.append(excluded)

    excluded_df = pd.concat(excluded_dfs, ignore_index=True) if excluded_dfs else pd.DataFrame()
    if not excluded_df.empty:
        print(
            '[合同迁移-一般流程] 反贿赂合同排除:',
            f'{len(excluded_df)} 条;',
            f'合同编号 {excluded_df["合同编号"].map(_contract_number_key).nunique()} 个',
        )
    return (*filtered_dfs, excluded_df)


def _forced_general_contract_query_numbers():
    numbers = []
    for contract_number in FORCED_GENERAL_CONTRACT_NUMBERS:
        numbers.extend(_forced_contract_number_variants(contract_number))
    return tuple(dict.fromkeys(numbers))


def read_forced_general_contract_source():
    query_numbers = _forced_general_contract_query_numbers()
    if not query_numbers:
        return pd.DataFrame()
    df = c.query_db(
        'FW',
        'vspn_xtyy',
        FORCED_GENERAL_CONTRACT_SOURCE_SQL,
        {
            'forced_contract_numbers': query_numbers,
            'anchor_contract_type_code': ANCHOR_CONTRACT_TYPE_CODE,
            'migration_status_codes': MIGRATION_STATUS_CODES,
        },
    )
    if df.empty:
        print('[合同迁移-一般流程] 强制追加合同: 未查到任何合同')
        return df

    df['_contract_key'] = df['合同编号'].map(_contract_number_key)
    df = df.drop_duplicates('_contract_key', keep='first').drop(columns=['_contract_key'])
    found_keys = set(df['合同编号'].map(_contract_number_key))
    missing = [
        contract_number
        for contract_number in FORCED_GENERAL_CONTRACT_NUMBERS
        if _contract_number_key(contract_number) not in found_keys
    ]
    if missing:
        print('[合同迁移-一般流程] 强制追加合同未查到:', '、'.join(missing))
    df['数据来源'] = '泛微(MCN)'
    df['强制追加导出'] = 'Y'
    print(f'[合同迁移-一般流程] 强制追加一般合同: {len(df)}/{len(FORCED_GENERAL_CONTRACT_NUMBERS)} 条')
    return df


def read_source():
    # ---- 源1: MCN 合同库 uf_htk ----
    c.validate_fw_fields(FW_TABLE, EXPECTED_FW_FIELDS)
    stats = _query_fw(STATS_SQL).iloc[0]
    print('[合同迁移-一般流程] MCN(uf_htk) 过滤: 合同类型<>主播协议 且 合同签署状态∈(审批完成, 已归档)')
    print(f"  合同库总数 {int(stats['all_count'] or 0)} 条; "
          f"主播协议排除 {int(stats['anchor_type_count'] or 0)} 条; "
          f"一般流程保留 {int(stats['kept_count'] or 0)} 条; "
          f"一般流程排除其他状态 {int(stats['excluded_status_count'] or 0)} 条")
    mcn_df = _timed('read_source/MCN取数(uf_htk)', lambda: _query_fw(SOURCE_SQL))
    mcn_df['数据来源'] = '泛微(MCN)'
    mcn_df['强制追加导出'] = ''
    print('[合同迁移-一般流程] MCN 主表行数:', len(mcn_df))
    forced_mcn_df = _timed('read_source/强制追加一般合同(uf_htk)', read_forced_general_contract_source)
    forced_keys = set(forced_mcn_df.get('合同编号', pd.Series(dtype=object)).map(_contract_number_key))
    if forced_keys:
        before_mcn = len(mcn_df)
        mcn_df = mcn_df[~mcn_df['合同编号'].map(lambda value: _contract_number_key(value) in forced_keys)].copy()
        print(f'[合同迁移-一般流程] 普通MCN结果剔除强制追加重复: {before_mcn - len(mcn_df)} 条')

    # ---- 源2: 赛事 合同审批台账 uf_htsp(htzt=0/1 审批中/归档); 与 MCN 重叠的编号让位给 MCN ----
    htsp_df = _timed('read_source/赛事取数(uf_htsp)', lambda: c.query_db(
        'FW', 'vspn_xtyy', SOURCE_SQL_HTSP, {'htsp_status_codes': HTSP_MIGRATION_STATUS_CODES}))
    htsp_df['数据来源'] = '泛微(赛事)'
    htsp_df['强制追加导出'] = ''
    mcn_df, htsp_df, anti_bribery_excluded_df = _exclude_anti_bribery_contracts(mcn_df, htsp_df)
    mcn_codes = set(mcn_df['合同编号'].map(_text))
    before = len(htsp_df)
    htsp_df = htsp_df[~htsp_df['合同编号'].map(_text).isin(mcn_codes)].copy()
    print(f'[合同迁移-一般流程] 赛事(uf_htsp) 反贿赂排除后行数 {before}; 去重(MCN优先)后 {len(htsp_df)}')

    resolved_mcn = _timed('read_source/解析MCN(%d行)' % len(mcn_df),
                          lambda: resolve_source_values(mcn_df, option_table=FW_TABLE))
    resolved_htsp = (
        _timed('read_source/解析赛事(%d行)' % len(htsp_df),
               lambda: resolve_source_values(htsp_df, option_table=FW_TABLE_HTSP))
        if not htsp_df.empty else htsp_df
    )
    resolved_forced_mcn = (
        _timed('read_source/解析强制追加一般合同(%d行)' % len(forced_mcn_df),
               lambda: resolve_source_values(forced_mcn_df, option_table=FW_TABLE))
        if not forced_mcn_df.empty else forced_mcn_df
    )

    merged = pd.concat([resolved_mcn, resolved_htsp, resolved_forced_mcn], ignore_index=True)
    merged, project_filter_excluded_df = _timed(
        'read_source/泛微项目编码白名单过滤',
        lambda: _filter_by_contract_project_whitelist(merged),
    )
    merged, approval_status_excluded_df = _timed(
        'read_source/合同审批状态过滤',
        lambda: _filter_by_approval_status(merged),
    )
    _merge_attrs(merged, [resolved_mcn, resolved_htsp, resolved_forced_mcn])
    merged.attrs['anti_bribery_excluded'] = anti_bribery_excluded_df
    merged.attrs['project_filter_excluded'] = project_filter_excluded_df
    merged.attrs['approval_status_excluded'] = approval_status_excluded_df
    merged = _timed('read_source/补充协议金额汇总', lambda: _apply_supplement_amount_rollup(merged))
    # 赛事收支计划明细(uf_htsp_dt4): 供付款/收款计划按 dt4 真实多期展开。
    saishi_ids = merged.loc[
        merged.get('数据来源', pd.Series('', index=merged.index)).map(_text).eq('泛微(赛事)'),
        'ID',
    ] if 'ID' in merged.columns else pd.Series(dtype=object)
    plan_map = (
        _timed('read_source/赛事收支计划dt4', lambda: load_htsp_plan_detail_map(saishi_ids))
        if not saishi_ids.empty else {}
    )
    merged.attrs['saishi_plan_map'] = plan_map
    print(f'[合同迁移-一般流程] 赛事收支计划明细: {sum(len(v) for v in plan_map.values())} 行 / {len(plan_map)} 合同')
    print('[合同迁移-一般流程] 合并后主表行数:', len(merged))
    return merged


# ============================ Sheet 构建 ============================
def build_main_output(source_df, headers):
    rows = []
    for source in source_df.to_dict('records'):
        row = _new_row(headers)
        amount = source['合同总额_解析']
        in_amount = source['收入总额_解析']
        out_amount = source['支出总额_解析']

        contract_name = _text(source['合同标题'])
        _set(row, 'contract_number（合同编码）', _text(source['合同编号']))
        _set(row, 'contract_name（合同名称）', contract_name)
        _set(row, '泛微项目编码', _text(source.get('泛微项目编号')))
        _set(row, '合同审批状态', _text(source.get('合同审批状态')))
        _set(row, '订单编号', _text(source.get('订单编号')))
        _set(row, 'contractCategory(智书框架合同类型)',
             '其他-保密协议' if '保密协议' in contract_name else source['合同分类'])
        _set(row, 'pay_type_code（收支类型）', source['收支类型'])
        _set(row, 'property_type_code（计价方式）', DEFAULT_PROPERTY_TYPE)
        _set(row, 'estimated_amount（预估金额）', amount)
        _set(row, 'in_amount（预估收入金额）', in_amount)
        _set(row, 'out_amount（预估支出金额）', out_amount)
        _set(row, 'amount（合同总额）', amount)
        _set(row, 'in_amount（收入总额）', in_amount)
        _set(row, 'out_amount（支出总额）', out_amount)
        _set(row, 'fixed_validity_code（合同期限类型）', DEFAULT_VALIDITY_TYPE)
        _set(row, 'start_date（合同期限-开始日期）', c.format_date(source['合同有效期起始时间']))
        _set(row, 'end_date（合同期限-结束日期）', c.format_date(source['合同有效期截止时间']))
        _set(row, 'remark（合同说明）', _text(source['合同摘要'])[:150])
        _set(row, 'custom_1001_948719050bfe402ab083c98e52fa71b2（合同执行人）飞书user_id',
             _text(source['合同执行人飞书ID']))
        _set(row, '合同执行人', _text(source['合同执行人员']))
        _set(row, '合同创建人', _text(source['合同创建人']))
        _set(row, '合同创建人user_id', _text(source['合同创建人user_id']))
        _set(row, '合同创建人状态', _text(source.get('合同创建人状态')))
        _set(row, 'custom_15_78cf503c57194e4fb8ad03ded1c4ad60（打印模式）', DEFAULT_PRINT_MODE)
        _set(row, 'custom_10_9a2a0e99771346c98bfb6cfb893e1bee（签署日期）', c.format_date(source['合同签订日期']))
        _set(row, 'custom_15_de8944334b104d52b28d9472ab0584ef（专项品类）', _text(source.get('专项品类')))
        _set(row, 'custom_13_c9805a6fe9f245ebbfeea13407277306（是否需要验收）', DEFAULT_ACCEPTANCE_REQUIRED)
        _set(row, 'custom_1012_cec7052f613b465980f23f7004e2f82c（采购金额）', out_amount if out_amount else '')
        _set(row, 'custom_15_e5f7b7cb17b34602adf790f0ac8d69a1（发票种类）', DEFAULT_INVOICE_TYPE)
        _set(row, 'custom_15_7b0d0e2f63a148729f929ba985c227c2（收入税率）',
             _format_tax_rate(source.get('收入税率')) or DEFAULT_TAX_RATE)
        _set(row, 'custom_15_b293866468ac4ab4bb11b5cb8c9bbb37（收入税目）', DEFAULT_TAX_ITEM)
        _set(row, 'custom_15_7b0d0e2f63a148729f929ba985c227c2（支出税率）',
             _format_tax_rate(source.get('支出税率')) or DEFAULT_TAX_RATE)
        _set(row, 'custom_15_b293866468ac4ab4bb11b5cb8c9bbb37（支出税目）', DEFAULT_TAX_ITEM)
        _set(row, 'custom_15_e46e1f9b6eb0469987f0656a999cdf09（银行手续费承担方）',
             DEFAULT_BANK_FEE_BEARER)
        deposit_amount = (_round_amount(source.get('押金')) or 0) + (_round_amount(source.get('保证金')) or 0)
        _set(row, 'custom_1012_7e2c970e63f648268eaefbd13d6bfc8f（押金/保证金）',
             deposit_amount if deposit_amount else '')
        _set(row, 'sign_type_code（先盖章方）', DEFAULT_FIRST_SEAL_PARTY)
        _set(row, 'sign_type_code（签约形式）', DEFAULT_SIGN_FORM)
        _set(row, 'seal_number（盖章份数）待确认必填项', DEFAULT_SEAL_NUMBER)
        _set(row, 'contract_files.contract_text（合同文本）', '')
        _set(row, 'custom_1012_7c79ac40b9ec4efb8be367a43f480d01（首款金额）', '')
        _set(row, 'custom_1012_98dfe42395274c53926c2c37b2dbd9a9（尾款金额）', '')
        rows.append(row)
    return pd.DataFrame(rows, columns=headers)


def build_relation_output(source_df, headers):
    relation_info_map = source_df.attrs.get('relation_info_map', {})
    rows = []
    source_rows = []
    for source in source_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        relation_items = source.get('关联合同信息') or []
        if not isinstance(relation_items, list):
            relation_items = []
        if not relation_items:
            relation_items = [
                {'id': relation_id, **relation_info_map.get(relation_id, {})}
                for relation_id in c.parse_browser_ids(source['关联框架协议ID'])
            ]
        relation_ids = []
        relation_numbers = []
        relation_titles = []
        relation_source_tables = []
        for info in relation_items:
            relation_id = _text(info.get('id'))
            relation_number = _text(info.get('number'))
            relation_title = _text(info.get('title'))
            relation_source_table = _text(info.get('source_table'))
            if relation_id:
                relation_ids.append(relation_id)
            if relation_number:
                relation_numbers.append(relation_number)
            if relation_title:
                relation_titles.append(relation_title)
            if relation_source_table:
                relation_source_tables.append(relation_source_table)

        relation_number_text = ';'.join(c.clean_text_values(relation_numbers))
        row = _new_row(headers)
        _set(row, 'contract_number（合同编码）', contract_number)
        _set(row, 'relation.relation_contracts（关联合同）', relation_number_text)
        _set(row, '框架合同编号', relation_number_text)
        rows.append(row)
        source_rows.append({
            'contract_number（合同编码）': contract_number,
            '泛微关联合同ID': ';'.join(c.clean_text_values(relation_ids)),
            '关联合同编号': relation_number_text,
            '关联合同名称': ';'.join(c.clean_text_values(relation_titles)),
            '关联合同来源表': ';'.join(c.clean_text_values(relation_source_tables)),
        })
    return pd.DataFrame(rows, columns=headers), pd.DataFrame(source_rows)


def _order_entries_for_source(source):
    order_codes = _split_multi_values(source.get('订单编号'))
    order_names = _split_joined_field(source.get('订单名称'))
    entries = []
    seen = set()
    for index, order_code in enumerate(order_codes):
        if order_code in seen:
            continue
        seen.add(order_code)
        entries.append({
            '订单编号': order_code,
            '订单名称': order_names[index] if index < len(order_names) else '',
        })
    if not entries and order_names:
        entries.append({'订单编号': '', '订单名称': order_names[0]})
    return entries


def build_related_order_output(source_df, headers):
    rows = []
    for source in source_df.to_dict('records'):
        for entry in _order_entries_for_source(source):
            order_code = _text(entry.get('订单编号'))
            if not order_code:
                continue
            row = _new_row(headers)
            _set(row, 'contract_number（合同编码）', _text(source['合同编号']))
            _set(row, 'custom_1024_90a78c8120994f95b2dbfedd297c7d81（相关单据-订单信息）', order_code)
            rows.append(row)
    return pd.DataFrame(rows, columns=headers)


def build_purchase_request_output(source_df, headers):
    rows = []
    for source in source_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        for code in _text(source.get('采购申请单编号')).split(';'):
            code = code.strip()
            if not code:
                continue
            row = _new_row(headers)
            _set(row, 'contract_number（合同编码）', contract_number)
            _set(row, 'custom_1024_7db9a8ee2b3d4a3f9d9835dd9fee69df（采购申请）', code)
            rows.append(row)
    return pd.DataFrame(rows, columns=headers)


def build_order_detail_output(source_df, headers):
    rows = []
    source_rows = []
    for source in source_df.to_dict('records'):
        for entry in _order_entries_for_source(source):
            order_code = _text(entry.get('订单编号'))
            order_name = _text(entry.get('订单名称'))
            if not (order_code or order_name):
                continue
            row = _new_row(headers)
            _set(row, 'contract_number（合同编码）', _text(source['合同编号']))
            _set(row, 'custom_1_5549b19faea641eeac924deada603c11（订单名称）', order_name)
            _set(row, 'custom_1_7f977c0d30064dd199434f706470c669（订单编号）', order_code)
            _set(row, 'custom_16_3171b080033943c9a98380f20e0895a8（成本中心）', '')
            start_date = c.format_date(source['合同有效期起始时间'])
            end_date = c.format_date(source['合同有效期截止时间'])
            period = f'{start_date}~{end_date}'.strip('~')
            _set(row, 'custom_12_d67e0d9472134b1cba5187e192bb2670（订单周期）',
                 _first_non_blank(period, c.format_date(source['合同签订日期'])))
            _set(row, 'custom_1001_b193503253664cf28b2ca1c3f57b68b3（项目经理）飞书user_id', '')
            _set(row, 'custom_1001_0a61d360dcad4265a68d2555d17e896e（日常费用组）飞书user_id', '')
            _set(row, 'custom_1001_f8f9114f511346f9adc7fabae012f17a（项目验收岗）飞书user_id', '')
            _set(row, 'custom_1001_8b8028ca466f4841bead801a9c7fedf2（项目预算岗）飞书user_id', '')
            _set(row, 'custom_1001_9072dcb2126e4051854da3927f742ab9（项目Sponsor-项目经理B）飞书user_id', '')
            _set(row, 'custom_15_622e96ab047c4f689d287a27066f7bcb（订单类型）', '')
            rows.append(row)
            source_row = dict(source)
            source_row['订单编号'] = order_code
            source_row['订单名称'] = order_name
            source_rows.append(source_row)
    return pd.DataFrame(rows, columns=headers), pd.DataFrame(source_rows)


def build_counterparty_output(source_df, headers):
    customer_info_map = source_df.attrs.get('customer_info_map', {})
    supplier_info_map = source_df.attrs.get('supplier_info_map', {})
    rows = []
    source_rows = []
    for source in source_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        for customer_id in c.parse_browser_ids(source['合同客户ID']):
            info = customer_info_map.get(customer_id, {})
            row = _new_row(headers)
            _set(row, 'contract_number（合同编码）', contract_number)
            _set(row, 'counter_party_code（对方主体编码）', info.get('code', ''))
            rows.append(row)
            source_rows.append({
                'contract_number（合同编码）': contract_number,
                '对方主体类型': '客户',
                '泛微对方ID': customer_id,
                '泛微对方名称': info.get('source_name', ''),
                '归并目标ID': info.get('target_id', ''),
                '归并/匹配方式': info.get('match_method', ''),
                '对方主体名称': info.get('name', ''),
                '对方主体编码': info.get('code', ''),
            })
        for supplier_id in c.parse_browser_ids(source['合同供应商ID']):
            info = supplier_info_map.get(supplier_id, {})
            row = _new_row(headers)
            _set(row, 'contract_number（合同编码）', contract_number)
            _set(row, 'counter_party_code（对方主体编码）', info.get('code', ''))
            rows.append(row)
            source_rows.append({
                'contract_number（合同编码）': contract_number,
                '对方主体类型': '供应商',
                '泛微对方ID': supplier_id,
                '泛微对方状态': info.get('status_code', ''),
                '泛微对方名称': info.get('source_name', ''),
                '归并目标ID': info.get('target_id', ''),
                '归并目标名称': info.get('target_name', ''),
                '归并/匹配方式': info.get('match_method', ''),
                '对方主体名称': info.get('name') or info.get('source_name', ''),
                '对方主体编码': info.get('code', ''),
            })
    return pd.DataFrame(rows, columns=headers), pd.DataFrame(source_rows)


def build_our_party_output(source_df, headers):
    company_info_map = source_df.attrs.get('company_info_map', {})
    rows = []
    source_rows = []
    for source in source_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        company_ids = c.parse_browser_ids(source['合同用印范围ID'])
        if not company_ids:
            company_ids = ['']
        for company_id in company_ids:
            info = company_info_map.get(company_id, {})
            row = _new_row(headers)
            _set(row, 'contract_number（合同编码）', contract_number)
            _set(row, 'our_party_code（我方主体编码）', info.get('code', ''))
            rows.append(row)
            source_rows.append({
                'contract_number（合同编码）': contract_number,
                '泛微我方主体ID': company_id,
                '我方主体名称': info.get('name', ''),
                '我方主体编码': info.get('code', ''),
            })
    return pd.DataFrame(rows, columns=headers), pd.DataFrame(source_rows)


def _first_counterparty_code(source, customer_info_map, supplier_info_map):
    for customer_id in c.parse_browser_ids(source.get('合同客户ID')):
        code = customer_info_map.get(customer_id, {}).get('code', '')
        if code:
            return code
    for supplier_id in c.parse_browser_ids(source.get('合同供应商ID')):
        code = supplier_info_map.get(supplier_id, {}).get('code', '')
        if code:
            return code
    return ''


PLAN_TYPE_COLLECTION = '0'  # uf_htsp_dt4.lx=0 收款(收入)
PLAN_TYPE_PAYMENT = '1'     # uf_htsp_dt4.lx=1 付款(支出)
_HTML_TAG_RE = re.compile(r'<[^>]+>')


def _clean_plan_desc(value):
    text = html.unescape(_text(value))
    text = re.sub(r'<br\s*/?>', ' ', text, flags=re.IGNORECASE)
    text = _HTML_TAG_RE.sub(' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def load_htsp_plan_detail_map(contract_ids):
    """uf_htsp.id -> 收支计划明细行(uf_htsp_dt4)。lx: 0=收款 1=付款。"""
    result = {}
    ids = c.clean_codes(_text(cid) for cid in contract_ids if _text(cid))
    if not ids:
        return result
    for batch in _chunked(sorted(set(ids))):
        df = c.query_db(
            'FW',
            'vspn_xtyy',
            'SELECT mainid, lx, rqjd, je, rmbje, bz FROM uf_htsp_dt4 '
            f'WHERE mainid IN ({c.in_placeholders(batch)}) ORDER BY mainid, id',
            batch,
        )
        for _, row in df.iterrows():
            mainid = c.format_code(row['mainid'])
            if not mainid:
                continue
            result.setdefault(mainid, []).append({
                'lx': c.format_code(row.get('lx')),
                'date': _text(row.get('rqjd')),
                'amount': _round_amount(row.get('je')),
                'desc': _clean_plan_desc(row.get('bz')),
            })
    return result


def _plan_row_id(source, kind, index):
    # 跨源唯一: 赛事/MCN 的 ID 可能数值相同, 用来源字母前缀区分。
    tag = 'S' if _text(source.get('数据来源')) == '泛微(赛事)' else 'M'
    return f'{tag}{_text(source.get("ID"))}-{kind}{index}'


def _htsp_plan_details(source, plan_map, plan_type):
    if _text(source.get('数据来源')) != '泛微(赛事)':
        return []
    return [
        detail
        for detail in plan_map.get(_text(source.get('ID')), [])
        if detail['lx'] == plan_type and detail['amount']
    ]


def build_payment_plan_output(source_df, headers):
    customer_info_map = source_df.attrs.get('customer_info_map', {})
    supplier_info_map = source_df.attrs.get('supplier_info_map', {})
    plan_map = source_df.attrs.get('saishi_plan_map', {})
    rows = []

    def add_row(source, amount, date, desc, line_id):
        row = _new_row(headers)
        _set(row, 'contract_number（合同编码）', _text(source['合同编号']))
        _set(row, 'payment_plan_list（付款计划）', '付款计划')
        _set(row, 'payment_plan_list[].payment_date（付款时间）', c.format_date(date))
        _set(row, 'payment_plan_list[].prepaid（是否预付）', DEFAULT_PREPAID)
        _set(row, 'payment_plan_list[].payment_amount（付款金额）', amount)
        _set(row, 'payment_plan_list[].payment_desc（付款说明）', desc[:80])
        _set(row, 'payment_plan_list[].payment_custom_attributes/custom_付款性质（付款性质）',
             DEFAULT_PAYMENT_NATURE)
        _set(row, 'payment_plan_list[].payment_counter_party[].counter_party_code（付款对象）',
             _first_counterparty_code(source, customer_info_map, supplier_info_map))
        _set(row, '付款计划行id(付款记录传的id)', line_id)
        rows.append(row)

    for source in source_df.to_dict('records'):
        if _text(source.get('是否补充协议')) == 'Y' and _text(source.get('主合同编号')):
            continue
        # 押金/保证金: 合同若有押金或保证金, 额外补一行付款计划, 金额为两者之和。
        deposit_amount = (_round_amount(source.get('押金')) or 0) + (_round_amount(source.get('保证金')) or 0)
        if deposit_amount:
            add_row(
                source, deposit_amount,
                _first_non_blank(source['合同有效期截止时间'], source['合同签订日期']),
                '押金/保证金',
                _plan_row_id(source, 'P', 'D'),
            )
        details = _htsp_plan_details(source, plan_map, PLAN_TYPE_PAYMENT)
        if details:
            # 赛事: 用 dt4 真实收支计划明细, 每行一条付款计划。
            for index, detail in enumerate(details, 1):
                add_row(
                    source, detail['amount'],
                    _first_non_blank(detail['date'], source['合同有效期截止时间'], source['合同签订日期']),
                    _first_non_blank(detail['desc'], _text(source['合同标题'])),
                    _plan_row_id(source, 'P', index),
                )
            continue
        if _number(source.get('补充协议数量')):
            amount = _round_amount(source.get('付款计划汇总金额'))
            if amount:
                add_row(
                    source, amount,
                    _first_non_blank(source['合同有效期截止时间'], source['合同签订日期']),
                    _first_non_blank(f"主合同及补充协议汇总:{_text(source.get('补充协议编号'))}", _text(source['合同标题'])),
                    _plan_row_id(source, 'P', 'SUM'),
                )
            continue
        # 兜底(无明细的赛事 / MCN): 按支出总额合成单行。
        amount = _round_amount(source['支出总额_解析'])
        if not amount:
            continue
        add_row(
            source, amount,
            _first_non_blank(source['合同有效期截止时间'], source['合同签订日期']),
            _text(source['合同标题']),
            _plan_row_id(source, 'P', 1),
        )
    return pd.DataFrame(rows, columns=headers)


def build_collection_plan_output(source_df, headers):
    customer_info_map = source_df.attrs.get('customer_info_map', {})
    supplier_info_map = source_df.attrs.get('supplier_info_map', {})
    plan_map = source_df.attrs.get('saishi_plan_map', {})
    rows = []

    def add_row(source, amount, date, desc, line_id):
        row = _new_row(headers)
        _set(row, 'contract_number（合同编码）', _text(source['合同编号']))
        _set(row, 'collection_plan_list（收款计划）', '收款计划')
        _set(row, 'collection_plan_list[].collection_date（收款时间）', c.format_date(date))
        _set(row, 'collection_plan_list[].collection_amount（收款金额）', amount)
        _set(row, 'collection_plan_list[].collection_desc（收款说明）', desc[:80])
        _set(row, 'collection_plan_list[].collection_counter_party[].counter_party_code（收款对象）',
             _first_counterparty_code(source, customer_info_map, supplier_info_map))
        _set(row, '收款计划行id(收款记录传的id)', line_id)
        rows.append(row)

    for source in source_df.to_dict('records'):
        if _text(source.get('是否补充协议')) == 'Y' and _text(source.get('主合同编号')):
            continue
        details = _htsp_plan_details(source, plan_map, PLAN_TYPE_COLLECTION)
        if details:
            for index, detail in enumerate(details, 1):
                add_row(
                    source, detail['amount'],
                    _first_non_blank(detail['date'], source['合同有效期截止时间'], source['合同签订日期']),
                    _first_non_blank(detail['desc'], _text(source['合同标题'])),
                    _plan_row_id(source, 'C', index),
                )
            continue
        if _number(source.get('补充协议数量')):
            amount = _round_amount(source.get('收款计划汇总金额'))
            if amount:
                add_row(
                    source, amount,
                    _first_non_blank(source['合同有效期截止时间'], source['合同签订日期']),
                    _first_non_blank(f"主合同及补充协议汇总:{_text(source.get('补充协议编号'))}", _text(source['合同标题'])),
                    _plan_row_id(source, 'C', 'SUM'),
                )
            continue
        amount = _round_amount(source['收入总额_解析'])
        if not amount:
            continue
        add_row(
            source, amount,
            _first_non_blank(source['合同有效期截止时间'], source['合同签订日期']),
            _text(source['合同标题']),
            _plan_row_id(source, 'C', 1),
        )
    return pd.DataFrame(rows, columns=headers)


def _build_attachment_sheet_output(manifest_df, headers, target_sheet, field_name):
    output_rows = []
    seen_names = set()
    for meta in manifest_df.to_dict('records'):
        if _text(meta.get('attachment_sheet')) != target_sheet:
            continue
        contract_number = _text(meta.get('contract_number（合同编码）'))
        attachment_name = _text(meta.get('attachment_name'))
        dedupe_key = (target_sheet, contract_number, attachment_name)
        if dedupe_key in seen_names:
            continue
        seen_names.add(dedupe_key)
        row = _new_row(headers)
        _set(row, 'contract_number（合同编码）', contract_number)
        _set(row, field_name, attachment_name)
        output_rows.append(row)
    return pd.DataFrame(output_rows, columns=headers)


def build_contract_attachment_output(source_df, contract_headers, other_headers):
    manifest_df, missing_df = build_contract_attachment_manifest(source_df)
    # 导入侧是「按合同的扁平附件名单」: 同一稿件可能重复出现,按
    # (目标sheet, 合同编码, 附件名) 去重。下载侧仍按稿件类型分文件夹保留。
    contract_output_df = _build_attachment_sheet_output(
        manifest_df,
        contract_headers,
        SHEET_CONTRACT_ATTACHMENT,
        'contract_files.contract_causes（合同附件）',
    )
    other_output_df = _build_attachment_sheet_output(
        manifest_df,
        other_headers,
        SHEET_OTHER_ATTACHMENT,
        'contract_files.contract_attachments（其他附件）',
    )
    if len(manifest_df) > 0:
        manifest_df = manifest_df.copy()
        manifest_df['status'] = 'listed_only'
        manifest_df['error'] = 'Excel任务仅生成附件名称清单;下载请单独运行 contract_general_attachments_db'
        print(f'[合同迁移-一般流程] 合同附件名称清单行数: {len(manifest_df)}; 未执行下载')

    return (
        contract_output_df,
        other_output_df,
        manifest_df,
        missing_df,
    )


def build_attachment_output(headers):
    return pd.DataFrame(columns=headers)


def build_status_breakdown():
    option_maps = c.build_fw_select_option_maps(FW_TABLE, ['htzt'])
    df = _query_fw(STATUS_BREAKDOWN_SQL)
    df['合同签署状态'] = df['合同签署状态ID'].map(lambda value: option_maps['htzt'].get(c.format_code(value), '空'))
    return df[['合同签署状态ID', '合同签署状态', '合同数']]


# ============================ 输出 / 异常 ============================
MAIN_ISSUE_SOURCE_FIELDS = {
    'contract_number（合同编码）': '合同编号',
    'contract_name（合同名称）': '合同标题',
    'contractCategory(智书框架合同类型)': '合同分类依据',
    'custom_1001_948719050bfe402ab083c98e52fa71b2（合同执行人）飞书user_id': '合同执行人员工号',
    'start_date（合同期限-开始日期）': '合同有效期起始时间',
    'end_date（合同期限-结束日期）': '合同有效期截止时间',
    'remark（合同说明）': '合同摘要',
}

ORDER_ISSUE_SOURCE_FIELDS = {
    'custom_1_5549b19faea641eeac924deada603c11（订单名称）': '项目编号',
    'custom_1_7f977c0d30064dd199434f706470c669（订单编号）': '项目编号',
}


def collect_missing_counterparty(counterparty_source_df):
    if counterparty_source_df.empty:
        return pd.DataFrame(columns=[
            'contract_number（合同编码）', '对方主体类型', '泛微对方ID', '泛微对方状态',
            '泛微对方名称', '归并目标ID', '归并目标名称', '归并/匹配方式', '对方主体名称',
        ])
    missing = counterparty_source_df[counterparty_source_df['对方主体编码'].astype(str).str.strip() == '']
    columns = [
        'contract_number（合同编码）', '对方主体类型', '泛微对方ID', '泛微对方状态',
        '泛微对方名称', '归并目标ID', '归并目标名称', '归并/匹配方式', '对方主体名称',
    ]
    return missing[[column for column in columns if column in missing.columns]].drop_duplicates()


def collect_missing_our_party(our_party_source_df):
    if our_party_source_df.empty:
        return pd.DataFrame(columns=['contract_number（合同编码）', '泛微我方主体ID', '我方主体名称'])
    missing = our_party_source_df[our_party_source_df['我方主体编码'].astype(str).str.strip() == '']
    return missing[['contract_number（合同编码）', '泛微我方主体ID', '我方主体名称']].drop_duplicates()


def collect_missing_relation(relation_source_df):
    if relation_source_df.empty:
        return pd.DataFrame(columns=['contract_number（合同编码）', '泛微关联合同ID', '关联合同编号', '关联合同名称'])
    missing = relation_source_df[relation_source_df['关联合同编号'].astype(str).str.strip() == '']
    return missing.drop_duplicates()


def _audit_df(source_df, columns):
    audit = source_df.reindex(columns=columns)
    for column in audit.columns:
        audit[column] = audit[column].map(_text)
    return audit


def build_category_audit_df(source_df):
    return _audit_df(source_df, [
        '数据来源', '合同编号', '合同标题', '合同类型ID', '合同类型', '合同二级类型ID', '合同二级类型',
        '流程类型ID', '流程名称', '项目编号', '泛微项目编号', '合同所属项目', '合同供应商名称',
        '成本中心平台', '收支类型', '合同分类', '合同分类依据', '合同金额', '合同预计收入', '合同预计支出',
    ])


def build_order_audit_df(source_df):
    return _audit_df(source_df, [
        '数据来源', '合同编号', '合同标题', '合同所属项目编号ID', '泛微项目编号', '泛微项目名称',
        '清洗后项目编号', '清洗后项目名称', '项目映射来源', '订单编号', '订单名称', '订单映射来源',
        '订单申请人员工编码', '申请人', '申请人工号', '申请人状态',
        '合同执行人员', '合同执行人员工号', '合同执行人状态',
        '合同创建人', '合同创建人工号', '合同创建人状态', '合同创建人调整方式',
    ])


def build_project_filter_excluded_df(source_df):
    excluded = source_df.attrs.get('project_filter_excluded')
    columns = [
        '数据来源', '合同编号', '合同标题', '合同所属项目编号ID', '泛微项目编号', '泛微项目名称',
        '项目白名单Sheet', '项目白名单校验',
    ]
    if excluded is None or excluded.empty:
        return pd.DataFrame(columns=columns)
    return _audit_df(excluded, columns)


def build_approval_status_excluded_df(source_df):
    excluded = source_df.attrs.get('approval_status_excluded')
    columns = [
        '数据来源', '合同编号', '合同标题', '合同签署状态ID', '合同流程ID',
        '流程类型ID', '流程名称', '合同审批状态', '审批状态过滤',
    ]
    if excluded is None or excluded.empty:
        return pd.DataFrame(columns=columns)
    return _audit_df(excluded, columns)


def build_anti_bribery_excluded_df(source_df):
    excluded = source_df.attrs.get('anti_bribery_excluded')
    columns = [
        '数据来源', '合同编号', '合同标题', '合同签署状态ID', '合同流程ID',
        '流程类型ID', '流程名称', '合同审批状态', '排除原因',
    ]
    if excluded is None or excluded.empty:
        return pd.DataFrame(columns=columns)
    result = _audit_df(excluded, [
        '数据来源', '合同编号', '合同标题', '合同签署状态ID', '合同流程ID',
        '流程类型ID', '流程名称', '合同审批状态',
    ])
    result['排除原因'] = '合同编号已由 contract_anti_bribery_db 处理'
    return result.reindex(columns=columns)


def build_supplement_amount_backtest_df(source_df):
    backtest_codes = ['H-DF2026060027-S01', 'H-DF2026050137-S01']
    rows = []
    by_contract = {
        _text(row.get('合同编号')): row
        for row in source_df.to_dict('records')
    }
    for child_code in backtest_codes:
        child = by_contract.get(child_code, {})
        main_code = _text(child.get('主合同编号')) if child else ''
        main = by_contract.get(main_code, {})
        rows.append({
            '回测补充协议编号': child_code,
            '补充协议是否在本次范围': 'Y' if child else 'N',
            '主合同编号': main_code,
            '主合同是否在本次范围': 'Y' if main else 'N',
            '主合同原支出金额(签名)': _number(main.get('支出总额_签名')) if main else '',
            '补充协议支出金额(签名)': _number(child.get('支出总额_签名')) if child else '',
            '主合同付款计划汇总金额': _number(main.get('付款计划汇总金额')) if main else '',
            '主合同补充协议数量': _number(main.get('补充协议数量')) if main else '',
            '主合同补充协议编号': _text(main.get('补充协议编号')) if main else '',
        })
    return pd.DataFrame(rows)


def run():
    headers_by_sheet = _template_headers()
    required_by_sheet, remarks_by_sheet = _read_general_required_rules(headers_by_sheet)

    source_df = _timed('阶段1: read_source(读数+解析+合并)', read_source)

    print('[计时] === 阶段2: 构建各 sheet ===', flush=True)
    main_output_df = _timed('build 字段模板', lambda: build_main_output(source_df, headers_by_sheet[SHEET_MAIN]))
    relation_output_df, relation_source_df = _timed(
        'build 关联合同', lambda: build_relation_output(source_df, headers_by_sheet[SHEET_RELATION]))
    related_order_output_df = _timed(
        'build 相关单据-订单', lambda: build_related_order_output(source_df, headers_by_sheet[SHEET_RELATED_ORDER]))
    purchase_request_output_df = _timed(
        'build 采购申请', lambda: build_purchase_request_output(source_df, headers_by_sheet[SHEET_PURCHASE_REQUEST]))
    order_detail_output_df, order_detail_source_df = _timed(
        'build 订单明细', lambda: build_order_detail_output(source_df, headers_by_sheet[SHEET_ORDER_DETAIL]))
    counterparty_output_df, counterparty_source_df = _timed(
        'build 对方信息', lambda: build_counterparty_output(source_df, headers_by_sheet[SHEET_COUNTERPARTY]))
    our_party_output_df, our_party_source_df = _timed(
        'build 我方主体', lambda: build_our_party_output(source_df, headers_by_sheet[SHEET_OUR_PARTY]))
    payment_plan_output_df = _timed(
        'build 付款计划', lambda: build_payment_plan_output(source_df, headers_by_sheet[SHEET_PAYMENT_PLAN]))
    collection_plan_output_df = _timed(
        'build 收款计划', lambda: build_collection_plan_output(source_df, headers_by_sheet[SHEET_COLLECTION_PLAN]))
    (
        contract_attachment_output_df,
        other_attachment_output_df,
        contract_attachment_meta_df,
        contract_attachment_missing_df,
    ) = _timed(
        'build 合同附件/其他附件(含清单)',
        lambda: build_contract_attachment_output(
            source_df,
            headers_by_sheet[SHEET_CONTRACT_ATTACHMENT],
            headers_by_sheet[SHEET_OTHER_ATTACHMENT],
        ),
    )

    print('[合同迁移-一般流程] 字段模板行数:', len(main_output_df))
    print('[合同迁移-一般流程] 关联合同行数:', len(relation_output_df))
    print('[合同迁移-一般流程] 相关单据-订单信息行数:', len(related_order_output_df))
    print('[合同迁移-一般流程] 订单信息明细行数:', len(order_detail_output_df))
    print('[合同迁移-一般流程] 对方信息行数:', len(counterparty_output_df))
    print('[合同迁移-一般流程] 我方主体列表行数:', len(our_party_output_df))
    print('[合同迁移-一般流程] 付款计划行数:', len(payment_plan_output_df))
    print('[合同迁移-一般流程] 收款计划行数:', len(collection_plan_output_df))
    print('[合同迁移-一般流程] 合同附件行数:', len(contract_attachment_output_df))
    print('[合同迁移-一般流程] 其他附件行数:', len(other_attachment_output_df))

    print('[计时] === 阶段3: 写出 Excel ===', flush=True)
    cat_audit = _timed('build 合同分类核对', lambda: build_category_audit_df(source_df))
    ord_audit = _timed('build 订单映射核对', lambda: build_order_audit_df(source_df))
    output_file = _timed('写出导入Excel(单次load/save)', lambda: _write_template_sheets_with_fallback(
        TEMPLATE_FILE, OUTPUT_FILE, {
            SHEET_MAIN: main_output_df,
            SHEET_RELATION: relation_output_df,
            SHEET_RELATED_ORDER: related_order_output_df,
            SHEET_PURCHASE_REQUEST: purchase_request_output_df,
            SHEET_ORDER_DETAIL: order_detail_output_df,
            SHEET_COUNTERPARTY: counterparty_output_df,
            SHEET_OUR_PARTY: our_party_output_df,
            SHEET_PAYMENT_PLAN: payment_plan_output_df,
            SHEET_COLLECTION_PLAN: collection_plan_output_df,
            SHEET_CONTRACT_ATTACHMENT: contract_attachment_output_df,
            SHEET_OTHER_ATTACHMENT: other_attachment_output_df,
        }, extra_sheets={'合同分类核对': cat_audit, '订单映射核对': ord_audit}))
    print('已写出:', output_file)

    exception_sheets = {
        '过滤状态分布': build_status_breakdown(),
        '默认值说明': pd.DataFrame([
            {'字段': 'property_type_code（计价方式）', '默认值': DEFAULT_PROPERTY_TYPE},
            {'字段': 'fixed_validity_code（合同期限类型）', '默认值': DEFAULT_VALIDITY_TYPE},
            {'字段': 'custom_13_c9805a6fe9f245ebbfeea13407277306（是否需要验收）', '默认值': DEFAULT_ACCEPTANCE_REQUIRED},
            {'字段': 'custom_15_78cf503c57194e4fb8ad03ded1c4ad60（打印模式）', '默认值': DEFAULT_PRINT_MODE},
            {'字段': 'custom_15_e5f7b7cb17b34602adf790f0ac8d69a1（发票种类）', '默认值': DEFAULT_INVOICE_TYPE},
            {'字段': 'custom_15_7b0d0e2f63a148729f929ba985c227c2（收入/支出税率）',
             '默认值': f'赛事取源值srsl/zcsl(百分数), 无源值时默认 {DEFAULT_TAX_RATE}'},
            {'字段': 'custom_15_b293866468ac4ab4bb11b5cb8c9bbb37（收入/支出税目）', '默认值': DEFAULT_TAX_ITEM},
            {'字段': 'custom_15_e46e1f9b6eb0469987f0656a999cdf09（银行手续费承担方）', '默认值': DEFAULT_BANK_FEE_BEARER},
            {'字段': 'sign_type_code（先盖章方）', '默认值': DEFAULT_FIRST_SEAL_PARTY},
            {'字段': 'sign_type_code（签约形式）', '默认值': DEFAULT_SIGN_FORM},
            {'字段': 'seal_number（盖章份数）待确认必填项', '默认值': DEFAULT_SEAL_NUMBER},
        ]),
        '字段模板_必输字段未达100%': _fill_summary(
            main_output_df,
            required_by_sheet[SHEET_MAIN],
            remarks_by_sheet[SHEET_MAIN],
        ),
        '关联合同_必输字段未达100%': _fill_summary(
            relation_output_df,
            required_by_sheet[SHEET_RELATION],
            remarks_by_sheet[SHEET_RELATION],
        ),
        '相关单据订单_必输字段未达100%': _fill_summary(
            related_order_output_df,
            required_by_sheet[SHEET_RELATED_ORDER],
            remarks_by_sheet[SHEET_RELATED_ORDER],
        ),
        '订单信息明细_必输字段未达100%': _fill_summary(
            order_detail_output_df,
            required_by_sheet[SHEET_ORDER_DETAIL],
            remarks_by_sheet[SHEET_ORDER_DETAIL],
        ),
        '对方信息_必输字段未达100%': _fill_summary(
            counterparty_output_df,
            required_by_sheet[SHEET_COUNTERPARTY],
            remarks_by_sheet[SHEET_COUNTERPARTY],
        ),
        '我方主体列表_必输字段未达100%': _fill_summary(
            our_party_output_df,
            required_by_sheet[SHEET_OUR_PARTY],
            remarks_by_sheet[SHEET_OUR_PARTY],
        ),
        '付款计划_必输字段未达100%': _fill_summary(
            payment_plan_output_df,
            required_by_sheet[SHEET_PAYMENT_PLAN],
            remarks_by_sheet[SHEET_PAYMENT_PLAN],
        ),
        '收款计划_必输字段未达100%': _fill_summary(
            collection_plan_output_df,
            required_by_sheet[SHEET_COLLECTION_PLAN],
            remarks_by_sheet[SHEET_COLLECTION_PLAN],
        ),
        '对方主体编码_未匹配': collect_missing_counterparty(counterparty_source_df),
        '我方主体编码_未匹配': collect_missing_our_party(our_party_source_df),
        '关联合同编号_未匹配': collect_missing_relation(relation_source_df),
        '合同附件下载清单': contract_attachment_meta_df,
        '合同附件DOCID_缺失映射': contract_attachment_missing_df,
        '泛微项目编码_非白名单': build_project_filter_excluded_df(source_df),
        '合同审批状态_已过滤': build_approval_status_excluded_df(source_df),
        '反贿赂合同排除': build_anti_bribery_excluded_df(source_df),
        '补充协议金额回测': build_supplement_amount_backtest_df(source_df),
    }
    exception_sheets.update({
        f'字段模板_{name}': df
        for name, df in _collect_missing_details(
            main_output_df,
            source_df,
            required_by_sheet[SHEET_MAIN],
            MAIN_ISSUE_SOURCE_FIELDS,
            'contract_number（合同编码）',
        ).items()
    })
    exception_sheets.update({
        f'订单信息明细_{name}': df
        for name, df in _collect_missing_details(
            order_detail_output_df,
            order_detail_source_df,
            required_by_sheet[SHEET_ORDER_DETAIL],
            ORDER_ISSUE_SOURCE_FIELDS,
            'contract_number（合同编码）',
        ).items()
    })
    legacy_order_issue_df = source_df[
        source_df.get('订单映射来源', pd.Series('', index=source_df.index)).map(_text) != ORDER_INIT_MAPPING_SOURCE
    ].copy()
    exception_sheets.update(c.collect_order_mapping_issues(
        legacy_order_issue_df,
        doc_col='合同编号',
        project_col='项目编号',
        project_id_col='合同所属项目编号ID',
        project_name_col='项目名称',
    ))

    exception_file = _write_exceptions_with_fallback(EXCEPTION_FILE, exception_sheets)
    if exception_file:
        print('已写出:', exception_file)


if __name__ == '__main__':
    run()
