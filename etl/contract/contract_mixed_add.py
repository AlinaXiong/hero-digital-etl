# -*- coding: utf-8 -*-
"""智书合同导入清单任务。

同一份业务清单里可能同时包含一般流程合同和主播流程合同。本任务只负责:

1. 仅读取本次上传的业务清单，并按合同编号去重;
2. 按 Excel 智书合同类型或泛微合同类型分流:主播类走 contract_anchor_db,其余走 contract_general_db;
3. 只读取输入清单前三列:合同编号用于定位源数据,关联业财订单/智书合同类型有值时优先采用,
   为空时沿用原清洗逻辑;第 4 列及之后全部忽略;
4. 复用两个主任务的解析与导出 builder,生成一个混合导入 Excel:一般流程占前 9 个 sheet,
   主播流程占后 4 个 sheet；并保留模板中紧随主播流程的两个空付款/收款计划 sheet。

运行方式::

    python run.py contract_mixed_add
    python run.py contract_mixed_add_all

可用环境变量覆盖输入文件::

    CONTRACT_MIXED_ADD_FILE

默认输入文件放在::

    resources/source/contract_mixed_add/技术导入数据清单.xlsx
"""
from __future__ import annotations

import os
import re
import json
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from etl.contract import contract_anchor_db as anchor
from etl.contract import contract_general_db as general
from etl.util import common as c


TASK_NAME = 'contract_mixed_add'
OUTPUT_DIR = c.OUT_DIR / TASK_NAME
SOURCE_DIR = c.SRC_DIR / TASK_NAME
DEFAULT_INPUT_FILE_NAME = '技术导入数据清单.xlsx'


def _default_input_file():
    if SOURCE_DIR.exists():
        xlsx_files = sorted(SOURCE_DIR.glob('*.xlsx'))
        if len(xlsx_files) == 1:
            return xlsx_files[0]
        if len(xlsx_files) > 1:
            raise RuntimeError(
                '智书合同导入清单输入目录中存在多个 Excel，请只保留一个导入清单，'
                '或设置 CONTRACT_MIXED_ADD_FILE 指向具体 .xlsx 文件: '
                + ', '.join(path.name for path in xlsx_files)
            )
    return SOURCE_DIR / DEFAULT_INPUT_FILE_NAME


def _input_file_from_directory(directory):
    xlsx_files = sorted(directory.glob('*.xlsx'))
    if len(xlsx_files) == 1:
        return xlsx_files[0]
    if not xlsx_files:
        return directory / DEFAULT_INPUT_FILE_NAME
    raise RuntimeError(
        '智书合同导入清单输入目录中存在多个 Excel，请只保留一个导入清单，'
        '或设置 CONTRACT_MIXED_ADD_FILE 指向具体 .xlsx 文件: '
        + ', '.join(path.name for path in xlsx_files)
    )


def _resolve_input_file(value=None):
    raw = '' if value is None else str(value).strip().strip('"')
    if not raw:
        return _default_input_file()
    path = Path(raw)
    if path.is_dir():
        return _input_file_from_directory(path)
    return path


INPUT_FILE = _resolve_input_file(os.getenv('CONTRACT_MIXED_ADD_FILE'))

DATE_SUFFIX = general.DATE_SUFFIX
MIXED_TEMPLATE_FILE = general.TEMPLATE_DIR / '智书合同字段-混合流程.xlsx'
INPUT_TEMPLATE_FILE = general.TEMPLATE_DIR / '混合合同增补业务清单模板.xlsx'
GENERAL_OUTPUT_FILE = OUTPUT_DIR / f'智书合同字段_一般流程_混合增补_{DATE_SUFFIX}.xlsx'
ANCHOR_OUTPUT_FILE = OUTPUT_DIR / f'智书合同字段_主播流程_混合增补_{DATE_SUFFIX}.xlsx'
MIXED_OUTPUT_FILE = OUTPUT_DIR / f'智书合同字段_混合增补_{DATE_SUFFIX}.xlsx'
MIXED_ATTACHMENT_ROOT = OUTPUT_DIR / f'混合增补合同附件_{DATE_SUFFIX}'
ARCHIVED_REQUEST_FILE = OUTPUT_DIR / f'智书合同同步请求_归档_9_{DATE_SUFFIX}.json'
OTHER_REQUEST_FILE = OUTPUT_DIR / f'智书合同同步请求_其他_0_{DATE_SUFFIX}.json'
APPROVE_TO_NODE_REQUEST_FILE = OUTPUT_DIR / f'智书合同自动审批请求_非归档_{DATE_SUFFIX}.json'
YECAI_SYNC_REQUEST_FILE = OUTPUT_DIR / f'智书合同同步业财请求_全量_{DATE_SUFFIX}.json'
AUDIT_FILE = OUTPUT_DIR / f'混合增补处理清单_{DATE_SUFFIX}.xlsx'

# ZhiShuSynServiceImpl.SheetRole uses fixed workbook indexes instead of sheet names.
# Keep these slots synchronized with the Java enum: general 0-8, anchor 9-12.
JAVA_GENERAL_SHEET_NAMES = (
    '一般流程主表',
    '关联合同',
    '相关订单-订单信息',
    '采购申请',
    '订单信息明细',
    '对方信息',
    '我方主体列表',
    '付款计划',
    '收款计划',
)
JAVA_ANCHOR_SHEET_NAMES = (
    '主播流程主表',
    '主播流程_对方信息',
    '主播流程_我方信息',
    '主播流程_费用明细',
)
# 混合模板还预置了两个主播付款/收款计划 sheet。当前 ETL 不填充数据，
# 但必须保留在费用明细之后，确保 Java 按固定索引读取时仍对应 13、14 位。
JAVA_ANCHOR_EMPTY_PLAN_SHEET_NAMES = (
    '主播流程_付款计划',
    '主播流程_收款计划',
)
JAVA_MIXED_SHEET_NAMES = (
    JAVA_GENERAL_SHEET_NAMES
    + JAVA_ANCHOR_SHEET_NAMES
    + JAVA_ANCHOR_EMPTY_PLAN_SHEET_NAMES
)
GENERAL_SOURCE_TO_JAVA_SHEET_NAMES = {
    general.SHEET_MAIN: JAVA_GENERAL_SHEET_NAMES[0],
    general.SHEET_RELATION: JAVA_GENERAL_SHEET_NAMES[1],
    general.SHEET_RELATED_ORDER: JAVA_GENERAL_SHEET_NAMES[2],
    general.SHEET_PURCHASE_REQUEST: JAVA_GENERAL_SHEET_NAMES[3],
    general.SHEET_ORDER_DETAIL: JAVA_GENERAL_SHEET_NAMES[4],
    general.SHEET_COUNTERPARTY: JAVA_GENERAL_SHEET_NAMES[5],
    general.SHEET_OUR_PARTY: JAVA_GENERAL_SHEET_NAMES[6],
    general.SHEET_PAYMENT_PLAN: JAVA_GENERAL_SHEET_NAMES[7],
    general.SHEET_COLLECTION_PLAN: JAVA_GENERAL_SHEET_NAMES[8],
}
ANCHOR_SOURCE_TO_JAVA_SHEET_NAMES = {
    anchor.SHEET_MAIN: JAVA_ANCHOR_SHEET_NAMES[0],
    anchor.SHEET_COUNTERPARTY: JAVA_ANCHOR_SHEET_NAMES[1],
    anchor.SHEET_OUR_PARTY: JAVA_ANCHOR_SHEET_NAMES[2],
    anchor.SHEET_FEE_DETAIL: JAVA_ANCHOR_SHEET_NAMES[3],
}

GENERAL_RELATION_BLANK_COLUMNS = (
    'relation.relation_contracts（关联合同）',
    '框架合同编号',
)
GENERAL_PURCHASE_REQUEST_BLANK_COLUMNS = (
    'custom_1024_7db9a8ee2b3d4a3f9d9835dd9fee69df（采购申请）',
)
ANCHOR_MAIN_EXECUTOR_COLUMN_INDEX = 20  # Excel T 列
ANCHOR_MAIN_EXECUTOR_OUTPUT_HEADER = '合同执行人'
COUNTERPARTY_CODE_COLUMN = 'counter_party_code（对方主体编码）'

CONTRACT_COLUMN_CANDIDATES = (
    '合同编号', '合同号', '合同编码', 'contract_number（合同编码）', 'contract_number',
)
STATUS_COLUMN_CANDIDATES = ('导入状态', '状态', '是否导入', '导入标记')
TYPE_COLUMN_CANDIDATES = ('智书合同类型', '合同类型', 'contractCategory(智书框架合同类型)')
ORDER_COLUMN_CANDIDATES = ('关联业财订单（必填）', '关联业财订单', '订单编号')
OLD_CODE_COLUMN_CANDIDATES = ('老泛微项目编码', '老泛微编码', '泛微编码', 'OA编号')

EXCEL_ORDER_OVERRIDE_SOURCE = 'Excel关联业财订单覆盖'
EXCEL_CATEGORY_OVERRIDE_SOURCE = 'Excel智书合同类型覆盖'
APPROVAL_NODE_NAME_OVERRIDES = {
    # 自动审批接口使用的节点名与泛微节点名存在一处历史叫法差异。
    '上传电子档': '上传电子版',
    '申请人确认签署类型': '申请人确认签约性质',
}

# 默认不排除任何合同。需要临时排除时再在这里显式添加合同编号。
EXCLUDED_CONTRACT_NUMBERS = frozenset()


def _text(value):
    return general._text(value)


def _contract_key(value):
    return general._contract_number_key(value)


def _first_existing_column(df, names, required=False):
    normalized = {general._normalize_field_name(column): column for column in df.columns}
    for name in names:
        column = normalized.get(general._normalize_field_name(name))
        if column:
            return column
    if required:
        raise KeyError(f'输入表缺少列: {" / ".join(names)}; 实际列: {list(df.columns)}')
    return None


def _split_contract_numbers(value):
    text = _text(value)
    if not text:
        return []
    parts = re.split(r'[\r\n,，;；]+', text)
    return [part.strip() for part in parts if _contract_key(part.strip())]


def _read_mixed_input(path):
    path = _resolve_input_file(path)
    if not path.exists():
        raise FileNotFoundError(
            '智书合同导入清单输入文件不存在: '
            f'{path}\n'
            f'请把唯一一份 Excel 放到 {SOURCE_DIR}，推荐文件名: {DEFAULT_INPUT_FILE_NAME}；'
            '或设置环境变量 CONTRACT_MIXED_ADD_FILE 指向具体文件/目录。'
        )
    if path.is_dir():
        raise IsADirectoryError(
            '智书合同导入清单输入路径是目录，且未找到可自动识别的 Excel: '
            f'{path}\n'
            f'请在该目录只放置一份 Excel，推荐文件名: {DEFAULT_INPUT_FILE_NAME}，'
            '或设置 CONTRACT_MIXED_ADD_FILE 指向具体 .xlsx 文件。'
        )

    # 输入清单的「创建人」仅是业务备注，禁止参与导入值映射。
    # 合同创建人必须由 general/anchor 原流程从泛微源数据解析并执行离职替换规则。
    sheets = pd.read_excel(path, sheet_name=None, dtype=object)
    rows = []
    skipped_sheets = []
    for sheet_name, raw in sheets.items():
        if raw.empty or len(raw.columns) < 1:
            skipped_sheets.append(sheet_name)
            continue
        contract_col = raw.columns[0]
        order_col = raw.columns[1] if len(raw.columns) > 1 else None
        type_col = raw.columns[2] if len(raw.columns) > 2 else None

        for excel_index, row in raw.iterrows():
            raw_contract = _text(row.get(contract_col))
            for contract_number in _split_contract_numbers(raw_contract):
                rows.append({
                    '输入文件': path.name,
                    'Sheet': sheet_name,
                    'Excel行号': int(excel_index) + 2,
                    '源表合同编号': raw_contract,
                    '合同编号': contract_number,
                    '合同key': _contract_key(contract_number),
                    '智书合同类型': _text(row.get(type_col)) if type_col else '',
                    '关联业财订单': _text(row.get(order_col)) if order_col else '',
                    '老泛微编码': '',
                    '导入状态': '',
                })

    if not rows:
        detail = f'; 已跳过空sheet: {", ".join(skipped_sheets)}' if skipped_sheets else ''
        raise RuntimeError(f'输入文件未读到任何合同编号: {path}{detail}')

    result = pd.DataFrame(rows)
    result = result[result['合同key'] != ''].copy()
    result['输入顺序'] = range(1, len(result) + 1)
    return result


def _apply_input_exclusions(input_df):
    result = input_df.copy()
    reasons = [[] for _ in range(len(result))]
    explicitly_excluded = result['合同key'].isin(EXCLUDED_CONTRACT_NUMBERS)
    for pos, is_excluded in enumerate(explicitly_excluded):
        if is_excluded:
            reasons[pos].append('用户指定不处理')

    result['排除原因'] = ['; '.join(items) for items in reasons]
    processable_mask = result['排除原因'].eq('')
    duplicate_mask = result.loc[processable_mask].duplicated('合同key', keep='first')
    duplicate_indexes = duplicate_mask[duplicate_mask].index
    result.loc[duplicate_indexes, '排除原因'] = '重复合同编号,本任务按合同编号仅生成一次'
    result['是否剔除'] = result['排除原因'].map(lambda value: 'Y' if _text(value) else 'N')
    return result


def _tuple_param(values):
    return tuple(dict.fromkeys(value for value in values if _text(value)))


def _query_contract_meta(contract_keys):
    keys = _tuple_param(contract_keys)
    if not keys:
        return {}
    frames = []
    for batch in general._chunked(list(keys), 500):
        df = c.query_db(
            'FW',
            'vspn_xtyy',
            'SELECT htbh AS `合同编号`, htlx AS `合同类型ID`, htzt AS `合同签署状态ID` '
            'FROM uf_htk WHERE htbh IN %(contract_codes)s',
            {'contract_codes': tuple(batch)},
        )
        frames.append(df)
    if not frames:
        return {}
    meta = pd.concat(frames, ignore_index=True)
    result = {}
    for _, row in meta.iterrows():
        key = _contract_key(row.get('合同编号'))
        result.setdefault(key, {
            '合同类型ID': c.format_code(row.get('合同类型ID')),
            '合同签署状态ID': c.format_code(row.get('合同签署状态ID')),
        })
    return result


def _is_excel_anchor_type(value):
    return '主播' in _text(value)


def _route_processable_rows(input_df):
    processable = input_df[input_df['排除原因'].eq('')].copy()
    if processable.empty:
        return processable
    meta = _query_contract_meta(processable['合同key'])
    processable['DB合同类型ID'] = processable['合同key'].map(
        lambda key: meta.get(key, {}).get('合同类型ID', ''))
    processable['DB合同签署状态ID'] = processable['合同key'].map(
        lambda key: meta.get(key, {}).get('合同签署状态ID', ''))
    processable['Excel类型含主播'] = processable['智书合同类型'].map(
        lambda value: 'Y' if _is_excel_anchor_type(value) else '')
    processable['DB类型为主播协议'] = processable['DB合同类型ID'].map(
        lambda value: 'Y' if c.format_code(value) == str(anchor.ANCHOR_CONTRACT_TYPE_CODE) else '')
    anchor_mask = processable['Excel类型含主播'].eq('Y') | processable['DB类型为主播协议'].eq('Y')
    processable['路由流程'] = anchor_mask.map({True: '主播流程', False: '一般流程'})
    processable['路由依据'] = processable.apply(_route_basis, axis=1)
    return processable


def _route_basis(row):
    basis = []
    if _text(row.get('Excel类型含主播')) == 'Y':
        basis.append('Excel智书合同类型含主播')
    if _text(row.get('DB类型为主播协议')) == 'Y':
        basis.append('泛微合同类型=主播协议')
    if basis:
        return '; '.join(basis)
    return '默认一般流程'


def _selected_sql_without_base_where(base_sql):
    marker = 'ORDER BY h.htbh, h.id'
    if marker not in base_sql:
        raise RuntimeError('源 SQL 结构已变化,找不到 ORDER BY 注入点')
    return base_sql.replace(marker, 'WHERE h.htbh IN %(contract_codes)s\n' + marker)


def _selected_anchor_sql():
    """查询路由到主播流程的指定合同，允许输入清单显式覆盖泛微合同类型。"""
    marker = 'ORDER BY h.htbh, h.id'
    type_and_status_filter = (
        'WHERE h.htlx = %(anchor_contract_type_code)s\n'
        '  AND h.htzt IN %(migration_status_codes)s'
    )
    if marker not in anchor.SOURCE_SQL or type_and_status_filter not in anchor.SOURCE_SQL:
        raise RuntimeError('主播源 SQL 结构已变化,找不到合同类型/状态筛选或 ORDER BY 注入点')

    # 混合增补的路由已由输入清单与泛微类型共同决定。输入清单显式标为
    # “主播”的一般合同也必须能使用主播模板，故此处仅保留状态和合同编号筛选。
    sql = anchor.SOURCE_SQL.replace(
        type_and_status_filter,
        'WHERE h.htzt IN %(migration_status_codes)s',
        1,
    )
    return sql.replace(marker, '  AND h.htbh IN %(contract_codes)s\n' + marker, 1)


def _query_selected_source(base_sql, contract_keys, source_label):
    keys = _tuple_param(contract_keys)
    if not keys:
        return pd.DataFrame()
    sql = _selected_sql_without_base_where(base_sql)
    frames = []
    for batch in general._chunked(list(keys), 500):
        frame = c.query_db('FW', 'vspn_xtyy', sql, {'contract_codes': tuple(batch)})
        if not frame.empty:
            frame['数据来源'] = source_label
            frame['强制追加导出'] = ''
            frames.append(frame)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def _query_selected_anchor_source(contract_keys):
    keys = _tuple_param(contract_keys)
    if not keys:
        return pd.DataFrame()
    sql = _selected_anchor_sql()
    frames = []
    for batch in general._chunked(list(keys), 500):
        frame = c.query_db(
            'FW',
            'vspn_xtyy',
            sql,
            {
                'migration_status_codes': anchor.MIGRATION_STATUS_CODES,
                'contract_codes': tuple(batch),
            },
        )
        if not frame.empty:
            frames.append(frame)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def _copy_attrs(target, source):
    target.attrs = dict(source.attrs)
    return target


def _join_order_override_field(items, field, unique=False):
    values = [_text(item.get(field, '')) for item in items]
    if unique:
        values = c.clean_text_values(values)
    return ';'.join(value for value in values if value)


def _order_override_items(order_value):
    items = []
    for order_code in general._split_multi_values(order_value):
        info = c.hand_order_info_for_order(order_code) or c.cleanable_order_info_for_order(order_code) or {}
        items.append({
            '订单编号': order_code,
            '订单标题': _text(info.get('订单标题')),
            '成本中心': _text(info.get('成本中心')),
            '订单开始日': c.format_date(info.get('订单开始日')),
            '订单结束日': c.format_date(info.get('订单结束日')),
            '映射来源': _text(info.get('映射来源')) or EXCEL_ORDER_OVERRIDE_SOURCE,
        })
    return items


def _route_override_lookup(route_df):
    if route_df.empty:
        return pd.DataFrame()
    return route_df.drop_duplicates('合同key', keep='first').set_index('合同key')


def _apply_excel_overrides(source_df, route_df):
    """Apply the input sheet's first-three-column overrides to resolved source rows."""
    if source_df.empty or route_df.empty:
        return source_df

    route = _route_override_lookup(route_df)
    result = source_df.copy()
    result.attrs = dict(source_df.attrs)
    for column in ('Excel覆盖_关联业财订单', 'Excel覆盖_智书合同类型'):
        if column not in result.columns:
            result[column] = ''

    for index, row in result.iterrows():
        key = _contract_key(row.get('合同编号'))
        if not key or key not in route.index:
            continue

        category = _text(route.at[key, '智书合同类型']) if '智书合同类型' in route.columns else ''
        if category:
            result.at[index, 'Excel覆盖_智书合同类型'] = category
            if '合同分类' in result.columns:
                result.at[index, '合同分类'] = category
            if '合同分类依据' in result.columns:
                result.at[index, '合同分类依据'] = f'{EXCEL_CATEGORY_OVERRIDE_SOURCE}: {category}'

        order_value = _text(route.at[key, '关联业财订单']) if '关联业财订单' in route.columns else ''
        if order_value:
            result.at[index, 'Excel覆盖_关联业财订单'] = order_value
            if '订单编号' in result.columns:
                items = _order_override_items(order_value)
                result.at[index, '订单编号'] = _join_order_override_field(items, '订单编号')
                if '订单名称' in result.columns:
                    result.at[index, '订单名称'] = _join_order_override_field(items, '订单标题')
                if '订单成本中心' in result.columns:
                    result.at[index, '订单成本中心'] = _join_order_override_field(items, '成本中心')
                if '订单开始日' in result.columns:
                    result.at[index, '订单开始日'] = _join_order_override_field(items, '订单开始日')
                if '订单结束日' in result.columns:
                    result.at[index, '订单结束日'] = _join_order_override_field(items, '订单结束日')
                if '订单映射来源' in result.columns:
                    result.at[index, '订单映射来源'] = _join_order_override_field(
                        items, '映射来源', unique=True)
    return result


def _resolve_general_sources(contract_keys):
    mcn = _query_selected_source(general.SOURCE_SQL_ALL, contract_keys, '泛微(MCN)')
    if not mcn.empty:
        mcn = mcn[mcn['合同类型ID'].map(c.format_code) != str(anchor.ANCHOR_CONTRACT_TYPE_CODE)].copy()
    event = _query_selected_source(general.SOURCE_SQL_HTSP_ALL, contract_keys, '泛微(赛事)')
    if not mcn.empty and not event.empty:
        mcn_keys = set(mcn['合同编号'].map(_contract_key))
        event = event[~event['合同编号'].map(_contract_key).isin(mcn_keys)].copy()

    resolved_mcn = (
        general.resolve_source_values(mcn, option_table=general.FW_TABLE)
        if not mcn.empty else pd.DataFrame()
    )
    resolved_event = (
        general.resolve_source_values(event, option_table=general.FW_TABLE_HTSP)
        if not event.empty else pd.DataFrame()
    )
    resolved = pd.concat([resolved_mcn, resolved_event], ignore_index=True)
    if resolved.empty:
        return resolved, mcn, event

    general._merge_attrs(resolved, [resolved_mcn, resolved_event])
    resolved = general._apply_supplement_amount_rollup(resolved)
    general._merge_attrs(resolved, [resolved_mcn, resolved_event])
    event_ids = resolved.loc[
        resolved.get('数据来源', pd.Series('', index=resolved.index)).map(_text).eq('泛微(赛事)'),
        'ID',
    ] if 'ID' in resolved.columns else pd.Series(dtype=object)
    resolved.attrs['saishi_plan_map'] = (
        general.load_htsp_plan_detail_map(event_ids) if not event_ids.empty else {}
    )
    return resolved, mcn, event


def _resolve_anchor_sources(contract_keys):
    raw = _query_selected_anchor_source(contract_keys)
    if raw.empty:
        return pd.DataFrame(), raw
    force_keep_contract_numbers = raw.loc[
        raw['合同类型ID'].map(c.format_code) != str(anchor.ANCHOR_CONTRACT_TYPE_CODE),
        '合同编号',
    ]
    resolved = anchor.resolve_source_values(
        raw,
        force_keep_contract_numbers=force_keep_contract_numbers,
    )
    return resolved, raw


def _timestamped_path(path):
    return path.with_name(f'{path.stem}_{datetime.now().strftime("%H%M%S")}{path.suffix}')


def _write_template_sheets_with_fallback(template_file, output_file, sheet_to_df, extra_sheets=None):
    try:
        return c.write_template_sheets(template_file, output_file, sheet_to_df, extra_sheets=extra_sheets)
    except PermissionError:
        fallback = _timestamped_path(output_file)
        print(f'输出文件被占用,改写到: {fallback}')
        return c.write_template_sheets(template_file, fallback, sheet_to_df, extra_sheets=extra_sheets)


def _mixed_template_headers():
    """读取混合模板的表头，以模板字段名覆盖同位置的历史 builder 字段名。"""
    workbook = load_workbook(MIXED_TEMPLATE_FILE, read_only=True, data_only=True)
    try:
        return {
            sheet_name: [_text(cell.value) for cell in next(
                workbook[sheet_name].iter_rows(min_row=1, max_row=1)
            )]
            for sheet_name in JAVA_GENERAL_SHEET_NAMES + JAVA_ANCHOR_SHEET_NAMES
        }
    finally:
        workbook.close()


def _apply_mixed_template_headers(sheet_to_df):
    """保持字段列顺序，将输出 DataFrame 的表头对齐至混合模板。"""
    template_headers = _mixed_template_headers()
    result = {}
    for sheet_name, output_df in sheet_to_df.items():
        headers = template_headers.get(sheet_name)
        if headers is None:
            raise KeyError(f'混合流程模板缺少 sheet: {sheet_name}')
        if len(output_df.columns) != len(headers):
            raise ValueError(
                f'混合流程模板字段数不匹配: {sheet_name}，'
                f'输出 {len(output_df.columns)} 列，模板 {len(headers)} 列'
            )
        aligned = output_df.copy()
        aligned.columns = headers
        result[sheet_name] = aligned
    return result


def _ensure_placeholder_sheets(workbook, sheet_names):
    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)


def _move_sheet_to_index(workbook, sheet_name, target_index):
    worksheet = workbook[sheet_name]
    current_index = workbook.index(worksheet)
    if current_index != target_index:
        workbook.move_sheet(worksheet, offset=target_index - current_index)


def _rename_general_sheets_for_sync(workbook):
    for source_name, target_name in GENERAL_SOURCE_TO_JAVA_SHEET_NAMES.items():
        if target_name not in workbook.sheetnames and source_name in workbook.sheetnames:
            workbook[source_name].title = target_name


def _rename_anchor_sheets_for_sync(workbook):
    for source_name, target_name in ANCHOR_SOURCE_TO_JAVA_SHEET_NAMES.items():
        if target_name not in workbook.sheetnames and source_name in workbook.sheetnames:
            workbook[source_name].title = target_name


def _align_sheet_order_for_zhishu_sync(output_file, flow_name):
    """Align workbook slots with ZhiShuSynServiceImpl.SheetRole indexes."""
    path = Path(output_file)
    workbook = load_workbook(path)
    if flow_name == '一般流程':
        _rename_general_sheets_for_sync(workbook)
        required_names = JAVA_GENERAL_SHEET_NAMES
        placeholder_names = JAVA_ANCHOR_SHEET_NAMES
        ordered_names = required_names + placeholder_names
        active_index = 0
    elif flow_name == '主播流程':
        _rename_anchor_sheets_for_sync(workbook)
        required_names = JAVA_ANCHOR_SHEET_NAMES
        placeholder_names = JAVA_GENERAL_SHEET_NAMES
        ordered_names = placeholder_names + required_names
        active_index = 9
    elif flow_name == '混合流程':
        _rename_general_sheets_for_sync(workbook)
        _rename_anchor_sheets_for_sync(workbook)
        required_names = JAVA_MIXED_SHEET_NAMES
        placeholder_names = ()
        ordered_names = required_names
        active_index = 0
    else:
        raise ValueError(f'不支持的智书同步流程: {flow_name}')

    missing_names = [name for name in required_names if name not in workbook.sheetnames]
    if missing_names:
        raise KeyError(f'{flow_name}输出缺少必需sheet: {missing_names}')

    _ensure_placeholder_sheets(workbook, placeholder_names)
    for target_index, sheet_name in enumerate(ordered_names):
        _move_sheet_to_index(workbook, sheet_name, target_index)
    for worksheet in list(workbook.worksheets):
        if worksheet.title not in ordered_names:
            workbook.remove(worksheet)
    workbook.active = active_index
    workbook.save(path)
    print(f'[智书合同导入清单] 已按Java SheetRole顺序整理{flow_name}文件: {path}')
    return path


def _blank_output_columns(output_df, columns):
    result = output_df.copy()
    for column in columns:
        if column in result.columns:
            result[column] = ''
    return result


def _counterparty_name_keys(*values):
    keys = []
    for value in values:
        key = c.normalize_name(value)
        if key and key not in ('nan', 'none') and key not in keys:
            keys.append(key)
    return keys


def _counterparty_info_names(info):
    if not info:
        return []
    return [
        info.get('source_name', ''),
        info.get('name', ''),
        info.get('target_name', ''),
        info.get('taxpayer_name', ''),
    ]


def _append_unique_code(codes_by_key, key, code):
    code = _text(code)
    if not key or not code:
        return
    codes = codes_by_key.setdefault(key, [])
    if code not in codes:
        codes.append(code)


def _join_counterparty_codes(*code_groups):
    codes = []
    for code_group in code_groups:
        if isinstance(code_group, (list, tuple, set)):
            raw_codes = code_group
        else:
            raw_codes = re.split(r'[;；]+', _text(code_group))
        for code in raw_codes:
            code = _text(code)
            if code and code not in codes:
                codes.append(code)
    return ';'.join(codes)


def _build_mixed_counterparty_code_map(source_df, include_anchor=False):
    if source_df.empty:
        return {}

    vendor_codes_by_key = {}
    customer_codes_by_key = {}
    names = []
    customer_info_map = source_df.attrs.get('customer_info_map', {})
    supplier_info_map = source_df.attrs.get('supplier_info_map', {})

    for info in customer_info_map.values():
        info_names = _counterparty_info_names(info)
        names.extend(info_names)
        for key in _counterparty_name_keys(*info_names):
            _append_unique_code(customer_codes_by_key, key, info.get('code', ''))

    for info in supplier_info_map.values():
        info_names = _counterparty_info_names(info)
        names.extend(info_names)
        for key in _counterparty_name_keys(*info_names):
            _append_unique_code(vendor_codes_by_key, key, info.get('code', ''))

    if include_anchor:
        anchor_vendor_by_id = source_df.attrs.get('anchor_vendor_by_id_number', {})
        anchor_vendor_by_name = source_df.attrs.get('anchor_vendor_by_name', {})
        for source in source_df.to_dict('records'):
            vendor_info = anchor.resolve_anchor_vendor_info(source, anchor_vendor_by_id, anchor_vendor_by_name)
            info_names = [
                vendor_info.get('source_name', ''),
                vendor_info.get('name', ''),
                _text(source.get('主播姓名')),
                _text(source.get('主播昵称')),
            ]
            names.extend(info_names)
            for key in _counterparty_name_keys(*info_names):
                _append_unique_code(vendor_codes_by_key, key, vendor_info.get('code', ''))

    vendor_info_by_name = c.build_hand_vendor_info_by_names(names)
    for key, info in vendor_info_by_name.items():
        _append_unique_code(vendor_codes_by_key, key, info.get('code', ''))

    customer_code_by_name = c.build_customer_map_for_names(names)
    for key, code in customer_code_by_name.items():
        _append_unique_code(customer_codes_by_key, key, code)

    return {
        key: _join_counterparty_codes(vendor_codes_by_key.get(key, []), customer_codes_by_key.get(key, []))
        for key in set(vendor_codes_by_key) & set(customer_codes_by_key)
        if vendor_codes_by_key.get(key) and customer_codes_by_key.get(key)
    }


def _counterparty_code_for_names(name_values, current_code, mixed_code_by_name):
    for key in _counterparty_name_keys(*name_values):
        mixed_code = mixed_code_by_name.get(key)
        if mixed_code:
            return mixed_code
    return current_code


def _apply_general_mixed_counterparty_codes(counterparty_df, source_df):
    if counterparty_df.empty or COUNTERPARTY_CODE_COLUMN not in counterparty_df.columns:
        return counterparty_df
    result = counterparty_df.copy()
    mixed_code_by_name = _build_mixed_counterparty_code_map(source_df)
    if not mixed_code_by_name:
        return result

    customer_info_map = source_df.attrs.get('customer_info_map', {})
    supplier_info_map = source_df.attrs.get('supplier_info_map', {})
    row_index = 0
    for source in source_df.to_dict('records'):
        for customer_id in c.parse_browser_ids(source.get('合同客户ID')):
            if row_index >= len(result):
                return result
            info = customer_info_map.get(customer_id, {})
            result.at[result.index[row_index], COUNTERPARTY_CODE_COLUMN] = _counterparty_code_for_names(
                _counterparty_info_names(info),
                result.iloc[row_index][COUNTERPARTY_CODE_COLUMN],
                mixed_code_by_name,
            )
            row_index += 1
        for supplier_id in c.parse_browser_ids(source.get('合同供应商ID')):
            if row_index >= len(result):
                return result
            info = supplier_info_map.get(supplier_id, {})
            result.at[result.index[row_index], COUNTERPARTY_CODE_COLUMN] = _counterparty_code_for_names(
                _counterparty_info_names(info),
                result.iloc[row_index][COUNTERPARTY_CODE_COLUMN],
                mixed_code_by_name,
            )
            row_index += 1
    return result


def _apply_anchor_mixed_counterparty_codes(counterparty_df, source_df):
    if counterparty_df.empty or COUNTERPARTY_CODE_COLUMN not in counterparty_df.columns:
        return counterparty_df
    result = counterparty_df.copy()
    mixed_code_by_name = _build_mixed_counterparty_code_map(source_df, include_anchor=True)
    if not mixed_code_by_name:
        return result

    customer_info_map = source_df.attrs.get('customer_info_map', {})
    supplier_info_map = source_df.attrs.get('supplier_info_map', {})
    anchor_vendor_by_id = source_df.attrs.get('anchor_vendor_by_id_number', {})
    anchor_vendor_by_name = source_df.attrs.get('anchor_vendor_by_name', {})
    row_index = 0
    for source in source_df.to_dict('records'):
        if row_index >= len(result):
            return result
        vendor_info = anchor.resolve_anchor_vendor_info(source, anchor_vendor_by_id, anchor_vendor_by_name)
        result.at[result.index[row_index], COUNTERPARTY_CODE_COLUMN] = _counterparty_code_for_names(
            [
                vendor_info.get('source_name', ''),
                vendor_info.get('name', ''),
                _text(source.get('主播姓名')),
                _text(source.get('主播昵称')),
            ],
            result.iloc[row_index][COUNTERPARTY_CODE_COLUMN],
            mixed_code_by_name,
        )
        row_index += 1
        for customer_id in c.parse_browser_ids(source.get('合同客户ID')):
            if row_index >= len(result):
                return result
            info = customer_info_map.get(customer_id, {})
            result.at[result.index[row_index], COUNTERPARTY_CODE_COLUMN] = _counterparty_code_for_names(
                _counterparty_info_names(info),
                result.iloc[row_index][COUNTERPARTY_CODE_COLUMN],
                mixed_code_by_name,
            )
            row_index += 1
        for supplier_id in c.parse_browser_ids(source.get('合同供应商ID')):
            if row_index >= len(result):
                return result
            info = supplier_info_map.get(supplier_id, {})
            result.at[result.index[row_index], COUNTERPARTY_CODE_COLUMN] = _counterparty_code_for_names(
                _counterparty_info_names(info),
                result.iloc[row_index][COUNTERPARTY_CODE_COLUMN],
                mixed_code_by_name,
            )
            row_index += 1
    return result


def _rename_anchor_main_executor_header(output_file):
    path = Path(output_file)
    workbook = load_workbook(path)
    renamed = False
    for sheet_name in (anchor.SHEET_MAIN, JAVA_ANCHOR_SHEET_NAMES[0]):
        if sheet_name in workbook.sheetnames:
            workbook[sheet_name].cell(
                row=1,
                column=ANCHOR_MAIN_EXECUTOR_COLUMN_INDEX,
            ).value = ANCHOR_MAIN_EXECUTOR_OUTPUT_HEADER
            renamed = True
    if renamed:
        workbook.save(path)
    return path


def _build_general_sheet_frames(source_df):
    headers = general._template_headers()
    if source_df.empty:
        sheets = {
            sheet_name: pd.DataFrame(columns=headers[sheet_name])
            for sheet_name in GENERAL_SOURCE_TO_JAVA_SHEET_NAMES
        }
        sheets[general.SHEET_RELATION] = _blank_output_columns(
            sheets[general.SHEET_RELATION],
            GENERAL_RELATION_BLANK_COLUMNS,
        )
        sheets[general.SHEET_PURCHASE_REQUEST] = _blank_output_columns(
            sheets[general.SHEET_PURCHASE_REQUEST],
            GENERAL_PURCHASE_REQUEST_BLANK_COLUMNS,
        )
        return headers, sheets

    relation_df, _ = general.build_relation_output(source_df, headers[general.SHEET_RELATION])
    relation_df = _blank_output_columns(relation_df, GENERAL_RELATION_BLANK_COLUMNS)
    order_detail_df, _ = general.build_order_detail_output(source_df, headers[general.SHEET_ORDER_DETAIL])
    counterparty_df, _ = general.build_counterparty_output(source_df, headers[general.SHEET_COUNTERPARTY])
    counterparty_df = _apply_general_mixed_counterparty_codes(counterparty_df, source_df)
    our_party_df, _ = general.build_our_party_output(source_df, headers[general.SHEET_OUR_PARTY])
    purchase_request_df = general.build_purchase_request_output(
        source_df,
        headers[general.SHEET_PURCHASE_REQUEST],
    )
    purchase_request_df = _blank_output_columns(
        purchase_request_df,
        GENERAL_PURCHASE_REQUEST_BLANK_COLUMNS,
    )
    return headers, {
        general.SHEET_MAIN: general.build_main_output(source_df, headers[general.SHEET_MAIN]),
        general.SHEET_RELATION: relation_df,
        general.SHEET_RELATED_ORDER: general.build_related_order_output(
            source_df, headers[general.SHEET_RELATED_ORDER]),
        general.SHEET_PURCHASE_REQUEST: purchase_request_df,
        general.SHEET_ORDER_DETAIL: order_detail_df,
        general.SHEET_COUNTERPARTY: counterparty_df,
        general.SHEET_OUR_PARTY: our_party_df,
        general.SHEET_PAYMENT_PLAN: general.build_payment_plan_output(
            source_df, headers[general.SHEET_PAYMENT_PLAN]),
        general.SHEET_COLLECTION_PLAN: general.build_collection_plan_output(
            source_df, headers[general.SHEET_COLLECTION_PLAN]),
    }


def _build_anchor_sheet_frames(source_df):
    headers = anchor._template_headers()
    if source_df.empty:
        return headers, {
            sheet_name: pd.DataFrame(columns=headers[sheet_name])
            for sheet_name in ANCHOR_SOURCE_TO_JAVA_SHEET_NAMES
        }

    counterparty_df, _ = anchor.build_counterparty_output(source_df, headers[anchor.SHEET_COUNTERPARTY])
    counterparty_df = _apply_anchor_mixed_counterparty_codes(counterparty_df, source_df)
    our_party_df, _ = anchor.build_our_party_output(source_df, headers[anchor.SHEET_OUR_PARTY])
    fee_detail_df, _ = anchor.build_fee_detail_output(source_df, headers[anchor.SHEET_FEE_DETAIL])
    return headers, {
        anchor.SHEET_MAIN: anchor.build_main_output(source_df, headers[anchor.SHEET_MAIN]),
        anchor.SHEET_COUNTERPARTY: counterparty_df,
        anchor.SHEET_OUR_PARTY: our_party_df,
        anchor.SHEET_FEE_DETAIL: fee_detail_df,
    }


def _build_attachment_audit(source_df, flow_module, headers):
    if source_df.empty:
        return pd.DataFrame(), pd.DataFrame()
    old_output_dir = flow_module.OUTPUT_DIR
    try:
        flow_module.OUTPUT_DIR = OUTPUT_DIR
        _, _, manifest_df, missing_df = flow_module.build_contract_attachment_output(
            source_df,
            headers[flow_module.SHEET_CONTRACT_ATTACHMENT],
            headers[flow_module.SHEET_OTHER_ATTACHMENT],
        )
    finally:
        flow_module.OUTPUT_DIR = old_output_dir
    return manifest_df, missing_df


def _write_mixed_workbook(general_source_df, anchor_source_df):
    if not MIXED_TEMPLATE_FILE.exists():
        raise FileNotFoundError(f'混合流程模板不存在: {MIXED_TEMPLATE_FILE}')

    MIXED_ATTACHMENT_ROOT.mkdir(parents=True, exist_ok=True)
    general_headers, general_sheets = _build_general_sheet_frames(general_source_df)
    anchor_headers, anchor_sheets = _build_anchor_sheet_frames(anchor_source_df)
    general_manifest_df, general_missing_df = _build_attachment_audit(
        general_source_df, general, general_headers)
    anchor_manifest_df, anchor_missing_df = _build_attachment_audit(
        anchor_source_df, anchor, anchor_headers)

    general_java_sheets = {
        GENERAL_SOURCE_TO_JAVA_SHEET_NAMES[source_name]: output_df
        for source_name, output_df in general_sheets.items()
    }
    anchor_java_sheets = {
        ANCHOR_SOURCE_TO_JAVA_SHEET_NAMES[source_name]: output_df
        for source_name, output_df in anchor_sheets.items()
    }
    template_sheets = _apply_mixed_template_headers(
        {**general_java_sheets, **anchor_java_sheets}
    )
    path = _write_template_sheets_with_fallback(
        MIXED_TEMPLATE_FILE,
        MIXED_OUTPUT_FILE,
        template_sheets,
    )
    path = _align_sheet_order_for_zhishu_sync(path, '混合流程')
    path = _rename_anchor_main_executor_header(path)
    print(f'[智书合同导入清单] 已生成混合导入文件: {path}')
    return (
        path,
        general_manifest_df,
        general_missing_df,
        anchor_manifest_df,
        anchor_missing_df,
    )


def _is_archived_for_sync(row):
    approval_status = _text(row.get('合同审批状态'))
    if approval_status == '归档':
        return True

    status_text = _text(row.get('合同签署状态'))
    status_id = c.format_code(row.get('合同签署状态ID'))
    if '归档' in status_text:
        return True

    # uf_htsp(赛事合同审批台账)的 htzt 独立码表:0=审批中,1=归档。
    if _text(row.get('数据来源')) == '泛微(赛事)':
        return status_id == '1'

    # uf_htk/主播合同使用通用码表:0=审批中,1=审批完成,2=已归档。
    return status_id == '2'


def _contract_numbers_by_archive_status(general_source_df, anchor_source_df):
    status_by_contract = {}
    archived_by_contract = {}
    for source_df in (general_source_df, anchor_source_df):
        if source_df.empty:
            continue
        for _, row in source_df.iterrows():
            contract_number = _text(row.get('合同编号'))
            if contract_number and contract_number not in status_by_contract:
                status_by_contract[contract_number] = _text(row.get('合同审批状态'))
                archived_by_contract[contract_number] = _is_archived_for_sync(row)
    archived = [
        contract_number
        for contract_number in status_by_contract
        if archived_by_contract.get(contract_number)
    ]
    other = [
        contract_number
        for contract_number in status_by_contract
        if not archived_by_contract.get(contract_number)
    ]
    return archived, other, status_by_contract


def _approval_node_groups_for_zero_status_contracts(status_by_contract, contract_numbers):
    groups = {}
    skipped = []
    for contract_number in contract_numbers:
        status = status_by_contract.get(contract_number, '')
        node_name = APPROVAL_NODE_NAME_OVERRIDES.get(status, status)
        if not node_name:
            skipped.append(contract_number)
            continue
        groups.setdefault(node_name, []).append(contract_number)
    return groups, skipped


def _write_json_file(path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', encoding='utf-8') as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)
        file.write('\n')
    return path


def _write_sync_request_files(general_source_df, anchor_source_df, mixed_path):
    archived, other, status_by_contract = _contract_numbers_by_archive_status(general_source_df, anchor_source_df)
    approve_node_groups, approve_skipped = _approval_node_groups_for_zero_status_contracts(
        status_by_contract,
        other,
    )
    file_path = Path(mixed_path).resolve().as_posix()
    fallback_root = MIXED_ATTACHMENT_ROOT.resolve().as_posix()
    archived_path = _write_json_file(ARCHIVED_REQUEST_FILE, {
        'filePath': file_path,
        'contractNumbers': archived,
        'contractFileFallbackRoot': fallback_root,
        'contractStatusCode': '9',
        'threadCount': 5,
        'batchSize': 10,
    })
    other_path = _write_json_file(OTHER_REQUEST_FILE, {
        'filePath': file_path,
        'contractNumbers': other,
        'contractFileFallbackRoot': fallback_root,
        'contractStatusCode': '0',
        'threadCount': 5,
        'batchSize': 10,
    })
    approve_path = _write_json_file(APPROVE_TO_NODE_REQUEST_FILE, approve_node_groups)
    yecai_path = _write_json_file(YECAI_SYNC_REQUEST_FILE, {
        'contractNumbers': list(status_by_contract.keys()),
        'threadCount': 5,
    })
    print(
        '[智书合同导入清单] 已生成同步请求:',
        f'归档(9) {len(archived)} 个 -> {archived_path};',
        f'其他(0) {len(other)} 个 -> {other_path};',
        f'0状态自动审批 {sum(len(items) for items in approve_node_groups.values())} 个 -> {approve_path};',
        f'同步业财 {len(status_by_contract)} 个 -> {yecai_path}',
    )
    if approve_skipped:
        print(f'[智书合同导入清单] 0状态自动审批跳过 {len(approve_skipped)} 个合同: 合同审批状态为空')
    return archived_path, other_path, approve_path, yecai_path


def _flow_scope_summary(flow_name, source_df, input_df, output_file):
    return pd.DataFrame([
        {'项目': '流程', '值': flow_name},
        {'项目': '输出文件', '值': output_file.name},
        {'项目': '输入明细行数', '值': len(input_df)},
        {'项目': '输入合同数', '值': input_df['合同key'].nunique() if len(input_df) else 0},
        {'项目': '导入合同数', '值': len(source_df)},
        {'项目': '输入来源', '值': str(INPUT_FILE)},
    ])


def _write_general_workbook(source_df, input_df):
    if source_df.empty:
        print('[智书合同导入清单] 一般流程无可生成合同,跳过一般流程导入文件')
        return None, pd.DataFrame(), pd.DataFrame()

    headers = general._template_headers()
    main_df = general.build_main_output(source_df, headers[general.SHEET_MAIN])
    relation_df, _ = general.build_relation_output(source_df, headers[general.SHEET_RELATION])
    relation_df = _blank_output_columns(relation_df, GENERAL_RELATION_BLANK_COLUMNS)
    related_order_df = general.build_related_order_output(source_df, headers[general.SHEET_RELATED_ORDER])
    purchase_request_df = general.build_purchase_request_output(source_df, headers[general.SHEET_PURCHASE_REQUEST])
    purchase_request_df = _blank_output_columns(
        purchase_request_df,
        GENERAL_PURCHASE_REQUEST_BLANK_COLUMNS,
    )
    order_detail_df, _ = general.build_order_detail_output(source_df, headers[general.SHEET_ORDER_DETAIL])
    counterparty_df, _ = general.build_counterparty_output(source_df, headers[general.SHEET_COUNTERPARTY])
    counterparty_df = _apply_general_mixed_counterparty_codes(counterparty_df, source_df)
    our_party_df, _ = general.build_our_party_output(source_df, headers[general.SHEET_OUR_PARTY])
    payment_df = general.build_payment_plan_output(source_df, headers[general.SHEET_PAYMENT_PLAN])
    collection_df = general.build_collection_plan_output(source_df, headers[general.SHEET_COLLECTION_PLAN])

    old_output_dir = general.OUTPUT_DIR
    try:
        general.OUTPUT_DIR = OUTPUT_DIR
        contract_attachment_df, other_attachment_df, manifest_df, missing_df = general.build_contract_attachment_output(
            source_df,
            headers[general.SHEET_CONTRACT_ATTACHMENT],
            headers[general.SHEET_OTHER_ATTACHMENT],
        )
    finally:
        general.OUTPUT_DIR = old_output_dir

    path = general._write_template_sheets_with_fallback(
        general.TEMPLATE_FILE,
        GENERAL_OUTPUT_FILE,
        {
            general.SHEET_MAIN: main_df,
            general.SHEET_RELATION: relation_df,
            general.SHEET_RELATED_ORDER: related_order_df,
            general.SHEET_PURCHASE_REQUEST: purchase_request_df,
            general.SHEET_ORDER_DETAIL: order_detail_df,
            general.SHEET_COUNTERPARTY: counterparty_df,
            general.SHEET_OUR_PARTY: our_party_df,
            general.SHEET_PAYMENT_PLAN: payment_df,
            general.SHEET_COLLECTION_PLAN: collection_df,
            general.SHEET_CONTRACT_ATTACHMENT: contract_attachment_df,
            general.SHEET_OTHER_ATTACHMENT: other_attachment_df,
        },
        extra_sheets={
            '处理范围': _flow_scope_summary('一般流程', source_df, input_df, GENERAL_OUTPUT_FILE),
            '输入清单': input_df,
            '合同分类核对': general.build_category_audit_df(source_df),
            '订单映射核对': general.build_order_audit_df(source_df),
        },
    )
    path = _align_sheet_order_for_zhishu_sync(path, '一般流程')
    print(f'[智书合同导入清单] 已生成一般流程导入文件: {path}')
    return path, manifest_df, missing_df


def _write_anchor_workbook(source_df, input_df):
    if source_df.empty:
        print('[智书合同导入清单] 主播流程无可生成合同,跳过主播流程导入文件')
        return None, pd.DataFrame(), pd.DataFrame()

    headers = anchor._template_headers()
    main_df = anchor.build_main_output(source_df, headers[anchor.SHEET_MAIN])
    counterparty_df, _ = anchor.build_counterparty_output(source_df, headers[anchor.SHEET_COUNTERPARTY])
    counterparty_df = _apply_anchor_mixed_counterparty_codes(counterparty_df, source_df)
    our_party_df, _ = anchor.build_our_party_output(source_df, headers[anchor.SHEET_OUR_PARTY])
    fee_detail_df, _ = anchor.build_fee_detail_output(source_df, headers[anchor.SHEET_FEE_DETAIL])

    old_output_dir = anchor.OUTPUT_DIR
    try:
        anchor.OUTPUT_DIR = OUTPUT_DIR
        contract_attachment_df, other_attachment_df, manifest_df, missing_df = anchor.build_contract_attachment_output(
            source_df,
            headers[anchor.SHEET_CONTRACT_ATTACHMENT],
            headers[anchor.SHEET_OTHER_ATTACHMENT],
        )
    finally:
        anchor.OUTPUT_DIR = old_output_dir

    path = _write_template_sheets_with_fallback(
        anchor.TEMPLATE_FILE,
        ANCHOR_OUTPUT_FILE,
        {
            anchor.SHEET_MAIN: main_df,
            anchor.SHEET_COUNTERPARTY: counterparty_df,
            anchor.SHEET_OUR_PARTY: our_party_df,
            anchor.SHEET_FEE_DETAIL: fee_detail_df,
            anchor.SHEET_CONTRACT_ATTACHMENT: contract_attachment_df,
            anchor.SHEET_OTHER_ATTACHMENT: other_attachment_df,
        },
        extra_sheets={
            '处理范围': _flow_scope_summary('主播流程', source_df, input_df, ANCHOR_OUTPUT_FILE),
            '输入清单': input_df,
        },
    )
    anchor._add_flow_audit_sheet(path, source_df)
    anchor._add_platform_audit_sheet(path, source_df)
    path = _align_sheet_order_for_zhishu_sync(path, '主播流程')
    path = _rename_anchor_main_executor_header(path)
    print(f'[智书合同导入清单] 已生成主播流程导入文件: {path}')
    return path, manifest_df, missing_df


def _source_keys(source_df):
    if source_df.empty or '合同编号' not in source_df.columns:
        return set()
    return set(source_df['合同编号'].map(_contract_key)) - {''}


def _build_unresolved_df(route_df, general_keys, anchor_keys, anchor_raw_keys):
    generated_keys = set(general_keys) | set(anchor_keys)
    rows = []
    for _, row in route_df.iterrows():
        key = row['合同key']
        if key in generated_keys:
            continue
        if row['路由流程'] == '主播流程':
            reason = (
                '主播流程解析后未保留,请看主播流程异常口径'
                if key in anchor_raw_keys else
                '主播源未找到或不满足主播流程SQL条件'
            )
        else:
            reason = '一般流程源未找到(uf_htk/uf_htsp均未命中)'
        rows.append({
            '合同编号': row['合同编号'],
            '合同key': key,
            '路由流程': row['路由流程'],
            '路由依据': row['路由依据'],
            '未生成原因': reason,
            'Excel行号': row['Excel行号'],
            '智书合同类型': row['智书合同类型'],
        })
    return pd.DataFrame(rows)


def _build_input_audit(input_df, route_df, general_keys, anchor_keys, unresolved_df):
    result = input_df.copy()
    route = route_df.drop_duplicates('合同key').set_index('合同key') if not route_df.empty else pd.DataFrame()
    unresolved = (
        unresolved_df.drop_duplicates('合同key').set_index('合同key')['未生成原因'].to_dict()
        if not unresolved_df.empty else {}
    )
    generated_keys = set(general_keys) | set(anchor_keys)
    for column in ('路由流程', '路由依据', 'DB合同类型ID', 'DB合同签署状态ID'):
        result[column] = result['合同key'].map(
            lambda key: _text(route.at[key, column]) if not route.empty and key in route.index else '')
    result['生成状态'] = result.apply(
        lambda row: _generation_status(row, generated_keys, unresolved),
        axis=1,
    )
    return result


def _generation_status(row, generated_keys, unresolved):
    if _text(row.get('排除原因')):
        return '未生成: ' + _text(row.get('排除原因'))
    key = row.get('合同key')
    if key in generated_keys:
        return '已生成'
    return '未生成: ' + unresolved.get(key, '未命中源数据')


def _write_audit_workbook(input_audit_df, route_df, unresolved_df,
                          general_manifest_df, general_missing_df, anchor_manifest_df, anchor_missing_df,
                          mixed_path):
    summary = pd.DataFrame([
        {'项目': '输入文件', '值': str(INPUT_FILE)},
        {'项目': '输入展开行数', '值': len(input_audit_df)},
        {'项目': '输入合同数', '值': input_audit_df['合同key'].nunique()},
        {'项目': '剔除行数', '值': int((input_audit_df['是否剔除'] == 'Y').sum())},
        {'项目': '待处理合同数', '值': route_df['合同key'].nunique() if len(route_df) else 0},
        {'项目': '一般流程合同数', '值': int((route_df.get('路由流程', pd.Series(dtype=str)) == '一般流程').sum())},
        {'项目': '主播流程合同数', '值': int((route_df.get('路由流程', pd.Series(dtype=str)) == '主播流程').sum())},
        {'项目': '混合流程输出', '值': str(mixed_path or '')},
    ])
    sheets = {
        '处理汇总': summary,
        '输入清单_处理结果': input_audit_df,
        '待处理_路由结果': route_df,
        '未生成清单': unresolved_df,
        '一般流程合同附件下载清单': general_manifest_df,
        '一般流程合同附件DOCID缺失': general_missing_df,
        '主播流程合同附件下载清单': anchor_manifest_df,
        '主播流程合同附件DOCID缺失': anchor_missing_df,
    }
    try:
        path = c.write_exceptions(AUDIT_FILE, sheets)
    except PermissionError:
        fallback = _timestamped_path(AUDIT_FILE)
        print(f'处理清单被占用,改写到: {fallback}')
        path = c.write_exceptions(fallback, sheets)
    if path:
        print(f'[智书合同导入清单] 已生成处理清单: {path}')
    return Path(path) if path else None


def run(suppress_audit=False):
    """生成混合导入结果；API 调用可跳过仅供审计的处理清单 Excel。"""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    input_df = _read_mixed_input(INPUT_FILE)
    input_df = _apply_input_exclusions(input_df)
    route_df = _route_processable_rows(input_df)

    general_keys = route_df.loc[route_df['路由流程'].eq('一般流程'), '合同key'] if not route_df.empty else []
    anchor_keys = route_df.loc[route_df['路由流程'].eq('主播流程'), '合同key'] if not route_df.empty else []

    print(
        '[智书合同导入清单] 输入:',
        f'{len(input_df)} 行 / {input_df["合同key"].nunique()} 个合同;',
        f'剔除 {(input_df["是否剔除"] == "Y").sum()} 行;',
        f'一般流程 {len(list(general_keys))} 个;',
        f'主播流程 {len(list(anchor_keys))} 个',
    )

    general_source_df, general_mcn_raw, general_event_raw = _resolve_general_sources(general_keys)
    anchor_source_df, anchor_raw_df = _resolve_anchor_sources(anchor_keys)
    general_source_df = _apply_excel_overrides(general_source_df, route_df)
    anchor_source_df = _apply_excel_overrides(anchor_source_df, route_df)

    (
        mixed_path,
        general_manifest_df,
        general_missing_df,
        anchor_manifest_df,
        anchor_missing_df,
    ) = _write_mixed_workbook(general_source_df, anchor_source_df)
    archived_request_path, other_request_path, approve_node_request_path, yecai_request_path = _write_sync_request_files(
        general_source_df,
        anchor_source_df,
        mixed_path,
    )

    general_output_keys = _source_keys(general_source_df)
    anchor_output_keys = _source_keys(anchor_source_df)
    anchor_raw_keys = _source_keys(anchor_raw_df)
    unresolved_df = _build_unresolved_df(route_df, general_output_keys, anchor_output_keys, anchor_raw_keys)
    input_audit_df = _build_input_audit(
        input_df,
        route_df,
        general_output_keys,
        anchor_output_keys,
        unresolved_df,
    )

    audit_path = None
    if suppress_audit:
        print('[智书合同导入清单] API 调用跳过处理清单 Excel')
    else:
        audit_path = _write_audit_workbook(
            input_audit_df,
            route_df,
            unresolved_df,
            general_manifest_df,
            general_missing_df,
            anchor_manifest_df,
            anchor_missing_df,
            mixed_path,
        )

    return [
        path
        for path in (
            mixed_path,
            audit_path,
            archived_request_path,
            other_request_path,
            approve_node_request_path,
            yecai_request_path,
        )
        if path
    ]


if __name__ == '__main__':
    run()
