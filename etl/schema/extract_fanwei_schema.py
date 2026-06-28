# -*- coding: utf-8 -*-
"""
读取泛微生产库所有数据库表的字段名称、中文名称、枚举值
输出到 Excel 文件

数据关系:
  workflow_bill          -> 表定义 (TABLENAME, DETAILTABLENAME)
  workflow_billfield     -> 字段定义 (BILLID→workflow_bill.ID, FIELDNAME, FIELDLABEL→htmllabelindex.ID)
  workflow_selectitem    -> 枚举值  (FIELDID→workflow_billfield.id, SELECTVALUE, SELECTNAME)
  htmllabelindex         -> 标签索引 (ID, INDEXDESC)
  htmllabelinfo          -> 标签多语言 (INDEXID→htmllabelindex.ID, LABELNAME, LANGUAGEID=7为简体中文)
"""
import re
import pymysql
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
import os

# ============================================================
# 数据库连接配置
# ============================================================
DB_CONFIG = {
    'host': 'gz-cdbrg-0qkmgfbf.sql.tencentcdb.com',
    'port': 58342,
    'user': 'bfd_no1',
    'password': 'vspo_te3pV4y3',
    'database': 'vspn_xtyy',
    'charset': 'utf8',
}

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'resources', 'reference')
os.makedirs(OUTPUT_DIR, exist_ok=True)


# ============================================================
# 工具函数：解析泛微多语言编码
# 实际数据库格式（通过 HEX 确认）:
#   ~`~`7 中文`~`8 English`~`9 繁體中文`~`~
#   开头: ~`~` (tilde-backtick-tilde-backtick)
#   分隔: `~`  (backtick-tilde-backtick)
#   结尾: `~`~ (backtick-tilde-backtick-tilde)
# ============================================================
def parse_multilang(text, lang_id=7):
    """
    解析泛微多语言字段.
    格式: ~`~`7 中文`~`8 English`~`9 繁体`~`~
    返回指定语言ID的文本; 如果不是多语言格式,直接返回原文本.
    """
    if not text or not isinstance(text, str):
        return ''

    t = text.strip()

    # 如果不是多语言格式,直接返回
    if '~`~`' not in t and '`~`' not in t:
        return t

    # 去掉开头的 ~`~`
    if t.startswith('~`~`'):
        t = t[4:]

    # 去掉结尾的 `~`~ (注意与开头不同!)
    if t.endswith('`~`~'):
        t = t[:-4]

    # 按 `~` (backtick-tilde-backtick) 分割各语言段
    # 例如: "7 中文`~`8 English`~`9 繁体" -> ["7 中文", "8 English", "9 繁体"]
    segments = t.split('`~`')

    result = {}
    for seg in segments:
        seg = seg.strip()
        # 匹配 "数字 文本"
        m = re.match(r'^(\d+)\s+(.*)', seg, re.DOTALL)
        if m:
            lid = int(m.group(1))
            val = m.group(2).strip()
            result[lid] = val

    if lang_id in result:
        return result[lang_id]

    # 如果没找到指定语言, 返回第一个找到的
    if result:
        return list(result.values())[0]

    return ''


# ============================================================
# 主提取逻辑
# ============================================================
def extract_all():
    conn = pymysql.connect(**DB_CONFIG)
    cursor = conn.cursor()

    # --------------------------------------------------------
    # 1. 构建标签缓存 (htmllabelindex + htmllabelinfo)
    # --------------------------------------------------------
    print('[1/4] 加载中文标签缓存...')
    label_cache = {}

    # 从 htmllabelindex 加载 (ID → INDEXDESC)
    cursor.execute('SELECT ID, INDEXDESC FROM htmllabelindex')
    for row in cursor.fetchall():
        label_cache[row[0]] = parse_multilang(row[1] or '')

    # 从 htmllabelinfo 加载中文翻译 (LANGUAGEID=7 简体中文)
    # 这会覆盖 htmllabelindex 中的值（更精确的中文）
    cursor.execute(
        'SELECT INDEXID, LABELNAME FROM htmllabelinfo WHERE LANGUAGEID=7'
    )
    for row in cursor.fetchall():
        label_cache[row[0]] = parse_multilang(row[1] or '')

    print(f'    已加载 {len(label_cache):,} 条标签')

    # --------------------------------------------------------
    # 2. 构建枚举值缓存 (workflow_selectitem)
    #    key = FIELDID (对应 workflow_billfield.id)
    # --------------------------------------------------------
    print('[2/4] 加载枚举值缓存...')
    select_items_cache = defaultdict(list)

    cursor.execute(
        'SELECT FIELDID, SELECTVALUE, SELECTNAME '
        'FROM workflow_selectitem '
        'ORDER BY FIELDID, LISTORDER'
    )
    for row in cursor.fetchall():
        field_id = row[0]
        value = row[1]
        name = parse_multilang(row[2] or '') if row[2] else ''
        select_items_cache[field_id].append((value, name))

    print(f'    已加载 {len(select_items_cache):,} 个字段的枚举值')

    # --------------------------------------------------------
    # 3. 获取所有 bills 和 fields
    # --------------------------------------------------------
    print('[3/4] 获取表定义和字段定义...')

    cursor.execute(
        'SELECT ID, NAMELABEL, TABLENAME, DETAILTABLENAME '
        'FROM workflow_bill ORDER BY ID'
    )
    bills = cursor.fetchall()
    print(f'    共 {len(bills):,} 个 bill（表定义）')

    cursor.execute(
        'SELECT id, BILLID, FIELDNAME, FIELDLABEL, FIELDHTMLTYPE, FIELDDBTYPE, DETAILTABLE '
        'FROM workflow_billfield ORDER BY BILLID, DSPORDER'
    )
    all_fields = cursor.fetchall()
    print(f'    共 {len(all_fields):,} 个字段')

    # 按 BILLID 分组
    fields_by_bill = defaultdict(list)
    for f in all_fields:
        fields_by_bill[f[1]].append(f)

    # --------------------------------------------------------
    # 4. 生成 Excel
    # --------------------------------------------------------
    print('[4/4] 生成 Excel 文件...')

    wb = openpyxl.Workbook()

    # --- 样式定义 ---
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # HTML类型映射
    html_type_map = {
        '1': '文本输入',
        '2': '多行文本',
        '3': '日期/时间',
        '4': '整数',
        '5': '小数',
        '6': '浏览按钮',
        '7': '复选框',
    }

    # ================================================================
    # Sheet 1: 表汇总
    # ================================================================
    ws_summary = wb.active
    ws_summary.title = '表汇总'

    summary_headers = ['序号', 'Bill ID', '表中文名称', '主表名', '明细表名', '字段数', '有枚举的字段数']
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ================================================================
    # Sheet 2: 字段明细
    # ================================================================
    ws_fields = wb.create_sheet('字段明细')
    field_headers = ['主表名', '明细表名', '表中文名', '字段名', '字段中文名', '字段类型(HTML)', '数据库类型', '是否有枚举', '枚举值(值=中文名)']
    for col, h in enumerate(field_headers, 1):
        cell = ws_fields.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ================================================================
    # Sheet 3: 枚举值明细
    # ================================================================
    ws_enum = wb.create_sheet('枚举值明细')
    enum_headers = ['主表名', '明细表名', '表中文名', '字段名', '字段中文名', '枚举值', '枚举中文名']
    for col, h in enumerate(enum_headers, 1):
        cell = ws_enum.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ================================================================
    # Sheet 4: 所有表名（从数据库查出的实际表）
    # ================================================================
    ws_tables = wb.create_sheet('数据库实际表')
    # 获取数据库中所有以 formtable_ 或 bill_ 或 uf_ 开头的表
    cursor.execute("SHOW TABLES")
    all_db_tables = [r[0] for r in cursor.fetchall()]
    # 筛选与业务相关的表
    biz_tables = [t for t in all_db_tables if
                  t.startswith('formtable_') or
                  t.startswith('bill_') or
                  t.startswith('uf_')]
    table_headers = ['序号', '表名', '所属 Bill', 'Bill 中文名']
    for col, h in enumerate(table_headers, 1):
        cell = ws_tables.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 构建 bill 表名到 bill info 的映射
    bill_by_table = {}  # table_name -> (bill_id, chinese_name, detail_table)
    for bill in bills:
        bill_id, label_id, table_name, detail_table = bill
        chinese_name = label_cache.get(label_id, '')
        if table_name:
            bill_by_table[table_name.lower()] = (bill_id, chinese_name, detail_table or '')
        if detail_table:
            bill_by_table[detail_table.lower()] = (bill_id, chinese_name, detail_table or '')

    row_table = 2
    for idx, t_name in enumerate(sorted(biz_tables)):
        t_lower = t_name.lower()
        bill_info = bill_by_table.get(t_lower, ('', '', ''))
        ws_tables.cell(row=row_table, column=1, value=idx + 1).alignment = cell_alignment
        ws_tables.cell(row=row_table, column=2, value=t_name).alignment = cell_alignment
        ws_tables.cell(row=row_table, column=3, value=bill_info[0] if bill_info[0] else '未匹配').alignment = cell_alignment
        ws_tables.cell(row=row_table, column=4, value=bill_info[1]).alignment = cell_alignment
        for c in range(1, 5):
            ws_tables.cell(row=row_table, column=c).border = thin_border
        row_table += 1

    ws_tables.column_dimensions['A'].width = 8
    ws_tables.column_dimensions['B'].width = 35
    ws_tables.column_dimensions['C'].width = 12
    ws_tables.column_dimensions['D'].width = 30
    ws_tables.freeze_panes = 'A2'
    ws_tables.auto_filter.ref = ws_tables.dimensions

    # ================================================================
    # 填充数据
    # ================================================================
    row_num_summary = 2
    row_num_field = 2
    row_enum = 2

    for idx, bill in enumerate(bills):
        bill_id, label_id, table_name, detail_table = bill

        # 获取中文名称
        chinese_name = label_cache.get(label_id, '')
        # 如果中文名还是多语言编码, 再解析一次
        chinese_name = parse_multilang(chinese_name)

        fields = fields_by_bill.get(bill_id, [])
        enum_count = 0

        for f in fields:
            f_id = f[0]         # workflow_billfield.id
            f_name = f[2]       # FIELDNAME
            f_label_id = f[3]   # FIELDLABEL → htmllabelindex.ID
            f_html_type = f[4]  # FIELDHTMLTYPE
            f_db_type = f[5]    # FIELDDBTYPE
            f_detail_table = f[6]  # DETAILTABLE

            # 字段中文名
            f_chinese = label_cache.get(f_label_id, '') if f_label_id else ''
            f_chinese = parse_multilang(f_chinese)

            # HTML类型中文
            html_type_cn = html_type_map.get(str(f_html_type), f'类型{f_html_type}')

            # 通过 workflow_billfield.id 查找枚举值
            # (workflow_selectitem.FIELDID = workflow_billfield.id)
            items = select_items_cache.get(f_id, [])
            has_enum = '是' if items else '否'

            enum_values_str = ''
            if items:
                enum_count += 1
                enum_values_str = '; '.join([f'{v}={n}' for v, n in items])

            # --- 写入字段明细 ---
            ws_fields.cell(row=row_num_field, column=1, value=table_name or '').alignment = cell_alignment
            ws_fields.cell(row=row_num_field, column=2, value=detail_table or '').alignment = cell_alignment
            ws_fields.cell(row=row_num_field, column=3, value=chinese_name).alignment = cell_alignment
            ws_fields.cell(row=row_num_field, column=4, value=f_name or '').alignment = cell_alignment
            ws_fields.cell(row=row_num_field, column=5, value=f_chinese).alignment = cell_alignment
            ws_fields.cell(row=row_num_field, column=6, value=html_type_cn).alignment = cell_alignment
            ws_fields.cell(row=row_num_field, column=7, value=f_db_type or '').alignment = cell_alignment
            ws_fields.cell(row=row_num_field, column=8, value=has_enum).alignment = cell_alignment
            ws_fields.cell(row=row_num_field, column=9, value=enum_values_str).alignment = cell_alignment

            for c in range(1, 10):
                ws_fields.cell(row=row_num_field, column=c).border = thin_border

            row_num_field += 1

            # --- 写入枚举值明细 ---
            if items:
                for val, name in items:
                    ws_enum.cell(row=row_enum, column=1, value=table_name or '').alignment = cell_alignment
                    ws_enum.cell(row=row_enum, column=2, value=detail_table or '').alignment = cell_alignment
                    ws_enum.cell(row=row_enum, column=3, value=chinese_name).alignment = cell_alignment
                    ws_enum.cell(row=row_enum, column=4, value=f_name or '').alignment = cell_alignment
                    ws_enum.cell(row=row_enum, column=5, value=f_chinese).alignment = cell_alignment
                    ws_enum.cell(row=row_enum, column=6, value=str(val) if val is not None else '').alignment = cell_alignment
                    ws_enum.cell(row=row_enum, column=7, value=name).alignment = cell_alignment

                    for c in range(1, 8):
                        ws_enum.cell(row=row_enum, column=c).border = thin_border
                    row_enum += 1

        # --- 写入汇总 ---
        ws_summary.cell(row=row_num_summary, column=1, value=idx + 1).alignment = cell_alignment
        ws_summary.cell(row=row_num_summary, column=2, value=bill_id).alignment = cell_alignment
        ws_summary.cell(row=row_num_summary, column=3, value=chinese_name).alignment = cell_alignment
        ws_summary.cell(row=row_num_summary, column=4, value=table_name or '').alignment = cell_alignment
        ws_summary.cell(row=row_num_summary, column=5, value=detail_table or '').alignment = cell_alignment
        ws_summary.cell(row=row_num_summary, column=6, value=len(fields)).alignment = cell_alignment
        ws_summary.cell(row=row_num_summary, column=7, value=enum_count).alignment = cell_alignment

        for c in range(1, 8):
            ws_summary.cell(row=row_num_summary, column=c).border = thin_border

        row_num_summary += 1

        if (idx + 1) % 100 == 0:
            print(f'    处理进度: {idx + 1}/{len(bills)}')

    # ================================================================
    # 调整列宽
    # ================================================================
    ws_summary.column_dimensions['A'].width = 6
    ws_summary.column_dimensions['B'].width = 10
    ws_summary.column_dimensions['C'].width = 35
    ws_summary.column_dimensions['D'].width = 32
    ws_summary.column_dimensions['E'].width = 32
    ws_summary.column_dimensions['F'].width = 10
    ws_summary.column_dimensions['G'].width = 14

    ws_fields.column_dimensions['A'].width = 28
    ws_fields.column_dimensions['B'].width = 28
    ws_fields.column_dimensions['C'].width = 28
    ws_fields.column_dimensions['D'].width = 24
    ws_fields.column_dimensions['E'].width = 24
    ws_fields.column_dimensions['F'].width = 16
    ws_fields.column_dimensions['G'].width = 18
    ws_fields.column_dimensions['H'].width = 10
    ws_fields.column_dimensions['I'].width = 65

    ws_enum.column_dimensions['A'].width = 28
    ws_enum.column_dimensions['B'].width = 28
    ws_enum.column_dimensions['C'].width = 28
    ws_enum.column_dimensions['D'].width = 24
    ws_enum.column_dimensions['E'].width = 24
    ws_enum.column_dimensions['F'].width = 14
    ws_enum.column_dimensions['G'].width = 45

    # 冻结首行
    ws_summary.freeze_panes = 'A2'
    ws_fields.freeze_panes = 'A2'
    ws_enum.freeze_panes = 'A2'

    # 自动筛选
    ws_summary.auto_filter.ref = f'A1:G{row_num_summary - 1}'
    ws_fields.auto_filter.ref = f'A1:I{row_num_field - 1}'
    ws_enum.auto_filter.ref = f'A1:G{row_enum - 1}'

    # ================================================================
    # 保存
    # ================================================================
    output_path = os.path.join(OUTPUT_DIR, '泛微数据库字段字典.xlsx')
    wb.save(output_path)
    print(f'\n===== 完成 =====')
    print(f'输出文件: {output_path}')
    print(f'  表汇总:    {row_num_summary - 2} 个表（bill定义）')
    print(f'  字段明细:  {row_num_field - 2} 个字段')
    print(f'  枚举值:    {row_enum - 2} 条')
    print(f'  数据库表:  {row_table - 2} 个业务表')

    # 统计有枚举的字段
    fields_with_enum = sum(
        1 for items in select_items_cache.values() if items
    )
    print(f'\n  其中有枚举值的字段: {fields_with_enum} 个')

    cursor.close()
    conn.close()


if __name__ == '__main__':
    extract_all()
