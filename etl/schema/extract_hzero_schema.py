# -*- coding: utf-8 -*-
"""
读取汉得 HZERO 中台库所有数据库表的字段名称、中文名称、枚举值
输出到 Excel 文件

数据关系:
  hzero_platform.hpfm_entity_table   -> 实体/表定义 (table_name, table_comment)
  hzero_platform.hpfm_entity_column  -> 字段定义  (entity_table_id, column_name, column_comment, lov_code)
  hzero_platform.hpfm_lov            -> 枚举定义  (lov_code, lov_name)
  hzero_platform.hpfm_lov_value      -> 枚举值    (lov_id/lov_code, value, meaning)
  hzero_platform.hpfm_lov_value_tl   -> 枚举值翻译 (lov_value_id, lang, meaning)
  hzero_platform.hpfm_prompt         -> 多语言提示词 (prompt_code, lang, description)
"""
import re
import pymysql
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
import os

# ============================================================
# 配置
# ============================================================
DB_CONFIG = {
    'host': 'nj-cdb-1in2euqz.sql.tencentcdb.com',
    'port': 25257,
    'user': 'root',
    'password': 'heroesport_1',
    'charset': 'utf8',
}

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'resources', 'reference')
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 业务数据库列表
BIZ_DATABASES = [
    'hfins', 'hfins_base', 'hfins_base_account', 'hfins_bgt',
    'hfins_cpln', 'hfins_ecm', 'hfins_esr', 'hfins_fm',
    'hfins_ids', 'hfins_dynamic_report',
]


def extract_all():
    conn = pymysql.connect(**DB_CONFIG)
    cursor = conn.cursor()

    # ================================================================
    # 1. 从 hzero_platform 获取实体元数据
    # ================================================================
    print('[1/5] 加载实体元数据...')
    cursor.execute('USE hzero_platform')

    # 实体表
    cursor.execute(
        'SELECT entity_table_id, service_name, table_name, table_schema, '
        '       table_comment, multi_language_table_name '
        'FROM hpfm_entity_table ORDER BY entity_table_id'
    )
    entities = cursor.fetchall()
    print(f'    实体表: {len(entities)} 个')

    # 实体字段
    cursor.execute(
        'SELECT entity_column_id, entity_table_id, field_name, column_name, '
        '       column_comment, jdbc_type, java_type, lov_code, pk_id_flag '
        'FROM hpfm_entity_column ORDER BY entity_table_id, entity_column_id'
    )
    entity_cols = cursor.fetchall()
    print(f'    实体字段: {len(entity_cols)} 个')

    # 按 entity_table_id 分组
    cols_by_entity = defaultdict(list)
    for c in entity_cols:
        cols_by_entity[c[1]].append(c)

    # ================================================================
    # 2. 加载 LOV 枚举值（含翻译）
    # ================================================================
    print('[2/5] 加载 LOV 枚举值...')

    # LOV 主表 (lov_code -> lov_id, lov_name)
    cursor.execute('SELECT lov_id, lov_code, lov_name FROM hpfm_lov')
    lov_info = {}  # lov_code -> (lov_id, lov_name)
    for r in cursor.fetchall():
        lov_info[r[1]] = (r[0], r[2] or '')

    # LOV 中文翻译 (lov_id -> lov_name_cn)
    cursor.execute(
        "SELECT lov_id, lov_name FROM hpfm_lov_tl WHERE lang='zh_CN'"
    )
    lov_name_cn = {}
    for r in cursor.fetchall():
        lov_name_cn[r[0]] = r[1] or ''

    # LOV 值
    cursor.execute(
        'SELECT lov_value_id, lov_id, lov_code, value, meaning, order_seq '
        'FROM hpfm_lov_value ORDER BY lov_id, order_seq'
    )
    lov_values_raw = cursor.fetchall()

    # 按 lov_code 分组
    lov_values = defaultdict(list)  # lov_code -> [(value, meaning), ...]
    for r in lov_values_raw:
        lov_values[r[2]].append((r[3], r[4] or ''))

    # LOV 值中文翻译 (lov_value_id -> meaning_cn)
    cursor.execute(
        "SELECT lov_value_id, meaning FROM hpfm_lov_value_tl WHERE lang='zh_CN'"
    )
    lov_value_cn = {}
    for r in cursor.fetchall():
        lov_value_cn[r[0]] = r[1] or ''

    # 构建 lov_code -> [(value, chinese_meaning), ...]
    lov_items = defaultdict(list)  # lov_code -> [(value, chinese_name), ...]
    for r in lov_values_raw:
        vid, lid, code, val, meaning, seq = r
        cn = lov_value_cn.get(vid, meaning or '')
        lov_items[code].append((val or '', cn))

    print(f'    LOV 定义: {len(lov_info)} 个')
    print(f'    LOV 值:   {len(lov_values_raw)} 条')
    print(f'    有枚举值的 LOV: {len(lov_items)} 个')

    # ================================================================
    # 3. 扫描实际数据库表结构
    # ================================================================
    print('[3/5] 扫描数据库实际表结构...')

    # 构建实体 table_name -> entity 映射
    entity_by_table = {}  # table_name.lower() -> (entity_id, service_name, table_comment, schema)
    for e in entities:
        eid, svc, tbl, sch, cmt, ml_tbl = e
        if tbl:
            entity_by_table[tbl.lower()] = (eid, svc, cmt or '', sch or '', ml_tbl or '')

    # 扫描所有业务数据库
    all_tables = []  # (database, table_name, table_comment, column_list)

    for db_name in BIZ_DATABASES:
        try:
            cursor.execute(f'USE `{db_name}`')
            cursor.execute('SHOW TABLES')
            tables = [r[0] for r in cursor.fetchall()]

            for tbl in tables:
                # 跳过系统表
                if tbl.startswith('databasechangelog') or tbl.startswith('undo_log'):
                    continue
                if tbl.endswith('_tl') or tbl.endswith('_bak') or tbl.endswith('_backup'):
                    continue

                try:
                    cursor.execute(f'SHOW FULL COLUMNS FROM `{tbl}`')
                    columns = cursor.fetchall()
                except:
                    continue

                # 获取表注释
                cursor.execute(f"""
                    SELECT TABLE_COMMENT FROM INFORMATION_SCHEMA.TABLES
                    WHERE TABLE_SCHEMA='{db_name}' AND TABLE_NAME='{tbl}'
                """)
                row = cursor.fetchone()
                table_comment = row[0] if row else ''

                all_tables.append((db_name, tbl, table_comment, columns))

            print(f'    {db_name}: {len(tables)} 个表')
        except Exception as e:
            print(f'    {db_name}: 跳过 ({e})')

    print(f'    共扫描 {len(all_tables)} 个业务表')

    # ================================================================
    # 4. 构建提示词缓存 (prompt_code -> chinese_description)
    # ================================================================
    print('[4/5] 加载多语言提示词...')
    cursor.execute('USE hzero_platform')

    prompt_cache = {}  # prompt_code -> chinese_description
    cursor.execute(
        "SELECT prompt_code, description FROM hpfm_prompt WHERE lang='zh_CN'"
    )
    for r in cursor.fetchall():
        prompt_cache[r[0]] = r[1] or ''

    print(f'    提示词: {len(prompt_cache)} 条 (zh_CN)')

    # ================================================================
    # 5. 生成 Excel
    # ================================================================
    print('[5/5] 生成 Excel 文件...')

    wb = openpyxl.Workbook()

    # 样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    # === Sheet 1: 表汇总 ===
    ws_summary = wb.active
    ws_summary.title = '表汇总'
    write_header(ws_summary, ['序号', '数据库', '表名', '表中文名', '实体ID',
                               '微服务', '字段数', '有枚举的字段数'])

    # === Sheet 2: 字段明细 ===
    ws_fields = wb.create_sheet('字段明细')
    write_header(ws_fields, ['数据库', '表名', '表中文名', '字段名', '字段中文名',
                              '数据类型', 'Java类型', '是否主键', 'LOV编码',
                              '是否有枚举', '枚举值'])

    # === Sheet 3: 枚举值明细 ===
    ws_enum = wb.create_sheet('枚举值明细')
    write_header(ws_enum, ['数据库', '表名', '表中文名', '字段名', '字段中文名',
                            '枚举值', '枚举中文名', 'LOV编码'])

    row_sum = 2
    row_field = 2
    row_enum = 2

    for idx, (db_name, tbl, tbl_comment, columns) in enumerate(all_tables):
        tbl_lower = tbl.lower()
        entity_info = entity_by_table.get(tbl_lower, None)

        # 表中文名：优先用实体元数据
        if entity_info:
            eid, svc, e_cmt, schema, ml_tbl = entity_info
            tbl_cn = e_cmt or tbl_comment or ''
            entity_id = eid
            service_name = svc
        else:
            tbl_cn = tbl_comment or ''
            entity_id = ''
            service_name = ''

        # 获取该实体对应的字段元数据
        entity_cols_map = {}
        if entity_id and entity_id in cols_by_entity:
            for ec in cols_by_entity[entity_id]:
                col_name = (ec[3] or '').lower()
                entity_cols_map[col_name] = ec

        enum_count = 0

        for col in columns:
            col_name = col[0]
            col_type = col[1]
            col_comment = col[8] or ''

            # 从实体元数据中取更准确的信息
            ec = entity_cols_map.get(col_name.lower(), None)
            if ec:
                ec_id, ec_tbl_id, field_name, ec_col_name, ec_comment, ec_jdbc, ec_java, ec_lov, ec_pk = ec
                col_cn = ec_comment or col_comment or ''
                java_type = ec_java or ''
                jdbc_type = ec_jdbc or ''
                lov_code = ec_lov or ''
                is_pk = '是' if ec_pk == 1 else '否'
            else:
                col_cn = col_comment
                java_type = ''
                jdbc_type = col_type
                lov_code = ''
                is_pk = ''

            # 默认值
            is_nullable = '否' if col[2] == 'NO' else '是'
            col_key = col[4] or ''
            col_default = col[5] or ''

            # 尝试从列注释中提取 LOV code
            # 常见格式: "值集：XXX", "syscode：XXX", "LOV:XXX", "值集编码:XXX"
            #          "（syscode：XXX）", "值集:XXX"
            if not lov_code and col_comment:
                m = re.search(r'(?:值集(?:编码)?|syscode|LOV|lov)[：:]\s*([A-Za-z0-9._]+)', col_comment)
                if m:
                    lov_code = m.group(1)

            # 查找枚举值
            items = []
            if lov_code:
                items = lov_items.get(lov_code, [])

            has_enum = '是' if items else '否'
            if items:
                enum_count += 1
                enum_str = '; '.join([f'{v}={n}' for v, n in items])
            else:
                enum_str = ''

            # --- 写入字段明细 ---
            ws_fields.cell(row=row_field, column=1, value=db_name).alignment = cell_alignment
            ws_fields.cell(row=row_field, column=2, value=tbl).alignment = cell_alignment
            ws_fields.cell(row=row_field, column=3, value=tbl_cn).alignment = cell_alignment
            ws_fields.cell(row=row_field, column=4, value=col_name).alignment = cell_alignment
            ws_fields.cell(row=row_field, column=5, value=col_cn).alignment = cell_alignment
            ws_fields.cell(row=row_field, column=6, value=jdbc_type or col_type).alignment = cell_alignment
            ws_fields.cell(row=row_field, column=7, value=java_type).alignment = cell_alignment
            ws_fields.cell(row=row_field, column=8, value=is_pk).alignment = cell_alignment
            ws_fields.cell(row=row_field, column=9, value=lov_code).alignment = cell_alignment
            ws_fields.cell(row=row_field, column=10, value=has_enum).alignment = cell_alignment
            ws_fields.cell(row=row_field, column=11, value=enum_str).alignment = cell_alignment
            for c in range(1, 12):
                ws_fields.cell(row=row_field, column=c).border = thin_border
            row_field += 1

            # --- 写入枚举值明细 ---
            if items:
                for val, name in items:
                    ws_enum.cell(row=row_enum, column=1, value=db_name).alignment = cell_alignment
                    ws_enum.cell(row=row_enum, column=2, value=tbl).alignment = cell_alignment
                    ws_enum.cell(row=row_enum, column=3, value=tbl_cn).alignment = cell_alignment
                    ws_enum.cell(row=row_enum, column=4, value=col_name).alignment = cell_alignment
                    ws_enum.cell(row=row_enum, column=5, value=col_cn).alignment = cell_alignment
                    ws_enum.cell(row=row_enum, column=6, value=val).alignment = cell_alignment
                    ws_enum.cell(row=row_enum, column=7, value=name).alignment = cell_alignment
                    ws_enum.cell(row=row_enum, column=8, value=lov_code).alignment = cell_alignment
                    for c in range(1, 9):
                        ws_enum.cell(row=row_enum, column=c).border = thin_border
                    row_enum += 1

        # --- 写入汇总 ---
        ws_summary.cell(row=row_sum, column=1, value=idx + 1).alignment = cell_alignment
        ws_summary.cell(row=row_sum, column=2, value=db_name).alignment = cell_alignment
        ws_summary.cell(row=row_sum, column=3, value=tbl).alignment = cell_alignment
        ws_summary.cell(row=row_sum, column=4, value=tbl_cn).alignment = cell_alignment
        ws_summary.cell(row=row_sum, column=5, value=entity_id).alignment = cell_alignment
        ws_summary.cell(row=row_sum, column=6, value=service_name).alignment = cell_alignment
        ws_summary.cell(row=row_sum, column=7, value=len(columns)).alignment = cell_alignment
        ws_summary.cell(row=row_sum, column=8, value=enum_count).alignment = cell_alignment
        for c in range(1, 9):
            ws_summary.cell(row=row_sum, column=c).border = thin_border
        row_sum += 1

        if (idx + 1) % 200 == 0:
            print(f'    处理进度: {idx + 1}/{len(all_tables)}')

    # ================================================================
    # === Sheet 4: LOV 值集汇总 ===
    # ================================================================
    ws_lov = wb.create_sheet('LOV值集汇总')
    write_header(ws_lov, ['LOV编码', 'LOV中文名', '值', '中文含义'])
    row_lov = 2
    for code in sorted(lov_items.keys()):
        lov_name = ''
        if code in lov_info:
            lid, ln = lov_info[code]
            lov_name = lov_name_cn.get(lid, ln)
        items = lov_items[code]
        for val, meaning in items:
            ws_lov.cell(row=row_lov, column=1, value=code).alignment = cell_alignment
            ws_lov.cell(row=row_lov, column=2, value=lov_name).alignment = cell_alignment
            ws_lov.cell(row=row_lov, column=3, value=val).alignment = cell_alignment
            ws_lov.cell(row=row_lov, column=4, value=meaning).alignment = cell_alignment
            for c in range(1, 5):
                ws_lov.cell(row=row_lov, column=c).border = thin_border
            row_lov += 1

    # ================================================================
    # 调整列宽
    # ================================================================
    ws_summary.column_dimensions['A'].width = 6
    ws_summary.column_dimensions['B'].width = 22
    ws_summary.column_dimensions['C'].width = 35
    ws_summary.column_dimensions['D'].width = 30
    ws_summary.column_dimensions['E'].width = 8
    ws_summary.column_dimensions['F'].width = 16
    ws_summary.column_dimensions['G'].width = 8
    ws_summary.column_dimensions['H'].width = 14

    ws_fields.column_dimensions['A'].width = 22
    ws_fields.column_dimensions['B'].width = 35
    ws_fields.column_dimensions['C'].width = 28
    ws_fields.column_dimensions['D'].width = 22
    ws_fields.column_dimensions['E'].width = 28
    ws_fields.column_dimensions['F'].width = 16
    ws_fields.column_dimensions['G'].width = 16
    ws_fields.column_dimensions['H'].width = 8
    ws_fields.column_dimensions['I'].width = 22
    ws_fields.column_dimensions['J'].width = 10
    ws_fields.column_dimensions['K'].width = 60

    ws_enum.column_dimensions['A'].width = 22
    ws_enum.column_dimensions['B'].width = 35
    ws_enum.column_dimensions['C'].width = 28
    ws_enum.column_dimensions['D'].width = 22
    ws_enum.column_dimensions['E'].width = 28
    ws_enum.column_dimensions['F'].width = 14
    ws_enum.column_dimensions['G'].width = 30
    ws_enum.column_dimensions['H'].width = 22

    ws_lov.column_dimensions['A'].width = 30
    ws_lov.column_dimensions['B'].width = 25
    ws_lov.column_dimensions['C'].width = 14
    ws_lov.column_dimensions['D'].width = 35

    # 冻结首行
    for ws in [ws_summary, ws_fields, ws_enum, ws_lov]:
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

    # 保存
    output_path = os.path.join(OUTPUT_DIR, '汉得数据库字段字典.xlsx')
    wb.save(output_path)
    print(f'\n===== 完成 =====')
    print(f'输出文件: {output_path}')
    print(f'  表汇总:      {row_sum - 2} 个表')
    print(f'  字段明细:    {row_field - 2} 条')
    print(f'  枚举值明细:  {row_enum - 2} 条')
    print(f'  LOV值集:     {row_lov - 2} 条')
    print(f'  有实体的表:  {sum(1 for e in entity_by_table.values() if e)} 个')

    cursor.close()
    conn.close()


if __name__ == '__main__':
    extract_all()
