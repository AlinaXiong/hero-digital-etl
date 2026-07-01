# -*- coding: utf-8 -*-
"""应收期初 —— 应收报账单(DB 直连版)。

处理流程:
1. 校验泛微字段字典,避免 SQL 字段名/含义写错。
2. 用固定 SQL 从泛微开票表 uf_xtyykp 取数,并左关联收款登记 uf_skdj 的汇总金额。
3. 只对必须跨表/跨系统的 ID 做批量解析,例如申请人、部门、公司主体、客户、合同、币种。
4. 按导入模版字段逐列生成输出,字段旁标注取值来源。

跑法:在项目根执行  python run.py ar_invoice_opening_db
"""
import os
import sys
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

if __package__ is None or __package__ == '':
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from etl.util import common as c
from etl.process.ap_prepayment_opening_db import (
    build_fw_project_code_map_for_ids,
    project_filter_codes,
    EVENT_PROJECT_SHEET,
    MCN_PROJECT_SHEET,
    MCN_PROJECT_TABLES,
)

# ============================ 文件 / 模板 ============================
TASK_NAME = 'ar_invoice_opening_db'
TEMPLATE_DIR = c.TPL_DIR / 'ar_invoice_opening'
OUTPUT_DIR = c.OUT_DIR / TASK_NAME
DATE_SUFFIX = c.today_suffix()

TEMPLATE_FILE = TEMPLATE_DIR / '应收报账单期初数据导入模板.xlsx'
OUTPUT_FILE = OUTPUT_DIR / f'英雄应收报账单期初数据导入_应收期初_{DATE_SUFFIX}.xlsx'
EXCEPTION_FILE = OUTPUT_DIR / f'未匹配清单_应收期初_{DATE_SUFFIX}.xlsx'
ISSUE_FIX_ENV = 'AR_INVOICE_ISSUE_FIX_XLSX'
DEFAULT_ISSUE_FIX_FILE = c.RULES_DIR / '问题处理-应收.xlsx'

TEMPLATE_SHEET = '应收报账单期初数据导入'
MCN_TEMPLATE_SHEET = '应收报账单期初数据导入-MCN'
RULE_SHEET = '应收期初'
RULE_TABLE = '应收报账单期初数据导入'
DOCUMENT_TYPE = 'OPEN10'
MANAGEMENT_COMPANY = 'Hero'
PAYER_TYPE = '客户'
INCOME_ITEM = '项目收款'
BUSINESS_TYPE_LOV = 'HERO.BUSINESS_TYPE'
INVOICE_TYPE_LOV = 'HERO.INVOICE_TYPE'
CONTRACT_INVOICE_MEANING = '合同'
PUBLIC_INVOICE_MEANING = '对公开票'
PLATFORM_PRE_INVOICE_MEANING = '平台预开票'
OUTPUT_COLUMNS = [
    '当前节点',
    '当前状况',
    '来源单据号',
    '应收报账单类型',
    '核算主体',
    '管理公司',
    '部门',
    '岗位',
    '申请人',
    '申请日期',
    '支付币种',
    '付款对象类型',
    '付款对象',
    '合同编号',
    '里程碑阶段',
    '平台',
    '业务类型编码',
    '开票类型编码',
    '核销金额',
    '头备注',
    '发票号',
    '自审批',
    '自审核',
    '凭证推送',
    '凭证日期',
    '行号',
    '收入分类',
    '收入项目',
    '数量',
    '单价',
    '金额',
    '税率类型',
    '税额',
    '行备注',
    *[f'头维度{i}' for i in range(1, 21)],
    '项目',
    '订单',
    *[f'行维度{i}' for i in range(3, 21)],
    '泛微项目编号',
]
MCN_OUTPUT_COLUMNS = [
    *OUTPUT_COLUMNS,
    '原泛微订单号',
]

# 问题清单里,目标字段缺失时带出的泛微源字段。
ISSUE_SOURCE_FIELD_MAP = {
    '来源单据号': '流程编号',
    '核算主体': '公司主体',
    '申请人': '申请人ID',
    '支付币种': '开票币种',
    '付款对象': '客户',
    '合同编号': '开票合同ID',
    '发票号': '发票号',
    '业务类型编码': '业务类型',
    '核销金额': '收款登记已收款金额',
    '金额': '开票金额（含税价）',
    '税率类型': '税率',
    '项目': '项目编号',
    '订单': '项目编号',
}

FW_INVOICE_TABLE = 'uf_xtyykp'
FW_RECEIPT_TABLE = 'uf_skdj'
INVOICE_NUMBER_SEPARATOR_RE = re.compile(r'[,，、;；|\s]+')
ORDER_FIX_SHEET = '订单映射_多候选'
TAX_FIX_SHEET = '缺失_税率类型'
_ISSUE_ORDER_FIX_CACHE = None
_ISSUE_TAX_FIX_CACHE = None


# ============================ 枚举 / 过滤口径 ============================
# 泛微 uf_xtyykp.kpzt: 开票状态。来源于 workflow_selectitem。
INVOICE_STATUS_MEANINGS = {
    0: '已开票',
    1: '已部分开票',
    2: '已红冲/废票',
    3: '开票失败',
}
VOID_FLAG_MEANINGS = {
    0: '是',
    1: '否',
}
# 泛微 uf_xtyykp.ywlx: 业务类型。0/1 来自 workflow_selectitem;2 在历史数据里表示空业务类型。
BUSINESS_TYPE_CODE_MEANINGS = {
    0: '外部公司',
    1: '外部个人',
    2: '',
}
BUSINESS_TYPE_MEANING = {
    '外部公司': '对公开票',
    '外部个人': '个人开票',
    '': '虚拟开票',
}

TAX_PREFERRED_DESCRIPTIONS = {
    0.00: ['0%销项税，中国', '0%税率', '0%'],
    0.01: ['1%税率(价外)', '1%'],
    0.03: ['3%税率(价外)', '3%'],
    0.06: ['6%销项税，中国', '6%税率', '6%'],
    0.09: ['9%税率(价内)', '9%销项税，中国', '9%'],
    0.13: ['13%税率(价外)', '13%销项税，中国', '13%'],
}

MCN_TAX_PREFERRED_DESCRIPTIONS = {
    0.00: ['0%销项税，中国', '0%税率', '0%'],
    0.01: ['1%税率(价外)', '1%'],
    0.03: ['3%税率(价外)', '3%'],
    0.06: ['6%销项税，中国', '6%税率', '6%'],
    0.09: ['9%销项税，中国', '9%税率(价内)', '9%'],
    0.13: ['13%销项税，中国', '13%税率(价外)', '13%'],
}

VOID_CODE = 0
COMPLETE_NODE_TYPE = 3

MCN_INVOICE_TABLE = 'formtable_main_28'
MCN_INVOICE_DETAIL_TABLE = 'formtable_main_28_dt1'
MCN_ORDER_INVOICE_TABLE = 'formtable_main_72'
MCN_ORDER_WORKFLOW_IDS = (73, 89, 336)
MCN_INVOICE_WORKFLOW_IDS = (50, 335)


# ============================ 泛微源 SQL ============================
# 只查开票表,并用子查询按「开票/预收单号」汇总收款登记金额。其他字典/维表解析放到后续批量查询中做。
# 字段含义由 EXPECTED_*_FIELDS + common.validate_fw_fields 在运行时校验。
SOURCE_SQL = """
SELECT
    m.id AS `ID`,
    rb.requestid AS `requestid`,
    m.lcbh AS `流程编号`,
    m.sqr AS `申请人ID`,
    m.sqrbm AS `申请人部门ID`,
    m.sqrq AS `申请日期`,
    m.kpht AS `开票合同ID`,
    m.xmbh AS `项目编号ID`,
    m.xmmc AS `项目名称`,
    m.gszt AS `公司主体ID`,
    m.cbzx AS `成本中心ID`,
    m.kh AS `客户ID`,
    m.kpjehsj AS `开票金额（含税价）`,
    m.sl AS `税率`,
    m.se AS `税额`,
    m.kpbz AS `开票币种ID`,
    m.kptxt AS `开票备注`,
    m.fphm AS `发票号`,
    m.ywlx AS `业务类型ID`,
    m.xmje AS `不含税金额（明细）`,
    m.xmjshj AS `价税合计（明细）`,
    m.semx AS `税额（明细）`,
    m.slmx AS `税率（明细）`,
    COALESCE(r.receipt_amount, 0) AS `收款登记已收款金额`
FROM uf_xtyykp m
LEFT JOIN workflow_requestbase rb ON rb.requestmark = m.lcbh
LEFT JOIN (
    SELECT
        kpysdh,
        SUM(COALESCE(bfqrjehj, 0)) AS receipt_amount
    FROM uf_skdj
    WHERE kpysdh IS NOT NULL AND TRIM(kpysdh) <> ''
    GROUP BY kpysdh
) r ON r.kpysdh = m.lcbh
WHERE (m.sfzf IS NULL OR m.sfzf <> %(void_code)s)
ORDER BY m.id
"""


# 仅用于打印过滤前后数量,便于核对口径。
STATS_SQL = """
SELECT
    COUNT(*) AS total_count,
    SUM(CASE WHEN m.sfzf = %(void_code)s THEN 1 ELSE 0 END) AS void_count,
    SUM(CASE
        WHEN (m.sfzf IS NULL OR m.sfzf <> %(void_code)s)
        THEN 1 ELSE 0 END) AS kept_count
FROM uf_xtyykp m
"""


# 运行前校验字段真实含义。主表字段 detail_table 用空字符串。
EXPECTED_INVOICE_FIELDS = {
    '': {
        'lcbh': '流程编号',
        'sqr': '申请人',
        'sqrbm': '申请人部门',
        'sqrq': '申请日期',
        'kpht': '开票合同',
        'xmbh': '项目编号',
        'xmmc': '项目名称',
        'gszt': '公司主体',
        'cbzx': '成本中心',
        'kh': '客户',
        'kpjehsj': '开票金额（含税价）',
        'sl': '税率',
        'se': '税额',
        'kpzt': '开票状态',
        'kpbz': '开票币种',
        'sfzf': '是否作废',
        'kptxt': '开票备注',
        'fphm': '发票号码',
        'ywlx': '业务类型',
        'xmje': '不含税金额（明细）',
        'xmjshj': '价税合计（明细）',
        'semx': '税额（明细）',
        'slmx': '税率（明细）',
    },
}
EXPECTED_RECEIPT_FIELDS = {
    '': {
        'kpysdh': '开票/预收单号',
        'bfqrjehj': '已收款金额',
    },
}

EXPECTED_MCN_INVOICE_FIELDS = {
    '': {
        'kpdh': '开票单号',
        'sqr': '申请人',
        'sqrbm': '申请人部门',
        'sqrq': '申请日期',
        'kptt': '公司主体',
        'kplx': '开票类型',
        'kh': '客户',
        'kpht': '开票合同',
        'pt': '平台',
    },
    MCN_INVOICE_DETAIL_TABLE: {
        'xmbh': '项目编号',
        'kpje': '开票金额',
        'sl': '税率',
        'se': '税额',
        'bhsje': '不含税金额',
    },
}

EXPECTED_MCN_ORDER_INVOICE_FIELDS = {
    '': {
        'kpdh': '开票单号',
        'sqr': '申请人',
        'sqrbm': '申请人部门',
        'sqrq': '申请日期',
        'gszt': '公司主体',
        'kplx': '开票类型',
        'kh': '客户',
        'kpht': '开票合同',
    },
    'formtable_main_72_dt4': {
        'ddbh': '订单编号',
        'szxm': '所属项目',
        'ddje': '开票金额',
        'pt': '平台',
        'sl': '税率',
        'se': '税额',
        'bhsje': '不含税金额',
    },
    'formtable_main_72_dt5': {
        'ptpqh': '订单号',
        'szxm': '所属项目',
        'ddje': '开票金额',
        'pt': '平台',
        'sl': '税率',
        'se': '税额',
        'bhsje': '不含税金额',
    },
    'formtable_main_72_dt6': {
        'ptpqh': '订单号',
        'szxm': '所属项目',
        'ddje': '开票金额',
        'pt': '平台',
        'sl': '税率',
        'se': '税额',
        'bhsje': '不含税金额',
    },
    'formtable_main_72_dt7': {
        'ddh': '订单号',
        'szxm': '所属项目',
        'dkje': '开票金额',
        'pt': '平台',
        'sl': '税率',
        'se': '税额',
        'bhsje': '不含税金额',
    },
}


MCN_SOURCE_SQL = """
SELECT
    '开票申请流程' AS `来源流程`,
    'formtable_main_28_dt1' AS `明细来源`,
    m.id AS `ID`,
    d.id AS `明细ID`,
    m.requestid AS `requestid`,
    rb.workflowid AS `workflowid`,
    rb.REQUESTNAME AS `标题`,
    m.kpdh AS `开票单号`,
    NULL AS `泛微订单编号`,
    NULL AS `泛微订单编号ID`,
    m.sqr AS `申请人ID`,
    m.sqrbm AS `申请人部门ID`,
    m.sqrq AS `申请日期`,
    m.kptt AS `公司主体ID`,
    m.kh AS `客户ID`,
    NULL AS `明细客户ID`,
    m.kpht AS `开票合同ID`,
    m.kplx AS `开票类型ID`,
    m.pt AS `平台ID`,
    d.xmbh AS `项目编号ID`,
    d.xmmc AS `项目名称`,
    d.kpje AS `开票金额`,
    d.sl AS `税率`,
    d.se AS `税额`,
    d.bhsje AS `不含税金额`,
    d.cbzx AS `成本中心ID`,
    m.bz AS `开票备注`
FROM formtable_main_28 m
JOIN workflow_requestbase rb ON rb.requestid = m.requestid
JOIN formtable_main_28_dt1 d ON d.mainid = m.id
WHERE rb.currentnodetype = %(complete_node_type)s
  AND rb.workflowid IN %(mcn_invoice_workflow_ids)s

UNION ALL

SELECT
    '开票申请流程（订单）' AS `来源流程`,
    'formtable_main_72_dt4' AS `明细来源`,
    m.id AS `ID`,
    d.id AS `明细ID`,
    m.requestid AS `requestid`,
    rb.workflowid AS `workflowid`,
    rb.REQUESTNAME AS `标题`,
    m.kpdh AS `开票单号`,
    COALESCE(NULLIF(od.ddbh, ''), d.ddbh) AS `泛微订单编号`,
    d.ddbh AS `泛微订单编号ID`,
    m.sqr AS `申请人ID`,
    m.sqrbm AS `申请人部门ID`,
    m.sqrq AS `申请日期`,
    m.gszt AS `公司主体ID`,
    m.kh AS `客户ID`,
    d.khmc AS `明细客户ID`,
    m.kpht AS `开票合同ID`,
    m.kplx AS `开票类型ID`,
    d.pt AS `平台ID`,
    d.szxm AS `项目编号ID`,
    d.xmmc AS `项目名称`,
    d.ddje AS `开票金额`,
    d.sl AS `税率`,
    d.se AS `税额`,
    d.bhsje AS `不含税金额`,
    m.cbzx AS `成本中心ID`,
    m.bz AS `开票备注`
FROM formtable_main_72 m
JOIN workflow_requestbase rb ON rb.requestid = m.requestid
JOIN formtable_main_72_dt4 d ON d.mainid = m.id
LEFT JOIN uf_ddk od ON od.id = d.ddbh
WHERE rb.currentnodetype = %(complete_node_type)s
  AND rb.workflowid IN %(mcn_order_workflow_ids)s

UNION ALL

SELECT
    '开票申请流程（订单）' AS `来源流程`,
    'formtable_main_72_dt5' AS `明细来源`,
    m.id AS `ID`,
    d.id AS `明细ID`,
    m.requestid AS `requestid`,
    rb.workflowid AS `workflowid`,
    rb.REQUESTNAME AS `标题`,
    m.kpdh AS `开票单号`,
    COALESCE(NULLIF(od.ddbh, ''), d.ptpqh) AS `泛微订单编号`,
    d.ptpqh AS `泛微订单编号ID`,
    m.sqr AS `申请人ID`,
    m.sqrbm AS `申请人部门ID`,
    m.sqrq AS `申请日期`,
    m.gszt AS `公司主体ID`,
    m.kh AS `客户ID`,
    d.khmc AS `明细客户ID`,
    m.kpht AS `开票合同ID`,
    m.kplx AS `开票类型ID`,
    d.pt AS `平台ID`,
    d.szxm AS `项目编号ID`,
    d.xmmc AS `项目名称`,
    d.ddje AS `开票金额`,
    d.sl AS `税率`,
    d.se AS `税额`,
    d.bhsje AS `不含税金额`,
    m.cbzx AS `成本中心ID`,
    m.bz AS `开票备注`
FROM formtable_main_72 m
JOIN workflow_requestbase rb ON rb.requestid = m.requestid
JOIN formtable_main_72_dt5 d ON d.mainid = m.id
LEFT JOIN formtable_main_76_dt2 od ON od.id = d.ptpqh
WHERE rb.currentnodetype = %(complete_node_type)s
  AND rb.workflowid IN %(mcn_order_workflow_ids)s

UNION ALL

SELECT
    '开票申请流程（订单）' AS `来源流程`,
    'formtable_main_72_dt6' AS `明细来源`,
    m.id AS `ID`,
    d.id AS `明细ID`,
    m.requestid AS `requestid`,
    rb.workflowid AS `workflowid`,
    rb.REQUESTNAME AS `标题`,
    m.kpdh AS `开票单号`,
    COALESCE(NULLIF(od.ddbh, ''), d.ptpqh) AS `泛微订单编号`,
    d.ptpqh AS `泛微订单编号ID`,
    m.sqr AS `申请人ID`,
    m.sqrbm AS `申请人部门ID`,
    m.sqrq AS `申请日期`,
    m.gszt AS `公司主体ID`,
    m.kh AS `客户ID`,
    d.khmc AS `明细客户ID`,
    m.kpht AS `开票合同ID`,
    m.kplx AS `开票类型ID`,
    d.pt AS `平台ID`,
    d.szxm AS `项目编号ID`,
    d.xmmc AS `项目名称`,
    d.ddje AS `开票金额`,
    d.sl AS `税率`,
    d.se AS `税额`,
    d.bhsje AS `不含税金额`,
    m.cbzx AS `成本中心ID`,
    m.bz AS `开票备注`
FROM formtable_main_72 m
JOIN workflow_requestbase rb ON rb.requestid = m.requestid
JOIN formtable_main_72_dt6 d ON d.mainid = m.id
LEFT JOIN formtable_main_79_dt2 od ON od.id = d.ptpqh
WHERE rb.currentnodetype = %(complete_node_type)s
  AND rb.workflowid IN %(mcn_order_workflow_ids)s

UNION ALL

SELECT
    '开票申请流程（订单）' AS `来源流程`,
    'formtable_main_72_dt7' AS `明细来源`,
    m.id AS `ID`,
    d.id AS `明细ID`,
    m.requestid AS `requestid`,
    rb.workflowid AS `workflowid`,
    rb.REQUESTNAME AS `标题`,
    m.kpdh AS `开票单号`,
    COALESCE(NULLIF(od.ddbh, ''), d.ddh) AS `泛微订单编号`,
    d.ddh AS `泛微订单编号ID`,
    m.sqr AS `申请人ID`,
    m.sqrbm AS `申请人部门ID`,
    m.sqrq AS `申请日期`,
    m.gszt AS `公司主体ID`,
    m.kh AS `客户ID`,
    d.khmc AS `明细客户ID`,
    m.kpht AS `开票合同ID`,
    m.kplx AS `开票类型ID`,
    d.pt AS `平台ID`,
    d.szxm AS `项目编号ID`,
    d.xmmc AS `项目名称`,
    d.dkje AS `开票金额`,
    d.sl AS `税率`,
    d.se AS `税额`,
    d.bhsje AS `不含税金额`,
    m.cbzx AS `成本中心ID`,
    m.bz AS `开票备注`
FROM formtable_main_72 m
JOIN workflow_requestbase rb ON rb.requestid = m.requestid
JOIN formtable_main_72_dt7 d ON d.mainid = m.id
LEFT JOIN formtable_main_76_dt2 od ON od.id = d.ddh
WHERE rb.currentnodetype = %(complete_node_type)s
  AND rb.workflowid IN %(mcn_order_workflow_ids)s
ORDER BY `开票单号`, `ID`, `明细来源`, `明细ID`
"""

MCN_RECEIPT_SQL = """
SELECT
    kpysdh AS `开票单号`,
    SUM(COALESCE(bfqrjehj, 0)) AS `收款登记已收款金额`
FROM uf_skdj
WHERE kpysdh IS NOT NULL
  AND TRIM(kpysdh) <> ''
GROUP BY kpysdh
"""

WORKFLOW_STATE_SQL = """
SELECT
    rb.requestid AS `requestid`,
    nb.nodename AS `当前节点`,
    rb.status AS `当前状况`
FROM workflow_requestbase rb
LEFT JOIN (
    SELECT requestid, MAX(nownodeid) AS nownodeid
    FROM workflow_nownode
    WHERE requestid IN %(requestids)s
    GROUP BY requestid
) nn ON nn.requestid = rb.requestid
LEFT JOIN workflow_nodebase nb
    ON nb.id = COALESCE(nn.nownodeid, rb.currentnodeid, rb.lastnodeid)
WHERE rb.requestid IN %(requestids)s
"""


# ============================ DB 查询小工具 ============================
def _query_fw(sql):
    """查询泛微库,统一带上本任务过滤参数。"""
    return c.query_db('FW', 'vspn_xtyy', sql, {
        'void_code': VOID_CODE,
    })


# ============================ 源值解析 ============================
def _text(value):
    if pd.isna(value):
        return ''
    text = str(value).strip()
    return '' if text in ('', 'nan', 'None', 'NaT') else text


def _first_browser_id(value):
    ids = c.parse_browser_ids(value)
    return ids[0] if ids else ''


def _lookup_first_browser_value(mapping, value):
    for item_id in c.parse_browser_ids(value):
        mapped = mapping.get(item_id, '')
        if mapped:
            return mapped
    return ''


def _chunks(values, size=1000):
    values = list(values)
    for start in range(0, len(values), size):
        yield values[start:start + size]


def _looks_like_browser_id_list(value):
    text = _text(value)
    if not text:
        return False
    ids = c.parse_browser_ids(text)
    return bool(ids) and all(item_id.isdigit() for item_id in ids)


def _build_mcn_order_code_map(table_name, order_ids):
    order_ids = c.clean_codes(
        item_id
        for value in order_ids
        for item_id in c.parse_browser_ids(value)
        if item_id.isdigit()
    )
    if not order_ids:
        return {}

    result = {}
    for batch in _chunks(order_ids):
        order_df = c.query_db(
            'FW',
            'vspn_xtyy',
            f'SELECT id, ddbh AS order_code FROM {table_name} '
            f'WHERE id IN ({c.in_placeholders(batch)})',
            batch,
        )
        for _, row in order_df.iterrows():
            order_id = c.format_code(row['id'])
            order_code = _text(row['order_code'])
            if order_id and order_code:
                result[order_id] = order_code
    return result


def _lookup_mcn_order_codes(order_map, value):
    order_codes = [
        _text(order_map.get(item_id, ''))
        for item_id in c.parse_browser_ids(value)
    ]
    return ','.join(order_code for order_code in order_codes if order_code)


def _resolve_mcn_order_numbers(df):
    """MCN 订单流程明细里的订单字段存 ID,输出列必须展示订单编号。"""
    if df.empty or '泛微订单编号' not in df.columns:
        return pd.Series('', index=df.index)

    display_values = df['泛微订单编号'].map(_text)
    raw_ids = df.get('泛微订单编号ID', pd.Series('', index=df.index)).map(_text)
    detail_sources = df.get('明细来源', pd.Series('', index=df.index)).map(_text)
    resolved_values = pd.Series('', index=df.index)

    source_table_map = {
        'formtable_main_72_dt4': 'uf_ddk',
        'formtable_main_72_dt5': 'formtable_main_76_dt2',
        'formtable_main_72_dt6': 'formtable_main_79_dt2',
        'formtable_main_72_dt7': 'formtable_main_76_dt2',
    }
    for detail_source, table_name in source_table_map.items():
        mask = detail_sources == detail_source
        if not mask.any():
            continue
        order_map = _build_mcn_order_code_map(table_name, raw_ids.loc[mask])
        if order_map:
            resolved_values.loc[mask] = raw_ids.loc[mask].map(
                lambda value: _lookup_mcn_order_codes(order_map, value))

    return pd.Series([
        resolved if resolved else ('' if _looks_like_browser_id_list(display) else display)
        for display, resolved in zip(display_values, resolved_values)
    ], index=df.index)


def attach_workflow_states(source_df):
    """按 requestid 补当前节点和当前状况，当前节点取 MAX(nownodeid) 对应名称。"""
    df = source_df.copy()
    if df.empty or 'requestid' not in df.columns:
        df['当前节点'] = ''
        df['当前状况'] = ''
        return df

    request_ids = sorted({
        c.format_code(value)
        for value in df['requestid']
        if c.format_code(value)
    })
    if not request_ids:
        df['当前节点'] = ''
        df['当前状况'] = ''
        return df

    state_frames = []
    for chunk in _chunks(request_ids):
        state_frames.append(c.query_db(
            'FW',
            'vspn_xtyy',
            WORKFLOW_STATE_SQL,
            {'requestids': tuple(chunk)},
        ))
    state_df = pd.concat(state_frames, ignore_index=True) if state_frames else pd.DataFrame()
    if state_df.empty:
        df['当前节点'] = ''
        df['当前状况'] = ''
        return df

    state_df['requestid_key'] = state_df['requestid'].map(c.format_code)
    state_df = state_df.drop_duplicates('requestid_key', keep='last')
    node_map = state_df.set_index('requestid_key')['当前节点'].map(c.clean_fw_select_name).to_dict()
    status_map = state_df.set_index('requestid_key')['当前状况'].map(_text).to_dict()
    request_keys = df['requestid'].map(c.format_code)
    df['当前节点'] = request_keys.map(node_map).fillna('')
    df['当前状况'] = request_keys.map(status_map).fillna('')
    return df


def _issue_fix_file():
    configured = os.getenv(ISSUE_FIX_ENV, '').strip()
    path = Path(configured) if configured else DEFAULT_ISSUE_FIX_FILE
    if path.exists():
        return path
    print(f'[应收期初-问题处理] 未找到问题处理文件: {path}')
    return None


def _read_issue_fix_sheet(sheet_name, required_columns):
    path = _issue_fix_file()
    if path is None:
        return pd.DataFrame(columns=list(required_columns))
    try:
        sheet_df = pd.read_excel(
            path,
            sheet_name=sheet_name,
            dtype=str,
            keep_default_na=False,
            engine='openpyxl',
        )
    except ValueError:
        print(f'[应收期初-问题处理] {path} 不存在 sheet: {sheet_name}')
        return pd.DataFrame(columns=list(required_columns))
    missing = set(required_columns) - set(sheet_df.columns)
    if missing:
        raise ValueError(f'{path.name}/{sheet_name} 缺少列: {sorted(missing)}')
    return sheet_df


def _load_issue_order_fix():
    """读取《问题处理-应收》订单多候选人工指定结果。"""
    global _ISSUE_ORDER_FIX_CACHE
    if _ISSUE_ORDER_FIX_CACHE is not None:
        return _ISSUE_ORDER_FIX_CACHE

    columns = ['来源单据编号', '泛微项目编号', '指定订单编号']
    sheet_df = _read_issue_fix_sheet(ORDER_FIX_SHEET, columns)
    rows = []
    for _, row in sheet_df.iterrows():
        doc_no = _text(row.get('来源单据编号'))
        project_code = _text(row.get('泛微项目编号'))
        order_code = _text(row.get('指定订单编号'))
        if doc_no and project_code and order_code:
            rows.append({
                '来源单据编号': doc_no,
                '泛微项目编号': project_code,
                '指定订单编号': order_code,
            })
    if rows:
        fix_df = pd.DataFrame(rows, columns=columns).drop_duplicates()
        conflict = (
            fix_df.groupby(['来源单据编号', '泛微项目编号'])['指定订单编号']
            .nunique()
            .reset_index(name='指定订单数')
        )
        conflict = conflict[conflict['指定订单数'] > 1]
        if len(conflict) > 0:
            raise ValueError(f'{ORDER_FIX_SHEET} 存在同一单据+项目指定多个订单: {conflict.head(10).to_dict("records")}')
        fix_df = fix_df.drop_duplicates(['来源单据编号', '泛微项目编号'], keep='first')
    else:
        fix_df = pd.DataFrame(columns=columns)

    print(f'[应收期初-问题处理] {ORDER_FIX_SHEET} 指定订单记录数: {len(fix_df)}')
    _ISSUE_ORDER_FIX_CACHE = fix_df
    return _ISSUE_ORDER_FIX_CACHE


def _load_issue_tax_fix():
    """读取《问题处理-应收》税率类型人工指定结果。"""
    global _ISSUE_TAX_FIX_CACHE
    if _ISSUE_TAX_FIX_CACHE is not None:
        return _ISSUE_TAX_FIX_CACHE

    columns = ['来源单据号', '指定税率类型']
    sheet_df = _read_issue_fix_sheet(TAX_FIX_SHEET, columns)
    rows = []
    for _, row in sheet_df.iterrows():
        doc_no = _text(row.get('来源单据号'))
        tax_type = _text(row.get('指定税率类型'))
        if doc_no and tax_type:
            rows.append({'来源单据号': doc_no, '指定税率类型': tax_type})
    if rows:
        fix_df = pd.DataFrame(rows, columns=columns).drop_duplicates()
        conflict = (
            fix_df.groupby('来源单据号')['指定税率类型']
            .nunique()
            .reset_index(name='指定税率类型数')
        )
        conflict = conflict[conflict['指定税率类型数'] > 1]
        if len(conflict) > 0:
            raise ValueError(f'{TAX_FIX_SHEET} 存在同一单据指定多个税率类型: {conflict.head(10).to_dict("records")}')
        fix_df = fix_df.drop_duplicates('来源单据号', keep='first')
    else:
        fix_df = pd.DataFrame(columns=columns)

    print(f'[应收期初-问题处理] {TAX_FIX_SHEET} 指定税率类型记录数: {len(fix_df)}')
    _ISSUE_TAX_FIX_CACHE = fix_df
    return _ISSUE_TAX_FIX_CACHE


def _invoice_numbers_text(value):
    """发票号码可能录成逗号、顿号、分号或换行分隔,输出统一用英文逗号。"""
    text = _text(value)
    if not text:
        return ''
    invoice_numbers = [part.strip() for part in INVOICE_NUMBER_SEPARATOR_RE.split(text) if part.strip()]
    return ','.join(invoice_numbers)


def resolve_business_type_name(value):
    """uf_xtyykp.ywlx 业务类型:0=外部公司,1=外部个人,2=空业务类型。"""
    code = c.format_code(value)
    return BUSINESS_TYPE_CODE_MEANINGS.get(int(code), '') if code.isdigit() else ''


def resolve_source_values(source_df):
    """基于主 SQL 返回的泛微 ID 字段补充输出需要的展示值。

    保留原始 ID 列,新增展示列:
    - 申请人ID -> 申请人 / 申请人工号
    - 申请人部门ID -> 申请人部门
    - 公司主体ID -> 公司主体
    - 客户ID -> 客户
    - 开票合同ID -> 开票合同
    - 开票币种ID -> 开票币种
    - 业务类型ID -> 业务类型
    """
    df = source_df.copy()
    employee_map = c.build_fw_employee_info_map_for_ids(df['申请人ID'])
    department_map = c.build_fw_department_name_map_for_ids(df['申请人部门ID'])
    company_map = c.build_fw_company_name_map_for_ids(df['公司主体ID'])
    customer_map = c.build_fw_customer_name_map_for_ids(df['客户ID'])
    contract_map = c.build_fw_contract_code_map_for_ids(df['开票合同ID'])
    currency_map = c.build_fw_currency_name_map_for_ids(df['开票币种ID'])
    cost_center_map = c.build_fw_cost_center_map_for_ids(df['成本中心ID'])
    cost_center_code_map = c.build_fw_cost_center_code_map_for_ids(df['成本中心ID'])
    project_map = build_fw_project_code_map_for_ids(df['项目编号ID'])

    # [开票表] sqr -> hrmresource / hrmjobtitles
    df['申请人'] = df['申请人ID'].map(lambda value: employee_map.get(c.format_code(value), {}).get('name', ''))
    df['申请人工号'] = df['申请人ID'].map(lambda value: employee_map.get(c.format_code(value), {}).get('code', ''))
    # [开票表] sqrbm -> hrmdepartment.DEPARTMENTNAME
    df['申请人部门'] = df['申请人部门ID'].map(lambda value: department_map.get(c.format_code(value), ''))
    # [开票表] gszt -> uf_gstt.gsmc
    df['公司主体'] = df['公司主体ID'].map(lambda value: company_map.get(c.format_code(value), ''))
    # [开票表] kh -> uf_khgys.khmc
    df['客户'] = df['客户ID'].map(lambda value: customer_map.get(_first_browser_id(value), ''))
    # [开票表] kpht -> uf_htsp.htbh
    df['开票合同'] = df['开票合同ID'].map(lambda value: contract_map.get(_first_browser_id(value), ''))
    # [开票表] xmbh -> uf_xtyyxmkp.xmbh
    df['项目编号'] = [
        _lookup_first_browser_value(project_map, project_value)
        or _text(project_value)
        or _lookup_first_browser_value(cost_center_code_map, cost_center_value)
        for project_value, cost_center_value in zip(df['项目编号ID'], df['成本中心ID'])
    ]
    # [开票表] kpbz -> fnacurrency.CURRENCYNAME
    df['开票币种'] = df['开票币种ID'].map(lambda value: currency_map.get(c.format_code(value), ''))
    # [开票表] cbzx -> uf_cbzx.mc(成本中心名称)
    df['成本中心'] = df['成本中心ID'].map(lambda value: _lookup_first_browser_value(cost_center_map, value))
    # [开票表] ywlx:0=外部公司,1=外部个人,2=空业务类型
    df['业务类型'] = df['业务类型ID'].map(resolve_business_type_name)
    return df


def resolve_mcn_source_values(source_df):
    """基于 MCN 开票旧流程明细补充输出需要的展示值。"""
    df = source_df.copy()
    if df.empty:
        df['申请人'] = ''
        df['申请人工号'] = ''
        df['申请人部门'] = ''
        df['公司主体'] = ''
        df['客户'] = ''
        df['开票合同'] = ''
        df['项目编号'] = ''
        df['开票类型'] = ''
        df['平台'] = ''
        df['成本中心'] = ''
        return df

    employee_map = c.build_fw_employee_info_map_for_ids(df['申请人ID'])
    department_map = c.build_fw_department_name_map_for_ids(df['申请人部门ID'])
    company_map = c.build_fw_company_name_map_for_ids(df['公司主体ID'])
    customer_map = c.build_fw_customer_name_map_for_ids(
        pd.concat([df['客户ID'], df['明细客户ID']], ignore_index=True))
    contract_map = _mcn_contract_code_map(df['开票合同ID'])
    project_map = build_fw_project_code_map_for_ids(df['项目编号ID'], MCN_PROJECT_TABLES)
    cost_center_map = c.build_fw_cost_center_map_for_ids(df['成本中心ID'])
    option_maps = c.build_fw_select_option_maps(MCN_INVOICE_TABLE, ['kplx', 'pt'])

    df['申请人'] = df['申请人ID'].map(lambda value: employee_map.get(c.format_code(value), {}).get('name', ''))
    df['申请人工号'] = df['申请人ID'].map(lambda value: employee_map.get(c.format_code(value), {}).get('code', ''))
    df['申请人部门'] = df['申请人部门ID'].map(lambda value: department_map.get(c.format_code(value), ''))
    df['公司主体'] = df['公司主体ID'].map(
        lambda value: _lookup_first_browser_value(company_map, value) or company_map.get(c.format_code(value), ''))
    df['客户'] = [
        _lookup_first_browser_value(customer_map, customer_value)
        or _lookup_first_browser_value(customer_map, detail_customer_value)
        for customer_value, detail_customer_value in zip(df['客户ID'], df['明细客户ID'])
    ]
    df['开票合同'] = df['开票合同ID'].map(lambda value: _lookup_first_browser_value(contract_map, value))
    df['项目编号'] = df['项目编号ID'].map(
        lambda value: _lookup_first_browser_value(project_map, value) or _text(value))
    df['开票类型'] = df['开票类型ID'].map(lambda value: option_maps.get('kplx', {}).get(c.format_code(value), ''))
    df['平台'] = df['平台ID'].map(lambda value: option_maps.get('pt', {}).get(c.format_code(value), ''))
    df['成本中心'] = df['成本中心ID'].map(lambda value: _lookup_first_browser_value(cost_center_map, value))
    df['泛微订单编号'] = _resolve_mcn_order_numbers(df)
    return df


# ============================ 模板输出 ============================
def _lookup_by_name(mapping, value):
    return '' if pd.isna(value) else mapping.get(c.normalize_name(value), '')


def _business_type_code(value, business_type_map):
    meaning = BUSINESS_TYPE_MEANING.get(_text(value), BUSINESS_TYPE_MEANING[''])
    return business_type_map.get(meaning, '')


def _normalize_tax_rate(value):
    if pd.isna(value):
        return None
    text = str(value).strip().replace('%', '')
    if not text or text in ('nan', 'None'):
        return None
    try:
        rate = float(text)
    except ValueError:
        return None
    if rate > 1:
        rate = rate / 100
    return round(rate, 4)


def _tax_description(value, tax_description_map):
    rate = _normalize_tax_rate(value)
    if rate is None:
        return ''
    return tax_description_map.get(rate, '')


def _mcn_invoice_type_code(value, invoice_type_map):
    """MCN 开票类型 -> 汉得 HERO.INVOICE_TYPE 编码。

    规则表口径:合同开票/订单开票按合同开票;平台预开票/平台结算单按平台预开票。
    泛微旧表里「平台订单」即订单开票口径。
    """
    meaning = _text(value)
    if meaning in ('平台预开票', '平台结算单'):
        return invoice_type_map.get(PLATFORM_PRE_INVOICE_MEANING, '')
    if meaning in ('合同开票', '订单开票', '平台订单'):
        return invoice_type_map.get(CONTRACT_INVOICE_MEANING, '')
    return invoice_type_map.get(CONTRACT_INVOICE_MEANING, '')


def _mcn_contract_code_map(contract_values):
    """MCN 合同浏览框 ID -> 合同编号。

    MCN 开票旧流程的 browser.ktl 主要落 uf_htk;历史/协同数据兜底查 uf_htsp。
    """
    contract_ids = c.clean_codes(
        contract_id
        for value in contract_values
        for contract_id in c.parse_browser_ids(value)
    )
    if not contract_ids:
        return {}

    result = {}
    for table in ('uf_htk', 'uf_htsp'):
        remaining = [contract_id for contract_id in contract_ids if contract_id not in result]
        if not remaining:
            break
        for batch in _chunks(remaining):
            try:
                contract_df = c.query_db(
                    'FW',
                    'vspn_xtyy',
                    f'SELECT id, htbh AS contract_code FROM {table} '
                    f'WHERE id IN ({c.in_placeholders(batch)})',
                    batch,
                )
            except Exception:
                continue
            for _, row in contract_df.iterrows():
                contract_id = c.format_code(row['id'])
                contract_code = _text(row['contract_code'])
                if contract_id and contract_code and contract_id not in result:
                    result[contract_id] = contract_code
    return result


def _apply_issue_order_fix(output_df):
    if output_df.empty:
        return output_df
    required = {'来源单据号', '泛微项目编号', '订单'}
    if not required.issubset(output_df.columns):
        return output_df
    fix_df = _load_issue_order_fix()
    if fix_df.empty:
        return output_df

    result = output_df.copy()
    keys = pd.DataFrame({
        '_row_pos': range(len(result)),
        '来源单据编号': result['来源单据号'].map(_text),
        '泛微项目编号': result['泛微项目编号'].map(_text),
    })
    merged = keys.merge(
        fix_df[['来源单据编号', '泛微项目编号', '指定订单编号']],
        on=['来源单据编号', '泛微项目编号'],
        how='left',
        sort=False,
    )
    matched = merged['指定订单编号'].map(_text) != ''
    if matched.any():
        row_positions = merged.loc[matched, '_row_pos'].astype(int).to_numpy()
        result.iloc[row_positions, result.columns.get_loc('订单')] = merged.loc[matched, '指定订单编号'].to_numpy()
        print(f'[应收期初-问题处理] 覆盖赛事订单编号 {int(matched.sum())} 行')
    return result


def _apply_issue_tax_fix(output_df):
    if output_df.empty or not {'来源单据号', '税率类型'}.issubset(output_df.columns):
        return output_df
    fix_df = _load_issue_tax_fix()
    if fix_df.empty:
        return output_df

    result = output_df.copy()
    tax_map = {
        _text(row['来源单据号']): _text(row['指定税率类型'])
        for _, row in fix_df.iterrows()
    }
    fixed_values = result['来源单据号'].map(lambda value: tax_map.get(_text(value), ''))
    matched = fixed_values.map(_text) != ''
    if matched.any():
        result.loc[matched, '税率类型'] = fixed_values.loc[matched]
        print(f'[应收期初-问题处理] 覆盖赛事税率类型 {int(matched.sum())} 行')
    return result


def _apply_event_issue_fixes(output_df):
    result = _apply_issue_order_fix(output_df)
    result = _apply_issue_tax_fix(result)
    return result


def _drop_fixed_order_issue_rows(sheets):
    fix_df = _load_issue_order_fix()
    issue_df = sheets.get('订单映射_多候选')
    if fix_df.empty or issue_df is None or issue_df.empty:
        return sheets

    fixed_keys = set(zip(fix_df['来源单据编号'].map(_text), fix_df['泛微项目编号'].map(_text)))
    keep_mask = [
        (_text(doc_no), _text(project_code)) not in fixed_keys
        for doc_no, project_code in zip(issue_df['来源单据编号'], issue_df['泛微项目编号'])
    ]
    removed = len(issue_df) - sum(keep_mask)
    if removed:
        sheets['订单映射_多候选'] = issue_df.loc[keep_mask].reset_index(drop=True)
        print(f'[应收期初-问题处理] 订单多候选清单剔除已指定 {removed} 行')
    return sheets


def build_output(invoice_df):
    """DB 源数据 -> 应收报账单期初导入模板 75 列。
    能从泛微 DB 原始字段/ID 直接取得的字段直接取;跨系统编码再查中台/值集。
    """
    # 核算主体: [开票表] gszt -> 公司主体名称 -> Hand hfac_accounting_entity.acc_entity_code。
    entity_map = c.build_accounting_entity_map_for_names(invoice_df['公司主体'])
    # 付款对象: [开票表] kh -> 泛微客户名称 -> Hand hfbs_system_customer.customer_code。
    customer_map = c.build_customer_map_for_names(invoice_df['客户'])
    # 业务类型/开票类型: Hand 值集。
    business_type_map = c.build_lov_meaning_map(BUSINESS_TYPE_LOV)
    invoice_type_map = c.build_lov_meaning_map(INVOICE_TYPE_LOV)
    contract_invoice_code = invoice_type_map.get(CONTRACT_INVOICE_MEANING, '')
    # 税率类型: 税率 -> Hand 税率类型描述。
    tax_description_map = c.build_tax_type_description_map(TAX_PREFERRED_DESCRIPTIONS)

    tax_rate_source = invoice_df['税率'].where(
        invoice_df['税率'].notna(), invoice_df['税率（明细）'])
    tax_amount = invoice_df['税额（明细）'].where(
        invoice_df['税额（明细）'].notna(), invoice_df['税额'])

    output_df = pd.DataFrame(index=invoice_df.index)

    # 固定值。
    output_df['应收报账单类型'] = DOCUMENT_TYPE
    output_df['管理公司'] = MANAGEMENT_COMPANY
    output_df['付款对象类型'] = PAYER_TYPE
    output_df['开票类型编码'] = contract_invoice_code
    output_df['收入项目'] = INCOME_ITEM

    # 泛微开票表直取/解析字段。
    output_df['当前节点'] = invoice_df.get(
        '当前节点', pd.Series('', index=invoice_df.index)).map(_text)
    output_df['当前状况'] = invoice_df.get(
        '当前状况', pd.Series('', index=invoice_df.index)).map(_text)
    output_df['来源单据号'] = invoice_df['流程编号']                  # [开票表] lcbh
    output_df['部门'] = invoice_df['申请人部门']                      # [开票表] sqrbm -> hrmdepartment
    output_df['岗位'] = ''                                           # 不涉及
    output_df['申请人'] = invoice_df['申请人工号']                    # [开票表] sqr -> hrmjobtitles.JOBTITLENAME
    output_df['申请日期'] = invoice_df['申请日期'].map(c.format_date)  # [开票表] sqrq
    output_df['支付币种'] = invoice_df['开票币种'].map(c.to_iso_currency)  # [开票表] kpbz -> ISO
    output_df['合同编号'] = invoice_df['开票合同'].where(
        invoice_df['开票合同'].notna(), '')                          # [开票表] kpht -> uf_htsp.htbh
    output_df['头备注'] = invoice_df['开票备注'].astype(str).where(
        invoice_df['开票备注'].notna(), '').str.slice(0, 150)         # [开票表] kptxt,截前150字符
    output_df['发票号'] = invoice_df['发票号'].map(_invoice_numbers_text)  # [开票表] fphm,多个发票号用英文逗号隔开

    # 当前口径没有直接可用的维度字段,按模板留空。
    output_df['里程碑阶段'] = ''
    output_df['平台'] = ''
    output_df['自审批'] = ''
    output_df['自审核'] = ''
    output_df['凭证推送'] = ''
    output_df['凭证日期'] = ''
    output_df['行号'] = ''
    output_df['收入分类'] = ''
    output_df['数量'] = ''
    output_df['单价'] = ''
    output_df['行备注'] = ''

    # 跨系统映射字段。
    output_df['核算主体'] = invoice_df['公司主体'].map(
        lambda value: _lookup_by_name(entity_map, value))             # [开票表] gszt -> Hand 核算主体编码
    output_df['付款对象'] = invoice_df['客户'].map(
        lambda value: _lookup_by_name(customer_map, value))           # [开票表] kh -> Hand 客户编码
    output_df['业务类型编码'] = invoice_df['业务类型'].map(
        lambda value: _business_type_code(value, business_type_map))  # [开票表] ywlx -> HERO.BUSINESS_TYPE

    # 金额和税率。
    output_df['核销金额'] = pd.to_numeric(
        invoice_df['收款登记已收款金额'], errors='coerce').fillna(0).map(c.round_amount)  # [收款登记] bfqrjehj按单号汇总
    output_df['金额'] = pd.to_numeric(
        invoice_df['开票金额（含税价）'], errors='coerce').map(c.round_amount)  # [开票表] kpjehsj
    output_df['税率类型'] = tax_rate_source.map(
        lambda value: _tax_description(value, tax_description_map))   # [开票表] sl/slmx -> Hand税率类型描述
    output_df['税额'] = pd.to_numeric(tax_amount, errors='coerce').map(c.round_amount)  # [开票表] semx/se

    for column in [f'头维度{i}' for i in range(1, 21)]:
        output_df[column] = ''
    project_codes = invoice_df['项目编号'].map(_text)
    output_df['项目'] = project_codes.map(lambda value: c.project_order_mapping_value(value, '项目编号'))
    output_df['订单'] = project_codes.map(lambda value: c.project_order_mapping_value(value, '订单编号'))
    for column in [f'行维度{i}' for i in range(3, 21)]:
        output_df[column] = ''
    output_df['泛微项目编号'] = project_codes

    # write_to_template 按 DataFrame 顺序写入模板,这里显式固定列序。
    output_df = output_df[OUTPUT_COLUMNS]
    return _apply_event_issue_fixes(output_df)


def _mcn_receipt_allocations(invoice_df, line_amounts):
    """把 MCN 收款登记头金额按开票单号顺序分摊到明细行,每行不超过本行金额。"""
    if invoice_df.empty:
        return pd.Series([], index=invoice_df.index)

    receipt_amounts = pd.to_numeric(
        invoice_df.get('收款登记已收款金额', pd.Series(0, index=invoice_df.index)),
        errors='coerce',
    ).fillna(0)
    line_limits = pd.to_numeric(line_amounts, errors='coerce').fillna(0)
    allocations = pd.Series(0.0, index=invoice_df.index)

    invoice_numbers = invoice_df.get('开票单号', pd.Series('', index=invoice_df.index)).map(_text)
    row_fallback_keys = pd.Series(
        [f'__row_{row_no}' for row_no in range(len(invoice_numbers))],
        index=invoice_df.index,
    )
    group_keys = invoice_numbers.where(invoice_numbers != '', row_fallback_keys)

    for _, row_index in group_keys.groupby(group_keys, sort=False).groups.items():
        remaining = float(receipt_amounts.loc[row_index].iloc[0])
        if remaining <= 0:
            continue
        for idx in row_index:
            line_limit = float(line_limits.loc[idx])
            if line_limit <= 0:
                continue
            allocated = min(line_limit, remaining)
            allocations.at[idx] = c.round_amount(allocated)
            remaining = round(remaining - allocated, 10)
            if remaining <= 0:
                break
    return allocations.map(c.round_amount)


def build_mcn_output(invoice_df):
    """MCN 开票旧流程明细 -> 应收报账单期初导入模板 75 列。"""
    entity_map = c.build_accounting_entity_map_for_names(invoice_df['公司主体'])
    customer_map = c.build_customer_map_for_names(invoice_df['客户'])
    business_type_map = c.build_lov_meaning_map(BUSINESS_TYPE_LOV)
    invoice_type_map = c.build_lov_meaning_map(INVOICE_TYPE_LOV)
    public_invoice_code = business_type_map.get(PUBLIC_INVOICE_MEANING, '')
    tax_description_map = c.build_tax_type_description_map(MCN_TAX_PREFERRED_DESCRIPTIONS)

    output_df = pd.DataFrame(index=invoice_df.index)

    # 固定值。
    output_df['应收报账单类型'] = DOCUMENT_TYPE
    output_df['管理公司'] = MANAGEMENT_COMPANY
    output_df['支付币种'] = 'CNY'
    output_df['付款对象类型'] = PAYER_TYPE
    output_df['业务类型编码'] = public_invoice_code
    output_df['收入项目'] = INCOME_ITEM

    # 泛微 MCN 开票申请流程/开票申请流程（订单）直取或解析字段。
    output_df['当前节点'] = invoice_df.get(
        '当前节点', pd.Series('', index=invoice_df.index)).map(_text)
    output_df['当前状况'] = invoice_df.get(
        '当前状况', pd.Series('', index=invoice_df.index)).map(_text)
    output_df['来源单据号'] = invoice_df['开票单号']
    output_df['部门'] = invoice_df['申请人部门']
    output_df['岗位'] = ''                                             # 不涉及
    output_df['申请人'] = invoice_df['申请人工号']
    output_df['申请日期'] = invoice_df['申请日期'].map(c.format_date)
    output_df['合同编号'] = invoice_df['开票合同'].where(invoice_df['开票合同'].notna(), '')
    output_df['平台'] = invoice_df['平台'].map(_text)

    # 规则表备注为「不涉及」或当前无 MCN 源字段的列留空。
    output_df['里程碑阶段'] = ''
    output_df['头备注'] = invoice_df.get(
        '标题', pd.Series('', index=invoice_df.index)).map(_text).str.slice(0, 150)
    output_df['发票号'] = ''
    output_df['自审批'] = ''
    output_df['自审核'] = ''
    output_df['凭证推送'] = ''
    output_df['凭证日期'] = ''
    output_df['行号'] = ''
    output_df['收入分类'] = ''
    output_df['数量'] = ''
    output_df['单价'] = ''
    output_df['行备注'] = ''

    # 跨系统映射字段。
    output_df['核算主体'] = invoice_df['公司主体'].map(lambda value: _lookup_by_name(entity_map, value))
    output_df['付款对象'] = invoice_df['客户'].map(lambda value: _lookup_by_name(customer_map, value))
    output_df['开票类型编码'] = invoice_df['开票类型'].map(
        lambda value: _mcn_invoice_type_code(value, invoice_type_map))

    # 金额和税率。
    line_amounts = pd.to_numeric(invoice_df['开票金额'], errors='coerce')
    output_df['核销金额'] = _mcn_receipt_allocations(invoice_df, line_amounts)
    output_df['金额'] = line_amounts.map(c.round_amount)
    output_df['税率类型'] = invoice_df['税率'].map(lambda value: _tax_description(value, tax_description_map))
    output_df['税额'] = pd.to_numeric(invoice_df['税额'], errors='coerce').map(c.round_amount)

    for column in [f'头维度{i}' for i in range(1, 21)]:
        output_df[column] = ''
    project_codes = invoice_df['项目编号'].map(_text)
    output_df['项目'] = project_codes.map(lambda value: c.project_order_mapping_value(value, '项目编号'))
    output_df['订单'] = project_codes.map(lambda value: c.project_order_mapping_value(value, '订单编号'))
    for column in [f'行维度{i}' for i in range(3, 21)]:
        output_df[column] = ''
    output_df['泛微项目编号'] = project_codes
    output_df['原泛微订单号'] = invoice_df.get(
        '泛微订单编号', pd.Series('', index=invoice_df.index)).map(_text)

    return output_df[MCN_OUTPUT_COLUMNS]


def _filter_by_project_whitelist(invoice_df, sheet_name):
    """按指定白名单 sheet 的「原泛微项目编码」过滤已解析项目编号的开票记录。

    应收期初当前仅处理赛事,传 EVENT_PROJECT_SHEET;后续 MCN 数据处理时另走
    MCN_PROJECT_SHEET,两部分白名单区分开,互不混用。
    """
    if invoice_df.empty:
        return invoice_df
    allowed_codes = project_filter_codes(sheet_name)
    mask = invoice_df['项目编号'].map(_text).isin(allowed_codes)
    filtered = invoice_df.loc[mask].copy()
    print(f"[应收期初-应收报账单-DB] 项目白名单过滤({sheet_name}): "
          f"{len(filtered)}/{len(invoice_df)} 行 (白名单项目 {len(allowed_codes)} 个)")
    return filtered


def read_invoice_source():
    """从 DB 读取未作废开票记录,解析后按项目白名单过滤。"""
    c.validate_fw_fields(FW_INVOICE_TABLE, EXPECTED_INVOICE_FIELDS)
    c.validate_fw_fields(FW_RECEIPT_TABLE, EXPECTED_RECEIPT_FIELDS)
    stats = _query_fw(STATS_SQL).iloc[0]
    void_name = VOID_FLAG_MEANINGS.get(VOID_CODE, '')
    print(f"[应收期初-应收报账单-DB] SQL过滤: 仅 是否作废≠{VOID_CODE}({void_name}); "
          f"不再过滤申请日期/开票状态,改由项目白名单过滤")
    print(f"  全部 {int(stats['total_count'] or 0)} 行; "
          f"剔除作废 {int(stats['void_count'] or 0)} 行; "
          f"SQL保留(未作废) {int(stats['kept_count'] or 0)} 行")

    source_df = attach_workflow_states(_query_fw(SOURCE_SQL))
    invoice_df = resolve_source_values(source_df)
    print('[应收期初-应收报账单-DB] 未作废开票记录行数:', len(invoice_df))
    # 应收期初当前仅处理赛事,只按赛事白名单过滤;MCN 部分后续单独按 MCN sheet 处理。
    invoice_df = _filter_by_project_whitelist(invoice_df, EVENT_PROJECT_SHEET)
    return invoice_df


def read_mcn_invoice_source():
    """从 DB 读取流程完成的 MCN 开票申请明细,解析后按 MCN 项目白名单过滤。"""
    c.validate_fw_fields(MCN_INVOICE_TABLE, EXPECTED_MCN_INVOICE_FIELDS)
    c.validate_fw_fields(MCN_ORDER_INVOICE_TABLE, EXPECTED_MCN_ORDER_INVOICE_FIELDS)
    c.validate_fw_fields(FW_RECEIPT_TABLE, EXPECTED_RECEIPT_FIELDS)

    params = {
        'complete_node_type': COMPLETE_NODE_TYPE,
        'mcn_invoice_workflow_ids': MCN_INVOICE_WORKFLOW_IDS,
        'mcn_order_workflow_ids': MCN_ORDER_WORKFLOW_IDS,
    }
    print('[应收期初-MCN开票-DB] SQL过滤: 仅保留开票申请流程/开票申请流程（订单）且流程完成')
    source_df = c.query_db('FW', 'vspn_xtyy', MCN_SOURCE_SQL, params)
    print('[应收期初-MCN开票-DB] 流程完成开票明细行数:', len(source_df))
    source_df = attach_workflow_states(source_df)

    receipt_df = c.query_db('FW', 'vspn_xtyy', MCN_RECEIPT_SQL)
    if not receipt_df.empty:
        source_df = source_df.merge(receipt_df, on='开票单号', how='left')
    else:
        source_df['收款登记已收款金额'] = 0
    source_df['收款登记已收款金额'] = source_df['收款登记已收款金额'].fillna(0)

    invoice_df = resolve_mcn_source_values(source_df)
    invoice_df = _filter_by_project_whitelist(invoice_df, MCN_PROJECT_SHEET)
    return invoice_df


def write_output_workbook(event_output_df, mcn_output_df):
    """一次写出赛事原 sheet 和新增 MCN sheet,保留模板表头/下拉页。"""
    workbook = load_workbook(TEMPLATE_FILE)
    source_sheet = workbook[TEMPLATE_SHEET]
    if MCN_TEMPLATE_SHEET in workbook.sheetnames:
        del workbook[MCN_TEMPLATE_SHEET]
    mcn_sheet = workbook.copy_worksheet(source_sheet)
    mcn_sheet.title = MCN_TEMPLATE_SHEET

    c._fill_sheet(source_sheet, event_output_df)
    c._fill_sheet(mcn_sheet, mcn_output_df)

    front_sheets = [TEMPLATE_SHEET, MCN_TEMPLATE_SHEET]
    front_set = set(front_sheets)
    workbook._sheets = [workbook[name] for name in front_sheets] + [
        sheet for sheet in workbook.worksheets if sheet.title not in front_set
    ]
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(OUTPUT_FILE)
        return OUTPUT_FILE
    except PermissionError:
        fallback_file = OUTPUT_FILE.with_name(
            f'{OUTPUT_FILE.stem}_{datetime.now().strftime("%H%M%S")}{OUTPUT_FILE.suffix}'
        )
        print(f'[应收期初] 输出文件被占用，已改写到: {fallback_file}')
        workbook.save(fallback_file)
        return fallback_file


def _prefixed_sheets(prefix, sheets):
    result = {}
    for sheet_name, sheet_df in sheets.items():
        prefixed = f'{prefix}_{sheet_name}'
        result[prefixed[:31]] = sheet_df
    return result


def _safe_report_fill(label, output_df, required_cols):
    print(label)
    if output_df.empty:
        print('  (输出 0 行,跳过填充率计算)')
        return
    c.report_fill(output_df, required_cols)


def run():
    # 1. SQL 直接查过滤后的赛事开票记录 + 收款登记汇总金额
    invoice_df = read_invoice_source()
    # 2. SQL 读取 MCN 开票申请流程/开票申请流程（订单）明细
    mcn_invoice_df = read_mcn_invoice_source()

    # 3. 构建输出
    output_df = build_output(invoice_df)
    mcn_output_df = build_mcn_output(mcn_invoice_df)
    print('[应收期初-应收报账单-DB] 赛事输出明细行数:', len(output_df))
    print('[应收期初-MCN开票-DB] MCN输出明细行数:', len(mcn_output_df))

    # 4. 填充率(必输字段以规则表「是否必填」=Y 为准)
    required_cols = c.required_columns(RULE_SHEET, RULE_TABLE)
    _safe_report_fill('[应收期初-应收报账单-DB] 赛事必输字段填充率:', output_df, required_cols)
    _safe_report_fill('[应收期初-MCN开票-DB] MCN必输字段填充率:', mcn_output_df, required_cols)

    # 5. 写模版
    output_path = write_output_workbook(output_df, mcn_output_df)
    print('已写出:', output_path)

    # 6. 问题清单
    event_sheets = {'必输字段未达100%': c.fill_summary(output_df, required_cols, RULE_SHEET, RULE_TABLE)}
    event_sheets.update(c.collect_field_issues(
        output_df, invoice_df, required_cols, ISSUE_SOURCE_FIELD_MAP, doc_col='来源单据号'))
    event_sheets.update(c.collect_order_mapping_issues(invoice_df))
    _drop_fixed_order_issue_rows(event_sheets)
    # 给各未匹配清单补「成本中心」「预算项」两列(应收有成本中心,无预算科目则预算项留空)。
    c.attach_budget_issue_columns(event_sheets, c.build_budget_issue_map(invoice_df))

    mcn_sheets = {'必输字段未达100%': c.fill_summary(mcn_output_df, required_cols, RULE_SHEET, RULE_TABLE)}
    mcn_sheets.update(c.collect_field_issues(
        mcn_output_df, mcn_invoice_df, required_cols, ISSUE_SOURCE_FIELD_MAP, doc_col='来源单据号'))
    mcn_sheets.update(c.collect_order_mapping_issues(mcn_invoice_df, doc_col='开票单号'))
    c.attach_budget_issue_columns(mcn_sheets, c.build_budget_issue_map(mcn_invoice_df, doc_col='开票单号'))

    sheets = {}
    sheets.update(_prefixed_sheets('赛事', event_sheets))
    sheets.update(_prefixed_sheets('MCN', mcn_sheets))
    c.write_exceptions(EXCEPTION_FILE, sheets)
    print('已写出:', EXCEPTION_FILE, '| 各清单条数:', {
        sheet_name: len(sheet_df) for sheet_name, sheet_df in sheets.items()
    })


if __name__ == '__main__':
    run()
