# -*- coding: utf-8 -*-
"""预付期初 —— 供应商预付款单 / 零工预付款单(DB 直连版)。

处理流程:
1. 校验泛微字段字典,避免 SQL 字段名/含义写错。
2. 从泛微库读取供应商预付 uf_yfkxx/uf_yfkxx_dt1,并从 uf_dgfktz_dt2 汇总对公付款冲销;
   零工付款读取 uf_lgptfk + 原流程收款人明细/预算项明细。
3. 只对必须跨表/跨系统的 ID 做批量解析,例如人员、公司主体、币种、预算科目、供应商、银行账号。
4. 按导入模版两个 tab 逐列生成输出,字段旁标注取值来源。

跑法:在项目根执行  python run.py ap_prepayment_opening_db
"""
import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

if __package__ is None or __package__ == '':
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from etl.util import common as c

# ============================ 文件 / 模板 ============================
TASK_NAME = 'ap_prepayment_opening_db'
TEMPLATE_DIR = c.TPL_DIR / 'ap_prepayment_opening'
OUTPUT_DIR = c.OUT_DIR / TASK_NAME
DATE_SUFFIX = c.today_suffix()

TEMPLATE_FILE = TEMPLATE_DIR / '英雄期初预付款单导入模版.xlsx'
OUTPUT_FILE = OUTPUT_DIR / f'英雄期初预付款单导入_预付期初全量_{DATE_SUFFIX}.xlsx'
EXCEPTION_FILE = OUTPUT_DIR / f'未匹配清单_预付期初全量_{DATE_SUFFIX}.xlsx'
SUPPLIER_VENDOR_MISSING_FILE = OUTPUT_DIR / f'Hand按ID查不到的供应商_预付期初_{DATE_SUFFIX}.xlsx'
MCN_SUPPLIER_VENDOR_MISSING_FILE = OUTPUT_DIR / f'Hand按ID查不到的供应商_MCN预付期初_{DATE_SUFFIX}.xlsx'

TEMPLATE_SHEET_SUPPLIER = '期初供应商&期初投资-赛事'
TEMPLATE_SHEET_GIG = '期初灵工-赛事'
TEMPLATE_SOURCE_SHEET_SUPPLIER = '期初供应商预付款单&期初投资付款单导入'
TEMPLATE_SOURCE_SHEET_GIG = '期初灵工预付款单导入'
SUPPLIER_SUMMARY_SHEET = '期初供应商汇总'
GIG_SUMMARY_SHEET = '期初零工汇总'
SHEET_MCN_PREPAYMENT = '期初供应商&期初投资-MCN预付款流程'
SHEET_MCN_PREPAYMENT_ORDER = '期初供应商&期初投资-MCN预付款流程(订单)'
SHEET_MCN_ANCHOR_PREPAYMENT = '期初供应商&期初投资-MCN主播相关预付款'
SHEET_MCN_OUTBOUND_PREPAYMENT = '期初灵工-MCN对外付款非直付'
SHEET_MCN_OUTBOUND_ORDER_PREPAYMENT = '期初灵工-MCN对外付款订单非直付'
SHEET_MCN_ANCHOR_PLATFORM_PREPAYMENT = '期初灵工-MCN主播付款非直付'
RULE_SHEET = '预付期初'
RULE_TABLE_SUPPLIER = '期初供应商预付款单&期初投资付款单导入'
RULE_TABLE_GIG = '期初灵工预付款单导入'

DOCUMENT_TYPE = 'JK01-2'
GIG_DOCUMENT_TYPE = 'PP01-2'
GIG_PLATFORM_VENDOR = {'云账户': 'V-C-CN-HR-PAY-0001', '赛利得': 'V-C-CN-OT-OTH-6573'}
PROJECT_FILTER_ENV = 'PROJECT_FILTER_XLSX'
PROJECT_FILTER_DEFAULT_FILE = c.RULES_DIR / '数据清洗涉及泛微项目编码_0624_分类.xlsx'
PROJECT_FILTER_FILE = Path(os.getenv(PROJECT_FILTER_ENV, '').strip() or PROJECT_FILTER_DEFAULT_FILE)
EVENT_PROJECT_SHEET = '赛事'
MCN_PROJECT_SHEET = 'MCN'
EVENT_PROJECT_TABLES = ('uf_xtyyxmkp', 'uf_xmkp', 'view_xmjkzb')
MCN_PROJECT_TABLES = ('uf_xmkp', 'view_xmjkzb', 'uf_xtyyxmkp')
SQL_BATCH_SIZE = 800
SOURCE_SYSTEM = 'FW'
DIRECT_PAYMENT_PLATFORM_CODE = '2'
PREPAYMENT_YES_CODE = '0'
PREPAYMENT_NO_CODE = '1'

SUPPLIER_OUTPUT_COLUMNS = [
    '来源系统',
    '来源单据编号',
    '申请日期',
    '单据类型',
    '申请人工号',
    '申请人姓名',
    '订单编号',
    '订单名称',
    '核算主体编号',
    '核算主体描述',
    '备注',
    '合同号',
    '合同收支计划行',
    '保证金标志',
    '收款方编码',
    '收款方描述',
    '银行账号',
    '计划付款日期',
    '银行转账备注',
    '费用项目编码',
    '费用项目描述',
    '主播房间号',
    '预付款支付币种',
    '预付款金额（支付币种）',
    '已到票核销金额（支付币种）',
    '已付未核（支付币种）',
    '泛微项目编号',
    '泛微费用项目编码',
]
GIG_OUTPUT_COLUMNS = [
    '来源系统',
    '来源单据编号',
    '申请日期',
    '单据类型',
    '申请人工号',
    '申请人姓名',
    '订单编号',
    '订单名称',
    '核算主体编号',
    '核算主体描述',
    '备注_单头',
    '灵工平台收款方编码',
    '合同号',
    '合同收支计划行',
    '保证金标志',
    '计划付款日期',
    '银行转账备注',
    '费用项目编码',
    '费用项目描述',
    '收款方类别',
    '收款方编码',
    '备注',
    '银行账号',
    '预付款支付币种',
    '预付款金额（支付币种）',
    '传送状态',
    '支付状态',
    '退款状态',
    '核销状态',
    '泛微项目编号',
    '泛微费用项目编码',
]

SUPPLIER_ISSUE_SOURCE_FIELDS = {
    '申请人工号': '填单人',
    '收款方编码': '付款对象',
    '核算主体编号': '开票单位',
    '费用项目编码': '预算科目',
    '预付款支付币种': '付款币种',
    '订单编号': '项目编号',
}
GIG_ISSUE_SOURCE_FIELDS = {
    '申请人工号': '经办人',
    '灵工平台收款方编码': '收款方文本',
    '核算主体编号': '公司主体',
    '费用项目编码': '预算科目',
    '收款方编码': '实际收款方',
    '订单编号': '项目编号',
}
MCN_SUPPLIER_ISSUE_SOURCE_FIELDS = {
    '申请人工号': '申请人',
    '收款方编码': '供应商',
    '核算主体编号': '公司主体',
    '费用项目编码': '预算科目',
    '预付款支付币种': '',
    '订单编号': '项目编号',
}
MCN_GIG_ISSUE_SOURCE_FIELDS = {
    '申请人工号': '申请人',
    '灵工平台收款方编码': '供应商',
    '核算主体编号': '公司主体',
    '费用项目编码': '预算科目',
    '收款方编码': '实际收款方',
    '订单编号': '项目编号',
}
GIG_NOT_INVOLVED_REQUIRED_COLUMNS = {'保证金标志'}
MCN_GIG_NOT_INVOLVED_REQUIRED_COLUMNS = {'保证金标志', '备注'}

FW_SUPPLIER_TABLE = 'uf_yfkxx'
FW_SUPPLIER_DETAIL_TABLE = 'uf_yfkxx_dt1'
FW_PAYMENT_OFFSET_TABLE = 'uf_dgfktz'
FW_PAYMENT_OFFSET_DETAIL_TABLE = 'uf_dgfktz_dt2'
FW_GIG_HEADER_TABLE = 'uf_lgptfk'
FW_GIG_WORKFLOW_TABLE = 'formtable_main_279'
FW_GIG_BUDGET_TABLE = 'formtable_main_279_dt3'
FW_GIG_RECIPIENT_TABLE = 'formtable_main_279_dt4'
FW_PROJECT_TABLE = 'uf_xtyyxmkp'
FW_MCN_PREPAYMENT_TABLE = 'formtable_main_29'
FW_MCN_PREPAYMENT_ORDER_TABLE = 'formtable_main_83'
FW_MCN_ANCHOR_TABLE = 'formtable_main_38'
FW_MCN_OUTBOUND_TABLE = 'formtable_main_33'
FW_MCN_OUTBOUND_ORDER_TABLE = 'formtable_main_66'
FW_MCN_PREPAYMENT_OFFSET_TABLE = 'formtable_main_30'
FW_MCN_PREPAYMENT_ORDER_OFFSET_TABLE = 'formtable_main_82'
FW_MCN_ANCHOR_OFFSET_TABLE = 'formtable_main_105'


# ============================ 枚举 / 过滤口径 ============================
DATE_FROM = '2022-01-01'
PROJECT_FALLBACK_DATE_FROM = '2026-01-01'
APPROVED_STATUS_CODE = 2
VOID_CODE = 0
DEPARTMENT_MODULE_CODE = '1'
DEPOSIT_PAYMENT_NATURE_CODES = {'0', '1'}  # uf_yfkxx.fkxz:0=押金,1=质保金

FLOW_STATUS_MEANINGS = {
    0: '未提交',
    1: '审批中',
    2: '审批完成',
}
VOID_FLAG_MEANINGS = {
    0: '是',
    1: '否',
}
PAYMENT_NATURE_MEANINGS = {
    0: '押金',
    1: '质保金',
    2: '一般',
}


# ============================ 泛微源 SQL ============================
SUPPLIER_SOURCE_SQL = """
SELECT
    m.id AS `ID`,
    d.id AS `明细ID`,
    m.lcbh AS `流程编号`,
    m.sqrq AS `申请日期`,
    m.tdr AS `填单人ID`,
    m.xmbh AS `项目编号ID`,
    m.kpdw AS `开票单位ID`,
    m.cbzx AS `成本中心ID`,
    cc.bh AS `成本中心编号`,
    m.bz AS `备注`,
    m.xght AS `相关合同ID`,
    m.fkdx AS `付款对象ID`,
    m.yhkh AS `银行卡号ID`,
    m.fkbz AS `付款币种ID`,
    m.fkje AS `付款金额`,
    m.fkxz AS `付款性质ID`,
    m.sycxtkje AS `剩余冲销/退款金额`,
    d.yfje AS `预付金额`,
    d.yskm AS `预算科目ID`,
    COALESCE(w.written_off_amount, 0) AS `冲销金额（支付币种-同预付单预算科目）`
FROM uf_yfkxx m
JOIN uf_yfkxx_dt1 d ON d.mainid = m.id
LEFT JOIN uf_cbzx cc ON cc.id = m.cbzx
LEFT JOIN (
    SELECT
        CAST(x.yfkxx AS CHAR) AS prepayment_id,
        CAST(x.yskm AS CHAR) AS budget_subject_id,
        SUM(ABS(COALESCE(x.cxje, 0))) AS written_off_amount
    FROM uf_dgfktz_dt2 x
    JOIN uf_dgfktz p ON p.id = x.mainid
    WHERE x.yfkxx IS NOT NULL
      AND x.yfkxx <> ''
    GROUP BY CAST(x.yfkxx AS CHAR), CAST(x.yskm AS CHAR)
) w ON w.prepayment_id = CAST(m.id AS CHAR)
   AND w.budget_subject_id = CAST(d.yskm AS CHAR)
WHERE (
      m.xmbh IN %(project_ids)s
      OR (
          (m.xmbh IS NULL OR CAST(m.xmbh AS CHAR) = '')
          AND m.sqrq >= %(project_fallback_date_from)s
          AND cc.bh IN %(project_codes)s
      )
  )
ORDER BY m.id, d.id
"""

SUPPLIER_STATS_SQL = """
SELECT
    COUNT(DISTINCT m.id) AS document_count,
    COUNT(*) AS detail_count,
    SUM(d.yfje) AS amount_total
FROM uf_yfkxx m
JOIN uf_yfkxx_dt1 d ON d.mainid = m.id
LEFT JOIN uf_cbzx cc ON cc.id = m.cbzx
WHERE (
      m.xmbh IN %(project_ids)s
      OR (
          (m.xmbh IS NULL OR CAST(m.xmbh AS CHAR) = '')
          AND m.sqrq >= %(project_fallback_date_from)s
          AND cc.bh IN %(project_codes)s
      )
  )
"""

GIG_RECIPIENT_SOURCE_SQL = """
SELECT
    h.id AS `建模付款ID`,
    w.id AS `流程付款ID`,
    d.id AS `收款人明细ID`,
    h.lcbh AS `流程编号`,
    h.sqrq AS `申请日期`,
    h.lcfsrq AS `流程发生日期`,
    w.requestId AS `流程请求ID`,
    h.lczsjqqid AS `流程转数据请求ID`,
    COALESCE(NULLIF(CAST(h.xmbh AS CHAR), ''), NULLIF(CAST(w.xmbh AS CHAR), '')) AS `项目编号ID`,
    COALESCE(NULLIF(h.xmmc, ''), NULLIF(w.xmmc, '')) AS `项目名称`,
    h.jbr AS `经办人ID`,
    h.gszt AS `公司主体ID`,
    h.cbzx AS `成本中心ID`,
    cc.bh AS `成本中心编号`,
    COALESCE(NULLIF(h.skfwb, ''), w.skfmc) AS `收款方文本`,
    h.bz AS `备注`,
    w.htmc AS `合同名称`,
    h.xght AS `相关合同ID`,
    h.yjfkrq AS `预计付款日期`,
    d.skf AS `实际收款方`,
    d.sjh AS `手机号`,
    d.sfzh AS `身份证号`,
    d.yhzh AS `银行账号`,
    d.je AS `税前应付金额`,
    d.sl AS `付给三方平台金额`,
    d.se AS `付款平台税额`,
    d.zzs AS `增值税`,
    d.fjs AS `附加税`,
    d.grsds AS `个人所得税`,
    d.ygdsje AS `预估到手金额`,
    d.gyszxx AS `库内外`,
    d.mainid AS `明细mainid`
FROM uf_lgptfk h
JOIN formtable_main_279 w
  ON CAST(w.requestId AS CHAR) = h.lczsjqqid
JOIN formtable_main_279_dt4 d
  ON d.mainid = w.id
LEFT JOIN uf_cbzx cc ON cc.id = h.cbzx
WHERE (
      COALESCE(NULLIF(CAST(h.xmbh AS CHAR), ''), NULLIF(CAST(w.xmbh AS CHAR), '')) IN %(project_ids)s
      OR (
          COALESCE(NULLIF(CAST(h.xmbh AS CHAR), ''), NULLIF(CAST(w.xmbh AS CHAR), '')) IS NULL
          AND h.sqrq >= %(project_fallback_date_from)s
          AND cc.bh IN %(project_codes)s
      )
  )
ORDER BY h.id, d.id
"""

GIG_BUDGET_SOURCE_SQL = """
SELECT
    h.id AS `建模付款ID`,
    w.id AS `流程付款ID`,
    b.id AS `预算明细ID`,
    h.lcbh AS `流程编号`,
    h.sqrq AS `申请日期`,
    COALESCE(NULLIF(CAST(h.xmbh AS CHAR), ''), NULLIF(CAST(w.xmbh AS CHAR), '')) AS `项目编号ID`,
    COALESCE(NULLIF(h.xmmc, ''), NULLIF(w.xmmc, '')) AS `项目名称`,
    cc.bh AS `成本中心编号`,
    b.yskm AS `预算科目ID`,
    b.fysx AS `费用事项`,
    b.fkje AS `预算项金额`,
    b.rmbje AS `预算项人民币金额`
FROM uf_lgptfk h
JOIN formtable_main_279 w
  ON CAST(w.requestId AS CHAR) = h.lczsjqqid
JOIN formtable_main_279_dt3 b
  ON b.mainid = w.id
LEFT JOIN uf_cbzx cc ON cc.id = h.cbzx
WHERE (
      COALESCE(NULLIF(CAST(h.xmbh AS CHAR), ''), NULLIF(CAST(w.xmbh AS CHAR), '')) IN %(project_ids)s
      OR (
          COALESCE(NULLIF(CAST(h.xmbh AS CHAR), ''), NULLIF(CAST(w.xmbh AS CHAR), '')) IS NULL
          AND h.sqrq >= %(project_fallback_date_from)s
          AND cc.bh IN %(project_codes)s
      )
  )
ORDER BY h.id, b.id
"""

GIG_STATS_SQL = """
SELECT
    COUNT(DISTINCT h.id) AS document_count,
    COUNT(*) AS detail_count,
    SUM(d.sl) AS amount_total
FROM uf_lgptfk h
JOIN formtable_main_279 w
  ON CAST(w.requestId AS CHAR) = h.lczsjqqid
JOIN formtable_main_279_dt4 d
  ON d.mainid = w.id
LEFT JOIN uf_cbzx cc ON cc.id = h.cbzx
WHERE (
      COALESCE(NULLIF(CAST(h.xmbh AS CHAR), ''), NULLIF(CAST(w.xmbh AS CHAR), '')) IN %(project_ids)s
      OR (
          COALESCE(NULLIF(CAST(h.xmbh AS CHAR), ''), NULLIF(CAST(w.xmbh AS CHAR), '')) IS NULL
          AND h.sqrq >= %(project_fallback_date_from)s
          AND cc.bh IN %(project_codes)s
      )
  )
"""

MCN_PREPAYMENT_SOURCE_SQL = """
SELECT
    'MCN预付款流程' AS `来源流程`,
    m.id AS `ID`,
    m.requestId AS `RequestID`,
    d.id AS `明细ID`,
    m.lcbh AS `流程编号`,
    rb.REQUESTNAME AS `标题`,
    m.sqrq AS `申请日期`,
    m.sqr AS `申请人ID`,
    m.gszt AS `公司主体ID`,
    COALESCE(m.cbzx1, m.cbzx) AS `成本中心ID`,
    cc.bh AS `成本中心编号`,
    m.gysmc AS `供应商ID`,
    m.yhzh AS `银行账号ID`,
    m.szht AS `主表合同ID`,
    d.szht AS `明细合同ID`,
    d.szxm AS `项目编号ID`,
    d.xmmc AS `项目名称`,
    d.fjxh AS `费用项编码`,
    d.fyx AS `费用项名称`,
    d.je AS `金额`,
    d.fjhzbid AS `主播房间号`,
    d.zbnc AS `主播昵称`,
    d.sfzh AS `身份证号`,
    d.skr AS `实际收款方`,
    d.sjh AS `手机号`,
    d.yhzh AS `明细银行账号`
FROM formtable_main_29 m
JOIN formtable_main_29_dt1 d ON d.mainid = m.id
LEFT JOIN workflow_requestbase rb ON rb.REQUESTID = m.requestId
LEFT JOIN uf_cbzx cc ON cc.id = COALESCE(m.cbzx1, m.cbzx)
WHERE (
      d.szxm IN %(project_ids)s
      OR (
          (d.szxm IS NULL OR CAST(d.szxm AS CHAR) = '')
          AND m.sqrq >= %(project_fallback_date_from)s
          AND cc.bh IN %(project_codes)s
      )
  )
ORDER BY m.id, d.id
"""

MCN_PREPAYMENT_ORDER_SOURCE_SQL = """
SELECT * FROM (
    SELECT
        'MCN预付款流程（订单）' AS `来源流程`,
        m.id AS `ID`,
        m.requestid AS `RequestID`,
        d.id AS `明细ID`,
        m.lcbh AS `流程编号`,
        rb.REQUESTNAME AS `标题`,
        m.sqrq AS `申请日期`,
        m.sqr AS `申请人ID`,
        m.gszt AS `公司主体ID`,
        m.cbzx AS `成本中心ID`,
        cc.bh AS `成本中心编号`,
        m.gysmc AS `供应商ID`,
        m.yhzh AS `银行账号ID`,
        m.szht AS `主表合同ID`,
        NULL AS `明细合同ID`,
        d.szxm AS `项目编号ID`,
        d.xmmc AS `项目名称`,
        d.fjxh AS `费用项编码`,
        d.fyx AS `费用项名称`,
        COALESCE(d.ddje, d.zcje, d.sdzc, d.jsje) AS `金额`,
        COALESCE(od.ddbh, d.ddbh) AS `泛微订单编号`,
        d.zbid AS `主播房间号`,
        d.zbnc AS `主播昵称`,
        d.sfzh AS `身份证号`,
        d.skr AS `实际收款方`,
        d.sjh AS `手机号`,
        d.yhzh AS `明细银行账号`
    FROM formtable_main_83 m
    JOIN formtable_main_83_dt3 d ON d.mainid = m.id
    LEFT JOIN uf_ddk od ON od.id = d.ddbh
    LEFT JOIN workflow_requestbase rb ON rb.REQUESTID = m.requestid
    LEFT JOIN uf_cbzx cc ON cc.id = m.cbzx
    UNION ALL
    SELECT
        'MCN预付款流程（订单）' AS `来源流程`,
        m.id AS `ID`,
        m.requestid AS `RequestID`,
        d.id AS `明细ID`,
        m.lcbh AS `流程编号`,
        rb.REQUESTNAME AS `标题`,
        m.sqrq AS `申请日期`,
        m.sqr AS `申请人ID`,
        m.gszt AS `公司主体ID`,
        m.cbzx AS `成本中心ID`,
        cc.bh AS `成本中心编号`,
        m.gysmc AS `供应商ID`,
        m.yhzh AS `银行账号ID`,
        m.szht AS `主表合同ID`,
        NULL AS `明细合同ID`,
        d.szxm AS `项目编号ID`,
        d.xmmc AS `项目名称`,
        d.fjxh AS `费用项编码`,
        d.fyx AS `费用项名称`,
        COALESCE(d.ddje, d.zcje, d.sdzc, d.jsje) AS `金额`,
        COALESCE(od.ddbh, d.ptpqh) AS `泛微订单编号`,
        d.zbid AS `主播房间号`,
        d.zbnc AS `主播昵称`,
        d.sfzh AS `身份证号`,
        d.skr AS `实际收款方`,
        d.sjh AS `手机号`,
        d.yhzh AS `明细银行账号`
    FROM formtable_main_83 m
    JOIN formtable_main_83_dt4 d ON d.mainid = m.id
    LEFT JOIN formtable_main_76_dt2 od ON od.id = d.ptpqh
    LEFT JOIN workflow_requestbase rb ON rb.REQUESTID = m.requestid
    LEFT JOIN uf_cbzx cc ON cc.id = m.cbzx
    UNION ALL
    SELECT
        'MCN预付款流程（订单）' AS `来源流程`,
        m.id AS `ID`,
        m.requestid AS `RequestID`,
        d.id AS `明细ID`,
        m.lcbh AS `流程编号`,
        rb.REQUESTNAME AS `标题`,
        m.sqrq AS `申请日期`,
        m.sqr AS `申请人ID`,
        m.gszt AS `公司主体ID`,
        m.cbzx AS `成本中心ID`,
        cc.bh AS `成本中心编号`,
        m.gysmc AS `供应商ID`,
        m.yhzh AS `银行账号ID`,
        m.szht AS `主表合同ID`,
        NULL AS `明细合同ID`,
        d.szxm AS `项目编号ID`,
        d.xmmc AS `项目名称`,
        d.fjxh AS `费用项编码`,
        d.fyx AS `费用项名称`,
        COALESCE(d.ddje, d.zcje, d.sdzc, d.jsje) AS `金额`,
        COALESCE(od.ddbh, d.ptpqh) AS `泛微订单编号`,
        d.zbid AS `主播房间号`,
        d.zbnc AS `主播昵称`,
        d.sfzh AS `身份证号`,
        d.skr AS `实际收款方`,
        d.sjh AS `手机号`,
        d.yhzh AS `明细银行账号`
    FROM formtable_main_83 m
    JOIN formtable_main_83_dt5 d ON d.mainid = m.id
    LEFT JOIN formtable_main_79_dt2 od ON od.id = d.ptpqh
    LEFT JOIN workflow_requestbase rb ON rb.REQUESTID = m.requestid
    LEFT JOIN uf_cbzx cc ON cc.id = m.cbzx
    UNION ALL
    SELECT
        'MCN预付款流程（订单）' AS `来源流程`,
        m.id AS `ID`,
        m.requestid AS `RequestID`,
        d.id AS `明细ID`,
        m.lcbh AS `流程编号`,
        rb.REQUESTNAME AS `标题`,
        m.sqrq AS `申请日期`,
        m.sqr AS `申请人ID`,
        m.gszt AS `公司主体ID`,
        m.cbzx AS `成本中心ID`,
        cc.bh AS `成本中心编号`,
        m.gysmc AS `供应商ID`,
        m.yhzh AS `银行账号ID`,
        m.szht AS `主表合同ID`,
        NULL AS `明细合同ID`,
        d.szxm AS `项目编号ID`,
        d.xmmc AS `项目名称`,
        d.fjxh AS `费用项编码`,
        d.fyx AS `费用项名称`,
        COALESCE(d.dkje, d.zcje, d.jsje) AS `金额`,
        COALESCE(od.ddbh, d.ddh) AS `泛微订单编号`,
        NULL AS `主播房间号`,
        NULL AS `主播昵称`,
        NULL AS `身份证号`,
        NULL AS `实际收款方`,
        NULL AS `手机号`,
        NULL AS `明细银行账号`
    FROM formtable_main_83 m
    JOIN formtable_main_83_dt6 d ON d.mainid = m.id
    LEFT JOIN formtable_main_76_dt2 od ON od.id = d.ddh
    LEFT JOIN workflow_requestbase rb ON rb.REQUESTID = m.requestid
    LEFT JOIN uf_cbzx cc ON cc.id = m.cbzx
) x
WHERE (
      x.`项目编号ID` IN %(project_ids)s
      OR (
          (x.`项目编号ID` IS NULL OR CAST(x.`项目编号ID` AS CHAR) = '')
          AND x.`申请日期` >= %(project_fallback_date_from)s
          AND x.`成本中心编号` IN %(project_codes)s
      )
  )
ORDER BY x.`ID`, x.`明细ID`
"""

MCN_ANCHOR_SOURCE_SELECTS = """
    SELECT
        'MCN主播相关付款流程' AS `来源流程`,
        m.id AS `ID`,
        m.requestId AS `RequestID`,
        d.id AS `明细ID`,
        m.lcbh AS `流程编号`,
        rb.REQUESTNAME AS `标题`,
        m.sqrq AS `申请日期`,
        m.sqr AS `申请人ID`,
        COALESCE(m.gszt, m.szgs) AS `公司主体ID`,
        COALESCE(d.cbzx, m.cbzx1, m.cbzxjs) AS `成本中心ID`,
        cc.bh AS `成本中心编号`,
        m.dfgsmc AS `供应商ID`,
        COALESCE(m.gysyhzh, m.yhzh) AS `银行账号ID`,
        m.szht AS `主表合同ID`,
        d.xght AS `明细合同ID`,
        d.xmbh AS `项目编号ID`,
        d.xmmc AS `项目名称`,
        d.fjxh AS `费用项编码`,
        d.fyx AS `费用项名称`,
        {amount_expr} AS `金额`,
        d.zbid AS `主播房间号`,
        d.zbnc AS `主播昵称`,
        d.sfzh AS `身份证号`,
        d.skr AS `实际收款方`,
        d.sjh AS `手机号`,
        d.yhzh AS `明细银行账号`,
        m.fkpt AS `付款平台ID`,
        m.sfyfk AS `是否预付ID`
    FROM formtable_main_38 m
    JOIN {detail_table} d ON d.mainid = m.id
    LEFT JOIN workflow_requestbase rb ON rb.REQUESTID = m.requestId
    LEFT JOIN uf_cbzx cc ON cc.id = COALESCE(d.cbzx, m.cbzx1, m.cbzxjs)
"""

MCN_ANCHOR_PREPAYMENT_SOURCE_SQL = """
SELECT * FROM (
""" + "\nUNION ALL\n".join([
    MCN_ANCHOR_SOURCE_SELECTS.format(detail_table='formtable_main_38_dt4', amount_expr='COALESCE(d.dkje, d.yfuje, d.skje, d.yfaje)'),
]) + """
) x
WHERE x.`是否预付ID` = %(prepayment_yes_code)s
  AND (
      x.`项目编号ID` IN %(project_ids)s
      OR (
          (x.`项目编号ID` IS NULL OR CAST(x.`项目编号ID` AS CHAR) = '')
          AND x.`申请日期` >= %(project_fallback_date_from)s
          AND x.`成本中心编号` IN %(project_codes)s
      )
  )
ORDER BY x.`ID`, x.`明细ID`
"""

MCN_ANCHOR_PLATFORM_SOURCE_SQL = """
SELECT * FROM (
""" + "\nUNION ALL\n".join([
    MCN_ANCHOR_SOURCE_SELECTS.format(detail_table='formtable_main_38_dt4', amount_expr='COALESCE(d.dkje, d.yfuje, d.skje, d.yfaje)'),
]) + """
) x
WHERE x.`是否预付ID` = %(prepayment_no_code)s
  AND (x.`付款平台ID` IS NULL OR CAST(x.`付款平台ID` AS CHAR) <> %(direct_payment_code)s)
  AND (
      x.`项目编号ID` IN %(project_ids)s
      OR (
          (x.`项目编号ID` IS NULL OR CAST(x.`项目编号ID` AS CHAR) = '')
          AND x.`申请日期` >= %(project_fallback_date_from)s
          AND x.`成本中心编号` IN %(project_codes)s
      )
  )
ORDER BY x.`ID`, x.`明细ID`
"""

MCN_OUTBOUND_FEE_SOURCE_SQL = """
SELECT
    'MCN对外付款流程' AS `来源流程`,
    m.id AS `ID`,
    m.requestId AS `RequestID`,
    d.id AS `明细ID`,
    m.lcbh AS `流程编号`,
    rb.REQUESTNAME AS `标题`,
    m.sqrq AS `申请日期`,
    m.sqr AS `申请人ID`,
    COALESCE(m.gszt, m.szgs) AS `公司主体ID`,
    COALESCE(m.cbzx2, m.cbzx) AS `成本中心ID`,
    cc.bh AS `成本中心编号`,
    m.dfgsmc AS `平台供应商ID`,
    m.yhzh AS `平台银行账号ID`,
    m.fkht AS `主表合同ID`,
    d.szht AS `明细合同ID`,
    d.szxm AS `项目编号ID`,
    d.xmmc AS `项目名称`,
    d.fjxh AS `费用项编码`,
    d.fyx AS `费用项名称`,
    d.je AS `金额`,
    d.fjhzbid AS `主播房间号`,
    d.zbnc AS `主播昵称`
FROM formtable_main_33 m
JOIN formtable_main_33_dt1 d ON d.mainid = m.id
LEFT JOIN workflow_requestbase rb ON rb.REQUESTID = m.requestId
LEFT JOIN uf_cbzx cc ON cc.id = COALESCE(m.cbzx2, m.cbzx)
WHERE (m.fkpt IS NULL OR CAST(m.fkpt AS CHAR) <> %(direct_payment_code)s)
  AND (
      d.szxm IN %(project_ids)s
      OR (
          (d.szxm IS NULL OR CAST(d.szxm AS CHAR) = '')
          AND m.sqrq >= %(project_fallback_date_from)s
          AND cc.bh IN %(project_codes)s
      )
  )
ORDER BY m.id, d.id
"""

MCN_OUTBOUND_RECIPIENT_SOURCE_SQL = """
SELECT
    m.id AS `ID`,
    d.id AS `收款人明细ID`,
    d.skr AS `实际收款方`,
    d.sfzh AS `身份证号`,
    d.yhzh AS `明细银行账号`,
    d.sjh AS `手机号`,
    d.skje AS `收款金额`,
    d.dkje AS `打款金额`
FROM formtable_main_33 m
JOIN formtable_main_33_dt2 d ON d.mainid = m.id
WHERE (m.fkpt IS NULL OR CAST(m.fkpt AS CHAR) <> %(direct_payment_code)s)
"""

MCN_OUTBOUND_ORDER_FEE_SOURCE_SQL = """
SELECT * FROM (
    SELECT
        'MCN对外付款流程（订单）' AS `来源流程`,
        m.id AS `ID`,
        m.requestid AS `RequestID`,
        d.id AS `明细ID`,
        m.lcbh AS `流程编号`,
        rb.REQUESTNAME AS `标题`,
        m.sqrq AS `申请日期`,
        m.sqr AS `申请人ID`,
        COALESCE(m.gszt, m.szgs) AS `公司主体ID`,
        m.cbzx AS `成本中心ID`,
        cc.bh AS `成本中心编号`,
        m.dfgsmc AS `平台供应商ID`,
        m.yhzh AS `平台银行账号ID`,
        m.fkht AS `主表合同ID`,
        NULL AS `明细合同ID`,
        d.szxm AS `项目编号ID`,
        d.xmmc AS `项目名称`,
        d.fjxh AS `费用项编码`,
        d.fyx AS `费用项名称`,
        d.ddje AS `金额`,
        COALESCE(od.ddbh, d.ddh) AS `泛微订单编号`,
        d.zbid AS `主播房间号`,
        d.zbnc AS `主播昵称`
    FROM formtable_main_66 m
    JOIN formtable_main_66_dt3 d ON d.mainid = m.id
    LEFT JOIN uf_ddk od ON od.id = d.ddh
    LEFT JOIN workflow_requestbase rb ON rb.REQUESTID = m.requestid
    LEFT JOIN uf_cbzx cc ON cc.id = m.cbzx
    WHERE (m.fkpt IS NULL OR CAST(m.fkpt AS CHAR) <> %(direct_payment_code)s)
    UNION ALL
    SELECT
        'MCN对外付款流程（订单）' AS `来源流程`,
        m.id AS `ID`,
        m.requestid AS `RequestID`,
        d.id AS `明细ID`,
        m.lcbh AS `流程编号`,
        rb.REQUESTNAME AS `标题`,
        m.sqrq AS `申请日期`,
        m.sqr AS `申请人ID`,
        COALESCE(m.gszt, m.szgs) AS `公司主体ID`,
        m.cbzx AS `成本中心ID`,
        cc.bh AS `成本中心编号`,
        m.dfgsmc AS `平台供应商ID`,
        m.yhzh AS `平台银行账号ID`,
        m.fkht AS `主表合同ID`,
        NULL AS `明细合同ID`,
        d.szxm AS `项目编号ID`,
        d.xmmc AS `项目名称`,
        d.fjxh AS `费用项编码`,
        d.fyx AS `费用项名称`,
        d.ddje AS `金额`,
        COALESCE(od.ddbh, d.ptpqh) AS `泛微订单编号`,
        d.zbid AS `主播房间号`,
        d.zbnc AS `主播昵称`
    FROM formtable_main_66 m
    JOIN formtable_main_66_dt4 d ON d.mainid = m.id
    LEFT JOIN formtable_main_76_dt2 od ON od.id = d.ptpqh
    LEFT JOIN workflow_requestbase rb ON rb.REQUESTID = m.requestid
    LEFT JOIN uf_cbzx cc ON cc.id = m.cbzx
    WHERE (m.fkpt IS NULL OR CAST(m.fkpt AS CHAR) <> %(direct_payment_code)s)
    UNION ALL
    SELECT
        'MCN对外付款流程（订单）' AS `来源流程`,
        m.id AS `ID`,
        m.requestid AS `RequestID`,
        d.id AS `明细ID`,
        m.lcbh AS `流程编号`,
        rb.REQUESTNAME AS `标题`,
        m.sqrq AS `申请日期`,
        m.sqr AS `申请人ID`,
        COALESCE(m.gszt, m.szgs) AS `公司主体ID`,
        m.cbzx AS `成本中心ID`,
        cc.bh AS `成本中心编号`,
        m.dfgsmc AS `平台供应商ID`,
        m.yhzh AS `平台银行账号ID`,
        m.fkht AS `主表合同ID`,
        NULL AS `明细合同ID`,
        d.szxm AS `项目编号ID`,
        d.xmmc AS `项目名称`,
        d.fjxh AS `费用项编码`,
        d.fyx AS `费用项名称`,
        d.ddje AS `金额`,
        COALESCE(od.ddbh, d.ptpqh) AS `泛微订单编号`,
        d.zbid AS `主播房间号`,
        d.zbmc AS `主播昵称`
    FROM formtable_main_66 m
    JOIN formtable_main_66_dt5 d ON d.mainid = m.id
    LEFT JOIN formtable_main_79_dt2 od ON od.id = d.ptpqh
    LEFT JOIN workflow_requestbase rb ON rb.REQUESTID = m.requestid
    LEFT JOIN uf_cbzx cc ON cc.id = m.cbzx
    WHERE (m.fkpt IS NULL OR CAST(m.fkpt AS CHAR) <> %(direct_payment_code)s)
    UNION ALL
    SELECT
        'MCN对外付款流程（订单）' AS `来源流程`,
        m.id AS `ID`,
        m.requestid AS `RequestID`,
        d.id AS `明细ID`,
        m.lcbh AS `流程编号`,
        rb.REQUESTNAME AS `标题`,
        m.sqrq AS `申请日期`,
        m.sqr AS `申请人ID`,
        COALESCE(m.gszt, m.szgs) AS `公司主体ID`,
        m.cbzx AS `成本中心ID`,
        cc.bh AS `成本中心编号`,
        m.dfgsmc AS `平台供应商ID`,
        m.yhzh AS `平台银行账号ID`,
        m.fkht AS `主表合同ID`,
        NULL AS `明细合同ID`,
        d.szxm AS `项目编号ID`,
        d.xmmc AS `项目名称`,
        d.fjxh AS `费用项编码`,
        d.fyx AS `费用项名称`,
        d.fkje AS `金额`,
        COALESCE(od.ddbh, d.ddh) AS `泛微订单编号`,
        NULL AS `主播房间号`,
        NULL AS `主播昵称`
    FROM formtable_main_66 m
    JOIN formtable_main_66_dt6 d ON d.mainid = m.id
    LEFT JOIN formtable_main_76_dt2 od ON od.id = d.ddh
    LEFT JOIN workflow_requestbase rb ON rb.REQUESTID = m.requestid
    LEFT JOIN uf_cbzx cc ON cc.id = m.cbzx
    WHERE (m.fkpt IS NULL OR CAST(m.fkpt AS CHAR) <> %(direct_payment_code)s)
) x
WHERE (
      x.`项目编号ID` IN %(project_ids)s
      OR (
          (x.`项目编号ID` IS NULL OR CAST(x.`项目编号ID` AS CHAR) = '')
          AND x.`申请日期` >= %(project_fallback_date_from)s
          AND x.`成本中心编号` IN %(project_codes)s
      )
  )
ORDER BY x.`ID`, x.`明细ID`
"""

MCN_OUTBOUND_ORDER_RECIPIENT_SOURCE_SQL = """
SELECT
    m.id AS `ID`,
    d.id AS `收款人明细ID`,
    d.skr AS `实际收款方`,
    d.sfzh AS `身份证号`,
    d.yhzh AS `明细银行账号`,
    d.sjh AS `手机号`,
    d.skje AS `收款金额`
FROM formtable_main_66 m
JOIN formtable_main_66_dt2 d ON d.mainid = m.id
WHERE (m.fkpt IS NULL OR CAST(m.fkpt AS CHAR) <> %(direct_payment_code)s)
"""

MCN_PREPAYMENT_OFFSET_SQL = """
SELECT
    m.yfklc AS `关联流程ID`,
    d.szxm AS `项目编号ID`,
    d.fjxh AS `费用项编码`,
    SUM(ABS(COALESCE(d.hsje, 0))) AS `已到票核销金额`
FROM formtable_main_30 m
JOIN formtable_main_30_dt1 d ON d.mainid = m.id
WHERE d.szxm IN %(project_ids)s
GROUP BY m.yfklc, d.szxm, d.fjxh
"""

MCN_PREPAYMENT_ORDER_OFFSET_SQL = """
SELECT `关联流程ID`, `项目编号ID`, `费用项编码`, SUM(`金额`) AS `已到票核销金额`
FROM (
    SELECT m.yfklc AS `关联流程ID`, d.szxm AS `项目编号ID`, d.fjxh AS `费用项编码`, ABS(COALESCE(d.ddje, 0)) AS `金额`
    FROM formtable_main_82 m JOIN formtable_main_82_dt2 d ON d.mainid = m.id
    UNION ALL
    SELECT m.yfklc AS `关联流程ID`, d.szxm AS `项目编号ID`, d.fjxh AS `费用项编码`, ABS(COALESCE(d.ddje, 0)) AS `金额`
    FROM formtable_main_82 m JOIN formtable_main_82_dt3 d ON d.mainid = m.id
    UNION ALL
    SELECT m.yfklc AS `关联流程ID`, d.szxm AS `项目编号ID`, d.fjxh AS `费用项编码`, ABS(COALESCE(d.ddje, 0)) AS `金额`
    FROM formtable_main_82 m JOIN formtable_main_82_dt4 d ON d.mainid = m.id
    UNION ALL
    SELECT m.yfklc AS `关联流程ID`, d.szxm AS `项目编号ID`, d.fjxh AS `费用项编码`, ABS(COALESCE(d.dkje, 0)) AS `金额`
    FROM formtable_main_82 m JOIN formtable_main_82_dt6 d ON d.mainid = m.id
) x
WHERE x.`项目编号ID` IN %(project_ids)s
GROUP BY `关联流程ID`, `项目编号ID`, `费用项编码`
"""

MCN_ANCHOR_OFFSET_SQL = """
SELECT
    m.jmyfklc AS `关联流程ID`,
    d.xmbh AS `项目编号ID`,
    d.fjxh AS `费用项编码`,
    SUM(ABS(COALESCE(d.dkje, d.yfuje, d.skje, d.yfaje, 0))) AS `已到票核销金额`
FROM formtable_main_105 m
JOIN formtable_main_105_dt4 d ON d.mainid = m.id
WHERE d.xmbh IN %(project_ids)s
GROUP BY m.jmyfklc, d.xmbh, d.fjxh
"""


# 运行前校验字段真实含义。主表字段 detail_table 用空字符串。
EXPECTED_SUPPLIER_FIELDS = {
    '': {
        'lcbh': '流程编号',
        'sqrq': '申请日期',
        'tdr': '填单人',
        'xmbh': '项目编号',
        'kpdw': '开票单位',
        'cbzx': '成本中心',
        'bz': '备注',
        'xght': '相关合同',
        'fkdx': '付款对象',
        'yhkh': '银行卡号',
        'fkbz': '付款币种',
        'fkje': '付款金额',
        'fkxz': '付款性质',
        'sycxtkje': '剩余冲销/退款金额',
        'lczt': '流程状态',
        'sfzf': '是否作废',
    },
    FW_SUPPLIER_DETAIL_TABLE: {
        'yfje': '预付金额',
        'yskm': '预算科目',
    },
}
EXPECTED_PAYMENT_OFFSET_FIELDS = {
    '': {
        'sqrq': '申请日期',
        'lcz': '流程状态',
        'sfzf': '是否作废',
    },
    FW_PAYMENT_OFFSET_DETAIL_TABLE: {
        'yfkxx': '预付款信息',
        'yskm': '预算科目',
        'cxje': '冲销金额',
    },
}
EXPECTED_GIG_HEADER_FIELDS = {
    '': {
        'lcbh': '流程编号',
        'sqrq': '申请日期',
        'lcfsrq': '流程发生日期',
        'lczsjqqid': '流程转数据请求ID',
        'xmbh': '项目编号',
        'xmmc': '项目名称',
        'jbr': '经办人',
        'gszt': '公司主体',
        'cbzx': '成本中心',
        'skfwb': '收款方-文本',
        'bz': '备注',
        'xght': '相关合同',
        'yjfkrq': '预计付款日期',
        'lczt': '流程状态',
        'sfzf': '是否作废',
    },
}
EXPECTED_GIG_WORKFLOW_FIELDS = {
    '': {
        'lcbh': '流程编号',
        'xmbh': '项目编号',
        'xmmc': '项目名称',
        'skfmc': '收款方名称',
        'htmc': '合同名称',
    },
    FW_GIG_BUDGET_TABLE: {
        'fysx': '费用事项',
        'yskm': '预算科目',
        'fkje': '付给三方平台金额',
        'rmbje': '付款人民币金额',
    },
    FW_GIG_RECIPIENT_TABLE: {
        'skf': '收款方',
        'sjh': '手机号',
        'sfzh': '身份证号',
        'yhzh': '银行账号',
        'je': '税前应付金额',
        'sl': '付给三方平台金额',
        'se': '付款平台税额',
        'zzs': '增值税',
        'fjs': '附加税',
        'grsds': '个人所得税',
        'ygdsje': '预估到手金额',
    },
}


# ============================ DB 查询小工具 ============================
def _query_fw(sql, params=None):
    return c.query_db('FW', 'vspn_xtyy', sql, params or {})


# ============================ 源值解析 ============================
def _text(value):
    if pd.isna(value):
        return ''
    text = str(value).strip()
    return '' if text in ('', 'nan', 'None', 'NaT') else text


def _fanwei_fee_code(raw_code):
    text = _text(raw_code)
    if '_' in text:
        text = text.rsplit('_', 1)[-1]
    return text


def _fanwei_fee_item_display(raw_value, raw_name=''):
    value = _text(raw_value)
    name = _text(raw_name)
    if name:
        code = _fanwei_fee_code(value)
        if code and not name.startswith(code):
            return f'{code}{name}'
        return name
    if '/' in value:
        return value.rsplit('/', 1)[-1]
    code = _fanwei_fee_code(value)
    fallback_name = load_mcn_fee_names().get(code, '')
    return f'{code}{fallback_name}' if code and fallback_name else code


def _id_number_key(value):
    return _text(value).replace(' ', '').upper()


def _lookup_first_browser_value(mapping, value):
    for item_id in c.parse_browser_ids(value):
        mapped = mapping.get(item_id, '')
        if mapped:
            return mapped
    return ''


def _first_non_blank(*values):
    for value in values:
        text = _text(value)
        if text:
            return text
    return ''


def _chunks(values, size=SQL_BATCH_SIZE):
    values = list(values)
    for start in range(0, len(values), size):
        yield values[start:start + size]


_PROJECT_FILTER_CACHE = None
_PROJECT_FILTER_ID_CACHE = {}
_MCN_FEE_NAME_CACHE = None
_MCN_FEE_SUBJECT_CACHE = None


def load_project_filter_codes():
    """读取本次清洗白名单:赛事/MCN sheet 的「原泛微项目编码」。"""
    global _PROJECT_FILTER_CACHE
    if _PROJECT_FILTER_CACHE is not None:
        return _PROJECT_FILTER_CACHE
    if not PROJECT_FILTER_FILE.exists():
        raise FileNotFoundError(f'项目白名单不存在: {PROJECT_FILTER_FILE}')

    result = {}
    for sheet_name in (EVENT_PROJECT_SHEET, MCN_PROJECT_SHEET):
        df = pd.read_excel(PROJECT_FILTER_FILE, sheet_name=sheet_name, dtype=str)
        code_column = '原泛微项目编码' if '原泛微项目编码' in df.columns else df.columns[0]
        codes = []
        for value in df[code_column]:
            codes.extend(c.split_fanwei_project_codes(value))
        result[sheet_name] = set(codes)
    print(
        '[预付期初-项目白名单] 使用:',
        PROJECT_FILTER_FILE,
        f"| 赛事 {len(result[EVENT_PROJECT_SHEET])} 个",
        f"| MCN {len(result[MCN_PROJECT_SHEET])} 个",
    )
    _PROJECT_FILTER_CACHE = result
    return result


def load_mcn_fee_names():
    global _MCN_FEE_NAME_CACHE
    if _MCN_FEE_NAME_CACHE is not None:
        return _MCN_FEE_NAME_CACHE

    df = pd.read_excel(c.RULE_XLSX, sheet_name='MCN泛微新旧科目映射底表', header=None, dtype=str)
    result = {}
    for _, row in df.iloc[2:].iterrows():
        code = _fanwei_fee_code(row[1] if 1 < len(row) else '')
        name = _text(row[2] if 2 < len(row) else '')
        if code and name:
            result[code] = name
    result.setdefault('JG', '外包费用')
    _MCN_FEE_NAME_CACHE = result
    return result


def load_mcn_fee_subjects():
    global _MCN_FEE_SUBJECT_CACHE
    if _MCN_FEE_SUBJECT_CACHE is not None:
        return _MCN_FEE_SUBJECT_CACHE

    df = pd.read_excel(c.RULE_XLSX, sheet_name='MCN泛微新旧科目映射底表', header=None, dtype=str)
    result = {}
    for _, row in df.iloc[2:].iterrows():
        old_code = _fanwei_fee_code(row[1] if 1 < len(row) else '')
        subject_code = _text(row[5] if 5 < len(row) else '')
        subject_name = _text(row[6] if 6 < len(row) else '')
        if old_code and subject_code:
            result[old_code] = (subject_code, subject_name)
    result.setdefault('JG', ('AB0103', '其他周边搭建'))
    _MCN_FEE_SUBJECT_CACHE = result
    return result


def project_filter_codes(sheet_name):
    return load_project_filter_codes().get(sheet_name, set())


def project_filter_ids(sheet_name, table_order):
    cache_key = (sheet_name, tuple(table_order))
    if cache_key in _PROJECT_FILTER_ID_CACHE:
        return _PROJECT_FILTER_ID_CACHE[cache_key]

    codes = sorted(project_filter_codes(sheet_name))
    result = []
    seen = set()
    for table in table_order:
        for batch in _chunks([code for code in codes if code]):
            try:
                project_df = c.query_db(
                    'FW',
                    'vspn_xtyy',
                    f'SELECT id, xmbh FROM {table} '
                    f'WHERE xmbh IN ({c.in_placeholders(batch)})',
                    batch,
                )
            except Exception:
                continue
            for _, row in project_df.iterrows():
                project_id = c.format_code(row['id'])
                if project_id and project_id not in seen:
                    seen.add(project_id)
                    result.append(project_id)
    _PROJECT_FILTER_ID_CACHE[cache_key] = tuple(result)
    print(f'[预付期初-项目白名单] {sheet_name} 反查项目ID {len(result)} 个')
    return _PROJECT_FILTER_ID_CACHE[cache_key]


def build_fw_project_info_map_for_ids(project_values, table_order=EVENT_PROJECT_TABLES):
    """泛微项目浏览框 ID -> 项目编号/名称,兼容赛事与 MCN 项目台账。"""
    project_ids = c.clean_codes(
        project_id
        for value in project_values
        for project_id in c.parse_browser_ids(value)
    )
    if not project_ids:
        return {}

    result = {}
    remaining = set(project_ids)
    for table in table_order:
        if not remaining:
            break
        ids = [project_id for project_id in project_ids if project_id in remaining]
        try:
            project_df = c.query_db(
                'FW',
                'vspn_xtyy',
                f'SELECT id, xmbh AS project_code, xmmc AS project_name FROM {table} '
                f'WHERE id IN ({c.in_placeholders(ids)})',
                ids,
            )
        except Exception:
            try:
                project_df = c.query_db(
                    'FW',
                    'vspn_xtyy',
                    f"SELECT id, xmbh AS project_code, '' AS project_name FROM {table} "
                    f'WHERE id IN ({c.in_placeholders(ids)})',
                    ids,
                )
            except Exception:
                continue
        for _, row in project_df.iterrows():
            project_id = c.format_code(row['id'])
            project_code = _text(row['project_code'])
            if project_id and project_code and project_id not in result:
                result[project_id] = {
                    'code': project_code,
                    'name': _text(row.get('project_name', '')),
                }
                remaining.discard(project_id)
    return result


def build_fw_project_code_map_for_ids(project_values, table_order=EVENT_PROJECT_TABLES):
    return {
        project_id: info.get('code', '')
        for project_id, info in build_fw_project_info_map_for_ids(project_values, table_order).items()
    }


def _resolve_project_codes(values, project_map):
    return values.map(lambda value: _lookup_first_browser_value(project_map, value) or _text(value))


def _with_resolved_project_fields(source_df, project_column='项目编号ID', table_order=EVENT_PROJECT_TABLES):
    df = source_df.copy()
    if project_column not in df.columns:
        df['项目编号'] = df['项目编号'] if '项目编号' in df.columns else ''
        df['项目名称'] = df['项目名称'] if '项目名称' in df.columns else ''
        return df

    project_info_map = build_fw_project_info_map_for_ids(df[project_column], table_order)
    project_code_map = {
        project_id: info.get('code', '')
        for project_id, info in project_info_map.items()
    }
    project_name_map = {
        project_id: info.get('name', '')
        for project_id, info in project_info_map.items()
    }
    mapped_project_codes = _resolve_project_codes(df[project_column], project_code_map)
    mapped_project_names = df[project_column].map(lambda value: _lookup_first_browser_value(project_name_map, value))
    existing_project_names = df['项目名称'] if '项目名称' in df.columns else pd.Series('', index=df.index)
    cost_center_codes = (
        df['成本中心编号'].map(_text)
        if '成本中心编号' in df.columns
        else pd.Series('', index=df.index)
    )
    df['项目编号'] = [
        project_code or cost_center_code
        for project_code, cost_center_code in zip(mapped_project_codes, cost_center_codes)
    ]
    df['项目名称'] = [
        _first_non_blank(existing_name, mapped_name)
        for existing_name, mapped_name in zip(existing_project_names, mapped_project_names)
    ]
    return df


def _filter_by_project_whitelist(source_df, project_column, sheet_name, table_order, log_label):
    if source_df.empty:
        return source_df
    allowed_codes = project_filter_codes(sheet_name)
    project_source_df = _with_resolved_project_fields(source_df, project_column, table_order)
    mask = project_source_df['项目编号'].map(_text).isin(allowed_codes)
    filtered = source_df.loc[mask].copy()
    print(f'[{log_label}] 项目白名单过滤: {len(filtered)}/{len(source_df)} 行')
    return filtered


def build_fw_bank_account_map_for_ids(bank_account_values):
    """供应商银行账号浏览框ID -> 银行账号文本。

    uf_yfkxx.yhkh 存 browser.gysyhxx 的明细 ID,对应 uf_khgys_dt1.id。
    """
    bank_account_ids = c.clean_codes(
        bank_account_id
        for value in bank_account_values
        for bank_account_id in c.parse_browser_ids(value)
    )
    if not bank_account_ids:
        return {}
    bank_df = c.query_db(
        'FW',
        'vspn_xtyy',
        'SELECT id, yhzh bank_account '
        'FROM uf_khgys_dt1 '
        f'WHERE id IN ({c.in_placeholders(bank_account_ids)})',
        bank_account_ids,
    )
    return {
        c.format_code(row['id']): _text(row['bank_account'])
        for _, row in bank_df.iterrows()
        if _text(row['bank_account'])
    }


def _selected_supplier_name(value, supplier_status_map):
    selected_id = c.choose_fw_supplier_id(c.parse_browser_ids(value), supplier_status_map)
    return supplier_status_map.get(selected_id, {}).get('name', '')


def resolve_supplier_source_values(source_df):
    """基于供应商预付 SQL 返回的泛微 ID 字段补充输出需要的展示值。"""
    df = source_df.copy()
    employee_map = c.build_fw_employee_info_map_for_ids(df['填单人ID'])
    company_map = c.build_fw_company_name_map_for_ids(df['开票单位ID'])
    currency_map = c.build_fw_currency_name_map_for_ids(df['付款币种ID'])
    cost_center_map = c.build_fw_cost_center_map_for_ids(df['成本中心ID'])
    subject_map = c.build_fw_budget_subject_path_map_for_ids(df['预算科目ID'])
    contract_map = c.build_fw_contract_code_map_for_ids(df['相关合同ID'])
    bank_account_map = build_fw_bank_account_map_for_ids(df['银行卡号ID'])
    supplier_status_map = c.build_fw_supplier_status_map(df['付款对象ID'])

    # [主表] tdr(填单人) -> hrmresource / hrmjobtitles
    df['填单人'] = df['填单人ID'].map(lambda value: employee_map.get(c.format_code(value), {}).get('name', ''))
    df['填单人工号'] = df['填单人ID'].map(lambda value: employee_map.get(c.format_code(value), {}).get('code', ''))
    # [主表] kpdw(开票单位) -> uf_gstt.gsmc(公司主体名称)
    df['开票单位'] = df['开票单位ID'].map(lambda value: company_map.get(c.format_code(value), ''))
    # [主表] fkbz(付款币种) -> fnacurrency.CURRENCYNAME(币种名称)
    df['付款币种'] = df['付款币种ID'].map(lambda value: currency_map.get(c.format_code(value), ''))
    # [主表] cbzx(成本中心) -> uf_cbzx.mc(成本中心名称)
    df['成本中心'] = df['成本中心ID'].map(lambda value: _lookup_first_browser_value(cost_center_map, value))
    # [明细表] yskm(预算科目) -> fnabudgetfeetype 层级路径
    df['预算科目'] = df['预算科目ID'].map(lambda value: subject_map.get(c.format_code(value), ''))
    # [主表] xght(相关合同) -> uf_htsp.htbh(合同编号)
    df['相关合同'] = df['相关合同ID'].map(lambda value: _lookup_first_browser_value(contract_map, value))
    # [主表] yhkh(银行卡号) -> uf_khgys_dt1.yhzh(银行账号)
    df['银行卡号'] = df['银行卡号ID'].map(lambda value: _lookup_first_browser_value(bank_account_map, value))
    # [主表] fkdx(付款对象) -> uf_khgys.khmc(供应商名称),仅用于描述兜底和异常清单。
    df['付款对象'] = df['付款对象ID'].map(lambda value: _selected_supplier_name(value, supplier_status_map))
    df = _with_resolved_project_fields(df, '项目编号ID')
    df['付款性质'] = df['付款性质ID'].map(
        lambda value: PAYMENT_NATURE_MEANINGS.get(int(c.format_code(value)), '') if c.format_code(value).isdigit() else '')
    return df


def resolve_gig_source_values(source_df):
    """基于零工 SQL 返回的泛微 ID 字段补充输出需要的展示值。"""
    df = source_df.copy()
    employee_map = c.build_fw_employee_info_map_for_ids(df['经办人ID'])
    company_map = c.build_fw_company_name_map_for_ids(df['公司主体ID'])
    cost_center_map = c.build_fw_cost_center_map_for_ids(df['成本中心ID'])
    contract_map = c.build_fw_contract_code_map_for_ids(df['相关合同ID'])

    # [建模头表] cbzx(成本中心) -> uf_cbzx.mc(成本中心名称)
    df['成本中心'] = df['成本中心ID'].map(lambda value: _lookup_first_browser_value(cost_center_map, value))
    # [建模头表] jbr(经办人) -> hrmresource / hrmjobtitles
    df['经办人'] = df['经办人ID'].map(lambda value: employee_map.get(c.format_code(value), {}).get('name', ''))
    df['经办人工号'] = df['经办人ID'].map(lambda value: employee_map.get(c.format_code(value), {}).get('code', ''))
    # [建模头表] gszt(公司主体) -> uf_gstt.gsmc(公司主体名称)
    df['公司主体'] = df['公司主体ID'].map(lambda value: company_map.get(c.format_code(value), ''))
    # 原流程表 htmc 为空时,用合同台账编号兜底。
    fallback_contract = df['相关合同ID'].map(lambda value: _lookup_first_browser_value(contract_map, value))
    df['合同名称'] = [
        name if _text(name) else fallback
        for name, fallback in zip(df['合同名称'], fallback_contract)
    ]
    df = _with_resolved_project_fields(df, '项目编号ID')
    return df


def resolve_gig_budget_values(budget_df):
    """基于零工原流程预算项明细补充预算科目路径。

    四合一 Excel 中「零工平台付款」的费用项目来源,对应原流程 formtable_main_279_dt3:
    - yskm(预算科目)
    - fkje(付给三方平台金额)
    """
    df = budget_df.copy()
    subject_map = c.build_fw_budget_subject_path_map_for_ids(df['预算科目ID'])
    df['预算科目'] = df['预算科目ID'].map(lambda value: subject_map.get(c.format_code(value), ''))
    return df


def merge_gig_budget_subjects(budget_df):
    """同一流程内先合并同类预算科目,保留预算科目首次出现顺序。"""
    if budget_df.empty:
        return budget_df.copy()

    df = budget_df.copy()
    df['_预算科目合并键'] = [
        c.format_code(subject_id) or c.remove_slashes(subject_path)
        for subject_id, subject_path in zip(df['预算科目ID'], df['预算科目'])
    ]

    merged_rows = []
    for _, group in df.groupby(['建模付款ID', '_预算科目合并键'], sort=False, dropna=False):
        row = group.iloc[0].copy()
        row['预算项金额'] = c.round_amount(
            pd.to_numeric(group['预算项金额'], errors='coerce').fillna(0).sum())
        if '预算项人民币金额' in group.columns:
            row['预算项人民币金额'] = c.round_amount(
                pd.to_numeric(group['预算项人民币金额'], errors='coerce').fillna(0).sum())
        row = row.drop(labels=['_预算科目合并键'], errors='ignore')
        merged_rows.append(row.to_dict())

    return pd.DataFrame(merged_rows)


def map_gig_budget_expense_items(budget_df):
    """把合并后的零工预算科目映射成新费用项目。"""
    if budget_df.empty:
        return budget_df.copy()

    df = budget_df.copy()
    subject_map = c.build_subject_map()
    mapped_items = df['预算科目'].map(
        lambda value: subject_map.get(c.remove_slashes(value), ('', '')))
    df['费用项目编码'] = mapped_items.map(lambda item: item[0])
    df['费用项目描述'] = mapped_items.map(lambda item: item[1])
    df['泛微费用项目编码'] = df['预算科目'].where(df['预算科目'].notna(), '')
    return df


def allocate_gig_budget_to_recipients(recipient_df, budget_df):
    """把原流程预算项分配到实际收款人明细。

    规则表「预付期初」R47 要求:从「对公&报销&零工&批量四合一」取预算科目并分配到
    「零工平台付款_实际收款人明细」。在 DB 中,四合一里「零工平台付款」对应
    formtable_main_279_dt3(预算项明细),实际收款人对应 formtable_main_279_dt4。

    分配口径:
    - 同一流程内先按 dt3 预算科目合并金额,保留该预算科目的首次出现顺序。
    - 合并后的预算科目再映射新费用项目,并以映射后的项目形成预算金额池。
    - 按 dt4 实际收款人明细顺序依次占用预算池; 收款人金额跨预算科目边界时拆成多行。
    - 预算项金额用完后才进入下一预算项; 不再按预算项比例拆分每个收款人。
    """
    rows = []
    raw_budget_count = len(budget_df)
    budget_df = map_gig_budget_expense_items(merge_gig_budget_subjects(budget_df))
    budget_by_payment = {
        key: group.copy()
        for key, group in budget_df.groupby('建模付款ID', dropna=False)
    }
    no_budget_count = 0
    multi_budget_payment_count = 0

    for payment_id, recipient_group in recipient_df.groupby('建模付款ID', sort=False, dropna=False):
        budget_group = budget_by_payment.get(payment_id)
        if budget_group is None or budget_group.empty:
            no_budget_count += len(recipient_group)
            for _, recipient_row in recipient_group.iterrows():
                row = recipient_row.to_dict()
                row.update({
                    '预算明细ID': '',
                    '预算科目ID': '',
                    '预算科目': '',
                    '费用项目编码': '',
                    '费用项目描述': '',
                    '泛微费用项目编码': '',
                    '预算项金额': '',
                    '预付款金额分摊': pd.to_numeric(recipient_row['付给三方平台金额'], errors='coerce'),
                })
                rows.append(row)
            continue

        budget_group = budget_group.copy()
        budget_group['_剩余预算项金额'] = (
            pd.to_numeric(budget_group['预算项金额'], errors='coerce')
            .fillna(0)
            .map(c.round_amount)
        )
        budget_total = float(budget_group['_剩余预算项金额'].sum())
        if len(budget_group) > 1:
            multi_budget_payment_count += 1

        if budget_total == 0:
            # 没有可用预算项金额时,保留收款人粒度并带第一条预算科目。
            first_budget = budget_group.iloc[0]
            for _, recipient_row in recipient_group.iterrows():
                row = recipient_row.to_dict()
                row.update({
                    '预算明细ID': first_budget.get('预算明细ID', ''),
                    '预算科目ID': first_budget.get('预算科目ID', ''),
                    '预算科目': first_budget.get('预算科目', ''),
                    '费用项目编码': first_budget.get('费用项目编码', ''),
                    '费用项目描述': first_budget.get('费用项目描述', ''),
                    '泛微费用项目编码': first_budget.get('泛微费用项目编码', ''),
                    '预算项金额': first_budget.get('预算项金额', ''),
                    '预付款金额分摊': pd.to_numeric(recipient_row['付给三方平台金额'], errors='coerce'),
                })
                rows.append(row)
            continue

        budget_records = [row.to_dict() for _, row in budget_group.iterrows()]
        budget_index = 0

        def append_allocated_row(recipient_row, budget_row, allocated_amount):
            row = recipient_row.to_dict()
            row.update({
                '预算明细ID': budget_row.get('预算明细ID', ''),
                '预算科目ID': budget_row.get('预算科目ID', ''),
                '预算科目': budget_row.get('预算科目', ''),
                '费用项目编码': budget_row.get('费用项目编码', ''),
                '费用项目描述': budget_row.get('费用项目描述', ''),
                '泛微费用项目编码': budget_row.get('泛微费用项目编码', ''),
                '预算项金额': budget_row.get('预算项金额', ''),
                '预付款金额分摊': allocated_amount,
            })
            rows.append(row)

        for _, recipient_row in recipient_group.iterrows():
            recipient_amount = pd.to_numeric(recipient_row['付给三方平台金额'], errors='coerce')
            current_budget_index = min(budget_index, len(budget_records) - 1)
            if pd.isna(recipient_amount):
                append_allocated_row(recipient_row, budget_records[current_budget_index], recipient_amount)
                continue

            remaining_recipient_amount = c.round_amount(recipient_amount)
            if remaining_recipient_amount <= 0:
                append_allocated_row(recipient_row, budget_records[current_budget_index], remaining_recipient_amount)
                continue

            while remaining_recipient_amount > 0:
                if budget_index >= len(budget_records):
                    append_allocated_row(recipient_row, budget_records[-1], remaining_recipient_amount)
                    break

                budget_row = budget_records[budget_index]
                remaining_budget_amount = budget_row['_剩余预算项金额']
                if remaining_budget_amount <= 0:
                    budget_index += 1
                    continue

                allocated_amount = c.round_amount(min(remaining_recipient_amount, remaining_budget_amount))
                append_allocated_row(recipient_row, budget_row, allocated_amount)

                remaining_recipient_amount = c.round_amount(remaining_recipient_amount - allocated_amount)
                budget_row['_剩余预算项金额'] = c.round_amount(remaining_budget_amount - allocated_amount)
                if budget_row['_剩余预算项金额'] <= 0:
                    budget_index += 1

    allocated_df = pd.DataFrame(rows)
    print('[预付期初-零工预付款-DB] 未匹配预算项的收款人明细行数:', no_budget_count)
    print('[预付期初-零工预付款-DB] 预算项按预算科目合并:', raw_budget_count, '->', len(budget_df))
    print('[预付期初-零工预付款-DB] 多预算科目顺序分配的单据数:', multi_budget_payment_count)
    print('[预付期初-零工预付款-DB] 预算项分配后明细行数:', len(allocated_df))
    return allocated_df


# ============================ 模板输出:供应商预付款 ============================
def allocate_supplier_written_off_amounts(merged_df):
    """按预付单+预算科目把对公付款冲销金额分摊到预付预算明细行。"""
    detail_amount = pd.to_numeric(merged_df['预付金额'], errors='coerce').fillna(0).map(c.round_amount)
    group_written_off = pd.to_numeric(
        merged_df['冲销金额（支付币种-同预付单预算科目）'], errors='coerce').fillna(0)
    settled_amount = pd.Series(0.0, index=merged_df.index, dtype='float64')

    for _, group in merged_df.groupby(['ID', '预算科目ID'], sort=False, dropna=False):
        indexes = list(group.index)
        if not indexes:
            continue

        written_off_amount = c.round_amount(group_written_off.loc[indexes].iloc[0])
        if len(indexes) == 1:
            settled_amount.loc[indexes[0]] = written_off_amount
            continue

        group_detail_amount = detail_amount.loc[indexes]
        detail_total = c.round_amount(group_detail_amount.sum())
        if detail_total != 0:
            allocations = [
                c.round_amount(written_off_amount * amount / detail_total)
                for amount in group_detail_amount
            ]
        else:
            per_row_amount = c.round_amount(written_off_amount / len(indexes))
            allocations = [per_row_amount] * len(indexes)

        diff = c.round_amount(written_off_amount - sum(allocations))
        allocations[-1] = c.round_amount(allocations[-1] + diff)
        settled_amount.loc[indexes] = allocations

    unsettled_amount = (detail_amount - settled_amount).map(c.round_amount)
    return detail_amount, settled_amount, unsettled_amount


def build_supplier_output(merged_df):
    """供应商预付 DB 源数据 -> 供应商预付款单导入模版 26 列。"""
    def lookup_by_name(mapping, value):
        return '' if pd.isna(value) else mapping.get(c.normalize_name(value), '')

    vendor_map = c.build_supplier_vendor_info_map_for_rows(
        merged_df['付款对象ID'],
        supplier_texts=merged_df['付款对象'],
        document_numbers=merged_df['流程编号'],
        missing_report_file=SUPPLIER_VENDOR_MISSING_FILE,
        log_prefix='[预付期初-供应商预付款-DB]',
    )
    entity_map = c.build_accounting_entity_map_for_names(merged_df['开票单位'])
    subject_map = c.build_subject_map()

    def lookup_vendor(index, field):
        return vendor_map.get(index, {}).get(field, '')

    def vendor_description(index, fallback_text):
        name = lookup_vendor(index, 'name')
        if name:
            return name
        return _text(fallback_text)

    def subject_item(subject_path, index):
        if pd.isna(subject_path):
            return ''
        return subject_map.get(c.remove_slashes(subject_path), ('', ''))[index]

    detail_amount, settled_amount, unsettled_amount = allocate_supplier_written_off_amounts(merged_df)
    is_deposit = merged_df['付款性质ID'].map(c.format_code).isin(DEPOSIT_PAYMENT_NATURE_CODES)

    output_df = pd.DataFrame(index=merged_df.index)

    # 固定值。
    output_df['来源系统'] = 'FW'
    output_df['单据类型'] = DOCUMENT_TYPE

    # 泛微主表/明细表直取或解析字段。
    output_df['来源单据编号'] = merged_df['流程编号']                    # [主表] lcbh(流程编号)
    output_df['申请日期'] = merged_df['申请日期'].map(c.format_date)     # [主表] sqrq(申请日期)
    output_df['申请人工号'] = merged_df['填单人工号']                    # [主表] tdr(填单人) -> hrmjobtitles.JOBTITLENAME(工号)
    output_df['申请人姓名'] = merged_df['填单人']                        # [主表] tdr(填单人) -> hrmresource.LASTNAME(姓名)
    output_df['备注'] = merged_df['备注'].astype(str).where(
        merged_df['备注'].notna(), '').str.slice(0, 150)                # [主表] bz(备注),导入限制截前150字符
    output_df['合同号'] = merged_df['相关合同'].where(
        merged_df['相关合同'].notna(), '')                              # [主表] xght(相关合同) -> uf_htsp.htbh(合同编号)
    output_df['保证金标志'] = ['是' if flag else '否' for flag in is_deposit]  # [主表] fkxz(付款性质) 0=押金/1=质保金 -> 是
    output_df['银行账号'] = merged_df['银行卡号'].where(
        merged_df['银行卡号'].notna(), '')                              # [主表] yhkh(银行卡号) -> uf_khgys_dt1.yhzh(银行账号)

    # 项目&订单清洗结果:泛微项目编号 -> 订单编号/订单标题。
    output_df['泛微项目编号'] = merged_df['项目编号'].map(_text)
    output_df['订单编号'] = merged_df['项目编号'].map(lambda value: c.project_order_mapping_value(value, '订单编号'))
    output_df['订单名称'] = merged_df['项目编号'].map(lambda value: c.project_order_mapping_value(value, '订单标题'))

    # 当前源数据没有直接可用的计划行/付款计划日期/银行备注/主播房间号,按模板留空。
    output_df['合同收支计划行'] = ''
    output_df['计划付款日期'] = ''
    output_df['银行转账备注'] = ''
    output_df['主播房间号'] = ''

    # 跨系统映射字段。
    output_df['核算主体编号'] = merged_df['开票单位'].map(
        lambda value: lookup_by_name(entity_map, value))                # [主表] kpdw(开票单位) -> Hand 核算主体编号
    output_df['核算主体描述'] = merged_df['开票单位']                    # [主表] kpdw(开票单位) -> uf_gstt.gsmc(公司主体名称)
    output_df['收款方编码'] = [lookup_vendor(index, 'code') for index in merged_df.index]  # [主表] fkdx(付款对象) -> Hand vender_code(供应商编码)
    output_df['收款方描述'] = [
        vendor_description(index, supplier_text)
        for index, supplier_text in zip(merged_df.index, merged_df['付款对象'])
    ]                                                                    # 优先 Hand description(供应商描述),兜底 [主表] fkdx(付款对象) -> uf_khgys.khmc(供应商名称)
    output_df['银行账号'] = c.resolve_hand_vendor_bank_accounts(
        output_df['收款方编码'], merged_df['银行卡号'])                  # 按收款方 Hand 供应商银行卡校验;为空/不匹配时取默认账号
    output_df['费用项目编码'] = merged_df['预算科目'].map(
        lambda value: subject_item(value, 0))                            # [明细] yskm(预算科目) -> 规则表费用项目编码
    output_df['费用项目描述'] = merged_df['预算科目'].map(
        lambda value: subject_item(value, 1))                            # [明细] yskm(预算科目) -> 规则表费用项目描述
    output_df['预付款支付币种'] = merged_df['付款币种'].map(c.to_iso_currency)  # [主表] fkbz(付款币种) -> ISO币种

    # 金额拆分。
    output_df['预付款金额（支付币种）'] = detail_amount.map(c.round_amount)       # [明细] yfje(预付金额)
    output_df['已到票核销金额（支付币种）'] = settled_amount.map(c.round_amount)  # [对公付款冲销] uf_dgfktz_dt2.cxje 转正后按预付单+预算科目汇总/分摊
    output_df['已付未核（支付币种）'] = unsettled_amount.map(c.round_amount)     # [明细] yfje(预付金额) - 已到票核销金额
    output_df['泛微费用项目编码'] = merged_df['预算科目'].where(
        merged_df['预算科目'].notna(), '')                               # [明细] yskm -> 原泛微预算科目路径
    return output_df[SUPPLIER_OUTPUT_COLUMNS]


# ============================ 模板输出:零工预付款 ============================
def gig_platform_vendor(name):
    """收款方文本 -> 灵工平台收款方编码。"""
    text = _text(name)
    for keyword, code in GIG_PLATFORM_VENDOR.items():
        if keyword in text:
            return code
    return ''


def gig_recipient_remark(name, id_number, phone):
    """收款人备注:姓名-身份证-手机号 拼接,限 30 字。"""
    parts = [_text(value) for value in (name, id_number, phone) if _text(value)]
    return '-'.join(parts)[:30]


def build_gig_output(merged_df):
    """零工 DB 源数据 -> 灵工预付款单导入模版 29 列。"""
    vendor_map = c.build_vendor_map()
    entity_map = c.build_accounting_entity_map_for_names(merged_df['公司主体'])
    has_mapped_expense_items = {'费用项目编码', '费用项目描述'}.issubset(merged_df.columns)
    subject_map = {} if has_mapped_expense_items else c.build_subject_map()

    def gig_payee_code(value):
        payee = _text(value)
        if not payee:
            return ''
        return vendor_map.get(c.normalize_name(payee)) or payee

    def subject_item(subject_path, index):
        if pd.isna(subject_path):
            return ''
        return subject_map.get(c.remove_slashes(subject_path), ('', ''))[index]

    output_df = pd.DataFrame(index=merged_df.index)

    # 固定值。
    output_df['来源系统'] = 'FW'
    output_df['单据类型'] = GIG_DOCUMENT_TYPE
    output_df['保证金标志'] = ''
    output_df['收款方类别'] = '供应商'
    output_df['预付款支付币种'] = 'CNY'
    output_df['传送状态'] = '传送成功'
    output_df['支付状态'] = '支付成功'
    output_df['退款状态'] = ''
    output_df['核销状态'] = '已核销'

    # 泛微头表/原流程明细字段。
    output_df['来源单据编号'] = merged_df['流程编号']                    # [建模头/原流程] lcbh(流程编号)
    output_df['申请日期'] = merged_df['申请日期'].map(c.format_date)     # [建模头] sqrq(申请日期)
    output_df['申请人工号'] = merged_df['经办人工号']                    # [建模头] jbr(经办人) -> hrmjobtitles.JOBTITLENAME(工号)
    output_df['申请人姓名'] = merged_df['经办人']                        # [建模头] jbr(经办人) -> hrmresource.LASTNAME(姓名)
    output_df['备注_单头'] = merged_df['备注'].astype(str).where(
        merged_df['备注'].notna(), '').str.slice(0, 150)                # [建模头] bz(备注)
    output_df['灵工平台收款方编码'] = merged_df['收款方文本'].map(gig_platform_vendor)  # [建模头] skfwb(收款方-文本)
    output_df['合同号'] = merged_df['合同名称'].where(
        merged_df['合同名称'].notna(), '')                              # [原流程] htmc(合同名称);为空时合同编号兜底
    output_df['计划付款日期'] = merged_df['预计付款日期'].map(c.format_date)  # [建模头] yjfkrq(预计付款日期)
    output_df['收款方编码'] = merged_df['实际收款方'].map(gig_payee_code)  # [原流程明细] dt4.skf(收款方)
    output_df['银行账号'] = c.resolve_hand_vendor_bank_accounts(
        output_df['收款方编码'], merged_df['银行账号'])                  # 按收款方 Hand 供应商银行卡校验;为空/不匹配时取默认账号
    output_df['备注'] = [
        gig_recipient_remark(name, id_number, phone)
        for name, id_number, phone in zip(merged_df['实际收款方'], merged_df['身份证号'], merged_df['手机号'])
    ]                                                                    # 姓名-身份证-手机号
    output_df['预付款金额（支付币种）'] = pd.to_numeric(
        merged_df['预付款金额分摊'], errors='coerce').map(c.round_amount)  # [原流程明细] dt4.sl(付给三方平台金额) 按合并预算科目顺序占用预算项金额

    # 项目&订单清洗结果:泛微项目编号 -> 订单编号/订单标题。
    output_df['泛微项目编号'] = merged_df['项目编号'].map(_text)
    output_df['订单编号'] = merged_df['项目编号'].map(lambda value: c.project_order_mapping_value(value, '订单编号'))
    output_df['订单名称'] = merged_df['项目编号'].map(lambda value: c.project_order_mapping_value(value, '订单标题'))
    output_df['合同收支计划行'] = ''
    output_df['银行转账备注'] = ''
    if has_mapped_expense_items:
        output_df['费用项目编码'] = merged_df['费用项目编码'].where(
            merged_df['费用项目编码'].notna(), '')                        # [合并预算科目] 已映射的新费用项目编码
        output_df['费用项目描述'] = merged_df['费用项目描述'].where(
            merged_df['费用项目描述'].notna(), '')                        # [合并预算科目] 已映射的新费用项目描述
    else:
        output_df['费用项目编码'] = merged_df['预算科目'].map(
            lambda value: subject_item(value, 0))                        # [原流程预算项明细] dt3.yskm(预算科目) -> 规则表费用项目编码
        output_df['费用项目描述'] = merged_df['预算科目'].map(
            lambda value: subject_item(value, 1))                        # [原流程预算项明细] dt3.yskm(预算科目) -> 规则表费用项目描述
    if '泛微费用项目编码' in merged_df.columns:
        output_df['泛微费用项目编码'] = merged_df['泛微费用项目编码'].where(
            merged_df['泛微费用项目编码'].notna(), '')
    else:
        output_df['泛微费用项目编码'] = merged_df['预算科目'].where(merged_df['预算科目'].notna(), '')

    # 跨系统映射字段。
    output_df['核算主体编号'] = merged_df['公司主体'].map(
        lambda value: entity_map.get(c.normalize_name(value), '') if _text(value) else '')  # [建模头] gszt(公司主体) -> Hand 核算主体编号
    output_df['核算主体描述'] = merged_df['公司主体']                    # [建模头] gszt(公司主体) -> uf_gstt.gsmc(公司主体名称)
    return output_df[GIG_OUTPUT_COLUMNS]


# ============================ MCN 通用解析/输出 ============================
def _resolve_company_name(company_map, value):
    return _lookup_first_browser_value(company_map, value) or company_map.get(c.format_code(value), '')


def _mcn_contract_code_map(contract_values):
    contract_ids = c.clean_codes(
        contract_id
        for value in contract_values
        for contract_id in c.parse_browser_ids(value)
    )
    if not contract_ids:
        return {}

    result = {}
    for table, code_col in (('uf_htk', 'htbh'), ('uf_htsp', 'htbh')):
        remaining = [contract_id for contract_id in contract_ids if contract_id not in result]
        if not remaining:
            break
        for batch in _chunks(remaining):
            try:
                contract_df = c.query_db(
                    'FW',
                    'vspn_xtyy',
                    f'SELECT id, {code_col} AS contract_code FROM {table} '
                    f'WHERE id IN ({c.in_placeholders(batch)})',
                    batch,
                )
            except Exception:
                continue
            for _, row in contract_df.iterrows():
                contract_id = c.format_code(row['id'])
                contract_code = _text(row['contract_code'])
                if contract_id and contract_code and contract_id not in result:
                    result[contract_id] = contract_code
    return result


def _mcn_fee_subject_key(raw_code, subject_lookup):
    code = _text(raw_code)
    if not code:
        return ''
    deepest = code.split('_', 1)[-1].strip()
    if not deepest:
        return c.remove_slashes(code)
    candidates = [c.remove_slashes(code), deepest]
    if len(deepest) >= 4:
        candidates.append(f'{deepest[:2]}{deepest[:3]}{deepest}')
    if len(deepest) >= 3:
        candidates.append(f'{deepest[:2]}{deepest}')
    return next((key for key in candidates if key in subject_lookup), c.remove_slashes(code))


def _mcn_fee_subject_item(subject_lookup, raw_code, index):
    key = _mcn_fee_subject_key(raw_code, subject_lookup)
    if not key:
        return ''
    item = subject_lookup.get(key)
    if item:
        return item[index]
    return load_mcn_fee_subjects().get(_fanwei_fee_code(raw_code), ('', ''))[index]


def build_hand_anchor_room_map(id_numbers):
    keys = c.clean_text_values(_id_number_key(value) for value in id_numbers)
    if not keys:
        return {}

    rows = []
    for batch in _chunks(keys):
        batch_rows = c.query_db(
            'ZT',
            'hfins',
            f'''
            SELECT certificate_number, anchor_id
            FROM (
                SELECT p.certificate_number, l.anchor_id
                FROM anchor_profile_list p
                JOIN anchor_platform_list_line l ON l.header_id = p.header_id
                WHERE REPLACE(UPPER(p.certificate_number), ' ', '') IN ({c.in_placeholders(batch)})
                  AND COALESCE(p.enable, 'Y') <> 'N'
                  AND COALESCE(l.anchor_id, '') <> ''
                UNION ALL
                SELECT p.certificate_number, l.anchor_id
                FROM anchor_profile_header p
                JOIN anchor_platform_line l ON l.header_id = p.header_id
                WHERE REPLACE(UPPER(p.certificate_number), ' ', '') IN ({c.in_placeholders(batch)})
                  AND COALESCE(p.delete_flag, 'N') <> 'Y'
                  AND COALESCE(l.anchor_id, '') <> ''
            ) x
            ''',
            list(batch) + list(batch),
        )
        rows.append(batch_rows)

    if not rows:
        return {}
    result = {}
    room_df = pd.concat(rows, ignore_index=True)
    for _, row in room_df.iterrows():
        key = _id_number_key(row['certificate_number'])
        room = _text(row['anchor_id'])
        if not key or not room:
            continue
        result.setdefault(key, [])
        if room not in result[key]:
            result[key].append(room)
    return {key: ';'.join(rooms) for key, rooms in result.items()}


def build_fw_anchor_room_map(room_values):
    room_ids = c.clean_codes(
        room_id
        for value in room_values
        for room_id in c.parse_browser_ids(value)
    )
    if not room_ids:
        return {}

    result = {}
    for batch in _chunks(room_ids):
        room_df = c.query_db(
            'FW',
            'vspn_xtyy',
            f'''
            SELECT id, zbid
            FROM uf_zbkp_dt1
            WHERE id IN ({c.in_placeholders(batch)})
              AND COALESCE(zbid, '') <> ''
            ''',
            batch,
        )
        for _, row in room_df.iterrows():
            room_id = c.format_code(row['id'])
            room_no = _text(row['zbid'])
            if room_id and room_no:
                result[room_id] = room_no
    return result


def _resolve_fw_anchor_room(value, room_map):
    mapped = _lookup_first_browser_value(room_map, value)
    return mapped or _text(value)


def resolve_anchor_room_numbers(source_df):
    hand_room_map = build_hand_anchor_room_map(source_df.get('身份证号', pd.Series(dtype=str)))
    fw_room_map = build_fw_anchor_room_map(source_df.get('主播房间号', pd.Series(dtype=str)))
    hand_hits = 0
    result = []
    for _, row in source_df.iterrows():
        hand_room = hand_room_map.get(_id_number_key(row.get('身份证号')))
        if hand_room:
            hand_hits += 1
            result.append(hand_room)
        else:
            result.append(_resolve_fw_anchor_room(row.get('主播房间号'), fw_room_map))
    print(f'[预付期初-MCN主播相关预付款] 主播房间号: 汉得身份证命中 {hand_hits}/{len(source_df)} 行')
    return result


def resolve_mcn_common_values(source_df):
    df = source_df.copy()
    if df.empty:
        return df

    employee_map = c.build_fw_employee_info_map_for_ids(df['申请人ID'])
    company_map = c.build_fw_company_name_map_for_ids(df['公司主体ID'])
    cost_center_map = c.build_fw_cost_center_map_for_ids(df['成本中心ID'])
    supplier_status_map = c.build_fw_supplier_status_map(df['供应商ID'])
    bank_account_map = c.build_fw_supplier_bank_account_map_for_ids(df['银行账号ID'])
    contract_values = pd.concat([df['主表合同ID'], df['明细合同ID']], ignore_index=True)
    contract_map = _mcn_contract_code_map(contract_values)

    project_df = _with_resolved_project_fields(df, '项目编号ID', MCN_PROJECT_TABLES)
    df['项目编号'] = project_df['项目编号']
    df['项目名称'] = [
        _first_non_blank(existing_name, mapped_name)
        for existing_name, mapped_name in zip(df['项目名称'], project_df['项目名称'])
    ]
    df['申请人'] = df['申请人ID'].map(lambda value: employee_map.get(c.format_code(value), {}).get('name', ''))
    df['申请人工号'] = df['申请人ID'].map(lambda value: employee_map.get(c.format_code(value), {}).get('code', ''))
    df['公司主体'] = df['公司主体ID'].map(lambda value: _resolve_company_name(company_map, value))
    df['成本中心'] = df['成本中心ID'].map(lambda value: _lookup_first_browser_value(cost_center_map, value))
    df['供应商'] = df['供应商ID'].map(lambda value: _selected_supplier_name(value, supplier_status_map))
    df['银行账号'] = df['银行账号ID'].map(lambda value: _lookup_first_browser_value(bank_account_map, value))
    df['合同号'] = [
        _first_non_blank(
            _lookup_first_browser_value(contract_map, detail_contract),
            _lookup_first_browser_value(contract_map, main_contract),
        )
        for main_contract, detail_contract in zip(df['主表合同ID'], df['明细合同ID'])
    ]
    df['预算科目'] = df['费用项编码'].map(_text)
    return df


def _offset_key(flow_id, project_id, fee_code):
    return (c.format_code(flow_id), c.format_code(project_id), _text(fee_code))


def apply_mcn_offset_amounts(source_df, offset_df, link_column):
    df = source_df.copy()
    if df.empty:
        return df

    offset_map = {}
    if not offset_df.empty:
        for _, row in offset_df.iterrows():
            key = _offset_key(row['关联流程ID'], row['项目编号ID'], row['费用项编码'])
            offset_map[key] = c.round_amount(offset_map.get(key, 0) + float(row['已到票核销金额'] or 0))

    amount = pd.to_numeric(df['金额'], errors='coerce').fillna(0).map(c.round_amount)
    group_offset = pd.Series(0.0, index=df.index, dtype='float64')
    for idx, row in df.iterrows():
        group_offset.loc[idx] = offset_map.get(_offset_key(row[link_column], row['项目编号ID'], row['费用项编码']), 0)

    settled_amount = pd.Series(0.0, index=df.index, dtype='float64')
    for _, group in df.groupby([link_column, '项目编号ID', '费用项编码'], sort=False, dropna=False):
        indexes = list(group.index)
        total_written_off = c.round_amount(group_offset.loc[indexes].iloc[0])
        if len(indexes) == 1:
            settled_amount.loc[indexes[0]] = total_written_off
            continue

        detail_amount = amount.loc[indexes]
        detail_total = c.round_amount(detail_amount.sum())
        if detail_total:
            allocations = [c.round_amount(total_written_off * item / detail_total) for item in detail_amount]
        else:
            allocations = [c.round_amount(total_written_off / len(indexes))] * len(indexes)
        diff = c.round_amount(total_written_off - sum(allocations))
        allocations[-1] = c.round_amount(allocations[-1] + diff)
        settled_amount.loc[indexes] = allocations

    df['已到票核销金额'] = settled_amount.map(c.round_amount)
    df['已付未核金额'] = (amount - settled_amount).map(c.round_amount)
    return df


def build_mcn_supplier_output(source_df, fill_anchor_room=False):
    if source_df.empty:
        return pd.DataFrame(columns=SUPPLIER_OUTPUT_COLUMNS)

    vendor_info_map = c.build_supplier_vendor_info_map_for_rows(
        source_df['供应商ID'],
        supplier_texts=source_df['供应商'],
        document_numbers=source_df['流程编号'],
        missing_report_file=MCN_SUPPLIER_VENDOR_MISSING_FILE,
        log_prefix='[预付期初-MCN供应商预付]',
    )
    entity_map = c.build_accounting_entity_map_for_names(source_df['公司主体'])
    subject_lookup = c.build_subject_map()
    amount = pd.to_numeric(source_df['金额'], errors='coerce')

    def vendor_field(index, field):
        return vendor_info_map.get(index, {}).get(field, '')

    output_df = pd.DataFrame(index=source_df.index)
    output_df['来源系统'] = SOURCE_SYSTEM
    output_df['来源单据编号'] = source_df['流程编号']
    output_df['申请日期'] = source_df['申请日期'].map(c.format_date)
    output_df['单据类型'] = DOCUMENT_TYPE
    output_df['申请人工号'] = source_df['申请人工号']
    output_df['申请人姓名'] = source_df['申请人']
    output_df['泛微项目编号'] = source_df['项目编号'].map(_text)
    output_df['订单编号'] = source_df['项目编号'].map(lambda value: c.project_order_mapping_value(value, '订单编号'))
    output_df['订单名称'] = source_df['项目编号'].map(lambda value: c.project_order_mapping_value(value, '订单标题'))
    output_df['核算主体编号'] = source_df['公司主体'].map(lambda value: entity_map.get(c.normalize_name(value), '') if _text(value) else '')
    output_df['核算主体描述'] = source_df['公司主体']
    output_df['备注'] = source_df['标题'].map(lambda value: _text(value)[:150])
    output_df['合同号'] = source_df['合同号']
    output_df['合同收支计划行'] = ''
    output_df['保证金标志'] = '否'
    output_df['收款方编码'] = [vendor_field(index, 'code') for index in source_df.index]
    output_df['收款方描述'] = [
        vendor_field(index, 'name') or supplier_name
        for index, supplier_name in zip(source_df.index, source_df['供应商'])
    ]
    output_df['银行账号'] = c.resolve_hand_vendor_bank_accounts(
        output_df['收款方编码'], source_df['银行账号'])
    output_df['计划付款日期'] = source_df['申请日期'].map(c.format_date)
    output_df['银行转账备注'] = ''
    output_df['费用项目编码'] = source_df['费用项编码'].map(lambda value: _mcn_fee_subject_item(subject_lookup, value, 0))
    output_df['费用项目描述'] = source_df['费用项编码'].map(lambda value: _mcn_fee_subject_item(subject_lookup, value, 1))
    output_df['主播房间号'] = resolve_anchor_room_numbers(source_df) if fill_anchor_room else ''
    output_df['预付款支付币种'] = 'CNY'
    output_df['预付款金额（支付币种）'] = amount.map(c.round_amount)
    output_df['已到票核销金额（支付币种）'] = source_df['已到票核销金额'].map(c.round_amount)
    output_df['已付未核（支付币种）'] = source_df['已付未核金额'].map(c.round_amount)
    output_df['泛微费用项目编码'] = [
        _fanwei_fee_item_display(code, name)
        for code, name in zip(source_df['费用项编码'], source_df.get('费用项名称', pd.Series('', index=source_df.index)))
    ]
    return output_df[SUPPLIER_OUTPUT_COLUMNS]


def build_individual_vendor_maps(source_df):
    id_numbers = c.clean_text_values(source_df.get('身份证号', pd.Series(dtype=str)))
    by_id_number = {}
    if id_numbers:
        placeholders = c.in_placeholders(id_numbers)
        vendor_df = c.query_db(
            'ZT',
            'hfins_base',
            'SELECT vender_id, vender_code, description, taxpayer_name, tax_id_number, taxpayer_number '
            'FROM hfbs_system_vender '
            f'WHERE tax_id_number IN ({placeholders}) OR taxpayer_number IN ({placeholders})',
            id_numbers + id_numbers,
        )
        for _, row in vendor_df.iterrows():
            code = _text(row['vender_code'])
            if not code:
                continue
            info = {
                'code': code,
                'name': _first_non_blank(row.get('description'), row.get('taxpayer_name')),
                'id': c.format_code(row.get('vender_id')),
            }
            for key in (_text(row.get('tax_id_number')), _text(row.get('taxpayer_number'))):
                if key and key not in by_id_number:
                    by_id_number[key] = info

    names = c.clean_text_values(source_df.get('实际收款方', pd.Series(dtype=str)))
    by_name = c.build_hand_vendor_info_by_names(names)
    return by_id_number, by_name


def resolve_individual_vendor_info(row, by_id_number, by_name):
    id_number = _text(row.get('身份证号'))
    info = by_id_number.get(id_number) if id_number else None
    if info and info.get('code'):
        return info
    name_key = c.normalize_name(row.get('实际收款方'))
    name_info = by_name.get(name_key) if name_key else None
    return name_info if name_info and name_info.get('code') else {}


def allocate_mcn_fee_to_recipients(fee_df, recipient_df):
    if fee_df.empty:
        return fee_df

    rows = []
    recipient_by_doc = {
        key: group.copy()
        for key, group in recipient_df.groupby('ID', dropna=False)
    } if not recipient_df.empty else {}

    for doc_id, fee_group in fee_df.groupby('ID', sort=False, dropna=False):
        recipient_group = recipient_by_doc.get(doc_id)
        if recipient_group is None or recipient_group.empty:
            for _, fee_row in fee_group.iterrows():
                row = fee_row.to_dict()
                row.update({'收款人明细ID': '', '实际收款方': '', '身份证号': '', '明细银行账号': '', '手机号': '', '预付款金额分摊': fee_row['金额']})
                rows.append(row)
            continue

        recipient_group = recipient_group.copy()
        recipient_amount = pd.to_numeric(recipient_group['收款金额'], errors='coerce').fillna(0).map(c.round_amount)
        if float(recipient_amount.abs().sum()) == 0:
            recipient_row = recipient_group.iloc[0]
            for _, fee_row in fee_group.iterrows():
                row = fee_row.to_dict()
                row.update({
                    '收款人明细ID': recipient_row.get('收款人明细ID', ''),
                    '实际收款方': recipient_row.get('实际收款方', ''),
                    '身份证号': recipient_row.get('身份证号', ''),
                    '明细银行账号': recipient_row.get('明细银行账号', ''),
                    '手机号': recipient_row.get('手机号', ''),
                    '预付款金额分摊': fee_row['金额'],
                })
                rows.append(row)
            continue

        recipient_records = []
        for (_, recipient_row), amount in zip(recipient_group.iterrows(), recipient_amount):
            record = recipient_row.to_dict()
            record['_剩余收款金额'] = c.round_amount(abs(amount))
            recipient_records.append(record)
        recipient_index = 0

        def append_row(fee_row, recipient_row, allocated_amount):
            row = fee_row.to_dict()
            row.update({
                '收款人明细ID': recipient_row.get('收款人明细ID', ''),
                '实际收款方': recipient_row.get('实际收款方', ''),
                '身份证号': recipient_row.get('身份证号', ''),
                '明细银行账号': recipient_row.get('明细银行账号', ''),
                '手机号': recipient_row.get('手机号', ''),
                '预付款金额分摊': allocated_amount,
            })
            rows.append(row)

        for _, fee_row in fee_group.iterrows():
            raw_fee_amount = pd.to_numeric(fee_row['金额'], errors='coerce')
            if pd.isna(raw_fee_amount):
                raw_fee_amount = 0
            amount_sign = -1 if raw_fee_amount < 0 else 1
            remaining_fee_amount = c.round_amount(abs(raw_fee_amount))
            if remaining_fee_amount <= 0:
                append_row(fee_row, recipient_records[min(recipient_index, len(recipient_records) - 1)], fee_row['金额'])
                continue
            while remaining_fee_amount > 0:
                if recipient_index >= len(recipient_records):
                    append_row(fee_row, recipient_records[-1], c.round_amount(remaining_fee_amount * amount_sign))
                    break
                recipient_row = recipient_records[recipient_index]
                remaining_recipient_amount = recipient_row['_剩余收款金额']
                if remaining_recipient_amount <= 0:
                    recipient_index += 1
                    continue
                allocated_amount = c.round_amount(min(remaining_fee_amount, remaining_recipient_amount))
                append_row(fee_row, recipient_row, c.round_amount(allocated_amount * amount_sign))
                remaining_fee_amount = c.round_amount(remaining_fee_amount - allocated_amount)
                recipient_row['_剩余收款金额'] = c.round_amount(remaining_recipient_amount - allocated_amount)
                if recipient_row['_剩余收款金额'] <= 0:
                    recipient_index += 1

    return pd.DataFrame(rows)


def _recipient_amount_text(value):
    amount = pd.to_numeric(value, errors='coerce')
    return '' if pd.isna(amount) else f'{c.round_amount(amount):.2f}'


def attach_mcn_recipient_notes(
        source_df, recipient_df, amount_column='收款金额', amount_label='收款金额'):
    """源费用/订单明细保持一行一条，并把同流程全部收款人汇总到备注。"""
    df = source_df.copy()
    if df.empty:
        return df

    notes_by_doc = {}
    if not recipient_df.empty:
        for doc_id, recipient_group in recipient_df.groupby('ID', sort=False, dropna=False):
            details = []
            for _, recipient_row in recipient_group.iterrows():
                payee = _text(recipient_row.get('实际收款方'))
                amount = _recipient_amount_text(recipient_row.get(amount_column))
                if payee or amount:
                    details.append(f'收款人：{payee}，{amount_label}：{amount}')
            notes_by_doc[doc_id] = '；'.join(details)
    df['收款人明细备注'] = df['ID'].map(notes_by_doc).fillna('')
    return df


def build_mcn_gig_output(
        source_df, anchor_payee_category=False, counterparty_company_payee=False):
    if source_df.empty:
        return pd.DataFrame(columns=GIG_OUTPUT_COLUMNS)

    entity_map = c.build_accounting_entity_map_for_names(source_df['公司主体'])
    subject_lookup = c.build_subject_map()
    if counterparty_company_payee:
        counterparty_vendor_map = c.build_hand_vendor_info_by_names(source_df['供应商'])
        payee_codes = source_df['供应商'].map(
            lambda value: counterparty_vendor_map.get(c.normalize_name(value), {}).get('code', '')
        )
    else:
        by_id_number, by_name = build_individual_vendor_maps(source_df)
        vendor_infos = [
            resolve_individual_vendor_info(row, by_id_number, by_name)
            for _, row in source_df.iterrows()
        ]
        payee_codes = pd.Series([
            info.get('code', '') or _text(payee)
            for info, payee in zip(vendor_infos, source_df['实际收款方'])
        ], index=source_df.index)
    if counterparty_company_payee or '预付款金额分摊' not in source_df.columns:
        amount_source = source_df['金额']
    else:
        amount_source = source_df['预付款金额分摊']
    amount = pd.to_numeric(amount_source, errors='coerce')

    output_df = pd.DataFrame(index=source_df.index)
    output_df['来源系统'] = SOURCE_SYSTEM
    output_df['来源单据编号'] = source_df['流程编号']
    output_df['申请日期'] = source_df['申请日期'].map(c.format_date)
    output_df['单据类型'] = GIG_DOCUMENT_TYPE
    output_df['申请人工号'] = source_df['申请人工号']
    output_df['申请人姓名'] = source_df['申请人']
    output_df['泛微项目编号'] = source_df['项目编号'].map(_text)
    output_df['订单编号'] = source_df['项目编号'].map(lambda value: c.project_order_mapping_value(value, '订单编号'))
    output_df['订单名称'] = source_df['项目编号'].map(lambda value: c.project_order_mapping_value(value, '订单标题'))
    output_df['核算主体编号'] = source_df['公司主体'].map(lambda value: entity_map.get(c.normalize_name(value), '') if _text(value) else '')
    output_df['核算主体描述'] = source_df['公司主体']
    output_df['备注_单头'] = source_df['标题'].map(lambda value: _text(value)[:150])
    output_df['灵工平台收款方编码'] = source_df['供应商'].map(gig_platform_vendor)
    output_df['合同号'] = source_df['合同号']
    output_df['合同收支计划行'] = ''
    output_df['保证金标志'] = ''
    output_df['计划付款日期'] = source_df['申请日期'].map(c.format_date)
    output_df['银行转账备注'] = ''
    output_df['费用项目编码'] = source_df['费用项编码'].map(lambda value: _mcn_fee_subject_item(subject_lookup, value, 0))
    output_df['费用项目描述'] = source_df['费用项编码'].map(lambda value: _mcn_fee_subject_item(subject_lookup, value, 1))
    output_df['收款方类别'] = '主播' if anchor_payee_category else '供应商'
    output_df['收款方编码'] = payee_codes
    output_df['备注'] = (
        source_df.get('收款人明细备注', pd.Series('', index=source_df.index))
        if counterparty_company_payee else ''
    )
    bank_account_source = (
        source_df['银行账号']
        if counterparty_company_payee
        else source_df['明细银行账号']
    )
    output_df['银行账号'] = c.resolve_hand_vendor_bank_accounts(
        output_df['收款方编码'], bank_account_source)
    output_df['预付款支付币种'] = 'CNY'
    output_df['预付款金额（支付币种）'] = amount.map(c.round_amount)
    output_df['传送状态'] = '传送成功'
    output_df['支付状态'] = '支付成功'
    output_df['退款状态'] = ''
    output_df['核销状态'] = '已核销'
    output_df['泛微费用项目编码'] = [
        _fanwei_fee_item_display(code, name)
        for code, name in zip(source_df['费用项编码'], source_df.get('费用项名称', pd.Series('', index=source_df.index)))
    ]
    return output_df[GIG_OUTPUT_COLUMNS]


def _add_fanwei_order_code_column(output_df, source_df):
    df = output_df.copy()
    order_codes = source_df.get('泛微订单编号', pd.Series('', index=df.index)).map(_text)
    order_codes = order_codes.reindex(df.index).fillna('')
    insert_at = df.columns.get_loc('泛微项目编号') + 1 if '泛微项目编号' in df.columns else len(df.columns)
    if '泛微订单编号' in df.columns:
        df['泛微订单编号'] = order_codes
    else:
        df.insert(insert_at, '泛微订单编号', order_codes)
    return df


def _fill_sheet(worksheet, output_df):
    for col_idx, column_name in enumerate(output_df.columns, start=1):
        worksheet.cell(row=1, column=col_idx).value = column_name
    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row)
    for row in output_df.itertuples(index=False, name=None):
        worksheet.append([c._clean_cell_value(v) for v in row])


def _copy_template_sheet(workbook, source_sheet_name, target_sheet_name):
    if target_sheet_name in workbook.sheetnames:
        del workbook[target_sheet_name]
    copied = workbook.copy_worksheet(workbook[source_sheet_name])
    copied.title = target_sheet_name
    return copied


def _move_sheets_to_front(workbook, sheet_names):
    sheet_name_set = set(sheet_names)
    ordered_sheets = [workbook[sheet_name] for sheet_name in sheet_names if sheet_name in workbook.sheetnames]
    remaining_sheets = [sheet for sheet in workbook.worksheets if sheet.title not in sheet_name_set]
    workbook._sheets = ordered_sheets + remaining_sheets


def _prepare_template_sheet(workbook, target_sheet_name, fallback_sheet_name):
    if target_sheet_name in workbook.sheetnames:
        return workbook[target_sheet_name]
    if fallback_sheet_name not in workbook.sheetnames:
        raise KeyError(f'Worksheet {target_sheet_name} / {fallback_sheet_name} does not exist.')
    worksheet = workbook[fallback_sheet_name]
    worksheet.title = target_sheet_name
    return worksheet


def _build_sheet_summary(sheet_to_df):
    summary_frames = []
    columns = None
    for sheet_name, output_df in sheet_to_df.items():
        df = output_df.copy()
        if columns is None:
            columns = list(df.columns)
        if df.empty:
            continue
        df.insert(0, '来源Sheet', sheet_name)
        summary_frames.append(df)
    if columns is None:
        columns = []
    if not summary_frames:
        return pd.DataFrame(columns=['来源Sheet', *columns])
    return pd.concat(summary_frames, ignore_index=True)


def _save_workbook(workbook, output_file):
    try:
        workbook.save(output_file)
        return output_file
    except PermissionError:
        fallback_file = output_file.with_name(
            f'{output_file.stem}_{datetime.now().strftime("%H%M%S")}{output_file.suffix}'
        )
        print(f'[预付期初] 输出文件被占用，已改写到: {fallback_file}')
        workbook.save(fallback_file)
        return fallback_file


def write_output_workbook(supplier_output_df, gig_output_df, mcn_supplier_outputs, mcn_gig_outputs):
    wb = load_workbook(TEMPLATE_FILE)
    supplier_sheet = _prepare_template_sheet(wb, TEMPLATE_SHEET_SUPPLIER, TEMPLATE_SOURCE_SHEET_SUPPLIER)
    gig_sheet = _prepare_template_sheet(wb, TEMPLATE_SHEET_GIG, TEMPLATE_SOURCE_SHEET_GIG)
    supplier_summary_df = _build_sheet_summary({
        TEMPLATE_SHEET_SUPPLIER: supplier_output_df,
        **mcn_supplier_outputs,
    })
    gig_summary_df = _build_sheet_summary({
        TEMPLATE_SHEET_GIG: gig_output_df,
        **mcn_gig_outputs,
    })
    _fill_sheet(_copy_template_sheet(wb, TEMPLATE_SHEET_SUPPLIER, SUPPLIER_SUMMARY_SHEET), supplier_summary_df)
    _fill_sheet(_copy_template_sheet(wb, TEMPLATE_SHEET_GIG, GIG_SUMMARY_SHEET), gig_summary_df)
    _fill_sheet(supplier_sheet, supplier_output_df)
    _fill_sheet(gig_sheet, gig_output_df)
    for sheet_name, output_df in mcn_supplier_outputs.items():
        _fill_sheet(_copy_template_sheet(wb, TEMPLATE_SHEET_SUPPLIER, sheet_name), output_df)
    for sheet_name, output_df in mcn_gig_outputs.items():
        _fill_sheet(_copy_template_sheet(wb, TEMPLATE_SHEET_GIG, sheet_name), output_df)
    _move_sheets_to_front(
        wb,
        [
            SUPPLIER_SUMMARY_SHEET,
            GIG_SUMMARY_SHEET,
            TEMPLATE_SHEET_SUPPLIER,
            *mcn_supplier_outputs.keys(),
            TEMPLATE_SHEET_GIG,
            *mcn_gig_outputs.keys(),
        ],
    )
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    return _save_workbook(wb, OUTPUT_FILE)


# ============================ 源读取 ============================
def read_supplier_source():
    """从 DB 直接读取赛事供应商预付,仅按项目白名单过滤。"""
    c.validate_fw_fields(FW_SUPPLIER_TABLE, EXPECTED_SUPPLIER_FIELDS)
    c.validate_fw_fields(FW_PAYMENT_OFFSET_TABLE, EXPECTED_PAYMENT_OFFSET_FIELDS)
    event_project_ids = project_filter_ids(EVENT_PROJECT_SHEET, EVENT_PROJECT_TABLES)
    if not event_project_ids:
        return pd.DataFrame()

    params = {
        'project_ids': event_project_ids,
        'project_codes': tuple(sorted(project_filter_codes(EVENT_PROJECT_SHEET))),
        'project_fallback_date_from': PROJECT_FALLBACK_DATE_FROM,
    }
    stats = _query_fw(SUPPLIER_STATS_SQL, params).iloc[0]
    print('[预付期初-供应商预付款-DB] SQL过滤: 仅保留赛事项目白名单; 不再过滤申请日期/流程状态/作废状态')
    print(f"  命中主表 {int(stats['document_count'] or 0)} 单 / 明细 {int(stats['detail_count'] or 0)} 行; "
          f"金额合计 {float(stats['amount_total'] or 0):.2f}")
    source_df = _query_fw(SUPPLIER_SOURCE_SQL, params)
    source_df = _filter_by_project_whitelist(
        source_df, '项目编号ID', EVENT_PROJECT_SHEET, EVENT_PROJECT_TABLES,
        '预付期初-供应商预付款-DB',
    )
    merged_df = resolve_supplier_source_values(source_df)
    print('[预付期初-供应商预付款-DB] SQL主子合并明细行数:', len(merged_df))
    return merged_df


def read_gig_source():
    """从 DB 直接读取赛事零工预付,仅按项目白名单过滤。"""
    c.validate_fw_fields(FW_GIG_HEADER_TABLE, EXPECTED_GIG_HEADER_FIELDS)
    c.validate_fw_fields(FW_GIG_WORKFLOW_TABLE, EXPECTED_GIG_WORKFLOW_FIELDS)
    event_project_ids = project_filter_ids(EVENT_PROJECT_SHEET, EVENT_PROJECT_TABLES)
    if not event_project_ids:
        return pd.DataFrame()

    params = {
        'project_ids': event_project_ids,
        'project_codes': tuple(sorted(project_filter_codes(EVENT_PROJECT_SHEET))),
        'project_fallback_date_from': PROJECT_FALLBACK_DATE_FROM,
    }
    stats = _query_fw(GIG_STATS_SQL, params).iloc[0]
    print('[预付期初-零工预付款-DB] SQL过滤: 仅保留赛事项目白名单; 不再过滤申请日期/流程状态/作废状态')
    print(f"  命中主表 {int(stats['document_count'] or 0)} 单 / 明细 {int(stats['detail_count'] or 0)} 行; "
          f"金额合计 {float(stats['amount_total'] or 0):.2f}")
    recipient_source_df = _query_fw(GIG_RECIPIENT_SOURCE_SQL, params)
    recipient_source_df = _filter_by_project_whitelist(
        recipient_source_df, '项目编号ID', EVENT_PROJECT_SHEET, EVENT_PROJECT_TABLES,
        '预付期初-零工预付款-收款人',
    )
    recipient_df = resolve_gig_source_values(recipient_source_df)
    print('[预付期初-零工预付款-DB] SQL收款人明细行数:', len(recipient_df))
    budget_source_df = _query_fw(GIG_BUDGET_SOURCE_SQL, params)
    budget_source_df = _filter_by_project_whitelist(
        budget_source_df, '项目编号ID', EVENT_PROJECT_SHEET, EVENT_PROJECT_TABLES,
        '预付期初-零工预付款-预算项',
    )
    budget_df = resolve_gig_budget_values(budget_source_df)
    print('[预付期初-零工预付款-DB] SQL预算项明细行数:', len(budget_df))
    return allocate_gig_budget_to_recipients(recipient_df, budget_df)


def _mcn_query_params():
    mcn_project_ids = project_filter_ids(MCN_PROJECT_SHEET, MCN_PROJECT_TABLES)
    if not mcn_project_ids:
        return None
    return {
        'project_ids': mcn_project_ids,
        'project_codes': tuple(sorted(project_filter_codes(MCN_PROJECT_SHEET))),
        'project_fallback_date_from': PROJECT_FALLBACK_DATE_FROM,
        'direct_payment_code': DIRECT_PAYMENT_PLATFORM_CODE,
        'prepayment_yes_code': PREPAYMENT_YES_CODE,
        'prepayment_no_code': PREPAYMENT_NO_CODE,
    }


def _prepare_mcn_platform_columns(source_df):
    df = source_df.copy()
    if '供应商ID' not in df.columns and '平台供应商ID' in df.columns:
        df['供应商ID'] = df['平台供应商ID']
    if '银行账号ID' not in df.columns and '平台银行账号ID' in df.columns:
        df['银行账号ID'] = df['平台银行账号ID']
    return df


def read_mcn_supplier_sources():
    params = _mcn_query_params()
    if not params:
        return {
            SHEET_MCN_PREPAYMENT: pd.DataFrame(),
            SHEET_MCN_PREPAYMENT_ORDER: pd.DataFrame(),
            SHEET_MCN_ANCHOR_PREPAYMENT: pd.DataFrame(),
        }

    specs = [
        (SHEET_MCN_PREPAYMENT, MCN_PREPAYMENT_SOURCE_SQL, MCN_PREPAYMENT_OFFSET_SQL, 'RequestID'),
        (SHEET_MCN_PREPAYMENT_ORDER, MCN_PREPAYMENT_ORDER_SOURCE_SQL, MCN_PREPAYMENT_ORDER_OFFSET_SQL, 'RequestID'),
        (SHEET_MCN_ANCHOR_PREPAYMENT, MCN_ANCHOR_PREPAYMENT_SOURCE_SQL, MCN_ANCHOR_OFFSET_SQL, 'ID'),
    ]
    source_by_sheet = {}
    for sheet_name, source_sql, offset_sql, link_column in specs:
        source_df = _query_fw(source_sql, params)
        source_df = _filter_by_project_whitelist(
            source_df, '项目编号ID', MCN_PROJECT_SHEET, MCN_PROJECT_TABLES,
            f'预付期初-{sheet_name}',
        )
        source_df = resolve_mcn_common_values(source_df)
        if not source_df.empty:
            source_df = apply_mcn_offset_amounts(source_df, _query_fw(offset_sql, params), link_column)
        source_by_sheet[sheet_name] = source_df
        print(f'[预付期初-{sheet_name}] SQL明细行数:', len(source_df))
    return source_by_sheet


def _read_mcn_platform_gig_source(
        sheet_name, fee_sql, recipient_sql, params, keep_source_rows=False,
        note_amount_column='收款金额', note_amount_label='收款金额'):
    fee_df = _prepare_mcn_platform_columns(_query_fw(fee_sql, params))
    fee_df = _filter_by_project_whitelist(
        fee_df, '项目编号ID', MCN_PROJECT_SHEET, MCN_PROJECT_TABLES,
        f'预付期初-{sheet_name}',
    )
    fee_df = resolve_mcn_common_values(fee_df)
    recipient_df = _query_fw(recipient_sql, params)
    if not fee_df.empty and not recipient_df.empty:
        recipient_df = recipient_df[recipient_df['ID'].isin(fee_df['ID'].dropna().unique())].copy()
    merged_df = (
        attach_mcn_recipient_notes(
            fee_df, recipient_df,
            amount_column=note_amount_column,
            amount_label=note_amount_label,
        )
        if keep_source_rows
        else allocate_mcn_fee_to_recipients(fee_df, recipient_df)
    )
    print(f'[预付期初-{sheet_name}] SQL明细行数:', len(merged_df))
    return merged_df


def read_mcn_gig_sources():
    params = _mcn_query_params()
    if not params:
        return {
            SHEET_MCN_OUTBOUND_PREPAYMENT: pd.DataFrame(),
            SHEET_MCN_OUTBOUND_ORDER_PREPAYMENT: pd.DataFrame(),
            SHEET_MCN_ANCHOR_PLATFORM_PREPAYMENT: pd.DataFrame(),
        }

    source_by_sheet = {
        SHEET_MCN_OUTBOUND_PREPAYMENT: _read_mcn_platform_gig_source(
            SHEET_MCN_OUTBOUND_PREPAYMENT,
            MCN_OUTBOUND_FEE_SOURCE_SQL,
            MCN_OUTBOUND_RECIPIENT_SOURCE_SQL,
            params,
            keep_source_rows=True,
            note_amount_column='打款金额',
            note_amount_label='打款金额',
        ),
        SHEET_MCN_OUTBOUND_ORDER_PREPAYMENT: _read_mcn_platform_gig_source(
            SHEET_MCN_OUTBOUND_ORDER_PREPAYMENT,
            MCN_OUTBOUND_ORDER_FEE_SOURCE_SQL,
            MCN_OUTBOUND_ORDER_RECIPIENT_SOURCE_SQL,
            params,
            keep_source_rows=True,
        ),
    }

    anchor_df = _query_fw(MCN_ANCHOR_PLATFORM_SOURCE_SQL, params)
    anchor_df = _filter_by_project_whitelist(
        anchor_df, '项目编号ID', MCN_PROJECT_SHEET, MCN_PROJECT_TABLES,
        f'预付期初-{SHEET_MCN_ANCHOR_PLATFORM_PREPAYMENT}',
    )
    anchor_df = resolve_mcn_common_values(anchor_df)
    if not anchor_df.empty:
        anchor_df['预付款金额分摊'] = anchor_df['金额']
    source_by_sheet[SHEET_MCN_ANCHOR_PLATFORM_PREPAYMENT] = anchor_df
    print(f'[预付期初-{SHEET_MCN_ANCHOR_PLATFORM_PREPAYMENT}] SQL明细行数:', len(anchor_df))
    return source_by_sheet


def _required_without(required_cols, excluded_cols):
    excluded = set(excluded_cols)
    return [column for column in required_cols if column not in excluded]


def _collect_exception_sheets(
        output_df, source_df, required_cols, rule_table, source_field_map, bank_source_col=None):
    sheets = {'必输字段未达100%': c.fill_summary(output_df, required_cols, RULE_SHEET, rule_table)}
    sheets.update(c.collect_field_issues(output_df, source_df, required_cols, source_field_map))
    if bank_source_col and bank_source_col in source_df.columns:
        bank_issues = c.collect_hand_vendor_bank_account_issues(output_df, source_df[bank_source_col])
        if not bank_issues.empty:
            sheets['银行账号_校验异常'] = bank_issues
    sheets.update(c.collect_order_mapping_issues(source_df))
    c.attach_budget_issue_columns(sheets, c.build_budget_issue_map(source_df))
    return sheets


def run():
    # 1. SQL 直接查过滤后的源数据
    supplier_merged_df = read_supplier_source()
    gig_merged_df = read_gig_source()
    mcn_supplier_sources = read_mcn_supplier_sources()
    mcn_gig_sources = read_mcn_gig_sources()

    # 2. 构建原 2 个 sheet + 追加 MCN 6 个 sheet 输出
    supplier_output_df = (
        build_supplier_output(supplier_merged_df)
        if not supplier_merged_df.empty else pd.DataFrame(columns=SUPPLIER_OUTPUT_COLUMNS)
    )
    print('[预付期初-供应商预付款-DB] 输出明细行数:', len(supplier_output_df))
    gig_output_df = (
        build_gig_output(gig_merged_df)
        if not gig_merged_df.empty else pd.DataFrame(columns=GIG_OUTPUT_COLUMNS)
    )
    print('[预付期初-零工预付款-DB] 输出明细行数:', len(gig_output_df))
    mcn_supplier_outputs = {
        sheet_name: build_mcn_supplier_output(
            source_df,
            fill_anchor_room=(sheet_name == SHEET_MCN_ANCHOR_PREPAYMENT),
        )
        for sheet_name, source_df in mcn_supplier_sources.items()
    }
    mcn_supplier_outputs[SHEET_MCN_PREPAYMENT_ORDER] = _add_fanwei_order_code_column(
        mcn_supplier_outputs[SHEET_MCN_PREPAYMENT_ORDER],
        mcn_supplier_sources[SHEET_MCN_PREPAYMENT_ORDER],
    )
    mcn_gig_outputs = {
        SHEET_MCN_OUTBOUND_PREPAYMENT: build_mcn_gig_output(
            mcn_gig_sources[SHEET_MCN_OUTBOUND_PREPAYMENT],
            anchor_payee_category=False,
            counterparty_company_payee=True,
        ),
        SHEET_MCN_OUTBOUND_ORDER_PREPAYMENT: build_mcn_gig_output(
            mcn_gig_sources[SHEET_MCN_OUTBOUND_ORDER_PREPAYMENT],
            anchor_payee_category=False,
            counterparty_company_payee=True,
        ),
        SHEET_MCN_ANCHOR_PLATFORM_PREPAYMENT: build_mcn_gig_output(
            mcn_gig_sources[SHEET_MCN_ANCHOR_PLATFORM_PREPAYMENT],
            anchor_payee_category=True,
        ),
    }
    mcn_gig_outputs[SHEET_MCN_OUTBOUND_ORDER_PREPAYMENT] = _add_fanwei_order_code_column(
        mcn_gig_outputs[SHEET_MCN_OUTBOUND_ORDER_PREPAYMENT],
        mcn_gig_sources[SHEET_MCN_OUTBOUND_ORDER_PREPAYMENT],
    )
    for sheet_name, output_df in {**mcn_supplier_outputs, **mcn_gig_outputs}.items():
        print(f'[预付期初-{sheet_name}] 输出明细行数:', len(output_df))

    # 3. 填充率(必输字段以规则表「是否必填」=Y 为准)
    supplier_required = c.required_columns(RULE_SHEET, RULE_TABLE_SUPPLIER)
    gig_required = c.required_columns(RULE_SHEET, RULE_TABLE_GIG)
    gig_required_effective = _required_without(gig_required, GIG_NOT_INVOLVED_REQUIRED_COLUMNS)
    mcn_gig_required = _required_without(gig_required, MCN_GIG_NOT_INVOLVED_REQUIRED_COLUMNS)
    print('— 供应商预付款 填充率 —')
    c.report_fill(supplier_output_df, supplier_required)
    print('— 灵工预付款 填充率 —')
    c.report_fill(gig_output_df, gig_required_effective)
    for sheet_name, output_df in mcn_supplier_outputs.items():
        print(f'— {sheet_name} 填充率 —')
        c.report_fill(output_df, supplier_required)
    for sheet_name, output_df in mcn_gig_outputs.items():
        print(f'— {sheet_name} 填充率 —')
        c.report_fill(output_df, mcn_gig_required)

    # 4. 写模版(原 2 个 tab 保留,MCN 6 个 tab 复制模板追加,lov 页保留)
    output_file = write_output_workbook(supplier_output_df, gig_output_df, mcn_supplier_outputs, mcn_gig_outputs)
    print('已写出:', output_file)

    # 5. 问题清单
    exception_sheets = {}
    supplier_sheets = _collect_exception_sheets(
        supplier_output_df, supplier_merged_df, supplier_required, RULE_TABLE_SUPPLIER,
        SUPPLIER_ISSUE_SOURCE_FIELDS, bank_source_col='银行卡号',
    )
    exception_sheets.update({f'供应商_{name}': df for name, df in supplier_sheets.items()})

    gig_sheets = _collect_exception_sheets(
        gig_output_df, gig_merged_df, gig_required_effective, RULE_TABLE_GIG,
        GIG_ISSUE_SOURCE_FIELDS, bank_source_col='银行账号',
    )
    exception_sheets.update({f'灵工_{name}': df for name, df in gig_sheets.items()})

    for sheet_name, output_df in mcn_supplier_outputs.items():
        source_df = mcn_supplier_sources[sheet_name]
        sheets = _collect_exception_sheets(
            output_df, source_df, supplier_required, RULE_TABLE_SUPPLIER,
            MCN_SUPPLIER_ISSUE_SOURCE_FIELDS, bank_source_col='银行账号',
        )
        exception_sheets.update({f'{sheet_name}_{name}': df for name, df in sheets.items()})

    for sheet_name, output_df in mcn_gig_outputs.items():
        source_df = mcn_gig_sources[sheet_name]
        sheets = _collect_exception_sheets(
            output_df, source_df, mcn_gig_required, RULE_TABLE_GIG,
            MCN_GIG_ISSUE_SOURCE_FIELDS, bank_source_col='明细银行账号',
        )
        exception_sheets.update({f'{sheet_name}_{name}': df for name, df in sheets.items()})

    c.write_exceptions(EXCEPTION_FILE, exception_sheets)
    print('已写出:', EXCEPTION_FILE, '| 各清单条数:', {k: len(v) for k, v in exception_sheets.items()})


if __name__ == '__main__':
    run()
