# -*- coding: utf-8 -*-
"""反商业贿赂协议签署情况 —— 把已签署的反商业贿赂协议供应商补登到台账。

规则:合同审批台账(MCN uf_htk + 赛事 uf_htsp)里,
    合同编号含 'H-P' 且 合同名称含 '贿赂'(兼容「反商业贿赂」「反贿赂」)的合同,
    视为反商业贿赂协议;把对应供应商补登到第一个 sheet 末尾。

不改动模板 data/templates/contract_anchor_db/反商业贿赂协议签署情况.xlsx,
而是以它为底稿,另存为带日期后缀的新文件
    output/anti_bribery_signers_db/反商业贿赂协议签署情况_<YYYYMMDD>.xlsx。

去重键 = (供应商名称, 合同编号):若该组合已记录则跳过,可重复跑(幂等)。
新增行尽量补全:供应商名称 / 合同编号 / 合同状态 / 签约时间 / 供应商ID /
    供应商汉得编码 / 供应商汉得名称。

跑法:在项目根执行  python run.py anti_bribery_signers_db
"""
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

if __package__ is None or __package__ == '':
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from etl import common as c
from etl.tasks import contract_general_db as cg


# ============================ 文件 / 过滤口径 ============================
TASK_NAME = 'anti_bribery_signers_db'
# 底稿(只读):已有的签署情况台账,提供已记录的去重基准与表头结构。
TEMPLATE_FILE = cg.TEMPLATE_DIR / '反商业贿赂协议签署情况.xlsx'
# 产出(带日期后缀):底稿内容 + 本次新增,模板本身不改动。
OUTPUT_DIR = c.OUT_DIR / TASK_NAME
OUTPUT_FILE = OUTPUT_DIR / f'反商业贿赂协议签署情况_{c.today_suffix()}.xlsx'

# 合同编号「带有 H-P」、合同名称「带有 贿赂」即视为反商业贿赂协议。
CONTRACT_NUMBER_LIKE = '%H-P%'   # UPPER(htbh) LIKE,兼容 H-P / H-PZ 等
CONTRACT_NAME_LIKE = '%贿赂%'     # 兼容「反商业贿赂协议」「反贿赂协议」

# 第一个 sheet 的列名(只写其中存在的列)。
HEADER_SUPPLIER_NAME = '供应商名称'
HEADER_CONTRACT_NO = '合同编号'
HEADER_CONTRACT_STATUS = '合同状态'
HEADER_SIGN_TIME = '签约时间'
HEADER_REMARK = '备注'
HEADER_SUPPLIER_ID = '供应商ID'
HEADER_HAND_CODE = '供应商汉得编码'
HEADER_HAND_NAME = '供应商汉得名称'

REMARK_CUSTOMER_FALLBACK = '无供应商，取客户'

# 取数:MCN 合同库 uf_htk(合同名称字段 htbt)。
SQL_HTK = """
SELECT
    h.htbh  AS `合同编号`,
    h.htbt  AS `合同标题`,
    h.htzt  AS `合同签署状态ID`,
    h.htgys AS `合同供应商ID`,
    h.htkh  AS `合同客户ID`,
    h.htqdrq AS `合同签订日期`
FROM uf_htk h
WHERE UPPER(h.htbh) LIKE %s AND h.htbt LIKE %s
ORDER BY h.htbh, h.id
"""

# 取数:赛事 合同审批台账 uf_htsp(合同名称字段 htmc,供应商字段 gys,客户字段 kh,签订日期 qdsj)。
SQL_HTSP = """
SELECT
    h.htbh AS `合同编号`,
    h.htmc AS `合同标题`,
    h.htzt AS `合同签署状态ID`,
    h.gys  AS `合同供应商ID`,
    h.kh   AS `合同客户ID`,
    h.qdsj AS `合同签订日期`
FROM uf_htsp h
WHERE UPPER(h.htbh) LIKE %s AND h.htmc LIKE %s
ORDER BY h.htbh, h.id
"""


# ============================ 小工具 ============================
def _text(value):
    if pd.isna(value):
        return ''
    text = str(value).strip()
    return '' if text in ('', 'nan', 'None', 'NaT') else text


def _norm_no(value):
    """合同编号去重键:去空白、转大写。"""
    return re.sub(r'\s+', '', _text(value)).upper()


def _dedupe_key(supplier_name, contract_no):
    """去重键 = (归一化供应商名称, 归一化合同编号)。

    供应商名称用 common.normalize_name,容忍历史数据里的空格/换行/标点差异。
    """
    return (c.normalize_name(supplier_name), _norm_no(contract_no))


def _join(parts):
    """多供应商时按 '; ' 拼接,去重且保留顺序、忽略空值。"""
    result, seen = [], set()
    for part in parts:
        text = _text(part)
        if text and text not in seen:
            seen.add(text)
            result.append(text)
    return '; '.join(result)


# ============================ 取数 / 解析 ============================
def _fetch_anti_bribery_contracts():
    """查两套合同台账,返回带「合同状态」「供应商各字段」的候选行列表。"""
    htk_df = c.query_db('FW', 'vspn_xtyy', SQL_HTK, [CONTRACT_NUMBER_LIKE, CONTRACT_NAME_LIKE])
    htk_df['数据来源'] = '泛微(MCN)'
    htsp_df = c.query_db('FW', 'vspn_xtyy', SQL_HTSP, [CONTRACT_NUMBER_LIKE, CONTRACT_NAME_LIKE])
    htsp_df['数据来源'] = '泛微(赛事)'
    print(f'[反商业贿赂] 命中合同:MCN(uf_htk) {len(htk_df)} 条; 赛事(uf_htsp) {len(htsp_df)} 条')

    # 合同状态名称:两表各自的 htzt 选项码表。
    htk_status = c.build_fw_select_option_maps(cg.FW_TABLE, ['htzt']).get('htzt', {})
    htsp_status = c.build_fw_select_option_maps(cg.FW_TABLE_HTSP, ['htzt']).get('htzt', {})

    df = pd.concat([htk_df, htsp_df], ignore_index=True)
    if df.empty:
        return []

    # 对方主体:复用一般流程的汉得匹配,拿到 泛微名称 / 汉得编码 / 汉得名称。
    # 反商业贿赂协议可能签给供应商,也可能签给客户;无供应商时回退取客户。
    supplier_info_map = cg.build_supplier_info_map_for_values(df['合同供应商ID'])
    customer_info_map = cg.build_customer_info_map_for_values(df['合同客户ID'])

    contracts = []
    customer_fallback = 0
    seen_contract = set()  # 同编号在两表都命中时,按 MCN 优先取一条
    for source in df.to_dict('records'):
        contract_no = _text(source['合同编号'])
        if contract_no and contract_no in seen_contract:
            continue
        seen_contract.add(contract_no)

        status_map = htk_status if source['数据来源'] == '泛微(MCN)' else htsp_status
        supplier_ids = c.parse_browser_ids(source['合同供应商ID'])
        supplier_infos = [supplier_info_map.get(supplier_id, {}) for supplier_id in supplier_ids]
        party_name = _join(info.get('source_name', '') for info in supplier_infos)

        if party_name:
            party_ids, infos, remark = supplier_ids, supplier_infos, ''
        else:
            # 无供应商 -> 取客户(合同客户ID)。
            party_ids = c.parse_browser_ids(source['合同客户ID'])
            infos = [customer_info_map.get(customer_id, {}) for customer_id in party_ids]
            party_name = _join(info.get('source_name', '') for info in infos)
            remark = REMARK_CUSTOMER_FALLBACK if party_name else ''
            if party_name:
                customer_fallback += 1

        contracts.append({
            HEADER_SUPPLIER_NAME: party_name,
            HEADER_CONTRACT_NO: contract_no,
            HEADER_CONTRACT_STATUS: status_map.get(c.format_code(source['合同签署状态ID']), ''),
            HEADER_SIGN_TIME: c.format_date(source['合同签订日期']),
            HEADER_REMARK: remark,
            HEADER_SUPPLIER_ID: _join(party_ids),
            HEADER_HAND_CODE: _join(info.get('code', '') for info in infos),
            HEADER_HAND_NAME: _join(info.get('name', '') for info in infos),
        })
    print(f'[反商业贿赂] 无供应商改取客户: {customer_fallback} 条')
    return contracts


# ============================ 写入第一个 sheet ============================
def _header_columns(worksheet):
    """第一个 sheet 的表头文本 -> 列号(1-based)。"""
    columns = {}
    for cell in worksheet[1]:
        text = _text(cell.value)
        if text and text not in columns:
            columns[text] = cell.column
    return columns


def _existing_keys_and_last_row(worksheet, columns):
    """已记录去重键集合,以及最后一行有内容的行号(供续写定位)。"""
    name_col = columns.get(HEADER_SUPPLIER_NAME)
    no_col = columns.get(HEADER_CONTRACT_NO)
    keys = set()
    last_row = 1  # 至少保留表头行
    for row in range(2, worksheet.max_row + 1):
        supplier_name = _text(worksheet.cell(row, name_col).value) if name_col else ''
        contract_no = _text(worksheet.cell(row, no_col).value) if no_col else ''
        if not (supplier_name or contract_no):
            continue
        keys.add(_dedupe_key(supplier_name, contract_no))
        last_row = row
    return keys, last_row


def _timestamped_path(path):
    return path.with_name(f'{path.stem}_{datetime.now().strftime("%H%M%S")}{path.suffix}')


def append_anti_bribery_signers(template_file=TEMPLATE_FILE, output_file=OUTPUT_FILE):
    """以模板为底稿,把命中的反商业贿赂协议供应商补登到第一个 sheet 末尾,
    按 (供应商名称, 合同编号) 去重,另存为带日期后缀的新文件;模板本身不改动。"""
    contracts = _fetch_anti_bribery_contracts()
    print(f'[反商业贿赂] 去重后候选合同(MCN 优先): {len(contracts)} 条')

    workbook = load_workbook(template_file)
    worksheet = workbook.worksheets[0]
    columns = _header_columns(worksheet)
    if HEADER_SUPPLIER_NAME not in columns and HEADER_CONTRACT_NO not in columns:
        raise RuntimeError(
            f'第一个 sheet「{worksheet.title}」未找到「{HEADER_SUPPLIER_NAME}」或「{HEADER_CONTRACT_NO}」列,无法写入')

    existing_keys, last_row = _existing_keys_and_last_row(worksheet, columns)
    next_row = last_row + 1
    appended = 0
    skipped = 0
    for contract in contracts:
        key = _dedupe_key(contract[HEADER_SUPPLIER_NAME], contract[HEADER_CONTRACT_NO])
        if key in existing_keys:
            skipped += 1
            continue
        existing_keys.add(key)
        for header, col in columns.items():
            if header in contract:
                worksheet.cell(next_row, col, contract[header])
        next_row += 1
        appended += 1

    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(output_file)
        saved_to = output_file
    except PermissionError:
        saved_to = _timestamped_path(output_file)
        print(f'产出文件被占用,改写到: {saved_to}')
        workbook.save(saved_to)

    print(f'[反商业贿赂] 新增 {appended} 条; 跳过(已记录) {skipped} 条; 模板未改动')
    print('已写出:', saved_to)
    return appended


def run():
    append_anti_bribery_signers()


if __name__ == '__main__':
    run()
