# -*- coding: utf-8 -*-
"""合同迁移 - 反商业贿赂协议(DB直连版)。

把模板中的「归档的反贿赂供应商」和「未归档的反贿赂签署协议」
清洗到「反商业贿赂协议」sheet; 附件下载清单复用一般流程附件取数,
但落盘目录按 合同编号/合同附件|其他附件。
"""
import os
import sys
from pathlib import Path

import pandas as pd

from etl import common as c
from etl.tasks import contract_general_db as base

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')


TASK_NAME = 'contract_anti_bribery_db'
TEMPLATE_FILE = base.TEMPLATE_DIR / '签署反商业贿赂协议6.25终版.xlsx'
OUTPUT_DIR = c.OUT_DIR / TASK_NAME
OUTPUT_FILE = OUTPUT_DIR / f'反商业贿赂协议_{base.DATE_SUFFIX}.xlsx'
EXCEPTION_FILE = OUTPUT_DIR / f'反商业贿赂协议_核对清单_{base.DATE_SUFFIX}.xlsx'
MANIFEST_FILE = OUTPUT_DIR / f'反商业贿赂协议合同附件下载清单_{base.DATE_SUFFIX}.xlsx'

SHEET_ARCHIVED = '归档的反贿赂供应商'
SHEET_UNARCHIVED = '未归档的反贿赂签署协议'
SHEET_TARGET = '反商业贿赂协议'

FIXED_CATEGORY = '其他-反商业贿赂协议'
FIXED_PAY_TYPE = '无金额'
FIXED_VALIDITY = '长期有效'
FIXED_SIGN_FORM = '纸质签约'
FIXED_PRINT_MODE = '黑白双面打印'


SOURCE_SQL_MCN_BY_CODES = """
SELECT
    h.id AS `ID`,
    h.htbh AS `合同编号`,
    h.htbt AS `合同标题`,
    h.htlx AS `合同类型ID`,
    h.htejlx AS `合同二级类型ID`,
    h.htzt AS `合同签署状态ID`,
    h.htlc AS `合同流程ID`,
    h.htcjrq AS `合同创建日期`,
    h.htqdrq AS `合同签订日期`,
    h.htyxqqssj AS `合同有效期起始时间`,
    h.htyxqjzsj AS `合同有效期截止时间`,
    h.htzhry AS `合同执行人员ID`,
    h.htyyfw AS `合同用印范围ID`,
    h.htkh AS `合同客户ID`,
    h.htgys AS `合同供应商ID`,
    h.htje AS `合同金额`,
    h.htyjsr AS `合同预计收入`,
    h.htyjzc AS `合同预计支出`,
    h.htzy AS `合同摘要`,
    h.htqdg AS `合同附件DOCID`,
    h.cbzx AS `成本中心ID`,
    h.glkjxy AS `关联框架协议ID`,
    h.bglx AS `变更类型ID`,
    h.bczzbhsc AS `补充/终止编号生成`,
    h.htszxmbh AS `合同所属项目编号ID`,
    h.htszxm AS `合同所属项目`,
    NULL AS `收入税率`,
    NULL AS `支出税率`,
    NULL AS `押金`,
    NULL AS `保证金`,
    NULL AS `采购申请单ID`,
    NULL AS `专项分类编码`,
    h.modedatacreater AS `合同创建人ID`,
    h.modedatacreater AS `申请人ID`,
    rb.workflowid AS `流程类型ID`,
    wb.workflowname AS `流程名称`,
    nb.nodename AS `合同审批状态`
FROM uf_htk h
LEFT JOIN workflow_requestbase rb ON rb.requestid = h.htlc
LEFT JOIN workflow_base wb ON wb.id = rb.workflowid
LEFT JOIN workflow_nownode nn ON nn.requestid = h.htlc
LEFT JOIN workflow_nodebase nb ON nb.id = COALESCE(rb.currentnodeid, nn.nownodeid, rb.lastnodeid)
WHERE h.htbh IN ({placeholders})
ORDER BY h.htbh, h.id
"""

SOURCE_SQL_HTSP_BY_CODES = """
SELECT
    h.id AS `ID`,
    h.htbh AS `合同编号`,
    h.htmc AS `合同标题`,
    h.htlx AS `合同类型ID`,
    h.htejfl AS `合同二级类型ID`,
    h.htzt AS `合同签署状态ID`,
    h.lcqqid AS `合同流程ID`,
    h.sqrq AS `合同创建日期`,
    h.qdsj AS `合同签订日期`,
    h.htksrq AS `合同有效期起始时间`,
    h.htjsrq AS `合同有效期截止时间`,
    h.htzhr AS `合同执行人员ID`,
    h.gszt AS `合同用印范围ID`,
    h.kh AS `合同客户ID`,
    h.gys AS `合同供应商ID`,
    NULL AS `合同金额`,
    h.srje AS `合同预计收入`,
    h.zcje AS `合同预计支出`,
    h.htzy AS `合同摘要`,
    h.htsxg AS `合同附件DOCID`,
    h.htcg AS `赛事初稿DOCID`,
    h.htqsg AS `赛事签署稿DOCID`,
    h.htsxg AS `赛事生效稿DOCID`,
    h.cbzx AS `成本中心ID`,
    h.glht AS `关联框架协议ID`,
    NULL AS `变更类型ID`,
    NULL AS `补充/终止编号生成`,
    h.xmbh AS `合同所属项目编号ID`,
    h.xmmc AS `合同所属项目`,
    h.srsl AS `收入税率`,
    h.zcsl AS `支出税率`,
    h.yj AS `押金`,
    h.bzj AS `保证金`,
    h.cgsqddx AS `采购申请单ID`,
    h.zxflcg AS `专项分类编码`,
    h.modedatacreater AS `合同创建人ID`,
    h.sqr AS `申请人ID`,
    rb.workflowid AS `流程类型ID`,
    wb.workflowname AS `流程名称`,
    nb.nodename AS `合同审批状态`
FROM uf_htsp h
LEFT JOIN workflow_requestbase rb ON rb.requestid = h.lcqqid
LEFT JOIN workflow_base wb ON wb.id = rb.workflowid
LEFT JOIN workflow_nownode nn ON nn.requestid = h.lcqqid
LEFT JOIN workflow_nodebase nb ON nb.id = COALESCE(rb.currentnodeid, nn.nownodeid, rb.lastnodeid)
WHERE h.htbh IN ({placeholders})
ORDER BY h.htbh, h.id
"""


def _text(value):
    return base._text(value)


def _headers(sheet_name=SHEET_TARGET):
    return base._sheet_headers(sheet_name) if base.TEMPLATE_FILE == TEMPLATE_FILE else _template_headers(sheet_name)


def _template_headers(sheet_name):
    from openpyxl import load_workbook

    wb = load_workbook(TEMPLATE_FILE, read_only=True, data_only=True)
    ws = wb[sheet_name]
    return [_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]


def _new_row(headers):
    return {header: '' for header in headers}


def _set(row, field_name, value):
    return base._set(row, field_name, value)


def read_template_supplier_rows():
    rows = []
    for sheet_name in (SHEET_ARCHIVED, SHEET_UNARCHIVED):
        df = pd.read_excel(TEMPLATE_FILE, sheet_name=sheet_name, dtype=object)
        df = df.where(df.notna(), '')
        df['来源sheet'] = sheet_name
        rows.append(df)
    source = pd.concat(rows, ignore_index=True)
    source['合同编号'] = source['合同编号'].map(_text)
    source = source[source['合同编号'] != ''].copy()
    # 两个来源 sheet 只作为合同编号池: 按合同编号去重,其他字段均从数据库取。
    grouped = (
        source.groupby('合同编号', as_index=False)
        .agg({
            '来源sheet': lambda values: ';'.join(c.clean_text_values(values)),
        })
    )
    return grouped


def _query_by_codes(sql_template, codes):
    frames = []
    for batch in base._chunked(codes):
        sql = sql_template.format(placeholders=c.in_placeholders(batch))
        frames.append(c.query_db('FW', 'vspn_xtyy', sql, batch))
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def _resolve_contract_rows(contract_codes):
    codes = c.clean_text_values(contract_codes)
    if not codes:
        return pd.DataFrame()

    mcn_df = _query_by_codes(SOURCE_SQL_MCN_BY_CODES, codes)
    mcn_df['数据来源'] = '泛微(MCN)'
    htsp_df = _query_by_codes(SOURCE_SQL_HTSP_BY_CODES, codes)
    htsp_df['数据来源'] = '泛微(赛事)'

    mcn_codes = set(mcn_df.get('合同编号', pd.Series(dtype=object)).map(_text))
    if not htsp_df.empty:
        htsp_df = htsp_df[~htsp_df['合同编号'].map(_text).isin(mcn_codes)].copy()

    resolved_mcn = (
        base._timed('解析反贿赂MCN(%d行)' % len(mcn_df),
                    lambda: base.resolve_source_values(mcn_df, option_table=base.FW_TABLE))
        if not mcn_df.empty else mcn_df
    )
    resolved_htsp = (
        base._timed('解析反贿赂赛事(%d行)' % len(htsp_df),
                    lambda: base.resolve_source_values(htsp_df, option_table=base.FW_TABLE_HTSP))
        if not htsp_df.empty else htsp_df
    )
    merged = pd.concat([resolved_mcn, resolved_htsp], ignore_index=True)
    base._merge_attrs(merged, [resolved_mcn, resolved_htsp])
    if not merged.empty:
        merged['合同分类'] = FIXED_CATEGORY
        merged['收支类型'] = FIXED_PAY_TYPE
        merged['合同总额_解析'] = 0
        merged['收入总额_解析'] = 0
        merged['支出总额_解析'] = 0
        merged['合同总额_签名'] = 0
        merged['收入总额_签名'] = 0
        merged['支出总额_签名'] = 0
    return merged


def read_source():
    template_rows = read_template_supplier_rows()
    contract_df = _resolve_contract_rows(template_rows['合同编号'])
    by_code = {
        _text(row.get('合同编号')): row
        for row in contract_df.to_dict('records')
    }
    return template_rows, contract_df, by_code


def _counterparty_values(source, customer_info_map, supplier_info_map):
    names = []
    codes = []
    for supplier_id in c.parse_browser_ids(source.get('合同供应商ID')):
        info = supplier_info_map.get(supplier_id, {})
        names.append(base._first_non_blank(info.get('name'), info.get('source_name')))
        codes.append(info.get('code', ''))
    if not codes:
        for customer_id in c.parse_browser_ids(source.get('合同客户ID')):
            info = customer_info_map.get(customer_id, {})
            names.append(base._first_non_blank(info.get('name'), info.get('source_name')))
            codes.append(info.get('code', ''))
    return ';'.join(c.clean_text_values(names)), ';'.join(c.clean_text_values(codes))


def _our_party_values(source, company_info_map):
    names = []
    codes = []
    for company_id in c.parse_browser_ids(source.get('合同用印范围ID')):
        info = company_info_map.get(company_id, {})
        names.append(info.get('name', ''))
        codes.append(info.get('code', ''))
    return ';'.join(c.clean_text_values(names)), ';'.join(c.clean_text_values(codes))


def build_main_output(template_rows, contract_df, by_code, headers):
    company_info_map = contract_df.attrs.get('company_info_map', {})
    customer_info_map = contract_df.attrs.get('customer_info_map', {})
    supplier_info_map = contract_df.attrs.get('supplier_info_map', {})

    rows = []
    for template_row in template_rows.to_dict('records'):
        contract_number = _text(template_row.get('合同编号'))
        source = by_code.get(contract_number, {})
        row = _new_row(headers)
        counterparty_name, counterparty_code = _counterparty_values(source, customer_info_map, supplier_info_map)
        our_party_name, our_party_code = _our_party_values(source, company_info_map)

        _set(row, 'contract_number（合同编码）', contract_number)
        _set(row, 'contract_name（合同名称）', base._first_non_blank(source.get('合同标题'), FIXED_CATEGORY))
        _set(row, '泛微合同状态', _text(source.get('合同审批状态')))
        _set(row, '合同状态', '')
        _set(row, 'contractCategory(智书框架合同类型)', FIXED_CATEGORY)
        _set(row, '收支类型', FIXED_PAY_TYPE)
        _set(row, 'fixed_validity_code（合同期限类型）', FIXED_VALIDITY)
        _set(row, 'remark（合同说明）', _text(source.get('合同摘要'))[:150])
        _set(row, '签署日期', c.format_date(source.get('合同签订日期')))
        _set(row, '合同申请人（名字）', base._first_non_blank(source.get('申请人'), source.get('合同创建人')))
        _set(row, '合同申请人（user_id)', base._first_non_blank(source.get('申请人user_id'), source.get('合同创建人user_id')))
        _set(row, '签约性质', FIXED_SIGN_FORM)
        _set(row, 'custom_15_78cf503c57194e4fb8ad03ded1c4ad60（打印模式）', FIXED_PRINT_MODE)
        _set(row, '对方信息', counterparty_name)
        _set(row, '对方信息id', counterparty_code)
        _set(row, '我方信息', our_party_name)
        _set(row, '我方信息id', our_party_code)
        _set(row, '合同文本（附件）', '')
        _set(row, '合同附件（附件）', '')
        _set(row, '其他附件（附件）', '')
        rows.append(row)
    return pd.DataFrame(rows, columns=headers)


def _anti_attachment_download_root():
    configured = os.getenv(base.ATTACHMENT_DOWNLOAD_ROOT_ENV, '').strip()
    if configured:
        return Path(configured)
    return OUTPUT_DIR / f'反商业贿赂协议合同附件_{base.DATE_SUFFIX}'


def build_contract_attachment_manifest(contract_df):
    if contract_df.empty:
        return pd.DataFrame(), pd.DataFrame()

    old_output_dir = base.OUTPUT_DIR
    try:
        base.OUTPUT_DIR = OUTPUT_DIR
        manifest_df, missing_df = base.build_contract_attachment_manifest(contract_df)
    finally:
        base.OUTPUT_DIR = old_output_dir

    if manifest_df.empty:
        return manifest_df, missing_df

    download_root = _anti_attachment_download_root()
    used_paths = set()
    manifest_df = manifest_df.copy()
    for index, row in manifest_df.iterrows():
        contract_dir = base._sanitize_path_part(
            row.get('contract_number（合同编码）'),
            f'contract_{_text(row.get("合同ID"))}',
        )
        target_sheet = _text(row.get('attachment_sheet')) or base.SHEET_OTHER_ATTACHMENT
        target_dir = download_root / contract_dir / base._sanitize_path_part(target_sheet, target_sheet)
        target_name = base._build_target_filename(row.get('attachment_name'), row.get('imagefileid'))
        target_path = target_dir / target_name
        if target_path in used_paths:
            target_path = target_dir / f'{target_path.stem}_{row.get("imagefileid")}{target_path.suffix}'
        used_paths.add(target_path)
        manifest_df.at[index, 'target_path'] = str(target_path)
    return manifest_df, missing_df


def write_outputs(main_output_df, contract_df, template_rows):
    output_file = base._write_template_sheets_with_fallback(
        TEMPLATE_FILE,
        OUTPUT_FILE,
        {SHEET_TARGET: main_output_df},
        extra_sheets={
            '反贿赂来源核对': template_rows,
            '泛微合同核对': base._audit_df(contract_df, [
                '数据来源', '合同编号', '合同标题', '合同签署状态', '合同审批状态',
                '申请人', '申请人工号', '申请人user_id',
                '合同创建人', '合同创建人user_id', '合同供应商名称', '合同摘要',
            ]),
        },
    )
    print('已写出:', output_file)
    return output_file


def write_exception_outputs(template_rows, contract_df, main_output_df, manifest_df, missing_df):
    found_codes = set(contract_df.get('合同编号', pd.Series(dtype=object)).map(_text))
    missing_contracts = template_rows[~template_rows['合同编号'].map(_text).isin(found_codes)].copy()
    exception_sheets = {
        '泛微未匹配合同': missing_contracts,
        '字段填充概览': pd.DataFrame([
            {'字段': column, '填充数': int((main_output_df[column].astype(str).str.strip() != '').sum()), '总数': len(main_output_df)}
            for column in main_output_df.columns
        ]),
        '固定值说明': pd.DataFrame([
            {'字段': 'contractCategory(智书框架合同类型)', '固定值': FIXED_CATEGORY},
            {'字段': '收支类型', '固定值': FIXED_PAY_TYPE},
            {'字段': 'fixed_validity_code（合同期限类型）', '固定值': FIXED_VALIDITY},
            {'字段': '签约性质', '固定值': FIXED_SIGN_FORM},
            {'字段': 'custom_15_78cf503c57194e4fb8ad03ded1c4ad60（打印模式）', '固定值': FIXED_PRINT_MODE},
            {'字段': '合同状态', '固定值': '留空,由法务填写'},
            {'字段': '合同文本/合同附件/其他附件', '固定值': '留空,附件通过下载清单落盘'},
        ]),
        '合同附件下载清单': manifest_df,
        '合同附件DOCID_缺失映射': missing_df,
    }
    exception_file = base._write_exceptions_with_fallback(EXCEPTION_FILE, exception_sheets)
    if exception_file:
        print('已写出:', exception_file)
    manifest_file = base._write_exceptions_with_fallback(MANIFEST_FILE, {
        '合同附件下载清单': manifest_df,
        '合同附件DOCID_缺失映射': missing_df,
    })
    if manifest_file:
        print('已写出:', manifest_file)
    return exception_file, manifest_file


def build_outputs():
    headers = _template_headers(SHEET_TARGET)
    template_rows, contract_df, by_code = base._timed('阶段1: 读取反贿赂来源+泛微解析', read_source)
    main_output_df = base._timed(
        '阶段2: 构建反商业贿赂协议sheet',
        lambda: build_main_output(template_rows, contract_df, by_code, headers),
    )
    manifest_df, missing_df = base._timed(
        '阶段3: 构建附件下载清单',
        lambda: build_contract_attachment_manifest(contract_df),
    )
    return template_rows, contract_df, main_output_df, manifest_df, missing_df


def run():
    template_rows, contract_df, main_output_df, manifest_df, missing_df = build_outputs()
    write_outputs(main_output_df, contract_df, template_rows)
    write_exception_outputs(template_rows, contract_df, main_output_df, manifest_df, missing_df)
    print('[反商业贿赂协议] 来源行数:', len(template_rows))
    print('[反商业贿赂协议] 泛微命中合同:', len(contract_df))
    print('[反商业贿赂协议] 输出行数:', len(main_output_df))
    print('[反商业贿赂协议] 附件清单行数:', len(manifest_df))


if __name__ == '__main__':
    run()
