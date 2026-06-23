# -*- coding: utf-8 -*-
"""合同迁移 —— 智书一般流程(DB 直连版)。

按「智书合同字段-一般流程.xlsx」生成一般流程导入模板:
    1. 字段模板
    2. 关联合同 / 相关单据-订单信息 / 采购申请 / 订单信息明细
    3. 对方信息 / 我方主体列表
    4. 付款计划 / 收款计划
    5. 合同附件(仅写附件名称) / 其他附件(模板保留); 文件下载请运行 contract_general_attachments_db

跑法:在项目根执行  python run.py contract_general_db
"""
import html
import os
import re
import sys
import time
import urllib.error
from functools import lru_cache
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

if __package__ is None or __package__ == '':
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from etl import common as c
from etl.tasks.ap_prepayment_opening_db import build_fw_project_code_map_for_ids


# ============================ 文件 / 模板 ============================
TASK_NAME = 'contract_general_db'
TEMPLATE_DIR = c.TPL_DIR / 'contract_anchor_db'
OUTPUT_DIR = c.OUT_DIR / TASK_NAME
DATE_SUFFIX = c.today_suffix()

TEMPLATE_FILE = TEMPLATE_DIR / '智书合同字段-一般流程.xlsx'
RULE_CSV = c.RULES_DIR / '业财项目_数据映射规则 - 合同数据映射规则-for法务.csv'
OUTPUT_FILE = OUTPUT_DIR / f'智书合同字段_一般流程_{DATE_SUFFIX}.xlsx'
EXCEPTION_FILE = OUTPUT_DIR / f'未匹配清单_一般流程_{DATE_SUFFIX}.xlsx'

SHEET_MAIN = '字段模板'
SHEET_OPTIONS = '选项'
SHEET_RELATION = '关联合同'
SHEET_RELATED_ORDER = '相关单据-订单信息'
SHEET_PURCHASE_REQUEST = '采购申请'
SHEET_ORDER_DETAIL = '订单信息明细'
SHEET_COUNTERPARTY = '对方信息'
SHEET_OUR_PARTY = '我方主体列表'
SHEET_PAYMENT_PLAN = '付款计划'
SHEET_COLLECTION_PLAN = '收款计划'
SHEET_CONTRACT_ATTACHMENT = '合同附件'
SHEET_OTHER_ATTACHMENT = '其他附件'

OUTPUT_SHEETS = (
    SHEET_MAIN,
    SHEET_RELATION,
    SHEET_RELATED_ORDER,
    SHEET_PURCHASE_REQUEST,
    SHEET_ORDER_DETAIL,
    SHEET_COUNTERPARTY,
    SHEET_OUR_PARTY,
    SHEET_PAYMENT_PLAN,
    SHEET_COLLECTION_PLAN,
    SHEET_CONTRACT_ATTACHMENT,
    SHEET_OTHER_ATTACHMENT,
)

FW_TABLE = 'uf_htk'
MIGRATION_STATUS_CODES = (1, 2)
ANCHOR_CONTRACT_TYPE_CODE = 3

DEFAULT_PROPERTY_TYPE = '固定总价'
DEFAULT_VALIDITY_TYPE = '固定期限'
DEFAULT_ACCEPTANCE_REQUIRED = '否'
DEFAULT_PRINT_MODE = '黑白单面打印'
DEFAULT_INVOICE_TYPE = '增值税专用发票'
# 税率按百分数口径(13 = 13%); 赛事源直接取 srsl/zcsl, 仅在源无值时用此兜底默认(6%)。
DEFAULT_TAX_RATE = '6'
DEFAULT_TAX_ITEM = '其他'
DEFAULT_BANK_FEE_BEARER = '各自承担'
DEFAULT_FIRST_SEAL_PARTY = '我方'
DEFAULT_SIGN_FORM = '纸质签约'
DEFAULT_SEAL_NUMBER = 2
DEFAULT_PREPAID = '否'
DEFAULT_PAYMENT_NATURE = '一般付款'

ATTACHMENT_COOKIE_ENV = 'WEAVER_CONTRACT_ATTACHMENT_COOKIE'
ATTACHMENT_BASE_URL_ENV = 'WEAVER_CONTRACT_ATTACHMENT_BASE_URL'
ATTACHMENT_LOGIN_USERID_ENV = 'WEAVER_CONTRACT_ATTACHMENT_LOGIN_USERID'
ATTACHMENT_AUTHORIZEMODE_ID_ENV = 'WEAVER_CONTRACT_ATTACHMENT_AUTHORIZEMODE_ID'
ATTACHMENT_AUTHORIZEFIELD_ID_ENV = 'WEAVER_CONTRACT_ATTACHMENT_AUTHORIZEFIELD_ID'
ATTACHMENT_DOWNLOAD_ROOT_ENV = 'WEAVER_CONTRACT_ATTACHMENT_DOWNLOAD_ROOT'
ATTACHMENT_DOWNLOAD_ENABLED_ENV = 'WEAVER_CONTRACT_ATTACHMENT_DOWNLOAD_ENABLED'
DEFAULT_ATTACHMENT_BASE_URL = 'http://oaportal.heroesports.com'
DEFAULT_ATTACHMENT_LOGIN_USERID = '3837'
DEFAULT_ATTACHMENT_AUTHORIZEMODE_ID = '5'
DEFAULT_ATTACHMENT_AUTHORIZEFIELD_ID = '6461'
ATTACHMENT_TYPE_DRAFT = '合同初稿'
ATTACHMENT_TYPE_REVISED = '合同修订稿'
ATTACHMENT_TYPE_SIGNED = '合同签署稿'
ATTACHMENT_TYPE_EFFECTIVE = '合同生效稿'

# 审批流程表单(formtable_main_*)里的三个稿件字段 -> 固定归类。
# OA「合同相关」显示名: 合同修订稿/合同签署版/合同生效版。
FORM_DOC_FIELD_TYPES = (
    ('htsxb', ATTACHMENT_TYPE_EFFECTIVE),  # 合同生效版(= uf_htk.htqdg)
    ('htqsb', ATTACHMENT_TYPE_SIGNED),     # 合同签署版
    ('htxdg', ATTACHMENT_TYPE_REVISED),    # 合同修订稿
)
# 单字段表单(如 formtable_main_26)没有上面三字段, 退回单一稿件字段, 归生效稿。
FORM_DOC_FALLBACK_FIELD = 'htqdg'

# 赛事源 uf_htsp 的稿件字段(逗号分隔 docid) -> 固定归类。
SAISHI_DOC_FIELD_TYPES = (
    ('赛事生效稿DOCID', ATTACHMENT_TYPE_EFFECTIVE),
    ('赛事签署稿DOCID', ATTACHMENT_TYPE_SIGNED),
    ('赛事初稿DOCID', ATTACHMENT_TYPE_DRAFT),
)
INVALID_PATH_CHARS = re.compile(r'[<>:"/\\|?*]+')
DOC_ID_SPLITTER = re.compile(r'[,\uff0c]+')


# ============================ 泛微源 SQL ============================
SOURCE_SQL = """
SELECT
    h.id AS `ID`,
    h.htbh AS `合同编号`,
    h.htbt AS `合同标题`,
    h.htlx AS `合同类型ID`,
    h.htejlx AS `合同二级类型ID`,
    h.htzt AS `合同签署状态ID`,
    h.htlc AS `合同流程ID`,
    h.htcjrq AS `合同创建日期`,
    h.htqdrq AS `合同签订日期`,
    h.htyxqqssj AS `合同有效期起始时间`,
    h.htyxqjzsj AS `合同有效期截止时间`,
    h.htzhry AS `合同执行人员ID`,
    h.htyyfw AS `合同用印范围ID`,
    h.htkh AS `合同客户ID`,
    h.htgys AS `合同供应商ID`,
    h.htje AS `合同金额`,
    h.htyjsr AS `合同预计收入`,
    h.htyjzc AS `合同预计支出`,
    h.htzy AS `合同摘要`,
    h.htqdg AS `合同附件DOCID`,
    h.cbzx AS `成本中心ID`,
    h.glkjxy AS `关联框架协议ID`,
    h.bglx AS `变更类型ID`,
    h.bczzbhsc AS `补充/终止编号生成`,
    h.htszxmbh AS `合同所属项目编号ID`,
    h.htszxm AS `合同所属项目`,
    NULL AS `收入税率`,
    NULL AS `支出税率`,
    NULL AS `押金`,
    NULL AS `保证金`,
    NULL AS `采购申请单ID`,
    rb.workflowid AS `流程类型ID`,
    wb.workflowname AS `流程名称`
FROM uf_htk h
LEFT JOIN workflow_requestbase rb ON rb.requestid = h.htlc
LEFT JOIN workflow_base wb ON wb.id = rb.workflowid
WHERE (h.htlx <> %(anchor_contract_type_code)s OR h.htlx IS NULL)
  AND h.htzt IN %(migration_status_codes)s
ORDER BY h.htbh, h.id
"""

STATS_SQL = """
SELECT
    COUNT(*) AS all_count,
    SUM(CASE WHEN htlx = %(anchor_contract_type_code)s THEN 1 ELSE 0 END) AS anchor_type_count,
    SUM(CASE
        WHEN (htlx <> %(anchor_contract_type_code)s OR htlx IS NULL)
         AND htzt IN %(migration_status_codes)s
        THEN 1 ELSE 0 END) AS kept_count,
    SUM(CASE
        WHEN (htlx <> %(anchor_contract_type_code)s OR htlx IS NULL)
         AND (htzt NOT IN %(migration_status_codes)s OR htzt IS NULL)
        THEN 1 ELSE 0 END) AS excluded_status_count
FROM uf_htk
"""

STATUS_BREAKDOWN_SQL = """
SELECT htzt AS `合同签署状态ID`, COUNT(*) AS `合同数`
FROM uf_htk
WHERE htlx <> %(anchor_contract_type_code)s OR htlx IS NULL
GROUP BY htzt
ORDER BY htzt
"""

# ---- 赛事源: 合同审批台账 uf_htsp(独立码表, htzt=1 归档) ----
FW_TABLE_HTSP = 'uf_htsp'
HTSP_MIGRATION_STATUS_CODES = (1,)  # 1=归档

# 字段对齐到 uf_htk 源的统一列名, 复用下游 resolve/builder。
SOURCE_SQL_HTSP = """
SELECT
    h.id AS `ID`,
    h.htbh AS `合同编号`,
    h.htmc AS `合同标题`,
    h.htlx AS `合同类型ID`,
    h.htejfl AS `合同二级类型ID`,
    h.htzt AS `合同签署状态ID`,
    h.lcqqid AS `合同流程ID`,
    h.sqrq AS `合同创建日期`,
    h.qdsj AS `合同签订日期`,
    h.htksrq AS `合同有效期起始时间`,
    h.htjsrq AS `合同有效期截止时间`,
    h.htzhr AS `合同执行人员ID`,
    h.gszt AS `合同用印范围ID`,
    h.kh AS `合同客户ID`,
    h.gys AS `合同供应商ID`,
    NULL AS `合同金额`,
    h.srje AS `合同预计收入`,
    h.zcje AS `合同预计支出`,
    h.htzy AS `合同摘要`,
    h.htsxg AS `合同附件DOCID`,
    h.htcg AS `赛事初稿DOCID`,
    h.htqsg AS `赛事签署稿DOCID`,
    h.htsxg AS `赛事生效稿DOCID`,
    h.cbzx AS `成本中心ID`,
    h.glht AS `关联框架协议ID`,
    NULL AS `变更类型ID`,
    NULL AS `补充/终止编号生成`,
    h.xmbh AS `合同所属项目编号ID`,
    h.xmmc AS `合同所属项目`,
    h.srsl AS `收入税率`,
    h.zcsl AS `支出税率`,
    h.yj AS `押金`,
    h.bzj AS `保证金`,
    h.cgsqddx AS `采购申请单ID`,
    rb.workflowid AS `流程类型ID`,
    wb.workflowname AS `流程名称`
FROM uf_htsp h
LEFT JOIN workflow_requestbase rb ON rb.requestid = h.lcqqid
LEFT JOIN workflow_base wb ON wb.id = rb.workflowid
WHERE h.htzt IN %(htsp_status_codes)s
ORDER BY h.htbh, h.id
"""

EXPECTED_FW_FIELDS = {
    '': {
        'htbh': '合同编号',
        'htbt': '合同标题',
        'htlx': '合同类型',
        'htejlx': '合同二级类型',
        'htzt': '合同签署状态',
        'htlc': '合同流程',
        'htqdrq': '合同签订日期',
        'htyxqqssj': '合同有效期起始时间',
        'htyxqjzsj': '合同有效期截止时间',
        'htzhry': '合同执行人员',
        'htyyfw': '合同用印范围',
        'htkh': '合同客户',
        'htgys': '合同供应商',
        'htje': '合同金额',
        'htyjsr': '合同预计收入',
        'htyjzc': '合同预计支出',
        'htzy': '合同摘要',
        'htqdg': '合同签定稿',
        'glkjxy': '关联框架协议',
        'bglx': '变更类型',
        'bczzbhsc': '补充/终止编号生成',
        'htszxmbh': '合同所属项目编号',
    },
}


# ============================ 小工具 ============================
def _query_fw(sql):
    return c.query_db('FW', 'vspn_xtyy', sql, {
        'anchor_contract_type_code': ANCHOR_CONTRACT_TYPE_CODE,
        'migration_status_codes': MIGRATION_STATUS_CODES,
    })


def _text(value):
    if pd.isna(value):
        return ''
    text = str(value).strip()
    return '' if text in ('', 'nan', 'None', 'NaT') else text


def _first_non_blank(*values):
    for value in values:
        text = _text(value)
        if text:
            return text
    return ''


def _number(value, default=0.0):
    if pd.isna(value) or _text(value) == '':
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _round_amount(value):
    return round(_number(value), 2)


def _format_tax_rate(value):
    """税率按源库百分数口径输出, 去掉多余的尾随零: 13.00->'13'、1.0000->'1'、13.5->'13.5'。

    空值返回 '', 由调用方决定是否回退默认税率。
    """
    text = _text(value)
    if not text:
        return ''
    try:
        number = float(text)
    except (TypeError, ValueError):
        return text
    return f'{number:.4f}'.rstrip('0').rstrip('.') or '0'


@lru_cache(maxsize=8192)
def _normalize_cached(text):
    return re.sub(r'\s+', '', text)


def _normalize_field_name(value):
    # 热路径: 模板表头/字段名在 5 万行 × 数十字段上反复归一化, 缓存避免重复正则。
    return _normalize_cached(_text(value))


def _first_browser_id(value):
    ids = c.parse_browser_ids(value)
    return ids[0] if ids else ''


def _lookup_first_browser_value(mapping, value):
    for item_id in c.parse_browser_ids(value):
        mapped = mapping.get(item_id, '')
        if mapped:
            return mapped
    return ''


def _sanitize_path_part(value, fallback):
    text = _text(value) or fallback
    text = html.unescape(text)
    text = INVALID_PATH_CHARS.sub('_', text)
    text = text.replace('\r', ' ').replace('\n', ' ').strip(' .')
    return (text or fallback)[:180]


def _to_int_id(value):
    text = _text(value)
    if not text:
        return None
    try:
        return int(float(text)) if re.fullmatch(r'\d+(?:\.0+)?', text) else None
    except (TypeError, ValueError):
        return None


def _split_docids(raw_value):
    text = _text(raw_value)
    if not text:
        return []
    result = []
    for part in DOC_ID_SPLITTER.split(text):
        part = part.strip()
        docid = _to_int_id(part)
        if docid is not None:
            result.append(docid)
    return result


def _timed(label, func):
    """跑一个步骤并打印耗时(秒)。用于定位性能瓶颈。"""
    start = time.perf_counter()
    print(f'[计时] ▶ {label} ...', flush=True)
    result = func()
    print(f'[计时] ✓ {label}: {time.perf_counter() - start:.1f}s', flush=True)
    return result


def _chunked(items, size=5000):
    batch = []
    for item in items:
        batch.append(item)
        if len(batch) >= size:
            yield batch
            batch = []
    if batch:
        yield batch


def _build_target_filename(filename, imagefileid):
    original = _sanitize_path_part(filename, str(imagefileid))
    if '.' not in Path(original).name:
        return f'{original}_{imagefileid}'
    return original


def _attachment_download_root():
    configured = os.getenv(ATTACHMENT_DOWNLOAD_ROOT_ENV, '').strip()
    if configured:
        return Path(configured)
    return OUTPUT_DIR / f'一般流程合同附件_{DATE_SUFFIX}'


def _attachment_download_enabled(cookie):
    flag = os.getenv(ATTACHMENT_DOWNLOAD_ENABLED_ENV, '').strip().lower()
    if flag in ('0', 'false', 'n', 'no', '否'):
        return False
    return bool(_text(cookie))


def _build_attachment_referer(base_url, imagefileid, docid):
    query = urllib.parse.urlencode({
        'pdfimagefileid': imagefileid,
        'authorizemodeId': os.getenv(ATTACHMENT_AUTHORIZEMODE_ID_ENV, DEFAULT_ATTACHMENT_AUTHORIZEMODE_ID),
        'authorizefieldid': os.getenv(ATTACHMENT_AUTHORIZEFIELD_ID_ENV, DEFAULT_ATTACHMENT_AUTHORIZEFIELD_ID),
        'docisLock': 'false',
        'formmode_authorize': 'formmode_authorize',
        'authorizeformmodebillId': docid,
        'f_weaver_belongto_usertype': '0',
        'f_weaver_belongto_userid': os.getenv(ATTACHMENT_LOGIN_USERID_ENV, DEFAULT_ATTACHMENT_LOGIN_USERID),
        'canDownload': 'true',
        'canPrint': 'true',
    })
    return f'{base_url.rstrip("/")}/docs/pdfview3.x/web/pdfViewer.jsp?&{query}'


def _download_attachment_file(meta, cookie):
    target_path = Path(meta['target_path'])
    if target_path.exists() and target_path.stat().st_size > 0:
        return 'skipped_exists', ''
    target_path.parent.mkdir(parents=True, exist_ok=True)
    base_url = os.getenv(ATTACHMENT_BASE_URL_ENV, DEFAULT_ATTACHMENT_BASE_URL)
    url = f'{base_url.rstrip("/")}/weaver/weaver.file.FileDownload?fileid={meta["imagefileid"]}'
    headers = {
        'Accept': '*/*',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Connection': 'keep-alive',
        'User-Agent': (
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/146.0.0.0 Safari/537.36'
        ),
        'Cookie': cookie,
        'Referer': _build_attachment_referer(base_url, meta['imagefileid'], meta['docid']),
    }
    temp_path = target_path.with_suffix(target_path.suffix + '.part')
    try:
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=120) as resp:
            data = resp.read()
            final_url = resp.geturl()
            content_type = resp.headers.get('Content-Type', '')
            if 'login' in final_url.lower():
                raise RuntimeError(f'跳转到登录页: {final_url}')
            if 'text/html' in content_type.lower() and not _text(meta['attachment_name']).lower().endswith(('.html', '.htm')):
                snippet = data[:200].decode('utf-8', errors='ignore')
                raise RuntimeError(f'返回 HTML,疑似无权限或会话失效: {snippet}')
            with open(temp_path, 'wb') as file:
                file.write(data)
        os.replace(temp_path, target_path)
        return 'downloaded', ''
    except (OSError, urllib.error.URLError, RuntimeError) as exc:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)
        return 'failed', str(exc)


def _load_attachment_maps(docids):
    """一次 JOIN 取齐附件三表, 返回 (docimage_map, imagefile_map, docdetail_map)。

    原先按 docid/imagefileid 分三趟批量查(docimagefile / imagefile / docdetail),
    对远程库是几百次往返。这里改成每个 docid 批次一条
    docimagefile ⋈ imagefile ⋈ docdetail 的 JOIN, 配合更大的批量, 往返从数百降到十几次。
    """
    docimage_map, imagefile_map, docdetail_map = {}, {}, {}
    docids = [docid for docid in docids if docid is not None]
    if not docids:
        return docimage_map, imagefile_map, docdetail_map
    for batch in _chunked(sorted(set(docids))):
        df = c.query_db(
            'FW',
            'vspn_xtyy',
            'SELECT d.DOCID, d.IMAGEFILEID, d.IMAGEFILENAME, d.DOCFILETYPE, d.VERSIONID, '
            'i.IMAGEFILENAME AS IMG_IMAGEFILENAME, i.IMAGEFILETYPE, i.FILESIZE, '
            'dd.DOCSUBJECT, dd.DOCEXTENDNAME, dd.DOCCREATEDATE, dd.DOCCREATETIME '
            'FROM docimagefile d '
            'LEFT JOIN imagefile i ON i.IMAGEFILEID = d.IMAGEFILEID '
            'LEFT JOIN docdetail dd ON dd.ID = d.DOCID '
            f'WHERE d.DOCID IN ({c.in_placeholders(batch)}) '
            'ORDER BY d.DOCID, d.IMAGEFILEID',
            batch,
        )
        for _, row in df.iterrows():
            docid = _to_int_id(row['DOCID'])
            imagefileid = _to_int_id(row['IMAGEFILEID'])
            if docid is None or imagefileid is None:
                continue
            docimage_map.setdefault(docid, []).append({
                'docid': docid,
                'imagefileid': imagefileid,
                'filename': _text(row.get('IMAGEFILENAME')),
                'docfiletype': _text(row.get('DOCFILETYPE')),
                'versionid': _text(row.get('VERSIONID')),
            })
            if imagefileid not in imagefile_map:
                imagefile_map[imagefileid] = {
                    'filename': _text(row.get('IMG_IMAGEFILENAME')),
                    'imagefiletype': _text(row.get('IMAGEFILETYPE')),
                    'filesize': _text(row.get('FILESIZE')),
                }
            if docid not in docdetail_map:
                docdetail_map[docid] = {
                    'subject': html.unescape(_text(row.get('DOCSUBJECT'))),
                    'extend_name': _text(row.get('DOCEXTENDNAME')),
                    'created_date': _text(row.get('DOCCREATEDATE')),
                    'created_time': _text(row.get('DOCCREATETIME')),
                }
    return docimage_map, imagefile_map, docdetail_map


def _load_workflow_shared_docids(request_ids):
    mapping = {}
    request_ids = c.clean_codes(request_id for request_id in request_ids if _text(request_id))
    if not request_ids:
        return mapping
    for batch in _chunked(sorted(set(request_ids))):
        df = c.query_db(
            'FW',
            'vspn_xtyy',
            'SELECT REQUESTID, DOCID, NODEID, id AS share_id '
            'FROM workflow_docshareinfo '
            f'WHERE REQUESTID IN ({c.in_placeholders(batch)}) '
            'ORDER BY REQUESTID, id',
            batch,
        )
        for _, row in df.iterrows():
            request_id = c.format_code(row['REQUESTID'])
            docid = _to_int_id(row['DOCID'])
            if not (request_id and docid is not None):
                continue
            mapping.setdefault(request_id, []).append({
                'docid': docid,
                'nodeid': _text(row.get('NODEID')),
                'share_id': _to_int_id(row.get('share_id')) or 0,
            })
    for request_id, items in list(mapping.items()):
        if not items:
            continue
        latest = max(items, key=lambda item: item.get('share_id') or 0)
        latest_nodeid = latest.get('nodeid')
        if latest_nodeid:
            mapping[request_id] = [item for item in items if item.get('nodeid') == latest_nodeid]
    return mapping


def _classify_attachment_type(doc_row, image_info, doc_info, explicit_type=''):
    # 主取数(审批流程表单三字段)已带固定归类, 直接采用。
    if explicit_type:
        return explicit_type
    # 仅 docshareinfo 兜底走到这里: 按文件名关键字判断, 默认归初稿/未分类。
    filename = _first_non_blank(doc_row.get('filename'), image_info.get('filename'), doc_info.get('subject'))
    normalized = _normalize_field_name(filename).lower()
    if '生效' in normalized:
        return ATTACHMENT_TYPE_EFFECTIVE
    if '签署' in normalized or '签定' in normalized or '签订' in normalized:
        return ATTACHMENT_TYPE_SIGNED
    return ATTACHMENT_TYPE_DRAFT


def _resolve_request_form_tables(request_ids):
    """requestid -> 审批流程表单数据表名(formtable_main_*)。"""
    mapping = {}
    request_ids = c.clean_codes(request_id for request_id in request_ids if _text(request_id))
    if not request_ids:
        return mapping
    for batch in _chunked(sorted(set(request_ids))):
        df = c.query_db(
            'FW',
            'vspn_xtyy',
            'SELECT rb.requestid AS rid, bill.tablename AS tablename '
            'FROM workflow_requestbase rb '
            'JOIN workflow_base wb ON wb.id = rb.workflowid '
            'JOIN workflow_bill bill ON bill.id = wb.formid '
            f'WHERE rb.requestid IN ({c.in_placeholders(batch)})',
            batch,
        )
        for _, row in df.iterrows():
            request_id = c.format_code(row['rid'])
            tablename = _text(row.get('tablename'))
            if request_id and tablename:
                mapping[request_id] = tablename
    return mapping


def _existing_columns(tablename, candidates):
    df = c.query_db(
        'FW',
        'vspn_xtyy',
        'SELECT COLUMN_NAME FROM information_schema.COLUMNS '
        f'WHERE TABLE_NAME = %s AND COLUMN_NAME IN ({c.in_placeholders(candidates)})',
        [tablename, *candidates],
    )
    return {str(name).lower() for name in df['COLUMN_NAME'].tolist()}


def _load_form_doc_fields(request_form_tables):
    """requestid -> [(attachment_type, docid), ...],来自审批流程表单的稿件字段。"""
    result = {}
    requests_by_table = {}
    for request_id, tablename in request_form_tables.items():
        requests_by_table.setdefault(tablename, []).append(request_id)

    typed_fields = [field for field, _ in FORM_DOC_FIELD_TYPES]
    type_by_field = dict(FORM_DOC_FIELD_TYPES)
    for tablename, request_ids in requests_by_table.items():
        columns = _existing_columns(tablename, [*typed_fields, FORM_DOC_FALLBACK_FIELD])
        present_typed = [field for field in typed_fields if field in columns]
        use_fallback = not present_typed and FORM_DOC_FALLBACK_FIELD in columns
        select_fields = present_typed or ([FORM_DOC_FALLBACK_FIELD] if use_fallback else [])
        if not select_fields:
            continue
        select_clause = ', '.join(['requestid AS rid', *select_fields])
        for batch in _chunked(sorted(set(request_ids))):
            df = c.query_db(
                'FW',
                'vspn_xtyy',
                f'SELECT {select_clause} FROM {tablename} '
                f'WHERE requestid IN ({c.in_placeholders(batch)})',
                batch,
            )
            for _, row in df.iterrows():
                request_id = c.format_code(row['rid'])
                if not request_id:
                    continue
                entries = []
                for field in select_fields:
                    attachment_type = type_by_field.get(field, ATTACHMENT_TYPE_EFFECTIVE)
                    for docid in _split_docids(row.get(field)):
                        entries.append((attachment_type, docid))
                if entries:
                    result[request_id] = entries
    return result


def _collect_attachment_docrefs(source_df):
    raw_effective_docids = {}
    docrefs_by_contract = {}
    all_docids = []

    for source in source_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        raw_effective_docids[contract_number] = _text(source.get('合同附件DOCID', ''))
        docrefs_by_contract.setdefault(contract_number, [])

    contracts_with_form_docs = set()

    # 赛事(uf_htsp): 附件 docid 直接挂在 htcg/htqsg/htsxg 字段, 各自固定归类。
    is_saishi = source_df.get('数据来源', pd.Series('', index=source_df.index)).map(_text) == '泛微(赛事)'
    for source in source_df[is_saishi].to_dict('records'):
        contract_number = _text(source['合同编号'])
        for field, attachment_type in SAISHI_DOC_FIELD_TYPES:
            for docid in _split_docids(source.get(field)):
                docrefs_by_contract[contract_number].append({
                    'docid': docid,
                    'source': 'htsp_field',
                    'attachment_type': attachment_type,
                    'share_id': 0,
                })
                all_docids.append(docid)
                contracts_with_form_docs.add(contract_number)

    # MCN(uf_htk): 附件来自审批流程表单三字段(修订稿/签署版/生效版), 各自固定归类。
    mcn_df = source_df[~is_saishi]
    request_form_tables = _resolve_request_form_tables(mcn_df['合同流程ID'].map(c.format_code))
    form_doc_fields = _load_form_doc_fields(request_form_tables)
    for source in mcn_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        request_id = c.format_code(source.get('合同流程ID'))
        for attachment_type, docid in form_doc_fields.get(request_id, []):
            docrefs_by_contract[contract_number].append({
                'docid': docid,
                'source': 'form_field',
                'attachment_type': attachment_type,
                'share_id': 0,
            })
            all_docids.append(docid)
            contracts_with_form_docs.add(contract_number)

    # 兜底: 仅当合同三字段全空时, 才回退 workflow_docshareinfo(按文件名启发式归类)。
    fallback_request_ids = [
        c.format_code(source.get('合同流程ID'))
        for source in mcn_df.to_dict('records')
        if _text(source['合同编号']) not in contracts_with_form_docs
    ]
    workflow_docids = _load_workflow_shared_docids(fallback_request_ids)
    for source in mcn_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        if contract_number in contracts_with_form_docs:
            continue
        request_id = c.format_code(source.get('合同流程ID'))
        for item in workflow_docids.get(request_id, []):
            docid = item['docid']
            docrefs_by_contract[contract_number].append({
                'docid': docid,
                'source': 'workflow_docshareinfo',
                'attachment_type': '',
                'nodeid': item.get('nodeid', ''),
                'share_id': item.get('share_id', 0),
            })
            all_docids.append(docid)

    return docrefs_by_contract, raw_effective_docids, all_docids


def build_contract_attachment_manifest(source_df):
    docrefs_by_contract, raw_effective_docids, all_docids = _timed(
        '  ├─附件docref收集(表单/赛事字段)', lambda: _collect_attachment_docrefs(source_df))
    print(f'[计时]   ├─附件 docid 总数: {len(all_docids)}', flush=True)
    docimage_map, imagefile_map, docdetail_map = _timed(
        '  ├─附件三表JOIN取数(_load_attachment_maps)', lambda: _load_attachment_maps(all_docids))

    rows = []
    missing_rows = []
    seen = set()
    used_paths = set()
    download_root = _attachment_download_root()

    _t0 = time.perf_counter()
    for source in source_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        contract_dir = _sanitize_path_part(contract_number, f'contract_{_text(source.get("ID"))}')
        for docref in docrefs_by_contract.get(contract_number, []):
            docid = docref['docid']
            doc_rows = docimage_map.get(docid, [])
            if not doc_rows:
                missing_rows.append({
                    'contract_number（合同编码）': contract_number,
                    '合同ID': _text(source.get('ID')),
                    '合同名称': _text(source.get('合同标题')),
                    'raw_docids': raw_effective_docids.get(contract_number, ''),
                    'docid': docid,
                    'source': docref.get('source', ''),
                    'status': 'missing_docimagefile',
                    'error': 'docimagefile 无记录',
                })
                continue

            for doc_row in doc_rows:
                imagefileid = doc_row['imagefileid']
                dedupe_key = (contract_number, docid, imagefileid)
                if dedupe_key in seen:
                    continue
                seen.add(dedupe_key)

                image_info = imagefile_map.get(imagefileid, {})
                doc_info = docdetail_map.get(docid, {})
                attachment_name = html.unescape(_first_non_blank(
                    doc_row.get('filename'),
                    image_info.get('filename'),
                    doc_info.get('subject'),
                    str(imagefileid),
                ))
                attachment_type = _classify_attachment_type(
                    doc_row, image_info, doc_info, docref.get('attachment_type', ''))
                target_dir = download_root / contract_dir / _sanitize_path_part(attachment_type, attachment_type)
                target_name = _build_target_filename(attachment_name, imagefileid)
                target_path = target_dir / target_name
                if target_path in used_paths:
                    target_path = target_dir / f'{target_path.stem}_{imagefileid}{target_path.suffix}'
                used_paths.add(target_path)
                rows.append({
                    'contract_number（合同编码）': contract_number,
                    '合同ID': _text(source.get('ID')),
                    '合同名称': _text(source.get('合同标题')),
                    'attachment_type': attachment_type,
                    'raw_docids': raw_effective_docids.get(contract_number, ''),
                    'docid': docid,
                    'imagefileid': imagefileid,
                    'attachment_name': attachment_name,
                    'imagefiletype': image_info.get('imagefiletype', ''),
                    'filesize': image_info.get('filesize', ''),
                    'target_path': str(target_path),
                    'source': docref.get('source', ''),
                    'nodeid': docref.get('nodeid', ''),
                    'share_id': docref.get('share_id', ''),
                    'doc_created_at': ' '.join(
                        item for item in [doc_info.get('created_date', ''), doc_info.get('created_time', '')] if item
                    ),
                    'status': 'pending',
                    'error': '',
                })

    print(f'[计时]   └─附件清单逐行构建({len(rows)}行): {time.perf_counter() - _t0:.1f}s', flush=True)
    return pd.DataFrame(rows), pd.DataFrame(missing_rows)


def _sheet_headers(sheet_name):
    wb = load_workbook(TEMPLATE_FILE, read_only=True, data_only=True)
    ws = wb[sheet_name]
    return [_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]


def _template_headers():
    return {sheet_name: _sheet_headers(sheet_name) for sheet_name in OUTPUT_SHEETS}


def _timestamped_path(path):
    return path.with_name(f'{path.stem}_{datetime.now().strftime("%H%M%S")}{path.suffix}')


def _write_template_sheets_with_fallback(template_file, output_file, sheet_to_df, extra_sheets=None):
    try:
        return c.write_template_sheets(template_file, output_file, sheet_to_df, extra_sheets)
    except PermissionError:
        fallback = _timestamped_path(output_file)
        print(f'输出文件被占用,改写到: {fallback}')
        return c.write_template_sheets(template_file, fallback, sheet_to_df, extra_sheets)


def _write_exceptions_with_fallback(output_file, sheets):
    try:
        return c.write_exceptions(output_file, sheets)
    except PermissionError:
        fallback = _timestamped_path(output_file)
        print(f'未匹配清单被占用,改写到: {fallback}')
        return c.write_exceptions(fallback, sheets)


def _new_row(headers):
    return {header: '' for header in headers}


def _set(row, field_name, value):
    normalized = _normalize_field_name(field_name)
    for header in row:
        if _normalize_field_name(header) == normalized:
            row[header] = value
            return
    raise KeyError(f'模板缺少字段: {field_name}')


def _read_general_required_rules(headers_by_sheet):
    raw = pd.read_csv(RULE_CSV, encoding='utf-8-sig').iloc[1:, :15].copy()
    raw.columns = [
        'module', 'flow', 'table_name', 'field_name', 'required', 'enum', 'note',
        'event_table', 'event_field', 'event_enum', 'event_note',
        'mcn_table', 'mcn_field', 'mcn_enum', 'mcn_note',
    ]
    for column in raw.columns:
        raw[column] = raw[column].where(raw[column].notna(), '')
    raw['table_name'] = raw['table_name'].replace('', pd.NA).ffill().fillna('')
    raw = raw[raw['flow'] == '一般流程']

    table_to_sheet = {
        '字段模板（主表）': SHEET_MAIN,
        '关联合同': SHEET_RELATION,
        '相关单据-订单信息': SHEET_RELATED_ORDER,
        '采购申请': SHEET_PURCHASE_REQUEST,
        '订单信息明细': SHEET_ORDER_DETAIL,
        '对方信息': SHEET_COUNTERPARTY,
        '我方主体列表': SHEET_OUR_PARTY,
        '付款计划': SHEET_PAYMENT_PLAN,
        '收款计划': SHEET_COLLECTION_PLAN,
        '合同附件': SHEET_CONTRACT_ATTACHMENT,
        '其他附件': SHEET_OTHER_ATTACHMENT,
    }
    required = {sheet_name: [] for sheet_name in headers_by_sheet}
    remarks = {sheet_name: {} for sheet_name in headers_by_sheet}
    header_lookup = {
        sheet_name: {_normalize_field_name(header): header for header in headers}
        for sheet_name, headers in headers_by_sheet.items()
    }

    for _, rule in raw.iterrows():
        sheet_name = table_to_sheet.get(_text(rule['table_name']))
        if not sheet_name:
            continue
        normalized_field = _normalize_field_name(rule['field_name'])
        actual_field = header_lookup[sheet_name].get(normalized_field)
        if not actual_field:
            continue
        note = _first_non_blank(rule['note'], rule['event_note'], rule['mcn_note'])
        if note and actual_field not in remarks[sheet_name]:
            remarks[sheet_name][actual_field] = note
        if _text(rule['required']).upper() == 'Y' and actual_field not in required[sheet_name]:
            required[sheet_name].append(actual_field)
    return required, remarks


def _fill_summary(output_df, required_cols, remarks=None):
    total = len(output_df)
    rows = []
    for column in required_cols:
        column_exists = column in output_df.columns
        filled = int((output_df[column].astype(str).str.strip() != '').sum()) if column_exists else 0
        if filled < total:
            rows.append({
                '必输字段': column,
                '填充数': filled,
                '缺失数': total - filled,
                '总数': total,
                '填充率': '0.00%' if total == 0 else f'{filled / total * 100:.2f}%',
                '备注': (remarks or {}).get(column, ''),
            })
    return pd.DataFrame(rows, columns=['必输字段', '填充数', '缺失数', '总数', '填充率', '备注'])


def _collect_missing_details(output_df, source_df, required_cols, source_field_map, doc_col):
    sheets = {}
    total = len(output_df)
    if total == 0 or doc_col not in output_df.columns:
        return sheets
    for column in required_cols:
        if column not in output_df.columns:
            continue
        blank_mask = output_df[column].astype(str).str.strip() == ''
        missing_count = int(blank_mask.sum())
        if not (0 < missing_count < total):
            continue
        data = {doc_col: output_df.loc[blank_mask, doc_col].astype(str)}
        source_field = source_field_map.get(column)
        if source_field and source_field in source_df.columns:
            data[f'泛微原表-{source_field}'] = source_df.loc[blank_mask, source_field].astype(str)
        sheets[f'缺失_{column[:22]}'] = pd.DataFrame(data).drop_duplicates().reset_index(drop=True)
    return sheets


# ============================ 映射构建 ============================
def build_fw_company_info_map_for_values(company_values):
    company_ids = c.clean_codes(
        company_id
        for value in company_values
        for company_id in c.parse_browser_ids(value)
    )
    if not company_ids:
        return {}
    company_name_map = c.build_fw_company_name_map_for_ids(company_ids)
    entity_map = c.build_accounting_entity_map_for_names(company_name_map.values())
    return {
        company_id: {
            'name': company_name,
            'code': entity_map.get(c.normalize_name(company_name), ''),
        }
        for company_id, company_name in company_name_map.items()
    }


def build_customer_info_map_for_values(customer_values):
    customer_name_map = c.build_fw_customer_name_map_for_ids(customer_values)
    same_customer_map = c.load_same_customer_mapping(log_prefix='[合同迁移-一般流程]')
    source_to_target = {
        customer_id: c.resolve_same_customer_id(customer_id, same_customer_map)
        for customer_id in customer_name_map
    }
    target_name_map = c.build_fw_customer_name_map_for_ids(source_to_target.values())
    customer_by_target_id = c.build_hand_customer_info_by_ids(source_to_target.values())
    customer_code_map = c.build_customer_map_for_names(
        list(customer_name_map.values()) + list(target_name_map.values()))

    result = {}
    for customer_id, target_id in source_to_target.items():
        source_name = customer_name_map.get(customer_id, '')
        target_name = target_name_map.get(target_id, '')
        customer_info = customer_by_target_id.get(target_id, {})
        code = _first_non_blank(
            customer_info.get('code', ''),
            customer_code_map.get(c.normalize_name(target_name), ''),
            customer_code_map.get(c.normalize_name(source_name), ''),
        )
        name = _first_non_blank(customer_info.get('name', ''), target_name, source_name)
        result[customer_id] = {
            'name': name,
            'code': code,
            'source_name': source_name,
            'target_id': target_id,
            'match_method': '客户归并ID' if target_id != customer_id else customer_info.get('match_method', '客户名称'),
        }
    return result


def build_supplier_info_map_for_values(supplier_values):
    supplier_status_map = c.build_fw_supplier_status_map(supplier_values)
    same_supplier_map = c.load_same_supplier_mapping(log_prefix='[合同迁移-一般流程]')
    name_match_cache = c.load_supplier_vendor_name_match_map(log_prefix='[合同迁移-一般流程]')
    source_to_target = {
        supplier_id: c.resolve_same_supplier_id(supplier_id, same_supplier_map)
        for supplier_id in supplier_status_map
    }
    target_status_map = c.build_fw_supplier_status_map(source_to_target.values())
    vendor_by_target_id = c.build_hand_vendor_info_by_ids(source_to_target.values())

    vendor_by_source_id = {}
    names_to_match = []
    disabled_unmatched_ids = []
    for supplier_id, target_id in source_to_target.items():
        supplier_info = supplier_status_map.get(supplier_id, {})
        vendor_info = vendor_by_target_id.get(target_id, {})
        if vendor_info.get('code'):
            vendor_by_source_id[supplier_id] = vendor_info
            continue
        if supplier_info.get('status_code') != '0':
            cached_info = name_match_cache.get(supplier_id, {})
            if cached_info.get('code'):
                vendor_by_source_id[supplier_id] = cached_info
                continue
            supplier_name = supplier_info.get('name', '')
            if supplier_name:
                names_to_match.append(supplier_name)
                disabled_unmatched_ids.append(supplier_id)

    vendor_by_name = c.build_hand_vendor_info_by_names(names_to_match)
    discovered_rows = []
    for supplier_id in disabled_unmatched_ids:
        supplier_info = supplier_status_map.get(supplier_id, {})
        supplier_name = supplier_info.get('name', '')
        vendor_info = vendor_by_name.get(c.normalize_name(supplier_name), {})
        if not vendor_info.get('code'):
            continue
        vendor_by_source_id[supplier_id] = vendor_info
        discovered_rows.append({
            '供应商泛微Id': supplier_id,
            '供应商名称': supplier_name,
            '泛微状态': supplier_info.get('status_code', ''),
            '匹配方式': '禁用供应商按名称匹配汉得供应商',
            'handVendorId': vendor_info.get('id', ''),
            'handVendorCode': vendor_info.get('code', ''),
            'handVendorName': vendor_info.get('name', ''),
            'handTaxpayerName': vendor_info.get('taxpayer_name', ''),
        })
    changed_count = c.append_supplier_vendor_name_matches(discovered_rows)
    if changed_count:
        print(f'[合同迁移-一般流程] 禁用供应商按名称匹配汉得并写入缓存: {changed_count} 条')

    result = {}
    for supplier_id, target_id in source_to_target.items():
        vendor_info = vendor_by_source_id.get(supplier_id) or vendor_by_target_id.get(target_id, {})
        supplier_info = supplier_status_map.get(supplier_id, {})
        target_supplier_info = supplier_status_map.get(target_id) or target_status_map.get(target_id, {})
        target_name = target_supplier_info.get('name', '')
        match_method = vendor_info.get('match_method', '')
        if target_id != supplier_id and vendor_info.get('code'):
            match_method = '供应商归并ID'
        result[supplier_id] = {
            'name': _first_non_blank(vendor_info.get('name', ''), target_name, supplier_info.get('name', '')),
            'code': vendor_info.get('code', ''),
            'source_name': supplier_info.get('name', ''),
            'target_id': target_id,
            'target_name': target_name,
            'status_code': supplier_info.get('status_code', ''),
            'match_method': match_method or ('供应商名称缓存' if name_match_cache.get(supplier_id) else '供应商ID'),
        }
    return result


def build_contract_info_map_for_ids(contract_values):
    contract_ids = c.clean_codes(
        contract_id
        for value in contract_values
        for contract_id in c.parse_browser_ids(value)
    )
    if not contract_ids:
        return {}
    contract_df = c.query_db(
        'FW',
        'vspn_xtyy',
        'SELECT id, htbh, htbt, htlx, htejlx, htszxmbh, htszxm '
        'FROM uf_htk '
        f'WHERE id IN ({c.in_placeholders(contract_ids)})',
        contract_ids,
    )
    option_maps = c.build_fw_select_option_maps(FW_TABLE, ['htlx', 'htejlx'])
    result = {}
    for _, row in contract_df.iterrows():
        contract_id = c.format_code(row['id'])
        if not contract_id:
            continue
        type_id = c.format_code(row.get('htlx'))
        secondary_id = c.format_code(row.get('htejlx'))
        result[contract_id] = {
            'id': contract_id,
            'number': _text(row.get('htbh')),
            'title': _text(row.get('htbt')),
            'type_id': type_id,
            'type': option_maps['htlx'].get(type_id, ''),
            'secondary_type_id': secondary_id,
            'secondary_type': option_maps['htejlx'].get(secondary_id, ''),
            'project_id': _text(row.get('htszxmbh')),
            'project_name': _text(row.get('htszxm')),
        }
    return result


def build_cost_center_platform_map(cost_center_values):
    """成本中心ID -> 一级成本中心(平台)名称。

    链路: 合同.cbzx -> uf_cbzx.id, uf_cbzx.sjcbzx -> uf_yjcbzx.id, uf_yjcbzx.mc。
    """
    cost_center_ids = c.clean_codes(
        cost_center_id
        for value in cost_center_values
        for cost_center_id in c.parse_browser_ids(value)
    )
    if not cost_center_ids:
        return {}
    result = {}
    for batch in _chunked(sorted(set(cost_center_ids))):
        df = c.query_db(
            'FW',
            'vspn_xtyy',
            'SELECT cbzx.id AS cbzx_id, yj.mc AS platform '
            'FROM uf_cbzx cbzx LEFT JOIN uf_yjcbzx yj ON yj.id = cbzx.sjcbzx '
            f'WHERE cbzx.id IN ({c.in_placeholders(batch)})',
            batch,
        )
        for _, row in df.iterrows():
            cost_center_id = c.format_code(row['cbzx_id'])
            platform = _text(row.get('platform'))
            if cost_center_id and platform:
                result[cost_center_id] = platform
    return result


def build_feishu_employee_id_map(code_values):
    """员工工号(V编号) -> 飞书 user_id。

    来源: 汉得 hfins_base.hfbs_employee.feishu_employee_id, 按 employee_code(员工编号)关联。
    泛微侧的工号取自 hrmjobtitles.JOBTITLENAME(即 `合同执行人员工号`), 与汉得 employee_code 同口径。
    """
    codes = c.clean_codes(_text(code) for code in code_values if _text(code))
    if not codes:
        return {}
    result = {}
    for batch in _chunked(sorted(set(codes))):
        df = c.query_db(
            'ZT',
            'hfins_base',
            'SELECT employee_code, feishu_employee_id FROM hfbs_employee '
            f'WHERE employee_code IN ({c.in_placeholders(batch)})',
            batch,
        )
        for _, row in df.iterrows():
            code = _text(row.get('employee_code'))
            feishu_id = _text(row.get('feishu_employee_id'))
            if code and feishu_id:
                result[code] = feishu_id
    return result


def build_purchase_request_code_map(values):
    """采购申请对象ID(uf_htsp.cgsqddx -> uf_cgspxx.id) -> 采购申请单编号(uf_cgspxx.dh)。"""
    request_ids = c.clean_codes(
        request_id
        for value in values
        for request_id in c.parse_browser_ids(value)
    )
    if not request_ids:
        return {}
    result = {}
    for batch in _chunked(sorted(set(request_ids))):
        df = c.query_db(
            'FW',
            'vspn_xtyy',
            f'SELECT id, dh FROM uf_cgspxx WHERE id IN ({c.in_placeholders(batch)})',
            batch,
        )
        for _, row in df.iterrows():
            request_id = c.format_code(row['id'])
            code = _text(row.get('dh'))
            if request_id and code:
                result[request_id] = code
    return result


# ============================ 分类 / 金额 ============================
def _contract_prefix(contract_number):
    text = _text(contract_number).upper()
    match = re.search(r'H-[A-Z]+', text)
    return match.group(0) if match else ''


def _mcn_pay_letter(contract_number):
    """MCN 代字式编号(主体缩写-S/F/O…)的收支字母。S=收入 F=支出 O=其他。"""
    text = _text(contract_number).upper()
    for segment in text.split('-'):
        match = re.fullmatch(r'([SFO])\d*', segment.strip())
        if match:
            return match.group(1)
    return ''


def _contains_any(text, keywords):
    text = _text(text)
    folded = text.casefold()
    return any(keyword and keyword.casefold() in folded for keyword in keywords)


# 一级成本中心(平台) -> 业务域。MD「合同数据迁移-二级分类映射规则」以成本中心为准。
COST_CENTER_PLATFORM_AREAS = {
    '国内赛事业务平台': 'event',
    '赛事商务产品平台': 'event',
    '海外业务平台': 'overseas',
    '全球战略平台': 'overseas',
    '全球生态事业部': 'overseas',
    '资产技术平台': 'asset',
    '商业化事业部': 'commercial',
    'IP衍生事业部': 'derivative',
    '视效事业部': 'visual',
    '虚拟体育事业部': 'visual',
    '运营人力平台': 'admin',
    '采购中心': 'admin',
}


def _business_area(row):
    # 优先用成本中心平台(MD 规则口径), 命中即返回。
    platform = _text(row.get('成本中心平台'))
    if platform:
        for name, area in COST_CENTER_PLATFORM_AREAS.items():
            if name in platform:
                return area
    # 兜底: 成本中心未命中时, 退回项目/标题关键字推断。
    corpus = ' '.join([
        _text(row.get('合同所属项目')),
        _text(row.get('项目名称')),
        _text(row.get('合同标题')),
        _text(row.get('合同二级类型')),
    ])
    if _contains_any(corpus, ['海外', '全球']):
        return 'overseas'
    if _contains_any(corpus, ['资管', '资产技术', '资产采购', '资产租赁', '设备采购', '设备租赁']):
        return 'asset'
    if _contains_any(corpus, ['商业化', '广告', '赞助']):
        return 'commercial'
    if _contains_any(corpus, ['衍生', 'IP']):
        return 'derivative'
    if _contains_any(corpus, ['视效', '视频制作', '视觉']):
        return 'visual'
    if _contains_any(corpus, ['运营人力', '行政', '人力', 'IT', '采购中心']):
        return 'admin'
    if _contains_any(corpus, ['赛事', '活动', '赛']):
        return 'event'
    return ''


def _area_secondary(area, pay_side, contract_shape):
    names = {
        ('event', '支出', '单次'): '国内赛事及活动支出',
        ('overseas', '支出', '单次'): '海外赛事业及活动支出',
        ('asset', '支出', '单次'): '资产采购租赁及经营支出',
        ('commercial', '支出', '单次'): '广告赞助支出',
        ('derivative', '支出', '单次'): '衍生品支出',
        ('visual', '支出', '单次'): '视效支出',
        ('admin', '支出', '单次'): '行政运营及人力支出',
        ('event', '支出', '框架'): '赛事及活动支出框架',
        ('overseas', '支出', '框架'): '赛事及活动支出框架',
        ('admin', '支出', '框架'): '赛事及活动支出框架',
        ('asset', '支出', '框架'): '资产采购租赁及经营支出框架',
        ('commercial', '支出', '框架'): '广告赞助支出框架',
        ('derivative', '支出', '框架'): '衍生品支出框架',
        ('visual', '支出', '框架'): '视效支出框架',
        ('event', '支出', '订单'): '赛事及活动支出订单',
        ('overseas', '支出', '订单'): '赛事及活动支出订单',
        ('asset', '支出', '订单'): '资产采购租赁及经营支出订单',
        ('commercial', '支出', '订单'): '广告赞助支出订单',
        ('derivative', '支出', '订单'): '衍生品支出订单',
        ('visual', '支出', '订单'): '视效支出订单',
        ('admin', '支出', '订单'): '行政运营及人力支出订单',
        ('event', '收入', '单次'): '赛事及活动收入',
        ('overseas', '收入', '单次'): '赛事及活动收入',
        ('asset', '收入', '单次'): '资产采购租赁及经营收入',
        ('commercial', '收入', '单次'): '广告赞助收入',
        ('derivative', '收入', '单次'): '衍生品收入',
        ('visual', '收入', '单次'): '视效收入',
        ('event', '收入', '框架'): '赛事及活动收入框架',
        ('overseas', '收入', '框架'): '赛事及活动收入框架',
        ('asset', '收入', '框架'): '资产采购租赁及经营收入',
        ('admin', '收入', '框架'): '资产采购租赁及经营收入',
        ('commercial', '收入', '框架'): '广告赞助收入',
        ('derivative', '收入', '框架'): '衍生品收入',
        ('visual', '收入', '框架'): '视效收入',
        ('event', '收入', '订单'): '赛事及活动收入订单',
        ('overseas', '收入', '订单'): '赛事及活动收入订单',
        ('asset', '收入', '订单'): '资产采购租赁及经营收入',
        ('admin', '收入', '订单'): '资产采购租赁及经营收入',
        ('commercial', '收入', '订单'): '广告赞助收入订单',
        ('derivative', '收入', '订单'): '衍生品收入订单',
        ('visual', '收入', '订单'): '视效收入订单',
    }
    return names.get((area, pay_side, contract_shape), '')


def resolve_contract_category(row):
    prefix = _contract_prefix(row.get('合同编号'))
    title = _text(row.get('合同标题'))
    supplier_name = _text(row.get('合同供应商名称'))
    area = _business_area(row)

    if prefix in ('H-DF', 'H-F'):
        if '云账户' in supplier_name or _contains_any(title, ['个人合作']):
            return '单次支出-个人合作支出'
        if _contains_any(title, ['奖金', '补贴', 'prize', 'subsidy']):
            return '单次支出-奖金补贴支出'
        return '单次支出-' + (_area_secondary(area, '支出', '单次') or '其他支出')
    if prefix == 'H-KF':
        return '框架支出-' + (_area_secondary(area, '支出', '框架') or '其他支出框架')
    if prefix == 'H-OF':
        return '订单支出-' + (_area_secondary(area, '支出', '订单') or '其他支出订单')
    if prefix in ('H-DS', 'H-S'):
        return '单次收入-' + (_area_secondary(area, '收入', '单次') or '其他收入')
    if prefix == 'H-KS':
        return '框架收入-' + (_area_secondary(area, '收入', '框架') or '其他收入')
    if prefix == 'H-OS':
        return '订单收入-' + (_area_secondary(area, '收入', '订单') or '其他收入')
    if prefix == 'H-DJ':
        return '单次收支-通用单次收支'
    if prefix == 'H-KJ':
        return '框架收支-通用框架收支'
    if prefix == 'H-N':
        return '内部合同-内部结转'
    if prefix == 'H-P' and '保密' in title:
        return '其他-保密协议'
    if prefix == 'H-P' and '反商业贿赂' in title:
        return '其他-反商业贿赂协议'
    if prefix in ('H-P', 'H-PZ'):
        return '其他-战略合作协议'
    if prefix == 'H-Q':
        return '其他-其他类型'

    # MCN 代字式编号(无 H- 前缀): 用代字 S/F 字母定收支, 成本中心平台细分二级。
    letter = _mcn_pay_letter(row.get('合同编号'))
    if letter == 'F':
        return '单次支出-' + (_area_secondary(area, '支出', '单次') or '其他支出')
    if letter == 'S':
        return '单次收入-' + (_area_secondary(area, '收入', '单次') or '其他收入')

    contract_type = _text(row.get('合同类型')) + _text(row.get('合同二级类型'))
    if '收入' in contract_type and '支出' not in contract_type:
        return '单次收入-其他收入'
    if '支出' in contract_type and '收入' not in contract_type:
        return '单次支出-其他支出'
    return '其他-其他类型'


def resolve_contract_category_basis(row):
    prefix = _contract_prefix(row.get('合同编号')) or '空'
    area = _business_area(row) or '未命中成本中心/项目关键词'
    return (
        f'合同编号前缀={prefix}; 成本中心平台={_text(row.get("成本中心平台")) or "空"}; '
        f'合同类型={_text(row.get("合同类型")) or "空"}; '
        f'合同二级类型={_text(row.get("合同二级类型")) or "空"}; '
        f'分类线索={area}; 规则来源=合同数据迁移-二级分类映射规则'
    )


def resolve_pay_type(row):
    prefix = _contract_prefix(row.get('合同编号'))
    if prefix in ('H-DS', 'H-KS', 'H-OS', 'H-S'):
        return '收入类'
    if prefix in ('H-DF', 'H-KF', 'H-OF', 'H-F'):
        return '支出类'
    if prefix in ('H-DJ', 'H-KJ', 'H-N'):
        return '既收又支'
    if prefix in ('H-P', 'H-PZ', 'H-Q'):
        return '无金额'

    letter = _mcn_pay_letter(row.get('合同编号'))
    if letter == 'S':
        return '收入类'
    if letter == 'F':
        return '支出类'

    in_amount = abs(_number(row.get('合同预计收入')))
    out_amount = abs(_number(row.get('合同预计支出')))
    if in_amount and out_amount:
        return '既收又支'
    if in_amount:
        return '收入类'
    if out_amount:
        return '支出类'
    return '无金额'


def resolve_amounts(row):
    pay_type = _text(row.get('收支类型'))
    contract_amount = abs(_number(row.get('合同金额')))
    in_amount = abs(_number(row.get('合同预计收入')))
    out_amount = abs(_number(row.get('合同预计支出')))
    if pay_type == '收入类' and not in_amount:
        in_amount = contract_amount
    if pay_type == '支出类' and not out_amount:
        out_amount = contract_amount
    if pay_type == '既收又支' and not (in_amount or out_amount) and contract_amount:
        in_amount = contract_amount
    if pay_type == '收入类':
        amount = in_amount
    elif pay_type == '支出类':
        amount = out_amount
    elif pay_type == '既收又支':
        amount = contract_amount or in_amount + out_amount
    else:
        amount = 0
    return round(amount, 2), round(in_amount, 2), round(out_amount, 2)


# ============================ 源值解析 ============================
def resolve_source_values(source_df, option_table=FW_TABLE):
    df = source_df.copy()
    option_maps = c.build_fw_select_option_maps(
        option_table,
        ['htlx', 'htejlx', 'htzt', 'bglx'],
    )
    employee_map = _timed('  ├─员工映射', lambda: c.build_fw_employee_info_map_for_ids(df['合同执行人员ID']))
    company_info_map = _timed('  ├─我方主体映射', lambda: build_fw_company_info_map_for_values(df['合同用印范围ID']))
    customer_info_map = _timed('  ├─客户映射', lambda: build_customer_info_map_for_values(df['合同客户ID']))
    supplier_info_map = _timed('  ├─供应商映射(汉得匹配)', lambda: build_supplier_info_map_for_values(df['合同供应商ID']))
    project_map = _timed('  ├─项目映射', lambda: build_fw_project_code_map_for_ids(df['合同所属项目编号ID']))
    relation_info_map = _timed('  ├─关联框架映射', lambda: build_contract_info_map_for_ids(df['关联框架协议ID']))
    cost_center_platform_map = _timed('  └─成本中心平台映射',
                                      lambda: build_cost_center_platform_map(df.get('成本中心ID', pd.Series(dtype=object))))

    df['成本中心平台'] = df.get('成本中心ID', pd.Series('', index=df.index)).map(
        lambda value: _lookup_first_browser_value(cost_center_platform_map, value))
    df['合同类型'] = df['合同类型ID'].map(lambda value: option_maps.get('htlx', {}).get(c.format_code(value), ''))
    df['合同二级类型'] = df['合同二级类型ID'].map(lambda value: option_maps.get('htejlx', {}).get(c.format_code(value), ''))
    df['合同签署状态'] = df['合同签署状态ID'].map(lambda value: option_maps.get('htzt', {}).get(c.format_code(value), ''))
    df['变更类型'] = df['变更类型ID'].map(lambda value: option_maps.get('bglx', {}).get(c.format_code(value), ''))
    df['流程名称'] = df['流程名称'].map(c.clean_fw_select_name)
    df['合同执行人员'] = df['合同执行人员ID'].map(
        lambda value: employee_map.get(c.format_code(value), {}).get('name', ''))
    df['合同执行人员工号'] = df['合同执行人员ID'].map(
        lambda value: employee_map.get(c.format_code(value), {}).get('code', ''))
    feishu_id_map = _timed('  ├─合同执行人飞书ID映射(汉得)',
                           lambda: build_feishu_employee_id_map(df['合同执行人员工号']))
    df['合同执行人飞书ID'] = df['合同执行人员工号'].map(lambda code: feishu_id_map.get(_text(code), ''))
    purchase_request_map = _timed('  ├─采购申请单编号映射',
                                  lambda: build_purchase_request_code_map(df.get('采购申请单ID', pd.Series(dtype=object))))
    df['采购申请单编号'] = df.get('采购申请单ID', pd.Series('', index=df.index)).map(
        lambda value: ';'.join(
            purchase_request_map[request_id]
            for request_id in c.parse_browser_ids(value)
            if purchase_request_map.get(request_id)
        ))
    df['泛微项目编号'] = df['合同所属项目编号ID'].map(
        lambda value: _lookup_first_browser_value(project_map, value) or _text(value))
    df['泛微项目名称'] = df['合同所属项目']
    _t0 = time.perf_counter()
    cleaned_project_info = df.apply(
        lambda row: c.cleaned_project_mapping(
            row.get('合同所属项目编号ID'),
            row.get('合同所属项目'),
            row.get('泛微项目编号'),
        ),
        axis=1,
    )
    df['项目编号'] = cleaned_project_info.map(
        lambda info: _text(info.get('项目编号')) if info else '')
    df['项目名称'] = cleaned_project_info.map(
        lambda info: _text(info.get('项目名称')) if info else '')
    df['项目映射来源'] = cleaned_project_info.map(
        lambda info: _text(info.get('映射来源')) if info else '')
    df['项目编号'] = df.apply(lambda row: _first_non_blank(row['项目编号'], row['泛微项目编号']), axis=1)
    df['项目名称'] = df.apply(lambda row: _first_non_blank(row['项目名称'], row['合同所属项目']), axis=1)
    df['订单编号'] = df['项目编号'].map(lambda value: c.project_order_mapping_value(value, '订单编号'))
    df['订单名称'] = df['项目编号'].map(lambda value: c.project_order_mapping_value(value, '订单标题'))
    df['清洗后项目编号'] = df['项目编号']
    df['清洗后项目名称'] = df['项目名称']
    df['订单映射来源'] = df['项目编号'].map(lambda value: c.project_order_mapping_value(value, '映射来源'))
    print(f'[计时] ✓   项目/订单映射(逐行apply): {time.perf_counter() - _t0:.1f}s', flush=True)

    def supplier_names(value):
        return '; '.join(
            supplier_info_map.get(supplier_id, {}).get('source_name', '')
            for supplier_id in c.parse_browser_ids(value)
            if supplier_info_map.get(supplier_id, {}).get('source_name', '')
        )

    _t0 = time.perf_counter()
    df['合同供应商名称'] = df['合同供应商ID'].map(supplier_names)
    df['合同分类'] = df.apply(resolve_contract_category, axis=1)
    df['合同分类依据'] = df.apply(resolve_contract_category_basis, axis=1)
    df['收支类型'] = df.apply(resolve_pay_type, axis=1)
    amount_tuples = df.apply(resolve_amounts, axis=1)
    df['合同总额_解析'] = amount_tuples.map(lambda item: item[0])
    df['收入总额_解析'] = amount_tuples.map(lambda item: item[1])
    df['支出总额_解析'] = amount_tuples.map(lambda item: item[2])
    print(f'[计时] ✓   分类/金额(逐行apply): {time.perf_counter() - _t0:.1f}s', flush=True)

    def relation_codes(value):
        rows = []
        for relation_id in c.parse_browser_ids(value):
            info = relation_info_map.get(relation_id, {})
            if info.get('number'):
                rows.append(info['number'])
        return ';'.join(rows)

    df['关联合同编号'] = df['关联框架协议ID'].map(relation_codes)
    df.attrs['company_info_map'] = company_info_map
    df.attrs['customer_info_map'] = customer_info_map
    df.attrs['supplier_info_map'] = supplier_info_map
    df.attrs['relation_info_map'] = relation_info_map
    return df


def _merge_attrs(target_df, source_dfs):
    """把多个已解析 df 的 attrs(各 id->info 映射)合并到 target_df.attrs。"""
    for key in ('company_info_map', 'customer_info_map', 'supplier_info_map', 'relation_info_map'):
        merged = {}
        for df in source_dfs:
            merged.update(df.attrs.get(key, {}))
        target_df.attrs[key] = merged
    return target_df


def read_source():
    # ---- 源1: MCN 合同库 uf_htk ----
    c.validate_fw_fields(FW_TABLE, EXPECTED_FW_FIELDS)
    stats = _query_fw(STATS_SQL).iloc[0]
    print('[合同迁移-一般流程] MCN(uf_htk) 过滤: 合同类型<>主播协议 且 合同签署状态∈(审批完成, 已归档)')
    print(f"  合同库总数 {int(stats['all_count'] or 0)} 条; "
          f"主播协议排除 {int(stats['anchor_type_count'] or 0)} 条; "
          f"一般流程保留 {int(stats['kept_count'] or 0)} 条; "
          f"一般流程排除其他状态 {int(stats['excluded_status_count'] or 0)} 条")
    mcn_df = _timed('read_source/MCN取数(uf_htk)', lambda: _query_fw(SOURCE_SQL))
    mcn_df['数据来源'] = '泛微(MCN)'
    print('[合同迁移-一般流程] MCN 主表行数:', len(mcn_df))

    # ---- 源2: 赛事 合同审批台账 uf_htsp(htzt=1 归档); 与 MCN 重叠的编号让位给 MCN ----
    htsp_df = _timed('read_source/赛事取数(uf_htsp)', lambda: c.query_db(
        'FW', 'vspn_xtyy', SOURCE_SQL_HTSP, {'htsp_status_codes': HTSP_MIGRATION_STATUS_CODES}))
    htsp_df['数据来源'] = '泛微(赛事)'
    mcn_codes = set(mcn_df['合同编号'].map(_text))
    before = len(htsp_df)
    htsp_df = htsp_df[~htsp_df['合同编号'].map(_text).isin(mcn_codes)].copy()
    print(f'[合同迁移-一般流程] 赛事(uf_htsp) 归档行数 {before}; 去重(MCN优先)后 {len(htsp_df)}')

    resolved_mcn = _timed('read_source/解析MCN(%d行)' % len(mcn_df),
                          lambda: resolve_source_values(mcn_df, option_table=FW_TABLE))
    resolved_htsp = (
        _timed('read_source/解析赛事(%d行)' % len(htsp_df),
               lambda: resolve_source_values(htsp_df, option_table=FW_TABLE_HTSP))
        if not htsp_df.empty else htsp_df
    )

    merged = pd.concat([resolved_mcn, resolved_htsp], ignore_index=True)
    _merge_attrs(merged, [resolved_mcn, resolved_htsp])
    # 赛事收支计划明细(uf_htsp_dt4): 供付款/收款计划按 dt4 真实多期展开。
    plan_map = (
        _timed('read_source/赛事收支计划dt4', lambda: load_htsp_plan_detail_map(resolved_htsp['ID']))
        if not resolved_htsp.empty else {}
    )
    merged.attrs['saishi_plan_map'] = plan_map
    print(f'[合同迁移-一般流程] 赛事收支计划明细: {sum(len(v) for v in plan_map.values())} 行 / {len(plan_map)} 合同')
    print('[合同迁移-一般流程] 合并后主表行数:', len(merged))
    return merged


# ============================ Sheet 构建 ============================
def build_main_output(source_df, headers):
    rows = []
    for source in source_df.to_dict('records'):
        row = _new_row(headers)
        amount = source['合同总额_解析']
        in_amount = source['收入总额_解析']
        out_amount = source['支出总额_解析']

        _set(row, 'contract_number（合同编码）', _text(source['合同编号']))
        _set(row, 'contract_name（合同名称）', _text(source['合同标题']))
        _set(row, 'contractCategory(智书框架合同类型)', source['合同分类'])
        _set(row, 'pay_type_code（收支类型）', source['收支类型'])
        _set(row, 'property_type_code（计价方式）', DEFAULT_PROPERTY_TYPE)
        _set(row, 'estimated_amount（预估金额）', amount)
        _set(row, 'in_amount（预估收入金额）', in_amount)
        _set(row, 'out_amount（预估支出金额）', out_amount)
        _set(row, 'amount（合同总额）', amount)
        _set(row, 'in_amount（收入总额）', in_amount)
        _set(row, 'out_amount（支出总额）', out_amount)
        _set(row, 'fixed_validity_code（合同期限类型）', DEFAULT_VALIDITY_TYPE)
        _set(row, 'start_date（合同期限-开始日期）', c.format_date(source['合同有效期起始时间']))
        _set(row, 'end_date（合同期限-结束日期）', c.format_date(source['合同有效期截止时间']))
        _set(row, 'remark（合同说明）', _text(source['合同摘要'])[:150])
        _set(row, 'custom_1001_948719050bfe402ab083c98e52fa71b2（合同执行人）飞书user_id',
             _text(source['合同执行人飞书ID']))
        _set(row, 'custom_15_78cf503c57194e4fb8ad03ded1c4ad60（打印模式）', DEFAULT_PRINT_MODE)
        _set(row, 'custom_10_9a2a0e99771346c98bfb6cfb893e1bee（签署日期）', c.format_date(source['合同签订日期']))
        _set(row, 'custom_15_de8944334b104d52b28d9472ab0584ef（专项品类）', '')
        _set(row, 'custom_13_c9805a6fe9f245ebbfeea13407277306（是否需要验收）', DEFAULT_ACCEPTANCE_REQUIRED)
        _set(row, 'custom_1012_cec7052f613b465980f23f7004e2f82c（采购金额）', out_amount if out_amount else '')
        _set(row, 'custom_15_e5f7b7cb17b34602adf790f0ac8d69a1（发票种类）', DEFAULT_INVOICE_TYPE)
        _set(row, 'custom_15_7b0d0e2f63a148729f929ba985c227c2（收入税率）',
             _format_tax_rate(source.get('收入税率')) or DEFAULT_TAX_RATE)
        _set(row, 'custom_15_b293866468ac4ab4bb11b5cb8c9bbb37（收入税目）', DEFAULT_TAX_ITEM)
        _set(row, 'custom_15_7b0d0e2f63a148729f929ba985c227c2（支出税率）',
             _format_tax_rate(source.get('支出税率')) or DEFAULT_TAX_RATE)
        _set(row, 'custom_15_b293866468ac4ab4bb11b5cb8c9bbb37（支出税目）', DEFAULT_TAX_ITEM)
        _set(row, 'custom_15_e46e1f9b6eb0469987f0656a999cdf09（银行手续费承担方）',
             DEFAULT_BANK_FEE_BEARER)
        deposit_amount = _round_amount(source.get('押金')) or _round_amount(source.get('保证金'))
        _set(row, 'custom_1012_7e2c970e63f648268eaefbd13d6bfc8f（押金/保证金）',
             deposit_amount if deposit_amount else '')
        _set(row, 'sign_type_code（先盖章方）', DEFAULT_FIRST_SEAL_PARTY)
        _set(row, 'sign_type_code（签约形式）', DEFAULT_SIGN_FORM)
        _set(row, 'seal_number（盖章份数）待确认必填项', DEFAULT_SEAL_NUMBER)
        _set(row, 'contract_files.contract_text（合同文本）', '')
        _set(row, 'custom_1012_7c79ac40b9ec4efb8be367a43f480d01（首款金额）', '')
        _set(row, 'custom_1012_98dfe42395274c53926c2c37b2dbd9a9（尾款金额）', '')
        rows.append(row)
    return pd.DataFrame(rows, columns=headers)


def build_relation_output(source_df, headers):
    relation_info_map = source_df.attrs.get('relation_info_map', {})
    rows = []
    source_rows = []
    for source in source_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        for relation_id in c.parse_browser_ids(source['关联框架协议ID']):
            info = relation_info_map.get(relation_id, {})
            row = _new_row(headers)
            _set(row, 'contract_number（合同编码）', contract_number)
            _set(row, 'relation.relation_contracts（关联合同）', info.get('number', ''))
            _set(row, '框架合同编号', info.get('number', ''))
            rows.append(row)
            source_rows.append({
                'contract_number（合同编码）': contract_number,
                '泛微关联合同ID': relation_id,
                '关联合同编号': info.get('number', ''),
                '关联合同名称': info.get('title', ''),
            })
    return pd.DataFrame(rows, columns=headers), pd.DataFrame(source_rows)


def build_related_order_output(source_df, headers):
    rows = []
    for source in source_df.to_dict('records'):
        order_code = _text(source['订单编号'])
        if not order_code:
            continue
        row = _new_row(headers)
        _set(row, 'contract_number（合同编码）', _text(source['合同编号']))
        _set(row, 'custom_1024_90a78c8120994f95b2dbfedd297c7d81（相关单据-订单信息）', order_code)
        rows.append(row)
    return pd.DataFrame(rows, columns=headers)


def build_purchase_request_output(source_df, headers):
    rows = []
    for source in source_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        for code in _text(source.get('采购申请单编号')).split(';'):
            code = code.strip()
            if not code:
                continue
            row = _new_row(headers)
            _set(row, 'contract_number（合同编码）', contract_number)
            _set(row, 'custom_1024_7db9a8ee2b3d4a3f9d9835dd9fee69df（采购申请）', code)
            rows.append(row)
    return pd.DataFrame(rows, columns=headers)


def build_order_detail_output(source_df, headers):
    rows = []
    source_rows = []
    for source in source_df.to_dict('records'):
        order_code = _text(source['订单编号'])
        order_name = _text(source['订单名称'])
        if not (order_code or order_name):
            continue
        row = _new_row(headers)
        _set(row, 'contract_number（合同编码）', _text(source['合同编号']))
        _set(row, 'custom_1_5549b19faea641eeac924deada603c11（订单名称）', order_name)
        _set(row, 'custom_1_7f977c0d30064dd199434f706470c669（订单编号）', order_code)
        _set(row, 'custom_16_3171b080033943c9a98380f20e0895a8（成本中心）', '')
        start_date = c.format_date(source['合同有效期起始时间'])
        end_date = c.format_date(source['合同有效期截止时间'])
        period = f'{start_date}~{end_date}'.strip('~')
        _set(row, 'custom_12_d67e0d9472134b1cba5187e192bb2670（订单周期）',
             _first_non_blank(period, c.format_date(source['合同签订日期'])))
        _set(row, 'custom_1001_b193503253664cf28b2ca1c3f57b68b3（项目经理）飞书user_id', '')
        _set(row, 'custom_1001_0a61d360dcad4265a68d2555d17e896e（日常费用组）飞书user_id', '')
        _set(row, 'custom_1001_f8f9114f511346f9adc7fabae012f17a（项目验收岗）飞书user_id', '')
        _set(row, 'custom_1001_8b8028ca466f4841bead801a9c7fedf2（项目预算岗）飞书user_id', '')
        _set(row, 'custom_1001_9072dcb2126e4051854da3927f742ab9（项目Sponsor-项目经理B）飞书user_id', '')
        _set(row, 'custom_15_622e96ab047c4f689d287a27066f7bcb（订单类型）', '')
        rows.append(row)
        source_rows.append(dict(source))
    return pd.DataFrame(rows, columns=headers), pd.DataFrame(source_rows)


def build_counterparty_output(source_df, headers):
    customer_info_map = source_df.attrs.get('customer_info_map', {})
    supplier_info_map = source_df.attrs.get('supplier_info_map', {})
    rows = []
    source_rows = []
    for source in source_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        for customer_id in c.parse_browser_ids(source['合同客户ID']):
            info = customer_info_map.get(customer_id, {})
            row = _new_row(headers)
            _set(row, 'contract_number（合同编码）', contract_number)
            _set(row, 'counter_party_code（对方主体编码）', info.get('code', ''))
            rows.append(row)
            source_rows.append({
                'contract_number（合同编码）': contract_number,
                '对方主体类型': '客户',
                '泛微对方ID': customer_id,
                '泛微对方名称': info.get('source_name', ''),
                '归并目标ID': info.get('target_id', ''),
                '归并/匹配方式': info.get('match_method', ''),
                '对方主体名称': info.get('name', ''),
                '对方主体编码': info.get('code', ''),
            })
        for supplier_id in c.parse_browser_ids(source['合同供应商ID']):
            info = supplier_info_map.get(supplier_id, {})
            row = _new_row(headers)
            _set(row, 'contract_number（合同编码）', contract_number)
            _set(row, 'counter_party_code（对方主体编码）', info.get('code', ''))
            rows.append(row)
            source_rows.append({
                'contract_number（合同编码）': contract_number,
                '对方主体类型': '供应商',
                '泛微对方ID': supplier_id,
                '泛微对方状态': info.get('status_code', ''),
                '泛微对方名称': info.get('source_name', ''),
                '归并目标ID': info.get('target_id', ''),
                '归并目标名称': info.get('target_name', ''),
                '归并/匹配方式': info.get('match_method', ''),
                '对方主体名称': info.get('name') or info.get('source_name', ''),
                '对方主体编码': info.get('code', ''),
            })
    return pd.DataFrame(rows, columns=headers), pd.DataFrame(source_rows)


def build_our_party_output(source_df, headers):
    company_info_map = source_df.attrs.get('company_info_map', {})
    rows = []
    source_rows = []
    for source in source_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        company_ids = c.parse_browser_ids(source['合同用印范围ID'])
        if not company_ids:
            company_ids = ['']
        for company_id in company_ids:
            info = company_info_map.get(company_id, {})
            row = _new_row(headers)
            _set(row, 'contract_number（合同编码）', contract_number)
            _set(row, 'our_party_code（我方主体编码）', info.get('code', ''))
            rows.append(row)
            source_rows.append({
                'contract_number（合同编码）': contract_number,
                '泛微我方主体ID': company_id,
                '我方主体名称': info.get('name', ''),
                '我方主体编码': info.get('code', ''),
            })
    return pd.DataFrame(rows, columns=headers), pd.DataFrame(source_rows)


def _first_counterparty_code(source, customer_info_map, supplier_info_map):
    for customer_id in c.parse_browser_ids(source.get('合同客户ID')):
        code = customer_info_map.get(customer_id, {}).get('code', '')
        if code:
            return code
    for supplier_id in c.parse_browser_ids(source.get('合同供应商ID')):
        code = supplier_info_map.get(supplier_id, {}).get('code', '')
        if code:
            return code
    return ''


PLAN_TYPE_COLLECTION = '0'  # uf_htsp_dt4.lx=0 收款(收入)
PLAN_TYPE_PAYMENT = '1'     # uf_htsp_dt4.lx=1 付款(支出)
_HTML_TAG_RE = re.compile(r'<[^>]+>')


def _clean_plan_desc(value):
    text = html.unescape(_text(value))
    text = re.sub(r'<br\s*/?>', ' ', text, flags=re.IGNORECASE)
    text = _HTML_TAG_RE.sub(' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def load_htsp_plan_detail_map(contract_ids):
    """uf_htsp.id -> 收支计划明细行(uf_htsp_dt4)。lx: 0=收款 1=付款。"""
    result = {}
    ids = c.clean_codes(_text(cid) for cid in contract_ids if _text(cid))
    if not ids:
        return result
    for batch in _chunked(sorted(set(ids))):
        df = c.query_db(
            'FW',
            'vspn_xtyy',
            'SELECT mainid, lx, rqjd, je, rmbje, bz FROM uf_htsp_dt4 '
            f'WHERE mainid IN ({c.in_placeholders(batch)}) ORDER BY mainid, id',
            batch,
        )
        for _, row in df.iterrows():
            mainid = c.format_code(row['mainid'])
            if not mainid:
                continue
            result.setdefault(mainid, []).append({
                'lx': c.format_code(row.get('lx')),
                'date': _text(row.get('rqjd')),
                'amount': _round_amount(row.get('je')),
                'desc': _clean_plan_desc(row.get('bz')),
            })
    return result


def _plan_row_id(source, kind, index):
    # 跨源唯一: 赛事/MCN 的 ID 可能数值相同, 用来源字母前缀区分。
    tag = 'S' if _text(source.get('数据来源')) == '泛微(赛事)' else 'M'
    return f'{tag}{_text(source.get("ID"))}-{kind}{index}'


def _htsp_plan_details(source, plan_map, plan_type):
    if _text(source.get('数据来源')) != '泛微(赛事)':
        return []
    return [
        detail
        for detail in plan_map.get(_text(source.get('ID')), [])
        if detail['lx'] == plan_type and detail['amount']
    ]


def build_payment_plan_output(source_df, headers):
    customer_info_map = source_df.attrs.get('customer_info_map', {})
    supplier_info_map = source_df.attrs.get('supplier_info_map', {})
    plan_map = source_df.attrs.get('saishi_plan_map', {})
    rows = []

    def add_row(source, amount, date, desc, line_id):
        row = _new_row(headers)
        _set(row, 'contract_number（合同编码）', _text(source['合同编号']))
        _set(row, 'payment_plan_list（付款计划）', '付款计划')
        _set(row, 'payment_plan_list[].payment_date（付款时间）', c.format_date(date))
        _set(row, 'payment_plan_list[].prepaid（是否预付）', DEFAULT_PREPAID)
        _set(row, 'payment_plan_list[].payment_amount（付款金额）', amount)
        _set(row, 'payment_plan_list[].payment_desc（付款说明）', desc[:80])
        _set(row, 'payment_plan_list[].payment_custom_attributes/custom_付款性质（付款性质）',
             DEFAULT_PAYMENT_NATURE)
        _set(row, 'payment_plan_list[].payment_counter_party[].counter_party_code（付款对象）',
             _first_counterparty_code(source, customer_info_map, supplier_info_map))
        _set(row, '付款计划行id(付款记录传的id)', line_id)
        rows.append(row)

    for source in source_df.to_dict('records'):
        details = _htsp_plan_details(source, plan_map, PLAN_TYPE_PAYMENT)
        if details:
            # 赛事: 用 dt4 真实收支计划明细, 每行一条付款计划。
            for index, detail in enumerate(details, 1):
                add_row(
                    source, detail['amount'],
                    _first_non_blank(detail['date'], source['合同有效期截止时间'], source['合同签订日期']),
                    _first_non_blank(detail['desc'], _text(source['合同标题'])),
                    _plan_row_id(source, 'P', index),
                )
            continue
        # 兜底(无明细的赛事 / MCN): 按支出总额合成单行。
        amount = _round_amount(source['支出总额_解析'])
        if not amount:
            continue
        add_row(
            source, amount,
            _first_non_blank(source['合同有效期截止时间'], source['合同签订日期']),
            _text(source['合同标题']),
            _plan_row_id(source, 'P', 1),
        )
    return pd.DataFrame(rows, columns=headers)


def build_collection_plan_output(source_df, headers):
    customer_info_map = source_df.attrs.get('customer_info_map', {})
    supplier_info_map = source_df.attrs.get('supplier_info_map', {})
    plan_map = source_df.attrs.get('saishi_plan_map', {})
    rows = []

    def add_row(source, amount, date, desc, line_id):
        row = _new_row(headers)
        _set(row, 'contract_number（合同编码）', _text(source['合同编号']))
        _set(row, 'collection_plan_list（收款计划）', '收款计划')
        _set(row, 'collection_plan_list[].collection_date（收款时间）', c.format_date(date))
        _set(row, 'collection_plan_list[].collection_amount（收款金额）', amount)
        _set(row, 'collection_plan_list[].collection_desc（收款说明）', desc[:80])
        _set(row, 'collection_plan_list[].collection_counter_party[].counter_party_code（收款对象）',
             _first_counterparty_code(source, customer_info_map, supplier_info_map))
        _set(row, '收款计划行id(收款记录传的id)', line_id)
        rows.append(row)

    for source in source_df.to_dict('records'):
        details = _htsp_plan_details(source, plan_map, PLAN_TYPE_COLLECTION)
        if details:
            for index, detail in enumerate(details, 1):
                add_row(
                    source, detail['amount'],
                    _first_non_blank(detail['date'], source['合同有效期截止时间'], source['合同签订日期']),
                    _first_non_blank(detail['desc'], _text(source['合同标题'])),
                    _plan_row_id(source, 'C', index),
                )
            continue
        amount = _round_amount(source['收入总额_解析'])
        if not amount:
            continue
        add_row(
            source, amount,
            _first_non_blank(source['合同有效期截止时间'], source['合同签订日期']),
            _text(source['合同标题']),
            _plan_row_id(source, 'C', 1),
        )
    return pd.DataFrame(rows, columns=headers)


def build_contract_attachment_output(source_df, headers):
    manifest_df, missing_df = build_contract_attachment_manifest(source_df)
    # 导入侧是「按合同的扁平附件名单」: 修订稿/签署版常是同一份文件(同名),按
    # (合同编码, 附件名) 去重, 避免同一文件重复计入。下载侧仍按稿件类型分文件夹保留。
    output_rows = []
    seen_names = set()
    for meta in manifest_df.to_dict('records'):
        contract_number = _text(meta.get('contract_number（合同编码）'))
        attachment_name = _text(meta.get('attachment_name'))
        dedupe_key = (contract_number, attachment_name)
        if dedupe_key in seen_names:
            continue
        seen_names.add(dedupe_key)
        row = _new_row(headers)
        _set(row, 'contract_number（合同编码）', contract_number)
        _set(row, 'contract_files.contract_causes（合同附件）', attachment_name)
        output_rows.append(row)
    if len(manifest_df) > 0:
        manifest_df = manifest_df.copy()
        manifest_df['status'] = 'listed_only'
        manifest_df['error'] = 'Excel任务仅生成附件名称清单;下载请单独运行 contract_general_attachments_db'
        print(f'[合同迁移-一般流程] 合同附件名称清单行数: {len(manifest_df)}; 未执行下载')

    return (
        pd.DataFrame(output_rows, columns=headers),
        manifest_df,
        missing_df,
    )


def build_attachment_output(headers):
    return pd.DataFrame(columns=headers)


def build_status_breakdown():
    option_maps = c.build_fw_select_option_maps(FW_TABLE, ['htzt'])
    df = _query_fw(STATUS_BREAKDOWN_SQL)
    df['合同签署状态'] = df['合同签署状态ID'].map(lambda value: option_maps['htzt'].get(c.format_code(value), '空'))
    return df[['合同签署状态ID', '合同签署状态', '合同数']]


# ============================ 输出 / 异常 ============================
MAIN_ISSUE_SOURCE_FIELDS = {
    'contract_number（合同编码）': '合同编号',
    'contract_name（合同名称）': '合同标题',
    'contractCategory(智书框架合同类型)': '合同分类依据',
    'custom_1001_948719050bfe402ab083c98e52fa71b2（合同执行人）飞书user_id': '合同执行人员工号',
    'start_date（合同期限-开始日期）': '合同有效期起始时间',
    'end_date（合同期限-结束日期）': '合同有效期截止时间',
    'remark（合同说明）': '合同摘要',
}

ORDER_ISSUE_SOURCE_FIELDS = {
    'custom_1_5549b19faea641eeac924deada603c11（订单名称）': '项目编号',
    'custom_1_7f977c0d30064dd199434f706470c669（订单编号）': '项目编号',
}


def collect_missing_counterparty(counterparty_source_df):
    if counterparty_source_df.empty:
        return pd.DataFrame(columns=[
            'contract_number（合同编码）', '对方主体类型', '泛微对方ID', '泛微对方状态',
            '泛微对方名称', '归并目标ID', '归并目标名称', '归并/匹配方式', '对方主体名称',
        ])
    missing = counterparty_source_df[counterparty_source_df['对方主体编码'].astype(str).str.strip() == '']
    columns = [
        'contract_number（合同编码）', '对方主体类型', '泛微对方ID', '泛微对方状态',
        '泛微对方名称', '归并目标ID', '归并目标名称', '归并/匹配方式', '对方主体名称',
    ]
    return missing[[column for column in columns if column in missing.columns]].drop_duplicates()


def collect_missing_our_party(our_party_source_df):
    if our_party_source_df.empty:
        return pd.DataFrame(columns=['contract_number（合同编码）', '泛微我方主体ID', '我方主体名称'])
    missing = our_party_source_df[our_party_source_df['我方主体编码'].astype(str).str.strip() == '']
    return missing[['contract_number（合同编码）', '泛微我方主体ID', '我方主体名称']].drop_duplicates()


def collect_missing_relation(relation_source_df):
    if relation_source_df.empty:
        return pd.DataFrame(columns=['contract_number（合同编码）', '泛微关联合同ID', '关联合同编号', '关联合同名称'])
    missing = relation_source_df[relation_source_df['关联合同编号'].astype(str).str.strip() == '']
    return missing.drop_duplicates()


def _audit_df(source_df, columns):
    audit = source_df.reindex(columns=columns)
    for column in audit.columns:
        audit[column] = audit[column].map(_text)
    return audit


def build_category_audit_df(source_df):
    return _audit_df(source_df, [
        '数据来源', '合同编号', '合同标题', '合同类型ID', '合同类型', '合同二级类型ID', '合同二级类型',
        '流程类型ID', '流程名称', '项目编号', '泛微项目编号', '合同所属项目', '合同供应商名称',
        '成本中心平台', '收支类型', '合同分类', '合同分类依据', '合同金额', '合同预计收入', '合同预计支出',
    ])


def build_order_audit_df(source_df):
    return _audit_df(source_df, [
        '数据来源', '合同编号', '合同标题', '合同所属项目编号ID', '泛微项目编号', '泛微项目名称',
        '清洗后项目编号', '清洗后项目名称', '项目映射来源', '订单编号', '订单名称', '订单映射来源',
    ])


def run():
    headers_by_sheet = _template_headers()
    required_by_sheet, remarks_by_sheet = _read_general_required_rules(headers_by_sheet)

    source_df = _timed('阶段1: read_source(读数+解析+合并)', read_source)

    print('[计时] === 阶段2: 构建各 sheet ===', flush=True)
    main_output_df = _timed('build 字段模板', lambda: build_main_output(source_df, headers_by_sheet[SHEET_MAIN]))
    relation_output_df, relation_source_df = _timed(
        'build 关联合同', lambda: build_relation_output(source_df, headers_by_sheet[SHEET_RELATION]))
    related_order_output_df = _timed(
        'build 相关单据-订单', lambda: build_related_order_output(source_df, headers_by_sheet[SHEET_RELATED_ORDER]))
    purchase_request_output_df = _timed(
        'build 采购申请', lambda: build_purchase_request_output(source_df, headers_by_sheet[SHEET_PURCHASE_REQUEST]))
    order_detail_output_df, order_detail_source_df = _timed(
        'build 订单明细', lambda: build_order_detail_output(source_df, headers_by_sheet[SHEET_ORDER_DETAIL]))
    counterparty_output_df, counterparty_source_df = _timed(
        'build 对方信息', lambda: build_counterparty_output(source_df, headers_by_sheet[SHEET_COUNTERPARTY]))
    our_party_output_df, our_party_source_df = _timed(
        'build 我方主体', lambda: build_our_party_output(source_df, headers_by_sheet[SHEET_OUR_PARTY]))
    payment_plan_output_df = _timed(
        'build 付款计划', lambda: build_payment_plan_output(source_df, headers_by_sheet[SHEET_PAYMENT_PLAN]))
    collection_plan_output_df = _timed(
        'build 收款计划', lambda: build_collection_plan_output(source_df, headers_by_sheet[SHEET_COLLECTION_PLAN]))
    contract_attachment_output_df, contract_attachment_meta_df, contract_attachment_missing_df = _timed(
        'build 合同附件(含清单)', lambda: build_contract_attachment_output(source_df, headers_by_sheet[SHEET_CONTRACT_ATTACHMENT]))
    other_attachment_output_df = build_attachment_output(headers_by_sheet[SHEET_OTHER_ATTACHMENT])

    print('[合同迁移-一般流程] 字段模板行数:', len(main_output_df))
    print('[合同迁移-一般流程] 关联合同行数:', len(relation_output_df))
    print('[合同迁移-一般流程] 相关单据-订单信息行数:', len(related_order_output_df))
    print('[合同迁移-一般流程] 订单信息明细行数:', len(order_detail_output_df))
    print('[合同迁移-一般流程] 对方信息行数:', len(counterparty_output_df))
    print('[合同迁移-一般流程] 我方主体列表行数:', len(our_party_output_df))
    print('[合同迁移-一般流程] 付款计划行数:', len(payment_plan_output_df))
    print('[合同迁移-一般流程] 收款计划行数:', len(collection_plan_output_df))
    print('[合同迁移-一般流程] 合同附件行数:', len(contract_attachment_output_df))

    print('[计时] === 阶段3: 写出 Excel ===', flush=True)
    cat_audit = _timed('build 合同分类核对', lambda: build_category_audit_df(source_df))
    ord_audit = _timed('build 订单映射核对', lambda: build_order_audit_df(source_df))
    output_file = _timed('写出导入Excel(单次load/save)', lambda: _write_template_sheets_with_fallback(
        TEMPLATE_FILE, OUTPUT_FILE, {
            SHEET_MAIN: main_output_df,
            SHEET_RELATION: relation_output_df,
            SHEET_RELATED_ORDER: related_order_output_df,
            SHEET_PURCHASE_REQUEST: purchase_request_output_df,
            SHEET_ORDER_DETAIL: order_detail_output_df,
            SHEET_COUNTERPARTY: counterparty_output_df,
            SHEET_OUR_PARTY: our_party_output_df,
            SHEET_PAYMENT_PLAN: payment_plan_output_df,
            SHEET_COLLECTION_PLAN: collection_plan_output_df,
            SHEET_CONTRACT_ATTACHMENT: contract_attachment_output_df,
            SHEET_OTHER_ATTACHMENT: other_attachment_output_df,
        }, extra_sheets={'合同分类核对': cat_audit, '订单映射核对': ord_audit}))
    print('已写出:', output_file)

    exception_sheets = {
        '过滤状态分布': build_status_breakdown(),
        '默认值说明': pd.DataFrame([
            {'字段': 'property_type_code（计价方式）', '默认值': DEFAULT_PROPERTY_TYPE},
            {'字段': 'fixed_validity_code（合同期限类型）', '默认值': DEFAULT_VALIDITY_TYPE},
            {'字段': 'custom_13_c9805a6fe9f245ebbfeea13407277306（是否需要验收）', '默认值': DEFAULT_ACCEPTANCE_REQUIRED},
            {'字段': 'custom_15_78cf503c57194e4fb8ad03ded1c4ad60（打印模式）', '默认值': DEFAULT_PRINT_MODE},
            {'字段': 'custom_15_e5f7b7cb17b34602adf790f0ac8d69a1（发票种类）', '默认值': DEFAULT_INVOICE_TYPE},
            {'字段': 'custom_15_7b0d0e2f63a148729f929ba985c227c2（收入/支出税率）',
             '默认值': f'赛事取源值srsl/zcsl(百分数), 无源值时默认 {DEFAULT_TAX_RATE}'},
            {'字段': 'custom_15_b293866468ac4ab4bb11b5cb8c9bbb37（收入/支出税目）', '默认值': DEFAULT_TAX_ITEM},
            {'字段': 'custom_15_e46e1f9b6eb0469987f0656a999cdf09（银行手续费承担方）', '默认值': DEFAULT_BANK_FEE_BEARER},
            {'字段': 'sign_type_code（先盖章方）', '默认值': DEFAULT_FIRST_SEAL_PARTY},
            {'字段': 'sign_type_code（签约形式）', '默认值': DEFAULT_SIGN_FORM},
            {'字段': 'seal_number（盖章份数）待确认必填项', '默认值': DEFAULT_SEAL_NUMBER},
        ]),
        '字段模板_必输字段未达100%': _fill_summary(
            main_output_df,
            required_by_sheet[SHEET_MAIN],
            remarks_by_sheet[SHEET_MAIN],
        ),
        '关联合同_必输字段未达100%': _fill_summary(
            relation_output_df,
            required_by_sheet[SHEET_RELATION],
            remarks_by_sheet[SHEET_RELATION],
        ),
        '相关单据订单_必输字段未达100%': _fill_summary(
            related_order_output_df,
            required_by_sheet[SHEET_RELATED_ORDER],
            remarks_by_sheet[SHEET_RELATED_ORDER],
        ),
        '订单信息明细_必输字段未达100%': _fill_summary(
            order_detail_output_df,
            required_by_sheet[SHEET_ORDER_DETAIL],
            remarks_by_sheet[SHEET_ORDER_DETAIL],
        ),
        '对方信息_必输字段未达100%': _fill_summary(
            counterparty_output_df,
            required_by_sheet[SHEET_COUNTERPARTY],
            remarks_by_sheet[SHEET_COUNTERPARTY],
        ),
        '我方主体列表_必输字段未达100%': _fill_summary(
            our_party_output_df,
            required_by_sheet[SHEET_OUR_PARTY],
            remarks_by_sheet[SHEET_OUR_PARTY],
        ),
        '付款计划_必输字段未达100%': _fill_summary(
            payment_plan_output_df,
            required_by_sheet[SHEET_PAYMENT_PLAN],
            remarks_by_sheet[SHEET_PAYMENT_PLAN],
        ),
        '收款计划_必输字段未达100%': _fill_summary(
            collection_plan_output_df,
            required_by_sheet[SHEET_COLLECTION_PLAN],
            remarks_by_sheet[SHEET_COLLECTION_PLAN],
        ),
        '对方主体编码_未匹配': collect_missing_counterparty(counterparty_source_df),
        '我方主体编码_未匹配': collect_missing_our_party(our_party_source_df),
        '关联合同编号_未匹配': collect_missing_relation(relation_source_df),
        '合同附件下载清单': contract_attachment_meta_df,
        '合同附件DOCID_缺失映射': contract_attachment_missing_df,
    }
    exception_sheets.update({
        f'字段模板_{name}': df
        for name, df in _collect_missing_details(
            main_output_df,
            source_df,
            required_by_sheet[SHEET_MAIN],
            MAIN_ISSUE_SOURCE_FIELDS,
            'contract_number（合同编码）',
        ).items()
    })
    exception_sheets.update({
        f'订单信息明细_{name}': df
        for name, df in _collect_missing_details(
            order_detail_output_df,
            order_detail_source_df,
            required_by_sheet[SHEET_ORDER_DETAIL],
            ORDER_ISSUE_SOURCE_FIELDS,
            'contract_number（合同编码）',
        ).items()
    })
    exception_sheets.update(c.collect_order_mapping_issues(
        source_df,
        doc_col='合同编号',
        project_col='项目编号',
        project_id_col='合同所属项目编号ID',
        project_name_col='项目名称',
    ))

    exception_file = _write_exceptions_with_fallback(EXCEPTION_FILE, exception_sheets)
    if exception_file:
        print('已写出:', exception_file)


if __name__ == '__main__':
    run()
