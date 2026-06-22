# -*- coding: utf-8 -*-
"""合同迁移 —— 智书主播流程(DB 直连版)。

按「智书合同字段-主播流程.xlsx」生成主播流程导入模板的 5 个 sheet:
    1. 字段模板
    2. 对方信息
    3. 我方信息
    4. 费用明细
    5. 选项(模板自带,不写数据)

跑法:在项目根执行  python run.py contract_anchor_db
"""
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

if __package__ is None or __package__ == '':
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from etl import common as c


# ============================ 文件 / 模板 ============================
TASK_NAME = 'contract_anchor_db'
TEMPLATE_DIR = c.TPL_DIR / TASK_NAME
OUTPUT_DIR = c.OUT_DIR / TASK_NAME
DATE_SUFFIX = c.today_suffix()

TEMPLATE_FILE = TEMPLATE_DIR / '智书合同字段-主播流程.xlsx'
RULE_CSV = c.RULES_DIR / '业财项目_数据映射规则 - 合同数据映射规则-for法务.csv'
OUTPUT_FILE = OUTPUT_DIR / f'智书合同字段_主播流程_{DATE_SUFFIX}.xlsx'
EXCEPTION_FILE = OUTPUT_DIR / f'未匹配清单_主播流程_{DATE_SUFFIX}.xlsx'

SHEET_MAIN = '字段模板'
SHEET_OPTIONS = '选项'
SHEET_COUNTERPARTY = '对方信息'
SHEET_OUR_PARTY = '我方信息'
SHEET_FEE_DETAIL = '费用明细'

FW_TABLE = 'uf_htk'
FW_ANCHOR_CARD_TABLE = 'uf_zbkp'
FW_ANCHOR_CARD_FALLBACK_TABLE = 'uf_zbkp_bak1'
FW_ANCHOR_CARD_DETAIL_TABLE = 'uf_zbkp_dt1'
FW_ANCHOR_CARD_DETAIL_FALLBACK_TABLE = 'uf_zbkp_dt1_bak1'
PROJECT_TABLES = ('uf_xtyyxmkp', 'uf_xmkp', 'view_xmjkzb')

MIGRATION_STATUS_CODES = (1, 2)
ANCHOR_CONTRACT_TYPE_CODE = 3
ANCHOR_SECONDARY_CODES = (10, 11, 12, 13, 14, 15)
SUPPLEMENT_WORKFLOW_IDS = {'48', '442'}
ANCHOR_APPROVAL_WORKFLOW_NAME = '主播协议审批流程'
CATEGORY_INDONESIA_LIVE = '主播专项-印尼直播'
CATEGORY_MALAYSIA_LIVE = '主播专项-马来直播'
CATEGORY_PLATFORM_ECONOMY = '主播专项-平台经纪'
CATEGORY_ANCHOR_ECONOMY = '主播专项-主播经纪'

DEFAULT_PROPERTY_TYPE = '固定总价'
DEFAULT_VALIDITY_TYPE = '固定期限'
DEFAULT_ACCEPTANCE_REQUIRED = '否'
DEFAULT_PRINT_MODE = '黑白双面打印'
DEFAULT_FIRST_SEAL_PARTY = '我方'
DEFAULT_SIGN_FORM = '纸质签约-不限制我方/对方先签约'
DEFAULT_SEAL_NUMBER = 2
DEFAULT_PAYBACK_PERIOD_MONTHS = 0


# ============================ 泛微源 SQL ============================
SOURCE_SQL = """
SELECT
    h.id AS `ID`,
    h.htbh AS `合同编号`,
    h.htbt AS `合同标题`,
    h.htlx AS `合同类型ID`,
    h.htejlx AS `合同二级类型ID`,
    h.htzt AS `合同签署状态ID`,
    h.htlc AS `合同流程ID`,
    h.htcjrq AS `合同创建日期`,
    h.htqdrq AS `合同签订日期`,
    h.htyxqqssj AS `合同有效期起始时间`,
    h.htyxqjzsj AS `合同有效期截止时间`,
    h.htzhry AS `合同执行人员ID`,
    h.htyyfw AS `合同用印范围ID`,
    h.htkh AS `合同客户ID`,
    h.htgys AS `合同供应商ID`,
    h.zbid AS `主播卡片ID`,
    h.zbxm AS `主播姓名`,
    h.zbnc AS `主播昵称`,
    h.zdmc AS `战队名称`,
    h.szpt AS `所属平台ID`,
    h.yjhbzqy AS `预计回本周期（月）`,
    h.htje AS `合同金额`,
    h.htyjsr AS `合同预计收入`,
    h.htyjzc AS `合同预计支出`,
    h.htzy AS `合同摘要`,
    h.glkjxy AS `关联框架协议ID`,
    h.bglx AS `变更类型ID`,
    h.bczzbhsc AS `补充/终止编号生成`,
    h.htszxmbh AS `合同所属项目编号ID`,
    h.htszxm AS `合同所属项目`,
    rb.workflowid AS `流程类型ID`,
    wb.workflowname AS `流程名称`
FROM uf_htk h
LEFT JOIN workflow_requestbase rb ON rb.requestid = h.htlc
LEFT JOIN workflow_base wb ON wb.id = rb.workflowid
WHERE h.htlx = %(anchor_contract_type_code)s
  AND h.htzt IN %(migration_status_codes)s
ORDER BY h.htbh, h.id
"""

STATS_SQL = """
SELECT
    COUNT(*) AS all_count,
    SUM(CASE WHEN htlx = %(anchor_contract_type_code)s THEN 1 ELSE 0 END) AS anchor_type_count,
    SUM(CASE
        WHEN htlx = %(anchor_contract_type_code)s
         AND htzt IN %(migration_status_codes)s
        THEN 1 ELSE 0 END) AS kept_count,
    SUM(CASE
        WHEN htlx = %(anchor_contract_type_code)s
         AND (htzt NOT IN %(migration_status_codes)s OR htzt IS NULL)
        THEN 1 ELSE 0 END) AS excluded_status_count,
    SUM(CASE
        WHEN htlx = %(anchor_contract_type_code)s
         AND htzt IN %(migration_status_codes)s
         AND (htje IS NULL OR htje = 0)
         AND COALESCE(htyjsr, 0) + COALESCE(htyjzc, 0) <> 0
        THEN 1 ELSE 0 END) AS amount_fallback_count
FROM uf_htk
"""

STATUS_BREAKDOWN_SQL = """
SELECT htzt AS `合同签署状态ID`, COUNT(*) AS `合同数`
FROM uf_htk
WHERE htlx = %(anchor_contract_type_code)s
GROUP BY htzt
ORDER BY htzt
"""

EXPECTED_FW_FIELDS = {
    '': {
        'htbh': '合同编号',
        'htbt': '合同标题',
        'htlx': '合同类型',
        'htejlx': '合同二级类型',
        'htzt': '合同签署状态',
        'htlc': '合同流程',
        'htqdrq': '合同签订日期',
        'htyxqqssj': '合同有效期起始时间',
        'htyxqjzsj': '合同有效期截止时间',
        'htzhry': '合同执行人员',
        'htyyfw': '合同用印范围',
        'htkh': '合同客户',
        'htgys': '合同供应商',
        'zbid': '主播ID',
        'zbxm': '主播姓名',
        'zbnc': '主播昵称',
        'zdmc': '战队名称',
        'szpt': '所属平台',
        'yjhbzqy': '预计回本周期（月）',
        'htje': '合同金额',
        'htyjsr': '合同预计收入',
        'htyjzc': '合同预计支出',
        'htzy': '合同摘要',
        'glkjxy': '关联框架协议',
        'bglx': '变更类型',
        'bczzbhsc': '补充/终止编号生成',
        'htszxmbh': '合同所属项目编号',
    },
}


# ============================ 小工具 ============================
def _query_fw(sql):
    return c.query_db('FW', 'vspn_xtyy', sql, {
        'anchor_contract_type_code': ANCHOR_CONTRACT_TYPE_CODE,
        'migration_status_codes': MIGRATION_STATUS_CODES,
    })


def _text(value):
    if pd.isna(value):
        return ''
    text = str(value).strip()
    return '' if text in ('', 'nan', 'None', 'NaT') else text


def _first_non_blank(*values):
    for value in values:
        text = _text(value)
        if text:
            return text
    return ''


def _first_browser_code(*values):
    for value in values:
        for item_id in c.parse_browser_ids(value):
            return item_id
        code = c.format_code(value)
        if code:
            return code
    return ''


def _first_browser_value(mapping, value):
    for item_id in c.parse_browser_ids(value):
        mapped = mapping.get(item_id, '')
        if mapped:
            return mapped
    return ''


def _excel_id_value(*values):
    code = _first_browser_code(*values)
    return int(code) if code.isdigit() else code


def _number(value, default=0.0):
    if pd.isna(value) or _text(value) == '':
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _round_amount(value):
    return round(_number(value), 2)


def _normalize_field_name(value):
    return re.sub(r'\s+', '', _text(value))


def _sheet_headers(sheet_name):
    wb = load_workbook(TEMPLATE_FILE, read_only=True, data_only=True)
    ws = wb[sheet_name]
    return [_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]


def _template_headers():
    return {
        sheet_name: _sheet_headers(sheet_name)
        for sheet_name in (SHEET_MAIN, SHEET_COUNTERPARTY, SHEET_OUR_PARTY, SHEET_FEE_DETAIL)
    }


def _timestamped_path(path):
    return path.with_name(f'{path.stem}_{datetime.now().strftime("%H%M%S")}{path.suffix}')


def _write_template_sheets_with_fallback(template_file, output_file, sheet_to_df):
    try:
        return c.write_template_sheets(template_file, output_file, sheet_to_df)
    except PermissionError:
        fallback = _timestamped_path(output_file)
        print(f'输出文件被占用,改写到: {fallback}')
        return c.write_template_sheets(template_file, fallback, sheet_to_df)


def _write_exceptions_with_fallback(output_file, sheets):
    try:
        return c.write_exceptions(output_file, sheets)
    except PermissionError:
        fallback = _timestamped_path(output_file)
        print(f'未匹配清单被占用,改写到: {fallback}')
        return c.write_exceptions(fallback, sheets)


def _add_flow_audit_sheet(output_file, source_df):
    """在产出文件追加流程口径核对 sheet,不改变智书导入模板字段。"""
    sheet_name = '流程名称核对'
    wb = load_workbook(output_file)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    headers = [
        '合同编号', '合同标题', '合同流程ID', '流程类型ID', '流程名称',
        '合同类型ID', '合同类型', '合同二级类型ID', '合同二级类型', '合同一级类型判定',
        '主合同匹配方式', '主合同编号', '主合同标题', '主合同流程类型ID', '主合同流程名称',
        '主合同类型ID', '主合同类型', '主合同二级类型ID', '主合同二级类型',
        '主合同一级类型判定', '主合同主播卡片ID', '主合同主播姓名', '主合同主播昵称',
        '主合同战队名称', '主合同所属平台ID', '主播信息来源',
        '最终主播卡片导入ID', '最终主播卡片编号', '最终主播身份证号',
        '主合同智书合同类型', '智书合同类型', '分类依据',
    ]
    ws.append(headers)
    for row in source_df.to_dict('records'):
        ws.append([
            _text(row.get('合同编号')),
            _text(row.get('合同标题')),
            _text(row.get('合同流程ID')),
            _text(row.get('流程类型ID')),
            _text(row.get('流程名称')),
            _text(row.get('合同类型ID')),
            _text(row.get('合同类型')),
            _text(row.get('合同二级类型ID')),
            _text(row.get('合同二级类型')),
            _text(row.get('合同一级类型判定')),
            _text(row.get('主合同匹配方式')),
            _text(row.get('主合同编号')),
            _text(row.get('主合同标题')),
            _text(row.get('主合同流程类型ID')),
            _text(row.get('主合同流程名称')),
            _text(row.get('主合同类型ID')),
            _text(row.get('主合同类型')),
            _text(row.get('主合同二级类型ID')),
            _text(row.get('主合同二级类型')),
            _text(row.get('主合同一级类型判定')),
            _text(row.get('主合同主播卡片ID')),
            _text(row.get('主合同主播姓名')),
            _text(row.get('主合同主播昵称')),
            _text(row.get('主合同战队名称')),
            _text(row.get('主合同所属平台ID')),
            _text(row.get('主播信息来源')),
            _text(row.get('主播卡片导入ID')),
            _text(row.get('主播卡片编号')),
            _text(row.get('主播身份证号码')),
            _text(row.get('主合同分类')),
            _text(row.get('合同分类')),
            _text(row.get('合同分类依据')),
        ])
    for column_cells in ws.columns:
        max_length = max(len(_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)
    wb.save(output_file)


def _add_platform_audit_sheet(output_file, source_df):
    """追加所属平台取值核对 sheet。"""
    sheet_name = '所属平台核对'
    wb = load_workbook(output_file)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    headers = [
        '合同编号', '合同标题', '主播信息来源', '原合同主播卡片ID',
        '原合同主播姓名', '原合同主播昵称', '主合同编号', '主合同标题',
        '主合同主播卡片ID', '主合同主播姓名', '主合同主播昵称',
        '主播卡片ID', '主播姓名', '主播昵称',
        '最终主播卡片导入ID', '最终主播卡片编号', '最终主播身份证号',
        '合同主表所属平台ID', '合同主表所属平台', '主播卡片平台明细',
        '合同标题平台命中', '最终所属平台ID', '最终所属平台', '取值依据',
    ]
    ws.append(headers)
    for row in source_df.to_dict('records'):
        ws.append([
            _text(row.get('合同编号')),
            _text(row.get('合同标题')),
            _text(row.get('主播信息来源')),
            _text(row.get('原合同主播卡片ID')),
            _text(row.get('原合同主播姓名')),
            _text(row.get('原合同主播昵称')),
            _text(row.get('主合同编号')),
            _text(row.get('主合同标题')),
            _text(row.get('主合同主播卡片ID')),
            _text(row.get('主合同主播姓名')),
            _text(row.get('主合同主播昵称')),
            _text(row.get('主播卡片ID')),
            _text(row.get('主播姓名')),
            _text(row.get('主播昵称')),
            _text(row.get('主播卡片导入ID')),
            _text(row.get('主播卡片编号')),
            _text(row.get('主播身份证号码')),
            _text(row.get('所属平台ID')),
            _text(row.get('合同主表所属平台')),
            _text(row.get('主播卡片平台明细')),
            _text(row.get('合同标题平台命中')),
            _text(row.get('所属平台ID_解析')),
            _text(row.get('所属平台')),
            _text(row.get('所属平台解析依据')),
        ])
    for column_cells in ws.columns:
        max_length = max(len(_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)
    wb.save(output_file)


def _new_row(headers):
    return {header: '' for header in headers}


_HEADER_LOOKUP_CACHE = {}
_CLEANED_PROJECT_CANDIDATES_CACHE = None
_CONTRACT_ORDER_MAPPING_CACHE = None


def _set(row, field_name, value):
    normalized = _normalize_field_name(field_name)
    cache_key = tuple(row.keys())
    lookup = _HEADER_LOOKUP_CACHE.get(cache_key)
    if lookup is None:
        lookup = {
            _normalize_field_name(header): header
            for header in row
        }
        _HEADER_LOOKUP_CACHE[cache_key] = lookup
    header = lookup.get(normalized)
    if header:
        row[header] = value
        return
    raise KeyError(f'模板缺少字段: {field_name}')


def build_fw_project_info_map_for_ids(project_values):
    """泛微项目浏览框 ID -> 泛微项目编号/名称,兼容历史项目表。"""
    project_ids = c.clean_codes(
        project_id
        for value in project_values
        for project_id in c.parse_browser_ids(value)
    )
    if not project_ids:
        return {}

    result = {}
    remaining = set(project_ids)
    for table in PROJECT_TABLES:
        if not remaining:
            break
        ids = [project_id for project_id in project_ids if project_id in remaining]
        try:
            project_df = c.query_db(
                'FW',
                'vspn_xtyy',
                f'SELECT id, xmbh AS project_code, xmmc AS project_name FROM {table} '
                f'WHERE id IN ({c.in_placeholders(ids)})',
                ids,
            )
        except Exception:
            try:
                project_df = c.query_db(
                    'FW',
                    'vspn_xtyy',
                    f"SELECT id, xmbh AS project_code, '' AS project_name FROM {table} "
                    f'WHERE id IN ({c.in_placeholders(ids)})',
                    ids,
                )
            except Exception:
                continue
        for _, row in project_df.iterrows():
            project_id = c.format_code(row['id'])
            project_code = _text(row['project_code'])
            if project_id and project_code and project_id not in result:
                result[project_id] = {
                    'code': project_code,
                    'name': _text(row.get('project_name', '')),
                }
                remaining.discard(project_id)
    return result


def _read_cleaned_sheet_with_header(path, sheet_name, required_columns):
    raw_df = pd.read_excel(path, sheet_name=sheet_name, dtype=str, keep_default_na=False, header=None)
    required = set(required_columns)
    for index, row in raw_df.iterrows():
        headers = [_text(value).replace('\n', '') for value in row.tolist()]
        if required.issubset(set(headers)):
            df = raw_df.iloc[index + 1:].copy()
            df.columns = headers
            return df
    return pd.DataFrame(columns=list(required_columns))


def _cleaned_mapping_file():
    _, _, mapping_file = c.load_project_order_mapping()
    return mapping_file


def load_cleaned_project_candidates():
    """0621 清洗表里的项目ID -> 原泛微项目编码候选。主播合同 xmk ID 与赛事项目 ID 会重号,需按项目名优先。"""
    global _CLEANED_PROJECT_CANDIDATES_CACHE
    if _CLEANED_PROJECT_CANDIDATES_CACHE is not None:
        return _CLEANED_PROJECT_CANDIDATES_CACHE

    mapping_file = _cleaned_mapping_file()
    if mapping_file is None:
        _CLEANED_PROJECT_CANDIDATES_CACHE = {}
        return _CLEANED_PROJECT_CANDIDATES_CACHE

    candidates = {}
    sheet_configs = [
        ('MCN专项项目_清洗前', 'MCN'),
        ('赛事专项项目_清洗前', '赛事'),
    ]
    for sheet_name, source in sheet_configs:
        try:
            project_df = _read_cleaned_sheet_with_header(
                mapping_file,
                sheet_name,
                ['id', 'prj_dim_value（项目编号）', 'project_name（项目名称）'],
            )
        except ValueError:
            continue
        for _, row in project_df.iterrows():
            project_id = c.format_code(row.get('id'))
            project_code = _text(row.get('prj_dim_value（项目编号）'))
            project_name = _text(row.get('project_name（项目名称）'))
            if not project_id or not project_code:
                continue
            candidates.setdefault(project_id, []).append({
                'code': project_code,
                'name': project_name,
                'source': source,
            })
    _CLEANED_PROJECT_CANDIDATES_CACHE = candidates
    return candidates


def choose_cleaned_project_info(project_id, project_name, candidates_map):
    candidates = candidates_map.get(c.format_code(project_id), [])
    if not candidates:
        return {}
    name_key = c.normalize_name(project_name)
    if name_key:
        for candidate in candidates:
            if c.normalize_name(candidate.get('name', '')) == name_key:
                return candidate
    for candidate in candidates:
        if candidate.get('source') == 'MCN':
            return candidate
    return candidates[0]


def load_contract_order_mapping():
    """从 0621 订单/结算表读取 原泛微项目编码 -> 唯一订单编号。"""
    global _CONTRACT_ORDER_MAPPING_CACHE
    if _CONTRACT_ORDER_MAPPING_CACHE is not None:
        return _CONTRACT_ORDER_MAPPING_CACHE

    mapping_file = _cleaned_mapping_file()
    if mapping_file is None:
        _CONTRACT_ORDER_MAPPING_CACHE = ({}, {})
        return _CONTRACT_ORDER_MAPPING_CACHE

    sheet_names = [
        '全量订单主表_清洗后',
        '全量订单明细行表_清洗后',
        '全量结算明细表_清洗后',
        '填协同关系表',
        '全量订单明细表（下单或报价行）_清洗后',
    ]
    rows = []
    for sheet_name in sheet_names:
        try:
            order_df = _read_cleaned_sheet_with_header(mapping_file, sheet_name, ['原泛微项目编码', '订单编号'])
        except ValueError:
            continue
        if '是否可洗流程' in order_df.columns:
            order_df = order_df[order_df['是否可洗流程'].map(_text).str.upper().str.contains('Y', na=False)]
        for _, row in order_df.iterrows():
            project_code = _text(row.get('原泛微项目编码'))
            order_code = _text(row.get('订单编号'))
            if project_code and order_code:
                rows.append({
                    '项目编号': project_code,
                    '订单编号': order_code,
                    '映射来源': sheet_name,
                })

    by_project = {}
    for row in rows:
        by_project.setdefault(row['项目编号'], []).append(row)

    safe_map = {}
    ambiguous_map = {}
    for project_code, items in by_project.items():
        order_codes = []
        seen = set()
        for item in items:
            order_code = item['订单编号']
            if order_code not in seen:
                seen.add(order_code)
                order_codes.append(order_code)
        if len(order_codes) == 1:
            safe_map[project_code] = order_codes[0]
        else:
            ambiguous_map[project_code] = order_codes
    _CONTRACT_ORDER_MAPPING_CACHE = (safe_map, ambiguous_map)
    return _CONTRACT_ORDER_MAPPING_CACHE


def contract_order_mapping_value(project_code):
    safe_map, _, = load_contract_order_mapping()
    return safe_map.get(_text(project_code), '') or c.project_order_mapping_value(project_code, '订单编号')


def _read_anchor_required_rules(headers_by_sheet):
    raw = pd.read_csv(RULE_CSV, encoding='utf-8-sig').iloc[1:, :15].copy()
    raw.columns = [
        'module', 'flow', 'table_name', 'field_name', 'required', 'enum', 'note',
        'event_table', 'event_field', 'event_enum', 'event_note',
        'mcn_table', 'mcn_field', 'mcn_enum', 'mcn_note',
    ]
    for column in raw.columns:
        raw[column] = raw[column].where(raw[column].notna(), '')
    raw['table_name'] = raw['table_name'].replace('', pd.NA).ffill().fillna('')
    raw = raw[raw['flow'] == '主播流程']

    table_to_sheet = {
        '字段模板（主表）': SHEET_MAIN,
        '主播流程_对方信息': SHEET_COUNTERPARTY,
        '主播流程_我方信息': SHEET_OUR_PARTY,
        '主播流程_费用明细': SHEET_FEE_DETAIL,
    }
    required = {sheet_name: [] for sheet_name in headers_by_sheet}
    remarks = {sheet_name: {} for sheet_name in headers_by_sheet}
    header_lookup = {
        sheet_name: {_normalize_field_name(header): header for header in headers}
        for sheet_name, headers in headers_by_sheet.items()
    }

    for _, rule in raw.iterrows():
        sheet_name = table_to_sheet.get(_text(rule['table_name']))
        if not sheet_name:
            continue
        normalized_field = _normalize_field_name(rule['field_name'])
        actual_field = header_lookup[sheet_name].get(normalized_field)
        if not actual_field:
            continue
        note = _first_non_blank(rule['note'], rule['mcn_note'])
        if note and actual_field not in remarks[sheet_name]:
            remarks[sheet_name][actual_field] = note
        if _text(rule['required']).upper() == 'Y' and actual_field not in required[sheet_name]:
            required[sheet_name].append(actual_field)
    return required, remarks


def _fill_summary(output_df, required_cols, remarks=None):
    total = len(output_df)
    rows = []
    for column in required_cols:
        column_exists = column in output_df.columns
        filled = int((output_df[column].astype(str).str.strip() != '').sum()) if column_exists else 0
        if filled < total:
            note = (remarks or {}).get(column, '')
            rows.append({
                '必输字段': column,
                '填充数': filled,
                '缺失数': total - filled,
                '总数': total,
                '填充率': '0.00%' if total == 0 else f'{filled / total * 100:.2f}%',
                '备注': note,
            })
    return pd.DataFrame(rows, columns=['必输字段', '填充数', '缺失数', '总数', '填充率', '备注'])


def _collect_missing_details(output_df, source_df, required_cols, source_field_map, doc_col):
    sheets = {}
    total = len(output_df)
    if total == 0 or doc_col not in output_df.columns:
        return sheets
    for column in required_cols:
        if column not in output_df.columns:
            continue
        blank_mask = output_df[column].astype(str).str.strip() == ''
        missing_count = int(blank_mask.sum())
        if not (0 < missing_count < total):
            continue
        data = {doc_col: output_df.loc[blank_mask, doc_col].astype(str)}
        source_field = source_field_map.get(column)
        if source_field and source_field in source_df.columns:
            data[f'泛微原表-{source_field}'] = source_df.loc[blank_mask, source_field].astype(str)
        sheets[f'缺失_{column[:22]}'] = pd.DataFrame(data).drop_duplicates().reset_index(drop=True)
    return sheets


def _resolve_first_browser_value(mapping, value, field):
    for item_id in c.parse_browser_ids(value):
        mapped = mapping.get(item_id, {})
        value = mapped.get(field, '') if isinstance(mapped, dict) else mapped
        if value:
            return value
    return ''


# ============================ 映射构建 ============================
def build_fw_company_info_map_for_values(company_values):
    company_ids = c.clean_codes(
        company_id
        for value in company_values
        for company_id in c.parse_browser_ids(value)
    )
    if not company_ids:
        return {}
    company_name_map = c.build_fw_company_name_map_for_ids(company_ids)
    entity_map = c.build_accounting_entity_map_for_names(company_name_map.values())
    return {
        company_id: {
            'name': company_name,
            'code': entity_map.get(c.normalize_name(company_name), ''),
        }
        for company_id, company_name in company_name_map.items()
    }


def build_customer_info_map_for_values(customer_values):
    customer_name_map = c.build_fw_customer_name_map_for_ids(customer_values)
    same_customer_map = c.load_same_customer_mapping(log_prefix='[合同迁移-主播流程]')
    source_to_target = {
        customer_id: c.resolve_same_customer_id(customer_id, same_customer_map)
        for customer_id in customer_name_map
    }
    target_name_map = c.build_fw_customer_name_map_for_ids(source_to_target.values())
    customer_by_target_id = c.build_hand_customer_info_by_ids(source_to_target.values())
    customer_code_map = c.build_customer_map_for_names(
        list(customer_name_map.values()) + list(target_name_map.values()))

    result = {}
    for customer_id, target_id in source_to_target.items():
        source_name = customer_name_map.get(customer_id, '')
        target_name = target_name_map.get(target_id, '')
        customer_info = customer_by_target_id.get(target_id, {})
        code = _first_non_blank(
            customer_info.get('code', ''),
            customer_code_map.get(c.normalize_name(target_name), ''),
            customer_code_map.get(c.normalize_name(source_name), ''),
        )
        name = _first_non_blank(customer_info.get('name', ''), target_name, source_name)
        result[customer_id] = {
            'name': name,
            'code': code,
            'source_name': source_name,
            'target_id': target_id,
            'match_method': '客户归并ID' if target_id != customer_id else customer_info.get('match_method', '客户名称'),
        }
    return result


def build_supplier_info_map_for_values(supplier_values):
    supplier_status_map = c.build_fw_supplier_status_map(supplier_values)
    same_supplier_map = c.load_same_supplier_mapping(log_prefix='[合同迁移-主播流程]')
    name_match_cache = c.load_supplier_vendor_name_match_map(log_prefix='[合同迁移-主播流程]')
    source_to_target = {
        supplier_id: c.resolve_same_supplier_id(supplier_id, same_supplier_map)
        for supplier_id in supplier_status_map
    }
    target_status_map = c.build_fw_supplier_status_map(source_to_target.values())
    vendor_by_target_id = c.build_hand_vendor_info_by_ids(source_to_target.values())

    vendor_by_source_id = {}
    names_to_match = []
    disabled_unmatched_ids = []
    for supplier_id, target_id in source_to_target.items():
        supplier_info = supplier_status_map.get(supplier_id, {})
        vendor_info = vendor_by_target_id.get(target_id, {})
        if vendor_info.get('code'):
            vendor_by_source_id[supplier_id] = vendor_info
            continue
        if supplier_info.get('status_code') != '0':
            cached_info = name_match_cache.get(supplier_id, {})
            if cached_info.get('code'):
                vendor_by_source_id[supplier_id] = cached_info
                continue
            supplier_name = supplier_info.get('name', '')
            if supplier_name:
                names_to_match.append(supplier_name)
                disabled_unmatched_ids.append(supplier_id)

    vendor_by_name = c.build_hand_vendor_info_by_names(names_to_match)
    discovered_rows = []
    for supplier_id in disabled_unmatched_ids:
        supplier_info = supplier_status_map.get(supplier_id, {})
        supplier_name = supplier_info.get('name', '')
        vendor_info = vendor_by_name.get(c.normalize_name(supplier_name), {})
        if not vendor_info.get('code'):
            continue
        vendor_by_source_id[supplier_id] = vendor_info
        discovered_rows.append({
            '供应商泛微Id': supplier_id,
            '供应商名称': supplier_name,
            '泛微状态': supplier_info.get('status_code', ''),
            '匹配方式': '禁用供应商按名称匹配汉得供应商',
            'handVendorId': vendor_info.get('id', ''),
            'handVendorCode': vendor_info.get('code', ''),
            'handVendorName': vendor_info.get('name', ''),
            'handTaxpayerName': vendor_info.get('taxpayer_name', ''),
        })
    changed_count = c.append_supplier_vendor_name_matches(discovered_rows)
    if changed_count:
        print(f'[合同迁移-主播流程] 禁用供应商按名称匹配汉得并写入缓存: {changed_count} 条')

    result = {}
    for supplier_id, target_id in source_to_target.items():
        vendor_info = vendor_by_source_id.get(supplier_id) or vendor_by_target_id.get(target_id, {})
        supplier_info = supplier_status_map.get(supplier_id, {})
        target_supplier_info = supplier_status_map.get(target_id) or target_status_map.get(target_id, {})
        target_name = target_supplier_info.get('name', '')
        match_method = vendor_info.get('match_method', '')
        if target_id != supplier_id and vendor_info.get('code'):
            match_method = '供应商归并ID'
        result[supplier_id] = {
            'name': _first_non_blank(vendor_info.get('name', ''), target_name, supplier_info.get('name', '')),
            'code': vendor_info.get('code', ''),
            'source_name': supplier_info.get('name', ''),
            'target_id': target_id,
            'target_name': target_name,
            'status_code': supplier_info.get('status_code', ''),
            'match_method': match_method or ('供应商名称缓存' if name_match_cache.get(supplier_id) else '供应商ID'),
        }
    return result


def build_anchor_card_info_map(card_values):
    card_ids = c.clean_codes(
        card_id
        for value in card_values
        for card_id in c.parse_browser_ids(value)
    )
    if not card_ids:
        return {}

    card_rows = []
    remaining_ids = set(card_ids)
    for table_name, source_label in (
            (FW_ANCHOR_CARD_TABLE, 'uf_zbkp'),
            (FW_ANCHOR_CARD_FALLBACK_TABLE, 'uf_zbkp_bak1')):
        if not remaining_ids:
            break
        ids = [card_id for card_id in card_ids if card_id in remaining_ids]
        card_df = c.query_db(
            'FW',
            'vspn_xtyy',
            'SELECT id, zbbh, xm, nc, sfzh, zdmc, gqqyj, gqqyjfcbl, bdxzmy, gsqyj '
            f'FROM {table_name} '
            f'WHERE id IN ({c.in_placeholders(ids)})',
            ids,
        )
        for _, row in card_df.iterrows():
            card_id = c.format_code(row['id'])
            if card_id and card_id in remaining_ids:
                row_data = row.to_dict()
                row_data['source_table'] = source_label
                card_rows.append(row_data)
                remaining_ids.discard(card_id)

    def detail_table_columns(table_name):
        column_df = c.query_db(
            'FW',
            'information_schema',
            "SELECT column_name FROM columns WHERE table_schema='vspn_xtyy' AND table_name=%s",
            [table_name],
        )
        return set(column_df['column_name'])

    detail_rows = []
    for table_name in (FW_ANCHOR_CARD_DETAIL_TABLE, FW_ANCHOR_CARD_DETAIL_FALLBACK_TABLE):
        columns = detail_table_columns(table_name)
        if not {'mainid', 'id', 'zbid'} <= columns:
            continue
        platform_expr = 'pt1 AS platform_code' if 'pt1' in columns else "'' AS platform_code"
        status_expr = 'zt AS status_code' if 'zt' in columns else "'' AS status_code"
        detail_df = c.query_db(
            'FW',
            'vspn_xtyy',
            'SELECT mainid, id, zbid AS room_id, '
            f'{platform_expr}, {status_expr} '
            f'FROM {table_name} '
            f'WHERE mainid IN ({c.in_placeholders(card_ids)}) '
            'ORDER BY mainid, id',
            card_ids,
        )
        detail_rows.extend(row.to_dict() for _, row in detail_df.iterrows())

    details_by_card = {}
    for row in detail_rows:
        card_id = c.format_code(row['mainid'])
        details_by_card.setdefault(card_id, []).append({
            'room_id': _text(row['room_id']),
            'platform_code': c.format_code(row['platform_code']),
            'status_code': c.format_code(row['status_code']),
        })

    result = {}
    for row in card_rows:
        card_id = c.format_code(row['id'])
        result[card_id] = _anchor_card_info_from_row(row, details_by_card.get(card_id, []))
    print(f'[合同迁移-主播流程] 主播卡片匹配: {len(result)}/{len(card_ids)}; '
          f'其中历史备份表 {sum(1 for item in result.values() if item.get("source_table") == "uf_zbkp_bak1")} 条')
    return result


def _anchor_card_info_from_row(row, details):
    return {
        'card_id': c.format_code(row.get('id', '')),
        'card_number': _text(row.get('zbbh')),
        'name': _text(row.get('xm')),
        'nickname': _text(row.get('nc')),
        'id_number': _text(row.get('sfzh')),
        'team_name': _text(row.get('zdmc')),
        'official_signing_bonus': _round_amount(row.get('gqqyj')) if _text(row.get('gqqyj')) else '',
        'official_signing_bonus_share_ratio': (
            _number(row.get('gqqyjfcbl'), default='') if _text(row.get('gqqyjfcbl')) else ''
        ),
        'base_salary_monthly': _round_amount(row.get('bdxzmy')) if _text(row.get('bdxzmy')) else '',
        'company_signing_bonus': _round_amount(row.get('gsqyj')) if _text(row.get('gsqyj')) else '',
        'source_table': _text(row.get('source_table', '')),
        'details': details,
    }


def _identity_key(name, nickname):
    name_key = c.normalize_name(name)
    nickname_key = c.normalize_name(nickname)
    return f'{name_key}|{nickname_key}' if name_key and nickname_key else ''


def build_anchor_card_info_map_by_identity(source_df):
    """合同旧卡片缺资料时,按合同里的主播姓名+昵称到现行卡片表补充。"""
    names = c.clean_text_values(source_df['主播姓名'])
    nicknames = c.clean_text_values(source_df['主播昵称'])
    if not names and not nicknames:
        return {}

    conditions = []
    params = []
    if nicknames:
        conditions.append(f"nc IN ({c.in_placeholders(nicknames)})")
        params.extend(nicknames)
    if names:
        conditions.append(f"xm IN ({c.in_placeholders(names)})")
        params.extend(names)

    card_df = c.query_db(
        'FW',
        'vspn_xtyy',
        'SELECT id, zbbh, xm, nc, sfzh, zdmc, gqqyj, gqqyjfcbl, bdxzmy, gsqyj, '
        "'uf_zbkp_by_identity' AS source_table "
        'FROM uf_zbkp '
        f'WHERE {" OR ".join(conditions)}',
        params,
    )
    if card_df.empty:
        return {}

    card_ids = c.clean_codes(card_df['id'])
    detail_df = c.query_db(
        'FW',
        'vspn_xtyy',
        'SELECT mainid, id, zbid AS room_id, pt1 AS platform_code, zt AS status_code '
        'FROM uf_zbkp_dt1 '
        f'WHERE mainid IN ({c.in_placeholders(card_ids)}) '
        'ORDER BY mainid, status_code, id',
        card_ids,
    )
    details_by_card = {}
    for _, row in detail_df.iterrows():
        card_id = c.format_code(row['mainid'])
        details_by_card.setdefault(card_id, []).append({
            'room_id': _text(row['room_id']),
            'platform_code': c.format_code(row['platform_code']),
            'status_code': c.format_code(row['status_code']),
        })

    exact_groups = {}
    nickname_groups = {}
    name_groups = {}
    for _, row in card_df.iterrows():
        card_id = c.format_code(row['id'])
        info = _anchor_card_info_from_row(row.to_dict(), details_by_card.get(card_id, []))
        exact_groups.setdefault(_identity_key(row['xm'], row['nc']), []).append(info)
        nickname_key = c.normalize_name(row['nc'])
        name_key = c.normalize_name(row['xm'])
        if nickname_key:
            nickname_groups.setdefault(nickname_key, []).append(info)
        if name_key:
            name_groups.setdefault(name_key, []).append(info)

    result = {}
    for index, row in source_df.iterrows():
        exact_key = _identity_key(row['主播姓名'], row['主播昵称'])
        nickname_key = c.normalize_name(row['主播昵称'])
        name_key = c.normalize_name(row['主播姓名'])
        candidates = exact_groups.get(exact_key, []) if exact_key else []
        if len(candidates) != 1 and nickname_key:
            candidates = nickname_groups.get(nickname_key, [])
        if len(candidates) != 1 and name_key:
            candidates = name_groups.get(name_key, [])
        if len(candidates) == 1:
            result[index] = candidates[0]

    print(f'[合同迁移-主播流程] 按主播姓名/昵称补充现行卡片: {len(result)} 行')
    return result


def _merge_card_info(primary, fallback):
    merged = dict(primary or {})
    fallback = fallback or {}
    for key, value in fallback.items():
        if key == 'details':
            if not merged.get('details'):
                merged[key] = value
            continue
        if not _text(merged.get(key, '')) and _text(value):
            merged[key] = value
    return merged


def choose_anchor_room_id(card_info, platform_code):
    details = card_info.get('details', []) if isinstance(card_info, dict) else []
    if not details:
        return ''
    platform_code = c.format_code(platform_code)

    def score(detail):
        return (
            0 if platform_code and detail.get('platform_code') == platform_code else 1,
            0 if detail.get('status_code') == '0' else 1,
        )

    chosen = min(details, key=score)
    return chosen.get('room_id', '')


def _card_platform_codes(card_info, platform_map):
    details = card_info.get('details', []) if isinstance(card_info, dict) else []
    result = []
    seen = set()
    for detail in details:
        code = c.format_code(detail.get('platform_code'))
        if code and code in platform_map and code not in seen:
            seen.add(code)
            result.append(code)
    return result


def _platform_aliases(platform_map):
    aliases = {}
    for code, name in platform_map.items():
        if name:
            aliases.setdefault(code, set()).add(name)
    extra = {
        '2': {'B站', 'B 站', '哔哩哔哩', 'bilibili', 'Bilibili', 'BILI', 'bili'},
        '5': {'网易CC', 'CC直播', '网易cc', 'cc直播'},
        '7': {'抖音', 'Douyin'},
        '12': {'YouTube', 'youtube', 'Youtube'},
        '18': {'小红书', 'RED'},
        '19': {'视频号', '微信视频号'},
        '20': {'YY', 'yy', '歪歪'},
        '22': {'QQ直播', 'qq直播'},
        '24': {'爱奇艺', 'iQIYI'},
    }
    for code, names in extra.items():
        aliases.setdefault(code, set()).update(names)
    return {code: sorted(names, key=len, reverse=True) for code, names in aliases.items()}


def _title_platform_candidates(title, platform_map):
    text = _text(title)
    if not text:
        return []
    matches = []
    for code, aliases in _platform_aliases(platform_map).items():
        best_pos = None
        best_name = ''
        for alias in aliases:
            pos = text.lower().find(alias.lower())
            if pos >= 0 and (best_pos is None or pos < best_pos):
                best_pos = pos
                best_name = platform_map.get(code, alias)
        if best_pos is not None:
            matches.append((best_pos, code, best_name))
    matches.sort(key=lambda item: item[0])
    result = []
    seen = set()
    for _, code, name in matches:
        if code not in seen:
            seen.add(code)
            result.append((code, name))
    return result


def _title_platform_match_text(row, platform_map):
    title_sources = []
    if _is_supplement_or_termination_flow(row) and _text(row.get('主合同标题')):
        title_sources.append(('主合同标题', row.get('主合同标题')))
    title_sources.append(('合同标题', row.get('合同标题')))

    parts = []
    seen = set()
    for source, title in title_sources:
        for code, name in _title_platform_candidates(title, platform_map):
            if code not in seen:
                seen.add(code)
                parts.append(f'{source}:{name}')
    return '/'.join(parts)


def _first_title_platform_candidate(row, platform_map):
    title_sources = []
    if _is_supplement_or_termination_flow(row) and _text(row.get('主合同标题')):
        title_sources.append(('主合同标题', row.get('主合同标题')))
    title_sources.append(('合同标题', row.get('合同标题')))

    for source, title in title_sources:
        candidates = _title_platform_candidates(title, platform_map)
        if candidates:
            code, name = candidates[0]
            return code, name, source
    return '', '', ''


def resolve_anchor_platform(row, card_info, platform_map):
    card_codes = _card_platform_codes(card_info, platform_map)
    if len(card_codes) == 1:
        code = card_codes[0]
        return code, platform_map.get(code, ''), '主播卡片平台信息唯一'

    title_code, title_name, title_source = _first_title_platform_candidate(row, platform_map)
    if title_code:
        if len(card_codes) > 1:
            basis = f'主播卡片平台多个({"/".join(platform_map.get(item, item) for item in card_codes)}); {title_source}命中{title_name}'
        else:
            basis = f'主播卡片平台为空; {title_source}命中{title_name}'
        return title_code, title_name, basis

    source_code = c.format_code(row.get('所属平台ID'))
    if source_code and source_code in platform_map:
        return source_code, platform_map.get(source_code, ''), '合同主表所属平台兜底'
    return '', '', '未识别平台'


def _is_supplement_or_termination_flow(row):
    flow_id = c.format_code(row.get('流程类型ID'))
    flow_name = c.clean_fw_select_name(row.get('流程名称'))
    return flow_id in SUPPLEMENT_WORKFLOW_IDS or '合同补充/终止流程' in flow_name


def _is_empty_flow(row):
    return not c.format_code(row.get('流程类型ID')) and not c.clean_fw_select_name(row.get('流程名称'))


def _contract_type_label(contract_type, secondary_type):
    text = f'{_text(contract_type)}{_text(secondary_type)}'
    if '经纪' in text:
        return '经纪合同'
    if '直播平台' in text:
        return '直播平台合同'
    if '联合运营' in text:
        return '联合运营合同'
    if '自媒体' in text:
        return '自媒体合同'
    return _first_non_blank(secondary_type, contract_type, '空')


def _category_from_contract_number(contract_number):
    code = re.sub(r'\s+', '', _text(contract_number)).upper()
    compact_code = re.sub(r'[^A-Z0-9]', '', code)
    if compact_code.startswith('VSMB') or 'VSMB' in compact_code:
        return CATEGORY_INDONESIA_LIVE
    if compact_code.startswith('VT') or re.search(r'(^|[-_/])VT', code):
        return CATEGORY_MALAYSIA_LIVE
    if 'P' in compact_code:
        return CATEGORY_PLATFORM_ECONOMY
    return CATEGORY_ANCHOR_ECONOMY


def _category_number_basis(contract_number):
    code = _text(contract_number) or '空'
    category = _category_from_contract_number(contract_number)
    if category == CATEGORY_INDONESIA_LIVE:
        return f'合同编号={code}; 前缀命中 VSMB'
    if category == CATEGORY_MALAYSIA_LIVE:
        return f'合同编号={code}; 前缀命中 VT'
    if category == CATEGORY_PLATFORM_ECONOMY:
        return f'合同编号={code}; 编号包含 P'
    return f'合同编号={code}; 未命中 VSMB/VT/P'


def _main_contract_code_candidates(contract_number):
    code = _text(contract_number)
    candidates = []
    while code:
        next_code = re.sub(r'-(?:S\d+|N)$', '', code, flags=re.IGNORECASE)
        if next_code == code:
            break
        if next_code and next_code not in candidates:
            candidates.append(next_code)
        code = next_code
    return list(reversed(candidates))


def _main_contract_info_from_row(row, option_maps, source):
    type_id = c.format_code(row.get('htlx'))
    secondary_id = c.format_code(row.get('htejlx'))
    workflow_id = c.format_code(row.get('workflowid'))
    workflow_name = c.clean_fw_select_name(row.get('workflowname'))
    contract_type = option_maps['htlx'].get(type_id, '')
    secondary_type = option_maps['htejlx'].get(secondary_id, '')
    type_label = _contract_type_label(contract_type, secondary_type)
    return {
        'source': source,
        'id': c.format_code(row.get('id')),
        'number': _text(row.get('htbh')),
        'title': _text(row.get('htbt')),
        'type_id': type_id,
        'type': contract_type,
        'secondary_type_id': secondary_id,
        'secondary_type': secondary_type,
        'workflow_id': workflow_id,
        'workflow_name': workflow_name,
        'type_label': type_label,
        'category': _category_from_contract_number(row.get('htbh')),
        'anchor_card_id': _first_browser_code(row.get('zbid')),
        'anchor_name': _text(row.get('zbxm')),
        'anchor_nickname': _text(row.get('zbnc')),
        'team_name': _text(row.get('zdmc')),
        'platform_id': c.format_code(row.get('szpt')),
    }


def build_main_contract_info_map(source_df, option_maps):
    supplement_df = source_df[source_df.apply(_is_supplement_or_termination_flow, axis=1)]
    if supplement_df.empty:
        return {}

    main_ids = c.clean_codes(
        main_id
        for value in supplement_df['关联框架协议ID']
        for main_id in c.parse_browser_ids(value)
    )
    main_codes = []
    seen_codes = set()
    for contract_number in supplement_df['合同编号']:
        for code in _main_contract_code_candidates(contract_number):
            if code not in seen_codes:
                seen_codes.add(code)
                main_codes.append(code)

    by_id = {}
    if main_ids:
        main_df = c.query_db(
            'FW',
            'vspn_xtyy',
            'SELECT h.id, h.htbh, h.htbt, h.htlx, h.htejlx, h.zbid, h.zbxm, h.zbnc, h.zdmc, h.szpt, '
            'rb.workflowid, wb.workflowname '
            'FROM uf_htk h '
            'LEFT JOIN workflow_requestbase rb ON rb.requestid = h.htlc '
            'LEFT JOIN workflow_base wb ON wb.id = rb.workflowid '
            f'WHERE h.id IN ({c.in_placeholders(main_ids)})',
            main_ids,
        )
        by_id = {
            c.format_code(row['id']): _main_contract_info_from_row(row, option_maps, '关联框架协议ID')
            for _, row in main_df.iterrows()
        }

    by_code = {}
    if main_codes:
        main_df = c.query_db(
            'FW',
            'vspn_xtyy',
            'SELECT h.id, h.htbh, h.htbt, h.htlx, h.htejlx, h.zbid, h.zbxm, h.zbnc, h.zdmc, h.szpt, '
            'rb.workflowid, wb.workflowname '
            'FROM uf_htk h '
            'LEFT JOIN workflow_requestbase rb ON rb.requestid = h.htlc '
            'LEFT JOIN workflow_base wb ON wb.id = rb.workflowid '
            f'WHERE h.htbh IN ({c.in_placeholders(main_codes)}) '
            'ORDER BY h.htbh, h.id',
            main_codes,
        )
        for _, row in main_df.iterrows():
            code = _text(row['htbh'])
            if code and code not in by_code:
                by_code[code] = _main_contract_info_from_row(row, option_maps, '合同编号去后缀')

    result = {}
    for index, row in supplement_df.iterrows():
        info = None
        for code in _main_contract_code_candidates(row.get('合同编号')):
            info = by_code.get(code)
            if info:
                break
        if not info:
            for main_id in c.parse_browser_ids(row.get('关联框架协议ID')):
                info = by_id.get(main_id)
                if info:
                    break
        if info:
            result[index] = info
    print(f'[合同迁移-主播流程] 补充/终止流程主合同匹配(优先多次去后缀到最底层): {len(result)}/{len(supplement_df)}')
    return result


def apply_main_contract_anchor_info(df, main_contract_info_map):
    df['主播信息来源'] = '当前合同'
    df['原合同主播卡片ID'] = df['主播卡片ID']
    df['原合同主播姓名'] = df['主播姓名']
    df['原合同主播昵称'] = df['主播昵称']
    df['原合同战队名称'] = df['战队名称']
    df['原合同所属平台ID'] = df['所属平台ID']

    for index, info in main_contract_info_map.items():
        if not info:
            continue
        changed = False
        replacements = {
            '主播卡片ID': info.get('anchor_card_id', ''),
            '主播姓名': info.get('anchor_name', ''),
            '主播昵称': info.get('anchor_nickname', ''),
            '战队名称': info.get('team_name', ''),
            '所属平台ID': info.get('platform_id', ''),
        }
        for column, value in replacements.items():
            if _text(value):
                df.at[index, column] = value
                changed = True
        if changed:
            df.at[index, '主播信息来源'] = f'主合同:{info.get("number", "")}'
        else:
            df.at[index, '主播信息来源'] = f'主合同:{info.get("number", "")}; 主合同主播信息为空'
    return df


# ============================ 源值解析 ============================
def resolve_contract_category(row):
    if _is_supplement_or_termination_flow(row):
        return _first_non_blank(
            row.get('主合同分类'),
            _category_from_contract_number(row.get('合同编号')),
        )
    return _category_from_contract_number(row.get('合同编号'))


def resolve_contract_category_basis(row):
    flow_name = c.clean_fw_select_name(row.get('流程名称'))
    flow_id = c.format_code(row.get('流程类型ID'))
    type_label = _text(row.get('合同一级类型判定'))
    secondary_name = _text(row.get('合同二级类型'))
    if _is_supplement_or_termination_flow(row):
        main_category = _text(row.get('主合同分类'))
        if main_category:
            return (
                f'流程类型={flow_id or "空"}; 流程名称={flow_name or "空"}; '
                f'按去后缀主合同{_text(row.get("主合同编号")) or "空"}; '
                f'主合同分类={main_category}; 子合同与主合同保持一致'
            )
        return (
            f'流程类型={flow_id or "空"}; 流程名称={flow_name or "空"}; '
            f'未找到主合同,按当前{_category_number_basis(row.get("合同编号"))}兜底'
        )
    return _category_number_basis(row.get('合同编号'))


def resolve_pay_type(in_amount, out_amount):
    has_in = abs(_number(in_amount)) > 0
    has_out = abs(_number(out_amount)) > 0
    if has_in and has_out:
        return '既收又支'
    if has_in:
        return '收入类'
    if has_out:
        return '支出类'
    return '无金额'


def resolve_contract_amount(row):
    contract_amount = _number(row.get('合同金额'))
    if abs(contract_amount) > 0:
        return round(contract_amount, 2)
    in_amount = abs(_number(row.get('合同预计收入')))
    out_amount = abs(_number(row.get('合同预计支出')))
    return round(in_amount + out_amount, 2)


def resolve_source_values(source_df):
    df = source_df.copy()
    option_maps = c.build_fw_select_option_maps(
        FW_TABLE,
        ['htlx', 'htejlx', 'htzt', 'szpt', 'bglx'],
    )
    employee_map = c.build_fw_employee_info_map_for_ids(df['合同执行人员ID'])
    company_info_map = build_fw_company_info_map_for_values(df['合同用印范围ID'])
    customer_info_map = build_customer_info_map_for_values(df['合同客户ID'])
    supplier_info_map = build_supplier_info_map_for_values(df['合同供应商ID'])
    project_info_map = build_fw_project_info_map_for_ids(df['合同所属项目编号ID'])
    cleaned_project_candidates = load_cleaned_project_candidates()
    project_code_map = {
        project_id: info.get('code', '')
        for project_id, info in project_info_map.items()
    }
    project_name_map = {
        project_id: info.get('name', '')
        for project_id, info in project_info_map.items()
    }

    df['合同类型'] = df['合同类型ID'].map(lambda value: option_maps['htlx'].get(c.format_code(value), ''))
    df['合同二级类型'] = df['合同二级类型ID'].map(lambda value: option_maps['htejlx'].get(c.format_code(value), ''))
    df['合同签署状态'] = df['合同签署状态ID'].map(lambda value: option_maps['htzt'].get(c.format_code(value), ''))
    df['合同主表所属平台'] = df['所属平台ID'].map(lambda value: option_maps['szpt'].get(c.format_code(value), ''))
    df['所属平台'] = df['合同主表所属平台']
    df['变更类型'] = df['变更类型ID'].map(lambda value: option_maps['bglx'].get(c.format_code(value), ''))
    df['流程名称'] = df['流程名称'].map(c.clean_fw_select_name)
    cleaned_project_info = [
        choose_cleaned_project_info(project_id, project_name, cleaned_project_candidates)
        for project_id, project_name in zip(df['合同所属项目编号ID'], df['合同所属项目'])
    ]
    df['合同所属项目编号'] = [
        _first_non_blank(
            info.get('code', ''),
            _first_browser_value(project_code_map, project_id),
            project_id,
        )
        for info, project_id in zip(cleaned_project_info, df['合同所属项目编号ID'])
    ]
    df['合同所属项目名称'] = [
        _first_non_blank(source_name, info.get('name', ''), _first_browser_value(project_name_map, value))
        for source_name, info, value in zip(df['合同所属项目'], cleaned_project_info, df['合同所属项目编号ID'])
    ]
    df['订单编号'] = df['合同所属项目编号'].map(contract_order_mapping_value)
    df['合同执行人员'] = df['合同执行人员ID'].map(
        lambda value: employee_map.get(c.format_code(value), {}).get('name', ''))
    df['合同一级类型判定'] = df.apply(
        lambda row: _contract_type_label(row['合同类型'], row['合同二级类型']),
        axis=1,
    )
    main_contract_info_map = build_main_contract_info_map(df, option_maps)
    df['主合同匹配方式'] = df.index.map(lambda index: main_contract_info_map.get(index, {}).get('source', ''))
    df['主合同编号'] = df.index.map(lambda index: main_contract_info_map.get(index, {}).get('number', ''))
    df['主合同标题'] = df.index.map(lambda index: main_contract_info_map.get(index, {}).get('title', ''))
    df['主合同流程类型ID'] = df.index.map(lambda index: main_contract_info_map.get(index, {}).get('workflow_id', ''))
    df['主合同流程名称'] = df.index.map(lambda index: main_contract_info_map.get(index, {}).get('workflow_name', ''))
    df['主合同类型ID'] = df.index.map(lambda index: main_contract_info_map.get(index, {}).get('type_id', ''))
    df['主合同类型'] = df.index.map(lambda index: main_contract_info_map.get(index, {}).get('type', ''))
    df['主合同二级类型ID'] = df.index.map(
        lambda index: main_contract_info_map.get(index, {}).get('secondary_type_id', ''))
    df['主合同二级类型'] = df.index.map(lambda index: main_contract_info_map.get(index, {}).get('secondary_type', ''))
    df['主合同一级类型判定'] = df.index.map(lambda index: main_contract_info_map.get(index, {}).get('type_label', ''))
    df['主合同分类'] = df.index.map(lambda index: main_contract_info_map.get(index, {}).get('category', ''))
    df['主合同主播卡片ID'] = df.index.map(lambda index: main_contract_info_map.get(index, {}).get('anchor_card_id', ''))
    df['主合同主播姓名'] = df.index.map(lambda index: main_contract_info_map.get(index, {}).get('anchor_name', ''))
    df['主合同主播昵称'] = df.index.map(lambda index: main_contract_info_map.get(index, {}).get('anchor_nickname', ''))
    df['主合同战队名称'] = df.index.map(lambda index: main_contract_info_map.get(index, {}).get('team_name', ''))
    df['主合同所属平台ID'] = df.index.map(lambda index: main_contract_info_map.get(index, {}).get('platform_id', ''))
    df = apply_main_contract_anchor_info(df, main_contract_info_map)
    df['合同主表所属平台'] = df['所属平台ID'].map(lambda value: option_maps['szpt'].get(c.format_code(value), ''))
    card_info_map = build_anchor_card_info_map(df['主播卡片ID'])
    identity_card_info_map = build_anchor_card_info_map_by_identity(df)

    def card_info(row):
        card_id = c.format_code(row['主播卡片ID'])
        return _merge_card_info(
            card_info_map.get(card_id, {}),
            identity_card_info_map.get(row.name, {}),
        )

    def card_field(row, field):
        return card_info(row).get(field, '')

    df['主播卡片导入ID'] = df.apply(
        lambda row: _excel_id_value(row['主播卡片ID'], card_field(row, 'card_id')),
        axis=1,
    )
    df['主播卡片编号'] = df.apply(lambda row: card_field(row, 'card_number'), axis=1)
    df['主播身份证号码'] = df.apply(lambda row: card_field(row, 'id_number'), axis=1)
    df['主播姓名_卡片'] = df.apply(lambda row: card_field(row, 'name'), axis=1)
    df['主播昵称_卡片'] = df.apply(lambda row: card_field(row, 'nickname'), axis=1)
    df['战队名称_卡片'] = df.apply(lambda row: card_field(row, 'team_name'), axis=1)
    df['官签签约金'] = df.apply(lambda row: card_field(row, 'official_signing_bonus'), axis=1)
    df['官签签约金分成比例'] = df.apply(lambda row: card_field(row, 'official_signing_bonus_share_ratio'), axis=1)
    df['固定底薪（每月）'] = df.apply(lambda row: card_field(row, 'base_salary_monthly'), axis=1)
    df['公司签约金'] = df.apply(lambda row: card_field(row, 'company_signing_bonus'), axis=1)
    platform_results = df.apply(
        lambda row: resolve_anchor_platform(row, card_info(row), option_maps['szpt']),
        axis=1,
    )
    df['主播卡片平台明细'] = df.apply(
        lambda row: '/'.join(option_maps['szpt'].get(code, code) for code in _card_platform_codes(card_info(row), option_maps['szpt'])),
        axis=1,
    )
    df['合同标题平台命中'] = df.apply(
        lambda row: _title_platform_match_text(row, option_maps['szpt']),
        axis=1,
    )
    df['所属平台ID_解析'] = platform_results.map(lambda item: item[0])
    df['所属平台'] = platform_results.map(lambda item: item[1])
    df['所属平台解析依据'] = platform_results.map(lambda item: item[2])
    df['房间号/主播ID'] = df.apply(
        lambda row: choose_anchor_room_id(
            card_info(row),
            row['所属平台ID_解析'],
        ),
        axis=1,
    )
    df['合同分类'] = df.apply(resolve_contract_category, axis=1)
    df['合同分类依据'] = df.apply(resolve_contract_category_basis, axis=1)
    df['收支类型'] = [
        resolve_pay_type(in_amount, out_amount)
        for in_amount, out_amount in zip(df['合同预计收入'], df['合同预计支出'])
    ]
    df['推导合同总额'] = df.apply(resolve_contract_amount, axis=1)
    df.attrs['company_info_map'] = company_info_map
    df.attrs['customer_info_map'] = customer_info_map
    df.attrs['supplier_info_map'] = supplier_info_map
    df.attrs['card_info_map'] = card_info_map
    df.attrs['identity_card_info_map'] = identity_card_info_map
    return df


def read_source():
    c.validate_fw_fields(FW_TABLE, EXPECTED_FW_FIELDS)
    stats = _query_fw(STATS_SQL).iloc[0]
    print('[合同迁移-主播流程] SQL过滤: 合同类型=主播协议 且 合同签署状态∈(审批完成, 已归档)')
    print(f"  合同库总数 {int(stats['all_count'] or 0)} 条; "
          f"主播协议 {int(stats['anchor_type_count'] or 0)} 条; "
          f"保留审批完成/已归档 {int(stats['kept_count'] or 0)} 条; "
          f"排除其他状态 {int(stats['excluded_status_count'] or 0)} 条; "
          f"合同金额为空但可由收入/支出回填 {int(stats['amount_fallback_count'] or 0)} 条")
    source_df = _query_fw(SOURCE_SQL)
    print('[合同迁移-主播流程] SQL主表行数:', len(source_df))
    return resolve_source_values(source_df)


# ============================ Sheet 构建 ============================
def build_main_output(source_df, headers):
    rows = []
    for source in source_df.to_dict('records'):
        row = _new_row(headers)
        contract_amount = source['推导合同总额']
        in_amount = _round_amount(source['合同预计收入'])
        out_amount = _round_amount(source['合同预计支出'])

        _set(row, 'contract_number（合同编码）', _text(source['合同编号']))
        _set(row, 'contract_name（合同名称）', _text(source['合同标题']))
        _set(row, '订单编号', _text(source['订单编号']))
        _set(row, 'contractCategory(智书框架合同类型)', source['合同分类'])
        _set(row, 'pay_type_code（收支类型）', source['收支类型'])
        _set(row, 'property_type_code（计价方式）', DEFAULT_PROPERTY_TYPE)
        _set(row, 'estimated_amount（预估金额）', contract_amount)
        _set(row, 'in_amount（预估收入金额）', in_amount)
        _set(row, 'out_amount（预估支出金额）', out_amount)
        _set(row, 'amount（合同总额）', contract_amount)
        _set(row, 'in_amount（收入总额）', in_amount)
        _set(row, 'out_amount（支出总额）', out_amount)
        _set(row, 'fixed_validity_code（合同期限类型）', DEFAULT_VALIDITY_TYPE)
        _set(row, 'start_date（合同期限-开始日期）', c.format_date(source['合同有效期起始时间']))
        _set(row, 'end_date（合同期限-结束日期）', c.format_date(source['合同有效期截止时间']))
        _set(row, 'remark（合同说明）', _text(source['合同摘要'])[:150])
        _set(row, 'custom_15_78cf503c57194e4fb8ad03ded1c4ad60（打印模式）', DEFAULT_PRINT_MODE)
        _set(row, 'custom_10_9a2a0e99771346c98bfb6cfb893e1bee（签署日期）', c.format_date(source['合同签订日期']))
        _set(row, 'custom_1001_948719050bfe402ab083c98e52fa71b2（合同执行人）', _text(source['合同执行人员']))
        _set(row, 'custom_1024_61820798c0f348658d8daa64f8b2aef9（主播卡片）',
             source['主播卡片导入ID'])
        _set(row, 'custom_1_ab6f99ee02e549469ec5b2d4a5a98452（主播姓名）',
             _first_non_blank(source['主播姓名'], source['主播姓名_卡片']))
        _set(row, 'custom_1_4fa3c71e706c489e94977935b512b0f6（主播昵称）',
             _first_non_blank(source['主播昵称'], source['主播昵称_卡片']))
        _set(row, 'custom_1_543d4d9106f34c31bf3f9397ded6ef28（房间号/主播ID）',
             _first_non_blank(source['房间号/主播ID'], source['主播卡片编号']))
        _set(row, 'custom_1_c97a63f71e1048aea384680a64aa3573（主播身份证号码）', source['主播身份证号码'])
        _set(row, 'custom_1_ba05d6fb71bc4a778b1eca63423a38bc（战队名称）',
             _first_non_blank(source['战队名称'], source['战队名称_卡片']))
        _set(row, 'custom_15_de8944334b104d52b28d9472ab0584ef（专项品类）', '')
        _set(row, 'custom_13_c9805a6fe9f245ebbfeea13407277306（是否需要验收）',
             DEFAULT_ACCEPTANCE_REQUIRED)
        _set(row, 'custom_5_d6aaf62f8568491c8f5824285e72499d（直播平台的虚拟礼物收益）', '')
        _set(row, 'custom_5_fc246ba4165145af92cffb2e6088c4b3（直播平台签约后每月基本合作费）', '')
        _set(row, 'custom_5_a8e7beeebc7b4253a9ec3c713d9d790a（自媒体平台账号商务收入）', '')
        _set(row, 'custom_5_c0a993cb06774ab38ea3e75187631ab5（自媒体平台支付的自媒体账号收入）', '')
        _set(row, 'custom_1012_dbf82175f1964048b83b18550f3bb8d1（官签签约金）', source['官签签约金'])
        _set(row, 'custom_5_f53355b8e15546a7a187575f07cf59ee（官签签约金分成比例）',
             source['官签签约金分成比例'])
        _set(row, 'custom_5_def7270057bd4913a1fd087b4b1f128e（本合同项下其他可分配收益）', '')
        _set(row, 'custom_1012_3c0986ffce8848caab9249df37e5d49e（固定底薪（每月））',
             source['固定底薪（每月）'])
        _set(row, 'custom_1012_07b5c0b4f40a414588ce86d43705f5c4（公司签约金）', source['公司签约金'])
        _set(row, 'custom_1012_65df3bda1aae46a1822c4d4531be5e25（保底费用）', '')
        _set(row, 'custom_1012_a01fe29e8faf471a92d0f002f22aea48（其他费用）', '')
        _set(row, 'custom_1012_27f3d0bf4d1d44cb82333cfe22443b65（基础服务费）', '')
        _set(row, 'sign_type_code（先盖章方）', DEFAULT_FIRST_SEAL_PARTY)
        _set(row, 'sign_type_code（签约形式）', DEFAULT_SIGN_FORM)
        _set(row, 'seal_number（盖章份数）', DEFAULT_SEAL_NUMBER)
        _set(row, 'contract_files.contract_text（合同文本）', '')
        _set(row, 'contract_files.contract_causes（合同附件）', '')
        _set(row, 'contract_files.contract_attachments（其他附件）', '')
        _set(row, 'custom_15_99b283c1e1374c02aaeacead8b336cd7（所属平台）', _text(source['所属平台']))
        _set(row, 'custom_1201_9ffd79c1e5da4492bfdab24bce0d93f8（费用明细）', '')
        _set(row, 'custom_5_de3f9a77f174445495453e444e984037（预计回本周期（月））',
             DEFAULT_PAYBACK_PERIOD_MONTHS)
        rows.append(row)
    return pd.DataFrame(rows, columns=headers)


def build_counterparty_output(source_df, headers):
    customer_info_map = source_df.attrs.get('customer_info_map', {})
    supplier_info_map = source_df.attrs.get('supplier_info_map', {})
    rows = []
    source_rows = []
    for source in source_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        anchor_counter_party_code = _text(source['主播卡片导入ID'])
        row = _new_row(headers)
        _set(row, 'contract_number（合同编码）', contract_number)
        _set(row, 'counter_party_code（对方主体编码）', anchor_counter_party_code)
        rows.append(row)
        source_rows.append({
            'contract_number（合同编码）': contract_number,
            '对方主体类型': '主播',
            '泛微对方ID': _text(source['主播卡片ID']),
            '泛微对方名称': _first_non_blank(source['主播姓名'], source['主播昵称']),
            '归并/匹配方式': _text(source['主播信息来源']),
            '对方主体名称': _first_non_blank(source['主播姓名'], source['主播昵称']),
            '对方主体编码': anchor_counter_party_code,
        })
        for customer_id in c.parse_browser_ids(source['合同客户ID']):
            info = customer_info_map.get(customer_id, {})
            row = _new_row(headers)
            _set(row, 'contract_number（合同编码）', contract_number)
            _set(row, 'counter_party_code（对方主体编码）', info.get('code', ''))
            rows.append(row)
            source_rows.append({
                'contract_number（合同编码）': contract_number,
                '对方主体类型': '客户',
                '泛微对方ID': customer_id,
                '泛微对方名称': info.get('source_name', ''),
                '归并目标ID': info.get('target_id', ''),
                '归并/匹配方式': info.get('match_method', ''),
                '对方主体名称': info.get('name', ''),
                '对方主体编码': info.get('code', ''),
            })
        for supplier_id in c.parse_browser_ids(source['合同供应商ID']):
            info = supplier_info_map.get(supplier_id, {})
            row = _new_row(headers)
            _set(row, 'contract_number（合同编码）', contract_number)
            _set(row, 'counter_party_code（对方主体编码）', info.get('code', ''))
            rows.append(row)
            source_rows.append({
                'contract_number（合同编码）': contract_number,
                '对方主体类型': '供应商',
                '泛微对方ID': supplier_id,
                '泛微对方状态': info.get('status_code', ''),
                '泛微对方名称': info.get('source_name', ''),
                '归并目标ID': info.get('target_id', ''),
                '归并目标名称': info.get('target_name', ''),
                '归并/匹配方式': info.get('match_method', ''),
                '对方主体名称': info.get('name') or info.get('source_name', ''),
                '对方主体编码': info.get('code', ''),
            })
    return pd.DataFrame(rows, columns=headers), pd.DataFrame(source_rows)


def build_our_party_output(source_df, headers):
    company_info_map = source_df.attrs.get('company_info_map', {})
    rows = []
    source_rows = []
    for source in source_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        company_ids = c.parse_browser_ids(source['合同用印范围ID'])
        if not company_ids:
            company_ids = ['']
        for company_id in company_ids:
            info = company_info_map.get(company_id, {})
            row = _new_row(headers)
            _set(row, 'contract_number（合同编码）', contract_number)
            _set(row, 'our_party_code（我方主体编码）', info.get('code', ''))
            rows.append(row)
            source_rows.append({
                'contract_number（合同编码）': contract_number,
                '泛微我方主体ID': company_id,
                '我方主体名称': info.get('name', ''),
                '我方主体编码': info.get('code', ''),
            })
    return pd.DataFrame(rows, columns=headers), pd.DataFrame(source_rows)


def build_fee_detail_output(source_df, headers):
    rows = []
    source_rows = []
    for source in source_df.to_dict('records'):
        contract_number = _text(source['合同编号'])
        row = _new_row(headers)
        _set(row, 'contract_number（合同编码）', contract_number)
        rows.append(row)
        source_rows.append({'contract_number（合同编码）': contract_number, '规则说明': '映射规则要求费用明细字段留空'})
    return pd.DataFrame(rows, columns=headers), pd.DataFrame(source_rows)


def build_status_breakdown():
    option_maps = c.build_fw_select_option_maps(FW_TABLE, ['htzt'])
    df = _query_fw(STATUS_BREAKDOWN_SQL)
    df['合同签署状态'] = df['合同签署状态ID'].map(lambda value: option_maps['htzt'].get(c.format_code(value), '空'))
    return df[['合同签署状态ID', '合同签署状态', '合同数']]


# ============================ 输出 / 异常 ============================
MAIN_ISSUE_SOURCE_FIELDS = {
    'contract_number（合同编码）': '合同编号',
    'contract_name（合同名称）': '合同标题',
    '订单编号': '合同所属项目编号',
    'contractCategory(智书框架合同类型)': '合同二级类型',
    'custom_1001_948719050bfe402ab083c98e52fa71b2（合同执行人）': '合同执行人员ID',
    'start_date（合同期限-开始日期）': '合同有效期起始时间',
    'end_date（合同期限-结束日期）': '合同有效期截止时间',
    'remark（合同说明）': '合同摘要',
    'custom_1_c97a63f71e1048aea384680a64aa3573（主播身份证号码）': '主播卡片ID',
    'custom_15_99b283c1e1374c02aaeacead8b336cd7（所属平台）': '所属平台ID',
}


def collect_missing_counterparty(counterparty_source_df):
    if counterparty_source_df.empty:
        return pd.DataFrame(columns=[
            'contract_number（合同编码）', '对方主体类型', '泛微对方ID', '泛微对方状态',
            '泛微对方名称', '归并目标ID', '归并目标名称', '归并/匹配方式', '对方主体名称',
        ])
    missing = counterparty_source_df[counterparty_source_df['对方主体编码'].astype(str).str.strip() == '']
    columns = [
        'contract_number（合同编码）', '对方主体类型', '泛微对方ID', '泛微对方状态',
        '泛微对方名称', '归并目标ID', '归并目标名称', '归并/匹配方式', '对方主体名称',
    ]
    return missing[[column for column in columns if column in missing.columns]].drop_duplicates()


def collect_missing_our_party(our_party_source_df):
    if our_party_source_df.empty:
        return pd.DataFrame(columns=['contract_number（合同编码）', '泛微我方主体ID', '我方主体名称'])
    missing = our_party_source_df[our_party_source_df['我方主体编码'].astype(str).str.strip() == '']
    return missing[['contract_number（合同编码）', '泛微我方主体ID', '我方主体名称']].drop_duplicates()


def collect_missing_anchor_card(source_df):
    blank_mask = (
        (source_df['主播卡片ID'].astype(str).str.strip() != '')
        & (source_df['主播卡片编号'].astype(str).str.strip() == '')
    )
    result = source_df.loc[blank_mask, [
        '合同编号', '合同标题', '主播卡片ID', '主播姓名', '主播昵称'
    ]].drop_duplicates()
    if not result.empty:
        result['说明'] = '合同库已有主播姓名/昵称,但主播卡片ID未在 uf_zbkp 或 uf_zbkp_bak1 中查到,无法补身份证/签约金等卡片字段'
    return result


def collect_missing_final_anchor_card(source_df):
    blank_mask = source_df['主播卡片导入ID'].astype(str).str.strip().isin(('', 'nan', 'None'))
    result = source_df.loc[blank_mask, [
        '合同编号', '合同标题', '主播信息来源', '原合同主播卡片ID', '原合同主播姓名', '原合同主播昵称',
        '主合同编号', '主合同标题', '主合同主播卡片ID', '主合同主播姓名', '主合同主播昵称',
        '主播卡片ID', '主播姓名', '主播昵称',
    ]].drop_duplicates()
    if not result.empty:
        result['说明'] = '最终未取得可导入的主播卡片ID; 子合同已优先使用主合同主播信息,仍为空时需人工核对主合同主播卡片或姓名昵称'
    return result


def run():
    headers_by_sheet = _template_headers()
    required_by_sheet, remarks_by_sheet = _read_anchor_required_rules(headers_by_sheet)

    source_df = read_source()

    main_output_df = build_main_output(source_df, headers_by_sheet[SHEET_MAIN])
    counterparty_output_df, counterparty_source_df = build_counterparty_output(
        source_df, headers_by_sheet[SHEET_COUNTERPARTY])
    our_party_output_df, our_party_source_df = build_our_party_output(
        source_df, headers_by_sheet[SHEET_OUR_PARTY])
    fee_detail_output_df, fee_detail_source_df = build_fee_detail_output(
        source_df, headers_by_sheet[SHEET_FEE_DETAIL])

    print('[合同迁移-主播流程] 字段模板行数:', len(main_output_df))
    print('[合同迁移-主播流程] 对方信息行数:', len(counterparty_output_df))
    print('[合同迁移-主播流程] 我方信息行数:', len(our_party_output_df))
    print('[合同迁移-主播流程] 费用明细行数:', len(fee_detail_output_df))

    output_file = _write_template_sheets_with_fallback(TEMPLATE_FILE, OUTPUT_FILE, {
        SHEET_MAIN: main_output_df,
        SHEET_COUNTERPARTY: counterparty_output_df,
        SHEET_OUR_PARTY: our_party_output_df,
        SHEET_FEE_DETAIL: fee_detail_output_df,
    })
    _add_flow_audit_sheet(output_file, source_df)
    _add_platform_audit_sheet(output_file, source_df)
    print('已写出:', output_file)

    exception_sheets = {
        '过滤状态分布': build_status_breakdown(),
        '默认值说明': pd.DataFrame([
            {'字段': 'property_type_code（计价方式）', '默认值': DEFAULT_PROPERTY_TYPE},
            {'字段': 'fixed_validity_code（合同期限类型）', '默认值': DEFAULT_VALIDITY_TYPE},
            {'字段': 'custom_13_c9805a6fe9f245ebbfeea13407277306（是否需要验收）', '默认值': DEFAULT_ACCEPTANCE_REQUIRED},
            {'字段': 'custom_15_78cf503c57194e4fb8ad03ded1c4ad60（打印模式）', '默认值': DEFAULT_PRINT_MODE},
            {'字段': 'sign_type_code（先盖章方）', '默认值': DEFAULT_FIRST_SEAL_PARTY},
            {'字段': 'sign_type_code（签约形式）', '默认值': DEFAULT_SIGN_FORM},
            {'字段': 'seal_number（盖章份数）', '默认值': DEFAULT_SEAL_NUMBER},
            {'字段': 'custom_5_de3f9a77f174445495453e444e984037（预计回本周期（月））', '默认值': DEFAULT_PAYBACK_PERIOD_MONTHS},
        ]),
        '字段模板_必输字段未达100%': _fill_summary(
            main_output_df,
            required_by_sheet[SHEET_MAIN],
            remarks_by_sheet[SHEET_MAIN],
        ),
        '对方信息_必输字段未达100%': _fill_summary(
            counterparty_output_df,
            required_by_sheet[SHEET_COUNTERPARTY],
            remarks_by_sheet[SHEET_COUNTERPARTY],
        ),
        '我方信息_必输字段未达100%': _fill_summary(
            our_party_output_df,
            required_by_sheet[SHEET_OUR_PARTY],
            remarks_by_sheet[SHEET_OUR_PARTY],
        ),
        '费用明细_必输字段未达100%': _fill_summary(
            fee_detail_output_df,
            required_by_sheet[SHEET_FEE_DETAIL],
            remarks_by_sheet[SHEET_FEE_DETAIL],
        ),
        '对方主体编码_未匹配': collect_missing_counterparty(counterparty_source_df),
        '我方主体编码_未匹配': collect_missing_our_party(our_party_source_df),
        '主播卡片补充信息_未匹配': collect_missing_anchor_card(source_df),
        '最终主播卡片_未匹配': collect_missing_final_anchor_card(source_df),
    }
    exception_sheets.update(c.collect_order_mapping_issues(
        source_df,
        doc_col='合同编号',
        project_col='合同所属项目编号',
        project_id_col='合同所属项目编号ID',
        project_name_col='合同所属项目名称',
    ))
    exception_sheets.update({
        f'字段模板_{name}': df
        for name, df in _collect_missing_details(
            main_output_df,
            source_df,
            required_by_sheet[SHEET_MAIN],
            MAIN_ISSUE_SOURCE_FIELDS,
            'contract_number（合同编码）',
        ).items()
    })

    exception_file = _write_exceptions_with_fallback(EXCEPTION_FILE, exception_sheets)
    if exception_file:
        print('已写出:', exception_file)


if __name__ == '__main__':
    run()
