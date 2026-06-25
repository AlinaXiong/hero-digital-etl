# -*- coding: utf-8 -*-
"""应付期初 —— 对公付款单 / 批量费用流程 / 只转入外部成本(DB 直连版)。

按《业财项目_数据映射规则.xlsx》-「应付期初」生成期初对公付款单导入模板的三个 sheet:
    1. 期初对公付款单导入
    2. 批量费用流程
    3. 只转入外部成本

跑法:在项目根执行  python run.py ap_payment_opening_extra_db
"""
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

if __package__ is None or __package__ == '':
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from etl import common as c
from etl.tasks import ap_payment_opening_db as base_payment


# ============================ 文件 / 模板 ============================
TASK_NAME = 'ap_payment_opening_extra_db'
TEMPLATE_DIR = c.TPL_DIR / 'ap_payment_opening'
OUTPUT_DIR = c.OUT_DIR / TASK_NAME
DATE_SUFFIX = c.today_suffix()

PROJECT_FILTER_FILE = Path.home() / 'Downloads' / '数据清洗涉及泛微项目编码_0624_分类.xlsx'

TEMPLATE_FILE = TEMPLATE_DIR / '英雄期初对公付款单导入模版.xlsx'
OUTPUT_FILE = OUTPUT_DIR / f'英雄期初对公付款单导入_应付期初_补充_{DATE_SUFFIX}.xlsx'
EXCEPTION_FILE = OUTPUT_DIR / f'未匹配清单_应付期初_补充_{DATE_SUFFIX}.xlsx'
BASE_SUPPLIER_VENDOR_MISSING_FILE = OUTPUT_DIR / f'Hand按ID查不到的供应商_期初对公付款单导入_{DATE_SUFFIX}.xlsx'
MCN_SUPPLIER_VENDOR_MISSING_FILE = OUTPUT_DIR / f'Hand按ID查不到的供应商_MCN对公付款_{DATE_SUFFIX}.xlsx'
BATCH_SUPPLIER_VENDOR_MISSING_FILE = OUTPUT_DIR / f'Hand按ID查不到的供应商_批量费用流程_{DATE_SUFFIX}.xlsx'

BASE_TEMPLATE_SHEET = '期初对公付款单导入'
SHEET_BATCH = '批量费用流程'
SHEET_EXTERNAL_COST = '只转入外部成本'
SHEET_MCN_OUTBOUND = 'MCN对外付款流程'
SHEET_MCN_ORDER = 'MCN对外付款流程(订单)'
SHEET_MCN_ANCHOR = 'MCN主播相关付款流程'
RULE_SHEET = '应付期初'
RULE_TABLE = '期初对公付款单'

DOCUMENT_TYPE = 'AP01-1'
DATE_FROM = '2026-01-01'
SOURCE_SYSTEM = 'FW'
VIRTUAL_VENDOR_NAME = '外部成本转移虚拟供应商'
DIRECT_PAYMENT_PLATFORM_CODE = '2'
NOT_PREPAYMENT_CODE = '1'
EVENT_PROJECT_SHEET = '赛事'
MCN_PROJECT_SHEET = 'MCN'
EVENT_PROJECT_TABLES = ('uf_xtyyxmkp', 'uf_xmkp', 'view_xmjkzb')
MCN_PROJECT_TABLES = ('uf_xmkp', 'view_xmjkzb', 'uf_xtyyxmkp')
SQL_BATCH_SIZE = 800

OUTPUT_COLUMNS = [
    '来源系统',
    '来源单据编号',
    '申请日期',
    '单据类型',
    '申请人工号',
    '申请人姓名',
    '订单编号',
    '订单名称',
    '核算主体编号',
    '核算主体描述',
    '备注',
    '合同号',
    '合同收支计划行',
    '收款方编码',
    '收款方描述',
    '银行账号',
    '计划付款日期',
    '银行转账备注',
    '实际已支付金额',
    '费用项目编码',
    '费用项目描述',
    '主播房间号',
    '报账币种',
    '报账金额（支付币种）',
    '泛微项目编号',
    '泛微费用项目编码',
]

BATCH_ISSUE_SOURCE_FIELDS = {
    '申请人工号': '申请人',
    '收款方编码': '供应商',
    '核算主体编号': '公司主体',
    '费用项目编码': '预算科目',
    '订单编号': '项目编号',
}
EXTERNAL_COST_ISSUE_SOURCE_FIELDS = {
    '申请人工号': '申请人',
    '收款方编码': '收款方描述',
    '核算主体编号': '公司主体',
    '费用项目编码': '预算科目',
    '订单编号': '项目编号',
}


# ============================ 泛微源 SQL ============================
BASE_EVENT_SOURCE_SQL = """
SELECT
    m.id AS `ID`,
    m.lcbh AS `流程编号`,
    m.sqrq AS `申请日期`,
    m.jbr AS `经办人ID`,
    m.xmbh AS `项目编号`,
    m.gszt AS `公司主体ID`,
    m.rzdw AS `成本中心ID`,
    m.bz AS `备注`,
    m.xght AS `相关合同ID`,
    m.gys AS `供应商ID`,
    m.gyswb AS `供应商-文本`,
    m.yhzh AS `银行账号`,
    m.yjfkrq AS `预计付款日期`,
    m.fkbz AS `付款币种ID`,
    m.zfzt AS `支付状态ID`,
    d.fkje AS `付款金额`,
    d.yskm AS `预算科目ID`
FROM uf_dgfktz m
JOIN uf_dgfktz_dt1 d ON d.mainid = m.id
WHERE m.lcly IN %(source_codes)s
  AND m.xmbh IN %(project_ids)s
ORDER BY m.id, d.id
"""

BASE_EVENT_STATS_SQL = """
SELECT
    COUNT(DISTINCT m.id) AS document_count,
    COUNT(*) AS detail_count,
    SUM(d.fkje) AS amount_total
FROM uf_dgfktz m
JOIN uf_dgfktz_dt1 d ON d.mainid = m.id
WHERE m.lcly IN %(source_codes)s
  AND m.xmbh IN %(project_ids)s
"""

BATCH_SOURCE_SQL = """
SELECT
    m.id AS `ID`,
    d.id AS `明细ID`,
    COALESCE(NULLIF(m.fybh, ''), LEFT(d.fymxbh, CHAR_LENGTH(d.fymxbh) - 4), d.fymxbh) AS `流程编号`,
    d.jlrq AS `申请日期`,
    d.jsr AS `申请人ID`,
    d.xmbh AS `项目编号ID`,
    m.gszt AS `公司主体ID`,
    COALESCE(NULLIF(CAST(d.cbzx AS CHAR), ''), CAST(m.cbzx AS CHAR)) AS `成本中心ID`,
    COALESCE(NULLIF(d.bz, ''), NULLIF(m.pcbz, ''), NULLIF(m.fypcbz, ''), NULLIF(m.fypcmc, '')) AS `备注`,
    m.dwfkdw AS `供应商ID`,
    d.je AS `金额`,
    COALESCE(d.yskm, d.fyxh) AS `预算科目ID`
FROM uf_plfy m
JOIN uf_plfy_dt1 d ON d.mainid = m.id
WHERE d.xmbh IN %(project_ids)s
ORDER BY d.jlrq, m.id, d.id
"""

BATCH_STATS_SQL = """
SELECT
    COUNT(*) AS kept_count,
    COUNT(DISTINCT m.id) AS document_count,
    SUM(d.je) AS amount_total
FROM uf_plfy m
JOIN uf_plfy_dt1 d ON d.mainid = m.id
WHERE d.xmbh IN %(project_ids)s
"""

EXTERNAL_COST_SOURCE_SQL = """
SELECT
    m.id AS `ID`,
    d.id AS `明细ID`,
    m.lcbh AS `流程编号`,
    m.sqrq AS `申请日期`,
    m.sqr AS `申请人ID`,
    m.zrxmbh AS `转入项目编号ID`,
    m.zcxmbh AS `转出项目编号ID`,
    m.zrxmbhmcnss AS `转入MCN赛事项目编号ID`,
    m.zcxmbhmcnss AS `转出MCN赛事项目编号ID`,
    m.jsxmmc AS `转入项目名称`,
    m.zcxmmc AS `转出项目名称`,
    m.zrgszt AS `转入公司主体ID`,
    m.zcgszt AS `转出公司主体ID`,
    m.zrcbzx AS `转入成本中心ID`,
    m.zccbzx AS `转出成本中心ID`,
    m.zrzcjehz AS `转入总金额`,
    m.zczcjehz AS `转出总金额`,
    m.fyd AS `费用单ID`,
    d.stzjid AS `费用明细视图ID`,
    d.mxid AS `费用明细ID`,
    d.zczcje AS `转出明细金额`,
    d.srje AS `转出收入金额`,
    v.lcbh AS `费用单据号`,
    COALESCE(v.yskm, d.yslx) AS `预算科目ID`,
    v.fkbz AS `付款币种ID`,
    v.fkje AS `费用原金额`,
    v.sywzje AS `费用剩余未占金额`,
    v.yzje AS `费用已占金额`,
    m.ly AS `来源类型`,
    CASE m.ly
        WHEN 5 THEN '赛事只转入外部成本'
        WHEN 2 THEN 'MCN只转入外部成本'
        ELSE CONCAT('内部收支来源', m.ly)
    END AS `内部收支来源`
FROM uf_xtyynbsz m
JOIN uf_xtyynbsz_dt10 d ON d.mainid = m.id
LEFT JOIN view_costlist_ys v
    ON v.id = COALESCE(NULLIF(d.stzjid, ''), NULLIF(d.mxid, ''))
WHERE m.ly IN (5, 2)
  AND (
      m.zrxmbh IN %(project_ids)s
      OR m.zcxmbh IN %(project_ids)s
      OR m.zrxmbhmcnss IN %(project_ids)s
      OR m.zcxmbhmcnss IN %(project_ids)s
  )
ORDER BY m.sqrq, m.id, d.id
"""

EXTERNAL_COST_STATS_SQL = """
SELECT
    COUNT(*) AS document_count,
    SUM(CASE WHEN ly = 5 THEN 1 ELSE 0 END) AS event_document_count,
    SUM(CASE WHEN ly = 2 THEN 1 ELSE 0 END) AS mcn_document_count,
    SUM(COALESCE(zrzcjehz, 0)) AS in_amount_total,
    SUM(COALESCE(zczcjehz, 0)) AS out_amount_total
FROM uf_xtyynbsz
WHERE ly IN (5, 2)
  AND (
      zrxmbh IN %(project_ids)s
      OR zcxmbh IN %(project_ids)s
      OR zrxmbhmcnss IN %(project_ids)s
      OR zcxmbhmcnss IN %(project_ids)s
  )
"""

MCN_OUTBOUND_SOURCE_SQL = """
SELECT
    'MCN对外付款流程' AS `来源流程`,
    m.id AS `ID`,
    m.requestId AS `RequestID`,
    d.id AS `明细ID`,
    m.lcbh AS `流程编号`,
    rb.REQUESTNAME AS `标题`,
    m.sqrq AS `申请日期`,
    m.sqr AS `申请人ID`,
    m.szgs AS `公司主体ID`,
    m.cbzx2 AS `成本中心ID`,
    m.dfgsmc AS `供应商ID`,
    m.yhzh AS `银行账号ID`,
    m.fkht AS `主表合同ID`,
    d.szht AS `明细合同ID`,
    d.szxm AS `项目编号ID`,
    d.xmmc AS `项目名称`,
    d.fjxh AS `费用项编码`,
    d.fyx AS `费用项名称`,
    d.je AS `金额`,
    d.fjhzbid AS `主播房间号`,
    d.zbnc AS `主播昵称`
FROM formtable_main_33 m
JOIN formtable_main_33_dt1 d ON d.mainid = m.id
LEFT JOIN workflow_requestbase rb ON rb.REQUESTID = m.requestId
WHERE m.fkpt = %(direct_payment_code)s
  AND d.szxm IN %(project_ids)s
ORDER BY m.id, d.id
"""

MCN_ORDER_SOURCE_SQL = """
SELECT * FROM (
    SELECT
        'MCN对外付款流程（订单）' AS `来源流程`,
        m.id AS `ID`,
        m.requestid AS `RequestID`,
        d.id AS `明细ID`,
        m.lcbh AS `流程编号`,
        rb.REQUESTNAME AS `标题`,
        m.sqrq AS `申请日期`,
        m.sqr AS `申请人ID`,
        m.szgs AS `公司主体ID`,
        m.cbzx AS `成本中心ID`,
        m.dfgsmc AS `供应商ID`,
        m.yhzh AS `银行账号ID`,
        m.fkht AS `主表合同ID`,
        NULL AS `明细合同ID`,
        d.szxm AS `项目编号ID`,
        d.xmmc AS `项目名称`,
        d.fjxh AS `费用项编码`,
        d.fyx AS `费用项名称`,
        COALESCE(d.ddje, d.zcje, d.sdzc, d.jsje) AS `金额`,
        d.zbid AS `主播房间号`,
        d.zbnc AS `主播昵称`
    FROM formtable_main_66 m
    JOIN formtable_main_66_dt3 d ON d.mainid = m.id
    LEFT JOIN workflow_requestbase rb ON rb.REQUESTID = m.requestid
    WHERE m.fkpt = %(direct_payment_code)s
    UNION ALL
    SELECT
        'MCN对外付款流程（订单）' AS `来源流程`,
        m.id AS `ID`,
        m.requestid AS `RequestID`,
        d.id AS `明细ID`,
        m.lcbh AS `流程编号`,
        rb.REQUESTNAME AS `标题`,
        m.sqrq AS `申请日期`,
        m.sqr AS `申请人ID`,
        m.szgs AS `公司主体ID`,
        m.cbzx AS `成本中心ID`,
        m.dfgsmc AS `供应商ID`,
        m.yhzh AS `银行账号ID`,
        m.fkht AS `主表合同ID`,
        NULL AS `明细合同ID`,
        d.szxm AS `项目编号ID`,
        d.xmmc AS `项目名称`,
        d.fjxh AS `费用项编码`,
        d.fyx AS `费用项名称`,
        COALESCE(d.ddje, d.zcje, d.sdzc, d.jsje) AS `金额`,
        d.zbid AS `主播房间号`,
        d.zbnc AS `主播昵称`
    FROM formtable_main_66 m
    JOIN formtable_main_66_dt4 d ON d.mainid = m.id
    LEFT JOIN workflow_requestbase rb ON rb.REQUESTID = m.requestid
    WHERE m.fkpt = %(direct_payment_code)s
    UNION ALL
    SELECT
        'MCN对外付款流程（订单）' AS `来源流程`,
        m.id AS `ID`,
        m.requestid AS `RequestID`,
        d.id AS `明细ID`,
        m.lcbh AS `流程编号`,
        rb.REQUESTNAME AS `标题`,
        m.sqrq AS `申请日期`,
        m.sqr AS `申请人ID`,
        m.szgs AS `公司主体ID`,
        m.cbzx AS `成本中心ID`,
        m.dfgsmc AS `供应商ID`,
        m.yhzh AS `银行账号ID`,
        m.fkht AS `主表合同ID`,
        NULL AS `明细合同ID`,
        d.szxm AS `项目编号ID`,
        d.xmmc AS `项目名称`,
        d.fjxh AS `费用项编码`,
        d.fyx AS `费用项名称`,
        COALESCE(d.ddje, d.zcje, d.sdzc, d.jsje) AS `金额`,
        d.zbid AS `主播房间号`,
        d.zbmc AS `主播昵称`
    FROM formtable_main_66 m
    JOIN formtable_main_66_dt5 d ON d.mainid = m.id
    LEFT JOIN workflow_requestbase rb ON rb.REQUESTID = m.requestid
    WHERE m.fkpt = %(direct_payment_code)s
    UNION ALL
    SELECT
        'MCN对外付款流程（订单）' AS `来源流程`,
        m.id AS `ID`,
        m.requestid AS `RequestID`,
        d.id AS `明细ID`,
        m.lcbh AS `流程编号`,
        rb.REQUESTNAME AS `标题`,
        m.sqrq AS `申请日期`,
        m.sqr AS `申请人ID`,
        m.szgs AS `公司主体ID`,
        m.cbzx AS `成本中心ID`,
        m.dfgsmc AS `供应商ID`,
        m.yhzh AS `银行账号ID`,
        m.fkht AS `主表合同ID`,
        NULL AS `明细合同ID`,
        d.szxm AS `项目编号ID`,
        d.xmmc AS `项目名称`,
        d.fjxh AS `费用项编码`,
        d.fyx AS `费用项名称`,
        COALESCE(d.fkje, d.zcje, d.jsje) AS `金额`,
        NULL AS `主播房间号`,
        NULL AS `主播昵称`
    FROM formtable_main_66 m
    JOIN formtable_main_66_dt6 d ON d.mainid = m.id
    LEFT JOIN workflow_requestbase rb ON rb.REQUESTID = m.requestid
    WHERE m.fkpt = %(direct_payment_code)s
) x
WHERE x.`项目编号ID` IN %(project_ids)s
ORDER BY x.`ID`, x.`明细ID`
"""

MCN_ANCHOR_SOURCE_SQL = """
SELECT
    'MCN主播相关付款流程' AS `来源流程`,
    m.id AS `ID`,
    m.requestId AS `RequestID`,
    d.id AS `明细ID`,
    m.lcbh AS `流程编号`,
    rb.REQUESTNAME AS `标题`,
    m.sqrq AS `申请日期`,
    m.sqr AS `申请人ID`,
    m.szgs AS `公司主体ID`,
    COALESCE(d.cbzx, m.cbzx1, m.cbzxjs) AS `成本中心ID`,
    m.dfgsmc AS `供应商ID`,
    COALESCE(m.gysyhzh, m.yhzh) AS `银行账号ID`,
    m.szht AS `主表合同ID`,
    d.xght AS `明细合同ID`,
    d.xmbh AS `项目编号ID`,
    d.xmmc AS `项目名称`,
    d.fjxh AS `费用项编码`,
    d.fyx AS `费用项名称`,
    COALESCE(d.dkje, d.yfuje, d.skje, d.yfaje) AS `金额`,
    d.zbid AS `主播房间号`,
    d.zbnc AS `主播昵称`
FROM formtable_main_38 m
JOIN formtable_main_38_dt4 d ON d.mainid = m.id
LEFT JOIN workflow_requestbase rb ON rb.REQUESTID = m.requestId
WHERE m.sfyfk = %(not_prepayment_code)s
  AND d.xmbh IN %(project_ids)s
ORDER BY m.id, d.id
"""


# ============================ 小工具 ============================
def _query_fw(sql, params=None):
    return c.query_db('FW', 'vspn_xtyy', sql, params or {})


def _text(value):
    if pd.isna(value):
        return ''
    text = str(value).strip()
    return '' if text in ('', 'nan', 'None', 'NaT') else text


def _first_browser_value(mapping, value):
    for item_id in c.parse_browser_ids(value):
        mapped = mapping.get(item_id, '')
        if mapped:
            return mapped
    return ''


def _lookup_by_name(mapping, value):
    text = _text(value)
    return mapping.get(c.normalize_name(text), '') if text else ''


def _first_non_blank(*values):
    for value in values:
        text = _text(value)
        if text:
            return text
    return ''


def _chunks(values, size=SQL_BATCH_SIZE):
    values = list(values)
    for start in range(0, len(values), size):
        yield values[start:start + size]


_PROJECT_FILTER_CACHE = None
_PROJECT_FILTER_ID_CACHE = {}


def load_project_filter_codes():
    """读取本次清洗白名单:赛事/MCN sheet 的「原泛微项目编码」。"""
    global _PROJECT_FILTER_CACHE
    if _PROJECT_FILTER_CACHE is not None:
        return _PROJECT_FILTER_CACHE
    if not PROJECT_FILTER_FILE.exists():
        raise FileNotFoundError(f'项目白名单不存在: {PROJECT_FILTER_FILE}')

    result = {}
    for sheet_name in (EVENT_PROJECT_SHEET, MCN_PROJECT_SHEET):
        df = pd.read_excel(PROJECT_FILTER_FILE, sheet_name=sheet_name, dtype=str)
        codes = []
        for value in df.iloc[:, 0]:
            codes.extend(c.split_fanwei_project_codes(value))
        result[sheet_name] = set(codes)
    print(
        '[应付期初-项目白名单] 使用:',
        PROJECT_FILTER_FILE,
        f"| 赛事 {len(result[EVENT_PROJECT_SHEET])} 个",
        f"| MCN {len(result[MCN_PROJECT_SHEET])} 个",
    )
    _PROJECT_FILTER_CACHE = result
    return result


def project_filter_codes(sheet_name):
    return load_project_filter_codes().get(sheet_name, set())


def project_filter_ids(sheet_name, table_order):
    cache_key = (sheet_name, tuple(table_order))
    if cache_key in _PROJECT_FILTER_ID_CACHE:
        return _PROJECT_FILTER_ID_CACHE[cache_key]

    codes = sorted(project_filter_codes(sheet_name))
    result = []
    seen = set()
    for table in table_order:
        remaining_codes = [code for code in codes if code]
        for batch in _chunks(remaining_codes):
            try:
                project_df = c.query_db(
                    'FW',
                    'vspn_xtyy',
                    f'SELECT id, xmbh FROM {table} '
                    f'WHERE xmbh IN ({c.in_placeholders(batch)})',
                    batch,
                )
            except Exception:
                continue
            for _, row in project_df.iterrows():
                project_id = c.format_code(row['id'])
                if project_id and project_id not in seen:
                    seen.add(project_id)
                    result.append(project_id)
    _PROJECT_FILTER_ID_CACHE[cache_key] = tuple(result)
    print(f'[应付期初-项目白名单] {sheet_name} 反查项目ID {len(result)} 个')
    return _PROJECT_FILTER_ID_CACHE[cache_key]


def _empty_source_df(columns):
    return pd.DataFrame(columns=columns)


def build_fw_project_info_map_for_ids(project_values, table_order=EVENT_PROJECT_TABLES):
    """泛微项目浏览框 ID -> 项目编号/名称,兼容 xmkp/xmjk/MCN赛事项目等来源。"""
    project_ids = c.clean_codes(
        project_id
        for value in project_values
        for project_id in c.parse_browser_ids(value)
    )
    if not project_ids:
        return {}

    result = {}
    remaining = set(project_ids)
    for table in table_order:
        if not remaining:
            break
        ids = [project_id for project_id in project_ids if project_id in remaining]
        try:
            project_df = c.query_db(
                'FW',
                'vspn_xtyy',
                f'SELECT id, xmbh AS project_code, xmmc AS project_name FROM {table} '
                f'WHERE id IN ({c.in_placeholders(ids)})',
                ids,
            )
        except Exception:
            try:
                project_df = c.query_db(
                    'FW',
                    'vspn_xtyy',
                    f"SELECT id, xmbh AS project_code, '' AS project_name FROM {table} "
                    f'WHERE id IN ({c.in_placeholders(ids)})',
                    ids,
                )
            except Exception:
                continue
        for _, row in project_df.iterrows():
            project_id = c.format_code(row['id'])
            project_code = _text(row['project_code'])
            if project_id and project_code and project_id not in result:
                result[project_id] = {
                    'code': project_code,
                    'name': _text(row.get('project_name', '')),
                }
                remaining.discard(project_id)
    return result


def build_fw_project_code_map_for_ids(project_values, table_order=EVENT_PROJECT_TABLES):
    return {
        project_id: info.get('code', '')
        for project_id, info in build_fw_project_info_map_for_ids(project_values, table_order).items()
    }


def _resolve_project_codes(values, project_map):
    return values.map(lambda value: _first_browser_value(project_map, value) or _text(value))


def _resolve_project_names(values, project_name_map):
    return values.map(lambda value: _first_browser_value(project_name_map, value))


def _with_resolved_project_fields(source_df, project_column='项目编号', table_order=EVENT_PROJECT_TABLES):
    df = source_df.copy()
    if project_column not in df.columns:
        df['项目编号'] = ''
        df['项目名称'] = df['项目名称'] if '项目名称' in df.columns else ''
        return df

    project_info_map = build_fw_project_info_map_for_ids(df[project_column], table_order)
    project_code_map = {
        project_id: info.get('code', '')
        for project_id, info in project_info_map.items()
    }
    project_name_map = {
        project_id: info.get('name', '')
        for project_id, info in project_info_map.items()
    }

    df['项目编号'] = _resolve_project_codes(df[project_column], project_code_map)
    mapped_project_names = _resolve_project_names(df[project_column], project_name_map)
    existing_project_names = df['项目名称'] if '项目名称' in df.columns else pd.Series('', index=df.index)
    df['项目名称'] = [
        _first_non_blank(existing_name, mapped_name)
        for existing_name, mapped_name in zip(existing_project_names, mapped_project_names)
    ]
    return df


def _filter_by_project_whitelist(source_df, project_column, sheet_name, table_order, log_label):
    if source_df.empty:
        return source_df
    allowed_codes = project_filter_codes(sheet_name)
    project_source_df = _with_resolved_project_fields(source_df, project_column, table_order)
    mask = project_source_df['项目编号'].map(_text).isin(allowed_codes)
    filtered = source_df.loc[mask].copy()
    print(f'[{log_label}] 项目白名单过滤: {len(filtered)}/{len(source_df)} 行')
    return filtered


def _apply_order_project_columns(output_df, source_df, table_order=EVENT_PROJECT_TABLES):
    """按预付期初口径补充泛微项目编号,并用项目&订单清洗表映射订单字段。"""
    df = output_df.copy()
    project_source_df = _with_resolved_project_fields(source_df, table_order=table_order)
    project_codes = project_source_df['项目编号'].map(_text)
    df['泛微项目编号'] = project_codes
    df['订单编号'] = project_codes.map(lambda value: c.project_order_mapping_value(value, '订单编号'))
    df['订单名称'] = project_codes.map(lambda value: c.project_order_mapping_value(value, '订单标题'))
    return df[OUTPUT_COLUMNS]


def _enrich_missing_order_issue(sheets, output_df, source_df):
    """让「缺失_订单编号」清单同时带出项目编号/名称,便于回查原单。"""
    sheet_name = '缺失_订单编号'
    if sheet_name not in sheets or '订单编号' not in output_df.columns:
        return sheets

    project_source_df = _with_resolved_project_fields(source_df)
    blank_mask = output_df['订单编号'].astype(str).str.strip() == ''
    data = {
        '来源单据编号': output_df.loc[blank_mask, '来源单据编号'].astype(str),
        '订单编号': output_df.loc[blank_mask, '订单编号'].astype(str),
        '泛微项目编号': project_source_df.loc[blank_mask, '项目编号'].astype(str),
        '泛微项目名称': project_source_df.loc[blank_mask, '项目名称'].astype(str),
    }
    sheets[sheet_name] = pd.DataFrame(data).drop_duplicates().reset_index(drop=True)
    return sheets


def _resolve_subject_paths(subject_ids):
    subject_map = c.build_fw_budget_subject_path_map_for_ids(subject_ids)
    return subject_ids.map(lambda value: subject_map.get(c.format_code(value), ''))


def _subject_item(subject_lookup, value, index):
    subject_path = _text(value)
    return subject_lookup.get(c.remove_slashes(subject_path), ('', ''))[index] if subject_path else ''


def _supplier_name(value, supplier_status_map):
    selected_id = c.choose_fw_supplier_id(c.parse_browser_ids(value), supplier_status_map)
    return supplier_status_map.get(selected_id, {}).get('name', '')


def _mcn_contract_code_map(contract_values):
    contract_ids = c.clean_codes(
        contract_id
        for value in contract_values
        for contract_id in c.parse_browser_ids(value)
    )
    if not contract_ids:
        return {}

    result = {}
    for table, code_col in (('uf_htk', 'htbh'), ('uf_htsp', 'htbh')):
        remaining = [contract_id for contract_id in contract_ids if contract_id not in result]
        if not remaining:
            break
        for batch in _chunks(remaining):
            try:
                contract_df = c.query_db(
                    'FW',
                    'vspn_xtyy',
                    f'SELECT id, {code_col} AS contract_code FROM {table} '
                    f'WHERE id IN ({c.in_placeholders(batch)})',
                    batch,
                )
            except Exception:
                continue
            for _, row in contract_df.iterrows():
                contract_id = c.format_code(row['id'])
                contract_code = _text(row['contract_code'])
                if contract_id and contract_code and contract_id not in result:
                    result[contract_id] = contract_code
    return result


def _mcn_fee_subject_key(raw_code, subject_lookup):
    code = _text(raw_code)
    if not code:
        return ''
    deepest = code.split('_', 1)[-1].strip()
    if not deepest:
        return c.remove_slashes(code)
    candidates = [c.remove_slashes(code), deepest]
    if len(deepest) >= 4:
        candidates.append(f'{deepest[:2]}{deepest[:3]}{deepest}')
    if len(deepest) >= 3:
        candidates.append(f'{deepest[:2]}{deepest}')
    return next((key for key in candidates if key in subject_lookup), c.remove_slashes(code))


def _mcn_fee_subject_item(subject_lookup, raw_code, index):
    key = _mcn_fee_subject_key(raw_code, subject_lookup)
    if not key:
        return ''
    return subject_lookup.get(key, ('', ''))[index]


def _resolve_company_name(company_map, value):
    return _first_browser_value(company_map, value) or company_map.get(c.format_code(value), '')


def _format_signed_amount(value, sign=1):
    if pd.isna(value):
        return ''
    return c.round_amount(float(value) * sign)


def _copy_template_sheet(workbook, target_sheet_name):
    if target_sheet_name in workbook.sheetnames:
        del workbook[target_sheet_name]
    source = workbook[BASE_TEMPLATE_SHEET]
    copied = workbook.copy_worksheet(source)
    copied.title = target_sheet_name
    return copied


def _fill_sheet(worksheet, output_df):
    for col_idx, column_name in enumerate(output_df.columns, start=1):
        worksheet.cell(row=1, column=col_idx).value = column_name
    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row)
    for _, row in output_df.iterrows():
        worksheet.append(['' if pd.isna(value) else value for value in row.tolist()])


def _move_sheets_to_front(workbook, sheet_names):
    sheet_name_set = set(sheet_names)
    ordered_sheets = [workbook[sheet_name] for sheet_name in sheet_names if sheet_name in workbook.sheetnames]
    remaining_sheets = [sheet for sheet in workbook.worksheets if sheet.title not in sheet_name_set]
    workbook._sheets = ordered_sheets + remaining_sheets


def write_output_workbook(base_output_df, batch_output_df, external_output_df, mcn_output_dfs=None):
    wb = load_workbook(TEMPLATE_FILE)
    _fill_sheet(wb[BASE_TEMPLATE_SHEET], base_output_df)
    _fill_sheet(_copy_template_sheet(wb, SHEET_BATCH), batch_output_df)
    _fill_sheet(_copy_template_sheet(wb, SHEET_EXTERNAL_COST), external_output_df)
    for sheet_name, output_df in (mcn_output_dfs or {}).items():
        _fill_sheet(_copy_template_sheet(wb, sheet_name), output_df)
    _move_sheets_to_front(
        wb,
        [BASE_TEMPLATE_SHEET, SHEET_BATCH, SHEET_EXTERNAL_COST, *(mcn_output_dfs or {}).keys()],
    )
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE


def _allocate_group_amounts(group, total_col):
    total = pd.to_numeric(group[total_col], errors='coerce').dropna()
    direct = pd.to_numeric(group.get('转出明细金额'), errors='coerce')
    if direct.notna().any():
        return direct.fillna(0)

    base = pd.to_numeric(group['费用已占金额'], errors='coerce')
    if base.isna().all() or float(base.abs().sum()) == 0:
        base = pd.to_numeric(group['费用原金额'], errors='coerce')
    if base.isna().all() or float(base.abs().sum()) == 0:
        base = pd.Series([1] * len(group), index=group.index, dtype='float64')
    base = base.fillna(0).abs()
    base_sum = float(base.sum())

    if len(total) > 0 and base_sum:
        return base / base_sum * float(total.iloc[0])
    return base


def allocate_external_cost_amounts(source_df):
    df = source_df.copy()
    if df.empty:
        df['转入金额'] = []
        df['转出金额'] = []
        return df

    in_amounts = []
    out_amounts = []
    for _, group in df.groupby('ID', sort=False):
        allocated_in = _allocate_group_amounts(group, '转入总金额')
        allocated_out = _allocate_group_amounts(group, '转出总金额')
        in_amounts.append(allocated_in)
        out_amounts.append(allocated_out)
    df['转入金额'] = pd.concat(in_amounts).reindex(df.index)
    df['转出金额'] = pd.concat(out_amounts).reindex(df.index)
    return df


# ============================ 批量费用流程 ============================
def read_base_source():
    """读取赛事对公付款,仅按赛事项目白名单过滤。"""
    base_payment.SUPPLIER_VENDOR_MISSING_FILE = BASE_SUPPLIER_VENDOR_MISSING_FILE
    event_project_ids = project_filter_ids(EVENT_PROJECT_SHEET, EVENT_PROJECT_TABLES)
    if not event_project_ids:
        return pd.DataFrame()

    params = {
        'source_codes': base_payment.SOURCE_CODES,
        'project_ids': event_project_ids,
    }
    stats = _query_fw(BASE_EVENT_STATS_SQL, params).iloc[0]
    print('[应付期初-供应商付款-DB] SQL过滤: 仅保留赛事项目白名单; 不再过滤申请日期/流程状态/作废状态')
    print(f"  命中主表 {int(stats['document_count'] or 0)} 单 / 明细 {int(stats['detail_count'] or 0)} 行; "
          f"金额合计 {float(stats['amount_total'] or 0):.2f}")
    source_df = base_payment.resolve_source_values(_query_fw(BASE_EVENT_SOURCE_SQL, params))
    source_df = _filter_by_project_whitelist(
        source_df, '项目编号', EVENT_PROJECT_SHEET, EVENT_PROJECT_TABLES,
        '应付期初-供应商付款-DB',
    )
    print('[应付期初-供应商付款-DB] SQL主子合并明细行数:', len(source_df))
    return source_df


def read_batch_source():
    event_project_ids = project_filter_ids(EVENT_PROJECT_SHEET, EVENT_PROJECT_TABLES)
    if not event_project_ids:
        return pd.DataFrame()
    params = {'project_ids': event_project_ids}
    stats = _query_fw(BATCH_STATS_SQL, params).iloc[0]
    print('[应付期初-批量费用流程] SQL过滤: 仅保留赛事项目白名单; 不再过滤确认/作废/记录日期')
    print(f"  保留批量费用 {int(stats['document_count'] or 0)} 单 / 明细 {int(stats['kept_count'] or 0)} 行; "
          f"金额合计 {float(stats['amount_total'] or 0):.2f}")
    source_df = _query_fw(BATCH_SOURCE_SQL, params)
    source_df = _filter_by_project_whitelist(
        source_df, '项目编号ID', EVENT_PROJECT_SHEET, EVENT_PROJECT_TABLES,
        '应付期初-批量费用流程',
    )
    print('[应付期初-批量费用流程] SQL明细行数:', len(source_df))
    return resolve_batch_values(source_df)


def resolve_batch_values(source_df):
    df = source_df.copy()
    employee_map = c.build_fw_employee_info_map_for_ids(df['申请人ID'])
    company_map = c.build_fw_company_name_map_for_ids(df['公司主体ID'])
    cost_center_map = c.build_fw_cost_center_map_for_ids(df['成本中心ID'])
    supplier_status_map = c.build_fw_supplier_status_map(df['供应商ID'])

    df['申请人'] = df['申请人ID'].map(lambda value: employee_map.get(c.format_code(value), {}).get('name', ''))
    df['申请人工号'] = df['申请人ID'].map(lambda value: employee_map.get(c.format_code(value), {}).get('code', ''))
    df['公司主体'] = df['公司主体ID'].map(lambda value: company_map.get(c.format_code(value), ''))
    df['成本中心'] = df['成本中心ID'].map(lambda value: _first_browser_value(cost_center_map, value))
    df['供应商'] = df['供应商ID'].map(lambda value: _supplier_name(value, supplier_status_map))
    df = _with_resolved_project_fields(df, '项目编号ID')
    df['预算科目'] = _resolve_subject_paths(df['预算科目ID'])
    return df


def build_batch_output(source_df):
    vendor_info_map = c.build_supplier_vendor_info_map_for_rows(
        source_df['供应商ID'],
        supplier_texts=source_df['供应商'],
        document_numbers=source_df['流程编号'],
        missing_report_file=BATCH_SUPPLIER_VENDOR_MISSING_FILE,
        log_prefix='[应付期初-批量费用流程]',
    )
    entity_map = c.build_accounting_entity_map_for_names(source_df['公司主体'])
    subject_lookup = c.build_subject_map()
    amount = pd.to_numeric(source_df['金额'], errors='coerce')

    def vendor_field(index, field):
        return vendor_info_map.get(index, {}).get(field, '')

    output_df = pd.DataFrame(index=source_df.index)
    output_df['来源系统'] = SOURCE_SYSTEM
    output_df['来源单据编号'] = source_df['流程编号']
    output_df['申请日期'] = source_df['申请日期'].map(c.format_date)
    output_df['单据类型'] = DOCUMENT_TYPE
    output_df['申请人工号'] = source_df['申请人工号']
    output_df['申请人姓名'] = source_df['申请人']
    output_df['核算主体编号'] = source_df['公司主体'].map(lambda value: _lookup_by_name(entity_map, value))
    output_df['核算主体描述'] = source_df['公司主体']
    output_df['备注'] = source_df['备注'].map(lambda value: _text(value)[:150])
    output_df['合同号'] = ''
    output_df['合同收支计划行'] = ''
    output_df['收款方编码'] = [vendor_field(index, 'code') for index in source_df.index]
    output_df['收款方描述'] = [
        vendor_field(index, 'name') or supplier_name
        for index, supplier_name in zip(source_df.index, source_df['供应商'])
    ]
    output_df['银行账号'] = ''
    output_df['计划付款日期'] = ''
    output_df['银行转账备注'] = ''
    output_df['实际已支付金额'] = amount.map(c.round_amount)
    output_df['费用项目编码'] = source_df['预算科目'].map(lambda value: _subject_item(subject_lookup, value, 0))
    output_df['费用项目描述'] = source_df['预算科目'].map(lambda value: _subject_item(subject_lookup, value, 1))
    output_df['主播房间号'] = ''
    output_df['报账币种'] = 'CNY'
    output_df['报账金额（支付币种）'] = amount.map(c.round_amount)
    output_df['泛微费用项目编码'] = source_df['预算科目'].where(source_df['预算科目'].notna(), '')
    output_df['银行账号'] = c.resolve_hand_vendor_bank_accounts(output_df['收款方编码'])
    return _apply_order_project_columns(output_df, source_df)


# ============================ MCN 对公付款流程 ============================
def read_mcn_payment_source():
    mcn_project_ids = project_filter_ids(MCN_PROJECT_SHEET, MCN_PROJECT_TABLES)
    if not mcn_project_ids:
        return pd.DataFrame()
    params = {
        'project_ids': mcn_project_ids,
        'direct_payment_code': DIRECT_PAYMENT_PLATFORM_CODE,
        'not_prepayment_code': NOT_PREPAYMENT_CODE,
    }
    source_parts = [
        _query_fw(MCN_OUTBOUND_SOURCE_SQL, params),
        _query_fw(MCN_ORDER_SOURCE_SQL, params),
        _query_fw(MCN_ANCHOR_SOURCE_SQL, params),
    ]
    source_df = pd.concat(source_parts, ignore_index=True)
    source_df = _filter_by_project_whitelist(
        source_df, '项目编号ID', MCN_PROJECT_SHEET, MCN_PROJECT_TABLES,
        '应付期初-MCN对公付款',
    )
    print('[应付期初-MCN对公付款] 三类流程明细行数:', len(source_df))
    return resolve_mcn_payment_values(source_df)


def resolve_mcn_payment_values(source_df):
    df = source_df.copy()
    if df.empty:
        return df

    employee_map = c.build_fw_employee_info_map_for_ids(df['申请人ID'])
    company_map = c.build_fw_company_name_map_for_ids(df['公司主体ID'])
    cost_center_map = c.build_fw_cost_center_map_for_ids(df['成本中心ID'])
    supplier_status_map = c.build_fw_supplier_status_map(df['供应商ID'])
    bank_account_map = c.build_fw_supplier_bank_account_map_for_ids(df['银行账号ID'])
    contract_values = pd.concat([df['主表合同ID'], df['明细合同ID']], ignore_index=True)
    contract_map = _mcn_contract_code_map(contract_values)

    project_df = _with_resolved_project_fields(df, '项目编号ID', MCN_PROJECT_TABLES)
    df['项目编号'] = project_df['项目编号']
    df['项目名称'] = [
        _first_non_blank(existing_name, mapped_name)
        for existing_name, mapped_name in zip(df['项目名称'], project_df['项目名称'])
    ]
    df['申请人'] = df['申请人ID'].map(lambda value: employee_map.get(c.format_code(value), {}).get('name', ''))
    df['申请人工号'] = df['申请人ID'].map(lambda value: employee_map.get(c.format_code(value), {}).get('code', ''))
    df['公司主体'] = df['公司主体ID'].map(lambda value: _resolve_company_name(company_map, value))
    df['成本中心'] = df['成本中心ID'].map(lambda value: _first_browser_value(cost_center_map, value))
    df['供应商'] = df['供应商ID'].map(lambda value: _supplier_name(value, supplier_status_map))
    df['银行账号'] = df['银行账号ID'].map(lambda value: _first_browser_value(bank_account_map, value))
    df['合同号'] = [
        _first_non_blank(
            _first_browser_value(contract_map, detail_contract),
            _first_browser_value(contract_map, main_contract),
        )
        for main_contract, detail_contract in zip(df['主表合同ID'], df['明细合同ID'])
    ]
    df['预算科目'] = df['费用项编码'].map(_text)
    return df


def build_mcn_payment_output(source_df):
    if source_df.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS), source_df

    vendor_info_map = c.build_supplier_vendor_info_map_for_rows(
        source_df['供应商ID'],
        supplier_texts=source_df['供应商'],
        document_numbers=source_df['流程编号'],
        missing_report_file=MCN_SUPPLIER_VENDOR_MISSING_FILE,
        log_prefix='[应付期初-MCN对公付款]',
    )
    entity_map = c.build_accounting_entity_map_for_names(source_df['公司主体'])
    subject_lookup = c.build_subject_map()
    amount = pd.to_numeric(source_df['金额'], errors='coerce')

    def vendor_field(index, field):
        return vendor_info_map.get(index, {}).get(field, '')

    output_df = pd.DataFrame(index=source_df.index)
    output_df['来源系统'] = SOURCE_SYSTEM
    output_df['来源单据编号'] = source_df['流程编号']
    output_df['申请日期'] = source_df['申请日期'].map(c.format_date)
    output_df['单据类型'] = DOCUMENT_TYPE
    output_df['申请人工号'] = source_df['申请人工号']
    output_df['申请人姓名'] = source_df['申请人']
    output_df['核算主体编号'] = source_df['公司主体'].map(lambda value: _lookup_by_name(entity_map, value))
    output_df['核算主体描述'] = source_df['公司主体']
    output_df['备注'] = source_df['标题'].map(lambda value: _text(value)[:150])
    output_df['合同号'] = source_df['合同号']
    output_df['合同收支计划行'] = ''
    output_df['收款方编码'] = [vendor_field(index, 'code') for index in source_df.index]
    output_df['收款方描述'] = [
        vendor_field(index, 'name') or supplier_name
        for index, supplier_name in zip(source_df.index, source_df['供应商'])
    ]
    output_df['银行账号'] = c.resolve_hand_vendor_bank_accounts(
        output_df['收款方编码'], source_df['银行账号'])
    output_df['计划付款日期'] = source_df['申请日期'].map(c.format_date)
    output_df['银行转账备注'] = ''
    output_df['实际已支付金额'] = amount.map(c.round_amount)
    output_df['费用项目编码'] = source_df['费用项编码'].map(
        lambda value: _mcn_fee_subject_item(subject_lookup, value, 0))
    output_df['费用项目描述'] = source_df['费用项编码'].map(
        lambda value: _mcn_fee_subject_item(subject_lookup, value, 1))
    output_df['主播房间号'] = source_df['主播房间号'].map(_text)
    output_df['报账币种'] = 'CNY'
    output_df['报账金额（支付币种）'] = amount.map(c.round_amount)
    output_df['泛微费用项目编码'] = source_df['费用项编码'].map(_text)
    return _apply_order_project_columns(output_df, source_df, MCN_PROJECT_TABLES), source_df


# ============================ 只转入外部成本 ============================
def read_external_cost_source():
    project_ids = tuple(dict.fromkeys(
        list(project_filter_ids(EVENT_PROJECT_SHEET, EVENT_PROJECT_TABLES))
        + list(project_filter_ids(MCN_PROJECT_SHEET, MCN_PROJECT_TABLES))
    ))
    if not project_ids:
        return pd.DataFrame()
    params = {'project_ids': project_ids}
    stats = _query_fw(EXTERNAL_COST_STATS_SQL, params).iloc[0]
    print('[应付期初-只转入外部成本] SQL过滤: 仅保留赛事/MCN项目白名单; 不再过滤流程状态/作废/申请日期')
    print(f"  保留内部收支 {int(stats['document_count'] or 0)} 单; "
          f"赛事 {int(stats['event_document_count'] or 0)} 单; "
          f"MCN {int(stats['mcn_document_count'] or 0)} 单; "
          f"转入金额合计 {float(stats['in_amount_total'] or 0):.2f}; "
          f"转出金额合计 {float(stats['out_amount_total'] or 0):.2f}")
    source_df = _query_fw(EXTERNAL_COST_SOURCE_SQL, params)
    print('[应付期初-只转入外部成本] SQL费用明细行数:', len(source_df))
    return resolve_external_cost_values(allocate_external_cost_amounts(source_df))


def resolve_external_cost_values(source_df):
    df = source_df.copy()
    employee_map = c.build_fw_employee_info_map_for_ids(df['申请人ID'])
    company_ids = pd.concat([df['转入公司主体ID'], df['转出公司主体ID']], ignore_index=True)
    company_map = c.build_fw_company_name_map_for_ids(company_ids)
    cost_center_ids = pd.concat([df['转入成本中心ID'], df['转出成本中心ID']], ignore_index=True)
    cost_center_map = c.build_fw_cost_center_map_for_ids(cost_center_ids)
    currency_map = c.build_fw_currency_name_map_for_ids(df['付款币种ID'])
    project_values = pd.concat([
        df['转入项目编号ID'], df['转出项目编号ID'],
        df['转入MCN赛事项目编号ID'], df['转出MCN赛事项目编号ID'],
    ], ignore_index=True)
    event_project_info_map = build_fw_project_info_map_for_ids(project_values, EVENT_PROJECT_TABLES)
    event_project_code_map = {
        project_id: info.get('code', '')
        for project_id, info in event_project_info_map.items()
    }
    event_project_name_map = {
        project_id: info.get('name', '')
        for project_id, info in event_project_info_map.items()
    }
    mcn_project_info_map = build_fw_project_info_map_for_ids(project_values, MCN_PROJECT_TABLES)
    mcn_project_code_map = {
        project_id: info.get('code', '')
        for project_id, info in mcn_project_info_map.items()
    }
    mcn_project_name_map = {
        project_id: info.get('name', '')
        for project_id, info in mcn_project_info_map.items()
    }

    def project_maps(source_type):
        return (
            (mcn_project_code_map, mcn_project_name_map)
            if c.format_code(source_type) == '2'
            else (event_project_code_map, event_project_name_map)
        )

    def project_code(source_type, project_value, mcn_project_value):
        code_map, _ = project_maps(source_type)
        return _first_non_blank(
            _first_browser_value(code_map, mcn_project_value),
            _first_browser_value(code_map, project_value),
            _text(mcn_project_value),
            _text(project_value),
        )

    def project_name(source_type, existing_name, project_value, mcn_project_value):
        _, name_map = project_maps(source_type)
        return _first_non_blank(
            existing_name,
            _first_browser_value(name_map, mcn_project_value),
            _first_browser_value(name_map, project_value),
        )

    df['申请人'] = df['申请人ID'].map(lambda value: employee_map.get(c.format_code(value), {}).get('name', ''))
    df['申请人工号'] = df['申请人ID'].map(lambda value: employee_map.get(c.format_code(value), {}).get('code', ''))
    df['转入公司主体'] = df['转入公司主体ID'].map(lambda value: _first_browser_value(company_map, value))
    df['转出公司主体'] = df['转出公司主体ID'].map(lambda value: _first_browser_value(company_map, value))
    df['转入成本中心'] = df['转入成本中心ID'].map(lambda value: _first_browser_value(cost_center_map, value))
    df['转出成本中心'] = df['转出成本中心ID'].map(lambda value: _first_browser_value(cost_center_map, value))
    df['转入项目编号'] = [
        project_code(source_type, in_project, in_mcn_project)
        for source_type, in_project, in_mcn_project
        in zip(df['来源类型'], df['转入项目编号ID'], df['转入MCN赛事项目编号ID'])
    ]
    df['转出项目编号'] = [
        project_code(source_type, out_project, out_mcn_project)
        for source_type, out_project, out_mcn_project
        in zip(df['来源类型'], df['转出项目编号ID'], df['转出MCN赛事项目编号ID'])
    ]
    df['转入项目名称'] = [
        project_name(source_type, in_project_name, in_project, in_mcn_project)
        for source_type, in_project_name, in_project, in_mcn_project
        in zip(df['来源类型'], df['转入项目名称'], df['转入项目编号ID'], df['转入MCN赛事项目编号ID'])
    ]
    df['转出项目名称'] = [
        project_name(source_type, out_project_name, out_project, out_mcn_project)
        for source_type, out_project_name, out_project, out_mcn_project
        in zip(df['来源类型'], df['转出项目名称'], df['转出项目编号ID'], df['转出MCN赛事项目编号ID'])
    ]
    df['付款币种'] = df['付款币种ID'].map(lambda value: currency_map.get(c.format_code(value), '人民币'))
    df['预算科目'] = _resolve_subject_paths(df['预算科目ID'])
    df['费用单据号'] = [
        _first_non_blank(fee_doc_no, fee_id, flow_no)
        for fee_doc_no, fee_id, flow_no in zip(df['费用单据号'], df['费用单ID'], df['流程编号'])
    ]
    return df


def _build_external_side_rows(source_df, side):
    is_in = side == 'in'
    entity_col = '转入公司主体' if is_in else '转出公司主体'
    cost_center_col = '转入成本中心' if is_in else '转出成本中心'
    project_col = '转入项目编号' if is_in else '转出项目编号'
    project_name_col = '转入项目名称' if is_in else '转出项目名称'
    amount_col = '转入金额' if is_in else '转出金额'
    sign = 1 if is_in else -1
    side_label = '转入方' if is_in else '转出方'

    df = source_df.copy()
    df['方向'] = side_label
    df['公司主体'] = df[entity_col]
    df['成本中心'] = df[cost_center_col]
    df['项目编号'] = df[project_col]
    df['项目名称'] = df[project_name_col]
    df['金额'] = pd.to_numeric(df[amount_col], errors='coerce') * sign
    return df


def build_external_cost_output(source_df):
    if source_df.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    expanded_df = pd.concat([
        _build_external_side_rows(source_df, 'in'),
        _build_external_side_rows(source_df, 'out'),
    ], ignore_index=True)
    allowed_projects = project_filter_codes(EVENT_PROJECT_SHEET) | project_filter_codes(MCN_PROJECT_SHEET)
    before_count = len(expanded_df)
    expanded_df = expanded_df[expanded_df['项目编号'].map(_text).isin(allowed_projects)].copy()
    print(f'[应付期初-只转入外部成本] 输出行项目白名单过滤: {len(expanded_df)}/{before_count} 行')
    if expanded_df.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS), expanded_df

    vendor_map = c.build_vendor_map()
    virtual_vendor_code = vendor_map.get(c.normalize_name(VIRTUAL_VENDOR_NAME), '')
    entity_map = c.build_accounting_entity_map_for_names(expanded_df['公司主体'])
    subject_lookup = c.build_subject_map()

    output_df = pd.DataFrame(index=expanded_df.index)
    output_df['来源系统'] = SOURCE_SYSTEM
    output_df['来源单据编号'] = expanded_df['流程编号']
    output_df['申请日期'] = expanded_df['申请日期'].map(c.format_date)
    output_df['单据类型'] = DOCUMENT_TYPE
    output_df['申请人工号'] = expanded_df['申请人工号']
    output_df['申请人姓名'] = expanded_df['申请人']
    output_df['核算主体编号'] = expanded_df['公司主体'].map(lambda value: _lookup_by_name(entity_map, value))
    output_df['核算主体描述'] = expanded_df['公司主体']
    output_df['备注'] = [
        f'{side}:{fee_doc_no}'[:150]
        for side, fee_doc_no in zip(expanded_df['方向'], expanded_df['费用单据号'])
    ]
    output_df['合同号'] = ''
    output_df['合同收支计划行'] = ''
    output_df['收款方编码'] = virtual_vendor_code
    output_df['收款方描述'] = VIRTUAL_VENDOR_NAME
    output_df['银行账号'] = c.resolve_hand_vendor_bank_accounts(output_df['收款方编码'])
    output_df['计划付款日期'] = ''
    output_df['银行转账备注'] = ''
    output_df['实际已支付金额'] = expanded_df['金额'].map(c.round_amount)
    output_df['费用项目编码'] = expanded_df['预算科目'].map(lambda value: _subject_item(subject_lookup, value, 0))
    output_df['费用项目描述'] = expanded_df['预算科目'].map(lambda value: _subject_item(subject_lookup, value, 1))
    output_df['主播房间号'] = ''
    output_df['报账币种'] = expanded_df['付款币种'].map(c.to_iso_currency)
    output_df['报账金额（支付币种）'] = expanded_df['金额'].map(c.round_amount)
    output_df['泛微费用项目编码'] = expanded_df['预算科目'].where(expanded_df['预算科目'].notna(), '')

    # 供问题清单复用。
    expanded_df['收款方描述'] = VIRTUAL_VENDOR_NAME
    return _apply_order_project_columns(output_df, expanded_df), expanded_df


def collect_external_cost_pair_check(expanded_df):
    """校验只转入外部成本:每个单据/明细的正数转入行应有对应负数转出行。"""
    if expanded_df.empty:
        return {
            '转入转出校验': pd.DataFrame([{
                '校验结果': '无数据',
                '说明': '只转入外部成本无输出行',
            }]),
        }

    df = expanded_df.copy()
    df['金额'] = pd.to_numeric(df['金额'], errors='coerce').fillna(0).round(2)
    df['配对明细ID'] = [
        _first_non_blank(detail_id, view_id, fee_doc_no)
        for detail_id, view_id, fee_doc_no in zip(df['明细ID'], df['费用明细视图ID'], df['费用单据号'])
    ]

    detail_rows = []
    failed_detail_keys = set()
    for (doc_no, detail_id), group in df.groupby(['流程编号', '配对明细ID'], sort=False):
        positive = group[group['金额'] > 0]
        negative = group[group['金额'] < 0]
        zero_count = int((group['金额'] == 0).sum())
        if len(positive) == 0 and len(negative) == 0:
            continue
        positive_total = round(float(positive['金额'].sum()), 2)
        negative_total = round(float(negative['金额'].sum()), 2)
        diff = round(positive_total + negative_total, 2)
        is_ok = (
            len(positive) == 1
            and len(negative) == 1
            and abs(diff) <= 0.01
        )
        if is_ok:
            continue

        failed_detail_keys.add((doc_no, detail_id))
        sample = group.iloc[0]
        detail_rows.append({
            '来源单据编号': doc_no,
            '配对明细ID': detail_id,
            '费用单据号': _text(sample.get('费用单据号', '')),
            '预算科目': _text(sample.get('预算科目', '')),
            '正数行数': len(positive),
            '负数行数': len(negative),
            '零金额行数': zero_count,
            '正数金额合计': positive_total,
            '负数金额合计': negative_total,
            '正负合计差额': diff,
            '校验结果': '通过' if is_ok else '不通过',
        })

    summary_rows = []
    for doc_no, group in df.groupby('流程编号', sort=False):
        positive = group[group['金额'] > 0]
        negative = group[group['金额'] < 0]
        zero_count = int((group['金额'] == 0).sum())
        positive_total = round(float(positive['金额'].sum()), 2)
        negative_total = round(float(negative['金额'].sum()), 2)
        diff = round(positive_total + negative_total, 2)
        detail_ids = set(zip(group['流程编号'], group['配对明细ID']))
        failed_count = len(detail_ids & failed_detail_keys)
        is_ok = (
            len(positive) == len(negative)
            and failed_count == 0
            and abs(diff) <= 0.01
        )
        summary_rows.append({
            '来源单据编号': doc_no,
            '正数行数': len(positive),
            '负数行数': len(negative),
            '零金额行数': zero_count,
            '正数金额合计': positive_total,
            '负数金额合计': negative_total,
            '正负合计差额': diff,
            '未配对明细数': failed_count,
            '校验结果': '通过' if is_ok else '不通过',
        })

    sheets = {
        '转入转出校验': pd.DataFrame(summary_rows),
    }
    if detail_rows:
        sheets['转入转出校验异常明细'] = pd.DataFrame(detail_rows)
    return sheets


def run():
    # 1. SQL 读取并解析三个来源
    base_source_df = read_base_source()
    mcn_payment_source_df = read_mcn_payment_source()
    batch_source_df = read_batch_source()
    external_source_df = read_external_cost_source()

    # 2. 构建六个 sheet 输出:原三类保持独立,MCN 三类追加在后。
    base_issue_source_df = _with_resolved_project_fields(base_source_df)
    base_output_df = _apply_order_project_columns(
        base_payment.build_output(base_source_df),
        base_issue_source_df,
    )
    if '申请人' not in base_issue_source_df.columns and '经办人' in base_issue_source_df.columns:
        base_issue_source_df['申请人'] = base_issue_source_df['经办人']
    elif '经办人' in base_issue_source_df.columns:
        base_issue_source_df['申请人'] = [
            _first_non_blank(applicant, handler)
            for applicant, handler in zip(base_issue_source_df.get('申请人', ''), base_issue_source_df['经办人'])
        ]
    if '供应商' not in base_issue_source_df.columns:
        base_issue_source_df['供应商'] = ''
    batch_output_df = build_batch_output(batch_source_df)
    external_output_df, external_issue_source_df = build_external_cost_output(external_source_df)

    mcn_sheet_specs = [
        (SHEET_MCN_OUTBOUND, 'MCN对外付款流程'),
        (SHEET_MCN_ORDER, 'MCN对外付款流程（订单）'),
        (SHEET_MCN_ANCHOR, 'MCN主播相关付款流程'),
    ]
    mcn_output_dfs = {}
    mcn_issue_source_dfs = {}
    for sheet_name, source_label in mcn_sheet_specs:
        source_part = mcn_payment_source_df[
            mcn_payment_source_df['来源流程'].map(_text) == source_label
        ].copy() if not mcn_payment_source_df.empty else pd.DataFrame()
        output_part, issue_part = build_mcn_payment_output(source_part)
        mcn_output_dfs[sheet_name] = output_part
        mcn_issue_source_dfs[sheet_name] = issue_part

    print('[应付期初-期初对公付款单导入] 输出明细行数:', len(base_output_df))
    print('[应付期初-批量费用流程] 输出明细行数:', len(batch_output_df))
    print('[应付期初-只转入外部成本] 输出明细行数:', len(external_output_df))
    for sheet_name, output_df in mcn_output_dfs.items():
        print(f'[应付期初-{sheet_name}] 输出明细行数:', len(output_df))

    # 3. 填充率
    required_cols = c.required_columns(RULE_SHEET, RULE_TABLE)
    print('— 期初对公付款单导入 填充率 —')
    c.report_fill(base_output_df, required_cols)
    print('— 批量费用流程 填充率 —')
    c.report_fill(batch_output_df, required_cols)
    print('— 只转入外部成本 填充率 —')
    c.report_fill(external_output_df, required_cols)
    for sheet_name, output_df in mcn_output_dfs.items():
        print(f'— {sheet_name} 填充率 —')
        c.report_fill(output_df, required_cols)

    # 4. 写入模板:同一个 Excel 内写入六个数据 sheet,lov 页保留。
    write_output_workbook(base_output_df, batch_output_df, external_output_df, mcn_output_dfs)
    print('已写出:', OUTPUT_FILE)

    # 5. 问题清单
    exception_sheets = {}

    base_sheets = {'必输字段未达100%': c.fill_summary(
        base_output_df, required_cols, RULE_SHEET, RULE_TABLE)}
    base_sheets.update(c.collect_field_issues(
        base_output_df, base_issue_source_df, required_cols, {
            '申请人工号': '申请人',
            '收款方编码': '供应商',
            '核算主体编号': '公司主体',
            '费用项目编码': '预算科目',
            '报账币种': '付款币种',
            '订单编号': '项目编号',
        }))
    base_bank_issues = c.collect_hand_vendor_bank_account_issues(
        base_output_df, base_issue_source_df['银行账号'])
    if not base_bank_issues.empty:
        base_sheets['银行账号_校验异常'] = base_bank_issues
    base_sheets = _enrich_missing_order_issue(base_sheets, base_output_df, base_issue_source_df)
    base_sheets.update(c.collect_order_mapping_issues(base_issue_source_df))
    c.attach_budget_issue_columns(base_sheets, c.build_budget_issue_map(base_issue_source_df))
    exception_sheets.update({f'期初对公付款单导入_{name}': df for name, df in base_sheets.items()})

    mcn_issue_fields = {
        '申请人工号': '申请人',
        '收款方编码': '供应商',
        '核算主体编号': '公司主体',
        '费用项目编码': '预算科目',
        '订单编号': '项目编号',
    }
    for sheet_name in mcn_output_dfs:
        output_df = mcn_output_dfs[sheet_name]
        issue_source_df = mcn_issue_source_dfs[sheet_name]
        mcn_sheets = {'必输字段未达100%': c.fill_summary(
            output_df, required_cols, RULE_SHEET, RULE_TABLE)}
        mcn_sheets.update(c.collect_field_issues(
            output_df, issue_source_df, required_cols, mcn_issue_fields))
        if '银行账号' in issue_source_df.columns:
            mcn_bank_issues = c.collect_hand_vendor_bank_account_issues(output_df, issue_source_df['银行账号'])
            if not mcn_bank_issues.empty:
                mcn_sheets['银行账号_校验异常'] = mcn_bank_issues
        mcn_sheets = _enrich_missing_order_issue(mcn_sheets, output_df, issue_source_df)
        mcn_sheets.update(c.collect_order_mapping_issues(issue_source_df))
        c.attach_budget_issue_columns(mcn_sheets, c.build_budget_issue_map(issue_source_df))
        exception_sheets.update({f'{sheet_name}_{name}': df for name, df in mcn_sheets.items()})

    batch_sheets = {'必输字段未达100%': c.fill_summary(
        batch_output_df, required_cols, RULE_SHEET, RULE_TABLE)}
    batch_sheets.update(c.collect_field_issues(
        batch_output_df, batch_source_df, required_cols, BATCH_ISSUE_SOURCE_FIELDS))
    batch_bank_issues = c.collect_hand_vendor_bank_account_issues(batch_output_df)
    if not batch_bank_issues.empty:
        batch_sheets['银行账号_校验异常'] = batch_bank_issues
    batch_sheets = _enrich_missing_order_issue(batch_sheets, batch_output_df, batch_source_df)
    batch_sheets.update(c.collect_order_mapping_issues(batch_source_df))
    c.attach_budget_issue_columns(batch_sheets, c.build_budget_issue_map(batch_source_df))
    exception_sheets.update({f'批量费用流程_{name}': df for name, df in batch_sheets.items()})

    external_sheets = {'必输字段未达100%': c.fill_summary(
        external_output_df, required_cols, RULE_SHEET, RULE_TABLE)}
    external_sheets.update(c.collect_field_issues(
        external_output_df, external_issue_source_df, required_cols, EXTERNAL_COST_ISSUE_SOURCE_FIELDS))
    external_bank_issues = c.collect_hand_vendor_bank_account_issues(external_output_df)
    if not external_bank_issues.empty:
        external_sheets['银行账号_校验异常'] = external_bank_issues
    external_sheets = _enrich_missing_order_issue(external_sheets, external_output_df, external_issue_source_df)
    external_sheets.update(c.collect_order_mapping_issues(external_issue_source_df))
    external_sheets.update(collect_external_cost_pair_check(external_issue_source_df))
    c.attach_budget_issue_columns(external_sheets, c.build_budget_issue_map(external_issue_source_df))
    exception_sheets.update({f'只转入外部成本_{name}': df for name, df in external_sheets.items()})

    c.write_exceptions(EXCEPTION_FILE, exception_sheets)
    print('已写出:', EXCEPTION_FILE, '| 各清单条数:', {k: len(v) for k, v in exception_sheets.items()})


if __name__ == '__main__':
    run()
